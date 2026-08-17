# -*- coding: utf-8 -*-
"""工时对账核心的合成数据回归测试。"""

import datetime
import os
import tempfile
import unittest

import openpyxl

from core import common_core as cc
from core import reconcile_core as rc


class TestReconcileComparison(unittest.TestCase):
    """验证内存比较阶段的异常顺序、总工时和逐日差异口径。"""

    def test_reconcile_keeps_roster_and_difference_order(self):
        """名单差异应排在工时差异之前，同一人总工时差异应先于逐日差异。"""

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.append(["姓名", "劳务公司", 1, 2, "出勤工时"])
        worksheet.append(["张三", "甲公司", 8, 8, 16])
        worksheet.append(["李四", "乙公司", 8, 8, 16])
        layout = {
            "name_col": 1,
            "comp_col": 2,
            "day_cols": {1: 3, 2: 4},
            "work_col": 5,
            "check_col": None,
            "data_start": 2,
        }
        labor = {
            "张三": {"days": {1: 8, 2: 7}, "total": 15, "source": "甲.xlsx"},
            "王五": {"days": {1: 8}, "total": 8, "source": "乙.xlsx"},
        }
        try:
            anomalies = rc.reconcile(worksheet, layout, labor)
        finally:
            workbook.close()

        self.assertEqual(
            [item["异常类型"] for item in anomalies],
            ["仅我司名单有", "仅劳务公司有", "总工时不一致", "逐日工时不一致"],
        )  # 异常按业务顺序排列
        self.assertEqual(anomalies[2]["差异"], 1.0)  # 总工时差异
        self.assertEqual(anomalies[3]["差异明细"], "2日:我司8/劳务7")  # 逐日差异明细


class TestReconcileWorkflow(unittest.TestCase):
    """使用临时 Excel 文件验证完整运行编排、人工姓名配对和输出。"""

    def setUp(self):
        """创建完全隔离的临时输入和输出目录。"""

        self.temp = tempfile.TemporaryDirectory()
        self.target_path = os.path.join(self.temp.name, "待对总表.xlsx")
        self.source_path = os.path.join(self.temp.name, "我司来源.xlsx")
        self.labor_path = os.path.join(self.temp.name, "劳务考勤.xlsx")
        self.out_dir = os.path.join(self.temp.name, "输出")  # 输出目录隔离
        self._write_target()
        self._write_source()
        self._write_labor()

    def tearDown(self):
        """释放临时目录及其中生成的工作簿。"""

        self.temp.cleanup()

    def _write_target(self):
        """生成包含十个日期列的待对总表模板。"""

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "总表"
        worksheet.append(
            ["姓名", "劳务公司", *range(1, 11), "出勤工时", "对账时间"]
        )
        worksheet.append(["张三", "甲公司", *([None] * 10), None, None])
        workbook.save(self.target_path)
        workbook.close()

    def _write_source(self):
        """生成张三连续十天、每天八小时的我司来源明细。"""

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.append(["姓名", "日期", "实际工作时间"])
        for day in range(1, 11):
            worksheet.append(["张三", datetime.date(2026, 8, day), 8])
        workbook.save(self.source_path)
        workbook.close()

    def _write_labor(self):
        """生成姓名写法不同但工时一致的劳务考勤表，用于验证人工配对。"""

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "考勤"
        worksheet.append(["姓名", *range(1, 11), "出勤工时"])
        worksheet.append(["张三劳务", *([8] * 10), 80])
        workbook.save(self.labor_path)
        workbook.close()

    def test_run_applies_alias_and_generates_outputs(self):
        """人工姓名配对应让双方成功对齐，并生成已填写表和可信度报告。"""

        progress = []
        result = rc.run(
            self.target_path,
            self.source_path,
            [self.labor_path],
            out_dir=self.out_dir,
            opts=cc.Options(),
            choices={"aliases": {"张三劳务": "张三"}},
            progress=progress.append,
        )

        self.assertEqual(result["anomalies"], [])  # 人工配对后无异常
        self.assertEqual(result["metrics"]["matched_pairs"], 1)  # 一对匹配
        self.assertEqual(result["metrics"]["only_us"], 0)
        self.assertEqual(result["metrics"]["only_labor"], 0)
        self.assertTrue(os.path.isfile(result["filled_path"]))  # 已填写表输出
        self.assertTrue(os.path.isfile(result["summary_path"]))  # 汇总报告输出
        self.assertEqual(progress[-1], 100)  # 进度最终到 100
        self.assertEqual(progress, sorted(progress))  # 进度单调递增

        filled = openpyxl.load_workbook(result["filled_path"], data_only=True)
        try:
            worksheet = filled["总表"]
            self.assertEqual(worksheet.cell(2, 13).value, 80)  # 出勤工时填写
            self.assertIsNotNone(worksheet.cell(2, 14).value)  # 对账时间填写
        finally:
            filled.close()


class TestLaborMergePolicy(unittest.TestCase):
    """验证重复姓名冲突策略仍保持既有覆盖语义。"""

    def test_first_policy_keeps_existing_value(self):
        """先者优先时应记录重复数量，但不能覆盖先读取的数据。"""

        merged = {"张三": {"total": 8, "days": {1: 8}, "source": "先.xlsx"}}
        incoming = {"张三": {"total": 7, "days": {1: 7}}}
        logs = []
        duplicates = rc._merge_labor_file(
            merged, incoming, "后.xlsx", "first", logs.append,
        )

        self.assertEqual(duplicates, 1)  # 重复计数
        self.assertEqual(merged["张三"]["total"], 8)  # 先者优先保留
        self.assertIn("先者优先", logs[0])


if __name__ == "__main__":
    unittest.main()
