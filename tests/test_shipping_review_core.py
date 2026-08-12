# -*- coding: utf-8 -*-
"""发运评审对比业务的合成数据回归测试。"""

from __future__ import annotations

import os
import tempfile
import unittest
from unittest import mock

import openpyxl

from core import shipping_review_core


class ShippingReviewCoreTests(unittest.TestCase):
    """验证过滤、双侧汇总、活动页签选择、状态分类和报告结构。"""

    def setUp(self):
        self.temp = tempfile.TemporaryDirectory(prefix="fyt_shipping_review_")
        self.package = self.path("包装日计划.xlsx")
        self.review = self.path("发运评审.xlsx")
        self.catalog = self.path("主数据.json")

    def tearDown(self):
        self.temp.cleanup()

    def path(self, name: str) -> str:
        return os.path.join(self.temp.name, name)

    @staticmethod
    def _save_book(path, sheets, *, active=0):
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
        for title, rows in sheets:
            worksheet = workbook.create_sheet(title)
            for row in rows:
                worksheet.append(row)
        workbook.active = active
        workbook.save(path)
        workbook.close()

    def _package_rows(self):
        return [
            ["说明", None, None, None],
            ["物料号", "物料描述", "BOX状态", "实际包装数量", "BOX号"],
            ["A-01", "螺栓", "已配载", 4, "BOX-1"],
            ["A-01", "螺栓", "翻包完成", 6, "BOX-2"],
            ["B-02", "垫片", "已作废", 99, "BOX-X"],
            ["B-02", "垫片", "已配载", 8, "BOX-3"],
            ["C-03", "", "已配载", 2, "BOX-4"],
        ]

    def _review_rows(self, *, a_total=10):
        return [
            ["Part No", "Chinese Name", "总数", "供应商"],
            ["A-01", "螺栓", 4, "甲"],
            ["A-01", "螺栓", a_total - 4, "乙"],
            ["B-02", "垫片", 8, "甲"],
            ["C-03", "", 2, "甲"],
        ]

    def _run(self, **kwargs):
        with mock.patch.dict(os.environ, {"FYT_CATALOG_PATH": self.catalog}):
            return shipping_review_core.run(
                self.package,
                self.review,
                out_dir=self.path("输出"),
                **kwargs,
            )

    def test_filters_obsolete_and_sums_duplicate_rows_on_both_sides(self):
        """包装重复 BOX 与评审多供应商拆行都应累加，已作废记录不得进入数量。"""

        self._save_book(self.package, [("包装日计划", self._package_rows())])
        self._save_book(self.review, [("评审", self._review_rows())])
        result = self._run()
        self.assertEqual(result["source_rows"], 5)
        self.assertEqual(result["kept_rows"], 4)
        self.assertEqual(result["obsolete_rows"], 1)
        self.assertEqual(result["package_materials"], 3)
        self.assertEqual(result["review_materials"], 3)
        self.assertEqual(result["counts"]["quantity_diff"], 0)
        self.assertEqual(result["counts"]["full_match"], 2)
        self.assertEqual(result["counts"]["name_issues"], 1)
        a_row = next(row for row in result["details"] if row["code"] == "A-01")
        self.assertEqual(a_row["package_quantity"], 10)
        self.assertEqual(a_row["review_quantity"], 10)
        self.assertEqual(a_row["review_rows"], 2)

    def test_saved_active_review_sheet_is_used(self):
        """未人工指定时必须读取工作簿保存时的活动页签，而不是默认首页。"""

        self._save_book(self.package, [("包装日计划", self._package_rows())])
        self._save_book(self.review, [
            ("首页", [["说明"], ["请查看活动工作表"]]),
            ("评审A", self._review_rows()),
        ], active=1)
        result = self._run()
        self.assertEqual(result["review_sheet"], "评审A")
        self.assertEqual(result["counts"]["quantity_diff"], 0)

    def test_requested_review_sheet_overrides_saved_active_sheet(self):
        """人工页签覆盖活动页签，适用于工作簿保存状态不可信的情况。"""

        self._save_book(self.package, [("包装日计划", self._package_rows())])
        self._save_book(self.review, [
            ("旧版", self._review_rows(a_total=9)),
            ("正式版", self._review_rows(a_total=10)),
        ], active=0)
        result = self._run(review_sheet="正式版")
        self.assertEqual(result["review_sheet"], "正式版")
        self.assertEqual(result["counts"]["quantity_diff"], 0)

    def test_quantity_difference_and_missing_name_have_distinct_status(self):
        """数量差异和名称缺失应独立统计，并在同一物料同时发生时生成组合状态。"""

        self._save_book(self.package, [("包装日计划", self._package_rows())])
        review_rows = self._review_rows(a_total=9)
        review_rows[3][1] = ""  # B-02 只缺评审名称，数量仍一致。
        self._save_book(self.review, [("评审", review_rows)])
        result = self._run()
        statuses = {row["code"]: row["status"] for row in result["details"]}
        self.assertEqual(statuses["A-01"], "数量差异")
        self.assertEqual(statuses["B-02"], "评审名称缺失")
        self.assertEqual(result["counts"]["quantity_diff"], 1)
        self.assertEqual(result["counts"]["name_issues"], 2)

    def test_report_contains_four_auditable_sheets(self):
        """正式报告必须包含总表、异常、包装透视和过滤审计四个页签。"""

        self._save_book(self.package, [("包装日计划", self._package_rows())])
        self._save_book(self.review, [("评审", self._review_rows())])
        result = self._run()
        self.assertTrue(os.path.isfile(result["report_path"]))
        workbook = openpyxl.load_workbook(result["report_path"], data_only=False)
        try:
            self.assertEqual(workbook.sheetnames, ["对比总表", "异常明细", "包装透视", "过滤审计"])
            self.assertEqual(workbook["过滤审计"]["A1"].value, "包装日计划过滤审计")
            values = [cell.value for row in workbook["过滤审计"].iter_rows() for cell in row]
            self.assertIn("BOX-X", values)
        finally:
            workbook.close()

    def test_missing_required_headers_reports_chinese_error(self):
        """输入缺列时应报告缺失业务字段，不得抛出索引异常。"""

        self._save_book(self.package, [("包装日计划", [["物料号", "物料描述", "实际包装数量"], ["A", "螺栓", 1]])])
        self._save_book(self.review, [("评审", self._review_rows())])
        with self.assertRaisesRegex(ValueError, "BOX状态"):
            self._run()


if __name__ == "__main__":
    unittest.main()
