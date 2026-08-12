"""日清报告聚合与输出回归测试。"""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from core import daily_report_core


class DailyReportCoreTests(unittest.TestCase):
    """保护日清快照聚合、模板字段和 Excel 多工作表输出结构。"""

    def test_build_snapshot_and_excel(self):
        """到料与现场问题应聚合为管理摘要，并写入完整日清工作簿。"""

        arrivals = [
            {
                "id": "job-a",
                "title": "到料明细一",
                "display_name": "计划员甲",
                "updated_at": "2026-08-05T02:10:00+00:00",
                "result": {"result": {
                    "results": [["26035-01", 2, 8, 10]],
                    "batches": [{
                        "batch_no": "26035-01", "missing_count": 2,
                        "arrived_count": 8, "total_count": 10,
                        "missing_materials": [{
                            "material_code": "A-01", "material_name": "固定螺栓",
                            "supplier": "供应商甲", "demand_quantity": 12,
                            "received_quantity": 9, "shortage_quantity": 3,
                        }],
                    }],
                }, "logs": [], "task_id": "a", "out_dir": "输出"},
            },
            {
                "id": "job-b",
                "title": "到料明细二",
                "display_name": "计划员乙",
                "updated_at": "2026-08-05T03:20:00+00:00",
                "result": {"results": [["26035-02", 0, 5, 5]]},
            },
        ]
        issues = [
            {
                "id": "issue-a",
                "cause": "防护罩松动",
                "primary_owner": "张工",
                "secondary_owner": "李工",
                "notes": "下班前复核",
                "created_at": "2026-08-05T04:00:00+00:00",
                "uploader": {"display_name": "车间甲"},
                "images": [{"id": "image-a"}, {"id": "image-b"}],
            },
            {
                "id": "issue-b",
                "cause": "物料摆放不规范",
                "primary_owner": "张工",
                "secondary_owner": "",
                "notes": "",
                "created_at": "2026-08-05T05:00:00+00:00",
                "uploader": {"display_name": "车间乙"},
                "images": [],
            },
        ]
        # 两批到料来自不同任务封装形态，验证投影层同时兼容嵌套和直接结果。
        snapshot = daily_report_core.build_snapshot(
            "2026-08-05", arrivals, issues, generated_at="2026-08-05T06:00:00+00:00",
        )
        self.assertEqual(snapshot["arrival"]["batch_count"], 2)
        self.assertEqual(snapshot["arrival"]["total_categories"], 15)
        self.assertEqual(snapshot["arrival"]["completion_rate"], 86.7)
        self.assertEqual(snapshot["arrival"]["missing_material_detail_count"], 1)
        self.assertEqual(snapshot["arrival"]["batches"][0]["missing_materials"][0]["shortage_quantity"], "3")
        self.assertEqual(snapshot["arrival"]["supplier_distribution"][0]["supplier"], "供应商甲")
        self.assertEqual(snapshot["arrival"]["batches"][0]["supplier_distribution"][0]["shortage_quantity"], 3)
        self.assertEqual(snapshot["workshop"]["issue_count"], 2)
        self.assertEqual(snapshot["workshop"]["image_count"], 2)
        self.assertEqual(snapshot["workshop"]["owner_distribution"][0], {"owner": "张工", "count": 2})

        with tempfile.TemporaryDirectory() as temp_name:
            result = daily_report_core.run(snapshot, out_dir=temp_name)
            target = Path(result["out_file"])
            self.assertTrue(target.is_file())
            workbook = load_workbook(target, data_only=True)
            try:
                for sheet_name in (
                    "日清概览", "每日到料", "未到物料", "安全检查日报", "现场问题",
                    "月度生产订单台账", "订单缺件明细", "订单危包明细", "零星订单明细",
                ):
                    self.assertIn(sheet_name, workbook.sheetnames)
                self.assertEqual(workbook["日清概览"]["A1"].value, "峰运通日清报告 · 2026-08-05")
                self.assertEqual(workbook["每日到料"]["A2"].value, "26035-01")
                self.assertEqual(workbook["未到物料"]["C2"].value, "A-01")
                self.assertEqual(workbook["未到物料"]["H2"].value, "3")
                self.assertEqual(workbook["现场问题"]["C2"].value, "防护罩松动")
            finally:
                workbook.close()

    def test_invalid_date_is_rejected(self):
        """不存在的日历日期不能进入日清快照或输出文件名。"""

        with self.assertRaisesRegex(ValueError, "YYYY-MM-DD"):
            daily_report_core.build_snapshot("2026-02-30", [], [])

    def test_template_issue_fields_and_attendance_summary(self):
        """问题模板扩展字段与参会/生产考勤摘要应原样进入快照。"""

        snapshot = daily_report_core.build_snapshot(
            "2026-08-05", [], [{
                "id": "issue-template", "cause": "包装破损", "primary_owner": "张工",
                "issue_source": "包装异常", "model": "VX11", "country": "宝腾",
                "batch_no": "GKMYR26027-06", "material_code": "A-01", "material_name": "侧围",
                "cause_analysis": "防护不足", "corrective_action": "补充防护", "external_inspection_owner": "王工",
                "tracking_status": "处理中", "images": [], "uploader": {"display_name": "车间"},
            }], attendance={
                "people": [
                    {"name": "参会甲", "person_type": "participant", "unit": "管理组", "shift": "白班", "present": False, "reason": "出差"},
                ],
                # 同一班组跨两班次，验证编制、出勤与差异按班次保留且总数正确聚合。
                "production_groups": [
                    {"group_id": 1, "group_name": "小件组", "shift_id": 11, "shift_name": "白班", "staffing_count": 14, "attendance_count": 17, "difference": -3, "note": "3 人支援白班"},
                    {"group_id": 1, "group_name": "小件组", "shift_id": 12, "shift_name": "夜班", "staffing_count": 14, "attendance_count": 14, "difference": 0, "note": ""},
                ],
                "present_count": 31, "absent_count": 1, "participant_present_count": 0,
                "participant_absent_count": 1, "participant_total": 1,
                "production_present_count": 31, "production_total": 28, "production_staffing_count": 28,
                "production_difference": -3, "production_shortage_count": 0,
                "production_group_count": 1, "production_shift_count": 2,
            },
        )
        issue = snapshot["workshop"]["issues"][0]
        self.assertEqual(issue["batch_no"], "GKMYR26027-06")
        self.assertEqual(issue["external_inspection_owner"], "王工")
        summary = snapshot["attendance"]["unit_summary"]
        self.assertEqual(summary[0]["difference"], 1)
        self.assertEqual(summary[0]["reasons"], ["参会甲：出差"])
        self.assertEqual(snapshot["attendance"]["production_groups"][0]["attendance_count"], 17)
        self.assertEqual(snapshot["attendance"]["production_difference"], -3)

    def test_production_ledger_tolerates_formatted_and_dirty_quantities(self):
        """月度台账应识别千分位数量，并跳过单个脏值而不中断整份日报。"""

        plans = [{
            "id": "plan-a",
            "original_name": "8月生产计划.xlsx",
            "summary": {"insights": {"order_ledger": {
                "formal_orders": [
                    {"order_no": "F-001", "month": "2026-08", "quantity": "1,200", "completed": True},
                    {"order_no": "F-002", "month": "2026-08", "quantity": "待确认", "completed": False},
                ],
                "sporadic_orders": [{
                    "order_no": "S-001", "month": "2026-08", "pallet_count": "2",
                    "volume_cbm": "1.2500", "completed": False, "shipment_dates": ["2026-08-11"],
                }],
            }}},
        }]
        snapshot = daily_report_core.build_snapshot(
            "2026-08-11", [], [], monthly_production_plans=plans,
        )
        ledger = snapshot["production_ledger"]
        self.assertEqual(ledger["formal_quantity"], 1200.0)
        self.assertEqual(ledger["sporadic_pallets"], 2.0)
        self.assertEqual(ledger["sporadic_volume_cbm"], 1.25)
        self.assertEqual(ledger["today_shipments"][0]["order_no"], "S-001")


if __name__ == "__main__":
    unittest.main()
