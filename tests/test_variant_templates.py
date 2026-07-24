# -*- coding: utf-8 -*-
"""变种模板识别回归测试。

真实用户的表格并不都照标准模板来:列名换同义词、加/减前后缀、表头前有标题行、
表头文字带换行或空格。这里用一批"非标准但真实"的变体喂给各功能的识别入口,
锁定自动识别能力,防止日后收紧匹配时悄悄退化。

只测识别(表头行/列角色),不写文件;合成数据即可,不依赖样本。
"""
import os
import tempfile
import datetime
import unittest
import warnings

import openpyxl

warnings.filterwarnings("ignore")


def _mk(rows, sheet="Sheet1"):
    """把二维 rows 写成临时 xlsx,返回路径。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet
    for r in rows:
        ws.append(r)
    fd, p = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(p)
    wb.close()
    return p


def _ws(rows):
    """构造已加载的 worksheet(供直接调 detect_layout)。"""
    p = _mk(rows)
    wb = openpyxl.load_workbook(p)
    return wb, wb.active


D = datetime.datetime


class TestAttendanceVariants(unittest.TestCase):
    """考勤源:姓名/日期/上班·下班打卡列的同义变体都应能识别整表。"""

    def _load_n(self, hdr):
        from core import attendance_core as ac
        p = _mk([hdr, ["张三", D(2026, 5, 1), "08:00", "17:00"],
                 ["张三", D(2026, 5, 2), "08:05", "17:10"]])
        return len(ac.load_source(p))

    def test_standard(self):
        self.assertEqual(self._load_n(["姓名", "日期", "上班1打卡时间", "下班1打卡时间"]), 2)

    def test_name_prefixed(self):        # 员工姓名
        self.assertEqual(self._load_n(["员工姓名", "日期", "上班1打卡时间", "下班1打卡时间"]), 2)

    def test_date_prefixed(self):        # 考勤日期
        self.assertEqual(self._load_n(["姓名", "考勤日期", "上班1打卡时间", "下班1打卡时间"]), 2)

    def test_punch_without_1(self):      # 上班打卡时间(无"1")
        self.assertEqual(self._load_n(["姓名", "日期", "上班打卡时间", "下班打卡时间"]), 2)

    def test_multi_punch_first_wins(self):
        # 有多次打卡列时,取第一次(上班1),不被"上班2打卡"覆盖
        from core import attendance_core as ac
        p = _mk([["姓名", "日期", "上班1打卡时间", "上班2打卡时间", "下班1打卡时间"],
                 ["张三", D(2026, 5, 1), "08:00", "12:00", "17:00"]])
        data = ac.load_source(p)
        # 取到的上班打卡应是 08:00(第一次),而非 12:00
        (on, off) = list(data.values())[0]
        self.assertEqual(str(on), "08:00")

    def test_fill_target_col_not_mismatched(self):
        # "上班时间(系统)"是填报目标列(无"打卡"),不应被当作源打卡列 -> 整表识别失败
        from core import attendance_core as ac
        p = _mk([["姓名", "日期", "上班时间（系统）", "下班时间（系统）"],
                 ["张三", D(2026, 5, 1), None, None]])
        with self.assertRaises(Exception):
            ac.load_source(p)


class TestReconcileVariants(unittest.TestCase):
    """工时对账源:工时列的同义变体 + 干扰列。"""

    def _detect(self, hdr):
        from core import reconcile_core as rc
        return rc._detect_source_header([hdr, ["张三", D(2026, 5, 1), 8]])

    def test_standard(self):
        self.assertIsNotNone(self._detect(["姓名", "日期", "实际工作时间"]))

    def test_shixi_gongshi(self):        # 实际工时
        self.assertIsNotNone(self._detect(["姓名", "日期", "实际工时"]))

    def test_bare_gongshi(self):         # 工时
        self.assertIsNotNone(self._detect(["姓名", "日期", "工时"]))

    def test_gongzuo_shichang(self):     # 工作时长
        self.assertIsNotNone(self._detect(["姓名", "日期", "工作时长"]))

    def test_name_prefixed(self):        # 员工姓名
        self.assertIsNotNone(self._detect(["员工姓名", "日期", "实际工作时间"]))

    def test_decoy_overtime_not_chosen(self):
        # "加班工时"(干扰)在前、"实际工作时间"(真列)在后 -> 应选真列
        from core import reconcile_core as rc
        det = rc._detect_source_header(
            [["姓名", "日期", "加班工时", "实际工作时间"], ["张三", D(2026, 5, 1), 2, 8]])
        self.assertIsNotNone(det)
        _, _, _, col_work = det
        self.assertEqual(col_work, 3)    # 第4列(0基3)=实际工作时间


class TestPurchaseVariants(unittest.TestCase):
    """采购列识别:编号/名称/数量的同义变体(用友/金蝶/SAP 常见用词)。"""

    def _detect(self, hdr):
        from core import purchase_core as pc
        wb, ws = _ws([["某某公司对账单"], hdr, ["A001", "货物", "规格1", "个", 10, "P1"]])
        try:
            return pc.detect_layout(ws)
        finally:
            wb.close()

    def test_standard(self):
        hr, col = self._detect(["材料编号", "材料名称", "规格", "单位", "采购数量", "批次号"])
        self.assertEqual(hr, 2)
        self.assertTrue({"no", "name", "qty"} <= set(col))

    def test_wuliao_bianma(self):        # 物料编码(此前 no 别名缺失)
        hr, col = self._detect(["物料编码", "物料名称", "规格", "单位", "数量", "批次号"])
        self.assertTrue({"no", "name", "qty"} <= set(col))

    def test_huohao(self):               # 货号
        hr, col = self._detect(["货号", "货物名称", "规格", "单位", "数量", "批次号"])
        self.assertTrue({"no", "name", "qty"} <= set(col))

    def test_wuliao_daima(self):         # 物料代码
        hr, col = self._detect(["物料代码", "物料名称", "规格", "单位", "数量", "批号"])
        self.assertTrue({"no", "name", "qty"} <= set(col))

    def test_cunhuo_bianma(self):        # 存货编码/结算数量
        hr, col = self._detect(["存货编码", "存货名称", "规格型号", "计量单位", "结算数量", "批号"])
        self.assertTrue({"no", "name", "qty"} <= set(col))


class TestDeliveryVariants(unittest.TestCase):
    """送货列识别:code 角色的同义变体。"""

    def _detect(self, hdr):
        from core import delivery_core as dc
        wb, ws = _ws([hdr, ["x"] * len(hdr)])
        try:
            return dc.detect_layout(ws)
        finally:
            wb.close()

    def test_standard_partcode(self):
        hr, col = self._detect(["批次号", "属性", "零部件代码", "零部件名称", "数量",
                                "供应商代码", "供应商名称"])
        self.assertIn("code", col)

    def test_sap_xiajie(self):           # 下阶物料
        hr, col = self._detect(["下阶物料", "下阶物料描述", "数量", "供应商代码", "供应商名称"])
        self.assertIn("code", col)

    def test_lingjian_bianhao(self):     # 零件编号(此前 code 别名缺失)
        hr, col = self._detect(["零件编号", "零件名称", "需求数量", "供方代码", "供方名称"])
        self.assertIn("code", col)

    def test_liaohao_pinming(self):      # 料号/品名
        hr, col = self._detect(["料号", "品名", "数量"])
        self.assertIn("code", col)


if __name__ == "__main__":
    unittest.main()


class TestDateRepresentations(unittest.TestCase):
    """同一天的多种存储形态都应归一到同键;尤其 Excel 序列号(丢日期格式的常见情形)。"""

    def test_norm_date_forms(self):
        from core.common_core import norm_date
        want = (2026, 5, 1)
        for v in [D(2026, 5, 1), datetime.date(2026, 5, 1), "20260501",
                  "2026-05-01", "2026/5/1", "2026.05.01",
                  46143, 46143.0, "46143", "2026-05-01 08:00:00"]:
            self.assertEqual(norm_date(v), want, "norm_date(%r)" % (v,))

    def test_day_of_forms(self):
        from core.common_core import day_of
        for v in [D(2026, 5, 1), 46143, 46143.0, "46143",
                  "2026-05-01", "20260501", "2026-05-01 08:00:00", 1, "1日"]:
            self.assertEqual(day_of(v), 1, "day_of(%r)" % (v,))

    def test_day_of_keeps_small_int_as_dom(self):
        # 1~31 仍按"当月第几天",不可被误当序列号
        from core.common_core import day_of
        self.assertEqual(day_of(15), 15)
        self.assertEqual(day_of(31), 31)

    def test_serial_out_of_range_rejected(self):
        # 越界数字不当日期(避免把普通数量/编号误读成日期)
        from core.common_core import norm_date
        self.assertIsNone(norm_date(100))
        self.assertIsNone(norm_date(999999))

    def test_reconcile_source_with_serial_dates(self):
        # 端到端:数据来源日期列为序列号,整表仍应正确按日聚合(此前会静默漏读)
        from core import reconcile_core as rc
        p = _mk([["姓名", "日期", "实际工作时间"],
                 ["张三", 46143, 8], ["张三", 46144, 7.5], ["张三", 46145, 8]])
        data, seen = {}, set()
        n = rc._load_source_one(p, data, seen)
        self.assertEqual(n, 3)
        self.assertAlmostEqual(sum(data["张三"].values()), 23.5)


class TestCodeRepresentations(unittest.TestCase):
    """同一编码不同存储形态(int/float/前导零/大小写)跨表应能对上。"""

    def test_norm_no_unifies(self):
        from core.purchase_core import norm_no
        for v in [123, 123.0, "123", "00123", " 123 "]:
            self.assertEqual(norm_no(v), "123", "norm_no(%r)" % (v,))
        self.assertEqual(norm_no("jbc0202"), "JBC0202")

    def test_purchase_match_across_code_types(self):
        # 我方文本'00123' vs 供方 int 123 / float 123.0 应匹配上
        from core import purchase_core as pc
        r1 = [{"no": "00123", "name": "货物A", "spec": "S1", "qty": 10,
               "batch": "B1", "note": "", "r": 3}]
        for other in (123, 123.0):
            r2 = [{"no": other, "name": "货物A", "spec": "s1", "qty": 10,
                   "batch": "B1", "note": "", "r": 3}]
            self.assertTrue(pc.match_rows(r1, r2), "vs %r" % (other,))


def _mkbook(sheets):
    """sheets: [(name, rows)] -> path。第一个即 sheet0/active。"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, rows in sheets:
        ws = wb.create_sheet(title=name)
        for r in rows:
            ws.append(r)
    fd, p = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(p); wb.close()
    return p


_PHDR = ["序号", "材料编号", "材料名称", "规格", "单位", "采购数量", "批次号"]
_DHDR = ["批次号", "属性", "零部件代码", "零部件名称", "数量", "供应商代码", "供应商名称"]


class TestStructuralVariants(unittest.TestCase):
    """标题行/空行/合计行/多子表等结构变体:应读对数据区,不漏不串。"""

    def test_purchase_title_blank_total(self):
        # 标题行 + 表头 + 数据 + 空行 + 数据 + 合计行
        from core import purchase_core as pc
        rows = [["某公司对账单", None, None, None, None, None, None], _PHDR,
                [1, "A001", "货A", "S1", "个", 10, "B1"],
                [2, "A002", "货B", "S2", "个", 20, "B1"],
                [None, None, None, None, None, None, None],
                [3, "A003", "货C", "S3", "个", 30, "B2"],
                ["合计", None, None, None, None, 60, None]]
        got, _ = pc.load_rows(_mkbook([("对账单", rows)]))
        self.assertEqual([r["no"] for r in got], ["A001", "A002", "A003"])

    def test_purchase_multisheet_cover_skipped(self):
        # sheet0 是封面/说明, 真实数据在 sheet1 -> 自动扫到
        from core import purchase_core as pc
        data = [_PHDR, [1, "A001", "货A", "S1", "个", 10, "B1"]]
        p = _mkbook([("说明", [["本表为对账说明", None]]), ("对账数据", data)])
        got, lay = pc.load_rows(p)
        self.assertEqual(lay["sheet"], "对账数据")
        self.assertEqual(len(got), 1)

    def test_purchase_explicit_sheet_respected(self):
        # 显式指定子表时, 不因 sheet0 有效就走兜底
        from core import purchase_core as pc
        p = _mkbook([("第一表", [_PHDR, [1, "A001", "货A", "S1", "个", 10, "B1"]]),
                     ("第二表", [_PHDR, [1, "Z999", "货Z", "S1", "个", 5, "B1"]])])
        got, lay = pc.load_rows(p, sheet="第二表")
        self.assertEqual(lay["sheet"], "第二表")
        self.assertEqual(got[0]["no"], "Z999")

    def test_delivery_blank_and_total_filtered(self):
        from core import delivery_core as dc
        rows = [_DHDR,
                ["B1", "自制", "P001", "件A", 10, "S01", "供A"],
                [None, None, None, None, None, None, None],
                ["B1", "外购", "P002", "件B", 20, "S02", "供B"],
                ["合计", None, None, None, 30, None, None]]
        got, _ = dc.load_sheet(_mkbook([("送货", rows)]))
        self.assertEqual([r["code"] for r in got], ["P001", "P002"])

    def test_delivery_multisheet_cover_skipped(self):
        from core import delivery_core as dc
        p = _mkbook([("封面", [["送货说明", None]]),
                     ("明细", [_DHDR, ["B1", "自制", "P001", "件A", 10, "S01", "供A"]])])
        got, lay = dc.load_sheet(p)
        self.assertEqual(lay["sheet"], "明细")
        self.assertEqual(len(got), 1)

    def test_attendance_title_blank_total(self):
        from core import attendance_core as ac
        rows = [["每日统计表", None, None, None],
                ["姓名", "日期", "上班1打卡时间", "下班1打卡时间"],
                ["张三", D(2026, 5, 1), "08:00", "17:00"],
                [None, None, None, None],
                ["张三", D(2026, 5, 2), "08:05", "17:10"],
                ["合计", "合计", None, None]]
        data = ac.load_source(_mkbook([("每日统计表", rows)]))
        self.assertEqual(len(data), 2)

    def test_attendance_merged_group_header(self):
        # 第1行分组(合并), 第2行真表头 -> 仍能定位并读数
        from core import attendance_core as ac
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "每日统计表"
        ws.append(["基本信息", None, None, "打卡信息", None])
        ws.append(["姓名", "工号", "日期", "上班1打卡时间", "下班1打卡时间"])
        ws.append(["张三", "-", D(2026, 5, 1), "08:00", "17:00"])
        ws.append(["张三", "-", D(2026, 5, 2), "08:05", "17:10"])
        ws.merge_cells("A1:C1"); ws.merge_cells("D1:E1")
        fd, p = tempfile.mkstemp(suffix=".xlsx"); os.close(fd)
        wb.save(p); wb.close()
        self.assertEqual(len(ac.load_source(p)), 2)


class TestDirtyContent(unittest.TestCase):
    """全角/千分位/零宽/全角冒号:真实数据中高频出现的"肉眼正确但机器错"的脏字符。"""

    # ---- 数字归一 ----
    def test_fullwidth_digits_and_period(self):
        from core.common_core import to_num
        self.assertAlmostEqual(to_num("１０"), 10.0)
        self.assertAlmostEqual(to_num("８．５"), 8.5)

    def test_thousands_separator(self):
        from core.common_core import to_num
        self.assertAlmostEqual(to_num("1,234"), 1234.0)
        self.assertAlmostEqual(to_num("1,234.5"), 1234.5)

    def test_zero_width_in_number(self):
        from core.common_core import to_num
        self.assertAlmostEqual(to_num("1​0"), 10.0)

    def test_units_still_rejected(self):
        # 单位粘连:保守处理,返回 None 不猜
        from core.common_core import to_num
        self.assertIsNone(to_num("10个"))
        self.assertIsNone(to_num("8.5小时"))

    def test_placeholders_still_none(self):
        from core.common_core import to_num
        self.assertIsNone(to_num("—"))
        self.assertIsNone(to_num("N/A"))
        self.assertIsNone(to_num("#N/A"))

    def test_pivot_num_fullwidth_and_thousands(self):
        from core.pivot_core import _num
        self.assertEqual(_num("８．５"), 8.5)
        self.assertEqual(_num("1,234"), 1234)

    # ---- 时间解析 ----
    def test_fullwidth_colon_time(self):
        from core.common_core import parse_time
        t = parse_time("08：30")   # 全角冒号,中文输入法常见
        self.assertEqual(t, datetime.time(8, 30))

    def test_half_colon_time_unchanged(self):
        from core.common_core import parse_time
        self.assertEqual(parse_time("08:30"), datetime.time(8, 30))
        self.assertEqual(parse_time("08:30:00"), datetime.time(8, 30))

    # ---- 编码归一 ----
    def test_zero_width_in_code(self):
        from core.delivery_core import norm_code
        from core.purchase_core import norm_no
        self.assertEqual(norm_code("A​1"), "A1")
        self.assertEqual(norm_no("A​1"), "A1")

    def test_fullwidth_letters_in_code(self):
        from core.purchase_core import norm_no
        from core.delivery_core import norm_code
        self.assertEqual(norm_no("ＡＢ1"), "AB1")
        self.assertEqual(norm_code("ＡＢ1"), "AB1")

    def test_leading_zeros_preserved_in_norm_code(self):
        # delivery 保留前导零,norm_code 不应丢
        from core.delivery_core import norm_code
        self.assertEqual(norm_code("00123"), "00123")

    def test_leading_zeros_stripped_in_norm_no(self):
        # purchase 去前导零
        from core.purchase_core import norm_no
        self.assertEqual(norm_no("00123"), "123")


class TestRobustness(unittest.TestCase):
    """病态输入不得裸崩;输出不得静默覆盖上一次结果。"""

    def test_corrupt_file_clear_error(self):
        # 损坏/伪装的 xlsx -> 清晰 ValueError, 不是底层 BadZipFile
        from core import purchase_core as pc
        import zipfile
        fd, p = tempfile.mkstemp(suffix=".xlsx"); os.close(fd)
        with open(p, "w", encoding="utf-8") as f:
            f.write("这不是Excel文件")
        with self.assertRaises(ValueError):
            pc.load_rows(p)
        # 确保不是 BadZipFile 逃逸出来
        try:
            pc.load_rows(p)
        except ValueError:
            pass
        except zipfile.BadZipFile:
            self.fail("BadZipFile 泄漏到上层, 应转成 ValueError")

    def test_header_only_no_crash(self):
        # 只有表头没数据 -> 0 行, 不崩
        from core import purchase_core as pc
        got, _ = pc.load_rows(_mkbook([("s", [_PHDR])]))
        self.assertEqual(got, [])

    def test_empty_sheet_clear_error(self):
        from core import purchase_core as pc
        with self.assertRaises(ValueError):
            pc.load_rows(_mkbook([("s", [])]))

    def test_duplicate_headers_no_crash(self):
        # 重复列名不应导致崩溃
        from core import purchase_core as pc
        rows = [_PHDR + ["采购数量"], [1, "A1", "货", "S", "个", 10, "B", 99]]
        got, _ = pc.load_rows(_mkbook([("s", rows)]))
        self.assertEqual(len(got), 1)

    # ---- 输出防覆盖 ----
    def test_unique_path_avoids_overwrite(self):
        from core.common_core import unique_path
        d = tempfile.mkdtemp()
        p1 = unique_path(os.path.join(d, "结果.xlsx"))
        self.assertEqual(os.path.basename(p1), "结果.xlsx")
        open(p1, "w").close()
        p2 = unique_path(os.path.join(d, "结果.xlsx"))
        self.assertEqual(os.path.basename(p2), "结果 (2).xlsx")
        open(p2, "w").close()
        p3 = unique_path(os.path.join(d, "结果.xlsx"))
        self.assertEqual(os.path.basename(p3), "结果 (3).xlsx")

    def test_out_path_dedup_same_minute(self):
        # 同一时间戳两次 out_path: 第二次自动避让
        from core.common_core import out_path
        d = tempfile.mkdtemp()
        a = out_path(d, "对账单", "_已填写", ts="20260501_1030")
        open(a, "w").close()
        b = out_path(d, "对账单", "_已填写", ts="20260501_1030")
        self.assertNotEqual(a, b)
        self.assertTrue(os.path.basename(b).endswith("(2).xlsx"))


class TestPivotVariants(unittest.TestCase):
    """透视:最终数量列别名 + 全角空格表头。锚点为"材料编号"所在行。"""

    def _final_col(self, hdr):
        from core import pivot_core as pc
        wb, ws = _ws([hdr, ["V1", "M001", "螺栓", "M8", 10, "个", 8]])
        try:
            blocks = pc.find_all_blocks(ws)
            self.assertTrue(blocks, "未锚定到数据块:%r" % (hdr,))
            return blocks[0]["cols"][6]        # cols[6]=final 列(1-based),0=缺失
        finally:
            wb.close()

    def test_jihua_shuliang(self):             # 计划数量(本轮新增别名)
        self.assertEqual(self._final_col(
            ["版本", "材料编号", "材料名称", "规格", "数量", "单位", "计划数量"]), 7)

    def test_jihua_caigou_shuliang(self):      # 计划采购数量
        self.assertEqual(self._final_col(
            ["版本", "材料编号", "材料名称", "规格", "数量", "单位", "计划采购数量"]), 7)

    def test_fullwidth_space_header(self):     # "最终　采购数量"含全角空格,_norm 应清掉
        self.assertEqual(self._final_col(
            ["版本", "材料编号", "材料名称", "规格", "数量", "单位", "最终　采购数量"]), 7)


class TestArrivalVariants(unittest.TestCase):
    """到货:需求列认"计划数量"、剩余列认"缺料"。返回 dict(demand/remain/…)。"""

    def _cols(self, hdr):
        from core import arrival_core as ac
        wb, ws = _ws([hdr, ["M001", "螺栓", 100, 20]])
        try:
            return ac.locate_columns(ws)
        finally:
            wb.close()

    def test_demand_jihua_shuliang(self):      # 计划数量 -> demand
        cols = self._cols(["物料编码", "物料名称", "计划数量", "剩余未收数"])
        self.assertEqual(cols["demand"], 3)

    def test_remain_quefeng(self):             # 缺料 -> remain
        cols = self._cols(["物料编码", "物料名称", "需求数", "缺料"])
        self.assertEqual(cols["remain"], 4)


class TestCompareVariants(unittest.TestCase):
    """比对:填满的标题横幅不得冒名顶替真表头(_detect_header_row)。"""

    def test_banner_row_not_chosen_as_header(self):
        from core import compare_core as cc
        # 第1行是填满的整句标题横幅(长文本、重复填充),第2行才是真表头
        p = _mk([["某某公司二〇二六年五月采购对账明细表（内部资料请勿外传）"] * 5,
                 ["物料编码", "物料名称", "数量", "单价", "金额"],
                 ["M001", "螺栓", 10, 2, 20]])
        headers, rows = cc.read_table(p)
        self.assertIn("物料编码", headers)
        self.assertEqual(len(rows), 1)         # 只有 1 条数据行,横幅未被当数据
        self.assertEqual(rows[0]["物料编码"], "M001")

    def test_read_headers_matches_read_table(self):
        from core import compare_core as cc
        p = _mk([["标题横幅占满一整行的长长长长长长长说明文字"] * 4,
                 ["编码", "名称", "数量", "单价"],
                 ["A1", "件", 1, 9]])
        self.assertEqual(cc.read_headers(p), cc.read_table(p)[0])
