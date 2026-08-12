# -*- coding: utf-8 -*-
"""到料明细多子表选择回归测试(合成簿,可移植)。

钉死:含数据的子表即使不是活动表/第一张,也能被 _pick_data_ws 选中,
不再静默读到空的干扰子表(如 'Sheet2')而把全部物料误判为已收。
"""
import os
import tempfile
import unittest
import warnings

import openpyxl

from core import arrival_core as A

warnings.filterwarnings("ignore", message="Workbook contains no default style")

# locate_columns 需要的表头(编码/需求/剩余未收 为关键列)
HDR = ["物料编码", "物料名称", "供应商信息", "需求数", "剩余未收数"]
# 剩余未收数须为非零数值才算"未收"(extract 视非数值/0 为已收),故用真数字
DATA = ["8892602000", "右前踏板", "北京丰达", 360, 12]


class _Tmp(unittest.TestCase):
    """提供隔离主数据库和可控制活动工作表的合成文件工厂。"""

    def setUp(self):
        """创建临时目录并把主数据学习结果隔离在其中。"""

        self._tmp = tempfile.mkdtemp(prefix="fyt_arr_")
        self._old_catalog = os.environ.get("FYT_CATALOG_PATH")
        os.environ["FYT_CATALOG_PATH"] = os.path.join(self._tmp, "catalog.json")

    def tearDown(self):
        """恢复主数据库环境并删除合成到料表。"""

        import shutil
        if self._old_catalog is None:
            os.environ.pop("FYT_CATALOG_PATH", None)
        else:
            os.environ["FYT_CATALOG_PATH"] = self._old_catalog
        shutil.rmtree(self._tmp, ignore_errors=True)

    def mk(self, sheets, active_idx=0):
        """sheets: [(名, [行...]) ...]; active_idx 指定活动表。"""
        p = os.path.join(self._tmp, "plan.xlsx")
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for name, rows in sheets:
            ws = wb.create_sheet(title=name)
            for row in rows:
                ws.append(row)
        wb.active = active_idx
        wb.save(p)
        return p


class TestPickDataWs(_Tmp):
    """验证到料业务不会盲目信任错误活动表。"""

    def _pick(self, sheets, active_idx=0):
        """生成工作簿并返回选中表名及纠偏日志。"""

        p = self.mk(sheets, active_idx)
        wb = openpyxl.load_workbook(p, data_only=True)
        logs = []
        ws = A._pick_data_ws(wb, log=logs.append)
        name = ws.title
        wb.close()
        return name, logs

    def test_single_sheet_unchanged(self):
        """单工作表直接使用且不产生多余纠偏日志。"""

        name, logs = self._pick([("零件到货计划", [HDR, DATA])])
        self.assertEqual(name, "零件到货计划")
        self.assertEqual([l for l in logs if "子表" in l], [])   # 单表无噪音

    def test_active_valid_is_used(self):
        """活动表结构有效时应保持用户选择。"""

        # 活动表(idx=1)有效 -> 用它
        name, logs = self._pick(
            [("Sheet2", [["空"], []]), ("零件到货计划", [HDR, DATA])],
            active_idx=1)
        self.assertEqual(name, "零件到货计划")

    def test_wrong_active_auto_corrects(self):
        """活动表无有效字段时应改读其他有效子表并记录原因。"""

        # 活动表=空 Sheet2(idx=0) -> 应自动改读有数据的子表
        name, logs = self._pick(
            [("Sheet2", [["空表"], ["x"]]), ("零件到货计划", [HDR, DATA])],
            active_idx=0)
        self.assertEqual(name, "零件到货计划")
        self.assertTrue(any("改读" in l for l in logs), "应记录纠偏")

    def test_extract_reads_correct_sheet(self):
        """端到端提取在错误活动表场景仍应读到真实未收物料。"""

        # 端到端:错误活动表下 extract_unreceived 仍能取到那 1 行未收料
        p = self.mk([("Sheet2", [["空表"], ["x"]]),
                     ("零件到货计划", [HDR, DATA])], active_idx=0)
        rows = A.extract_unreceived(p)
        self.assertEqual(len(rows), 1)         # 若误读空表会得 0
        self.assertEqual(rows[0][0], "8892602000")

    def test_full_plan_counts_hidden_rows_and_supports_manual_total(self):
        """完整源表应自动统计总类数，并把隐藏行中的非零未收数纳入明细。"""

        path = self.mk([("零件到货计划", [
            HDR,
            ["A-01", "已到物料", "供应商甲", 10, 0],
            ["B-02", "隐藏缺料", "供应商乙", 12, 2],
            ["C-03", "负数缺料", "供应商丙", 8, -1],
        ])])
        workbook = openpyxl.load_workbook(path)
        workbook.active.row_dimensions[3].hidden = True
        workbook.save(path)
        workbook.close()

        inspection = A.inspect_plan(path)
        self.assertEqual(inspection["total"], 3)
        self.assertEqual(inspection["hidden"], 1)
        self.assertEqual(
            [item[0] for item in inspection["materials"]],
            ["B-02", "C-03"],
        )

        automatic, _ = A.build_batches([{
            "path": path, "batch_no": "AUTO", "remark": "", "include": True,
        }], A.DEFAULT_TOP_LABEL)
        self.assertEqual(automatic[0]["total"], 3)
        self.assertEqual(automatic[0]["auto_total"], 3)

        overridden, _ = A.build_batches([{
            "path": path, "batch_no": "MANUAL", "total": 5,
            "remark": "人工核对", "include": True,
        }], A.DEFAULT_TOP_LABEL)
        self.assertEqual(overridden[0]["total"], 5)
        self.assertEqual(overridden[0]["auto_total"], 3)

    def test_result_batches_keep_missing_material_and_quantity_gap(self):
        """结构化批次结果必须保留物料明细和实收到需求的数量缺口。"""

        batches = [{
            "batch_no": "26035-01",
            "total": 10,
            "materials": [["A-01", "固定螺栓", "供应商甲", 12, 3]],
        }]
        details = A.build_result_batches(batches, [("26035-01", 1, 9, 10)])
        self.assertEqual(details[0]["missing_count"], 1)
        material = details[0]["missing_materials"][0]
        self.assertEqual(material["material_code"], "A-01")
        self.assertEqual(material["received_quantity"], 9)
        self.assertEqual(material["shortage_quantity"], 3)

    def test_finished_report_reads_embedded_batch_metrics(self):
        """直接上传成品日报时应从各横向批次块解析总类数、到货数和缺料。"""

        path = os.path.join(self._tmp, "20260805每日主料到料明细.xlsx")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        # 每个批次占七列，模拟现行成品每日主料到料明细的横向排版。
        for start, batch_no, total, arrived, missing, code in (
            (3, "GMMYR26163A", 474, 473, 1, "A-01"),
            (12, "GMIDR26178A", 478, 477, 1, "B-02"),
        ):
            sheet.cell(1, start, "截止16点的数据")
            sheet.cell(2, start, batch_no)
            sheet.cell(3, start, "主料总共类")
            sheet.cell(3, start + 2, total)
            sheet.cell(4, start, "到货数量")
            sheet.cell(4, start + 2, arrived)
            sheet.cell(5, start, "差异")
            sheet.cell(5, start + 2, missing)
            for offset, title in enumerate(("序号", "物料编码", "物料名称", "供应商信息", "需求数", "剩余未收数", "备注")):
                sheet.cell(7, start + offset, title)
            sheet.cell(8, start, 1)
            sheet.cell(8, start + 1, code)
            sheet.cell(8, start + 2, "测试物料")
            sheet.cell(8, start + 3, "供应商甲")
            sheet.cell(8, start + 4, 10)
            sheet.cell(8, start + 5, -2)
        workbook.save(path)
        workbook.close()

        result = A.analyze_finished_report(path)
        self.assertEqual(result["report_date"], "2026-08-05")
        self.assertEqual(result["results"], [
            ("GMMYR26163A", 1, 473, 474),
            ("GMIDR26178A", 1, 477, 478),
        ])
        self.assertEqual(result["batches"][0]["missing_materials"][0]["shortage_quantity"], 2)
        self.assertEqual(result["batches"][0]["missing_materials"][0]["received_quantity"], 8)


if __name__ == "__main__":
    unittest.main()
