# -*- coding: utf-8 -*-
"""考勤月度归档：多表汇总、表头识别与月份推断。"""
from __future__ import annotations

import os
import tempfile
import unittest

import openpyxl

from core import attendance_archive_core


def _write_attendance(path, rows, headers=None):
    """生成可切换标准与别名表头的考勤来源表。"""

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "考勤"
    worksheet.append(headers or ["员工姓名", "考勤日期", "实际工作时间", "加班", "异常原因"])
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()  # 关闭工作簿释放文件


class AttendanceArchiveCoreTests(unittest.TestCase):
    """验证多车间考勤按月份、人员和日期合并后的归档口径。"""

    def setUp(self):
        """构造两份表头不同且人员日期有交叉的考勤文件。"""

        self.temp = tempfile.TemporaryDirectory(prefix="fyt_att_archive_")
        self.file_a = self.path("车间A_已填写.xlsx")
        self.file_b = self.path("车间B_已填写.xlsx")
        _write_attendance(self.file_a, [
            ["张三", "2026-08-03", 8, 0, ""],
            ["张三", "2026-08-04", 8, 2, "迟到 5 分钟"],
            ["李四", "2026-08-03", 7.5, 0, ""],
        ])
        _write_attendance(self.file_b, [
            ["张三", "2026-08-05", 8, 1, ""],
            ["李四", "2026-08-04", 8, 0, ""],
            ["王五", "2026-08-04", 8, 0, ""],
        ], headers=["姓名", "日期", "工时", "加班工时"])  # 别名表头验证识别

    def tearDown(self):
        """删除合成输入与月度归档输出。"""

        self.temp.cleanup()

    def path(self, name):
        """返回当前用例临时目录下路径。"""

        return os.path.join(self.temp.name, name)

    def test_archive_aggregates_persons_and_month(self):
        """归档应汇总出勤天数、工时、加班、异常并保留六条每日明细。"""

        out_dir = self.path("输出")
        result = attendance_archive_core.archive([self.file_a, self.file_b], out_dir=out_dir)
        self.assertEqual(result["month"], "2026-08")  # 月份从数据日期推断
        self.assertEqual(result["persons"], 3)  # 三个人员
        self.assertEqual(result["days"], 6)  # 六条明细
        self.assertTrue(os.path.isfile(result["path"]))  # 归档表落盘

        workbook = openpyxl.load_workbook(result["path"], data_only=True)
        summary = workbook["月度汇总"]
        rows = {str(row[0]): row for row in summary.iter_rows(min_row=2, values_only=True) if row[0]}
        self.assertEqual(rows["张三"][1], 3)      # 出勤 3 天
        self.assertEqual(rows["张三"][2], 24.0)   # 总工时
        self.assertEqual(rows["张三"][3], 3.0)    # 加班
        self.assertEqual(rows["张三"][4], 1)      # 异常 1 次
        self.assertEqual(rows["李四"][1], 2)
        self.assertEqual(rows["王五"][1], 1)
        detail = workbook["每日明细"]
        detail_rows = [row for row in detail.iter_rows(min_row=2, values_only=True) if row[0]]  # 跳过空行
        self.assertEqual(len(detail_rows), 6)  # 六条明细全保留
        workbook.close()

    def test_archive_requires_valid_inputs(self):
        """空输入或无考勤字段的工作簿应明确拒绝。"""

        with self.assertRaises(ValueError):
            attendance_archive_core.archive([])  # 空输入拒绝
        bad = self.path("无表头.xlsx")
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.append(["材料编号", "数量"])
        worksheet.append(["JBC001", 3])
        workbook.save(bad)
        workbook.close()
        with self.assertRaises(ValueError) as context:
            attendance_archive_core.archive([bad], out_dir=self.path("输出2"))
        self.assertIn("识别", str(context.exception))  # 错误信息指出表头无法识别


if __name__ == "__main__":
    unittest.main()
