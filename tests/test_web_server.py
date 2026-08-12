"""局域网 Web 服务的认证、任务和文件权限回归测试。"""

from __future__ import annotations

import json
import hashlib
import os
from io import BytesIO
import socket
import tempfile
import http.server
import threading
import time
import unittest
import urllib.error
import urllib.parse
import urllib.request
import uuid
import zipfile
from datetime import datetime, timedelta, timezone
from pathlib import Path
from unittest import mock

from openpyxl import Workbook, load_workbook
from PIL import Image

import web_server


class WebServerTests(unittest.TestCase):
    """使用临时 SQLite 与 HTTP 端口验证 Web 服务主链路。"""

    def setUp(self):
        """为每个用例创建隔离数据根、首次管理员库和随机本机 HTTP 端口。"""

        self.temp = tempfile.TemporaryDirectory()
        self.original = (web_server.DATA_ROOT, web_server.DB_PATH, web_server.STATIC_ROOT)
        # 同时替换数据、数据库和静态目录，任何上传、输出或备份都只能落入临时目录。
        web_server.DATA_ROOT = Path(self.temp.name)
        web_server.DB_PATH = web_server.DATA_ROOT / "accounts.sqlite3"
        web_server.STATIC_ROOT = web_server.DATA_ROOT / "dist"
        os.environ["FYT_ADMIN_PASSWORD"] = "admin123456"
        web_server.init_db()
        # 端口 0 由系统分配空闲端口，避免并行或残留测试进程造成固定端口冲突。
        self.server = web_server.ThreadingHTTPServer(("127.0.0.1", 0), web_server.Handler)
        self.thread = threading.Thread(target=self.server.serve_forever, daemon=True)
        self.thread.start()
        self.base = f"http://127.0.0.1:{self.server.server_port}"
        self.admin = self.call("/api/auth/login", {"username": "admin", "password": "admin123456"})[1]["token"]

    def tearDown(self):
        """停止测试服务器、恢复模块级路径并删除全部临时运行数据。"""

        os.environ.pop("FYT_ADMIN_PASSWORD", None)
        self.server.shutdown()
        self.server.server_close()
        self.thread.join(timeout=2)
        web_server.DATA_ROOT, web_server.DB_PATH, web_server.STATIC_ROOT = self.original
        self.temp.cleanup()

    def call(self, path, payload=None, token="", raw=None, headers=None, method=None):
        """发送测试 API 请求，并把正常及 HTTP 错误响应统一解析为状态码和正文。"""

        # raw 用于文件上传等二进制请求；payload 则按服务端默认 UTF-8 JSON 编码。
        data = raw if raw is not None else None if payload is None else json.dumps(payload, ensure_ascii=False).encode()
        request = urllib.request.Request(self.base + path, data=data, method=method or ("POST" if data is not None else "GET"))
        request.add_header("Content-Type", "application/json" if raw is None else "application/octet-stream")
        if token:
            request.add_header("X-Session-Token", token)
        for key, value in (headers or {}).items():
            request.add_header(key, value)
        try:
            with urllib.request.urlopen(request, timeout=10) as response:
                body = response.read()
                return response.status, json.loads(body) if "json" in (response.headers.get("Content-Type") or "") else body
        except urllib.error.HTTPError as error:
            # 权限拒绝和参数错误是待断言的正常测试结果，不应由 urllib 提前中断用例。
            body = error.read()
            return error.code, json.loads(body)

    def wait_job(self, job_id):
        """短间隔轮询持久化任务，超时意味着后台任务生命周期发生回归。"""

        # 八十次乘五十毫秒覆盖合成小文件任务，同时让失败测试在约四秒内给出结果。
        for _ in range(80):
            status, payload = self.call(f"/api/jobs/{job_id}", token=self.admin)
            self.assertEqual(status, 200)
            job = payload["job"]
            if job["status"] not in ("queued", "running"):
                return job
            time.sleep(0.05)
        self.fail("任务未在测试时间内结束")

    def test_requires_login(self):
        """健康检查允许匿名访问，其余业务总览必须拒绝未登录请求。"""

        status, health = self.call("/api/health")
        self.assertEqual(status, 200)
        self.assertEqual(health["version"], web_server.VERSION)
        status, payload = self.call("/api/overview")
        self.assertEqual(status, 401)
        self.assertEqual(payload["error"], "请先登录")

    def test_http_context_json_and_static_cache_policy(self):
        """验证损坏 JSON、Cookie 会话、SPA 回退和分层静态缓存策略。"""

        status, payload = self.call(
            "/api/auth/login",
            token="",
            raw=b"[",
            headers={"Content-Length": "1"},
        )
        self.assertEqual(status, 400)
        self.assertEqual(payload["error"], "请求内容不是有效 JSON")

        login_data = json.dumps({
            "username": "admin",
            "password": "admin123456",
        }).encode("utf-8")
        login_request = urllib.request.Request(
            self.base + "/api/auth/login",
            data=login_data,
            headers={"Content-Type": "application/json"},
        )
        with urllib.request.urlopen(login_request, timeout=10) as response:
            cookie = str(response.headers["Set-Cookie"]).split(";", 1)[0]
        me_request = urllib.request.Request(
            self.base + "/api/auth/me",
            headers={"Cookie": cookie},
        )
        with urllib.request.urlopen(me_request, timeout=10) as response:
            self.assertEqual(json.loads(response.read())["user"]["username"], "admin")

        status, missing = self.call("/")
        self.assertEqual(status, 404)
        self.assertIn("前端尚未构建", missing["error"])

        assets = web_server.STATIC_ROOT / "assets"
        assets.mkdir(parents=True)
        (web_server.STATIC_ROOT / "index.html").write_text(
            "<main>峰运通</main>", encoding="utf-8",
        )
        (assets / "app-123.js").write_text("window.fyt = true", encoding="utf-8")
        (web_server.STATIC_ROOT / "logo.txt").write_text("FYT", encoding="utf-8")

        with urllib.request.urlopen(self.base + "/", timeout=10) as response:
            self.assertEqual(
                response.headers["Cache-Control"],
                "no-store, must-revalidate, no-transform",
            )
            self.assertIn("峰运通", response.read().decode("utf-8"))
        with urllib.request.urlopen(self.base + "/assets/app-123.js", timeout=10) as response:
            self.assertEqual(
                response.headers["Cache-Control"],
                "public, max-age=31536000, immutable",
            )
        with urllib.request.urlopen(self.base + "/logo.txt", timeout=10) as response:
            self.assertEqual(response.headers["Cache-Control"], "public, max-age=604800")
        with urllib.request.urlopen(self.base + "/client/route", timeout=10) as response:
            self.assertEqual(
                response.headers["Cache-Control"],
                "no-store, must-revalidate, no-transform",
            )
            self.assertIn("峰运通", response.read().decode("utf-8"))

    def test_text_task_and_upload_download(self):
        """覆盖任务创建、二进制上传、重命名、版本下载、预览和工作台统计主链路。"""

        status, created = self.call("/api/jobs", {
            "action": "text.transform", "title": "测试文本", "payload": {
                "text": "乙\n甲\n乙", "operation": "dedup",
            },
        }, token=self.admin)
        self.assertEqual(status, 202)
        job = self.wait_job(created["job_id"])
        self.assertEqual(job["status"], "completed")
        self.assertEqual(job["result"]["text"], "乙\n甲")

        upload_query = urllib.parse.urlencode({"name": "原文件.txt", "group": "test-group"})
        status, uploaded = self.call(
            f"/api/files/upload?{upload_query}",
            token=self.admin, raw=b"sample\n", headers={"Content-Length": "7"},
        )
        self.assertEqual(status, 201)
        status, created = self.call("/api/jobs", {
            "action": "rename.apply", "title": "测试重命名", "payload": {
                "paths": [uploaded["handle"]], "rule": {"prefix": "新-"},
            },
        }, token=self.admin)
        self.assertEqual(status, 202)
        job = self.wait_job(created["job_id"])
        self.assertEqual(job["status"], "completed")
        self.assertEqual(len(job["files"]), 1)
        request = urllib.request.Request(self.base + job["files"][0]["url"], headers={"X-Session-Token": self.admin})
        with urllib.request.urlopen(request, timeout=10) as response:
            self.assertEqual(response.read(), b"sample\n")
            self.assertEqual(response.headers["Cache-Control"], "no-store")
            self.assertTrue(response.headers["Content-Disposition"].startswith("attachment;"))
        # 路由表必须优先识别历史版本文件，不能落入普通任务文件下载。
        version_url = job["versions"][0]["files"][0]["url"]
        request = urllib.request.Request(self.base + version_url, headers={"X-Session-Token": self.admin})
        with urllib.request.urlopen(request, timeout=10) as response:
            self.assertEqual(response.read(), b"sample\n")
        status, preview = self.call(job["files"][0]["url"] + "/preview", token=self.admin)
        self.assertEqual(status, 200)
        self.assertEqual(preview["rows"], [["sample"]])

        status, board = self.call("/api/dashboard", token=self.admin)
        self.assertEqual(status, 200)
        self.assertEqual(board["metrics"]["completed_jobs"], 2)
        self.assertEqual(len(board["trend"]), 7)
        self.assertIn("text", {item["key"] for item in board["feature_usage"]})
        self.assertGreaterEqual(len(board["recent_files"]), 1)

    def test_arrival_scan_and_manual_total_override(self):
        """Web 到料扫描应统计完整源表，正式任务必须采用人工覆盖后的总类数。"""

        stream = BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "订单 TEST2601 批次"
        sheet.append(["物料编码", "物料名称", "供应商信息", "需求数", "剩余未收数"])
        sheet.append(["A-01", "已到物料", "供应商甲", 10, 0])
        sheet.append(["B-02", "隐藏缺料", "供应商乙", 12, 2])
        sheet.append(["C-03", "负数缺料", "供应商丙", 8, -1])
        sheet.row_dimensions[4].hidden = True
        workbook.save(stream)
        workbook.close()
        content = stream.getvalue()

        query = urllib.parse.urlencode({"name": "送货计划.xlsx", "group": "arrival-scan"})
        status, uploaded = self.call(
            f"/api/files/upload?{query}", token=self.admin, raw=content,
            headers={"Content-Length": str(len(content))},
        )
        self.assertEqual(status, 201)

        status, scanned = self.call(
            "/api/arrival/scan", {"paths": [uploaded["handle"]]}, token=self.admin,
        )
        self.assertEqual(status, 200)
        self.assertEqual(len(scanned["rows"]), 1)
        row = scanned["rows"][0]
        self.assertEqual(row["batch_no"], "TEST2601")
        self.assertEqual(row["total"], 3)
        self.assertEqual(row["auto_total"], 3)
        self.assertEqual(row["missing_count"], 2)
        self.assertTrue(str(row["path"]).startswith("upload:"))
        self.assertNotIn(str(self.temp.name), json.dumps(row, ensure_ascii=False))

        row["total"] = 5
        row["remark"] = "人工确认总类数"
        status, created = self.call("/api/jobs", {
            "action": "web.arrival",
            "title": "到料人工参数回归",
            "payload": {"rows": [row], "top_label": "截止 16 点"},
        }, token=self.admin)
        self.assertEqual(status, 202)
        job = self.wait_job(created["job_id"])
        self.assertEqual(job["status"], "completed", job.get("error"))
        result = job["result"]["result"]
        self.assertEqual(result["results"][0][3], 5)
        self.assertEqual(result["batches"][0]["total_count"], 5)
        self.assertEqual(result["batches"][0]["missing_count"], 2)

    def test_daily_report_admin_scope_result_projection_and_export(self):
        """日清看板只能由管理员读取，并只汇总当天且属于管理员视角的业务投影。"""

        self.assertEqual(self.call("/api/auth/register", {
            "username": "daily_member", "display_name": "计划员乙", "password": "password123",
        })[0], 201)
        member = next(
            item for item in self.call("/api/admin/users", token=self.admin)[1]["users"]
            if item["username"] == "daily_member"
        )
        self.assertEqual(
            self.call(f"/api/admin/users/{member['id']}/approve", {}, token=self.admin)[0],
            200,
        )
        member_token = self.call("/api/auth/login", {
            "username": "daily_member", "password": "password123",
        })[1]["token"]

        report_date = web_server.business_today().isoformat()
        start, _ = web_server.business_day_bounds(web_server.business_today())
        completed_at = (datetime.fromisoformat(start) + timedelta(hours=2)).isoformat(timespec="seconds")
        previous_at = (datetime.fromisoformat(start) - timedelta(seconds=1)).isoformat(timespec="seconds")
        created_at = completed_at
        rows = [
            ("daily-arrival-admin", 1, "管理员到料", {
                "results": [["26035-01", 2, 8, 10]],
                "batches": [{
                    "batch_no": "26035-01", "missing_count": 2,
                    "arrived_count": 8, "total_count": 10,
                    "missing_materials": [{
                        "material_code": "A-01", "material_name": "固定螺栓",
                        "supplier": "供应商甲", "demand_quantity": 12,
                        "received_quantity": 9, "shortage_quantity": 3,
                    }],
                }],
            }, completed_at),
            ("daily-arrival-member", member["id"], "成员到料", {
                "results": [["26035-02", 0, 5, 5]],
            }, completed_at),
            ("daily-arrival-old", 1, "前一日到料", {
                "results": [["OLD", 1, 1, 2]],
            }, previous_at),
        ]
        with web_server.DB_LOCK, web_server.db() as connection:
            for job_id, user_id, title, result_payload, updated_at in rows:
                connection.execute(
                    "INSERT INTO web_jobs(id, user_id, action, title, status, progress, logs, result, "
                    "files, cancelled, payload, created_at, updated_at) "
                    "VALUES (?, ?, 'web.arrival', ?, 'completed', 100, '[]', ?, '[]', 0, '{}', ?, ?)",
                    (job_id, user_id, title, json.dumps({"result": result_payload, "logs": [], "task_id": job_id, "out_dir": "到料明细"}, ensure_ascii=False), created_at, updated_at),
                )
            connection.execute(
                "INSERT INTO workshop_issues(id, user_id, issue_date, cause, primary_owner, "
                "secondary_owner, notes, status, created_at, updated_at) "
                "VALUES ('daily-issue', ?, ?, '工位防护罩松动', '张工', '李工', '下班前复核', "
                "'published', ?, ?)",
                (member["id"], report_date, completed_at, completed_at),
            )

        status, denied = self.call(
            f"/api/daily-report?date={report_date}", token=member_token,
        )
        self.assertEqual(status, 403)
        self.assertIn("管理员", denied["error"])

        status, snapshot = self.call(
            f"/api/daily-report?date={report_date}", token=self.admin,
        )
        self.assertEqual(status, 200)
        self.assertEqual(snapshot["arrival"]["job_count"], 2)
        self.assertEqual(snapshot["arrival"]["batch_count"], 2)
        self.assertEqual(snapshot["arrival"]["total_categories"], 15)
        self.assertEqual(snapshot["arrival"]["completion_rate"], 86.7)
        self.assertEqual(snapshot["arrival"]["missing_material_detail_count"], 1)
        self.assertEqual(snapshot["arrival"]["batches"][0]["missing_materials"][0]["material_code"], "A-01")
        self.assertEqual(snapshot["workshop"]["issue_count"], 1)
        self.assertEqual(snapshot["workshop"]["owner_distribution"][0], {"owner": "张工", "count": 1})
        self.assertNotIn(str(self.temp.name), json.dumps(snapshot, ensure_ascii=False))

        status, jobs = self.call("/api/jobs", token=self.admin)
        self.assertEqual(status, 200)
        projected = next(item for item in jobs["jobs"] if item["id"] == "daily-arrival-admin")
        self.assertEqual(projected["presentation"]["kind"], "arrival")
        self.assertEqual(projected["presentation"]["metrics"][1]["value"], "80.0%")
        self.assertEqual(projected["presentation"]["sections"][1]["rows"][0]["shortage_quantity"], "3")

        status, content = self.call(
            f"/api/daily-report/export?date={report_date}", token=self.admin,
        )
        self.assertEqual(status, 200)
        workbook = load_workbook(BytesIO(content), data_only=True)
        try:
            for sheet_name in (
                "日清概览", "每日到料", "未到物料", "安全检查日报", "现场问题",
                "月度生产订单台账", "订单缺件明细", "订单危包明细", "零星订单明细",
            ):
                self.assertIn(sheet_name, workbook.sheetnames)
            self.assertEqual(workbook["每日到料"]["A2"].value, "26035-01")
            self.assertEqual(workbook["未到物料"]["C2"].value, "A-01")
            self.assertEqual(workbook["未到物料"]["H2"].value, "3")
            self.assertEqual(workbook["现场问题"]["C2"].value, "工位防护罩松动")
        finally:
            workbook.close()

    def test_daily_report_manual_attendance_briefs_and_production_plan(self):
        """人工考勤、事项和生产计划应共同进入日清快照及导出报告。"""

        report_date = web_server.business_today().isoformat()
        status, created = self.call("/api/admin/daily-people", {
            "name": "林楠栗", "person_type": "participant", "unit": "现场负责人", "shift": "", "sort_order": 1, "active": True,
        }, token=self.admin)
        self.assertEqual(status, 201)
        person = created["person"]
        status, rejected = self.call("/api/admin/daily-people", {
            "name": "生产甲", "person_type": "production", "unit": "小件组", "shift": "白班", "sort_order": 1, "active": True,
        }, token=self.admin)
        self.assertEqual(status, 400)
        self.assertIn("按班组维护", rejected["error"])
        status, created_group = self.call("/api/admin/daily-production-groups", {
            "name": "小件组", "sort_order": 1, "active": True,
            "shifts": [
                {"name": "白班", "staffing_count": 14, "sort_order": 1, "active": True},
                {"name": "夜班", "staffing_count": 14, "sort_order": 2, "active": True},
            ],
        }, token=self.admin)
        self.assertEqual(status, 201)
        production_group = created_group["group"]
        self.assertEqual(production_group["staffing_count"], 28)
        self.assertEqual([item["name"] for item in production_group["shifts"]], ["白班", "夜班"])

        status, attendance = self.call(
            f"/api/admin/daily-attendance?date={report_date}", token=self.admin,
        )
        self.assertEqual(status, 200)
        self.assertEqual(len(attendance["attendance"]), 1)
        self.assertTrue(all(item["present"] for item in attendance["attendance"]))
        self.assertEqual(attendance["production_groups"][0]["group_name"], "小件组")
        self.assertEqual(len(attendance["production_groups"]), 2)
        shifts = {item["shift_name"]: item for item in attendance["production_groups"]}
        self.assertEqual(shifts["白班"]["staffing_count"], 14)

        status, saved = self.call("/api/admin/daily-attendance", {
            "date": report_date,
            "records": [
                {"person_id": person["id"], "present": True, "status": "present", "reason": ""},
            ],
            "production_groups": [
                {"shift_id": shifts["白班"]["shift_id"], "attendance_count": 17, "note": "3 人支援小件白班"},
                {"shift_id": shifts["夜班"]["shift_id"], "attendance_count": 14, "note": ""},
            ],
        }, token=self.admin)
        self.assertEqual(status, 200)
        self.assertEqual(saved["production_groups"][0]["attendance_count"], 17)
        self.assertEqual(saved["production_groups"][0]["difference"], -3)
        self.assertEqual(saved["production_groups"][0]["note"], "3 人支援小件白班")

        status, brief = self.call("/api/admin/daily-brief-items", {
            "report_date": report_date, "category": "escalation", "unit": "采购",
            "owner": "户子丹", "title": "关键物料未按计划到场", "description": "需要供应商升级处理",
            "due_date": report_date, "progress": "已联系供应商", "status": "in_progress",
        }, token=self.admin)
        self.assertEqual(status, 201)
        self.assertEqual(brief["item"]["category"], "escalation")
        self.assertEqual(brief["item"]["description"], "")
        self.assertEqual(brief["item"]["due_date"], "")
        self.assertEqual(brief["item"]["progress"], "")
        self.assertEqual(brief["item"]["status"], "open")
        status, _payload = self.call("/api/admin/daily-brief-items", {
            "report_date": report_date, "category": "safety", "unit": "安环",
            "owner": "张工", "title": "旧安全事项", "status": "open",
        }, token=self.admin)
        self.assertEqual(status, 400)
        status, todo = self.call("/api/admin/daily-brief-items", {
            "report_date": report_date, "category": "meeting_todo", "unit": "生产",
            "owner": "李工", "title": "确认次日排产", "description": "会后完成核对",
            "due_date": report_date, "progress": "待确认", "status": "in_progress",
        }, token=self.admin)
        self.assertEqual(status, 201)
        self.assertEqual(todo["item"]["due_date"], report_date)

        stream = BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "计划汇总"
        sheet.append(["班次", "计划", "实际", "差异"])
        sheet.append(["白班", 45, 41, -4])
        workbook.save(stream)
        workbook.close()
        plan_content = stream.getvalue()
        query = urllib.parse.urlencode({"name": "当天生产计划.xlsx", "date": report_date})
        status, uploaded = self.call(
            f"/api/admin/daily-production-plans?{query}", token=self.admin, raw=plan_content,
            headers={"Content-Length": str(len(plan_content))},
        )
        self.assertEqual(status, 201)
        self.assertEqual(uploaded["plan"]["summary"]["sheet_count"], 1)
        plan_id = uploaded["plan"]["id"]
        with web_server.DB_LOCK, web_server.db() as connection:
            plan_file = Path(connection.execute(
                "SELECT path FROM daily_production_plans WHERE id = ?", (plan_id,),
            ).fetchone()["path"])

        status, snapshot = self.call(f"/api/daily-report?date={report_date}", token=self.admin)
        self.assertEqual(status, 200)
        self.assertEqual(snapshot["attendance"]["present_count"], 32)
        self.assertEqual(snapshot["attendance"]["absent_count"], 0)
        self.assertEqual(snapshot["attendance"]["production_present_count"], 31)
        self.assertEqual(snapshot["attendance"]["production_staffing_count"], 28)
        self.assertEqual(snapshot["attendance"]["production_difference"], -3)
        self.assertEqual(snapshot["attendance"]["production_group_count"], 1)
        self.assertEqual(snapshot["attendance"]["production_shift_count"], 2)
        self.assertEqual(snapshot["attendance"]["production_groups"][0]["group_name"], "小件组")
        self.assertEqual(snapshot["brief_items"][0]["title"], "关键物料未按计划到场")
        self.assertEqual(snapshot["production_plans"][0]["original_name"], "当天生产计划.xlsx")

        status, exported = self.call(
            f"/api/daily-report/export?date={report_date}", token=self.admin,
        )
        self.assertEqual(status, 200)
        workbook = load_workbook(BytesIO(exported), data_only=True)
        try:
            self.assertIn("每日考勤", workbook.sheetnames)
            self.assertIn("生产出勤", workbook.sheetnames)
            self.assertEqual(workbook["生产出勤"]["A2"].value, "小件组")
            self.assertEqual(workbook["生产出勤"]["B2"].value, "白班")
            self.assertEqual(workbook["生产出勤"]["C2"].value, 14)
            self.assertEqual(workbook["生产出勤"]["D2"].value, 17)
            self.assertEqual(workbook["生产出勤"]["E2"].value, -3)
            self.assertEqual(workbook["生产出勤"]["C4"].value, 28)
            self.assertEqual(workbook["生产出勤"]["D4"].value, 31)
            self.assertEqual(workbook["生产出勤"]["E4"].value, -3)
            self.assertIn("重点事项与通报", workbook.sheetnames)
            self.assertIn("生产计划", workbook.sheetnames)
        finally:
            workbook.close()

        status, removed = self.call(
            f"/api/admin/daily-production-plans/{plan_id}", token=self.admin, method="DELETE",
        )
        self.assertEqual(status, 200)
        self.assertIn("回收站", removed["message"])
        self.assertFalse(plan_file.parent.exists())
        self.assertEqual(
            self.call(f"/api/daily-report?date={report_date}", token=self.admin)[1]["production_plans"],
            [],
        )
        trash = self.call("/api/admin/trash", token=self.admin)[1]["trash"]
        item = next(row for row in trash if row["kind"] == "daily_production_plan")
        self.assertEqual(item["label"], "当天生产计划.xlsx")
        self.assertEqual(self.call(
            f"/api/admin/trash/{item['id']}/restore", {}, token=self.admin,
        )[0], 200)
        restored_plan = self.call(
            f"/api/daily-report?date={report_date}", token=self.admin,
        )[1]["production_plans"][0]
        self.assertEqual(restored_plan["id"], plan_id)
        self.assertEqual(restored_plan["original_name"], "当天生产计划.xlsx")
        self.assertTrue(plan_file.is_file())

        self.assertEqual(self.call("/api/admin/daily-people")[0], 401)

    def test_daily_source_uploads_feed_arrival_and_safety_dashboard(self):
        """管理员直接上传成品到料与安全检查表后，看板应无需重跑业务模块即可展示。"""

        report_date = web_server.business_today().isoformat()

        arrival_stream = BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        sheet["C1"] = "截止16点的数据"
        sheet["C2"] = "TEST2608"
        sheet["C3"] = "主料总共类"
        sheet["E3"] = 3
        sheet["C4"] = "到货数量"
        sheet["E4"] = 2
        sheet["C5"] = "差异"
        sheet["E5"] = 1
        for column, title in enumerate(
            ["序号", "物料编码", "物料名称", "供应商信息", "需求数", "剩余未收数", "备注"],
            start=3,
        ):
            sheet.cell(7, column, title)
        for column, value in enumerate([1, "A-01", "测试物料", "供应商甲", 10, -2, ""], start=3):
            sheet.cell(8, column, value)
        workbook.save(arrival_stream)
        workbook.close()
        arrival_content = arrival_stream.getvalue()
        arrival_query = urllib.parse.urlencode({
            "kind": "arrival", "date": report_date,
            "name": f"{report_date.replace('-', '')}每日主料到料明细.xlsx",
        })
        status, arrival_upload = self.call(
            f"/api/admin/daily-source-uploads?{arrival_query}", token=self.admin,
            raw=arrival_content, headers={"Content-Length": str(len(arrival_content))},
        )
        self.assertEqual(status, 201)
        self.assertEqual(arrival_upload["upload"]["kind"], "arrival")

        safety_stream = BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "安全检查日报"
        sheet["H1"] = datetime.strptime(report_date, "%Y-%m-%d").date()
        sheet.append(["检查类别", "序号", "检查项目", "安全标准要求", "检查结果", "问题描述", "整改措施", "责任人"])
        sheet.append(["人员安全", 1, "劳保用品", "按规定佩戴", "合格", "佩戴规范", "持续检查", "张工"])
        sheet.append(["设备安全", 2, "交叉作业", "作业区域隔离", "不合格", "隔离带缺失", "当天补齐", "李工"])
        workbook.save(safety_stream)
        workbook.close()
        safety_content = safety_stream.getvalue()
        safety_query = urllib.parse.urlencode({
            "kind": "safety", "date": report_date, "name": "安全检查记录.xlsx",
        })
        status, safety_upload = self.call(
            f"/api/admin/daily-source-uploads?{safety_query}", token=self.admin,
            raw=safety_content, headers={"Content-Length": str(len(safety_content))},
        )
        self.assertEqual(status, 201)
        self.assertEqual(safety_upload["upload"]["summary"]["unqualified_count"], 1)

        status, snapshot = self.call(f"/api/daily-report?date={report_date}", token=self.admin)
        self.assertEqual(status, 200)
        self.assertEqual(snapshot["arrival"]["upload_count"], 1)
        self.assertEqual(snapshot["arrival"]["batch_count"], 1)
        self.assertEqual(snapshot["arrival"]["missing_material_detail_count"], 1)
        self.assertEqual(snapshot["arrival"]["batches"][0]["missing_materials"][0]["shortage_quantity"], "2")
        self.assertEqual(snapshot["arrival"]["batches"][0]["missing_materials"][0]["received_quantity"], "8")
        self.assertEqual(snapshot["safety_checks"]["total_checks"], 2)
        self.assertEqual(snapshot["safety_checks"]["unqualified_count"], 1)
        self.assertEqual(len(snapshot["source_uploads"]), 2)

        source_id = safety_upload["upload"]["id"]
        status, removed = self.call(
            f"/api/admin/daily-source-uploads/{source_id}", token=self.admin, method="DELETE",
        )
        self.assertEqual(status, 200)
        self.assertIn("回收站", removed["message"])
        trash = self.call("/api/admin/trash", token=self.admin)[1]["trash"]
        item = next(row for row in trash if row["kind"] == "daily_source_upload")
        self.assertEqual(self.call(
            f"/api/admin/trash/{item['id']}/restore", {}, token=self.admin,
        )[0], 200)
        restored = self.call(f"/api/daily-report?date={report_date}", token=self.admin)[1]
        self.assertEqual(restored["safety_checks"]["unqualified_count"], 1)

    def test_legacy_production_group_attendance_migrates_to_default_shift(self):
        """旧版仅按班组保存的生产出勤记录应无损迁移到默认班次。"""

        report_date = web_server.business_today().isoformat()
        created = web_server.now_iso()
        with web_server.DB_LOCK, web_server.db() as connection:
            admin_id = int(connection.execute(
                "SELECT id FROM users WHERE username = 'admin'",
            ).fetchone()["id"])
            group_id = int(connection.execute(
                "INSERT INTO daily_production_groups(name, sort_order, active, created_at, updated_at) "
                "VALUES (?, ?, 1, ?, ?)",
                ("历史装配组", 99, created, created),
            ).lastrowid)
            connection.execute(
                "INSERT INTO daily_production_attendance"
                "(report_date, group_id, attendance_count, note, updated_by, updated_at) "
                "VALUES (?, ?, ?, ?, ?, ?)",
                (report_date, group_id, 7, "旧版班组记录", admin_id, created),
            )
        web_server.init_db()

        status, attendance = self.call(
            f"/api/admin/daily-attendance?date={report_date}", token=self.admin,
        )
        self.assertEqual(status, 200)
        migrated = next(item for item in attendance["production_groups"] if item["group_id"] == group_id)
        self.assertEqual(migrated["shift_name"], "白班")
        self.assertEqual(migrated["staffing_count"], 0)
        self.assertEqual(migrated["attendance_count"], 7)
        self.assertEqual(migrated["difference"], -7)
        groups = self.call("/api/admin/daily-production-groups", token=self.admin)[1]["groups"]
        group = next(item for item in groups if item["id"] == group_id)
        self.assertEqual(group["shifts"][0]["name"], "白班")

    def test_production_staffing_update_refreshes_today_but_keeps_history(self):
        """班组编制调整应刷新当天差异，但不得改写已经形成的历史日清数据。"""

        today = web_server.business_today().isoformat()
        previous_day = (web_server.business_today() - timedelta(days=1)).isoformat()
        status, created = self.call("/api/admin/daily-production-groups", {
            "name": "编制快照组", "sort_order": 1, "active": True,
            "shifts": [{"name": "白班", "staffing_count": 10, "sort_order": 1, "active": True}],
        }, token=self.admin)
        self.assertEqual(status, 201)
        group = created["group"]
        shift = group["shifts"][0]
        for report_date in (previous_day, today):
            status, _ = self.call("/api/admin/daily-attendance", {
                "date": report_date, "records": [],
                "production_groups": [{"shift_id": shift["id"], "attendance_count": 8, "note": ""}],
            }, token=self.admin)
            self.assertEqual(status, 200)

        status, updated = self.call(
            f"/api/admin/daily-production-groups/{group['id']}", {
                "name": group["name"], "sort_order": group["sort_order"], "active": True,
                "shifts": [{
                    "id": shift["id"], "name": shift["name"], "staffing_count": 12,
                    "sort_order": shift["sort_order"], "active": True,
                }],
            }, token=self.admin, method="PATCH",
        )
        self.assertEqual(status, 200)
        self.assertEqual(updated["group"]["staffing_count"], 12)
        today_row = self.call(
            f"/api/admin/daily-attendance?date={today}", token=self.admin,
        )[1]["production_groups"][0]
        history_row = self.call(
            f"/api/admin/daily-attendance?date={previous_day}", token=self.admin,
        )[1]["production_groups"][0]
        self.assertEqual(today_row["staffing_count"], 12)
        self.assertEqual(today_row["difference"], 4)
        self.assertEqual(history_row["staffing_count"], 10)
        self.assertEqual(history_row["difference"], 2)

    def test_templates_search_preflight_and_retry(self):
        """模板搜索、执行前检查和失败重试必须复用同一任务参数与用户隔离目录。"""

        status, created = self.call("/api/templates", {
            "name": "常用文本去重", "action": "text.transform",
            "payload": {"options": {"operation": "dedup"}},
        }, token=self.admin)
        self.assertEqual(status, 201)
        template_id = created["id"]
        status, templates = self.call("/api/templates", token=self.admin)
        self.assertEqual(status, 200)
        self.assertEqual(templates["templates"][0]["name"], "常用文本去重")
        status, _ = self.call(f"/api/templates/{template_id}", {"name": "文本去重模板", "payload": {"options": {"operation": "sort"}}}, token=self.admin, method="PATCH")
        self.assertEqual(status, 200)
        status, preflight = self.call("/api/jobs/preflight", {"action": "text.transform", "payload": {"text": "甲\n乙", "operation": "dedup"}}, token=self.admin)
        self.assertEqual(status, 200)
        self.assertTrue(preflight["ok"])
        status, created = self.call("/api/jobs", {"action": "text.transform", "title": "可搜索任务", "payload": {"text": "甲\n甲", "operation": "dedup"}}, token=self.admin)
        self.assertEqual(status, 202)
        job = self.wait_job(created["job_id"])
        self.assertEqual(len(job["versions"]), 1)
        status, found = self.call("/api/search?q=" + urllib.parse.quote("可搜索"), token=self.admin)
        self.assertEqual(status, 200)
        self.assertEqual(found["jobs"][0]["id"], job["id"])
        with web_server.DB_LOCK, web_server.db() as connection:
            connection.execute("UPDATE web_jobs SET status = 'failed', error = '模拟失败' WHERE id = ?", (job["id"],))
        status, retried = self.call(f"/api/jobs/{job['id']}/retry", token=self.admin, payload={})
        self.assertEqual(status, 202)
        retry = self.wait_job(retried["job_id"])
        self.assertEqual(retry["status"], "completed")
        self.assertEqual(retry["retry_of"], job["id"])
        status, _ = self.call(f"/api/templates/{template_id}", token=self.admin, method="DELETE")
        self.assertEqual(status, 200)

    def test_compare_review_continues_same_job(self):
        """对比业务人工复核应在原任务上继续执行，不能绕过计划或创建孤立任务。"""

        def book_bytes(rows):
            """把合成行序列化为人工复核上传所需的工作簿字节。"""

            book = Workbook()
            sheet = book.active
            for row in rows:
                sheet.append(row)
            stream = __import__("io").BytesIO()
            book.save(stream)
            return stream.getvalue()

        handles = []
        for name, content in (("a.xlsx", [["编号", "数量", "单价"], ["A-1", 1, 10]]),
                              ("b.xlsx", [["编号", "数量", "单价"], ["A-1", 2, 12]])):
            query = urllib.parse.urlencode({"name": name, "group": "review-group"})
            status, uploaded = self.call(
                f"/api/files/upload?{query}", token=self.admin,
                raw=book_bytes(content), headers={"Content-Length": str(len(book_bytes(content)))},
            )
            self.assertEqual(status, 201)
            handles.append(uploaded["handle"])

        status, created = self.call("/api/jobs", {
            "action": "web.compare.review", "title": "表格比对复核", "payload": {
                "file1": [handles[0]], "file2": [handles[1]],
            },
        }, token=self.admin)
        self.assertEqual(status, 202)
        prepared = self.wait_job(created["job_id"])
        self.assertEqual(prepared["status"], "completed")
        self.assertTrue(prepared["review_pending"])
        self.assertIn("编号", prepared["result"]["common"])

        status, board = self.call("/api/dashboard", token=self.admin)
        self.assertEqual(status, 200)
        usage_keys = {item["key"] for item in board["feature_usage"]}
        self.assertIn("compare", usage_keys)
        self.assertNotIn("web", usage_keys)
        self.assertEqual(board["metrics"]["completed_jobs"], 0)
        self.assertEqual(board["metrics"]["running_jobs"], 1)
        self.assertEqual(board["status_breakdown"]["review"], 1)
        self.assertTrue(board["recent_jobs"][0]["review_pending"])

        status, resumed = self.call(f"/api/jobs/{created['job_id']}/review", {
            "choices": {"key": "编号", "columns": ["数量"]},
        }, token=self.admin)
        self.assertEqual(status, 202)
        self.assertEqual(resumed["job_id"], created["job_id"])
        completed = self.wait_job(created["job_id"])
        self.assertEqual(completed["status"], "completed")
        self.assertFalse(completed["review_pending"])
        self.assertEqual(completed["result"]["result"]["counts"]["diffs"], 1)
        self.assertEqual(completed["result"]["result"]["columns"], ["数量"])
        self.assertEqual(completed["presentation"]["parameters"][1]["value"], "数量")
        self.assertEqual(len(completed["files"]), 1)

    def test_supplier_batch_review_selects_suppliers_and_excludes_original(self):
        """供应商批次复核只生成用户勾选供应商，并始终排除原厂内容。"""

        book = Workbook()
        sheet = book.active
        sheet.title = "批次清单"
        sheet.append(["材料编号", "材料名称", "规格", "单位", "最终采购数量", "供应商名称"])
        sheet.append(["JBC-001", "密封圈", "DN50", "个", 3, "供应商甲"])
        sheet.append(["JBC-002", "原厂密封圈", "DN50", "个", 5, "供应商原厂"])
        sheet.append(["JBC-003", "螺栓", "M8", "套", 2, "供应商乙"])
        stream = __import__("io").BytesIO()
        book.save(stream)
        content = stream.getvalue()

        query = urllib.parse.urlencode({"name": "批次清单.xlsx", "group": "supplier-batch-review"})
        status, uploaded = self.call(
            f"/api/files/upload?{query}", token=self.admin,
            raw=content, headers={"Content-Length": str(len(content))},
        )
        self.assertEqual(status, 201)

        status, created = self.call("/api/jobs", {
            "action": "web.supplier_batch.review", "title": "供应商批次表复核",
            "payload": {"batch_paths": [uploaded["handle"]]},
        }, token=self.admin)
        self.assertEqual(status, 202)
        prepared = self.wait_job(created["job_id"])
        self.assertEqual(prepared["status"], "completed")
        self.assertTrue(prepared["review_pending"])
        self.assertEqual(
            {item["name"] for item in prepared["result"]["result"]["suppliers"]},
            {"供应商甲", "供应商乙"},
        )
        self.assertEqual(prepared["result"]["result"]["excluded_original_count"], 1)

        status, rejected = self.call(f"/api/jobs/{created['job_id']}/review", {
            "choices": {"suppliers": ["供应商乙"]},
        }, token=self.admin)
        self.assertEqual(status, 400)
        self.assertIn("交付日期", rejected["error"])

        status, resumed = self.call(f"/api/jobs/{created['job_id']}/review", {
            "choices": {
                "suppliers": ["供应商乙"],
                "batch_dates": {"批次清单": "8.7"},
            },
        }, token=self.admin)
        self.assertEqual(status, 202)
        self.assertEqual(resumed["job_id"], created["job_id"])
        completed = self.wait_job(created["job_id"])
        self.assertEqual(completed["status"], "completed")
        self.assertFalse(completed["review_pending"])
        result = completed["result"]["result"]
        self.assertEqual(result["suppliers"], ["供应商乙"])
        self.assertEqual(result["batch_dates"], {"批次清单": "8.7"})
        self.assertEqual(result["generated"], 1)
        self.assertEqual(result["rows"], 1)
        self.assertEqual(result["excluded_original_count"], 1)
        self.assertEqual(len(completed["files"]), 1)

    def test_admin_data_account_and_notifications(self):
        """管理员可维护账号、查看数据摘要并向指定用户或全局发布通知。"""

        status, _ = self.call("/api/auth/register", {
            "username": "member_one", "display_name": "成员一", "password": "password123",
        })
        self.assertEqual(status, 201)
        status, users = self.call("/api/admin/users", token=self.admin)
        self.assertEqual(status, 200)
        member = next(item for item in users["users"] if item["username"] == "member_one")
        status, _ = self.call(f"/api/admin/users/{member['id']}/approve", token=self.admin, payload={})
        self.assertEqual(status, 200)
        status, _ = self.call(f"/api/admin/users/{member['id']}", {
            "display_name": "成员一（已更新）", "status": "approved",
        }, token=self.admin, method="PATCH")
        self.assertEqual(status, 200)
        status, _ = self.call("/api/admin/messages", {
            "user_id": member["id"], "title": "定向提醒", "content": "请查看本周数据。",
        }, token=self.admin)
        self.assertEqual(status, 201)
        status, _ = self.call("/api/admin/announcements", {
            "title": "全局公告", "content": "系统将在周末维护。",
        }, token=self.admin)
        self.assertEqual(status, 201)
        status, data = self.call("/api/admin/data", token=self.admin)
        self.assertEqual(status, 200)
        self.assertGreaterEqual(data["summary"]["approved_users"], 2)
        member_token = self.call("/api/auth/login", {"username": "member_one", "password": "password123"})[1]["token"]
        status, _ = self.call("/api/admin/data", token=member_token)
        self.assertEqual(status, 403)
        status, board = self.call("/api/dashboard", token=member_token)
        self.assertEqual(status, 200)
        self.assertEqual({item["title"] for item in board["notifications"]}, {"定向提醒", "全局公告"})
        status, inbox = self.call("/api/notifications", token=member_token)
        self.assertEqual(status, 200)
        self.assertEqual(inbox["unread_count"], 2)
        announcement = next(item for item in inbox["notifications"] if item["kind"] == "announcement")
        status, _ = self.call(f"/api/notifications/announcement/{announcement['id']}/read", token=member_token, payload={})
        self.assertEqual(status, 200)
        status, inbox = self.call("/api/notifications", token=member_token)
        self.assertEqual(inbox["unread_count"], 1)
        status, _ = self.call("/api/notifications/read-all", token=member_token, payload={})
        self.assertEqual(status, 200)
        status, inbox = self.call("/api/notifications", token=member_token)
        self.assertEqual(inbox["unread_count"], 0)
        status, _ = self.call(f"/api/admin/users/{member['id']}", token=self.admin, method="DELETE")
        self.assertEqual(status, 200)
        status, _ = self.call("/api/auth/me", token=member_token)
        self.assertEqual(status, 401)

    def test_admin_role_access_sessions_and_audit(self):
        """角色授予、管理入口、会话撤销和审计记录必须形成一致的管理员闭环。"""

        status, _ = self.call("/api/auth/register", {
            "username": "manager_one", "display_name": "管理员候选", "password": "password123",
        })
        self.assertEqual(status, 201)
        status, users = self.call("/api/admin/users", token=self.admin)
        self.assertEqual(status, 200)
        member = next(item for item in users["users"] if item["username"] == "manager_one")
        primary = next(item for item in users["users"] if item["username"] == "admin")
        status, _ = self.call(f"/api/admin/users/{member['id']}/approve", token=self.admin, payload={})
        self.assertEqual(status, 200)
        member_token = self.call(
            "/api/auth/login", {"username": "manager_one", "password": "password123"},
        )[1]["token"]

        status, payload = self.call(
            f"/api/admin/users/{member['id']}/role", {"role": "admin"}, token=self.admin,
        )
        self.assertEqual(status, 200)
        self.assertIn("管理员权限", payload["message"])
        status, _ = self.call("/api/admin/data", token=member_token)
        self.assertEqual(status, 200)

        status, _ = self.call(
            f"/api/admin/users/{member['id']}/role", {"role": "user"}, token=member_token,
        )
        self.assertEqual(status, 400)
        status, _ = self.call(
            f"/api/admin/users/{primary['id']}/role", {"role": "user"}, token=member_token,
        )
        self.assertEqual(status, 400)

        status, _ = self.call(
            f"/api/admin/users/{member['id']}/role", {"role": "user"}, token=self.admin,
        )
        self.assertEqual(status, 200)
        status, _ = self.call("/api/admin/data", token=member_token)
        self.assertEqual(status, 403)
        status, payload = self.call(
            f"/api/admin/users/{member['id']}/sessions/revoke", {}, token=self.admin,
        )
        self.assertEqual(status, 200)
        self.assertIn("1 个登录会话", payload["message"])
        status, _ = self.call("/api/auth/me", token=member_token)
        self.assertEqual(status, 401)

        status, _ = self.call(
            f"/api/admin/users/{member['id']}/access", {"enabled": False}, token=self.admin,
        )
        self.assertEqual(status, 200)
        status, payload = self.call(
            "/api/auth/login", {"username": "manager_one", "password": "password123"},
        )
        self.assertEqual(status, 403)
        self.assertIn("暂停", payload["error"])
        status, data = self.call("/api/admin/data", token=self.admin)
        self.assertEqual(status, 200)
        self.assertEqual(data["summary"]["disabled_users"], 1)

        status, _ = self.call(
            f"/api/admin/users/{member['id']}/access", {"enabled": True}, token=self.admin,
        )
        self.assertEqual(status, 200)
        restored_token = self.call(
            "/api/auth/login", {"username": "manager_one", "password": "password123"},
        )[1]["token"]
        status, _ = self.call("/api/admin/audit", token=restored_token)
        self.assertEqual(status, 403)
        status, payload = self.call("/api/admin/audit", token=self.admin)
        self.assertEqual(status, 200)
        actions = {item["action"] for item in payload["audit"]}
        self.assertTrue({
            "grant_admin", "revoke_admin", "revoke_sessions", "disable_user", "enable_user",
        }.issubset(actions))

    def test_login_lock_password_change_and_device_sessions(self):
        """连续登录失败触发锁定，改密后旧设备会话应按安全规则失效。"""

        for attempt in range(web_server.LOGIN_FAILURE_LIMIT):
            status, payload = self.call(
                "/api/auth/login", {"username": "admin", "password": "wrong-password1"},
            )
            self.assertEqual(status, 429 if attempt == web_server.LOGIN_FAILURE_LIMIT - 1 else 401)
        status, payload = self.call(
            "/api/auth/login", {"username": "admin", "password": "admin123456"},
        )
        self.assertEqual(status, 429)
        self.assertIn("分钟后重试", payload["error"])

        with web_server.DB_LOCK, web_server.db() as connection:
            connection.execute("DELETE FROM login_attempts")
        status, second = self.call(
            "/api/auth/login", {"username": "admin", "password": "admin123456"},
            headers={"User-Agent": "FYT-Test-Device-2"},
        )
        self.assertEqual(status, 200)
        second_token = second["token"]
        status, payload = self.call("/api/auth/sessions", token=second_token)
        self.assertEqual(status, 200)
        self.assertEqual(len(payload["sessions"]), 2)
        self.assertEqual(sum(bool(item["current"]) for item in payload["sessions"]), 1)

        status, payload = self.call("/api/auth/password", {
            "current_password": "admin123456", "new_password": "NewPassword2026",
        }, token=second_token)
        self.assertEqual(status, 200)
        self.assertIn("其他设备已退出", payload["message"])
        self.assertEqual(self.call("/api/auth/me", token=self.admin)[0], 401)
        self.assertEqual(self.call("/api/auth/me", token=second_token)[0], 200)
        self.assertEqual(self.call(
            "/api/auth/login", {"username": "admin", "password": "admin123456"},
        )[0], 401)
        self.assertEqual(self.call(
            "/api/auth/login", {"username": "admin", "password": "NewPassword2026"},
        )[0], 200)

    def test_admin_password_reset_revokes_target_sessions(self):
        """管理员重置他人密码后必须撤销该账号全部现有会话。"""

        self.call("/api/auth/register", {
            "username": "reset_member", "display_name": "待重置成员", "password": "password123",
        })
        member = next(item for item in self.call("/api/admin/users", token=self.admin)[1]["users"] if item["username"] == "reset_member")
        self.call(f"/api/admin/users/{member['id']}/approve", {}, token=self.admin)
        member_token = self.call(
            "/api/auth/login", {"username": "reset_member", "password": "password123"},
        )[1]["token"]
        status, payload = self.call(
            f"/api/admin/users/{member['id']}/password",
            {"password": "ResetPassword2026"}, token=self.admin,
        )
        self.assertEqual(status, 200)
        self.assertIn("1 个登录会话", payload["message"])
        self.assertEqual(self.call("/api/auth/me", token=member_token)[0], 401)
        self.assertEqual(self.call(
            "/api/auth/login", {"username": "reset_member", "password": "ResetPassword2026"},
        )[0], 200)

    def test_report_center_and_batch_track_include_result_files(self):
        """报表中心与批次跟踪应关联可下载结果文件，而不是只显示任务元数据。"""

        job_id = "report-batch-job"
        batch_name = "GKMYR26027-06"
        output = (
            web_server.DATA_ROOT / "users" / "1" / "jobs" / job_id
            / "outputs" / f"{batch_name}到料结果.xlsx"
        )
        output.parent.mkdir(parents=True)
        workbook = Workbook()
        workbook.active.append(["批次号", batch_name])
        workbook.save(output)
        timestamp = web_server.now_iso()
        with web_server.DB_LOCK, web_server.db() as connection:
            connection.execute(
                "INSERT INTO web_jobs(id, user_id, action, title, status, progress, logs, result, "
                "files, cancelled, payload, created_at, updated_at) "
                "VALUES (?, 1, 'web.arrival', '每日到料结果', 'completed', 100, '[]', ?, "
                "'[]', 0, '{}', ?, ?)",
                (
                    job_id,
                    json.dumps({"report_path": str(output)}, ensure_ascii=False),
                    timestamp,
                    timestamp,
                ),
            )

        query = urllib.parse.quote(batch_name)
        status, tracked = self.call(f"/api/batch-track?q={query}", token=self.admin)
        self.assertEqual(status, 200)
        self.assertEqual(len(tracked["items"]), 1)
        self.assertEqual(tracked["items"][0]["job_id"], job_id)
        self.assertEqual(tracked["items"][0]["files"], [output.name])

        status, report = self.call("/api/reports?range=7d", token=self.admin)
        self.assertEqual(status, 200)
        self.assertTrue(report["name"].startswith("业务报表_近7天_"))
        status, content = self.call(report["url"], token=self.admin)
        self.assertEqual(status, 200)
        exported = load_workbook(BytesIO(content), data_only=True)
        self.assertEqual(exported.sheetnames, ["汇总", "明细"])
        self.assertEqual(exported["汇总"]["B5"].value, 1)
        self.assertEqual(exported["明细"]["C2"].value, "每日到料结果")

        with mock.patch("web_backend.services.reports.datetime") as mocked_datetime:
            mocked_datetime.now.return_value = datetime(2026, 8, 10, 9, 0, 0)
            weekly = web_server.auto_weekly_report_if_due()
            self.assertTrue(weekly.startswith("已生成周报 "))
            self.assertEqual(web_server.auto_weekly_report_if_due(), "")
        state = json.loads(
            (web_server.DATA_ROOT / "auto_weekly_report_state.json").read_text(
                encoding="utf-8",
            )
        )
        self.assertEqual(state["last_weekly"], "20260810")
        with web_server.DB_LOCK, web_server.db() as connection:
            notification_count = connection.execute(
                "SELECT COUNT(*) FROM messages WHERE title = '本周业务报表已生成'",
            ).fetchone()[0]
        self.assertEqual(notification_count, 1)

        self.assertEqual(
            self.call("/api/reports?range=unknown", token=self.admin)[0],
            400,
        )
        invalid_path = urllib.parse.quote("../accounts.sqlite3")
        self.assertEqual(
            self.call(f"/api/reports/download?path={invalid_path}", token=self.admin)[0],
            404,
        )

    def test_role_matrix_and_workshop_edit_scope(self):
        """成员、班组长和管理员的导航/API 权限及现场问题编辑范围必须匹配角色矩阵。"""

        accounts = {}
        for username, display_name, role in (
            ("role_leader", "班组长甲", "team_leader"),
            ("role_member", "业务成员甲", "user"),
        ):
            self.assertEqual(self.call("/api/auth/register", {
                "username": username, "display_name": display_name, "password": "password123",
            })[0], 201)
            account = next(
                item for item in self.call("/api/admin/users", token=self.admin)[1]["users"]
                if item["username"] == username
            )
            self.assertEqual(
                self.call(f"/api/admin/users/{account['id']}/approve", {}, token=self.admin)[0],
                200,
            )
            if role != "user":
                status, payload = self.call(
                    f"/api/admin/users/{account['id']}/role",
                    {"role": role}, token=self.admin,
                )
                self.assertEqual(status, 200)
                self.assertIn("班组长", payload["message"])
            accounts[username] = self.call("/api/auth/login", {
                "username": username, "password": "password123",
            })[1]["token"]

        leader = accounts["role_leader"]
        member = accounts["role_member"]
        self.assertEqual(self.call("/api/auth/me", token=leader)[1]["user"]["role"], "team_leader")
        self.assertEqual(self.call("/api/daily-report?date=2026-08-08", token=leader)[0], 403)
        self.assertEqual(self.call("/api/daily-report?date=2026-08-08", token=member)[0], 403)
        query = urllib.parse.quote("不存在的批次")
        self.assertEqual(self.call(f"/api/batch-track?q={query}", token=leader)[0], 200)
        self.assertEqual(self.call(f"/api/batch-track?q={query}", token=member)[0], 403)
        self.assertEqual(self.call("/api/reports?range=7d", token=leader)[0], 403)
        self.assertEqual(self.call("/api/reports?range=7d", token=member)[0], 403)
        self.assertEqual(self.call("/api/library/files", token=leader)[0], 200)
        self.assertEqual(self.call("/api/library/files", token=member)[0], 403)

        issue_date = datetime.now().strftime("%Y-%m-%d")
        issue_payload = {
            "issue_date": issue_date,
            "category": "error_proofing",
            "cause": "班组长发布的现场问题",
            "happened_at": "08:30",
            "tracking_status": "处理中",
            "responsible_person": "班组长甲",
        }
        status, created = self.call("/api/workshop/issues", issue_payload, token=leader)
        self.assertEqual(status, 201)
        issue_id = created["issue"]["id"]
        self.assertEqual(self.call(f"/api/workshop/issues/{issue_id}/publish", {}, token=leader)[0], 200)
        published_issue = next(
            item for item in self.call(f"/api/workshop/issues?date={issue_date}", token=leader)[1]["issues"]
            if item["id"] == issue_id
        )
        listed = self.call(f"/api/workshop/issues?date={issue_date}", token=member)[1]
        visible = next(item for item in listed["issues"] if item["id"] == issue_id)
        self.assertFalse(visible["permissions"]["can_edit"])
        self.assertFalse(visible["permissions"]["can_resolve"])
        self.assertEqual(self.call(
            f"/api/workshop/issues/{issue_id}",
            {"cause": "普通成员不能修改班组长已发布问题", "expected_updated_at": visible["updated_at"]},
            token=member, method="PATCH",
        )[0], 403)
        self.assertEqual(self.call(
            f"/api/workshop/issues/{issue_id}",
            {"cause": "班组长已修正现场问题", "expected_updated_at": published_issue["updated_at"]},
            token=leader, method="PATCH",
        )[0], 200)

    def test_shared_library_permissions_pagination_and_restore(self):
        """共享数据库需验证角色权限、分页、修改时间、软删除和恢复流程。"""

        tokens = {}
        for username, display_name in (("library_a", "资料甲"), ("library_b", "资料乙")):
            self.assertEqual(self.call("/api/auth/register", {
                "username": username, "display_name": display_name, "password": "password123",
            })[0], 201)
            account = next(
                item for item in self.call("/api/admin/users", token=self.admin)[1]["users"]
                if item["username"] == username
            )
            self.assertEqual(
                self.call(f"/api/admin/users/{account['id']}/approve", {}, token=self.admin)[0],
                200,
            )
            self.assertEqual(
                self.call(
                    f"/api/admin/users/{account['id']}/role",
                    {"role": "team_leader"}, token=self.admin,
                )[0],
                200,
            )
            tokens[username] = self.call("/api/auth/login", {
                "username": username, "password": "password123",
            })[1]["token"]

        def upload(name, content, scope="team"):
            """上传共享资料并返回服务端记录，允许测试切换共享范围。"""

            query = urllib.parse.urlencode({
                "name": name, "scope": scope, "description": f"{name}说明",
            })
            status, payload = self.call(
                f"/api/library/files?{query}", token=tokens["library_a"], raw=content,
                headers={"Content-Length": str(len(content))},
            )
            self.assertEqual(status, 201)
            return payload["file"]

        team_file = upload("团队资料.txt", b"team-v1")
        private_file = upload("个人资料.txt", b"private", "private")
        second_team = upload("团队资料二.txt", b"team-v2")
        self.assertEqual(team_file["uploader"]["display_name"], "资料甲")
        self.assertEqual(team_file["scope"], "team")
        self.assertTrue(team_file["permissions"]["can_edit"])

        status, listed = self.call(
            "/api/library/files?page=1&page_size=1", token=tokens["library_b"],
        )
        self.assertEqual(status, 200)
        self.assertEqual(listed["pagination"], {"page": 1, "page_size": 1, "total": 2, "pages": 2})
        self.assertEqual(len(listed["files"]), 1)
        self.assertFalse(listed["files"][0]["permissions"]["can_edit"])
        self.assertEqual(listed["summary"]["visible_count"], 2)

        status, content = self.call(
            f"/api/library/files/{team_file['id']}/download", token=tokens["library_b"],
        )
        self.assertEqual((status, content), (200, b"team-v1"))
        self.assertEqual(self.call(
            f"/api/library/files/{private_file['id']}/download", token=tokens["library_b"],
        )[0], 404)
        self.assertEqual(self.call(
            f"/api/library/files/{team_file['id']}",
            {"name": "越权修改.txt", "scope": "team"},
            token=tokens["library_b"], method="PATCH",
        )[0], 403)

        status, updated = self.call(
            f"/api/library/files/{team_file['id']}",
            {"name": "团队资料-更新.txt", "description": "已核对", "scope": "team"},
            token=tokens["library_a"], method="PATCH",
        )
        self.assertEqual(status, 200)
        self.assertEqual(updated["file"]["updated_by"]["username"], "library_a")
        self.assertTrue(updated["file"]["updated_at"])

        replace_query = urllib.parse.urlencode({"name": "团队资料-替换.txt"})
        status, replaced = self.call(
            f"/api/library/files/{team_file['id']}/content?{replace_query}",
            token=tokens["library_a"], raw=b"replacement",
            headers={"Content-Length": "11"},
        )
        self.assertEqual(status, 200)
        self.assertEqual(replaced["file"]["size"], 11)
        self.assertEqual(self.call(
            f"/api/library/files/{team_file['id']}/download", token=tokens["library_b"],
        ), (200, b"replacement"))

        status, admin_view = self.call(
            "/api/library/files?scope=private", token=self.admin,
        )
        self.assertEqual(status, 200)
        self.assertEqual([item["id"] for item in admin_view["files"]], [private_file["id"]])
        self.assertTrue(admin_view["files"][0]["permissions"]["can_delete"])

        self.assertEqual(self.call(
            f"/api/library/files/{second_team['id']}", token=tokens["library_a"], method="DELETE",
        )[0], 200)
        trash = self.call("/api/admin/trash", token=self.admin)[1]["trash"]
        deleted = next(item for item in trash if item["label"] == "团队资料二.txt")
        self.assertEqual(deleted["kind"], "library_file")
        self.assertEqual(self.call(
            f"/api/admin/trash/{deleted['id']}/restore", {}, token=self.admin,
        )[0], 200)
        self.assertEqual(self.call(
            f"/api/library/files/{second_team['id']}/download", token=tokens["library_b"],
        ), (200, b"team-v2"))

    def test_library_classification_filter_override_and_replacement(self):
        """自动分类、筛选、人工覆盖和同名替换必须保持文件与索引一致。"""

        def workbook_bytes(headers, row):
            """构造单表分类样本字节，便于覆盖替换前后的类别变化。"""

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "数据表"
            sheet.append(headers)
            sheet.append(row)
            stream = BytesIO()
            workbook.save(stream)
            return stream.getvalue()

        supplier_content = workbook_bytes(
            ["批次号", "属性", "下阶物料", "下阶物料描述", "供应商代码", "供应商名称", "合计", "库区"],
            ["GK1", "KD", "8892602000", "右前踏板", "100079", "北京丰达", 360, "M62"],
        )
        query = urllib.parse.urlencode({"name": "供应商明细.xlsx", "scope": "team"})
        status, uploaded = self.call(
            f"/api/library/files?{query}", token=self.admin, raw=supplier_content,
            headers={"Content-Length": str(len(supplier_content))},
        )
        self.assertEqual(status, 201)
        file_id = uploaded["file"]["id"]
        self.assertEqual(uploaded["file"]["category"], "deliv_supp")
        self.assertIn("deliv_supp", uploaded["file"]["categories"])

        status, filtered = self.call(
            "/api/library/files?category=deliv_supp", token=self.admin,
        )
        self.assertEqual(status, 200)
        self.assertEqual([item["id"] for item in filtered["files"]], [file_id])

        status, edited = self.call(
            f"/api/library/files/{file_id}", {"category": "pivot_src"},
            token=self.admin, method="PATCH",
        )
        self.assertEqual(status, 200)
        self.assertEqual(edited["file"]["category"], "pivot_src")
        self.assertEqual(edited["file"]["confidence"], 100)

        pivot_content = workbook_bytes(
            ["版本序号", "材料编号", "材料名称", "规格", "数量", "单位", "最终采购数量"],
            [1, "MAT001", "纸箱", "600x400", 10, "个", 120],
        )
        replace_query = urllib.parse.urlencode({"name": "采购数据.xlsx"})
        status, replaced = self.call(
            f"/api/library/files/{file_id}/content?{replace_query}",
            token=self.admin, raw=pivot_content,
            headers={"Content-Length": str(len(pivot_content))},
        )
        self.assertEqual(status, 200)
        self.assertEqual(replaced["file"]["category"], "pivot_src")
        self.assertGreaterEqual(replaced["file"]["confidence"], 50)

    def test_library_category_migration_keeps_secondary_labels(self):
        """分类索引迁移后应保留多标签中的次要业务分类，避免搜索能力退化。"""

        legacy_id = "legacy-category-row"
        timestamp = web_server.now_iso()
        with web_server.DB_LOCK, web_server.db() as connection:
            connection.execute(
                "INSERT INTO library_files(id, owner_id, name, path, size, content_type, description, scope, "
                "created_at, updated_at, updated_by, category, categories, confidence, signals, sheet, category_sheets) "
                "VALUES (?, 1, ?, ?, 0, 'application/octet-stream', '', 'team', ?, ?, 1, ?, ?, 80, '[]', '', '{}')",
                (legacy_id, "旧文件.xlsx", str(self.temp.name), timestamp, timestamp, "pivot_src", "[\"deliv_supp\"]"),
            )
        web_server.init_db()
        with web_server.DB_LOCK, web_server.db() as connection:
            row = connection.execute(
                "SELECT categories FROM library_files WHERE id = ?", (legacy_id,)
            ).fetchone()
            labels = {
                item["category"] for item in connection.execute(
                    "SELECT category FROM library_file_categories WHERE file_id = ?", (legacy_id,)
                ).fetchall()
            }
        self.assertEqual(set(json.loads(row["categories"])), {"pivot_src", "deliv_supp"})
        self.assertEqual(labels, {"pivot_src", "deliv_supp"})

    def test_workshop_daily_issue_publish_permissions_export_and_restore(self):
        """现场问题草稿、图片、发布权限、按日展示、模板导出和回收站恢复应完整串联。"""

        tokens = {}
        for username, display_name in (("workshop_a", "车间甲"), ("workshop_b", "车间乙")):
            self.assertEqual(self.call("/api/auth/register", {
                "username": username, "display_name": display_name, "password": "password123",
            })[0], 201)
            account = next(
                item for item in self.call("/api/admin/users", token=self.admin)[1]["users"]
                if item["username"] == username
            )
            self.assertEqual(
                self.call(f"/api/admin/users/{account['id']}/approve", {}, token=self.admin)[0],
                200,
            )
            if username == "workshop_a":
                self.assertEqual(
                    self.call(
                        f"/api/admin/users/{account['id']}/role",
                        {"role": "team_leader"}, token=self.admin,
                    )[0],
                    200,
                )
            tokens[username] = self.call("/api/auth/login", {
                "username": username, "password": "password123",
            })[1]["token"]

        issue_date = datetime.now().strftime("%Y-%m-%d")
        future_date = (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%d")
        status, rejected = self.call("/api/workshop/issues", {
            "issue_date": future_date, "cause": "未来问题", "primary_owner": "负责人",
        }, token=tokens["workshop_a"])
        self.assertEqual(status, 400)
        self.assertIn("不能晚于今天", rejected["error"])

        status, invalid_category = self.call("/api/workshop/issues", {
            "issue_date": issue_date,
            "cause": "旧分类不应继续使用",
            "category": "quality",
        }, token=tokens["workshop_a"])
        self.assertEqual(status, 400)
        self.assertIn("主料异常", invalid_category["error"])

        status, wrong_fields = self.call("/api/workshop/issues", {
            "issue_date": issue_date,
            "cause": "主料异常不应接收海外字段",
            "category": "main_material",
            "discoverer": "李工",
            "country": "印度尼西亚",
        }, token=tokens["workshop_a"])
        self.assertEqual(status, 400)
        self.assertIn("不使用以下字段", wrong_fields["error"])

        status, created = self.call("/api/workshop/issues", {
            "issue_date": issue_date,
            "cause": "装配工位防护罩松动",
            "category": "main_material",
            "severity": "important",
            "model": "VX11",
            "batch_no": "GKMYR26027-06",
            "team": "小件组",
            "material_code": "A-01",
            "material_name": "防护罩",
            "cause_analysis": "紧固件松动",
            "corrective_action": "重新紧固并复核",
            "external_inspection_owner": "王工",
            "discoverer": "李工",
            "issue_level": "B",
            "quantity": "2",
            "issue_type": "装配",
            "completion_date": issue_date,
            "recurring": "否",
            "carrier": "承运商甲",
            "supplier": "供应商甲",
        }, token=tokens["workshop_a"])
        self.assertEqual(status, 201)
        issue_id = created["issue"]["id"]
        self.assertEqual(created["issue"]["batch_no"], "GKMYR26027-06")
        self.assertEqual(created["issue"]["external_inspection_owner"], "王工")
        self.assertEqual(created["issue"]["status"], "draft")
        self.assertEqual(self.call(
            f"/api/workshop/issues?date={issue_date}", token=tokens["workshop_b"],
        )[1]["issues"], [])
        self.assertEqual(self.call(
            f"/api/workshop/issues/{issue_id}/publish", {}, token=tokens["workshop_a"],
        )[0], 400)

        bad_query = urllib.parse.urlencode({"name": "损坏图片.jpg"})
        status, invalid = self.call(
            f"/api/workshop/issues/{issue_id}/images?{bad_query}",
            token=tokens["workshop_a"], raw=b"not-an-image",
            headers={"Content-Length": "12"},
        )
        self.assertEqual(status, 400)
        self.assertIn("损坏", invalid["error"])

        stream = BytesIO()
        Image.new("RGB", (80, 60), "#3d7df0").save(stream, format="PNG")
        image_content = stream.getvalue()
        for index in range(web_server.MAX_WORKSHOP_IMAGES):
            image_query = urllib.parse.urlencode({"name": f"现场-{index + 1}.png"})
            status, uploaded = self.call(
                f"/api/workshop/issues/{issue_id}/images?{image_query}",
                token=tokens["workshop_a"], raw=image_content,
                headers={"Content-Length": str(len(image_content))},
            )
            self.assertEqual(status, 201)
            self.assertEqual(len(uploaded["issue"]["images"]), index + 1)
        status, limited = self.call(
            f"/api/workshop/issues/{issue_id}/images?{bad_query}",
            token=tokens["workshop_a"], raw=image_content,
            headers={"Content-Length": str(len(image_content))},
        )
        self.assertEqual(status, 400)
        self.assertIn("最多上传", limited["error"])

        status, published = self.call(
            f"/api/workshop/issues/{issue_id}/publish", {}, token=tokens["workshop_a"],
        )
        self.assertEqual(status, 200)
        self.assertEqual(published["issue"]["status"], "published")
        status, published_again = self.call(
            f"/api/workshop/issues/{issue_id}/publish", {}, token=tokens["workshop_a"],
        )
        self.assertEqual(status, 200)
        self.assertEqual(published_again["issue"]["status"], "published")
        status, listed = self.call(
            f"/api/workshop/issues?date={issue_date}", token=tokens["workshop_b"],
        )
        self.assertEqual(status, 200)
        self.assertEqual(listed["summary"], {"issue_count": 1, "image_count": 8, "open_count": 1, "resolved_count": 0})
        self.assertFalse(listed["issues"][0]["permissions"]["can_delete"])
        image_url = listed["issues"][0]["images"][0]["url"]
        status, downloaded_image = self.call(image_url, token=tokens["workshop_b"])
        self.assertEqual(status, 200)
        with Image.open(BytesIO(downloaded_image)) as normalized:
            self.assertEqual(normalized.size, (80, 60))
        self.assertEqual(self.call(
            f"/api/workshop/issues/{issue_id}", token=tokens["workshop_b"], method="DELETE",
        )[0], 403)

        status, exported = self.call(
            f"/api/workshop/issues/export?date={issue_date}", token=tokens["workshop_b"],
        )
        self.assertEqual(status, 200)
        workbook = load_workbook(BytesIO(exported))
        self.assertEqual(workbook.sheetnames, ["主料异常", "辅料异常", "包装异常", "海外历史记录", "防错异常", "问题一览表"])
        main_sheet = workbook["主料异常"]
        self.assertEqual(main_sheet["A1"].value, "问题编号")
        self.assertEqual(main_sheet["B1"].value, "主料故障信息")
        self.assertEqual(main_sheet["H2"].value, "故障描述")
        self.assertEqual(main_sheet["H3"].value, "装配工位防护罩松动")
        self.assertEqual(main_sheet["D3"].value, "GKMYR26027-06")
        self.assertEqual(main_sheet["T3"].value, "王工")
        self.assertEqual(main_sheet["T1"].value, None)
        self.assertEqual(main_sheet["T2"].value, "外检责任人")
        self.assertEqual(len(main_sheet._images), 8)
        self.assertEqual(workbook["辅料异常"]["B1"].value, "问题源")
        self.assertEqual(workbook["海外历史记录"]["A1"].value, "海外问题统计")
        self.assertEqual(workbook["防错异常"]["A2"].value, "发生时间")
        self.assertEqual(workbook["问题一览表"]["S2"].value, "供应商")
        self.assertEqual(workbook["问题一览表"].max_row, 2)

        status, removed = self.call(
            f"/api/workshop/issues/{issue_id}", token=tokens["workshop_a"], method="DELETE",
        )
        self.assertEqual(status, 200)
        self.assertIn("回收站", removed["message"])
        trash = self.call("/api/admin/trash", token=self.admin)[1]["trash"]
        item = next(row for row in trash if row["kind"] == "workshop_issue")
        self.assertEqual(self.call(
            f"/api/admin/trash/{item['id']}/restore", {}, token=self.admin,
        )[0], 200)
        restored = self.call(
            f"/api/workshop/issues?date={issue_date}", token=tokens["workshop_b"],
        )[1]
        self.assertEqual(restored["summary"], {"issue_count": 1, "image_count": 8, "open_count": 1, "resolved_count": 0})
        self.assertEqual(restored["issues"][0]["category"], "main_material")
        self.assertEqual(restored["issues"][0]["severity"], "important")

    def test_workshop_error_proofing_publishes_without_images(self):
        """防错异常按模板不强制图片，不能套用其余四类问题的上传限制。"""

        issue_date = datetime.now().strftime("%Y-%m-%d")
        status, created = self.call("/api/workshop/issues", {
            "issue_date": issue_date,
            "category": "error_proofing",
            "cause": "PDA 扫描到相邻二维码",
            "happened_at": f"{issue_date}T09:30",
            "batch_no": "GKIDR26004-02",
            "material_name": "后减振器总成",
            "material_code": "6608116475",
            "cause_analysis": "扫码方向未对准",
            "corrective_action": "质检班长现场核验并解除异常",
            "tracking_status": "校验完成",
            "handling_time": f"{issue_date}T09:40",
            "responsible_person": "黄林",
            "updated_by_name": "杨晓云",
            "notes": "无需上传照片",
        }, token=self.admin)
        self.assertEqual(status, 201)
        issue_id = created["issue"]["id"]
        self.assertEqual(created["issue"]["primary_owner"], "黄林")
        self.assertEqual(created["issue"]["images"], [])

        status, published = self.call(
            f"/api/workshop/issues/{issue_id}/publish", {}, token=self.admin,
        )
        self.assertEqual(status, 200)
        self.assertEqual(published["issue"]["status"], "published")

        status, exported = self.call(
            f"/api/workshop/issues/export?date={issue_date}", token=self.admin,
        )
        self.assertEqual(status, 200)
        workbook = load_workbook(BytesIO(exported))
        proofing_sheet = workbook["防错异常"]
        headers = [cell.value for cell in proofing_sheet[2]]
        self.assertNotIn("故障照片", headers)
        self.assertEqual(proofing_sheet["E3"].value, "PDA 扫描到相邻二维码")
        self.assertEqual(proofing_sheet["L3"].value, "无需上传照片")
        self.assertEqual(len(proofing_sheet._images), 0)

    def test_workshop_issue_export_supports_date_range_and_rejects_reverse_range(self):
        """现场问题导出支持默认日、日期区间，并明确拒绝结束早于开始的范围。"""

        today = web_server.business_today()
        dates = [today - timedelta(days=1), today]
        for index, issue_day in enumerate(dates, 1):
            issue_date = issue_day.isoformat()
            status, created = self.call("/api/workshop/issues", {
                "issue_date": issue_date,
                "category": "error_proofing",
                "cause": f"日期范围问题-{index}",
                "happened_at": f"{issue_date}T09:30",
                "batch_no": f"B-{index:03d}",
                "material_name": "测试物料",
                "material_code": f"M-{index:03d}",
                "cause_analysis": "测试原因",
                "corrective_action": "测试措施",
                "tracking_status": "待处理",
                "responsible_person": "测试负责人",
                "updated_by_name": "测试记录人",
            }, token=self.admin)
            self.assertEqual(status, 201)
            self.assertEqual(self.call(
                f"/api/workshop/issues/{created['issue']['id']}/publish", {}, token=self.admin,
            )[0], 200)

        start_date, end_date = (item.isoformat() for item in dates)
        status, exported = self.call(
            f"/api/workshop/issues/export?start_date={start_date}&end_date={end_date}",
            token=self.admin,
        )
        self.assertEqual(status, 200)
        workbook = load_workbook(BytesIO(exported))
        proofing_sheet = workbook["防错异常"]
        causes = {proofing_sheet.cell(row, 5).value for row in range(3, proofing_sheet.max_row + 1)}
        self.assertEqual(causes, {"日期范围问题-1", "日期范围问题-2"})
        self.assertEqual(workbook["问题一览表"].max_row, 2)

        status, reversed_range = self.call(
            f"/api/workshop/issues/export?start_date={end_date}&end_date={start_date}",
            token=self.admin,
        )
        self.assertEqual(status, 400)
        self.assertIn("开始日期不能晚于结束日期", reversed_range["error"])

    def test_workshop_published_issue_edit_resolve_and_reopen(self):
        """已发布问题应支持乐观并发编辑、填写闭环说明、重新打开和权限校验。"""

        issue_date = datetime.now().strftime("%Y-%m-%d")
        status, created = self.call("/api/workshop/issues", {
            "issue_date": issue_date,
            "category": "error_proofing",
            "cause": "原始问题描述",
            "happened_at": f"{issue_date}T09:30",
            "batch_no": "B-001",
            "material_name": "测试物料",
            "material_code": "M-001",
            "cause_analysis": "原始原因",
            "corrective_action": "原始措施",
            "tracking_status": "待处理",
            "handling_time": "",
            "responsible_person": "王工",
            "updated_by_name": "李工",
            "notes": "原始备注",
        }, token=self.admin)
        self.assertEqual(status, 201)
        issue_id = created["issue"]["id"]
        self.assertEqual(self.call(f"/api/workshop/issues/{issue_id}/publish", {}, token=self.admin)[0], 200)

        listed = self.call(f"/api/workshop/issues?date={issue_date}", token=self.admin)[1]
        current = listed["issues"][0]
        status, edited = self.call(f"/api/workshop/issues/{issue_id}", {
            "expected_updated_at": current["updated_at"],
            "issue_date": issue_date,
            "category": "error_proofing",
            "cause": "修正后的问题描述",
            "happened_at": f"{issue_date}T09:30",
            "batch_no": "B-001",
            "material_name": "测试物料",
            "material_code": "M-001",
            "cause_analysis": "修正后的原因",
            "corrective_action": "修正后的措施",
            "tracking_status": "处理中",
            "handling_time": "",
            "responsible_person": "王工",
            "updated_by_name": "李工",
            "notes": "修正后的备注",
        }, token=self.admin, method="PATCH")
        self.assertEqual(status, 200)
        self.assertEqual(edited["issue"]["cause"], "修正后的问题描述")
        self.assertEqual(edited["issue"]["resolution_status"], "open")

        status, missing_note = self.call(
            f"/api/workshop/issues/{issue_id}/resolve", {}, token=self.admin,
        )
        self.assertEqual(status, 400)
        self.assertIn("解决情况", missing_note["error"])
        status, resolved = self.call(
            f"/api/workshop/issues/{issue_id}/resolve",
            {"expected_updated_at": edited["issue"]["updated_at"], "resolution_note": "已补齐物料并完成现场复核"},
            token=self.admin,
        )
        self.assertEqual(status, 200)
        self.assertEqual(resolved["issue"]["resolution_status"], "resolved")
        self.assertEqual(resolved["issue"]["resolution_note"], "已补齐物料并完成现场复核")
        summary = self.call(f"/api/workshop/issues?date={issue_date}", token=self.admin)[1]["summary"]
        self.assertEqual(summary["open_count"], 0)
        self.assertEqual(summary["resolved_count"], 1)

        status, reopened = self.call(
            f"/api/workshop/issues/{issue_id}/reopen",
            {"expected_updated_at": resolved["issue"]["updated_at"]}, token=self.admin,
        )
        self.assertEqual(status, 200)
        self.assertEqual(reopened["issue"]["resolution_status"], "open")
        self.assertEqual(reopened["issue"]["resolution_note"], "已补齐物料并完成现场复核")

    def test_slow_workshop_upload_does_not_block_other_uploads(self):
        """慢速图片上传不能持有全局请求锁并阻塞其他用户的并发上传。"""

        issue_date = datetime.now().strftime("%Y-%m-%d")
        issue_ids = []
        for cause in ("慢连接问题", "正常连接问题"):
            status, created = self.call("/api/workshop/issues", {
                "issue_date": issue_date, "cause": cause, "primary_owner": "测试负责人",
            }, token=self.admin)
            self.assertEqual(status, 201)
            issue_ids.append(created["issue"]["id"])

        stream = BytesIO()
        Image.new("RGB", (96, 72), "#3d7df0").save(stream, format="JPEG")
        image_content = stream.getvalue()
        slow_query = urllib.parse.urlencode({"name": "慢连接.jpg"})
        slow_socket = socket.create_connection(("127.0.0.1", self.server.server_port), timeout=5)
        slow_socket.settimeout(5)
        headers = (
            f"POST /api/workshop/issues/{issue_ids[0]}/images?{slow_query} HTTP/1.0\r\n"
            f"Host: 127.0.0.1\r\nX-Session-Token: {self.admin}\r\n"
            f"Content-Type: application/octet-stream\r\nContent-Length: {len(image_content)}\r\n\r\n"
        ).encode("ascii")
        split_at = max(1, len(image_content) // 3)
        try:
            slow_socket.sendall(headers + image_content[:split_at])
            time.sleep(0.1)
            fast_query = urllib.parse.urlencode({"name": "正常连接.jpg"})
            started = time.monotonic()
            status, uploaded = self.call(
                f"/api/workshop/issues/{issue_ids[1]}/images?{fast_query}",
                token=self.admin, raw=image_content,
                headers={"Content-Length": str(len(image_content))},
            )
            elapsed = time.monotonic() - started
            self.assertEqual(status, 201)
            self.assertEqual(len(uploaded["issue"]["images"]), 1)
            self.assertLess(elapsed, 2)
            slow_socket.sendall(image_content[split_at:])
            while slow_socket.recv(4096):
                pass
        finally:
            slow_socket.close()

        self.assertEqual(self.call(
            f"/api/workshop/issues/{issue_ids[1]}/publish", {}, token=self.admin,
        )[0], 200)

    def test_workshop_stale_draft_cleanup_removes_isolated_images(self):
        """定期维护应删除过期草稿及孤立图片，同时保留仍被有效问题引用的文件。"""

        issue_date = datetime.now().strftime("%Y-%m-%d")
        status, created = self.call("/api/workshop/issues", {
            "issue_date": issue_date, "cause": "未完成草稿", "primary_owner": "测试负责人",
        }, token=self.admin)
        self.assertEqual(status, 201)
        issue_id = created["issue"]["id"]
        stream = BytesIO()
        Image.new("RGB", (24, 18), "white").save(stream, format="JPEG")
        image_content = stream.getvalue()
        query = urllib.parse.urlencode({"name": "草稿.jpg"})
        self.assertEqual(self.call(
            f"/api/workshop/issues/{issue_id}/images?{query}", token=self.admin,
            raw=image_content, headers={"Content-Length": str(len(image_content))},
        )[0], 201)
        folder = web_server.workshop_issue_dir(1, issue_id)
        self.assertTrue(folder.is_dir())
        now = datetime.now(timezone.utc).replace(microsecond=0)
        old_time = (now - timedelta(hours=25)).isoformat(timespec="seconds")
        with web_server.DB_LOCK, web_server.db() as connection:
            connection.execute(
                "UPDATE workshop_issues SET updated_at = ? WHERE id = ?", (old_time, issue_id)
            )
        self.assertEqual(web_server.cleanup_stale_workshop_drafts(current_time=now), 1)
        self.assertFalse(folder.exists())
        with web_server.DB_LOCK, web_server.db() as connection:
            self.assertIsNone(connection.execute(
                "SELECT id FROM workshop_issues WHERE id = ?", (issue_id,)
            ).fetchone())

    def test_upload_trash_restore_and_permanent_delete(self):
        """普通上传进入回收站后可恢复，永久删除则同时清理文件与元数据。"""

        query = urllib.parse.urlencode({"name": "可恢复资料.txt", "group": "trash-test"})
        status, uploaded = self.call(
            f"/api/files/upload?{query}", token=self.admin, raw=b"recoverable",
            headers={"Content-Length": "11"},
        )
        self.assertEqual(status, 201)
        status, payload = self.call(
            f"/api/admin/uploads/{uploaded['handle']}", token=self.admin, method="DELETE",
        )
        self.assertEqual(status, 200)
        self.assertIn("回收站", payload["message"])
        status, trash = self.call("/api/admin/trash", token=self.admin)
        self.assertEqual(status, 200)
        item = trash["trash"][0]
        self.assertEqual(item["kind"], "upload")
        status, _ = self.call(
            f"/api/admin/trash/{item['id']}/restore", {}, token=self.admin,
        )
        self.assertEqual(status, 200)
        data = self.call("/api/admin/data", token=self.admin)[1]
        self.assertIn(uploaded["handle"], {row["handle"] for row in data["uploads"]})

        self.call(f"/api/admin/uploads/{uploaded['handle']}", token=self.admin, method="DELETE")
        item = self.call("/api/admin/trash", token=self.admin)[1]["trash"][0]
        status, _ = self.call(
            f"/api/admin/trash/{item['id']}", token=self.admin, method="DELETE",
        )
        self.assertEqual(status, 200)
        self.assertEqual(self.call("/api/admin/trash", token=self.admin)[1]["trash"], [])

    def test_job_trash_restores_record_and_result_file(self):
        """任务软删除与恢复必须成对处理数据库记录和所属结果文件。"""

        job_id = "recoverable-job"
        output = web_server.DATA_ROOT / "users" / "1" / "jobs" / job_id / "outputs" / "结果.txt"
        output.parent.mkdir(parents=True)
        output.write_bytes(b"result")
        timestamp = web_server.now_iso()
        with web_server.DB_LOCK, web_server.db() as connection:
            connection.execute(
                "INSERT INTO web_jobs(id, user_id, assignee_id, action, title, status, progress, logs, result, error, files, cancelled, payload, created_at, updated_at) "
                "VALUES (?, 1, 1, 'text.transform', '可恢复任务', 'completed', 100, '[]', '{}', NULL, ?, 0, '{}', ?, ?)",
                (job_id, json.dumps([{"name": "结果.txt", "path": str(output), "size": 6}], ensure_ascii=False), timestamp, timestamp),
            )
        status, payload = self.call(
            f"/api/admin/jobs/{job_id}", token=self.admin, method="DELETE",
        )
        self.assertEqual(status, 200)
        self.assertIn("回收站", payload["message"])
        item = self.call("/api/admin/trash", token=self.admin)[1]["trash"][0]
        self.assertEqual(item["kind"], "job")
        self.assertFalse(output.exists())
        self.assertEqual(self.call(
            f"/api/admin/trash/{item['id']}/restore", {}, token=self.admin,
        )[0], 200)
        self.assertTrue(output.is_file())
        with web_server.DB_LOCK, web_server.db() as connection:
            restored = connection.execute(
                "SELECT assignee_id FROM web_jobs WHERE id = ?", (job_id,),
            ).fetchone()
        self.assertEqual(restored["assignee_id"], 1)
        request = urllib.request.Request(
            self.base + f"/api/jobs/{job_id}/files/0", headers={"X-Session-Token": self.admin},
        )
        with urllib.request.urlopen(request, timeout=10) as response:
            self.assertEqual(response.read(), b"result")

    def test_storage_maintenance_limits_outputs_and_purges_expired_trash(self):
        """每用户只保留最近二十次输出，并按保留期清理回收站而不越权删除。"""

        now = datetime.now(timezone.utc).replace(microsecond=0)
        for index in range(22):
            job_id = f"retention-job-{index:02d}"
            output = web_server.DATA_ROOT / "users" / "1" / "jobs" / job_id / "outputs" / "结果.txt"
            output.parent.mkdir(parents=True)
            output.write_text(str(index), encoding="utf-8")
            timestamp = (now - timedelta(minutes=22 - index)).isoformat(timespec="seconds")
            files = json.dumps(
                [{"name": "结果.txt", "path": str(output), "size": output.stat().st_size}],
                ensure_ascii=False,
            )
            with web_server.DB_LOCK, web_server.db() as connection:
                connection.execute(
                    "INSERT INTO web_jobs(id, user_id, action, title, status, progress, logs, result, error, files, cancelled, payload, created_at, updated_at) "
                    "VALUES (?, 1, 'text.transform', ?, 'completed', 100, '[]', '{}', NULL, ?, 0, '{}', ?, ?)",
                    (job_id, f"输出任务 {index}", files, timestamp, timestamp),
                )
                if index == 0:
                    connection.execute(
                        "INSERT INTO web_job_versions(job_id, user_id, version, result, files, status, created_at) "
                        "VALUES (?, 1, 1, '{}', ?, 'completed', ?)",
                        (job_id, files, timestamp),
                    )

        review_id = "retention-review"
        review_output = web_server.DATA_ROOT / "users" / "1" / "jobs" / review_id / "outputs" / "复核.txt"
        review_output.parent.mkdir(parents=True)
        review_output.write_text("review", encoding="utf-8")
        review_files = json.dumps(
            [{"name": "复核.txt", "path": str(review_output), "size": 6}], ensure_ascii=False,
        )
        old_time = (now - timedelta(days=60)).isoformat(timespec="seconds")
        with web_server.DB_LOCK, web_server.db() as connection:
            connection.execute(
                "INSERT INTO web_jobs(id, user_id, action, title, status, progress, logs, result, error, files, cancelled, payload, created_at, updated_at) "
                "VALUES (?, 1, 'web.compare.review', '等待人工复核', 'completed', 100, '[]', '{}', NULL, ?, 0, '{}', ?, ?)",
                (review_id, review_files, old_time, old_time),
            )
            connection.execute(
                "INSERT INTO trash_items(id, kind, label, record_json, original_path, size, deleted_at) "
                "VALUES ('expired-trash', 'upload', '过期资料', '{}', 'users/1/uploads/expired.txt', 1, ?)",
                ((now - timedelta(days=31)).isoformat(timespec="seconds"),),
            )
            connection.execute(
                "INSERT INTO trash_items(id, kind, label, record_json, original_path, size, deleted_at) "
                "VALUES ('recent-trash', 'upload', '近期资料', '{}', 'users/1/uploads/recent.txt', 1, ?)",
                ((now - timedelta(days=29)).isoformat(timespec="seconds"),),
            )
        for trash_id in ("expired-trash", "recent-trash"):
            payload = web_server.DATA_ROOT / "trash" / trash_id / "payload"
            payload.parent.mkdir(parents=True)
            payload.write_text(trash_id, encoding="utf-8")

        report = web_server.run_storage_maintenance(
            output_limit=20, trash_retention_days=30, current_time=now,
        )
        self.assertEqual(report["moved_outputs"], 2)
        self.assertEqual(report["purged_trash"], 1)
        self.assertEqual(report["trash_cleanup_failures"], 0)
        self.assertFalse((web_server.DATA_ROOT / "trash" / "expired-trash").exists())
        self.assertTrue((web_server.DATA_ROOT / "trash" / "recent-trash" / "payload").is_file())

        with web_server.DB_LOCK, web_server.db() as connection:
            active_ids = {
                row["id"] for row in connection.execute("SELECT id FROM web_jobs").fetchall()
            }
            trash_rows = connection.execute("SELECT id, label FROM trash_items").fetchall()
        self.assertNotIn("retention-job-00", active_ids)
        self.assertNotIn("retention-job-01", active_ids)
        self.assertIn("retention-job-02", active_ids)
        self.assertIn(review_id, active_ids)
        self.assertTrue(review_output.is_file())

        archived = next(row for row in trash_rows if row["label"] == "输出任务 0")
        status, _ = self.call(
            f"/api/admin/trash/{archived['id']}/restore", {}, token=self.admin,
        )
        self.assertEqual(status, 200)
        restored_output = (
            web_server.DATA_ROOT / "users" / "1" / "jobs" / "retention-job-00" / "outputs" / "结果.txt"
        )
        self.assertTrue(restored_output.is_file())
        with web_server.DB_LOCK, web_server.db() as connection:
            versions = connection.execute(
                "SELECT COUNT(*) AS count FROM web_job_versions WHERE job_id = 'retention-job-00'"
            ).fetchone()["count"]
        self.assertEqual(versions, 1)

    def test_backup_verification_and_restore(self):
        """管理员备份需可校验、恢复并在失败路径保持当前数据库可用。"""

        query = urllib.parse.urlencode({"name": "备份资料.txt", "group": "backup-test"})
        status, uploaded = self.call(
            f"/api/files/upload?{query}", token=self.admin, raw=b"backup-content",
            headers={"Content-Length": "14"},
        )
        self.assertEqual(status, 201)
        status, created = self.call("/api/admin/backups", {}, token=self.admin)
        self.assertEqual(status, 201)
        backup_id = created["backup"]["id"]
        backup_path = web_server.DATA_ROOT / "backups" / f"{backup_id}.zip"
        manifest = web_server.verify_web_backup(backup_path)
        self.assertEqual(manifest["backup_id"], backup_id)
        self.call(f"/api/admin/uploads/{uploaded['handle']}", token=self.admin, method="DELETE")

        status, restored = self.call(
            f"/api/admin/backups/{backup_id}/restore",
            {"confirmation": "恢复备份"}, token=self.admin,
        )
        self.assertEqual(status, 200)
        self.assertTrue(restored["safety_backup_id"])
        self.assertEqual(self.call("/api/auth/me", token=self.admin)[0], 401)
        new_admin = self.call(
            "/api/auth/login", {"username": "admin", "password": "admin123456"},
        )[1]["token"]
        data = self.call("/api/admin/data", token=new_admin)[1]
        self.assertIn(uploaded["handle"], {row["handle"] for row in data["uploads"]})

    def test_backup_verification_rejects_duplicate_manifest_paths(self):
        """备份清单不得用重复路径掩盖同一文件的多次登记。"""

        backup_path = web_server.DATA_ROOT / "backups" / "duplicate-manifest.zip"
        backup_path.parent.mkdir(parents=True, exist_ok=True)
        database_payload = b"synthetic-database"
        entry = {
            "path": "database/accounts.sqlite3",
            "size": len(database_payload),
            "sha256": hashlib.sha256(database_payload).hexdigest(),
        }
        manifest = {"format": 1, "files": [entry, dict(entry)]}
        with zipfile.ZipFile(backup_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            archive.writestr("database/accounts.sqlite3", database_payload)
            archive.writestr(
                "manifest.json", json.dumps(manifest, ensure_ascii=False).encode("utf-8"),
            )
        with self.assertRaisesRegex(ValueError, "重复路径"):
            web_server.verify_web_backup(backup_path)

    def test_admin_master_data_upload_review_merge_and_permissions(self):
        """主数据上传必须经过冲突复核、确认合并和管理员权限约束。"""

        status, _ = self.call("/api/auth/register", {
            "username": "catalog_user", "display_name": "普通成员", "password": "password123",
        })
        self.assertEqual(status, 201)
        with web_server.DB_LOCK, web_server.db() as connection:
            member_id = connection.execute(
                "SELECT id FROM users WHERE username = ?", ("catalog_user",),
            ).fetchone()["id"]
        self.assertEqual(self.call(
            f"/api/admin/users/{member_id}/approve", {}, token=self.admin,
        )[0], 200)
        member_token = self.call(
            "/api/auth/login", {"username": "catalog_user", "password": "password123"},
        )[1]["token"]
        self.assertEqual(self.call("/api/admin/master-data/imports", token=member_token)[0], 403)

        # 现有逐条维护 POST 路由必须可用，并写入 Web 数据根而不是用户文档目录。
        status, catalog = self.call("/api/admin/catalog", {
            "op": "upsert_supplier", "name": "已有供应商", "code": "OLD01",
        }, token=self.admin)
        self.assertEqual(status, 200)
        self.assertEqual(catalog["suppliers"]["已有供应商"], "OLD01")
        self.assertTrue((web_server.DATA_ROOT / "catalog.json").is_file())

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "主数据"
        sheet.append(["物料编码", "物料名称", "规格型号", "单位", "供应商名称", "供应商编码"])
        sheet.append(["M001", "铁箱", "100x50", "个", "众瀚", "GYS01"])
        stream = BytesIO()
        workbook.save(stream)
        workbook.close()
        raw = stream.getvalue()
        query = urllib.parse.urlencode({"name": "管理员主数据.xlsx"})
        status, created = self.call(
            f"/api/admin/master-data/imports?{query}", token=self.admin, raw=raw,
            headers={"Content-Length": str(len(raw))},
        )
        self.assertEqual(status, 201)
        batch = created["batch"]
        self.assertEqual(batch["status"], "ready_to_confirm")
        self.assertEqual(batch["candidate_count"], 5)
        source_dir = (
            web_server.DATA_ROOT / "users" / str(batch["uploader_id"])
            / "master-data-imports" / batch["id"]
        )
        self.assertTrue((source_dir / "管理员主数据.xlsx").is_file())

        status, listing = self.call("/api/admin/master-data/imports", token=self.admin)
        self.assertEqual(status, 200)
        self.assertEqual(listing["summary"]["ready_to_confirm"], 1)
        self.assertEqual(self.call(
            f"/api/admin/master-data/imports/{batch['id']}/confirm", {}, token=self.admin,
        )[1]["batch"]["status"], "ready")
        merged = self.call(
            f"/api/admin/master-data/imports/{batch['id']}/merge", {}, token=self.admin,
        )[1]
        self.assertEqual(merged["batch"]["status"], "merged")
        catalog = self.call("/api/admin/catalog", token=self.admin)[1]
        self.assertEqual(catalog["suppliers"]["众瀚"], "GYS01")
        self.assertEqual(catalog["materials"]["M001"]["name"], "铁箱")

        duplicate_status, duplicate = self.call(
            f"/api/admin/master-data/imports?{query}", token=self.admin, raw=raw,
            headers={"Content-Length": str(len(raw))},
        )
        self.assertEqual(duplicate_status, 409)
        self.assertIn("已经上传", duplicate["error"])
        self.assertEqual(self.call("/api/admin/master-data/export", token=self.admin)[0], 200)

        conflict_book = Workbook()
        conflict_sheet = conflict_book.active
        conflict_sheet.append(["物料编码", "物料名称"])
        conflict_sheet.append(["M001", "冲突候选名称"])
        conflict_stream = BytesIO()
        conflict_book.save(conflict_stream)
        conflict_book.close()
        conflict_raw = conflict_stream.getvalue()
        conflict_query = urllib.parse.urlencode({"name": "冲突主数据.xlsx"})
        status, conflict_created = self.call(
            f"/api/admin/master-data/imports?{conflict_query}",
            token=self.admin,
            raw=conflict_raw,
            headers={"Content-Length": str(len(conflict_raw))},
        )
        self.assertEqual(status, 201)
        conflict_batch = conflict_created["batch"]
        self.assertEqual(conflict_batch["status"], "needs_review")
        status, detail = self.call(
            f"/api/admin/master-data/imports/{conflict_batch['id']}", token=self.admin,
        )
        self.assertEqual(status, 200)
        conflict = next(
            item for item in detail["batch"]["candidates"]
            if item["relation_type"] == "material_name"
        )
        status, resolved = self.call(
            f"/api/admin/master-data/imports/{conflict_batch['id']}/resolve",
            {"candidate_id": conflict["id"], "decision": "keep_current"},
            token=self.admin,
        )
        self.assertEqual(status, 200)
        self.assertEqual(resolved["batch"]["status"], "ready_to_confirm")
        status, rejected = self.call(
            f"/api/admin/master-data/imports/{conflict_batch['id']}/reject",
            {}, token=self.admin,
        )
        self.assertEqual(status, 200)
        self.assertEqual(rejected["batch"]["status"], "rejected")


    def test_upload_handle_rejects_cross_user_and_outside_paths(self):
        """上传句柄既不能跨账号复用，也不能通过伪造路径逃出所属上传根。"""

        status, _ = self.call("/api/auth/register", {
            "username": "path_user", "display_name": "路径测试", "password": "password123",
        })
        self.assertEqual(status, 201)
        with web_server.DB_LOCK, web_server.db() as connection:
            member = connection.execute(
                "SELECT id FROM users WHERE username = ?", ("path_user",),
            ).fetchone()
        self.assertEqual(self.call(f"/api/admin/users/{member['id']}/approve", token=self.admin, payload={})[0], 200)
        member_token = self.call(
            "/api/auth/login", {"username": "path_user", "password": "password123"},
        )[1]["token"]

        query = urllib.parse.urlencode({"name": "越权样本.txt", "group": "traversal-group"})
        status, uploaded = self.call(
            f"/api/files/upload?{query}", token=self.admin, raw=b"secret",
            headers={"Content-Length": "6"},
        )
        self.assertEqual(status, 201)

        # 跨用户引用管理员的上传句柄必须被拒绝
        status, payload = self.call("/api/jobs", {
            "action": "rename.apply", "title": "跨用户引用", "payload": {
                "paths": [uploaded["handle"]], "rule": {"prefix": "新-"},
            },
        }, token=member_token)
        self.assertEqual(status, 400)
        self.assertIn("不属于当前账号", payload["error"])

        # 普通账号提交 DATA_ROOT 之外的绝对路径必须被拒绝
        outside = web_server.DATA_ROOT.parent / "机密文件.txt"
        outside.write_text("敏感内容", encoding="utf-8")
        try:
            status, payload = self.call("/api/jobs", {
                "action": "rename.apply", "title": "越权路径", "payload": {
                    "paths": [str(outside)], "rule": {"prefix": "新-"},
                },
            }, token=member_token)
            self.assertEqual(status, 400)
            self.assertIn("不属于当前账号", payload["error"])
            status, payload = self.call("/api/jobs/preflight", {
                "action": "web.compare", "payload": {"file1": [str(outside)], "file2": [str(outside)]},
            }, token=member_token)
            self.assertEqual(status, 400)
            self.assertIn("不属于当前账号", payload["error"])
        finally:
            outside.unlink(missing_ok=True)

    def test_job_file_download_rejects_path_outside_owned_roots(self):
        """任务文件下载只允许任务所属输出根内路径，数据库伪造绝对路径也必须被拒绝。"""

        query = urllib.parse.urlencode({"name": "结果样本.txt", "group": "download-root"})
        status, uploaded = self.call(
            f"/api/files/upload?{query}", token=self.admin, raw=b"sample",
            headers={"Content-Length": "6"},
        )
        self.assertEqual(status, 201)
        status, created = self.call("/api/jobs", {
            "action": "rename.apply", "title": "下载校验", "payload": {
                "paths": [uploaded["handle"]], "rule": {"prefix": "新-"},
            },
        }, token=self.admin)
        self.assertEqual(status, 202)
        job = self.wait_job(created["job_id"])
        self.assertEqual(job["status"], "completed")
        self.assertEqual(len(job["files"]), 1)

        # 把结果记录篡改为任务目录之外的路径，下载必须 404
        outside = web_server.DATA_ROOT.parent / "越权下载.txt"
        outside.write_text("机密", encoding="utf-8")
        try:
            with web_server.DB_LOCK, web_server.db() as connection:
                connection.execute(
                    "UPDATE web_jobs SET files = ? WHERE id = ?",
                    (json.dumps([{"name": "越权下载.txt", "path": str(outside), "size": outside.stat().st_size}]), job["id"]),
                )
            status, payload = self.call(job["files"][0]["url"], token=self.admin)
            self.assertEqual(status, 404)
            self.assertIn("结果文件不存在", payload["error"])
            status, payload = self.call(job["files"][0]["url"] + "/preview", token=self.admin)
            self.assertEqual(status, 404)
            self.assertIn("结果文件不存在", payload["error"])
        finally:
            outside.unlink(missing_ok=True)

    def test_daily_auto_backup_and_rollover(self):
        """跨业务日维护应每天只自动备份一次并正确推进日切状态。"""

        # 当天第一次：创建自动备份
        report = web_server.auto_backup_if_due()
        self.assertTrue(report.startswith("已创建自动备份 auto-"))
        backups = list((web_server.DATA_ROOT / "backups").glob("auto-*.zip"))
        self.assertEqual(len(backups), 1)
        # 同一天再次调用：跳过
        self.assertEqual(web_server.auto_backup_if_due(), "")
        # 滚动保留：制造超龄自动备份后触发清理（默认保留 7 份）
        keep = web_server.AUTO_BACKUP_KEEP
        for _ in range(keep + 2):
            path = web_server.DATA_ROOT / "backups" / f"auto-old-{uuid.uuid4().hex}.zip"
            path.write_bytes(b"fake")
        (web_server.DATA_ROOT / "auto_backup_state.json").unlink(missing_ok=True)
        web_server.auto_backup_if_due()
        remaining = sorted((web_server.DATA_ROOT / "backups").glob("auto-*.zip"))
        self.assertLessEqual(len(remaining), keep + 1)
        # 手动备份不受滚动清理影响
        manual = web_server.DATA_ROOT / "backups" / "manual-1.zip"
        manual.write_bytes(b"fake")
        web_server.auto_backup_if_due()
        self.assertTrue((web_server.DATA_ROOT / "backups" / "manual-1.zip").exists())

    def test_download_action_written_to_audit(self):
        """受保护文件下载属于管理行为，成功后必须写入可追溯审计记录。"""

        upload_query = urllib.parse.urlencode({"name": "审计文件.txt", "group": "audit-test"})
        status, uploaded = self.call(
            f"/api/files/upload?{upload_query}",
            token=self.admin, raw=b"audit\n", headers={"Content-Length": "6"},
        )
        self.assertEqual(status, 201)
        status, created = self.call("/api/jobs", {
            "action": "rename.apply", "title": "审计下载", "payload": {
                "paths": [uploaded["handle"]], "rule": {"prefix": "新-"},
            },
        }, token=self.admin)
        self.assertEqual(status, 202)
        job = self.wait_job(created["job_id"])
        self.assertEqual(job["status"], "completed")
        self.assertEqual(len(job["files"]), 1)
        request = urllib.request.Request(
            self.base + job["files"][0]["url"],
            headers={"X-Session-Token": self.admin},
        )
        with urllib.request.urlopen(request, timeout=10) as response:
            self.assertEqual(response.status, 200)
        _, payload = self.call("/api/admin/audit", token=self.admin)
        actions = [row["action"] for row in payload["audit"]]
        self.assertTrue(any(action.startswith("download_job:") for action in actions))

    def test_login_failure_lock_configurable(self):
        """登录失败阈值和锁定时长应服从配置，同时保持原子计数。"""

        original = (web_server.LOGIN_FAILURE_LIMIT, web_server.LOGIN_LOCK_SECONDS)
        try:
            web_server.LOGIN_FAILURE_LIMIT = 3
            web_server.LOGIN_LOCK_SECONDS = 300
            for _ in range(3):
                status, _ = self.call("/api/auth/login", {"username": "admin", "password": "wrong-pass-123"})
                self.assertIn(status, (401, 429))
            status, payload = self.call("/api/auth/login", {"username": "admin", "password": "wrong-pass-123"})
            self.assertEqual(status, 429)
            self.assertIn("重试", payload["error"])
            # 正确密码也被锁定拦截
            status, _ = self.call("/api/auth/login", {"username": "admin", "password": "admin123456"})
            self.assertEqual(status, 429)
        finally:
            web_server.LOGIN_FAILURE_LIMIT, web_server.LOGIN_LOCK_SECONDS = original


    def test_webhook_notify_on_job_completion(self):
        """任务完成通知应异步调用配置的 Webhook，失败不能反向改变任务成功状态。"""

        received = []
        class Receiver(http.server.BaseHTTPRequestHandler):
            """记录任务通知请求体的最小本机 Webhook 接收器。"""

            def do_POST(self):
                """读取请求体后返回成功，供异步通知线程完成调用。"""

                length = int(self.headers.get("Content-Length") or 0)
                received.append(self.rfile.read(length))
                self.send_response(200)
                self.end_headers()
                self.wfile.write(b"ok")
            def log_message(self, *args):
                """关闭访问日志，避免后台接收器污染测试输出。"""

                pass
        receiver = http.server.ThreadingHTTPServer(("127.0.0.1", 0), Receiver)
        receiver_thread = threading.Thread(target=receiver.serve_forever, daemon=True)
        receiver_thread.start()
        original = web_server.NOTIFY_WEBHOOK_URL
        web_server.NOTIFY_WEBHOOK_URL = f"http://127.0.0.1:{receiver.server_port}/hook"
        try:
            status, created = self.call("/api/jobs", {
                "action": "text.transform", "title": "推送测试任务", "payload": {
                    "text": "甲\n乙", "operation": "dedup",
                },
            }, token=self.admin)
            self.assertEqual(status, 202)
            job = self.wait_job(created["job_id"])
            self.assertEqual(job["status"], "completed")
            deadline = time.monotonic() + 8
            while not received and time.monotonic() < deadline:
                time.sleep(0.1)
            self.assertTrue(received, "webhook 未收到推送")
            body = json.loads(received[0].decode("utf-8"))
            self.assertEqual(body["msgtype"], "text")
            self.assertIn("推送测试任务", body["text"]["content"])
        finally:
            web_server.NOTIFY_WEBHOOK_URL = original
            receiver.shutdown()
            receiver.server_close()


    def test_share_link_anonymous_download_and_revoke(self):
        """分享链接允许受限匿名下载，撤销或过期后必须立即失效。"""

        upload_query = urllib.parse.urlencode({"name": "分享文件.txt", "group": "share-test"})
        status, uploaded = self.call(
            f"/api/files/upload?{upload_query}",
            token=self.admin, raw=b"share-me", headers={"Content-Length": "8"},
        )
        self.assertEqual(status, 201)
        status, created = self.call("/api/jobs", {
            "action": "rename.apply", "title": "分享测试", "payload": {
                "paths": [uploaded["handle"]], "rule": {"prefix": "新-"},
            },
        }, token=self.admin)
        self.assertEqual(status, 202)
        job = self.wait_job(created["job_id"])
        self.assertEqual(job["status"], "completed")

        # 创建分享
        status, share = self.call("/api/shares", {
            "job_id": created["job_id"], "file_index": 0, "expires_in_days": 7,
        }, token=self.admin)
        self.assertEqual(status, 200)
        self.assertTrue(share["url"].startswith("/api/shares/"))
        # 匿名下载（不带 token）
        request = urllib.request.Request(self.base + share["url"])
        with urllib.request.urlopen(request, timeout=10) as response:
            self.assertEqual(response.read(), b"share-me")
        # 撤销后匿名下载被拒
        status, _ = self.call(f"/api/shares/{share['token']}", token=self.admin, method="DELETE")
        self.assertEqual(status, 200)
        with self.assertRaises(urllib.error.HTTPError) as context:
            urllib.request.urlopen(urllib.request.Request(self.base + share["url"]), timeout=10)
        self.assertEqual(context.exception.code, 410)
        # 未登录不能撤销分享
        status, _ = self.call(f"/api/shares/{share['token']}", token="", method="DELETE")
        self.assertIn(status, (401, 403))


    def test_assign_job_to_user_and_notify(self):
        """管理员分派任务后目标用户应获得权限与通知，其他账号仍不可访问。"""

        status, created = self.call("/api/jobs", {
            "action": "text.transform", "title": "指派测试", "payload": {
                "text": "行\n列", "operation": "dedup",
            },
        }, token=self.admin)
        self.assertEqual(status, 202)
        job = self.wait_job(created["job_id"])
        self.assertEqual(job["status"], "completed")
        me = self.call("/api/auth/me", token=self.admin)[1]["user"]
        status, payload = self.call(
            f"/api/jobs/{created['job_id']}/assign",
            {"assignee_id": me["id"]}, token=self.admin,
        )
        self.assertEqual(status, 200)
        data = self.call("/api/admin/data", token=self.admin)[1]
        job_row = next(row for row in data["jobs"] if row["id"] == created["job_id"])
        self.assertEqual(job_row["assignee_id"], me["id"])
        self.assertEqual(job_row["assignee_display_name"], me["display_name"])
        # 被指派账号的消息中心收到提醒
        notifications = self.call("/api/notifications", token=self.admin)[1]
        self.assertTrue(any("需要你确认" in item["title"] for item in notifications["notifications"]))
        # 取消指派
        status, _ = self.call(
            f"/api/jobs/{created['job_id']}/assign",
            {"assignee_id": None}, token=self.admin,
        )
        self.assertEqual(status, 200)
        data = self.call("/api/admin/data", token=self.admin)[1]
        job_row = next(row for row in data["jobs"] if row["id"] == created["job_id"])
        self.assertIsNone(job_row["assignee_id"])

if __name__ == "__main__":
    unittest.main()
