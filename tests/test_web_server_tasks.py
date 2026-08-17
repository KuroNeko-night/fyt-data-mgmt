# -*- coding: utf-8 -*-
"""Web 服务任务流水线、日清业务与模板复核回归测试。"""
from __future__ import annotations

import json
import urllib.parse
import urllib.request
from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path
from unittest import mock

from openpyxl import Workbook, load_workbook

import web_server

from tests.web_server_test_base import WebServerTestBase


class WebServerTaskTests(WebServerTestBase):
    """任务创建、上传下载、到料、日清、模板与人工复核主链路。"""

    def test_text_task_and_upload_download(self):
        """覆盖任务创建、二进制上传、重命名、版本下载、预览和工作台统计主链路。"""

        status, created = self.call("/api/jobs", {
            "action": "text.transform", "title": "测试文本", "payload": {
                "text": "乙\n甲\n乙", "operation": "dedup",
            },
        }, token=self.admin)
        self.assertEqual(status, 202)  # 任务受理
        job = self.wait_job(created["job_id"])
        self.assertEqual(job["status"], "completed")  # 任务完成
        self.assertEqual(job["result"]["text"], "乙\n甲")  # 去重结果

        upload_query = urllib.parse.urlencode({"name": "原文件.txt", "group": "test-group"})
        status, uploaded = self.call(
            f"/api/files/upload?{upload_query}",
            token=self.admin, raw=b"sample\n", headers={"Content-Length": "7"},
        )
        self.assertEqual(status, 201)  # 上传成功
        status, created = self.call("/api/jobs", {
            "action": "rename.apply", "title": "测试重命名", "payload": {
                "paths": [uploaded["handle"]], "rule": {"prefix": "新-"},
            },
        }, token=self.admin)
        self.assertEqual(status, 202)  # 任务受理
        job = self.wait_job(created["job_id"])
        self.assertEqual(job["status"], "completed")  # 任务完成
        self.assertEqual(len(job["files"]), 1)  # 一个输出
        request = urllib.request.Request(self.base + job["files"][0]["url"], headers={"X-Session-Token": self.admin})
        with urllib.request.urlopen(request, timeout=10) as response:
            self.assertEqual(response.read(), b"sample\n")  # 下载内容一致
            self.assertEqual(response.headers["Cache-Control"], "no-store")  # 下载禁缓存
            self.assertTrue(response.headers["Content-Disposition"].startswith("attachment;"))  # 附件头
        # 路由表必须优先识别历史版本文件，不能落入普通任务文件下载。
        version_url = job["versions"][0]["files"][0]["url"]
        request = urllib.request.Request(self.base + version_url, headers={"X-Session-Token": self.admin})
        with urllib.request.urlopen(request, timeout=10) as response:
            self.assertEqual(response.read(), b"sample\n")  # 版本文件可下载
        status, preview = self.call(job["files"][0]["url"] + "/preview", token=self.admin)
        self.assertEqual(status, 200)  # 预览成功
        self.assertEqual(preview["rows"], [["sample"]])

        status, board = self.call("/api/dashboard", token=self.admin)
        self.assertEqual(status, 200)  # 工作台可访问
        self.assertEqual(board["metrics"]["completed_jobs"], 2)  # 完成任务计数
        self.assertEqual(len(board["trend"]), 7)  # 七日趋势
        self.assertIn("text", {item["key"] for item in board["feature_usage"]})  # 功能使用统计
        self.assertGreaterEqual(len(board["recent_files"]), 1)  # 最近文件

    def test_arrival_scan_and_manual_total_override(self):
        """Web 到料扫描应统计完整源表，正式任务必须采用人工覆盖后的总类数。"""

        stream = BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "订单 TEST2601 批次"
        sheet.append(["物料编码", "物料名称", "供应商信息", "需求数", "剩余未收数"])
        sheet.append(["A-01", "已到物料", "供应商甲", 10, 0])
        sheet.append(["B-02", "隐藏缺料", "供应商乙", 12, 2])
        sheet.append(["C-03", "负数缺料", "供应商丙", 8, -1])
        sheet.row_dimensions[4].hidden = True  # 隐藏缺料行
        workbook.save(stream)
        workbook.close()
        content = stream.getvalue()

        query = urllib.parse.urlencode({"name": "送货计划.xlsx", "group": "arrival-scan"})
        status, uploaded = self.call(
            f"/api/files/upload?{query}", token=self.admin, raw=content,
            headers={"Content-Length": str(len(content))},
        )
        self.assertEqual(status, 201)

        status, scanned = self.call(
            "/api/arrival/scan", {"paths": [uploaded["handle"]]}, token=self.admin,
        )
        self.assertEqual(status, 200)  # 扫描成功
        self.assertEqual(len(scanned["rows"]), 1)  # 一个批次
        row = scanned["rows"][0]
        self.assertEqual(row["batch_no"], "TEST2601")  # 批次号识别
        self.assertEqual(row["total"], 3)  # 总类数
        self.assertEqual(row["auto_total"], 3)  # 自动总类数
        self.assertEqual(row["missing_count"], 2)  # 缺料条数
        self.assertTrue(str(row["path"]).startswith("upload:"))  # 使用上传句柄
        self.assertNotIn(str(self.temp.name), json.dumps(row, ensure_ascii=False))  # 不泄露服务器路径

        row["total"] = 5
        row["remark"] = "人工确认总类数"
        status, created = self.call("/api/jobs", {
            "action": "web.arrival",
            "title": "到料人工参数回归",
            "payload": {"rows": [row], "top_label": "截止 16 点"},
        }, token=self.admin)
        self.assertEqual(status, 202)  # 任务受理
        job = self.wait_job(created["job_id"])
        self.assertEqual(job["status"], "completed", job.get("error"))  # 任务完成
        result = job["result"]["result"]
        self.assertEqual(result["results"][0][3], 5)  # 人工总类数覆盖
        self.assertEqual(result["batches"][0]["total_count"], 5)
        self.assertEqual(result["batches"][0]["missing_count"], 2)  # 缺料保留

    def test_daily_report_admin_scope_result_projection_and_export(self):
        """日清看板只能由管理员读取，并只汇总当天且属于管理员视角的业务投影。"""

        self.assertEqual(self.call("/api/auth/register", {
            "username": "daily_member", "display_name": "计划员乙", "password": "password123",
        })[0], 201)
        member = next(
            item for item in self.call("/api/admin/users", token=self.admin)[1]["users"]
            if item["username"] == "daily_member"
        )
        self.assertEqual(
            self.call(f"/api/admin/users/{member['id']}/approve", {}, token=self.admin)[0],
            200,
        )
        member_token = self.call("/api/auth/login", {
            "username": "daily_member", "password": "password123",
        })[1]["token"]

        report_date = web_server.business_today().isoformat()
        start, _ = web_server.business_day_bounds(web_server.business_today())
        completed_at = (datetime.fromisoformat(start) + timedelta(hours=2)).isoformat(timespec="seconds")
        previous_at = (datetime.fromisoformat(start) - timedelta(seconds=1)).isoformat(timespec="seconds")
        created_at = completed_at
        rows = [
            ("daily-arrival-admin", 1, "管理员到料", {
                "results": [["26035-01", 2, 8, 10]],
                "batches": [{
                    "batch_no": "26035-01", "missing_count": 2,
                    "arrived_count": 8, "total_count": 10,
                    "missing_materials": [{
                        "material_code": "A-01", "material_name": "固定螺栓",
                        "supplier": "供应商甲", "demand_quantity": 12,
                        "received_quantity": 9, "shortage_quantity": 3,
                    }],
                }],
            }, completed_at),
            ("daily-arrival-member", member["id"], "成员到料", {
                "results": [["26035-02", 0, 5, 5]],
            }, completed_at),
            ("daily-arrival-old", 1, "前一日到料", {
                "results": [["OLD", 1, 1, 2]],
            }, previous_at),
        ]
        with web_server.DB_LOCK, web_server.db() as connection:
            for job_id, user_id, title, result_payload, updated_at in rows:
                connection.execute(
                    "INSERT INTO web_jobs(id, user_id, action, title, status, progress, logs, result, "
                    "files, cancelled, payload, created_at, updated_at) "
                    "VALUES (?, ?, 'web.arrival', ?, 'completed', 100, '[]', ?, '[]', 0, '{}', ?, ?)",
                    (job_id, user_id, title, json.dumps({"result": result_payload, "logs": [], "task_id": job_id, "out_dir": "到料明细"}, ensure_ascii=False), created_at, updated_at),
                )
            connection.execute(
                "INSERT INTO workshop_issues(id, user_id, issue_date, cause, primary_owner, "
                "secondary_owner, notes, status, created_at, updated_at) "
                "VALUES ('daily-issue', ?, ?, '工位防护罩松动', '张工', '李工', '下班前复核', "
                "'published', ?, ?)",
                (member["id"], report_date, completed_at, completed_at),
            )

        status, denied = self.call(
            f"/api/daily-report?date={report_date}", token=member_token,
        )
        self.assertEqual(status, 403)
        self.assertIn("管理员", denied["error"])

        status, snapshot = self.call(
            f"/api/daily-report?date={report_date}", token=self.admin,
        )
        self.assertEqual(status, 200)
        self.assertEqual(snapshot["arrival"]["job_count"], 2)
        self.assertEqual(snapshot["arrival"]["batch_count"], 2)
        self.assertEqual(snapshot["arrival"]["total_categories"], 15)
        self.assertEqual(snapshot["arrival"]["completion_rate"], 86.7)
        self.assertEqual(snapshot["arrival"]["missing_material_detail_count"], 1)
        self.assertEqual(snapshot["arrival"]["batches"][0]["missing_materials"][0]["material_code"], "A-01")
        self.assertEqual(snapshot["workshop"]["issue_count"], 1)
        self.assertEqual(snapshot["workshop"]["owner_distribution"][0], {"owner": "张工", "count": 1})
        self.assertNotIn(str(self.temp.name), json.dumps(snapshot, ensure_ascii=False))

        status, jobs = self.call("/api/jobs", token=self.admin)
        self.assertEqual(status, 200)
        projected = next(item for item in jobs["jobs"] if item["id"] == "daily-arrival-admin")
        self.assertEqual(projected["presentation"]["kind"], "arrival")
        self.assertEqual(projected["presentation"]["metrics"][1]["value"], "80.0%")
        self.assertEqual(projected["presentation"]["sections"][1]["rows"][0]["shortage_quantity"], "3")

        status, content = self.call(
            f"/api/daily-report/export?date={report_date}", token=self.admin,
        )
        self.assertEqual(status, 200)
        workbook = load_workbook(BytesIO(content), data_only=True)
        try:
            for sheet_name in (
                "日清概览", "每日到料", "未到物料", "安全检查日报", "现场问题",
                "月度生产订单台账", "订单缺件明细", "订单危包明细", "零星订单明细",
            ):
                self.assertIn(sheet_name, workbook.sheetnames)
            self.assertEqual(workbook["每日到料"]["A2"].value, "26035-01")
            self.assertEqual(workbook["未到物料"]["C2"].value, "A-01")
            self.assertEqual(workbook["未到物料"]["H2"].value, "3")
            self.assertEqual(workbook["现场问题"]["C2"].value, "工位防护罩松动")
        finally:
            workbook.close()

    def test_daily_report_manual_attendance_briefs_and_production_plan(self):
        """人工考勤、事项和生产计划应共同进入日清快照及导出报告。"""

        report_date = web_server.business_today().isoformat()
        status, created = self.call("/api/admin/daily-people", {
            "name": "林楠栗", "person_type": "participant", "unit": "现场负责人", "shift": "", "sort_order": 1, "active": True,
        }, token=self.admin)
        self.assertEqual(status, 201)
        person = created["person"]
        status, rejected = self.call("/api/admin/daily-people", {
            "name": "生产甲", "person_type": "production", "unit": "小件组", "shift": "白班", "sort_order": 1, "active": True,
        }, token=self.admin)
        self.assertEqual(status, 400)
        self.assertIn("按班组维护", rejected["error"])
        status, created_group = self.call("/api/admin/daily-production-groups", {
            "name": "小件组", "sort_order": 1, "active": True,
            "shifts": [
                {"name": "白班", "staffing_count": 14, "sort_order": 1, "active": True},
                {"name": "夜班", "staffing_count": 14, "sort_order": 2, "active": True},
            ],
        }, token=self.admin)
        self.assertEqual(status, 201)
        production_group = created_group["group"]
        self.assertEqual(production_group["staffing_count"], 28)
        self.assertEqual([item["name"] for item in production_group["shifts"]], ["白班", "夜班"])

        status, attendance = self.call(
            f"/api/admin/daily-attendance?date={report_date}", token=self.admin,
        )
        self.assertEqual(status, 200)
        self.assertEqual(len(attendance["attendance"]), 1)
        self.assertTrue(all(item["present"] for item in attendance["attendance"]))
        self.assertEqual(attendance["production_groups"][0]["group_name"], "小件组")
        self.assertEqual(len(attendance["production_groups"]), 2)
        shifts = {item["shift_name"]: item for item in attendance["production_groups"]}
        self.assertEqual(shifts["白班"]["staffing_count"], 14)

        status, saved = self.call("/api/admin/daily-attendance", {
            "date": report_date,
            "records": [
                {"person_id": person["id"], "present": True, "status": "present", "reason": ""},
            ],
            "production_groups": [
                {"shift_id": shifts["白班"]["shift_id"], "attendance_count": 17, "note": "3 人支援小件白班"},
                {"shift_id": shifts["夜班"]["shift_id"], "attendance_count": 14, "note": ""},
            ],
        }, token=self.admin)
        self.assertEqual(status, 200)
        self.assertEqual(saved["production_groups"][0]["attendance_count"], 17)
        self.assertEqual(saved["production_groups"][0]["difference"], -3)
        self.assertEqual(saved["production_groups"][0]["note"], "3 人支援小件白班")

        status, brief = self.call("/api/admin/daily-brief-items", {
            "report_date": report_date, "category": "escalation", "unit": "采购",
            "owner": "户子丹", "title": "关键物料未按计划到场", "description": "需要供应商升级处理",
            "due_date": report_date, "progress": "已联系供应商", "status": "in_progress",
        }, token=self.admin)
        self.assertEqual(status, 201)
        self.assertEqual(brief["item"]["category"], "escalation")
        self.assertEqual(brief["item"]["description"], "")
        self.assertEqual(brief["item"]["due_date"], "")
        self.assertEqual(brief["item"]["progress"], "")
        self.assertEqual(brief["item"]["status"], "open")
        status, _payload = self.call("/api/admin/daily-brief-items", {
            "report_date": report_date, "category": "safety", "unit": "安环",
            "owner": "张工", "title": "旧安全事项", "status": "open",
        }, token=self.admin)
        self.assertEqual(status, 400)
        status, todo = self.call("/api/admin/daily-brief-items", {
            "report_date": report_date, "category": "meeting_todo", "unit": "生产",
            "owner": "李工", "title": "确认次日排产", "description": "会后完成核对",
            "due_date": report_date, "progress": "待确认", "status": "in_progress",
        }, token=self.admin)
        self.assertEqual(status, 201)
        self.assertEqual(todo["item"]["due_date"], report_date)

        stream = BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "计划汇总"
        sheet.append(["班次", "计划", "实际", "差异"])
        sheet.append(["白班", 45, 41, -4])
        workbook.save(stream)
        workbook.close()
        plan_content = stream.getvalue()
        query = urllib.parse.urlencode({"name": "当天生产计划.xlsx", "date": report_date})
        status, uploaded = self.call(
            f"/api/admin/daily-production-plans?{query}", token=self.admin, raw=plan_content,
            headers={"Content-Length": str(len(plan_content))},
        )
        self.assertEqual(status, 201)
        self.assertEqual(uploaded["plan"]["summary"]["sheet_count"], 1)
        plan_id = uploaded["plan"]["id"]
        with web_server.DB_LOCK, web_server.db() as connection:
            plan_file = Path(connection.execute(
                "SELECT path FROM daily_production_plans WHERE id = ?", (plan_id,),
            ).fetchone()["path"])

        status, snapshot = self.call(f"/api/daily-report?date={report_date}", token=self.admin)
        self.assertEqual(status, 200)
        self.assertEqual(snapshot["attendance"]["present_count"], 32)
        self.assertEqual(snapshot["attendance"]["absent_count"], 0)
        self.assertEqual(snapshot["attendance"]["production_present_count"], 31)
        self.assertEqual(snapshot["attendance"]["production_staffing_count"], 28)
        self.assertEqual(snapshot["attendance"]["production_difference"], -3)
        self.assertEqual(snapshot["attendance"]["production_group_count"], 1)
        self.assertEqual(snapshot["attendance"]["production_shift_count"], 2)
        self.assertEqual(snapshot["attendance"]["production_groups"][0]["group_name"], "小件组")
        self.assertEqual(snapshot["brief_items"][0]["title"], "关键物料未按计划到场")
        self.assertEqual(snapshot["production_plans"][0]["original_name"], "当天生产计划.xlsx")

        status, exported = self.call(
            f"/api/daily-report/export?date={report_date}", token=self.admin,
        )
        self.assertEqual(status, 200)
        workbook = load_workbook(BytesIO(exported), data_only=True)
        try:
            self.assertIn("每日考勤", workbook.sheetnames)
            self.assertIn("生产出勤", workbook.sheetnames)
            self.assertEqual(workbook["生产出勤"]["A2"].value, "小件组")
            self.assertEqual(workbook["生产出勤"]["B2"].value, "白班")
            self.assertEqual(workbook["生产出勤"]["C2"].value, 14)
            self.assertEqual(workbook["生产出勤"]["D2"].value, 17)
            self.assertEqual(workbook["生产出勤"]["E2"].value, -3)
            self.assertEqual(workbook["生产出勤"]["C4"].value, 28)
            self.assertEqual(workbook["生产出勤"]["D4"].value, 31)
            self.assertEqual(workbook["生产出勤"]["E4"].value, -3)
            self.assertIn("重点事项与通报", workbook.sheetnames)
            self.assertIn("生产计划", workbook.sheetnames)
        finally:
            workbook.close()

        status, removed = self.call(
            f"/api/admin/daily-production-plans/{plan_id}", token=self.admin, method="DELETE",
        )
        self.assertEqual(status, 200)
        self.assertIn("回收站", removed["message"])
        self.assertFalse(plan_file.parent.exists())
        self.assertEqual(
            self.call(f"/api/daily-report?date={report_date}", token=self.admin)[1]["production_plans"],
            [],
        )
        trash = self.call("/api/admin/trash", token=self.admin)[1]["trash"]
        item = next(row for row in trash if row["kind"] == "daily_production_plan")
        self.assertEqual(item["label"], "当天生产计划.xlsx")
        self.assertEqual(self.call(
            f"/api/admin/trash/{item['id']}/restore", {}, token=self.admin,
        )[0], 200)
        restored_plan = self.call(
            f"/api/daily-report?date={report_date}", token=self.admin,
        )[1]["production_plans"][0]
        self.assertEqual(restored_plan["id"], plan_id)
        self.assertEqual(restored_plan["original_name"], "当天生产计划.xlsx")
        self.assertTrue(plan_file.is_file())

        self.assertEqual(self.call("/api/admin/daily-people")[0], 401)

    def test_daily_source_uploads_feed_arrival_and_safety_dashboard(self):
        """管理员直接上传成品到料与安全检查表后，看板应无需重跑业务模块即可展示。"""

        report_date = web_server.business_today().isoformat()

        arrival_stream = BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        sheet["C1"] = "截止16点的数据"
        sheet["C2"] = "TEST2608"
        sheet["C3"] = "主料总共类"
        sheet["E3"] = 3
        sheet["C4"] = "到货数量"
        sheet["E4"] = 2
        sheet["C5"] = "差异"
        sheet["E5"] = 1
        for column, title in enumerate(
            ["序号", "物料编码", "物料名称", "供应商信息", "需求数", "剩余未收数", "备注"],
            start=3,
        ):
            sheet.cell(7, column, title)
        for column, value in enumerate([1, "A-01", "测试物料", "供应商甲", 10, -2, ""], start=3):
            sheet.cell(8, column, value)
        workbook.save(arrival_stream)
        workbook.close()
        arrival_content = arrival_stream.getvalue()
        arrival_query = urllib.parse.urlencode({
            "kind": "arrival", "date": report_date,
            "name": f"{report_date.replace('-', '')}每日主料到料明细.xlsx",
        })
        status, arrival_upload = self.call(
            f"/api/admin/daily-source-uploads?{arrival_query}", token=self.admin,
            raw=arrival_content, headers={"Content-Length": str(len(arrival_content))},
        )
        self.assertEqual(status, 201)
        self.assertEqual(arrival_upload["upload"]["kind"], "arrival")

        safety_stream = BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "安全检查日报"
        sheet["H1"] = datetime.strptime(report_date, "%Y-%m-%d").date()
        sheet.append(["检查类别", "序号", "检查项目", "安全标准要求", "检查结果", "问题描述", "整改措施", "责任人"])
        sheet.append(["人员安全", 1, "劳保用品", "按规定佩戴", "合格", "佩戴规范", "持续检查", "张工"])
        sheet.append(["设备安全", 2, "交叉作业", "作业区域隔离", "不合格", "隔离带缺失", "当天补齐", "李工"])
        workbook.save(safety_stream)
        workbook.close()
        safety_content = safety_stream.getvalue()
        safety_query = urllib.parse.urlencode({
            "kind": "safety", "date": report_date, "name": "安全检查记录.xlsx",
        })
        status, safety_upload = self.call(
            f"/api/admin/daily-source-uploads?{safety_query}", token=self.admin,
            raw=safety_content, headers={"Content-Length": str(len(safety_content))},
        )
        self.assertEqual(status, 201)
        self.assertEqual(safety_upload["upload"]["summary"]["unqualified_count"], 1)

        status, snapshot = self.call(f"/api/daily-report?date={report_date}", token=self.admin)
        self.assertEqual(status, 200)
        self.assertEqual(snapshot["arrival"]["upload_count"], 1)
        self.assertEqual(snapshot["arrival"]["batch_count"], 1)
        self.assertEqual(snapshot["arrival"]["missing_material_detail_count"], 1)
        self.assertEqual(snapshot["arrival"]["batches"][0]["missing_materials"][0]["shortage_quantity"], "2")
        self.assertEqual(snapshot["arrival"]["batches"][0]["missing_materials"][0]["received_quantity"], "8")
        self.assertEqual(snapshot["safety_checks"]["total_checks"], 2)
        self.assertEqual(snapshot["safety_checks"]["unqualified_count"], 1)
        self.assertEqual(len(snapshot["source_uploads"]), 2)

        source_id = safety_upload["upload"]["id"]
        status, removed = self.call(
            f"/api/admin/daily-source-uploads/{source_id}", token=self.admin, method="DELETE",
        )
        self.assertEqual(status, 200)
        self.assertIn("回收站", removed["message"])
        trash = self.call("/api/admin/trash", token=self.admin)[1]["trash"]
        item = next(row for row in trash if row["kind"] == "daily_source_upload")
        self.assertEqual(self.call(
            f"/api/admin/trash/{item['id']}/restore", {}, token=self.admin,
        )[0], 200)
        restored = self.call(f"/api/daily-report?date={report_date}", token=self.admin)[1]
        self.assertEqual(restored["safety_checks"]["unqualified_count"], 1)

    def test_legacy_production_group_attendance_migrates_to_default_shift(self):
        """旧版仅按班组保存的生产出勤记录应无损迁移到默认班次。"""

        report_date = web_server.business_today().isoformat()
        created = web_server.now_iso()
        with web_server.DB_LOCK, web_server.db() as connection:
            admin_id = int(connection.execute(
                "SELECT id FROM users WHERE username = 'admin'",
            ).fetchone()["id"])
            group_id = int(connection.execute(
                "INSERT INTO daily_production_groups(name, sort_order, active, created_at, updated_at) "
                "VALUES (?, ?, 1, ?, ?)",
                ("历史装配组", 99, created, created),
            ).lastrowid)
            connection.execute(
                "INSERT INTO daily_production_attendance"
                "(report_date, group_id, attendance_count, note, updated_by, updated_at) "
                "VALUES (?, ?, ?, ?, ?, ?)",
                (report_date, group_id, 7, "旧版班组记录", admin_id, created),
            )
        web_server.init_db()

        status, attendance = self.call(
            f"/api/admin/daily-attendance?date={report_date}", token=self.admin,
        )
        self.assertEqual(status, 200)
        migrated = next(item for item in attendance["production_groups"] if item["group_id"] == group_id)
        self.assertEqual(migrated["shift_name"], "白班")
        self.assertEqual(migrated["staffing_count"], 0)
        self.assertEqual(migrated["attendance_count"], 7)
        self.assertEqual(migrated["difference"], -7)
        groups = self.call("/api/admin/daily-production-groups", token=self.admin)[1]["groups"]
        group = next(item for item in groups if item["id"] == group_id)
        self.assertEqual(group["shifts"][0]["name"], "白班")

    def test_production_staffing_update_refreshes_today_but_keeps_history(self):
        """班组编制调整应刷新当天差异，但不得改写已经形成的历史日清数据。"""

        today = web_server.business_today().isoformat()
        previous_day = (web_server.business_today() - timedelta(days=1)).isoformat()
        status, created = self.call("/api/admin/daily-production-groups", {
            "name": "编制快照组", "sort_order": 1, "active": True,
            "shifts": [{"name": "白班", "staffing_count": 10, "sort_order": 1, "active": True}],
        }, token=self.admin)
        self.assertEqual(status, 201)
        group = created["group"]
        shift = group["shifts"][0]
        for report_date in (previous_day, today):
            status, _ = self.call("/api/admin/daily-attendance", {
                "date": report_date, "records": [],
                "production_groups": [{"shift_id": shift["id"], "attendance_count": 8, "note": ""}],
            }, token=self.admin)
            self.assertEqual(status, 200)

        status, updated = self.call(
            f"/api/admin/daily-production-groups/{group['id']}", {
                "name": group["name"], "sort_order": group["sort_order"], "active": True,
                "shifts": [{
                    "id": shift["id"], "name": shift["name"], "staffing_count": 12,
                    "sort_order": shift["sort_order"], "active": True,
                }],
            }, token=self.admin, method="PATCH",
        )
        self.assertEqual(status, 200)
        self.assertEqual(updated["group"]["staffing_count"], 12)
        today_row = self.call(
            f"/api/admin/daily-attendance?date={today}", token=self.admin,
        )[1]["production_groups"][0]
        history_row = self.call(
            f"/api/admin/daily-attendance?date={previous_day}", token=self.admin,
        )[1]["production_groups"][0]
        self.assertEqual(today_row["staffing_count"], 12)
        self.assertEqual(today_row["difference"], 4)
        self.assertEqual(history_row["staffing_count"], 10)
        self.assertEqual(history_row["difference"], 2)

    def test_templates_search_preflight_and_retry(self):
        """模板搜索、执行前检查和失败重试必须复用同一任务参数与用户隔离目录。"""

        status, created = self.call("/api/templates", {
            "name": "常用文本去重", "action": "text.transform",
            "payload": {"options": {"operation": "dedup"}},
        }, token=self.admin)
        self.assertEqual(status, 201)
        template_id = created["id"]
        status, templates = self.call("/api/templates", token=self.admin)
        self.assertEqual(status, 200)
        self.assertEqual(templates["templates"][0]["name"], "常用文本去重")
        status, _ = self.call(f"/api/templates/{template_id}", {"name": "文本去重模板", "payload": {"options": {"operation": "sort"}}}, token=self.admin, method="PATCH")
        self.assertEqual(status, 200)
        status, preflight = self.call("/api/jobs/preflight", {"action": "text.transform", "payload": {"text": "甲\n乙", "operation": "dedup"}}, token=self.admin)
        self.assertEqual(status, 200)
        self.assertTrue(preflight["ok"])
        status, created = self.call("/api/jobs", {"action": "text.transform", "title": "可搜索任务", "payload": {"text": "甲\n甲", "operation": "dedup"}}, token=self.admin)
        self.assertEqual(status, 202)
        job = self.wait_job(created["job_id"])
        self.assertEqual(len(job["versions"]), 1)
        status, found = self.call("/api/search?q=" + urllib.parse.quote("可搜索"), token=self.admin)
        self.assertEqual(status, 200)
        self.assertEqual(found["jobs"][0]["id"], job["id"])
        with web_server.DB_LOCK, web_server.db() as connection:
            connection.execute("UPDATE web_jobs SET status = 'failed', error = '模拟失败' WHERE id = ?", (job["id"],))
        status, retried = self.call(f"/api/jobs/{job['id']}/retry", token=self.admin, payload={})
        self.assertEqual(status, 202)
        retry = self.wait_job(retried["job_id"])
        self.assertEqual(retry["status"], "completed")
        self.assertEqual(retry["retry_of"], job["id"])
        status, _ = self.call(f"/api/templates/{template_id}", token=self.admin, method="DELETE")
        self.assertEqual(status, 200)

    def test_compare_review_continues_same_job(self):
        """对比业务人工复核应在原任务上继续执行，不能绕过计划或创建孤立任务。"""

        def book_bytes(rows):
            """把合成行序列化为人工复核上传所需的工作簿字节。"""

            book = Workbook()
            sheet = book.active
            for row in rows:
                sheet.append(row)
            stream = __import__("io").BytesIO()
            book.save(stream)
            return stream.getvalue()

        handles = []
        for name, content in (("a.xlsx", [["编号", "数量", "单价"], ["A-1", 1, 10]]),
                              ("b.xlsx", [["编号", "数量", "单价"], ["A-1", 2, 12]])):
            query = urllib.parse.urlencode({"name": name, "group": "review-group"})
            status, uploaded = self.call(
                f"/api/files/upload?{query}", token=self.admin,
                raw=book_bytes(content), headers={"Content-Length": str(len(book_bytes(content)))},
            )
            self.assertEqual(status, 201)
            handles.append(uploaded["handle"])

        status, created = self.call("/api/jobs", {
            "action": "web.compare.review", "title": "表格比对复核", "payload": {
                "file1": [handles[0]], "file2": [handles[1]],
            },
        }, token=self.admin)
        self.assertEqual(status, 202)
        prepared = self.wait_job(created["job_id"])
        self.assertEqual(prepared["status"], "completed")
        self.assertTrue(prepared["review_pending"])
        self.assertIn("编号", prepared["result"]["common"])

        status, board = self.call("/api/dashboard", token=self.admin)
        self.assertEqual(status, 200)
        usage_keys = {item["key"] for item in board["feature_usage"]}
        self.assertIn("compare", usage_keys)
        self.assertNotIn("web", usage_keys)
        self.assertEqual(board["metrics"]["completed_jobs"], 0)
        self.assertEqual(board["metrics"]["running_jobs"], 1)
        self.assertEqual(board["status_breakdown"]["review"], 1)
        self.assertTrue(board["recent_jobs"][0]["review_pending"])

        status, resumed = self.call(f"/api/jobs/{created['job_id']}/review", {
            "choices": {"key": "编号", "columns": ["数量"]},
        }, token=self.admin)
        self.assertEqual(status, 202)
        self.assertEqual(resumed["job_id"], created["job_id"])
        completed = self.wait_job(created["job_id"])
        self.assertEqual(completed["status"], "completed")
        self.assertFalse(completed["review_pending"])
        self.assertEqual(completed["result"]["result"]["counts"]["diffs"], 1)
        self.assertEqual(completed["result"]["result"]["columns"], ["数量"])
        self.assertEqual(completed["presentation"]["parameters"][1]["value"], "数量")
        self.assertEqual(len(completed["files"]), 1)

    def test_supplier_batch_review_selects_suppliers_and_excludes_original(self):
        """供应商批次复核只生成用户勾选供应商，并始终排除原厂内容。"""

        book = Workbook()
        sheet = book.active
        sheet.title = "批次清单"
        sheet.append(["材料编号", "材料名称", "规格", "单位", "最终采购数量", "供应商名称"])
        sheet.append(["JBC-001", "密封圈", "DN50", "个", 3, "供应商甲"])
        sheet.append(["JBC-002", "原厂密封圈", "DN50", "个", 5, "供应商原厂"])
        sheet.append(["JBC-003", "螺栓", "M8", "套", 2, "供应商乙"])
        stream = __import__("io").BytesIO()
        book.save(stream)
        content = stream.getvalue()

        query = urllib.parse.urlencode({"name": "批次清单.xlsx", "group": "supplier-batch-review"})
        status, uploaded = self.call(
            f"/api/files/upload?{query}", token=self.admin,
            raw=content, headers={"Content-Length": str(len(content))},
        )
        self.assertEqual(status, 201)

        status, created = self.call("/api/jobs", {
            "action": "web.supplier_batch.review", "title": "供应商批次表复核",
            "payload": {"batch_paths": [uploaded["handle"]]},
        }, token=self.admin)
        self.assertEqual(status, 202)
        prepared = self.wait_job(created["job_id"])
        self.assertEqual(prepared["status"], "completed")
        self.assertTrue(prepared["review_pending"])
        self.assertEqual(
            {item["name"] for item in prepared["result"]["result"]["suppliers"]},
            {"供应商甲", "供应商乙"},
        )
        self.assertEqual(prepared["result"]["result"]["excluded_original_count"], 1)

        status, rejected = self.call(f"/api/jobs/{created['job_id']}/review", {
            "choices": {"suppliers": ["供应商乙"]},
        }, token=self.admin)
        self.assertEqual(status, 400)
        self.assertIn("交付日期", rejected["error"])

        status, resumed = self.call(f"/api/jobs/{created['job_id']}/review", {
            "choices": {
                "suppliers": ["供应商乙"],
                "batch_dates": {"批次清单": "8.7"},
            },
        }, token=self.admin)
        self.assertEqual(status, 202)
        self.assertEqual(resumed["job_id"], created["job_id"])
        completed = self.wait_job(created["job_id"])
        self.assertEqual(completed["status"], "completed")
        self.assertFalse(completed["review_pending"])
        result = completed["result"]["result"]
        self.assertEqual(result["suppliers"], ["供应商乙"])
        self.assertEqual(result["batch_dates"], {"批次清单": "8.7"})
        self.assertEqual(result["generated"], 1)
        self.assertEqual(result["rows"], 1)
        self.assertEqual(result["excluded_original_count"], 1)
        self.assertEqual(len(completed["files"]), 1)

    def test_report_center_and_batch_track_include_result_files(self):
        """报表中心与批次跟踪应关联可下载结果文件，而不是只显示任务元数据。"""

        job_id = "report-batch-job"
        batch_name = "GKMYR26027-06"
        output = (
            web_server.DATA_ROOT / "users" / "1" / "jobs" / job_id
            / "outputs" / f"{batch_name}到料结果.xlsx"
        )
        output.parent.mkdir(parents=True)
        workbook = Workbook()
        workbook.active.append(["批次号", batch_name])
        workbook.save(output)
        timestamp = web_server.now_iso()
        with web_server.DB_LOCK, web_server.db() as connection:
            connection.execute(
                "INSERT INTO web_jobs(id, user_id, action, title, status, progress, logs, result, "
                "files, cancelled, payload, created_at, updated_at) "
                "VALUES (?, 1, 'web.arrival', '每日到料结果', 'completed', 100, '[]', ?, "
                "'[]', 0, '{}', ?, ?)",
                (
                    job_id,
                    json.dumps({"report_path": str(output)}, ensure_ascii=False),
                    timestamp,
                    timestamp,
                ),
            )

        query = urllib.parse.quote(batch_name)
        status, tracked = self.call(f"/api/batch-track?q={query}", token=self.admin)
        self.assertEqual(status, 200)
        self.assertEqual(len(tracked["items"]), 1)
        self.assertEqual(tracked["items"][0]["job_id"], job_id)
        self.assertEqual(tracked["items"][0]["files"], [output.name])

        status, report = self.call("/api/reports?range=7d", token=self.admin)
        self.assertEqual(status, 200)
        self.assertTrue(report["name"].startswith("业务报表_近7天_"))
        status, content = self.call(report["url"], token=self.admin)
        self.assertEqual(status, 200)
        exported = load_workbook(BytesIO(content), data_only=True)
        self.assertEqual(exported.sheetnames, ["汇总", "明细"])
        self.assertEqual(exported["汇总"]["B5"].value, 1)
        self.assertEqual(exported["明细"]["C2"].value, "每日到料结果")

        with mock.patch("web_backend.services.reports.datetime") as mocked_datetime:
            mocked_datetime.now.return_value = datetime(2026, 8, 10, 9, 0, 0)
            weekly = web_server.auto_weekly_report_if_due()
            self.assertTrue(weekly.startswith("已生成周报 "))
            self.assertEqual(web_server.auto_weekly_report_if_due(), "")
        state = json.loads(
            (web_server.DATA_ROOT / "auto_weekly_report_state.json").read_text(
                encoding="utf-8",
            )
        )
        self.assertEqual(state["last_weekly"], "20260810")
        with web_server.DB_LOCK, web_server.db() as connection:
            notification_count = connection.execute(
                "SELECT COUNT(*) FROM messages WHERE title = '本周业务报表已生成'",
            ).fetchone()[0]
        self.assertEqual(notification_count, 1)

        self.assertEqual(
            self.call("/api/reports?range=unknown", token=self.admin)[0],
            400,
        )
        invalid_path = urllib.parse.quote("../accounts.sqlite3")
        self.assertEqual(
            self.call(f"/api/reports/download?path={invalid_path}", token=self.admin)[0],
            404,
        )


if __name__ == "__main__":
    unittest.main()
