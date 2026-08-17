# -*- coding: utf-8 -*-
"""票货匹配：发票台账与采购明细按供应商比对。"""
from __future__ import annotations

import os
import tempfile
import unittest

import openpyxl

from core import invoice_match_core


def _write_invoice(path, rows):
    """生成带合并标题和正式发票字段的最小台账。"""

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "202608"  # 表名表示所属月份
    worksheet.merge_cells("A1:J1")
    worksheet.cell(1, 1, "增值税发票")
    worksheet.append(["序号", "发票号码", "开票日期", "销售方名称", "费用项目",
                      "不含税金额（元）", "税额（元）", "价税合计（元）", "税率/征收方式", "备注"])
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()


def _write_purchase(path, rows):
    """生成采购导入模板，并保留前两列空白以贴近真实表结构。"""

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "采购计划导入模板"
    worksheet.append(["仓库编号", "采购员编号", "产品编号", "产品名称", "规格",
                      "供应商编码", "供应商", "数量", "单据明细备注", "预计到货日期"])
    for row in rows:
        worksheet.append([None, None] + list(row))  # 前两列留空贴近真实模板
    workbook.save(path)
    workbook.close()


class InvoiceMatchCoreTests(unittest.TestCase):
    """验证供应商集合匹配、状态输出和输入结构错误提示。"""

    def setUp(self):
        """构造两个共有、一个单边采购和一个单边发票供应商。"""

        self.temp = tempfile.TemporaryDirectory(prefix="fyt_invoice_match_")
        self.invoice = self.path("8月统计增值税发票.xlsx")
        self.purchase = self.path("26036-02.xlsx")
        _write_invoice(self.invoice, [
            [1, "111", "2026-08-01", "客供件", "保护膜", 100, 13, 113, "13%", ""],
            [2, "112", "2026-08-02", "众瀚", "铁箱", 200, 26, 226, "13%", ""],
            [3, "113", "2026-08-03", "独立新供应商", "珍珠棉", 50, 6.5, 56.5, "13%", ""],
        ])
        _write_purchase(self.purchase, [
            ["8891167589", "保护膜", "0.08×200", "GYS1", "客供件", 2.08],
            ["JBC0101A0001", "铁箱", "2280x1480x830", "GYS2", "众瀚", 5],
            ["JBC0099", "物料", "1x1", "GYS9", "瑞敏", 3],
        ])

    def tearDown(self):
        """删除合成发票、采购表和输出目录。"""

        self.temp.cleanup()

    def path(self, name):
        """返回当前用例临时目录下的文件路径。"""

        return os.path.join(self.temp.name, name)

    def test_match_classifies_suppliers(self):
        """供应商应分别归入正常、无票采购和有发票无采购，并写入汇总表。"""

        out_dir = self.path("输出")
        result = invoice_match_core.match([self.invoice], [self.purchase], out_dir=out_dir)
        self.assertEqual(result["matched"], 2)  # 两边都有的供应商
        self.assertEqual(result["no_invoice"], 1)  # 有采购无发票
        self.assertEqual(result["no_purchase"], 1)  # 有发票无采购
        self.assertEqual(result["no_invoice_suppliers"], ["瑞敏"])  # 单边采购名单
        self.assertEqual(result["no_purchase_suppliers"], ["独立新供应商"])  # 单边发票名单
        self.assertTrue(os.path.isfile(result["path"]))  # 汇总表落盘

        workbook = openpyxl.load_workbook(result["path"], data_only=True)
        summary = workbook.active
        rows = [row for row in summary.iter_rows(min_row=2, values_only=True) if any(row)]  # 跳过表头与空行
        by_supplier = {str(row[0]): row[3] for row in rows}
        self.assertEqual(by_supplier["客供件"], "正常")  # 共有供应商标正常
        self.assertEqual(by_supplier["众瀚"], "正常")
        self.assertEqual(by_supplier["瑞敏"], "无票采购")  # 采购单边状态
        self.assertEqual(by_supplier["独立新供应商"], "有发票无采购")  # 发票单边状态
        self.assertEqual(by_supplier["客供件"], "正常")
        workbook.close()

    def test_missing_inputs_raise(self):
        """任一侧没有输入文件时都应在处理前给出业务错误。"""

        with self.assertRaises(ValueError):
            invoice_match_core.match([], [self.purchase])  # 发票为空拒绝
        with self.assertRaises(ValueError):
            invoice_match_core.match([self.invoice], [])  # 采购为空拒绝

    def test_missing_columns_raise(self):
        """采购表缺少供应商列时错误应明确指出缺失字段。"""

        bad = self.path("无供应商列.xlsx")
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.append(["材料编号", "材料名称", "规格", "单位", "数量"])  # 缺少供应商列
        worksheet.append(["JBC001", "铁箱", "1x1", "个", 3])
        workbook.save(bad)
        workbook.close()
        with self.assertRaises(ValueError) as context:
            invoice_match_core.match([self.invoice], [bad], out_dir=self.path("输出2"))
        self.assertIn("供应商", str(context.exception))  # 错误信息指出缺列


if __name__ == "__main__":
    unittest.main()
