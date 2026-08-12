"""生产计划、月度订单台账和零星订单结构化分析回归测试。"""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from core import daily_production_plan_core


class DailyProductionPlanCoreTests(unittest.TestCase):
    """使用合成 Excel 验证日清生产图表所需的解析口径。"""

    def test_analyze_builds_safe_preview(self):
        """多工作表预览应保留顺序并把单元格值安全转换为前端文本。"""

        with tempfile.TemporaryDirectory() as temp_name:
            target = Path(temp_name) / "生产计划.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "计划汇总"
            sheet.append(["班次", "计划", "实际", "差异"])
            sheet.append(["白班", 45, 41, -4])
            detail = workbook.create_sheet("批次明细")
            detail.append(["批次", "班组", "数量"])
            detail.append(["GKMYR26027-06", "小件组", 18])
            workbook.save(target)
            workbook.close()

            result = daily_production_plan_core.analyze(target)

        self.assertEqual(result["sheet_count"], 2)
        self.assertEqual(result["row_count"], 4)
        self.assertEqual(result["sheets"][0]["preview"][1], ["白班", "45", "41", "-4"])

    def test_rejects_unsupported_file(self):
        """日清生产资料只接受 xlsx，旧 xls 必须在解析前明确拒绝。"""

        with tempfile.TemporaryDirectory() as temp_name:
            target = Path(temp_name) / "计划.xls"
            target.write_bytes(b"legacy")
            with self.assertRaisesRegex(ValueError, "xlsx"):
                daily_production_plan_core.analyze(target)

    def test_converts_excel_serial_date_and_detects_shipping_sheet(self):
        """Excel 日期序列应转为业务日期，同时识别正式订单与生产计划表。"""

        with tempfile.TemporaryDirectory() as temp_name:
            target = Path(temp_name) / "订单号发运统计.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "正式订单"
            sheet.append(["序号", "日期", "订单号", "国家", "类型", "数量", "发运完成时间"])
            sheet.append([1, "8月份", "GMMYR26163A", "宝腾", "CKD", 120, ""])
            plan = workbook.create_sheet("生产计划")
            # 46235、46236 是连续 Excel 日期序列，避免测试依赖 openpyxl 自动日期格式。
            plan.append(["日期", 46235, 46236])
            plan.append(["班次", "白班", "夜班"])
            plan.append(["计划", 45, 40])
            plan.append(["实际", 37, 43])
            plan.append(["差异", -8, 3])
            workbook.save(target)
            workbook.close()

            result = daily_production_plan_core.analyze(target)

        self.assertEqual(result["sheets"][0]["kind"], "正式订单")
        plan_sheet = result["sheets"][1]
        self.assertEqual(plan_sheet["kind"], "生产计划")
        self.assertEqual(plan_sheet["preview"][0][1], "2026-08-01")

    def test_builds_production_insights_for_daily_matrix(self):
        """横向日班/夜班矩阵应聚合计划、实际、差异、班组和批次产量。"""

        with tempfile.TemporaryDirectory() as temp_name:
            target = Path(temp_name) / "生产计划矩阵.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Sheet1"
            # 空前导列与成对日期列模拟现行生产计划模板，而非人为简化成标准数据库表。
            sheet.append([None, None, "日期", 46235, None, 46236, None])
            sheet.append([None, None, "班次", "白班", "夜班", "白班", "夜班"])
            sheet.append([None, None, "计划", 45, 40, 45, 45])
            sheet.append([None, None, "实际", 37, 43, 34, 44])
            sheet.append([None, None, "差异", -8, 3, -11, -1])
            sheet.append([None, None, "班组/CASE", 46235, None, 46236, None])
            sheet.append([None, None, None, "白班", "夜班", "白班", "夜班"])
            sheet.append([None, None, None, "批次A", "批次A", "批次B", "批次B"])
            sheet.append([None, None, "小件组", 12, 13, 8, 7])
            sheet.append([None, None, "大件组", 20, 20, 15, 17])
            workbook.save(target)
            workbook.close()

            result = daily_production_plan_core.analyze(target, report_date="2026-08-01")

        insights = result["insights"]
        self.assertEqual(insights["focus_date"], "2026-08-01")
        self.assertEqual(insights["plan_total"], 85)
        self.assertEqual(insights["actual_total"], 80)
        self.assertEqual(insights["difference_total"], -5)
        self.assertEqual(insights["team_summary"][0]["team"], "大件组")
        self.assertEqual(insights["batch_summary"][0]["batch"], "批次A")
        self.assertEqual(len(result["sheets"][0]["table_rows"]), 10)

    def test_unreported_shift_is_not_treated_as_shortfall(self):
        """尚未填报实际产量的班次应单独提示，不能虚构为计划欠产。"""

        with tempfile.TemporaryDirectory() as temp_name:
            target = Path(temp_name) / "生产计划待填报.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append([None, None, "日期", 46237, None])
            sheet.append([None, None, "班次", "白班", "夜班"])
            sheet.append([None, None, "计划", 45, 45])
            sheet.append([None, None, "实际", 41, None])
            sheet.append([None, None, "差异", -4, None])
            workbook.save(target)
            workbook.close()

            result = daily_production_plan_core.analyze(target, report_date="2026-08-03")

        insights = result["insights"]
        self.assertEqual(insights["plan_total"], 90)
        self.assertEqual(insights["actual_total"], 41)
        self.assertEqual(insights["difference_total"], -4)
        self.assertEqual(insights["reported_plan_total"], 45)
        self.assertEqual(insights["unreported_plan_total"], 45)
        self.assertEqual(insights["unreported_shift_count"], 1)
        self.assertFalse(insights["shift_summary"][1]["actual_reported"])
        self.assertTrue(any("夜班尚未填报实际产量" in item for item in insights["highlights"]))
        self.assertNotIn("夜班较计划少完成", "；".join(insights["highlights"]))

    def test_order_ledger_groups_merged_rows_and_sporadic_pallets(self):
        """订单主行与续行应合并缺件/危包，零星订单续行托数也需累计。"""

        with tempfile.TemporaryDirectory() as temp_name:
            target = Path(temp_name) / "订单号发运统计.xlsx"
            workbook = Workbook()
            formal = workbook.active
            formal.title = "正式订单"
            formal.append(["订单管理清单"])
            formal.append(["序号", "日期", "订单号", "国家", "类型", "数量", "发运完成时间", "缺件", None, None, None, None, "危包", None, None, None, None, "缺件实际完时间", "危包实际完时间", "整柜实际完时间", "订单是否关闭", "备注"])
            formal.append([None, None, None, None, None, None, None, "物料号", "零件名称", "零件数量", "缺件发运订单号", "发运完成时间", "物料号", "零件名称", "零件数量", "危包发运订单号", "发运完成时间"])
            formal.append([1, "8月份", "ORDER-001", "宝腾", "KD", 120, "2026-08-04", "M-01", "缺件甲", 2, "MISS-01", "", "H-01", "危包甲", 1, "HAZ-01", "2026-08-05", "", "", "", "否", "跟踪中"])
            # 第二行故意省略订单主字段，模拟合并单元格导出后只在首行保留订单号。
            formal.append([None, None, None, None, None, None, None, "M-02", "缺件乙", 1, "MISS-02", "2026-08-06"])
            sporadic = workbook.create_sheet("零星订单")
            sporadic.append(["订单号", "运输方式", "国家", "订单类型", "是否拼箱", "柜量", "柜型", "托数", "长MM", "宽MM", "高MM", "体积（CBM)", "发运时间", "提货司机车牌", "提货司机姓名", "提货司机电话", "备注"])
            sporadic.append(["SP-001", "空运", "宝腾", "领料", "否", None, None, 1, 1000, 1000, 1000, 1, 8.4, "渝A1", "张师傅", "13800000000", "已完结"])
            # 零星订单的第二托盘沿用上一订单，验证续行不会被错误识别为独立空订单。
            sporadic.append([None, None, None, None, None, None, None, 2, 1000, 1000, 500, 1])
            workbook.save(target)
            workbook.close()

            result = daily_production_plan_core.analyze(target, report_date="2026-08-06")

        ledger = result["insights"]["order_ledger"]
        self.assertEqual(len(ledger["formal_orders"]), 1)
        self.assertEqual(len(ledger["formal_orders"][0]["missing_parts"]), 2)
        self.assertEqual(ledger["formal_orders"][0]["outstanding_missing_count"], 1)
        self.assertEqual(len(ledger["sporadic_orders"]), 1)
        self.assertEqual(ledger["sporadic_orders"][0]["pallet_count"], 3)
        self.assertEqual(ledger["sporadic_orders"][0]["shipment_date"], "2026-08-04")


if __name__ == "__main__":
    unittest.main()
