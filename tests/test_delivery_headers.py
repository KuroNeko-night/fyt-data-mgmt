# -*- coding: utf-8 -*-
"""送货计划表头识别回归测试(合成簿,可移植)。

钉死本轮修复:
- SAP 列名 下阶物料/下阶物料描述 能被 detect_layout 映射为 code/cname;
- 「委外供应商属性」不再被当成供应商名称列(含'属性'的列排除);
- list_sheets 列子表、非 xlsx 返回 []。

只在临时目录造簿并读取,对用户数据零副作用。
"""
import os
import tempfile
import unittest
import warnings

import openpyxl

from core import delivery_core as D

warnings.filterwarnings("ignore", message="Workbook contains no default style")


class _Tmp(unittest.TestCase):
    """提供隔离主数据库与多工作表合成器。"""

    def setUp(self):
        """创建临时送货资料根。"""

        self._tmp = tempfile.mkdtemp(prefix="fyt_dlv_")
        self._old_catalog = os.environ.get("FYT_CATALOG_PATH")
        os.environ["FYT_CATALOG_PATH"] = os.path.join(self._tmp, "catalog.json")

    def tearDown(self):
        """恢复环境并删除合成文件。"""

        import shutil
        if self._old_catalog is None:
            os.environ.pop("FYT_CATALOG_PATH", None)
        else:
            os.environ["FYT_CATALOG_PATH"] = self._old_catalog
        shutil.rmtree(self._tmp, ignore_errors=True)

    def mk(self, name, sheets):
        """按字典顺序生成多工作表 xlsx。"""

        p = os.path.join(self._tmp, name)
        wb = openpyxl.Workbook()
        first = True
        for sn, rows in sheets.items():
            ws = wb.active if first else wb.create_sheet()
            ws.title = sn
            first = False
            for row in rows:
                ws.append(row)
        wb.save(p)
        return p


class TestDetectLayout(_Tmp):
    """保护 SAP 与经典送货表头到内部字段的映射。"""

    def _layout(self, header):
        """用一行表头和一行占位数据执行布局检测。"""

        p = self.mk("t.xlsx", {"S": [header, ["x"] * len(header)]})
        wb = openpyxl.load_workbook(p)
        hr, cols = D.detect_layout(wb["S"])
        wb.close()
        return hr, cols

    def test_sap_xiajie_maps_code_and_name(self):
        """SAP 下阶物料和描述应分别映射编码与名称。"""

        # SAP KD 清单用「下阶物料/下阶物料描述」而非零部件代码
        hr, cols = self._layout(["上阶物料", "下阶物料", "下阶物料描述", "数量"])
        self.assertIsNotNone(hr)
        self.assertIn("code", cols)
        self.assertIn("cname", cols)
        # code 应指向「下阶物料」列(第2列),而非「上阶物料」
        self.assertEqual(cols["code"], 2)

    def test_supplier_attr_not_taken_as_supplier_name(self):
        """包含“供应商”字样的属性列不能抢占真正供应商名称列。"""

        # 「委外供应商属性」含'供应商'子串,但不是供应商名称列,不能被抢
        hr, cols = self._layout(
            ["下阶物料", "委外供应商属性", "供应商代码", "供应商名称", "数量"])
        self.assertIsNotNone(hr)
        # 供应商名称列必须是第4列「供应商名称」,不是第2列「委外供应商属性」。
        # 硬断言(不加 if):若守卫失效 sup_name 会指向 2,此处即失败。
        self.assertIn("sup_name", cols)
        self.assertEqual(cols["sup_name"], 4)

    def test_classic_code_still_works(self):
        """扩展 SAP 别名后不得破坏经典零部件代码表头。"""

        # 不回退:经典「零部件代码」仍可识别
        hr, cols = self._layout(["零部件代码", "零部件名称", "数量", "供应商代码"])
        self.assertIsNotNone(hr)
        self.assertIn("code", cols)


class TestListSheets(_Tmp):
    """验证前端工作表选择接口的顺序和错误回退。"""

    def test_lists_all_sheets_in_order(self):
        """xlsx 工作表应按原始顺序完整列出。"""

        p = self.mk("multi.xlsx", {"Sheet1": [["a"]], "BOM": [["b"]],
                                   "发运清单": [["c"]]})
        self.assertEqual(D.list_sheets(p), ["Sheet1", "BOM", "发运清单"])

    def test_non_xlsx_returns_empty(self):
        """不支持的旧 xls 文件安全返回空列表。"""

        p = os.path.join(self._tmp, "x.xls")
        with open(p, "w") as f:
            f.write("not a real xls")
        self.assertEqual(D.list_sheets(p), [])

    def test_missing_file_returns_empty(self):
        """文件不存在时工作表枚举不抛底层异常。"""

        self.assertEqual(D.list_sheets(os.path.join(self._tmp, "nope.xlsx")), [])


if __name__ == "__main__":
    unittest.main()
