# -*- coding: utf-8 -*-
"""采购计划导入：模板输出、供应商编码映射与“原厂”排除回归。"""
from __future__ import annotations

import os
import tempfile
import unittest

import openpyxl

from core import purchase_plan_core


def _write_template(path, suppliers, extra_rows=None):
    """生成带供应商编码对照和可选额外行的采购导入模板。"""

    workbook = openpyxl.Workbook()
    main = workbook.active
    main.title = "采购计划导入模板"
    main.append(["仓库编号", "采购员编号", "产品编号", "产品名称", "规格",
                 "供应商编码", "供应商", "数量", "单据明细备注", "预计到货日期"])
    main.append([None] * 10)
    for row in (extra_rows or []):
        main.append(row)
    sheet = workbook.create_sheet("Sheet1")
    sheet.append([None, None, None])
    for index, (name, code) in enumerate(suppliers, start=1):
        sheet.append([None, name, code])
    workbook.save(path)
    workbook.close()


def _write_batch(path, rows):
    """生成采购批次清单，保留真实模板的前导空行。"""

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "汇总"
    worksheet.append([None] * 6)
    worksheet.append([None] * 6)
    worksheet.append(["材料编号", "材料名称", "规格", "单位", "求和项:最终采购数量", "供应商"])
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()


class PurchasePlanCoreTests(unittest.TestCase):
    """验证采购计划导入、供应商编码、差异表和批次识别。"""

    def setUp(self):
        """创建模板、批次来源和隔离主数据库。"""

        self.temp = tempfile.TemporaryDirectory(prefix="fyt_purchase_plan_")
        os.environ["FYT_CATALOG_PATH"] = os.path.join(self.temp.name, "catalog.json")  # 主数据库隔离
        self.template = self.path("2采购计划导入模板1.xlsx")
        self.batch_a = self.path("GKMYR26036-02辅料清单总表.xlsx")
        self.batch_b = self.path("GMIDR26178A辅料清单总表.xlsx")
        _write_template(self.template, [
            ("客供件", "GYS26062300001-1"), ("众瀚", "GYS26062300002"),
            ("文恒", "GYS26062300003"), ("吉致", "GYS26062300007"),
        ])
        _write_batch(self.batch_a, [
            ["8891167589", "保护膜（车身油漆保护膜德莎）", "0.08×200", "卷", 2.08, "客供件"],
            ["JBC0101A0001", "铁箱-标箱1(NL-3RB)", "2280x1480x830", "套", 5, "众瀚"],
            ["1", "多层板", "2210*1410*5", "张", 6, "文恒"],
            ["JBC001", "纸箱（原厂）", "560x360x260", "个", 3, "吉致"],
            ["JBC002", "托盘", "1200x1000", "个", 0, "众瀚"],
            ["JBC003", "珍珠棉", "100x50", "张", 2, "#N/A"],
            ["(空白)", "(空白)", "", "", None, ""],
            ["总计", "", "", "", 100, "0"],
        ])
        _write_batch(self.batch_b, [
            ["JBC004", "铁箱", "1000x800", "套", 4, "吉致"],
            ["JBC005", "纸箱(原厂)", "600x400", "个", 2, "吉致"],
        ])

    def tearDown(self):
        """移除主数据库覆盖并删除合成输入输出。"""

        os.environ.pop("FYT_CATALOG_PATH", None)
        self.temp.cleanup()

    def path(self, name):
        """返回当前用例临时目录下路径。"""

        return os.path.join(self.temp.name, name)

    def test_run_generates_batch_files(self):
        """多个批次应各生成独立导入文件，填入供应商编码、数量和交付日期。"""

        out_dir = self.path("输出")
        result = purchase_plan_core.run(
            [self.template], [self.batch_a, self.batch_b], out_dir=out_dir,
        )
        self.assertEqual(result["generated"], 2)  # 两个批次文件
        self.assertEqual(result["rows"], 4)  # 有效数据行数
        self.assertEqual(result["excluded_original_count"], 2)  # 原厂排除
        self.assertEqual(
            [os.path.basename(path) for path in result["files"]],
            ["26036-02.xlsx", "26178A.xlsx"],
        )  # 文件名按批次号

        workbook = openpyxl.load_workbook(result["files"][0], data_only=True)
        self.assertEqual(workbook.sheetnames, ["采购计划导入模板", "Sheet1"])
        main = workbook["采购计划导入模板"]
        self.assertEqual(
            [cell.value for cell in main[1]][:10],
            ["仓库编号", "采购员编号", "产品编号", "产品名称", "规格",
             "供应商编码", "供应商", "数量", "单据明细备注", "预计到货日期"],
        )
        rows = [row for row in main.iter_rows(min_row=2, values_only=True)
                if any(cell is not None for cell in row)]
        self.assertEqual(len(rows), 3)  # 三条有效数据
        first = rows[0]
        # 仓库编号、采购员编号、预计到货日期留空
        self.assertIsNone(first[0])
        self.assertIsNone(first[1])
        self.assertIsNone(first[8])
        self.assertEqual(first[2], "8891167589")  # 物料编码
        self.assertEqual(first[3], "保护膜（车身油漆保护膜德莎）")  # 物料名称
        self.assertEqual(first[4], "0.08×200")  # 规格
        self.assertEqual(first[5], "GYS26062300001-1")  # 供应商编码
        self.assertEqual(first[6], "客供件")  # 供应商名称
        self.assertEqual(first[7], 2.08)  # 数量
        # 数据区字体统一宋体 10
        for row in main.iter_rows(min_row=2, max_row=4, max_col=10):
            for cell in row:
                if cell.value is not None:
                    self.assertEqual(cell.font.name, "宋体")  # 字体
                    self.assertEqual(cell.font.size, 10)  # 字号
        workbook.close()

    def test_run_keeps_supplier_code_sheet(self):
        out_dir = self.path("输出2")
        result = purchase_plan_core.run([self.template], [self.batch_a], out_dir=out_dir)
        workbook = openpyxl.load_workbook(result["files"][0], data_only=True)
        sheet = workbook["Sheet1"]
        codes = {(str(row[1]).strip() if row[1] else "", str(row[2]).strip() if row[2] else "")
                 for row in sheet.iter_rows(min_row=1, values_only=True)}
        self.assertIn(("客供件", "GYS26062300001-1"), codes)
        workbook.close()

    def test_missing_supplier_code_raises(self):
        batch = self.path("GMMYR26163A辅料清单总表.xlsx")
        _write_batch(batch, [["JBC006", "珍珠棉", "100x50", "张", 9, "未知供应商"]])
        with self.assertRaises(ValueError) as context:
            purchase_plan_core.run([self.template], [batch], out_dir=self.path("输出3"))
        self.assertIn("缺少供应商", str(context.exception))  # 缺编码报错

    def test_catalog_supplements_missing_supplier_code(self):
        """模板缺少供应商编码时可从正式主库补齐，但不得臆造未知供应商。"""

        from core import material_catalog
        material_catalog.learn_suppliers({"未知供应商": "GYS999"})
        batch = self.path("GMMYR26163A辅料清单总表.xlsx")
        _write_batch(batch, [["JBC006", "珍珠棉", "100x50", "张", 9, "未知供应商"]])
        result = purchase_plan_core.run([self.template], [batch], out_dir=self.path("输出补全"))
        workbook = openpyxl.load_workbook(result["files"][0], data_only=True)
        main = workbook["采购计划导入模板"]
        self.assertEqual(main.cell(2, 6).value, "GYS999")  # 主库补编码
        workbook.close()
        # 运行后材料与供应商自动学习进档案
        self.assertEqual(material_catalog.resolve_supplier_code("未知供应商"), "GYS999")  # 供应商入库
        self.assertIn("JBC006", material_catalog.load()["materials"])  # 材料入库

    def test_template_without_code_sheet_raises(self):
        bad_template = self.path("无代码模板.xlsx")
        workbook = openpyxl.Workbook()
        workbook.active.title = "采购计划导入模板"
        workbook.save(bad_template)
        workbook.close()
        with self.assertRaises(ValueError) as context:
            purchase_plan_core.run([bad_template], [self.batch_a], out_dir=self.path("输出4"))
        self.assertIn("供应商代码子表", str(context.exception))  # 缺代码子表报错

    def test_unrecognized_batch_name_raises(self):
        batch = self.path("无法识别.xlsx")
        _write_batch(batch, [["JBC006", "珍珠棉", "100x50", "张", 9, "吉致"]])
        with self.assertRaises(ValueError) as context:
            purchase_plan_core.run([self.template], [batch], out_dir=self.path("输出5"))
        self.assertIn("批次号", str(context.exception))  # 批次号不可识别报错

    def test_empty_selection_raises(self):
        with self.assertRaises(ValueError):
            purchase_plan_core.run([], [self.batch_a])  # 模板为空拒绝
        with self.assertRaises(ValueError):
            purchase_plan_core.run([self.template], [])  # 批次为空拒绝

    def test_keeps_prefilled_warehouse_purchaser_and_arrival_date(self):
        """输出复制模板已有仓库、采购员和预计到货日期，不用默认值覆盖。"""

        prefilled = self.path("预填模板.xlsx")
        _write_template(prefilled, [
            ("客供件", "GYS26062300001-1"), ("众瀚", "GYS26062300002"),
        ], extra_rows=[
            ["CK26062300001", "U26052800001", None, None, None, None, None, None, None, "2026-08-04"],
            ["CK26062300002", "U26052800002", None, None, None, None, None, None, None, None],
            ["CK_OLD", "U_OLD", None, None, None, None, None, None, None, "2026-01-01"],
        ])
        batch = self.path("GKMYR26036-05辅料清单总表.xlsx")
        _write_batch(batch, [
            ["8891167589", "保护膜", "0.08×200", "卷", 2.08, "客供件"],
            ["JBC0101A0001", "铁箱", "2280x1480x830", "套", 5, "众瀚"],
            ["JBC0101A0002", "托盘", "1200x1000", "个", 1, "众瀚"],
        ])
        out_dir = self.path("输出预填")
        result = purchase_plan_core.run([prefilled], [batch], out_dir=out_dir)
        workbook = openpyxl.load_workbook(result["files"][0], data_only=True)
        main = workbook["采购计划导入模板"]
        # 模板第 2 行（对应输出行 2）未填写 → 保留列为空，新数据正常写入
        self.assertIsNone(main.cell(2, 1).value)  # 空列保持空
        self.assertIsNone(main.cell(2, 10).value)
        self.assertEqual(main.cell(2, 3).value, "8891167589")  # 新数据写入
        # 模板第 3 行已填写的仓库/采购员/到货日期 → 按行原样保留
        self.assertEqual(main.cell(3, 1).value, "CK26062300001")
        self.assertEqual(main.cell(3, 2).value, "U26052800001")
        self.assertEqual(main.cell(3, 10).value, "2026-08-04")
        self.assertEqual(main.cell(3, 3).value, "JBC0101A0001")
        # 模板第 4 行填写了仓库/采购员但到货日期为空 → 保留已填、空仍空
        self.assertEqual(main.cell(4, 1).value, "CK26062300002")
        self.assertEqual(main.cell(4, 2).value, "U26052800002")
        self.assertIsNone(main.cell(4, 10).value)
        # 模板第 5 行超出新数据行数 → 保留列一并清空，不残留上一批数据
        self.assertIsNone(main.cell(5, 1).value)
        self.assertIsNone(main.cell(5, 2).value)
        self.assertIsNone(main.cell(5, 10).value)
        self.assertIsNone(main.cell(5, 3).value)
        workbook.close()

    def test_duplicate_batch_names_generate_unique_files(self):
        """重复批次显示名仍应生成唯一文件路径，避免后一个结果覆盖前一个。"""

        duplicate = self.path("副本-GKMYR26036-02辅料清单总表.xlsx")
        _write_batch(duplicate, [
            ["JBC099", "第二份材料", "10x10", "件", 3, "众瀚"],
        ])
        result = purchase_plan_core.run(
            [self.template], [self.batch_a, duplicate], out_dir=self.path("输出重名"))
        self.assertEqual(
            [os.path.basename(path) for path in result["files"]],
            ["26036-02.xlsx", "26036-02 (2).xlsx"],
        )  # 重名追加序号
        self.assertTrue(all(os.path.isfile(path) for path in result["files"]))  # 两个文件都存在

    def test_failed_batch_does_not_commit_outputs_or_catalog(self):
        bad = self.path("GMMYR26163A辅料清单总表.xlsx")
        _write_batch(bad, [["JBC006", "珍珠棉", "100x50", "张", 9, "未知供应商"]])
        out_dir = self.path("输出事务")
        with self.assertRaises(ValueError):
            purchase_plan_core.run([self.template], [self.batch_a, bad], out_dir=out_dir)
        self.assertEqual([name for name in os.listdir(out_dir) if name.endswith(".xlsx")], [])  # 无半成品输出
        self.assertEqual(purchase_plan_core.material_catalog.load()["materials"], {})  # 主数据库无写入
        self.assertEqual(purchase_plan_core.material_catalog.load()["suppliers"], {})

    def test_supplier_codes_after_row_200_are_loaded(self):
        """供应商编码对照不能只扫描前两百行，否则大型模板尾部关系会丢失。"""

        template = self.path("长供应商表模板.xlsx")
        suppliers = [("供应商%03d" % index, "GYS%03d" % index) for index in range(1, 202)]
        _write_template(template, suppliers)
        batch = self.path("GKMYR26099-01辅料清单总表.xlsx")
        _write_batch(batch, [["JBC200", "长表材料", "20x20", "件", 1, "供应商201"]])
        result = purchase_plan_core.run([template], [batch], out_dir=self.path("输出长表"))
        workbook = openpyxl.load_workbook(result["files"][0], data_only=True)
        self.assertEqual(workbook["采购计划导入模板"].cell(2, 6).value, "GYS201")  # 两百行后关系仍生效
        workbook.close()


if __name__ == "__main__":
    unittest.main()
