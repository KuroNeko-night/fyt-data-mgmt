# -*- coding: utf-8 -*-
"""对账单制作：批次扫描（多块并排/缺表头/标黄排除）、供应商识别与分组生成。"""
from __future__ import annotations

import os
import tempfile
import unittest

import openpyxl
from openpyxl.styles import PatternFill

from core import reconcile_statement_core


def _write_file(path, blocks):
    """blocks: [{batch, sheet, col, header(bool), rows, yellow:[相对数据行号]}]
    header=True 时在数据前插一行表头；rows=[(code,name,spec,unit,qty,supplier)]"""
    workbook = openpyxl.Workbook()
    worksheets = {}
    for block in blocks:
        if block["sheet"] in worksheets:
            worksheet = worksheets[block["sheet"]]
        elif not worksheets:
            worksheet = workbook.active
            worksheet.title = block["sheet"]
        else:
            worksheet = workbook.create_sheet(block["sheet"])
        worksheets[block["sheet"]] = worksheet
        row = int(block.get("row", 1))
        worksheet.cell(row, block["col"], "%s交付日期7.6" % block["batch"])
        row += 1
        if block.get("header", True):
            for offset, name in enumerate(("材料编号", "材料名称", "规格", "\n单位", "求和项:最终采购数量", "供应商")):
                worksheet.cell(row, block["col"] + offset, name)
            row += 1
        for index, values in enumerate(block["rows"], start=1):
            for offset, value in enumerate(values):
                cell = worksheet.cell(row, block["col"] + offset, value)
                if index in block.get("yellow", ()):
                    cell.fill = PatternFill("solid", fgColor="FFFF00")
            row += 1
    workbook.save(path)
    workbook.close()


class ReconcileStatementCoreTests(unittest.TestCase):
    """验证对账单扫描、黄行排除、供应商分组和输出路径安全。"""

    def setUp(self):
        """构造含批次块、横纵布局和黄色排除行的采购清单。"""

        self.temp = tempfile.TemporaryDirectory(prefix="fyt_rs_")
        self.old_catalog = os.environ.get("FYT_CATALOG_PATH")
        os.environ["FYT_CATALOG_PATH"] = os.path.join(self.temp.name, "catalog.json")  # 主数据库隔离
        self.file_a = os.path.join(self.temp.name, "（已下单）甲供应商7月采购清单明细.xlsx")
        _write_file(self.file_a, [
            {"batch": "GKMYR26027-01", "sheet": "01", "col": 1, "rows": [
                ("JBC0101A0001", "铁箱-标箱1", "2280x1480x830", "套", 6, "甲"),
                ("JBC0101A0002", "铁箱-标箱2", "2280x1480x620", "套", 4, "甲"),
                ("JBC0101A0003", "铁箱-标箱3", "2280x1480x420", "套", 2, "甲"),
            ], "yellow": (2,)},
            {"batch": "GKMYR26027-02", "sheet": "02", "col": 1, "rows": [
                ("JBC0101A0001", "铁箱-标箱1", "2280x1480x830", "套", 10, "甲"),
            ]},
        ])
        self.file_b = os.path.join(self.temp.name, "清单未命名.xlsx")
        _write_file(self.file_b, [
            {"batch": "GKMYR26027-01", "sheet": "左右", "col": 1, "rows": [
                ("8891167589", "保护膜", "0.08×200", "卷", 2, "客供件"),
                ("JBC0101A0337", "循环铁箱", "2280x1480x830", "套", 6, "客供件"),
            ], "yellow": (1,)},
            {"batch": "GKMYR26027-02", "sheet": "左右", "col": 9, "rows": [
                ("JBC0412X0056", "循环托盘", "1140x1480x130", "套", 27, "客供件"),
            ]},
        ])

    def tearDown(self):
        """恢复主数据库环境并删除对账单输出。"""

        if self.old_catalog is None:
            os.environ.pop("FYT_CATALOG_PATH", None)
        else:
            os.environ["FYT_CATALOG_PATH"] = self.old_catalog
        self.temp.cleanup()

    def test_scan_detects_batches_and_yellow_rows(self):
        """扫描应识别批次块及黄色人工排除行，并返回复核所需结构。"""

        result = reconcile_statement_core.scan([self.file_a, self.file_b])
        files = {item["name"]: item for item in result["files"]}
        self.assertEqual(files["（已下单）甲供应商7月采购清单明细.xlsx"]["supplier"], "甲供应商")  # 文件名识别供应商
        # 文件名识别不出时，从数据行供应商列兜底（众数）
        self.assertEqual(files["清单未命名.xlsx"]["supplier"], "客供件")  # 数据行兜底
        batches = files["（已下单）甲供应商7月采购清单明细.xlsx"]["batches"]
        self.assertEqual([b["batch"] for b in batches], ["GKMYR26027-01", "GKMYR26027-02"])  # 两个批次
        self.assertEqual(batches[0]["rows"], 3)  # 三行数据
        self.assertEqual(batches[0]["excluded_rows"], 1)  # 黄行排除
        # 并排双块：同一 sheet 两个批次都被识别
        side_batches = files["清单未命名.xlsx"]["batches"]
        self.assertEqual([b["batch"] for b in side_batches], ["GKMYR26027-01", "GKMYR26027-02"])
        self.assertEqual(side_batches[0]["excluded_rows"], 1)

    def test_build_groups_by_supplier_and_excludes_yellow(self):
        """生成对账单按供应商分文件，黄色行不得进入汇总数量。"""

        scan = reconcile_statement_core.scan([self.file_a, self.file_b])
        selected = ["1:GKMYR26027-01", "1:GKMYR26027-02", "2:GKMYR26027-01"]
        out_dir = os.path.join(self.temp.name, "输出")
        result = reconcile_statement_core.build(
            [self.file_a, self.file_b], selected, "202607", out_dir=out_dir)
        names = {item["name"]: item for item in result["files"]}
        self.assertEqual(set(names), {"甲供应商202607月对单明细.xlsx", "客供件202607月对单明细.xlsx"})  # 按供应商分文件
        self.assertEqual(names["甲供应商202607月对单明细.xlsx"]["rows"], 3)  # 3+1-1(标黄)
        self.assertEqual(names["客供件202607月对单明细.xlsx"]["rows"], 1)  # 2-1(标黄)

        workbook = openpyxl.load_workbook(names["甲供应商202607月对单明细.xlsx"]["path"], data_only=True)
        worksheet = workbook.active
        self.assertEqual(worksheet.cell(1, 1).value, "202607月甲供应商对账单")  # 标题
        self.assertEqual([c.value for c in worksheet[2]], ["序号", "材料编号", "材料名称", "规格", "单位", "采购数量", "批次号", "备注"])  # 固定表头
        rows = list(worksheet.iter_rows(min_row=3, values_only=True))
        codes = [row[1] for row in rows]
        self.assertNotIn("JBC0101A0002", codes)          # 标黄行被排除
        self.assertEqual(rows[2][6], "GKMYR26027-02")    # 批次号列
        self.assertEqual(rows[2][0], 3)                  # 序号连续
        workbook.close()

    def test_build_supplier_override_and_validation(self):
        """人工供应商覆盖只接受已扫描批次，并在输出分组中生效。"""

        scan = reconcile_statement_core.scan([self.file_b])
        selected = ["1:GKMYR26027-01"]
        out_dir = os.path.join(self.temp.name, "输出2")
        result = reconcile_statement_core.build(
            [self.file_b], selected, "7", supplier_map={"1": "丙供应商"}, out_dir=out_dir)
        self.assertEqual(result["files"][0]["name"], "丙供应商7月对单明细.xlsx")  # 人工覆盖生效
        with self.assertRaises(ValueError):
            reconcile_statement_core.build([self.file_b], [], "202607", out_dir=out_dir)  # 空选择拒绝
        with self.assertRaises(ValueError):
            reconcile_statement_core.build([self.file_b], selected, "", out_dir=out_dir)  # 空月份拒绝

    def test_vertical_blocks_without_headers_do_not_mix_rows(self):
        """纵向批次续块缺少重复表头时仍需保持批次边界，不能串行混入上一块。"""

        path = os.path.join(self.temp.name, "纵向清单.xlsx")
        _write_file(path, [
            {"batch": "BATCH01", "sheet": "纵向", "col": 1, "row": 1, "header": False,
             "rows": [("JBC0101A0001", "第一批", "A", "件", 1, "甲")]},
            {"batch": "BATCH02", "sheet": "纵向", "col": 1, "row": 3, "header": False,
             "rows": [("JBC0101A0002", "第二批", "B", "件", 2, "甲")]},
        ])
        result = reconcile_statement_core.scan([path])
        batches = result["files"][0]["batches"]
        self.assertEqual([(item["batch"], item["rows"]) for item in batches], [
            ("BATCH01", 1), ("BATCH02", 1),
        ])  # 无表头续块仍保持边界

    def test_three_horizontal_blocks_and_second_row_are_all_detected(self):
        """同表三组横向批次及每组第二数据行都必须被扫描和导出。"""

        path = os.path.join(self.temp.name, "客供件三列清单.xlsx")
        blocks = []
        for batch_index, (row, column) in enumerate((
            (1, 1), (1, 9), (1, 17),
            (20, 1), (20, 9), (20, 17),
        ), start=1):
            blocks.append({
                "batch": "GKMYR26027-%02d" % batch_index,
                "sheet": "Sheet1",
                "col": column,
                "row": row,
                "rows": [(
                    "JBC0101A%04d" % batch_index,
                    "第%d批数据" % batch_index,
                    "规格%d" % batch_index,
                    "件",
                    batch_index,
                    "客供件",
                )],
            })
        _write_file(path, blocks)

        result = reconcile_statement_core.scan([path])
        batches = result["files"][0]["batches"]
        self.assertEqual(
            [(item["batch"], item["rows"]) for item in batches],
            [("GKMYR26027-%02d" % index, 1) for index in range(1, 7)],
        )  # 六块全识别

        out_dir = os.path.join(self.temp.name, "三列输出")
        built = reconcile_statement_core.build(
            [path], ["1:GKMYR26027-03"], "202607", out_dir=out_dir)
        self.assertEqual(built["total_rows"], 1)  # 导出指定批次一行
        workbook = openpyxl.load_workbook(built["files"][0]["path"], data_only=True)
        try:
            self.assertEqual(workbook.active.cell(3, 2).value, "JBC0101A0003")  # 物料编码
            self.assertEqual(workbook.active.cell(3, 7).value, "GKMYR26027-03")  # 批次号
        finally:
            workbook.close()

    def test_supplier_override_cannot_escape_output_directory(self):
        selected = ["1:GKMYR26027-01"]
        out_dir = os.path.join(self.temp.name, "安全输出")
        result = reconcile_statement_core.build(
            [self.file_b], selected, "../202607", supplier_map={"1": "../越界"}, out_dir=out_dir)
        target = result["files"][0]["path"]
        self.assertEqual(os.path.commonpath([out_dir, target]), out_dir)  # 路径不越界
        self.assertNotIn("..", os.path.basename(target))  # 文件名无上级引用
        self.assertEqual(result["out_dir"], out_dir)

    def test_supplier_from_filename(self):
        self.assertEqual(reconcile_statement_core.supplier_from_filename("（已下单）众瀚7月采购清单明细.xlsx"), "众瀚")  # 中文括号
        self.assertEqual(reconcile_statement_core.supplier_from_filename("(已下单）圣直7月采购清单明细.xlsx"), "圣直")  # 混合括号
        self.assertIsNone(reconcile_statement_core.supplier_from_filename("采购清单.xlsx"))  # 无标记返回空
        self.assertIsNone(reconcile_statement_core.supplier_from_filename("（已下单）7月采购清单明细.xlsx"))  # 无供应商名


if __name__ == "__main__":
    unittest.main()
