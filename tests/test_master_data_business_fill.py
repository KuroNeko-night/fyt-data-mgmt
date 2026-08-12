# -*- coding: utf-8 -*-
"""主数据库缺失字段补全：相关业务统一只补空值的端到端回归。"""
from __future__ import annotations

import os
import tempfile
import unittest

import openpyxl

from core import arrival_core, delivery_core, material_catalog, pivot_core
from core import purchase_core, purchase_plan_core, reconcile_statement_core
from core import supplier_batch_core


def _save_rows(path: str, title: str, rows) -> None:
    """把二维合成数据写成单工作表 xlsx，供多个业务复用。"""

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = title
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()


class MasterDataBusinessFillTests(unittest.TestCase):
    """验证所有已接入业务只用主数据库补空值，不覆盖源表显式内容。"""

    def setUp(self):
        """为每个用例创建独立主数据库和输出目录。"""

        self.temp = tempfile.TemporaryDirectory(prefix="fyt_master_fill_")
        self.old_catalog = os.environ.get("FYT_CATALOG_PATH")
        os.environ["FYT_CATALOG_PATH"] = self.path("主数据.json")

    def tearDown(self):
        """恢复调用前主数据库路径并删除合成业务文件。"""

        if self.old_catalog is None:
            os.environ.pop("FYT_CATALOG_PATH", None)
        else:
            os.environ["FYT_CATALOG_PATH"] = self.old_catalog
        self.temp.cleanup()

    def path(self, name: str) -> str:
        """返回当前测试临时根下路径。"""

        return os.path.join(self.temp.name, name)

    def seed(self, code="JBC900", name="主库材料", spec="100×50",
             unit="件", supplier="主库供应商", supplier_code="GYS900"):
        """写入一组完整物料与供应商关系，作为各业务补全基准。"""

        material_catalog.upsert_supplier(supplier, supplier_code)
        material_catalog.upsert_material(
            code, name, spec=spec, unit=unit, supplier=supplier)

    def test_arrival_fills_missing_name_and_supplier(self):
        """每日到料应按规范化编码补齐名称和供应商，并记录补全计数。"""

        self.seed(code="100")
        source = self.path("订单26001批次送货计划.xlsx")
        _save_rows(source, "计划", [
            ["物料编码", "物料名称", "供应商信息", "需求数", "剩余未收数"],
            [100.0, "", "", 8, 2],
        ])
        counts = {}
        batches, _memory = arrival_core.build_batches(
            [{"path": source, "batch_no": "26001", "total": 1, "include": True}],
            "截止16点", resolver=material_catalog.CatalogResolver(), fill_counts=counts)
        self.assertEqual(batches[0]["materials"][0][1:3], ["主库材料", "主库供应商"])
        self.assertEqual(counts, {"name": 1, "supplier": 1})

    def test_delivery_fills_name_supplier_and_supplier_code(self):
        """送货计划输出应补齐物料名称、供应商编码和供应商名称。"""

        self.seed(code="100")
        source = self.path("物料清单.xlsx")
        _save_rows(source, "BOM", [
            ["物料编码", "物料名称", "数量"],
            [100.0, "", 3],
        ])
        result = delivery_core.run(source, out_dir=self.path("送货输出"))
        workbook = openpyxl.load_workbook(result["plan_path"], data_only=True)
        try:
            row = [workbook.active.cell(3, column).value for column in range(2, 6)]
        finally:
            workbook.close()
        self.assertEqual(row, [100, "主库材料", "GYS900", "主库供应商"])
        self.assertTrue(result["supplier_used"])

    def test_supplier_batch_uses_catalog_after_current_and_history_sources(self):
        """当前表与历史表仍无供应商时，批次表分析才回退使用正式主库。"""

        self.seed()
        source = self.path("GMIDR26199A辅料清单总表.xlsx")
        _save_rows(source, "汇总", [
            [None] * 6,
            [None] * 6,
            ["材料编号", "材料名称", "规格", "单位", "求和项:最终采购数量", "供应商"],
            ["JBC900", "", "", "", 2, ""],
        ])
        plan = supplier_batch_core.analyze([source])
        self.assertEqual(plan["suppliers"][0]["name"], "主库供应商")
        self.assertEqual(plan["unmatched_count"], 0)
        self.assertEqual(plan["batches"][0]["rows"], 1)

    def test_purchase_plan_and_diff_fill_catalog_fields(self):
        """采购导入与实收差异两个输出都应复用同一主数据补全结果。"""

        self.seed()
        template = self.path("采购计划导入模板.xlsx")
        workbook = openpyxl.Workbook()
        main = workbook.active
        main.title = "采购计划导入模板"
        main.append(["仓库编号", "采购员编号", "产品编号", "产品名称", "规格",
                     "供应商编码", "供应商", "数量", "备注", "预计到货日期"])
        main.append([None] * 10)
        suppliers = workbook.create_sheet("Sheet1")
        suppliers.append([None, "其他供应商", "OTHER"])
        workbook.save(template)
        workbook.close()

        batch = self.path("GMIDR26199A辅料清单总表.xlsx")
        _save_rows(batch, "汇总", [
            [None] * 8,
            [None] * 8,
            ["材料编号", "材料名称", "规格", "单位", "最终采购数量", "供应商", "实收", "差异"],
            ["JBC900", "", "", "", 5, "", 4, -1],
        ])
        result = purchase_plan_core.run([template], [batch], out_dir=self.path("采购计划输出"))
        workbook = openpyxl.load_workbook(result["files"][0], data_only=True)
        try:
            values = [workbook["采购计划导入模板"].cell(2, column).value
                      for column in range(3, 9)]
        finally:
            workbook.close()
        self.assertEqual(values, ["JBC900", "主库材料", "100×50", "GYS900", "主库供应商", 5])

        diff = purchase_plan_core.diff([batch], out_dir=self.path("差异输出"))
        workbook = openpyxl.load_workbook(diff["path"], data_only=True)
        try:
            values = [workbook.active.cell(2, column).value for column in range(2, 6)]
        finally:
            workbook.close()
        self.assertEqual(values, ["JBC900", "主库材料", "100×50", "件"])

    def test_pivot_fills_fields_before_clustering(self):
        """销售表聚类前必须先补齐名称、规格和单位，避免同物料被错误拆组。"""

        self.seed(code="M900")
        source = self.path("销售输入.xlsx")
        _save_rows(source, "包装方案汇总", [
            ["版本序号", "材料编号", "材料名称", "规格", "数量", "单位", "最终采购数量"],
            [1, "M900", "", "", 1, "", 3],
        ])
        plan = pivot_core.analyze_workbooks([source])
        row = plan["sheets"][0]["kept"][0]
        self.assertEqual(
            [row[pivot_core.F_NAME], row[pivot_core.F_SPEC], row[pivot_core.F_UNIT]],
            ["主库材料", "100×50", "件"],
        )

    def test_reconcile_statement_fills_output_and_supplier_group(self):
        """对账单扫描应按主库归供应商组，输出明细也需补齐物料字段。"""

        self.seed()
        source = self.path("采购清单.xlsx")
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.cell(1, 1, "BATCH01交付日期8.5")
        headers = ["材料编号", "材料名称", "规格", "单位", "采购数量", "供应商"]
        for column, value in enumerate(headers, start=1):
            worksheet.cell(2, column, value)
        worksheet.cell(3, 1, "JBC900")
        worksheet.cell(3, 5, 2)
        workbook.save(source)
        workbook.close()

        scanned = reconcile_statement_core.scan([source])
        self.assertEqual(scanned["files"][0]["supplier"], "主库供应商")
        built = reconcile_statement_core.build(
            [source], ["1:BATCH01"], "8", out_dir=self.path("对账单输出"))
        workbook = openpyxl.load_workbook(built["files"][0]["path"], data_only=True)
        try:
            values = [workbook.active.cell(3, column).value for column in range(2, 6)]
        finally:
            workbook.close()
        self.assertEqual(values, ["JBC900", "主库材料", "100×50", "件"])

    def test_purchase_reconciliation_fills_before_matching_and_in_output_copy(self):
        """采购对账应在匹配前补全双方，并把补全值写入输出副本。"""

        self.seed()
        left = self.path("我方.xlsx")
        right = self.path("供方.xlsx")
        rows = [
            ["材料编号", "材料名称", "规格", "单位", "数量"],
            ["JBC900", "", "", "", 2],
        ]
        _save_rows(left, "Sheet1", rows)
        _save_rows(right, "Sheet1", rows)
        result = purchase_core.run(left, right, out_dir=self.path("采购对账输出"))
        self.assertEqual(len(result["pairs"]), 1)
        workbook = openpyxl.load_workbook(result["out1"], data_only=True)
        try:
            values = [workbook.active.cell(2, column).value for column in range(2, 5)]
        finally:
            workbook.close()
        self.assertEqual(values, ["主库材料", "100×50", "件"])


if __name__ == "__main__":
    unittest.main()
