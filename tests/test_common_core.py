# -*- coding: utf-8 -*-
"""common_core 纯解析/工具函数单元测试。"""
import datetime
import unittest

from core import common_core as cc


class TestNormName(unittest.TestCase):
    def test_removes_all_spaces(self):
        self.assertEqual(cc.norm_name("张 三"), "张三")
        self.assertEqual(cc.norm_name("  李四  "), "李四")
        self.assertEqual(cc.norm_name("a\tb c"), "abc")

    def test_none_and_number(self):
        self.assertEqual(cc.norm_name(None), "")
        self.assertEqual(cc.norm_name(123), "123")


class TestNormDate(unittest.TestCase):
    def test_datetime_and_date(self):
        self.assertEqual(cc.norm_date(datetime.datetime(2026, 5, 1, 8, 0)), (2026, 5, 1))
        self.assertEqual(cc.norm_date(datetime.date(2026, 12, 31)), (2026, 12, 31))

    def test_string_forms(self):
        self.assertEqual(cc.norm_date("20260501"), (2026, 5, 1))
        self.assertEqual(cc.norm_date("2026-05-01"), (2026, 5, 1))
        self.assertEqual(cc.norm_date("2026/5/1"), (2026, 5, 1))
        self.assertEqual(cc.norm_date("2026.5.1"), (2026, 5, 1))

    def test_invalid(self):
        self.assertIsNone(cc.norm_date(None))
        self.assertIsNone(cc.norm_date(""))
        self.assertIsNone(cc.norm_date("-"))
        self.assertIsNone(cc.norm_date("not a date"))


class TestDayOf(unittest.TestCase):
    def test_various(self):
        self.assertEqual(cc.day_of(datetime.date(2026, 5, 17)), 17)
        self.assertEqual(cc.day_of(5), 5)
        self.assertEqual(cc.day_of("20260517"), 17)
        self.assertEqual(cc.day_of("5日"), 5)

    def test_out_of_range(self):
        self.assertIsNone(cc.day_of(0))
        self.assertIsNone(cc.day_of(32))
        self.assertIsNone(cc.day_of("abc"))


class TestParseTime(unittest.TestCase):
    def test_time_and_datetime(self):
        self.assertEqual(cc.parse_time(datetime.time(8, 30)), datetime.time(8, 30))
        self.assertEqual(cc.parse_time(datetime.datetime(2026, 5, 1, 9, 15)),
                         datetime.time(9, 15))

    def test_string(self):
        self.assertEqual(cc.parse_time("08:30"), datetime.time(8, 30))
        self.assertEqual(cc.parse_time("8:05:30"), datetime.time(8, 5, 30))

    def test_invalid(self):
        self.assertIsNone(cc.parse_time(None))
        self.assertIsNone(cc.parse_time("-"))
        self.assertIsNone(cc.parse_time("0830"))   # 无冒号
        self.assertIsNone(cc.parse_time("aa:bb"))


class TestRoundHalfHour(unittest.TestCase):
    def test_up(self):
        self.assertEqual(cc.round_half_hour(datetime.time(7, 56), "up"), datetime.time(8, 0))
        self.assertEqual(cc.round_half_hour(datetime.time(7, 31), "up"), datetime.time(8, 0))
        self.assertEqual(cc.round_half_hour(datetime.time(7, 30), "up"), datetime.time(7, 30))
        self.assertEqual(cc.round_half_hour(datetime.time(7, 1), "up"), datetime.time(7, 30))

    def test_down(self):
        self.assertEqual(cc.round_half_hour(datetime.time(8, 13), "down"), datetime.time(8, 0))
        self.assertEqual(cc.round_half_hour(datetime.time(8, 24), "down"), datetime.time(8, 0))
        self.assertEqual(cc.round_half_hour(datetime.time(8, 30), "down"), datetime.time(8, 30))
        self.assertEqual(cc.round_half_hour(datetime.time(8, 59), "down"), datetime.time(8, 30))

    def test_none(self):
        self.assertIsNone(cc.round_half_hour(None, "up"))


class TestHoursHelpers(unittest.TestCase):
    def test_to_hours(self):
        self.assertAlmostEqual(cc.to_hours(datetime.time(8, 30)), 8.5)
        self.assertAlmostEqual(cc.to_hours(datetime.time(9, 15, 36)), 9.26, places=2)
        self.assertIsNone(cc.to_hours(None))

    def test_fmt_time(self):
        self.assertEqual(cc.fmt_time(datetime.time(8, 5)), "08:05")
        self.assertEqual(cc.fmt_time(None), "")

    def test_parse_rest(self):
        self.assertEqual(cc.parse_rest(1.5), 1.5)
        self.assertEqual(cc.parse_rest("2"), 2.0)
        self.assertEqual(cc.parse_rest(None), 0.0)
        self.assertEqual(cc.parse_rest("-"), 0.0)
        self.assertEqual(cc.parse_rest("abc"), 0.0)


class TestToNum(unittest.TestCase):
    def test_numbers(self):
        self.assertEqual(cc.to_num(9), 9.0)
        self.assertEqual(cc.to_num("8.5"), 8.5)

    def test_skip_marks(self):
        self.assertIsNone(cc.to_num("休"))
        self.assertIsNone(cc.to_num("假"))
        self.assertIsNone(cc.to_num(""))
        self.assertIsNone(cc.to_num(None))

    def test_custom_skip(self):
        skip = {"NA"}
        self.assertIsNone(cc.to_num("NA", skip=skip))
        # 自定义集合不含"休"时，"休"不再被跳过，但也不是数字 -> None
        self.assertIsNone(cc.to_num("休", skip=skip))


class TestOptions(unittest.TestCase):
    def test_defaults(self):
        o = cc.Options()
        self.assertEqual(o.workday_hours, 9.0)
        self.assertTrue(o.overtime)
        self.assertEqual(o.conflict, "last")
        self.assertIn("休", o.skip_set())

    def test_conflict_normalized(self):
        self.assertEqual(cc.Options(conflict="bogus").conflict, "last")
        self.assertEqual(cc.Options(conflict="first").conflict, "first")

    def test_per_file_map(self):
        o = cc.Options(columns={"a.xlsx": {"sheet": "总表",
                                           "roles": {"name": 0}}})
        self.assertEqual(o.resolve_sheet("/x/y/a.xlsx"), "总表")
        self.assertEqual(o.resolve_roles("a.xlsx"), {"name": 0})
        self.assertIsNone(o.resolve_sheet("other.xlsx"))

    def test_summary_is_string(self):
        self.assertIsInstance(cc.Options().summary(), str)


class TestLoadDataOnly(unittest.TestCase):
    """load_data_only 跳过透视缓存后,单元格值必须与常规 data_only 逐格一致。

    额外防回归:openpyxl 若升级改了内部结构,monkeypatch 应安全退化而非报错/丢数。"""

    def test_values_match_plain_data_only(self):
        import os
        import shutil
        import tempfile
        import openpyxl
        d = tempfile.mkdtemp()
        p = os.path.join(d, "t.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["编码", "数量", "小计"])
        ws.append(["M01", 3, 6])
        ws.append(["M02", 5, 10])
        wb.save(p)
        wb.close()
        try:
            def cells(wb):
                return [[list(r) for r in ws.iter_rows(values_only=True)]
                        for ws in wb.worksheets]
            w1 = openpyxl.load_workbook(p, data_only=True)
            a = cells(w1); w1.close()
            w2 = cc.load_data_only(p)
            b = cells(w2); w2.close()
            self.assertEqual(a, b)
        finally:
            shutil.rmtree(d, ignore_errors=True)

    def test_parser_property_restored(self):
        # 上下文退出后必须还原全局属性,不能污染后续普通读取
        from openpyxl.reader.workbook import WorkbookParser
        before = WorkbookParser.pivot_caches
        with cc._skip_pivot_cache_parse():
            pass
        self.assertIs(WorkbookParser.pivot_caches, before)


if __name__ == "__main__":
    unittest.main()
