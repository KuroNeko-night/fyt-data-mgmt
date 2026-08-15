# -*- coding: utf-8 -*-
"""Web 服务文件库、上传/回收站、备份与路径安全回归测试。"""
from __future__ import annotations

import hashlib
import json
import urllib.parse
import urllib.request
import zipfile
from datetime import datetime, timedelta, timezone
from io import BytesIO

from openpyxl import Workbook

import web_server

from tests.web_server_test_base import WebServerTestBase


class WebServerFileTests(WebServerTestBase):
    """文件库、上传句柄、回收站、备份、主数据导入与路径越权防护。"""

    def test_shared_library_permissions_pagination_and_restore(self):
        """共享数据库需验证角色权限、分页、修改时间、软删除和恢复流程。"""

        tokens = {}
        for username, display_name in (("library_a", "资料甲"), ("library_b", "资料乙")):
            self.assertEqual(self.call("/api/auth/register", {
                "username": username, "display_name": display_name, "password": "password123",
            })[0], 201)
            account = next(
                item for item in self.call("/api/admin/users", token=self.admin)[1]["users"]
                if item["username"] == username
            )
            self.assertEqual(
                self.call(f"/api/admin/users/{account['id']}/approve", {}, token=self.admin)[0],
                200,
            )
            self.assertEqual(
                self.call(
                    f"/api/admin/users/{account['id']}/role",
                    {"role": "team_leader"}, token=self.admin,
                )[0],
                200,
            )
            tokens[username] = self.call("/api/auth/login", {
                "username": username, "password": "password123",
            })[1]["token"]

        def upload(name, content, scope="team"):
            """上传共享资料并返回服务端记录，允许测试切换共享范围。"""

            query = urllib.parse.urlencode({
                "name": name, "scope": scope, "description": f"{name}说明",
            })
            status, payload = self.call(
                f"/api/library/files?{query}", token=tokens["library_a"], raw=content,
                headers={"Content-Length": str(len(content))},
            )
            self.assertEqual(status, 201)
            return payload["file"]

        team_file = upload("团队资料.txt", b"team-v1")
        private_file = upload("个人资料.txt", b"private", "private")
        second_team = upload("团队资料二.txt", b"team-v2")
        self.assertEqual(team_file["uploader"]["display_name"], "资料甲")
        self.assertEqual(team_file["scope"], "team")
        self.assertTrue(team_file["permissions"]["can_edit"])

        status, listed = self.call(
            "/api/library/files?page=1&page_size=1", token=tokens["library_b"],
        )
        self.assertEqual(status, 200)
        self.assertEqual(listed["pagination"], {"page": 1, "page_size": 1, "total": 2, "pages": 2})
        self.assertEqual(len(listed["files"]), 1)
        self.assertFalse(listed["files"][0]["permissions"]["can_edit"])
        self.assertEqual(listed["summary"]["visible_count"], 2)

        status, content = self.call(
            f"/api/library/files/{team_file['id']}/download", token=tokens["library_b"],
        )
        self.assertEqual((status, content), (200, b"team-v1"))
        self.assertEqual(self.call(
            f"/api/library/files/{private_file['id']}/download", token=tokens["library_b"],
        )[0], 404)
        self.assertEqual(self.call(
            f"/api/library/files/{team_file['id']}",
            {"name": "越权修改.txt", "scope": "team"},
            token=tokens["library_b"], method="PATCH",
        )[0], 403)

        status, updated = self.call(
            f"/api/library/files/{team_file['id']}",
            {"name": "团队资料-更新.txt", "description": "已核对", "scope": "team"},
            token=tokens["library_a"], method="PATCH",
        )
        self.assertEqual(status, 200)
        self.assertEqual(updated["file"]["updated_by"]["username"], "library_a")
        self.assertTrue(updated["file"]["updated_at"])

        replace_query = urllib.parse.urlencode({"name": "团队资料-替换.txt"})
        status, replaced = self.call(
            f"/api/library/files/{team_file['id']}/content?{replace_query}",
            token=tokens["library_a"], raw=b"replacement",
            headers={"Content-Length": "11"},
        )
        self.assertEqual(status, 200)
        self.assertEqual(replaced["file"]["size"], 11)
        self.assertEqual(self.call(
            f"/api/library/files/{team_file['id']}/download", token=tokens["library_b"],
        ), (200, b"replacement"))

        status, admin_view = self.call(
            "/api/library/files?scope=private", token=self.admin,
        )
        self.assertEqual(status, 200)
        self.assertEqual([item["id"] for item in admin_view["files"]], [private_file["id"]])
        self.assertTrue(admin_view["files"][0]["permissions"]["can_delete"])

        self.assertEqual(self.call(
            f"/api/library/files/{second_team['id']}", token=tokens["library_a"], method="DELETE",
        )[0], 200)
        trash = self.call("/api/admin/trash", token=self.admin)[1]["trash"]
        deleted = next(item for item in trash if item["label"] == "团队资料二.txt")
        self.assertEqual(deleted["kind"], "library_file")
        self.assertEqual(self.call(
            f"/api/admin/trash/{deleted['id']}/restore", {}, token=self.admin,
        )[0], 200)
        self.assertEqual(self.call(
            f"/api/library/files/{second_team['id']}/download", token=tokens["library_b"],
        ), (200, b"team-v2"))

    def test_library_classification_filter_override_and_replacement(self):
        """自动分类、筛选、人工覆盖和同名替换必须保持文件与索引一致。"""

        def workbook_bytes(headers, row):
            """构造单表分类样本字节，便于覆盖替换前后的类别变化。"""

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "数据表"
            sheet.append(headers)
            sheet.append(row)
            stream = BytesIO()
            workbook.save(stream)
            return stream.getvalue()

        supplier_content = workbook_bytes(
            ["批次号", "属性", "下阶物料", "下阶物料描述", "供应商代码", "供应商名称", "合计", "库区"],
            ["GK1", "KD", "8892602000", "右前踏板", "100079", "北京丰达", 360, "M62"],
        )
        query = urllib.parse.urlencode({"name": "供应商明细.xlsx", "scope": "team"})
        status, uploaded = self.call(
            f"/api/library/files?{query}", token=self.admin, raw=supplier_content,
            headers={"Content-Length": str(len(supplier_content))},
        )
        self.assertEqual(status, 201)
        file_id = uploaded["file"]["id"]
        self.assertEqual(uploaded["file"]["category"], "deliv_supp")
        self.assertIn("deliv_supp", uploaded["file"]["categories"])

        status, filtered = self.call(
            "/api/library/files?category=deliv_supp", token=self.admin,
        )
        self.assertEqual(status, 200)
        self.assertEqual([item["id"] for item in filtered["files"]], [file_id])

        status, edited = self.call(
            f"/api/library/files/{file_id}", {"category": "pivot_src"},
            token=self.admin, method="PATCH",
        )
        self.assertEqual(status, 200)
        self.assertEqual(edited["file"]["category"], "pivot_src")
        self.assertEqual(edited["file"]["confidence"], 100)

        pivot_content = workbook_bytes(
            ["版本序号", "材料编号", "材料名称", "规格", "数量", "单位", "最终采购数量"],
            [1, "MAT001", "纸箱", "600x400", 10, "个", 120],
        )
        replace_query = urllib.parse.urlencode({"name": "采购数据.xlsx"})
        status, replaced = self.call(
            f"/api/library/files/{file_id}/content?{replace_query}",
            token=self.admin, raw=pivot_content,
            headers={"Content-Length": str(len(pivot_content))},
        )
        self.assertEqual(status, 200)
        self.assertEqual(replaced["file"]["category"], "pivot_src")
        self.assertGreaterEqual(replaced["file"]["confidence"], 50)

    def test_library_category_migration_keeps_secondary_labels(self):
        """分类索引迁移后应保留多标签中的次要业务分类，避免搜索能力退化。"""

        legacy_id = "legacy-category-row"
        timestamp = web_server.now_iso()
        with web_server.DB_LOCK, web_server.db() as connection:
            connection.execute(
                "INSERT INTO library_files(id, owner_id, name, path, size, content_type, description, scope, "
                "created_at, updated_at, updated_by, category, categories, confidence, signals, sheet, category_sheets) "
                "VALUES (?, 1, ?, ?, 0, 'application/octet-stream', '', 'team', ?, ?, 1, ?, ?, 80, '[]', '', '{}')",
                (legacy_id, "旧文件.xlsx", str(self.temp.name), timestamp, timestamp, "pivot_src", "[\"deliv_supp\"]"),
            )
        web_server.init_db()
        with web_server.DB_LOCK, web_server.db() as connection:
            row = connection.execute(
                "SELECT categories FROM library_files WHERE id = ?", (legacy_id,)
            ).fetchone()
            labels = {
                item["category"] for item in connection.execute(
                    "SELECT category FROM library_file_categories WHERE file_id = ?", (legacy_id,)
                ).fetchall()
            }
        self.assertEqual(set(json.loads(row["categories"])), {"pivot_src", "deliv_supp"})
        self.assertEqual(labels, {"pivot_src", "deliv_supp"})

    def test_upload_trash_restore_and_permanent_delete(self):
        """普通上传进入回收站后可恢复，永久删除则同时清理文件与元数据。"""

        query = urllib.parse.urlencode({"name": "可恢复资料.txt", "group": "trash-test"})
        status, uploaded = self.call(
            f"/api/files/upload?{query}", token=self.admin, raw=b"recoverable",
            headers={"Content-Length": "11"},
        )
        self.assertEqual(status, 201)
        status, payload = self.call(
            f"/api/admin/uploads/{uploaded['handle']}", token=self.admin, method="DELETE",
        )
        self.assertEqual(status, 200)
        self.assertIn("回收站", payload["message"])
        status, trash = self.call("/api/admin/trash", token=self.admin)
        self.assertEqual(status, 200)
        item = trash["trash"][0]
        self.assertEqual(item["kind"], "upload")
        status, _ = self.call(
            f"/api/admin/trash/{item['id']}/restore", {}, token=self.admin,
        )
        self.assertEqual(status, 200)
        data = self.call("/api/admin/data", token=self.admin)[1]
        self.assertIn(uploaded["handle"], {row["handle"] for row in data["uploads"]})

        self.call(f"/api/admin/uploads/{uploaded['handle']}", token=self.admin, method="DELETE")
        item = self.call("/api/admin/trash", token=self.admin)[1]["trash"][0]
        status, _ = self.call(
            f"/api/admin/trash/{item['id']}", token=self.admin, method="DELETE",
        )
        self.assertEqual(status, 200)
        self.assertEqual(self.call("/api/admin/trash", token=self.admin)[1]["trash"], [])

    def test_job_trash_restores_record_and_result_file(self):
        """任务软删除与恢复必须成对处理数据库记录和所属结果文件。"""

        job_id = "recoverable-job"
        output = web_server.DATA_ROOT / "users" / "1" / "jobs" / job_id / "outputs" / "结果.txt"
        output.parent.mkdir(parents=True)
        output.write_bytes(b"result")
        timestamp = web_server.now_iso()
        with web_server.DB_LOCK, web_server.db() as connection:
            connection.execute(
                "INSERT INTO web_jobs(id, user_id, assignee_id, action, title, status, progress, logs, result, error, files, cancelled, payload, created_at, updated_at) "
                "VALUES (?, 1, 1, 'text.transform', '可恢复任务', 'completed', 100, '[]', '{}', NULL, ?, 0, '{}', ?, ?)",
                (job_id, json.dumps([{"name": "结果.txt", "path": str(output), "size": 6}], ensure_ascii=False), timestamp, timestamp),
            )
        status, payload = self.call(
            f"/api/admin/jobs/{job_id}", token=self.admin, method="DELETE",
        )
        self.assertEqual(status, 200)
        self.assertIn("回收站", payload["message"])
        item = self.call("/api/admin/trash", token=self.admin)[1]["trash"][0]
        self.assertEqual(item["kind"], "job")
        self.assertFalse(output.exists())
        self.assertEqual(self.call(
            f"/api/admin/trash/{item['id']}/restore", {}, token=self.admin,
        )[0], 200)
        self.assertTrue(output.is_file())
        with web_server.DB_LOCK, web_server.db() as connection:
            restored = connection.execute(
                "SELECT assignee_id FROM web_jobs WHERE id = ?", (job_id,),
            ).fetchone()
        self.assertEqual(restored["assignee_id"], 1)
        request = urllib.request.Request(
            self.base + f"/api/jobs/{job_id}/files/0", headers={"X-Session-Token": self.admin},
        )
        with urllib.request.urlopen(request, timeout=10) as response:
            self.assertEqual(response.read(), b"result")

    def test_storage_maintenance_limits_outputs_and_purges_expired_trash(self):
        """每用户只保留最近二十次输出，并按保留期清理回收站而不越权删除。"""

        now = datetime.now(timezone.utc).replace(microsecond=0)
        for index in range(22):
            job_id = f"retention-job-{index:02d}"
            output = web_server.DATA_ROOT / "users" / "1" / "jobs" / job_id / "outputs" / "结果.txt"
            output.parent.mkdir(parents=True)
            output.write_text(str(index), encoding="utf-8")
            timestamp = (now - timedelta(minutes=22 - index)).isoformat(timespec="seconds")
            files = json.dumps(
                [{"name": "结果.txt", "path": str(output), "size": output.stat().st_size}],
                ensure_ascii=False,
            )
            with web_server.DB_LOCK, web_server.db() as connection:
                connection.execute(
                    "INSERT INTO web_jobs(id, user_id, action, title, status, progress, logs, result, error, files, cancelled, payload, created_at, updated_at) "
                    "VALUES (?, 1, 'text.transform', ?, 'completed', 100, '[]', '{}', NULL, ?, 0, '{}', ?, ?)",
                    (job_id, f"输出任务 {index}", files, timestamp, timestamp),
                )
                if index == 0:
                    connection.execute(
                        "INSERT INTO web_job_versions(job_id, user_id, version, result, files, status, created_at) "
                        "VALUES (?, 1, 1, '{}', ?, 'completed', ?)",
                        (job_id, files, timestamp),
                    )

        review_id = "retention-review"
        review_output = web_server.DATA_ROOT / "users" / "1" / "jobs" / review_id / "outputs" / "复核.txt"
        review_output.parent.mkdir(parents=True)
        review_output.write_text("review", encoding="utf-8")
        review_files = json.dumps(
            [{"name": "复核.txt", "path": str(review_output), "size": 6}], ensure_ascii=False,
        )
        old_time = (now - timedelta(days=60)).isoformat(timespec="seconds")
        with web_server.DB_LOCK, web_server.db() as connection:
            connection.execute(
                "INSERT INTO web_jobs(id, user_id, action, title, status, progress, logs, result, error, files, cancelled, payload, created_at, updated_at) "
                "VALUES (?, 1, 'web.compare.review', '等待人工复核', 'completed', 100, '[]', '{}', NULL, ?, 0, '{}', ?, ?)",
                (review_id, review_files, old_time, old_time),
            )
            connection.execute(
                "INSERT INTO trash_items(id, kind, label, record_json, original_path, size, deleted_at) "
                "VALUES ('expired-trash', 'upload', '过期资料', '{}', 'users/1/uploads/expired.txt', 1, ?)",
                ((now - timedelta(days=31)).isoformat(timespec="seconds"),),
            )
            connection.execute(
                "INSERT INTO trash_items(id, kind, label, record_json, original_path, size, deleted_at) "
                "VALUES ('recent-trash', 'upload', '近期资料', '{}', 'users/1/uploads/recent.txt', 1, ?)",
                ((now - timedelta(days=29)).isoformat(timespec="seconds"),),
            )
        for trash_id in ("expired-trash", "recent-trash"):
            payload = web_server.DATA_ROOT / "trash" / trash_id / "payload"
            payload.parent.mkdir(parents=True)
            payload.write_text(trash_id, encoding="utf-8")

        report = web_server.run_storage_maintenance(
            output_limit=20, trash_retention_days=30, current_time=now,
        )
        self.assertEqual(report["moved_outputs"], 2)
        self.assertEqual(report["purged_trash"], 1)
        self.assertEqual(report["trash_cleanup_failures"], 0)
        self.assertFalse((web_server.DATA_ROOT / "trash" / "expired-trash").exists())
        self.assertTrue((web_server.DATA_ROOT / "trash" / "recent-trash" / "payload").is_file())

        with web_server.DB_LOCK, web_server.db() as connection:
            active_ids = {
                row["id"] for row in connection.execute("SELECT id FROM web_jobs").fetchall()
            }
            trash_rows = connection.execute("SELECT id, label FROM trash_items").fetchall()
        self.assertNotIn("retention-job-00", active_ids)
        self.assertNotIn("retention-job-01", active_ids)
        self.assertIn("retention-job-02", active_ids)
        self.assertIn(review_id, active_ids)
        self.assertTrue(review_output.is_file())

        archived = next(row for row in trash_rows if row["label"] == "输出任务 0")
        status, _ = self.call(
            f"/api/admin/trash/{archived['id']}/restore", {}, token=self.admin,
        )
        self.assertEqual(status, 200)
        restored_output = (
            web_server.DATA_ROOT / "users" / "1" / "jobs" / "retention-job-00" / "outputs" / "结果.txt"
        )
        self.assertTrue(restored_output.is_file())
        with web_server.DB_LOCK, web_server.db() as connection:
            versions = connection.execute(
                "SELECT COUNT(*) AS count FROM web_job_versions WHERE job_id = 'retention-job-00'"
            ).fetchone()["count"]
        self.assertEqual(versions, 1)

    def test_backup_verification_and_restore(self):
        """管理员备份需可校验、恢复并在失败路径保持当前数据库可用。"""

        query = urllib.parse.urlencode({"name": "备份资料.txt", "group": "backup-test"})
        status, uploaded = self.call(
            f"/api/files/upload?{query}", token=self.admin, raw=b"backup-content",
            headers={"Content-Length": "14"},
        )
        self.assertEqual(status, 201)
        status, created = self.call("/api/admin/backups", {}, token=self.admin)
        self.assertEqual(status, 201)
        backup_id = created["backup"]["id"]
        backup_path = web_server.DATA_ROOT / "backups" / f"{backup_id}.zip"
        manifest = web_server.verify_web_backup(backup_path)
        self.assertEqual(manifest["backup_id"], backup_id)
        self.call(f"/api/admin/uploads/{uploaded['handle']}", token=self.admin, method="DELETE")

        status, restored = self.call(
            f"/api/admin/backups/{backup_id}/restore",
            {"confirmation": "恢复备份"}, token=self.admin,
        )
        self.assertEqual(status, 200)
        self.assertTrue(restored["safety_backup_id"])
        self.assertEqual(self.call("/api/auth/me", token=self.admin)[0], 401)
        new_admin = self.call(
            "/api/auth/login", {"username": "admin", "password": "admin123456"},
        )[1]["token"]
        data = self.call("/api/admin/data", token=new_admin)[1]
        self.assertIn(uploaded["handle"], {row["handle"] for row in data["uploads"]})

    def test_backup_verification_rejects_duplicate_manifest_paths(self):
        """备份清单不得用重复路径掩盖同一文件的多次登记。"""

        backup_path = web_server.DATA_ROOT / "backups" / "duplicate-manifest.zip"
        backup_path.parent.mkdir(parents=True, exist_ok=True)
        database_payload = b"synthetic-database"
        entry = {
            "path": "database/accounts.sqlite3",
            "size": len(database_payload),
            "sha256": hashlib.sha256(database_payload).hexdigest(),
        }
        manifest = {"format": 1, "files": [entry, dict(entry)]}
        with zipfile.ZipFile(backup_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            archive.writestr("database/accounts.sqlite3", database_payload)
            archive.writestr(
                "manifest.json", json.dumps(manifest, ensure_ascii=False).encode("utf-8"),
            )
        with self.assertRaisesRegex(ValueError, "重复路径"):
            web_server.verify_web_backup(backup_path)

    def test_admin_master_data_upload_review_merge_and_permissions(self):
        """主数据上传必须经过冲突复核、确认合并和管理员权限约束。"""

        status, _ = self.call("/api/auth/register", {
            "username": "catalog_user", "display_name": "普通成员", "password": "password123",
        })
        self.assertEqual(status, 201)
        with web_server.DB_LOCK, web_server.db() as connection:
            member_id = connection.execute(
                "SELECT id FROM users WHERE username = ?", ("catalog_user",),
            ).fetchone()["id"]
        self.assertEqual(self.call(
            f"/api/admin/users/{member_id}/approve", {}, token=self.admin,
        )[0], 200)
        member_token = self.call(
            "/api/auth/login", {"username": "catalog_user", "password": "password123"},
        )[1]["token"]
        self.assertEqual(self.call("/api/admin/master-data/imports", token=member_token)[0], 403)

        # 现有逐条维护 POST 路由必须可用，并写入 Web 数据根而不是用户文档目录。
        status, catalog = self.call("/api/admin/catalog", {
            "op": "upsert_supplier", "name": "已有供应商", "code": "OLD01",
        }, token=self.admin)
        self.assertEqual(status, 200)
        self.assertEqual(catalog["suppliers"]["已有供应商"], "OLD01")
        self.assertTrue((web_server.DATA_ROOT / "catalog.json").is_file())

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "主数据"
        sheet.append(["物料编码", "物料名称", "规格型号", "单位", "供应商名称", "供应商编码"])
        sheet.append(["M001", "铁箱", "100x50", "个", "众瀚", "GYS01"])
        stream = BytesIO()
        workbook.save(stream)
        workbook.close()
        raw = stream.getvalue()
        query = urllib.parse.urlencode({"name": "管理员主数据.xlsx"})
        status, created = self.call(
            f"/api/admin/master-data/imports?{query}", token=self.admin, raw=raw,
            headers={"Content-Length": str(len(raw))},
        )
        self.assertEqual(status, 201)
        batch = created["batch"]
        self.assertEqual(batch["status"], "ready_to_confirm")
        self.assertEqual(batch["candidate_count"], 5)
        source_dir = (
            web_server.DATA_ROOT / "users" / str(batch["uploader_id"])
            / "master-data-imports" / batch["id"]
        )
        self.assertTrue((source_dir / "管理员主数据.xlsx").is_file())

        status, listing = self.call("/api/admin/master-data/imports", token=self.admin)
        self.assertEqual(status, 200)
        self.assertEqual(listing["summary"]["ready_to_confirm"], 1)
        self.assertEqual(self.call(
            f"/api/admin/master-data/imports/{batch['id']}/confirm", {}, token=self.admin,
        )[1]["batch"]["status"], "ready")
        merged = self.call(
            f"/api/admin/master-data/imports/{batch['id']}/merge", {}, token=self.admin,
        )[1]
        self.assertEqual(merged["batch"]["status"], "merged")
        catalog = self.call("/api/admin/catalog", token=self.admin)[1]
        self.assertEqual(catalog["suppliers"]["众瀚"], "GYS01")
        self.assertEqual(catalog["materials"]["M001"]["name"], "铁箱")

        duplicate_status, duplicate = self.call(
            f"/api/admin/master-data/imports?{query}", token=self.admin, raw=raw,
            headers={"Content-Length": str(len(raw))},
        )
        self.assertEqual(duplicate_status, 409)
        self.assertIn("已经上传", duplicate["error"])
        self.assertEqual(self.call("/api/admin/master-data/export", token=self.admin)[0], 200)

        conflict_book = Workbook()
        conflict_sheet = conflict_book.active
        conflict_sheet.append(["物料编码", "物料名称"])
        conflict_sheet.append(["M001", "冲突候选名称"])
        conflict_stream = BytesIO()
        conflict_book.save(conflict_stream)
        conflict_book.close()
        conflict_raw = conflict_stream.getvalue()
        conflict_query = urllib.parse.urlencode({"name": "冲突主数据.xlsx"})
        status, conflict_created = self.call(
            f"/api/admin/master-data/imports?{conflict_query}",
            token=self.admin,
            raw=conflict_raw,
            headers={"Content-Length": str(len(conflict_raw))},
        )
        self.assertEqual(status, 201)
        conflict_batch = conflict_created["batch"]
        self.assertEqual(conflict_batch["status"], "needs_review")
        status, detail = self.call(
            f"/api/admin/master-data/imports/{conflict_batch['id']}", token=self.admin,
        )
        self.assertEqual(status, 200)
        conflict = next(
            item for item in detail["batch"]["candidates"]
            if item["relation_type"] == "material_name"
        )
        status, resolved = self.call(
            f"/api/admin/master-data/imports/{conflict_batch['id']}/resolve",
            {"candidate_id": conflict["id"], "decision": "keep_current"},
            token=self.admin,
        )
        self.assertEqual(status, 200)
        self.assertEqual(resolved["batch"]["status"], "ready_to_confirm")
        status, rejected = self.call(
            f"/api/admin/master-data/imports/{conflict_batch['id']}/reject",
            {}, token=self.admin,
        )
        self.assertEqual(status, 200)
        self.assertEqual(rejected["batch"]["status"], "rejected")

    def test_upload_handle_rejects_cross_user_and_outside_paths(self):
        """上传句柄既不能跨账号复用，也不能通过伪造路径逃出所属上传根。"""

        status, _ = self.call("/api/auth/register", {
            "username": "path_user", "display_name": "路径测试", "password": "password123",
        })
        self.assertEqual(status, 201)
        with web_server.DB_LOCK, web_server.db() as connection:
            member = connection.execute(
                "SELECT id FROM users WHERE username = ?", ("path_user",),
            ).fetchone()
        self.assertEqual(self.call(f"/api/admin/users/{member['id']}/approve", token=self.admin, payload={})[0], 200)
        member_token = self.call(
            "/api/auth/login", {"username": "path_user", "password": "password123"},
        )[1]["token"]

        query = urllib.parse.urlencode({"name": "越权样本.txt", "group": "traversal-group"})
        status, uploaded = self.call(
            f"/api/files/upload?{query}", token=self.admin, raw=b"secret",
            headers={"Content-Length": "6"},
        )
        self.assertEqual(status, 201)

        # 跨用户引用管理员的上传句柄必须被拒绝
        status, payload = self.call("/api/jobs", {
            "action": "rename.apply", "title": "跨用户引用", "payload": {
                "paths": [uploaded["handle"]], "rule": {"prefix": "新-"},
            },
        }, token=member_token)
        self.assertEqual(status, 400)
        self.assertIn("不属于当前账号", payload["error"])

        # 普通账号提交 DATA_ROOT 之外的绝对路径必须被拒绝
        outside = web_server.DATA_ROOT.parent / "机密文件.txt"
        outside.write_text("敏感内容", encoding="utf-8")
        try:
            status, payload = self.call("/api/jobs", {
                "action": "rename.apply", "title": "越权路径", "payload": {
                    "paths": [str(outside)], "rule": {"prefix": "新-"},
                },
            }, token=member_token)
            self.assertEqual(status, 400)
            self.assertIn("不属于当前账号", payload["error"])
            status, payload = self.call("/api/jobs/preflight", {
                "action": "web.compare", "payload": {"file1": [str(outside)], "file2": [str(outside)]},
            }, token=member_token)
            self.assertEqual(status, 400)
            self.assertIn("不属于当前账号", payload["error"])
        finally:
            outside.unlink(missing_ok=True)

    def test_job_file_download_rejects_path_outside_owned_roots(self):
        """任务文件下载只允许任务所属输出根内路径，数据库伪造绝对路径也必须被拒绝。"""

        query = urllib.parse.urlencode({"name": "结果样本.txt", "group": "download-root"})
        status, uploaded = self.call(
            f"/api/files/upload?{query}", token=self.admin, raw=b"sample",
            headers={"Content-Length": "6"},
        )
        self.assertEqual(status, 201)
        status, created = self.call("/api/jobs", {
            "action": "rename.apply", "title": "下载校验", "payload": {
                "paths": [uploaded["handle"]], "rule": {"prefix": "新-"},
            },
        }, token=self.admin)
        self.assertEqual(status, 202)
        job = self.wait_job(created["job_id"])
        self.assertEqual(job["status"], "completed")
        self.assertEqual(len(job["files"]), 1)

        # 把结果记录篡改为任务目录之外的路径，下载必须 404
        outside = web_server.DATA_ROOT.parent / "越权下载.txt"
        outside.write_text("机密", encoding="utf-8")
        try:
            with web_server.DB_LOCK, web_server.db() as connection:
                connection.execute(
                    "UPDATE web_jobs SET files = ? WHERE id = ?",
                    (json.dumps([{"name": "越权下载.txt", "path": str(outside), "size": outside.stat().st_size}]), job["id"]),
                )
            status, payload = self.call(job["files"][0]["url"], token=self.admin)
            self.assertEqual(status, 404)
            self.assertIn("结果文件不存在", payload["error"])
            status, payload = self.call(job["files"][0]["url"] + "/preview", token=self.admin)
            self.assertEqual(status, 404)
            self.assertIn("结果文件不存在", payload["error"])
        finally:
            outside.unlink(missing_ok=True)


if __name__ == "__main__":
    unittest.main()
