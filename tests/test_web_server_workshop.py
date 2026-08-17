# -*- coding: utf-8 -*-
"""Web 服务现场问题（workshop）发布、权限、导出与清理回归测试。"""
from __future__ import annotations

import socket
import time
import urllib.parse
from datetime import datetime, timedelta, timezone
from io import BytesIO

from openpyxl import load_workbook
from PIL import Image

import web_server

from tests.web_server_test_base import WebServerTestBase


class WebServerWorkshopTests(WebServerTestBase):
    """现场问题五类模板、图片约束、导出与草稿维护的 HTTP 回归。"""

    def test_workshop_daily_issue_publish_permissions_export_and_restore(self):
        """现场问题草稿、图片、发布权限、按日展示、模板导出和回收站恢复应完整串联。"""

        tokens = {}
        for username, display_name in (("workshop_a", "车间甲"), ("workshop_b", "车间乙")):
            self.assertEqual(self.call("/api/auth/register", {
                "username": username, "display_name": display_name, "password": "password123",
            })[0], 201)  # 注册
            account = next(
                item for item in self.call("/api/admin/users", token=self.admin)[1]["users"]
                if item["username"] == username
            )
            self.assertEqual(
                self.call(f"/api/admin/users/{account['id']}/approve", {}, token=self.admin)[0],
                200,
            )  # 审核
            if username == "workshop_a":
                self.assertEqual(
                    self.call(
                        f"/api/admin/users/{account['id']}/role",
                        {"role": "team_leader"}, token=self.admin,
                    )[0],
                    200,
                )  # 甲为班组长
            tokens[username] = self.call("/api/auth/login", {
                "username": username, "password": "password123",
            })[1]["token"]  # 登录

        issue_date = datetime.now().strftime("%Y-%m-%d")  # 业务日
        future_date = (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%d")  # 未来日
        status, rejected = self.call("/api/workshop/issues", {
            "issue_date": future_date, "cause": "未来问题", "primary_owner": "负责人",
        }, token=tokens["workshop_a"])
        self.assertEqual(status, 400)  # 未来日期拒绝
        self.assertIn("不能晚于今天", rejected["error"])

        status, invalid_category = self.call("/api/workshop/issues", {
            "issue_date": issue_date,
            "cause": "旧分类不应继续使用",
            "category": "quality",
        }, token=tokens["workshop_a"])
        self.assertEqual(status, 400)  # 旧分类拒绝
        self.assertIn("主料异常", invalid_category["error"])

        status, wrong_fields = self.call("/api/workshop/issues", {
            "issue_date": issue_date,
            "cause": "主料异常不应接收海外字段",
            "category": "main_material",
            "discoverer": "李工",
            "country": "印度尼西亚",
        }, token=tokens["workshop_a"])
        self.assertEqual(status, 400)  # 模板外字段拒绝
        self.assertIn("不使用以下字段", wrong_fields["error"])

        status, created = self.call("/api/workshop/issues", {
            "issue_date": issue_date,
            "cause": "装配工位防护罩松动",
            "category": "main_material",
            "severity": "important",
            "model": "VX11",
            "batch_no": "GKMYR26027-06",
            "team": "小件组",
            "material_code": "A-01",
            "material_name": "防护罩",
            "cause_analysis": "紧固件松动",
            "corrective_action": "重新紧固并复核",
            "external_inspection_owner": "王工",
            "discoverer": "李工",
            "issue_level": "B",
            "quantity": "2",
            "issue_type": "装配",
            "completion_date": issue_date,
            "recurring": "否",
            "carrier": "承运商甲",
            "supplier": "供应商甲",
        }, token=tokens["workshop_a"])
        self.assertEqual(status, 201)  # 创建草稿成功
        issue_id = created["issue"]["id"]
        self.assertEqual(created["issue"]["batch_no"], "GKMYR26027-06")  # 模板字段保留
        self.assertEqual(created["issue"]["external_inspection_owner"], "王工")
        self.assertEqual(created["issue"]["status"], "draft")  # 初始草稿
        self.assertEqual(self.call(
            f"/api/workshop/issues?date={issue_date}", token=tokens["workshop_b"],
        )[1]["issues"], [])  # 他人草稿不可见
        self.assertEqual(self.call(
            f"/api/workshop/issues/{issue_id}/publish", {}, token=tokens["workshop_a"],
        )[0], 400)  # 无图片不可发布

        bad_query = urllib.parse.urlencode({"name": "损坏图片.jpg"})
        status, invalid = self.call(
            f"/api/workshop/issues/{issue_id}/images?{bad_query}",
            token=tokens["workshop_a"], raw=b"not-an-image",
            headers={"Content-Length": "12"},
        )
        self.assertEqual(status, 400)  # 非图片拒绝
        self.assertIn("损坏", invalid["error"])

        stream = BytesIO()
        Image.new("RGB", (80, 60), "#3d7df0").save(stream, format="PNG")  # 生成合法 PNG
        image_content = stream.getvalue()
        for index in range(web_server.MAX_WORKSHOP_IMAGES):
            image_query = urllib.parse.urlencode({"name": f"现场-{index + 1}.png"})
            status, uploaded = self.call(
                f"/api/workshop/issues/{issue_id}/images?{image_query}",
                token=tokens["workshop_a"], raw=image_content,
                headers={"Content-Length": str(len(image_content))},
            )
            self.assertEqual(status, 201)  # 图片上传成功
            self.assertEqual(len(uploaded["issue"]["images"]), index + 1)  # 图片计数递增
        status, limited = self.call(
            f"/api/workshop/issues/{issue_id}/images?{bad_query}",
            token=tokens["workshop_a"], raw=image_content,
            headers={"Content-Length": str(len(image_content))},
        )
        self.assertEqual(status, 400)
        self.assertIn("最多上传", limited["error"])

        status, published = self.call(
            f"/api/workshop/issues/{issue_id}/publish", {}, token=tokens["workshop_a"],
        )
        self.assertEqual(status, 200)
        self.assertEqual(published["issue"]["status"], "published")
        status, published_again = self.call(
            f"/api/workshop/issues/{issue_id}/publish", {}, token=tokens["workshop_a"],
        )
        self.assertEqual(status, 200)
        self.assertEqual(published_again["issue"]["status"], "published")
        status, listed = self.call(
            f"/api/workshop/issues?date={issue_date}", token=tokens["workshop_b"],
        )
        self.assertEqual(status, 200)
        self.assertEqual(listed["summary"], {"issue_count": 1, "image_count": 8, "open_count": 1, "resolved_count": 0})
        self.assertFalse(listed["issues"][0]["permissions"]["can_delete"])
        image_url = listed["issues"][0]["images"][0]["url"]
        status, downloaded_image = self.call(image_url, token=tokens["workshop_b"])
        self.assertEqual(status, 200)
        with Image.open(BytesIO(downloaded_image)) as normalized:
            self.assertEqual(normalized.size, (80, 60))
        self.assertEqual(self.call(
            f"/api/workshop/issues/{issue_id}", token=tokens["workshop_b"], method="DELETE",
        )[0], 403)

        status, exported = self.call(
            f"/api/workshop/issues/export?date={issue_date}", token=tokens["workshop_b"],
        )
        self.assertEqual(status, 200)
        workbook = load_workbook(BytesIO(exported))
        self.assertEqual(workbook.sheetnames, ["主料异常", "辅料异常", "包装异常", "海外历史记录", "防错异常", "问题一览表"])
        main_sheet = workbook["主料异常"]
        self.assertEqual(main_sheet["A1"].value, "问题编号")
        self.assertEqual(main_sheet["B1"].value, "主料故障信息")
        self.assertEqual(main_sheet["H2"].value, "故障描述")
        self.assertEqual(main_sheet["H3"].value, "装配工位防护罩松动")
        self.assertEqual(main_sheet["D3"].value, "GKMYR26027-06")
        self.assertEqual(main_sheet["T3"].value, "王工")
        self.assertEqual(main_sheet["T1"].value, None)
        self.assertEqual(main_sheet["T2"].value, "外检责任人")
        self.assertEqual(len(main_sheet._images), 8)
        self.assertEqual(workbook["辅料异常"]["B1"].value, "问题源")
        self.assertEqual(workbook["海外历史记录"]["A1"].value, "海外问题统计")
        self.assertEqual(workbook["防错异常"]["A2"].value, "发生时间")
        self.assertEqual(workbook["问题一览表"]["S2"].value, "供应商")
        self.assertEqual(workbook["问题一览表"].max_row, 2)

        status, removed = self.call(
            f"/api/workshop/issues/{issue_id}", token=tokens["workshop_a"], method="DELETE",
        )
        self.assertEqual(status, 200)
        self.assertIn("回收站", removed["message"])
        trash = self.call("/api/admin/trash", token=self.admin)[1]["trash"]
        item = next(row for row in trash if row["kind"] == "workshop_issue")
        self.assertEqual(self.call(
            f"/api/admin/trash/{item['id']}/restore", {}, token=self.admin,
        )[0], 200)
        restored = self.call(
            f"/api/workshop/issues?date={issue_date}", token=tokens["workshop_b"],
        )[1]
        self.assertEqual(restored["summary"], {"issue_count": 1, "image_count": 8, "open_count": 1, "resolved_count": 0})
        self.assertEqual(restored["issues"][0]["category"], "main_material")
        self.assertEqual(restored["issues"][0]["severity"], "important")

    def test_workshop_error_proofing_publishes_without_images(self):
        """防错异常按模板不强制图片，不能套用其余四类问题的上传限制。"""

        issue_date = datetime.now().strftime("%Y-%m-%d")
        status, created = self.call("/api/workshop/issues", {
            "issue_date": issue_date,
            "category": "error_proofing",
            "cause": "PDA 扫描到相邻二维码",
            "happened_at": f"{issue_date}T09:30",
            "batch_no": "GKIDR26004-02",
            "material_name": "后减振器总成",
            "material_code": "6608116475",
            "cause_analysis": "扫码方向未对准",
            "corrective_action": "质检班长现场核验并解除异常",
            "tracking_status": "校验完成",
            "handling_time": f"{issue_date}T09:40",
            "responsible_person": "黄林",
            "updated_by_name": "杨晓云",
            "notes": "无需上传照片",
        }, token=self.admin)
        self.assertEqual(status, 201)
        issue_id = created["issue"]["id"]
        self.assertEqual(created["issue"]["primary_owner"], "黄林")
        self.assertEqual(created["issue"]["images"], [])

        status, published = self.call(
            f"/api/workshop/issues/{issue_id}/publish", {}, token=self.admin,
        )
        self.assertEqual(status, 200)
        self.assertEqual(published["issue"]["status"], "published")

        status, exported = self.call(
            f"/api/workshop/issues/export?date={issue_date}", token=self.admin,
        )
        self.assertEqual(status, 200)
        workbook = load_workbook(BytesIO(exported))
        proofing_sheet = workbook["防错异常"]
        headers = [cell.value for cell in proofing_sheet[2]]
        self.assertNotIn("故障照片", headers)
        self.assertEqual(proofing_sheet["E3"].value, "PDA 扫描到相邻二维码")
        self.assertEqual(proofing_sheet["L3"].value, "无需上传照片")
        self.assertEqual(len(proofing_sheet._images), 0)

    def test_workshop_issue_export_supports_date_range_and_rejects_reverse_range(self):
        """现场问题导出支持默认日、日期区间，并明确拒绝结束早于开始的范围。"""

        today = web_server.business_today()
        dates = [today - timedelta(days=1), today]
        for index, issue_day in enumerate(dates, 1):
            issue_date = issue_day.isoformat()
            status, created = self.call("/api/workshop/issues", {
                "issue_date": issue_date,
                "category": "error_proofing",
                "cause": f"日期范围问题-{index}",
                "happened_at": f"{issue_date}T09:30",
                "batch_no": f"B-{index:03d}",
                "material_name": "测试物料",
                "material_code": f"M-{index:03d}",
                "cause_analysis": "测试原因",
                "corrective_action": "测试措施",
                "tracking_status": "待处理",
                "responsible_person": "测试负责人",
                "updated_by_name": "测试记录人",
            }, token=self.admin)
            self.assertEqual(status, 201)
            self.assertEqual(self.call(
                f"/api/workshop/issues/{created['issue']['id']}/publish", {}, token=self.admin,
            )[0], 200)

        start_date, end_date = (item.isoformat() for item in dates)
        status, exported = self.call(
            f"/api/workshop/issues/export?start_date={start_date}&end_date={end_date}",
            token=self.admin,
        )
        self.assertEqual(status, 200)
        workbook = load_workbook(BytesIO(exported))
        proofing_sheet = workbook["防错异常"]
        causes = {proofing_sheet.cell(row, 5).value for row in range(3, proofing_sheet.max_row + 1)}
        self.assertEqual(causes, {"日期范围问题-1", "日期范围问题-2"})
        self.assertEqual(workbook["问题一览表"].max_row, 2)

        status, reversed_range = self.call(
            f"/api/workshop/issues/export?start_date={end_date}&end_date={start_date}",
            token=self.admin,
        )
        self.assertEqual(status, 400)
        self.assertIn("开始日期不能晚于结束日期", reversed_range["error"])

    def test_workshop_published_issue_edit_resolve_and_reopen(self):
        """已发布问题应支持乐观并发编辑、填写闭环说明、重新打开和权限校验。"""

        issue_date = datetime.now().strftime("%Y-%m-%d")
        status, created = self.call("/api/workshop/issues", {
            "issue_date": issue_date,
            "category": "error_proofing",
            "cause": "原始问题描述",
            "happened_at": f"{issue_date}T09:30",
            "batch_no": "B-001",
            "material_name": "测试物料",
            "material_code": "M-001",
            "cause_analysis": "原始原因",
            "corrective_action": "原始措施",
            "tracking_status": "待处理",
            "handling_time": "",
            "responsible_person": "王工",
            "updated_by_name": "李工",
            "notes": "原始备注",
        }, token=self.admin)
        self.assertEqual(status, 201)
        issue_id = created["issue"]["id"]
        self.assertEqual(self.call(f"/api/workshop/issues/{issue_id}/publish", {}, token=self.admin)[0], 200)

        listed = self.call(f"/api/workshop/issues?date={issue_date}", token=self.admin)[1]
        current = listed["issues"][0]
        status, edited = self.call(f"/api/workshop/issues/{issue_id}", {
            "expected_updated_at": current["updated_at"],
            "issue_date": issue_date,
            "category": "error_proofing",
            "cause": "修正后的问题描述",
            "happened_at": f"{issue_date}T09:30",
            "batch_no": "B-001",
            "material_name": "测试物料",
            "material_code": "M-001",
            "cause_analysis": "修正后的原因",
            "corrective_action": "修正后的措施",
            "tracking_status": "处理中",
            "handling_time": "",
            "responsible_person": "王工",
            "updated_by_name": "李工",
            "notes": "修正后的备注",
        }, token=self.admin, method="PATCH")
        self.assertEqual(status, 200)
        self.assertEqual(edited["issue"]["cause"], "修正后的问题描述")
        self.assertEqual(edited["issue"]["resolution_status"], "open")

        status, missing_note = self.call(
            f"/api/workshop/issues/{issue_id}/resolve", {}, token=self.admin,
        )
        self.assertEqual(status, 400)
        self.assertIn("解决情况", missing_note["error"])
        status, resolved = self.call(
            f"/api/workshop/issues/{issue_id}/resolve",
            {"expected_updated_at": edited["issue"]["updated_at"], "resolution_note": "已补齐物料并完成现场复核"},
            token=self.admin,
        )
        self.assertEqual(status, 200)
        self.assertEqual(resolved["issue"]["resolution_status"], "resolved")
        self.assertEqual(resolved["issue"]["resolution_note"], "已补齐物料并完成现场复核")
        summary = self.call(f"/api/workshop/issues?date={issue_date}", token=self.admin)[1]["summary"]
        self.assertEqual(summary["open_count"], 0)
        self.assertEqual(summary["resolved_count"], 1)

        status, reopened = self.call(
            f"/api/workshop/issues/{issue_id}/reopen",
            {"expected_updated_at": resolved["issue"]["updated_at"]}, token=self.admin,
        )
        self.assertEqual(status, 200)
        self.assertEqual(reopened["issue"]["resolution_status"], "open")
        self.assertEqual(reopened["issue"]["resolution_note"], "已补齐物料并完成现场复核")

    def test_slow_workshop_upload_does_not_block_other_uploads(self):
        """慢速图片上传不能持有全局请求锁并阻塞其他用户的并发上传。"""

        issue_date = datetime.now().strftime("%Y-%m-%d")
        issue_ids = []
        for cause in ("慢连接问题", "正常连接问题"):
            status, created = self.call("/api/workshop/issues", {
                "issue_date": issue_date, "cause": cause, "primary_owner": "测试负责人",
            }, token=self.admin)
            self.assertEqual(status, 201)
            issue_ids.append(created["issue"]["id"])

        stream = BytesIO()
        Image.new("RGB", (96, 72), "#3d7df0").save(stream, format="JPEG")
        image_content = stream.getvalue()
        slow_query = urllib.parse.urlencode({"name": "慢连接.jpg"})
        slow_socket = socket.create_connection(("127.0.0.1", self.server.server_port), timeout=5)
        slow_socket.settimeout(5)
        headers = (
            f"POST /api/workshop/issues/{issue_ids[0]}/images?{slow_query} HTTP/1.0\r\n"
            f"Host: 127.0.0.1\r\nX-Session-Token: {self.admin}\r\n"
            f"Content-Type: application/octet-stream\r\nContent-Length: {len(image_content)}\r\n\r\n"
        ).encode("ascii")
        split_at = max(1, len(image_content) // 3)
        try:
            slow_socket.sendall(headers + image_content[:split_at])
            time.sleep(0.1)
            fast_query = urllib.parse.urlencode({"name": "正常连接.jpg"})
            started = time.monotonic()
            status, uploaded = self.call(
                f"/api/workshop/issues/{issue_ids[1]}/images?{fast_query}",
                token=self.admin, raw=image_content,
                headers={"Content-Length": str(len(image_content))},
            )
            elapsed = time.monotonic() - started
            self.assertEqual(status, 201)
            self.assertEqual(len(uploaded["issue"]["images"]), 1)
            self.assertLess(elapsed, 2)
            slow_socket.sendall(image_content[split_at:])
            while slow_socket.recv(4096):
                pass
        finally:
            slow_socket.close()

        self.assertEqual(self.call(
            f"/api/workshop/issues/{issue_ids[1]}/publish", {}, token=self.admin,
        )[0], 200)

    def test_workshop_stale_draft_cleanup_removes_isolated_images(self):
        """定期维护应删除过期草稿及孤立图片，同时保留仍被有效问题引用的文件。"""

        issue_date = datetime.now().strftime("%Y-%m-%d")
        status, created = self.call("/api/workshop/issues", {
            "issue_date": issue_date, "cause": "未完成草稿", "primary_owner": "测试负责人",
        }, token=self.admin)
        self.assertEqual(status, 201)
        issue_id = created["issue"]["id"]
        stream = BytesIO()
        Image.new("RGB", (24, 18), "white").save(stream, format="JPEG")
        image_content = stream.getvalue()
        query = urllib.parse.urlencode({"name": "草稿.jpg"})
        self.assertEqual(self.call(
            f"/api/workshop/issues/{issue_id}/images?{query}", token=self.admin,
            raw=image_content, headers={"Content-Length": str(len(image_content))},
        )[0], 201)
        folder = web_server.workshop_issue_dir(1, issue_id)
        self.assertTrue(folder.is_dir())
        now = datetime.now(timezone.utc).replace(microsecond=0)
        old_time = (now - timedelta(hours=25)).isoformat(timespec="seconds")
        with web_server.DB_LOCK, web_server.db() as connection:
            connection.execute(
                "UPDATE workshop_issues SET updated_at = ? WHERE id = ?", (old_time, issue_id)
            )
        self.assertEqual(web_server.cleanup_stale_workshop_drafts(current_time=now), 1)
        self.assertFalse(folder.exists())
        with web_server.DB_LOCK, web_server.db() as connection:
            self.assertIsNone(connection.execute(
                "SELECT id FROM workshop_issues WHERE id = ?", (issue_id,)
            ).fetchone())


if __name__ == "__main__":
    unittest.main()
