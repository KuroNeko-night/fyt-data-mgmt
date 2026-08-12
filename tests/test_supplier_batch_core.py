# -*- coding: utf-8 -*-
"""供应商批次表扫描、复核筛选与“原厂”排除回归。"""
from __future__ import annotations

import os
import tempfile
import unittest

import openpyxl

from core import supplier_batch_core


class SupplierBatchCoreTests(unittest.TestCase):
    """验证供应商扫描、原厂排除、人工选择和逐批交付日期。"""

    def setUp(self):
        """创建包含多个供应商、原厂行和批次的合成来源。"""

        self.temp = tempfile.TemporaryDirectory(prefix="fyt_supplier_batch_")
        self.old_catalog = os.environ.get("FYT_CATALOG_PATH")
        os.environ["FYT_CATALOG_PATH"] = os.path.join(self.temp.name, "catalog.json")
        self.batch_a = self.path("GMIDR26178A辅料清单总表.xlsx")
        self.batch_b = self.path("GMIDR26178B辅料清单总表.xlsx")
        self.history = self.path("供应商丙8月采购清单明细.xlsx")
        self._write_batch(self.batch_a, [
            ["JBC001", "纸箱（原厂）", "500x300", "个", 5, "供应商甲"],
            ["JBC002", "铁箱", "1000x800", "套", 3, "供应商甲"],
            ["JBC003", "珍珠棉", "100x50", "张", 2, "#N/A"],
            ["JBC004", "托盘", "1200x1000", "个", 0, "供应商甲"],
        ])
        self._write_batch(self.batch_b, [
            ["JBC002", "铁箱", "1000x800", "套", 4, "供应商乙"],
            ["JBC005", "纸箱(原厂)", "600x400", "个", 2, "供应商乙"],
        ])
        self._write_batch(self.history, [
            ["JBC003", "珍珠棉", "100x50", "张", 9, ""],
        ])

    def tearDown(self):
        """恢复主数据库环境并删除批次表输出。"""

        if self.old_catalog is None:
            os.environ.pop("FYT_CATALOG_PATH", None)
        else:
            os.environ["FYT_CATALOG_PATH"] = self.old_catalog
        self.temp.cleanup()

    def path(self, name):
        """返回当前用例临时目录下路径。"""

        return os.path.join(self.temp.name, name)

    @staticmethod
    def _write_batch(path, rows):
        """写入符合供应商批次分析列结构的工作簿。"""

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "汇总"
        worksheet.append([None] * 6)
        worksheet.append([None] * 6)
        worksheet.append(["材料编号", "材料名称", "规格", "单位", "求和项:最终采购数量", "供应商"])
        for row in rows:
            worksheet.append(row)
        workbook.save(path)
        workbook.close()

    def test_analyze_scans_suppliers_and_excludes_original_items(self):
        """分析应汇总有效供应商，排除原厂与零数量行，并用历史资料补空供应商。"""

        plan = supplier_batch_core.analyze(
            [self.batch_a, self.batch_b], [self.history],
        )
        counts = {item["name"]: item["rows"] for item in plan["suppliers"]}
        self.assertEqual(counts, {"供应商乙": 2, "供应商丙": 1})
        self.assertEqual(plan["excluded_original_count"], 2)
        self.assertEqual(plan["unmatched_count"], 0)
        self.assertGreaterEqual(plan["supplier_conflicts"], 1)
        self.assertEqual([item["batch"] for item in plan["batches"]], ["GMIDR26178A", "GMIDR26178B"])

    def test_run_generates_only_selected_supplier_without_original_items(self):
        """执行阶段只输出人工勾选供应商，成品批次表中不得出现原厂物料。"""

        out_dir = self.path("输出")
        result = supplier_batch_core.run(
            [self.batch_a, self.batch_b],
            [self.history],
            selected_suppliers=["供应商丙"],
            batch_dates={"GMIDR26178A": "8.7", "GMIDR26178B": "8.10"},
            out_dir=out_dir,
        )
        self.assertEqual(result["generated"], 1)
        self.assertEqual(result["rows"], 1)
        self.assertEqual(result["excluded_original_count"], 2)
        self.assertEqual(result["batch_dates"], {"GMIDR26178A": "8.7", "GMIDR26178B": "8.10"})
        self.assertEqual(len(result["files"]), 1)
        self.assertIn("供应商丙", os.path.basename(result["files"][0]))

        workbook = openpyxl.load_workbook(result["files"][0], data_only=True)
        try:
            values = [
                cell.value
                for worksheet in workbook.worksheets
                for row in worksheet.iter_rows()
                for cell in row
                if cell.value is not None
            ]
        finally:
            workbook.close()
        self.assertIn("JBC003", values)
        self.assertIn("GMIDR26178A交付日期8.7", values)
        self.assertFalse(any("原厂" in str(value) for value in values))

    def test_run_rejects_empty_or_unknown_supplier_selection(self):
        with self.assertRaisesRegex(ValueError, "至少选择一个"):
            supplier_batch_core.run(
                [self.batch_a], [self.history], selected_suppliers=[], out_dir=self.path("空选择"),
            )
        with self.assertRaisesRegex(ValueError, "不在本次扫描结果"):
            supplier_batch_core.run(
                [self.batch_a], [self.history], selected_suppliers=["不存在供应商"], out_dir=self.path("错误选择"),
            )

    def test_run_requires_complete_known_batch_dates(self):
        """每个已识别批次都必须填写交付日期，未知批次键也应被拒绝。"""

        common = {
            "batch_paths": [self.batch_a, self.batch_b],
            "history_paths": [self.history],
            "selected_suppliers": ["供应商乙"],
        }
        with self.assertRaisesRegex(ValueError, "GMIDR26178B"):
            supplier_batch_core.run(
                **common,
                batch_dates={"GMIDR26178A": "8.7"},
                out_dir=self.path("缺少日期"),
            )
        with self.assertRaisesRegex(ValueError, "GMIDR26178A"):
            supplier_batch_core.run(
                **common,
                batch_dates={"GMIDR26178A": "", "GMIDR26178B": "8.10"},
                out_dir=self.path("空日期"),
            )
        with self.assertRaisesRegex(ValueError, "未知批次"):
            supplier_batch_core.run(
                **common,
                batch_dates={
                    "GMIDR26178A": "8.7",
                    "GMIDR26178B": "8.10",
                    "GMIDR99999": "8.20",
                },
                out_dir=self.path("未知批次"),
            )


if __name__ == "__main__":
    unittest.main()
