# -*- coding: utf-8 -*-
"""报表中心任务聚合与 Excel 输出回归测试。"""

from __future__ import annotations

import os
import tempfile
import unittest

import openpyxl

from core import report_center_core


class ReportCenterCoreTests(unittest.TestCase):
    """保护状态统计、中文标题和客户字段裁剪规则。"""

    def setUp(self):
        """为每个测试创建隔离输出目录。"""

        self.temp = tempfile.TemporaryDirectory(prefix="fyt_report_center_")
        self.output = os.path.join(self.temp.name, "业务报表.xlsx")  # 输出到临时目录

    def tearDown(self):
        """删除合成报表。"""

        self.temp.cleanup()

    def test_build_report_groups_statuses_and_hides_internal_fields(self):
        """汇总应区分完成与失败，明细不得输出任务 ID 和绝对路径。"""

        items = [
            {
                "feature": "delivery",
                "title": "送货计划",
                "status": "completed",
                "started_at": "2026-08-13T08:00:00",
                "files": 2,
                "task_id": "internal-task",
                "output_dir": r"C:\private\output",
            },
            {
                "feature": "delivery",
                "title": "送货计划重试",
                "status": "failed",
                "started_at": "2026-08-13T09:00:00",
                "files": 0,
            },
            {
                "feature": "attendance",
                "title": "考勤填报",
                "status": "running",
                "started_at": "2026-08-13T10:00:00",
                "files": 0,
            },
        ]

        rows = report_center_core.build_report(items, self.output, "本日")

        self.assertEqual(rows, 3)  # 三条任务都进入报表
        workbook = openpyxl.load_workbook(self.output, data_only=True)
        try:
            self.assertEqual(workbook.sheetnames, ["汇总", "明细"])  # 固定两个工作表
            summary = workbook["汇总"]
            self.assertEqual(summary["B5"].value, 3)  # 任务总数
            self.assertEqual(summary["B6"].value, 1)  # 完成数
            self.assertEqual(summary["B7"].value, 1)  # 失败数
            detail_values = [cell.value for row in workbook["明细"].iter_rows() for cell in row]
            self.assertIn("送货计划", detail_values)  # 客户可见标题保留
            self.assertNotIn("internal-task", detail_values)  # 内部任务 ID 不输出
            self.assertNotIn(r"C:\private\output", detail_values)  # 绝对路径不输出
        finally:
            workbook.close()


if __name__ == "__main__":
    unittest.main()
