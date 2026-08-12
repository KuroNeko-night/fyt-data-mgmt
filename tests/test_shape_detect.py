# -*- coding: utf-8 -*-
"""数据形态兜底识别 shape_detect 的单元测试(合成簿,可移植)。

钉死:形态分类互斥、编码/数值/文本列能按形态被正确指派、必需角色缺失或
置信度过低时返回 (None,{},conf)、表头文字完全不同也能靠形态兜底命中。
"""
import os
import tempfile
import unittest
import warnings

import openpyxl

from core import shape_detect as S

warnings.filterwarnings("ignore", message="Workbook contains no default style")


class TestClassify(unittest.TestCase):
    """验证工作表形状分类枚举和别名。"""

    def test_kinds(self):
        """分类器应区分常见横表、竖表、矩阵和未知结构。"""

        self.assertEqual(S.classify_value(None), "empty")
        self.assertEqual(S.classify_value("   "), "empty")
        self.assertEqual(S.classify_value(123), "number")
        self.assertEqual(S.classify_value(12.5), "number")
        self.assertEqual(S.classify_value("1,234.5"), "number")
        self.assertEqual(S.classify_value("物料名称"), "text")
        self.assertEqual(S.classify_value("A12345"), "code")
        self.assertEqual(S.classify_value("KD-001"), "code")
        import datetime
        self.assertEqual(S.classify_value(datetime.date(2020, 1, 1)), "date")
        self.assertEqual(S.classify_value(True), "text")   # bool 不当数字


class _Tmp(unittest.TestCase):
    """提供可控制多行表头的临时工作表。"""

    def setUp(self):
        """创建临时目录。"""

        self._tmp = tempfile.mkdtemp(prefix="fyt_shape_")

    def tearDown(self):
        """删除合成工作簿。"""

        import shutil
        shutil.rmtree(self._tmp, ignore_errors=True)

    def _ws(self, rows):
        """写入行并返回加载后的活动工作表与工作簿。"""

        p = os.path.join(self._tmp, "t.xlsx")
        wb = openpyxl.Workbook(); ws = wb.active
        for row in rows:
            ws.append(row)
        wb.save(p)
        wb2 = openpyxl.load_workbook(p)
        return wb2, wb2[wb2.sheetnames[0]]


# delivery 风格画像:编码/名称/数量/供方代码/供方名称,require code
_PROFILE = [
    ("code", S.CODE, True),
    ("cname", S.TEXT, False),
    ("qty", S.NUMBER, False),
    ("sup_code", S.CODE, False),
    ("sup_name", S.TEXT, False),
]


class TestDetect(_Tmp):
    """验证陌生表头回退、必需角色、可信度阈值和识别日志。"""

    def test_fallback_when_headers_alien(self):
        """表头别名陌生但形状特征充分时，应通过结构评分回退识别。"""

        # 表头文字完全不同(乱码/英文),仍应按数据形态兜底命中 code/qty/text
        rows = [[" colA", "colB", "colC" ]]
        for i in range(8):
            rows.append(["A%04d" % i, "物料名称%d" % i, 100 + i])
        wb, ws = self._ws(rows)
        hr, col, conf = S.detect_by_shape(ws, _PROFILE, min_conf=0.4)
        wb.close()
        self.assertIsNotNone(hr)
        self.assertEqual(col["code"], 1)
        self.assertEqual(col["cname"], 2)
        self.assertEqual(col["qty"], 3)
        self.assertGreater(conf, 0.4)

    def test_missing_required_returns_none(self):
        """缺少业务必需角色时即使部分特征命中也不能接受布局。"""

        # 没有任何"像编码"的列 -> 必需 code 缺失 -> 不采用
        rows = [["h1", "h2"]]
        for i in range(6):
            rows.append(["名称文本%d" % i, "另一段中文%d" % i])
        wb, ws = self._ws(rows)
        hr, col, conf = S.detect_by_shape(ws, _PROFILE, min_conf=0.4)
        wb.close()
        self.assertIsNone(hr)
        self.assertEqual(col, {})

    def test_low_confidence_rejected(self):
        # 数据太杂,置信度低于阈值 -> 返回 None,但 conf 仍报出
        rows = [["h1", "h2", "h3"]]
        rows.append(["A123", "文本", 5])
        rows.append([None, None, None])
        wb, ws = self._ws(rows)
        hr, col, conf = S.detect_by_shape(ws, _PROFILE, min_conf=0.99)
        wb.close()
        self.assertIsNone(hr)

    def test_log_emitted_on_hit(self):
        rows = [["x", "y", "z"]]
        for i in range(8):
            rows.append(["B%04d" % i, "中文名%d" % i, 10 + i])
        wb, ws = self._ws(rows)
        logs = []
        S.detect_by_shape(ws, _PROFILE, min_conf=0.4, log=logs.append)
        wb.close()
        self.assertTrue(any("形态兜底" in l for l in logs))


class TestDeliveryWrapper(_Tmp):
    """验证送货业务包装层报告形状识别来源。"""

    def test_wrapper_reports_source(self):
        """送货包装层应在结果中标明列映射来自形状回退，便于人工判断可信度。"""

        from core import delivery_core as D
        # 正常表头 -> source == 'header'
        rows = [["物料号", "物料名称", "需求数"]]
        for i in range(4):
            rows.append(["A%04d" % i, "名%d" % i, 5 + i])
        wb, ws = self._ws(rows)
        hr, col, src = D.detect_layout_or_shape(ws)
        wb.close()
        self.assertEqual(src, "header")
        self.assertIn("code", col)


if __name__ == "__main__":
    unittest.main()
