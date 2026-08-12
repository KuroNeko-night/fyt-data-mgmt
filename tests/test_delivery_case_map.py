# -*- coding: utf-8 -*-
"""送货计划参考 CASE、托数与班组映射回归测试。"""
import os
import tempfile
import unittest

import openpyxl

from core import delivery_core


class DeliveryCaseMapTests(unittest.TestCase):
    """验证参考明细页选择、重复编码优先级和空记录处理。"""

    def setUp(self):
        """创建隔离的临时工作簿目录。"""
        self.temp = tempfile.TemporaryDirectory(prefix="fyt_delivery_case_")

    def tearDown(self):
        """删除临时参考计划。"""
        self.temp.cleanup()

    def path(self, name):
        """返回临时目录内的文件路径。"""
        return os.path.join(self.temp.name, name)

    def test_selects_detail_sheet_and_keeps_first_valid_record(self):
        """应跳过透视页，重复编码保留首条，空记录不阻止后续有效记录。"""
        path = self.path("参考计划.xlsx")
        workbook = openpyxl.Workbook()
        summary = workbook.active
        summary.title = "透视汇总"
        summary.append(["班组", "计数项:物料编码"])
        detail = workbook.create_sheet("零件到货计划")
        detail.append(["往期送货计划"])
        detail.append(["物料编码", "CASE", "CASE托数", "班组"])
        detail.append(["JBC001", "CASE-A", 2, "一组"])
        detail.append(["JBC001", "CASE-B", 3, "二组"])
        detail.append(["JBC002", "", None, ""])
        detail.append(["JBC002", "CASE-C", None, "三组"])
        detail.append([100.0, "CASE-D", 1, "四组"])
        workbook.save(path)
        workbook.close()

        logs = []
        mapping = delivery_core.build_case_map(path, log=logs.append)
        self.assertEqual(mapping["JBC001"], ("CASE-A", 2, "一组"))
        self.assertEqual(mapping["JBC002"], ("CASE-C", None, "三组"))
        self.assertEqual(mapping["100"], ("CASE-D", 1, "四组"))
        self.assertEqual(len(mapping), 3)
        self.assertIn("零件到货计划", logs[-1])

    def test_missing_detail_sheet_returns_empty_mapping(self):
        """只有透视统计页时应返回空映射并给出可读提示。"""
        path = self.path("无明细.xlsx")
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.append(["班组", "计数项:物料编码"])
        workbook.save(path)
        workbook.close()

        logs = []
        self.assertEqual(delivery_core.build_case_map(path, log=logs.append), {})
        self.assertIn("未找到", logs[-1])


if __name__ == "__main__":
    unittest.main()
