# -*- coding: utf-8 -*-
"""采购汇总静态输出与横向子表直读回归测试。"""

import os
import tempfile
import unittest
import zipfile

import openpyxl

from core import pivot_core


class PivotStaticSummaryTests(unittest.TestCase):
    """验证新流程只输出清洗子表和聚合主表。"""

    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.source = os.path.join(self.temp.name, "采购输入.xlsx")
        self.output = os.path.join(self.temp.name, "采购汇总.xlsx")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "包装方案汇总"
        # 左右两个横向子表，中间留一列空白；普通“数量”列存在但不进入六字段业务记录。
        headers = [
            "版本序号", "材料编号", "材料名称", "规格", "数量", "单位", "最终采购数量", "",
            "版本序号", "材料编号", "材料名称", "规格", "数量", "单位", "最终采购数量",
        ]
        sheet.append(headers)
        sheet.append([1, "M001", "螺栓", "M8", 1, "个", 2, "", 1, "M001", "螺栓", "M8", 1, "个", 3])
        sheet.append([1, "M002", "垫片", "M8", 1, "个", 4, "", 1, "M003", "螺母", "M8", 1, "个", 5])
        workbook.save(self.source)
        workbook.close()

    def tearDown(self):
        self.temp.cleanup()

    def test_side_by_side_blocks_are_read_directly_and_aggregated(self):
        """两块输入不经中间拼表，M001 应直接聚合为 5。"""
        plan = pivot_core.analyze_workbooks([self.source])
        self.assertEqual(len(plan["sheets"][0]["kept"]), 4)

        result = pivot_core.run(self.source, out_dir=self.temp.name)
        self.assertEqual(result["groups"], 3)
        self.assertEqual(result["total"], 14)
        self.assertTrue(result["source_check"]["passed"])
        self.assertEqual(result["source_check"]["status"], "通过")
        self.assertEqual(result["source_check"]["source_total"], 14)
        self.assertEqual(result["source_check"]["output_total"], 14)
        self.assertEqual(result["source_check"]["difference"], 0)

        workbook = openpyxl.load_workbook(result["out"], data_only=False)
        try:
            self.assertEqual(workbook.sheetnames, ["采购汇总", "清洗数据"])
            summary = workbook["采购汇总"]
            clean = workbook["清洗数据"]
            self.assertEqual(
                [summary.cell(1, column).value for column in range(1, 10)],
                ["材料编号", "材料名称", "规格", "单位", "最终采购数量", "供应商", "差异", "实收", "日期"],
            )
            self.assertNotIn("汇总", [summary.cell(1, column).value for column in range(1, 10)])
            self.assertNotIn("求和项:最终采购数量", [summary.cell(1, column).value for column in range(1, 10)])
            self.assertEqual(summary.cell(2, 1).value, "M001")
            self.assertEqual(summary.cell(2, 5).value, 5)
            self.assertEqual(summary.cell(2, 7).value, '=IF(H2="","",H2-E2)')
            self.assertEqual(summary.cell(7, 1).value, "来源最终采购数")
            self.assertEqual(summary.cell(7, 2).value, 14)
            self.assertEqual(summary.cell(7, 4).value, "最终表采购数")
            self.assertEqual(summary.cell(7, 5).value, 14)
            self.assertEqual(summary.cell(7, 8).value, "通过")
            self.assertEqual(
                [clean.cell(1, column).value for column in range(1, 6)],
                ["材料编号", "材料名称", "规格", "单位", "最终采购数量"],
            )
            self.assertEqual(clean.max_column, 5)
        finally:
            workbook.close()

        # 普通工作簿不应再写入 pivotCache、pivotTable 等 OOXML 部件。
        with zipfile.ZipFile(result["out"]) as archive:
            self.assertFalse(any("pivot" in name.lower() for name in archive.namelist()))

    def test_source_total_mismatch_is_reported_as_a_serious_self_check_issue(self):
        """源记录有数量但因缺少物料号未进入主表时，自检必须明确失败。"""
        workbook = openpyxl.load_workbook(self.source)
        try:
            sheet = workbook["包装方案汇总"]
            sheet.append([1, None, "缺少编码的物料", "M8", 1, "个", 6])
            workbook.save(self.source)
        finally:
            workbook.close()

        result = pivot_core.run(self.source, out_dir=self.temp.name)
        check = result["source_check"]
        self.assertFalse(check["passed"])
        self.assertEqual(check["status"], "异常")
        self.assertEqual(check["source_total"], 20)
        self.assertEqual(check["output_total"], 14)
        self.assertEqual(check["difference"], -6)
        self.assertTrue(any(
            level == "严重" and "最终采购数汇总自检异常" in message
            for level, message in result["issues"]
        ))

    def test_unparseable_source_quantity_cannot_pass_the_self_check(self):
        """非空数量无法解析时不得因为两侧数值碰巧相同而误报通过。"""
        check = pivot_core._quantity_check(14, 14, source_unparsed=1)
        self.assertFalse(check["passed"])
        self.assertEqual(check["status"], "无法确认")


if __name__ == "__main__":
    unittest.main()
