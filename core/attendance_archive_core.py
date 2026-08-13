# -*- coding: utf-8 -*-
"""
考勤月度归档核心
================
读取一份或多份已经填写完成的考勤表，自动识别姓名、日期、工时、加班和异常列，
按姓名聚合出勤天数、总工时、加班、异常次数和日均工时，并生成“月度汇总 + 每日
明细”两页报告。双端只负责选择文件和展示结果，归档口径集中在本模块维护。

日期支持 Excel datetime、带年份的年月日文本以及不带年份的月日文本；无年份时使用
服务器当前年份。归档月份取全部有效记录中出现次数最多的年月，因此混入少量跨月行
时仍按主体月份命名，但每日明细不会丢弃其他月份。出勤天数按姓名下不同日期去重，
工时、加班和异常则按每条输入记录累计；调用方应避免重复上传同一考勤文件。

本功能读取公式缓存值，执行前会提示未刷新的公式。它不重新计算班次或加班，输入表
中的工时结果是唯一依据；如需纠正打卡，应先通过考勤填表功能生成正确结果后再归档。
"""
from __future__ import annotations

import os
import re
from collections import defaultdict
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from . import common_core, paths, settings

EXCEL_SUFFIXES = {".xlsx", ".xlsm"}

_SUMMARY_HEADERS = ["姓名", "出勤天数", "总工时（小时）", "加班（小时）", "异常次数", "日均工时"]
_SUMMARY_WIDTHS = {"A": 14, "B": 10, "C": 15, "D": 12, "E": 10, "F": 10}
_DETAIL_HEADERS = ["姓名", "日期", "工时（小时）", "加班（小时）", "异常"]
_DETAIL_WIDTHS = {"A": 14, "B": 12, "C": 12, "D": 12, "E": 22}

_ROLE_ALIASES = {
    # 别名从业务中最常见的短名称到具体名称排列；每个角色只接受首个命中列。
    "name": ("姓名", "员工姓名", "姓名/工号"),
    "date": ("日期", "考勤日期", "出勤日期"),
    "hours": ("工时", "实际工时", "工作时间", "算出工时"),
    "ot": ("加班", "加班工时", "加班时长"),
    "abnormal": ("异常", "异常原因", "备注"),
}


def _text(value) -> str:
    """将单元格值转为去首尾空白的文本，空值统一为空串。"""
    return "" if value is None else str(value).strip()


def _number(value) -> float | None:
    """解析普通数值和带千分位文本，无法作为工时时返回 ``None``。

    布尔值不参与数值转换，避免 ``True`` 被统计为一小时。调用方把 ``None`` 按零
    处理，以保留日期和人员记录，同时不凭空增加工时。
    """
    if isinstance(value, bool) or value is None:
        return None
    try:
        return float(str(value).replace(",", ""))
    except (TypeError, ValueError):
        return None


def _normalize_date(value) -> str:
    """把支持的 Excel 日期或中文日期文本归一化为 ``YYYY-MM-DD``。

    完整日期可使用横线、斜线或“年月”分隔；只有月日时采用当前年份。正则只提取
    日期主体，允许单元格附带星期等说明；最终仍经 ``datetime`` 校验，非法日期返回
    空串并由读取层跳过。
    """
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    text = _text(value)
    if not text:
        return ""
    for pattern in (
        r"(\d{4})[-/年](\d{1,2})[-/月](\d{1,2})",
        r"(\d{1,2})[-/月](\d{1,2})[-/日]",
    ):
        match = re.search(pattern, text)
        if match:
            groups = match.groups()
            if len(groups) == 3:
                # 完整日期直接使用源年份，避免归档跨年数据时被当前时间影响。
                year, month, day = int(groups[0]), int(groups[1]), int(groups[2])
            else:
                month, day = int(groups[0]), int(groups[1])
                # 缺少年份的历史模板没有更多依据，只能按运行时当前年份解释。
                year = datetime.now().year
            try:
                return datetime(year, month, day).strftime("%Y-%m-%d")
            except ValueError:
                return ""
    return ""


def _detect_layout(worksheet) -> tuple[int, dict[str, int]] | None:
    """在页签前十二行中选择字段最完整的考勤表头。

    姓名、日期和工时是必要角色，加班与异常可选。同一页签可能在标题区出现部分字段，
    因此不是找到第一行就返回，而是比较候选角色数量，选择信息最完整的一行。返回
    openpyxl 的 1 基行列号；没有完整必要字段时返回 ``None``。
    """
    best = None
    scan_rows = min(12, worksheet.max_row or 12)
    for row_index in range(1, scan_rows + 1):
        columns: dict[str, int] = {}
        for cell in worksheet[row_index]:
            text = _text(cell.value).replace("\n", "")
            if not text:
                continue
            for role, aliases in _ROLE_ALIASES.items():
                if role in columns:
                    # 一个业务角色只使用该行中的首个命中列，保持模板行为可预测。
                    continue
                for alias in aliases:
                    if alias in text:
                        columns[role] = cell.column
                        break
        if {"name", "date", "hours"}.issubset(columns):
            score = len(columns)
            if best is None or score > best[1]:
                # 分数相同保留更靠上的候选，符合常规模板的首个正式表头习惯。
                best = (row_index, score, columns)
    if best is None:
        return None
    return best[0], best[2]


def _attendance_cell(values, columns: dict[str, int], role: str):
    """按考勤角色读取单行值；短行或可选列缺失时统一返回空值。"""

    column = columns.get(role)
    return values[column - 1] if column and column <= len(values) else None


def _parse_attendance_row(values, columns: dict[str, int]) -> dict[str, object] | None:
    """把一行表格值转换成标准考勤记录，无法确认人员或日期时返回 ``None``。"""

    # 空行、汇总行不能进入逐人统计；汇总行即使带日期也不代表实际出勤。
    if all(value is None or _text(value) == "" for value in values):
        return None
    name = _text(_attendance_cell(values, columns, "name"))
    if not name or name in ("合计", "总计", "小计"):
        return None
    date = _normalize_date(_attendance_cell(values, columns, "date"))
    if not date:
        # 日期是出勤天数和主体月份的依据，非法日期不能用“今天”兜底，否则会污染月度归档。
        return None
    hours = _number(_attendance_cell(values, columns, "hours"))
    overtime = _number(_attendance_cell(values, columns, "ot"))
    return {
        "name": name,
        "date": date,
        "hours": hours or 0.0,
        "ot": overtime or 0.0,
        "abnormal": _text(_attendance_cell(values, columns, "abnormal")),
    }


def _read_attendance(path: str, log=None) -> list[dict[str, object]]:
    """读取一本考勤表内所有可识别页签的有效每日记录。

    封面和说明页因识别不到必要字段而跳过。空行、合计行和无效日期不进入结果；缺少
    加班或异常列时使用零和空串，缺失/非法工时也按零保留。整本工作簿没有任何有效
    记录时抛出带文件名的业务错误。
    """
    common_core.warn_if_uncached(path, log, what="工时")
    # 只按顺序读取公式计算值，不修改样式；read_only 适合月度批量归档的大数据量。
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        rows: list[dict[str, object]] = []
        for worksheet in workbook.worksheets:
            layout = _detect_layout(worksheet)
            if layout is None:
                continue
            header_row, columns = layout
            max_column = max(columns.values())
            # 只读取到最后一个必要/可选角色列，避免异常使用范围带来的空列开销。
            for values in worksheet.iter_rows(
                min_row=header_row + 1, max_col=max_column, values_only=True):
                record = _parse_attendance_row(values, columns)
                if record is not None:
                    rows.append(record)
        if not rows:
            raise ValueError("未在 %s 中识别到 姓名/日期/工时 列" % os.path.basename(path))
        return rows
    finally:
        workbook.close()


def _validated_archive_files(files) -> list[str]:
    """规范化并一次性验证全部归档输入，避免读取一半后才发现坏路径。"""
    normalized = [
        os.path.abspath(str(value)) for value in (files or []) if str(value).strip()
    ]
    if not normalized:
        raise ValueError("请选择考勤填报表")
    for path in normalized:
        if not os.path.isfile(path):
            raise FileNotFoundError("找不到文件：%s" % path)
        if os.path.splitext(path)[1].lower() not in EXCEL_SUFFIXES:
            raise ValueError("仅支持 xlsx 或 xlsm 文件：%s" % os.path.basename(path))
    return normalized


def _merge_attendance_row(
    row: dict[str, object],
    per_person: dict[str, dict[str, object]],
    detail_rows: list[tuple[str, str, float, float, str]],
    months: dict[str, int],
) -> None:
    """把一条有效考勤记录合并到人员汇总、月份计数和原始明细。"""
    name = str(row["name"])
    date = str(row["date"])
    hours = float(row["hours"])
    overtime = float(row["ot"])
    abnormal = str(row["abnormal"])
    stats = per_person[name]
    stats["days"].add(date)  # 日期集合只负责出勤天数去重，同日多条工时仍全部累计。
    stats["hours"] += hours
    stats["ot"] += overtime
    if abnormal:
        stats["abnormal"] += 1  # 这里只统计非空异常次数，不重新解释文本严重程度。
    months[date[:7]] += 1  # 主体月份按有效明细行计数，少量跨月补卡不会改变主要月份。
    detail_rows.append((name, date, hours, overtime, abnormal))


def _aggregate_attendance(
    files: list[str], log=None, progress=None,
) -> tuple[dict[str, dict[str, object]], list[tuple[str, str, float, float, str]], str]:
    """读取全部考勤文件并返回人员汇总、明细行和主体月份。"""
    per_person: dict[str, dict[str, object]] = defaultdict(
        lambda: {"days": set(), "hours": 0.0, "ot": 0.0, "abnormal": 0},
    )
    detail_rows: list[tuple[str, str, float, float, str]] = []
    months: dict[str, int] = defaultdict(int)
    for index, path in enumerate(files, start=1):
        for row in _read_attendance(path, log=log):
            _merge_attendance_row(row, per_person, detail_rows, months)
        if log:
            log("已读取 %s" % os.path.basename(path))
        if progress:
            progress(10 + round(index / len(files) * 70))  # 读取阶段沿用历史 10%～80% 进度区间。
    if not per_person:
        raise ValueError("考勤表中没有有效记录")
    month_label = max(months, key=months.get) if months else datetime.now().strftime("%Y-%m")
    return per_person, detail_rows, month_label


def _archive_output_dir(files: list[str], out_dir) -> str:
    """按显式目录或全局输出策略解析归档目标目录。"""
    if out_dir is None:
        current = settings.get_settings()
        # 源旁输出以首份输入定位；固定目录等策略仍由 paths 单一事实源处理。
        return paths.resolve_output_dir(
            "attendance_archive", src_path=files[0], **current.output_kwargs(),
        )
    target = os.path.abspath(str(out_dir))
    os.makedirs(target, exist_ok=True)
    return target


def _report_styles() -> tuple[Border, PatternFill, Font, Font]:
    """创建两张报表页共同复用的边框、表头和正文字体。"""
    thin = Side(style="thin", color="9AA5B1")
    return (
        Border(left=thin, right=thin, top=thin, bottom=thin),
        PatternFill("solid", fgColor="EAF1FF"),
        Font(name="宋体", size=10, bold=True),
        Font(name="宋体", size=10),
    )


def _prepare_report_sheet(
    sheet, headers: list[str], widths: dict[str, int],
    border: Border, head_fill: PatternFill, head_font: Font,
) -> None:
    """设置报表页列宽和统一表头格式。"""
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for column, name in enumerate(headers, start=1):
        cell = sheet.cell(1, column, name)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border


def _write_report_row(sheet, row_index: int, values: list[object], font: Font, border: Border) -> None:
    """写入一行普通报表值并应用共享正文样式。"""
    for column, value in enumerate(values, start=1):
        cell = sheet.cell(row_index, column, value)
        cell.font = font
        cell.border = border


def _write_summary_sheet(
    sheet, per_person: dict[str, dict[str, object]],
    border: Border, head_fill: PatternFill, head_font: Font, cell_font: Font,
) -> None:
    """写入按人员汇总的出勤天数、工时和异常指标。"""
    sheet.title = "月度汇总"
    _prepare_report_sheet(
        sheet, _SUMMARY_HEADERS, _SUMMARY_WIDTHS, border, head_fill, head_font,
    )
    for row_index, name in enumerate(
        sorted(per_person, key=lambda item: item.encode("utf-8")), start=2,
    ):
        stats = per_person[name]
        days = len(stats["days"])
        # 日均工时以唯一出勤日为分母，同日多条明细不会把平均值错误压低。
        average = round(float(stats["hours"]) / days, 2) if days else 0
        _write_report_row(sheet, row_index, [
            name, days, round(float(stats["hours"]), 2), round(float(stats["ot"]), 2),
            int(stats["abnormal"]), average,
        ], cell_font, border)


def _write_detail_sheet(
    sheet, detail_rows: list[tuple[str, str, float, float, str]],
    border: Border, head_fill: PatternFill, head_font: Font, cell_font: Font,
) -> None:
    """写入排序稳定、保留全部源记录的每日明细页。"""
    sheet.title = "每日明细"
    _prepare_report_sheet(
        sheet, _DETAIL_HEADERS, _DETAIL_WIDTHS, border, head_fill, head_font,
    )
    detail_rows.sort(key=lambda item: (item[0].encode("utf-8"), item[1]))
    for row_index, (name, date, hours, overtime, abnormal) in enumerate(detail_rows, start=2):
        _write_report_row(
            sheet, row_index,
            [name, date, round(hours, 2), round(overtime, 2), abnormal],
            cell_font, border,
        )


def _save_archive_report(
    out_dir: str, month_label: str, per_person: dict[str, dict[str, object]],
    detail_rows: list[tuple[str, str, float, float, str]],
) -> str:
    """组装两页考勤工作簿并保存到唯一输出路径。"""
    border, head_fill, head_font, cell_font = _report_styles()
    workbook = openpyxl.Workbook()
    try:
        _write_summary_sheet(
            workbook.active, per_person, border, head_fill, head_font, cell_font,
        )
        _write_detail_sheet(
            workbook.create_sheet(), detail_rows, border, head_fill, head_font, cell_font,
        )
        target = common_core.unique_path(os.path.join(
            out_dir, "考勤月度汇总_%s.xlsx" % month_label,
        ))
        workbook.save(target)
        return target
    finally:
        workbook.close()  # 写盘失败也释放工作簿及其底层临时资源。


def archive(files, out_dir=None, log=None, progress=None) -> dict[str, object]:
    """汇总多份考勤表并生成月度统计工作簿。

    输入在读取前统一校验存在性和扩展名。每条记录都会进入每日明细并累计工时；每人
    出勤天数使用日期集合去重，因而同日多条记录只算一天但工时会相加。主体月份按
    有效记录数量最多的 ``YYYY-MM`` 决定。进度在校验后从 10% 开始，读取阶段占到
    80%；输出路径遵循全局设置。

    返回报告路径、主体月份、人员数和明细记录数，字段 ``days`` 为历史协议名称，
    实际表示每日明细行数而不是全体唯一自然日数量。
    """
    files = _validated_archive_files(files)
    if progress:
        progress(10)
    per_person, detail_rows, month_label = _aggregate_attendance(files, log, progress)
    out_dir = _archive_output_dir(files, out_dir)
    target = _save_archive_report(out_dir, month_label, per_person, detail_rows)
    if log:
        log("已生成月度汇总：%d 人、%d 条出勤记录（%s）。"
            % (len(per_person), len(detail_rows), month_label))
    return {
        "out_dir": out_dir, "path": target, "month": month_label,
        "persons": len(per_person), "days": len(detail_rows),
    }
