# -*- coding: utf-8 -*-
"""
安全检查日报解析核心
====================
读取管理员上传的安全检查 xlsx/xlsm，从表格内容提取报告日期、检查类别、检查项目、
标准、结果、问题描述、整改措施和责任人，并把工作表中的嵌入图片按锚点行关联到检查
记录。返回值可直接持久化到日清资料并由看板展示，无需前端重新分析 Excel。

表头必须同时包含七个规定字段；“序号”可选。合并单元格常只在类别首行保存值，因此
解析时会向下沿用最近的非空检查类别。结果文本按包含“不合格”优先、再包含“合格”
统计，其他文本计为待确认。图片使用 openpyxl 当前提供的内部集合和数据读取接口，
只支持常见 PNG/JPEG；无法提取时明确报错，避免数据库记录指向不存在的图片。

``analyze`` 可仅分析元数据，也可把图片写入指定目录；``run`` 负责解析统一输出目录。
本模块不修改上传文件，也不把安全检查混入日清事项类别。
"""

from __future__ import annotations

import datetime as dt
import os
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from . import paths, settings


SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm"}
_REQUIRED_HEADERS = {
    "检查类别", "检查项目", "安全标准要求", "检查结果", "问题描述", "整改措施", "责任人",
}
_DATE_RE = re.compile(r"^\d{4}[-/.]\d{1,2}[-/.]\d{1,2}$")


def _text(value: Any) -> str:
    """将单元格值转为去首尾空白的展示文本，空值统一为空串。"""
    return "" if value is None else str(value).strip()


def _date_text(value: Any, epoch: Any) -> str:
    """把支持的日期值归一化为 ISO 日期，无法确认时返回空串。

    支持 datetime/date、合理范围内的 Excel 日期序列号，以及 ``YYYY-MM-DD``、斜线或
    点号分隔文本。序列号必须结合工作簿 epoch 解释，兼容 1900/1904 日期制；数值范围
    限制可避免把普通检查数量误当日期。
    """
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, (int, float)) and not isinstance(value, bool) and 20000 < value < 80000:
        try:
            # from_excel 可能返回 date 或 datetime，统一只保留业务日期部分。
            converted = from_excel(value, epoch=epoch)
            return converted.date().isoformat() if isinstance(converted, dt.datetime) else converted.isoformat()
        except (TypeError, ValueError, OverflowError):
            return ""
    text = _text(value)
    if not _DATE_RE.match(text):
        return ""
    for separator in ("/", "."):
        text = text.replace(separator, "-")
    try:
        return dt.datetime.strptime(text, "%Y-%m-%d").date().isoformat()
    except ValueError:
        return ""


def _header_row(worksheet) -> tuple[int, dict[str, int]]:
    """在页签前十二行定位包含全部规定字段的表头。

    返回 1 基表头行号和“原始表头文本 -> 列号”映射。规定字段采用精确名称匹配，
    防止把说明文字误当业务列；扫描失败时抛出用户可理解的模板错误。
    """
    for row_index in range(1, min(worksheet.max_row or 1, 12) + 1):
        # 同名表头按字典规则保留最后一列；规范模板不应出现重复规定字段。
        values = {_text(worksheet.cell(row_index, column).value): column for column in range(1, worksheet.max_column + 1)}
        if _REQUIRED_HEADERS.issubset(values):
            return row_index, values
    raise ValueError("没有识别到安全检查日报表头")


def _image_anchor_row(image: Any) -> int:
    """取得 openpyxl 图片左上锚点对应的 1 基工作表行号，失败返回零。"""
    anchor = getattr(image, "anchor", None)
    marker = getattr(anchor, "_from", None)
    try:
        # DrawingML 锚点行号为 0 基，记录结构统一使用 openpyxl 的 1 基行号。
        return int(marker.row) + 1
    except (AttributeError, TypeError, ValueError):
        return 0


def _save_images(worksheet, image_dir: Path | None, records_by_row: dict[int, dict[str, object]]) -> list[dict[str, object]]:
    """提取页签图片并按锚点行补充所属检查记录摘要。

    ``image_dir`` 为 ``None`` 时只生成图片元数据，不写文件；否则先确保目录存在，并
    以稳定的 ``safety-NNN`` 编号保存 PNG/JPEG。openpyxl 的 ``_images``/``_data``
    属于内部接口，但当前没有等价公开 API，因此访问均采用容错包装。锚点未对应数据
    行时仍保留图片，只把类别和项目留空。
    """
    images: list[dict[str, object]] = []
    if image_dir is not None:
        image_dir.mkdir(parents=True, exist_ok=True)
    # 图片列表保持工作表绘图对象顺序，使同一文件重复分析时 ID 稳定。
    for index, image in enumerate(getattr(worksheet, "_images", []), start=1):
        row = _image_anchor_row(image)
        record = records_by_row.get(row, {})
        image_format = str(getattr(image, "format", "png") or "png").lower()
        extension = "jpg" if image_format in {"jpg", "jpeg"} else "png"
        image_id = f"safety-{index:03d}"
        file_name = f"{image_id}.{extension}"
        if image_dir is not None:
            try:
                # 直接写原始图像字节，不重新编码，避免质量损失和额外图像依赖。
                (image_dir / file_name).write_bytes(image._data())
            except (OSError, AttributeError, TypeError) as exc:
                raise ValueError("安全检查图片提取失败") from exc
        images.append({
            "id": image_id,
            "file_name": file_name,
            "row": row,
            "category": _text(record.get("category")),
            "check_item": _text(record.get("check_item")),
            "width": int(getattr(image, "width", 0) or 0),
            "height": int(getattr(image, "height", 0) or 0),
        })
    return images


def _safety_source_path(path: str | os.PathLike[str]) -> Path:
    """解析并校验安全检查日报路径，只允许当前可完整读取图片的格式。"""
    target = Path(path).resolve()
    if target.suffix.lower() not in SUPPORTED_EXTENSIONS:
        raise ValueError("安全检查日报仅支持 .xlsx 或 .xlsm 文件")
    if not target.is_file():
        raise ValueError("安全检查日报文件不存在")
    return target


def _open_safety_workbook(target: Path):
    """以可读取公式缓存和嵌入图片的模式打开日报。"""
    try:
        # 图片不会在 read_only 模式完整提供；外部链接不参与日报分析，显式关闭可减少开销。
        return load_workbook(target, data_only=True, read_only=False, keep_links=False)
    except Exception as exc:
        raise ValueError("安全检查日报无法读取，请确认文件没有损坏") from exc


def _safety_worksheet(workbook):
    """选取首张尺寸足以容纳规范字段的页签。"""
    worksheet = next(
        (sheet for sheet in workbook.worksheets if sheet.max_row > 1 and sheet.max_column >= 8),
        None,
    )
    if worksheet is None:
        raise ValueError("安全检查日报没有可读取的工作表")
    return worksheet


def _report_date_from_first_row(worksheet, epoch: Any) -> str:
    """从规范模板首行取得第一个可验证日期。"""
    for cell in worksheet[1]:
        report_date = _date_text(cell.value, epoch)
        if report_date:
            return report_date
    return ""


def _safety_record(worksheet, row_index: int, columns: dict[str, int], category: str) -> dict[str, object]:
    """把一行规范字段投影成日清看板可直接使用的检查记录。"""
    return {
        "row": row_index,
        "category": category or "未分类",
        "sequence": (
            _text(worksheet.cell(row_index, columns["序号"]).value)
            if columns.get("序号") else ""
        ),
        "check_item": _text(worksheet.cell(row_index, columns["检查项目"]).value),
        "standard": _text(worksheet.cell(row_index, columns["安全标准要求"]).value),
        "result": _text(worksheet.cell(row_index, columns["检查结果"]).value) or "未填写",
        "problem_description": _text(worksheet.cell(row_index, columns["问题描述"]).value),
        "corrective_action": _text(worksheet.cell(row_index, columns["整改措施"]).value),
        "owner": _text(worksheet.cell(row_index, columns["责任人"]).value),
    }


def _safety_records(worksheet, header_row: int, columns: dict[str, int]) -> tuple[list[dict[str, object]], dict[int, dict[str, object]]]:
    """读取检查明细，并沿用合并单元格最近一次出现的检查类别。"""
    current_category = ""
    records: list[dict[str, object]] = []
    records_by_row: dict[int, dict[str, object]] = {}
    for row_index in range(header_row + 1, worksheet.max_row + 1):
        category = _text(worksheet.cell(row_index, columns["检查类别"]).value)
        if category:
            current_category = category  # 合并类别单元格只有左上角有值，后续明细沿用最近类别。
        check_item = _text(worksheet.cell(row_index, columns["检查项目"]).value)
        result = _text(worksheet.cell(row_index, columns["检查结果"]).value)
        if not check_item and not result:
            continue  # 分隔行、图片占位和说明行不能制造空检查记录。
        record = _safety_record(worksheet, row_index, columns, current_category)
        records.append(record)
        records_by_row[row_index] = record  # 图片锚点以原工作表行号关联，不能改用结果列表序号。
    if not records:
        raise ValueError("安全检查日报没有可展示的检查记录")
    return records, records_by_row


def _result_kind(result: object) -> str:
    """把自由文本结果归为合格、不合格或待确认，优先识别“不合格”。"""
    text = str(result or "")
    if "不合格" in text:
        return "unqualified"  # “不合格”包含“合格”子串，必须先判断。
    if "合格" in text:
        return "qualified"
    return "pending"


def _category_summary(records: list[dict[str, object]]) -> list[dict[str, object]]:
    """按检查类别统计总数、合格数和不合格数，保持类别首次出现顺序。"""
    categories: dict[str, dict[str, object]] = {}
    for record in records:
        category = str(record["category"])
        item = categories.setdefault(
            category, {"category": category, "total": 0, "qualified": 0, "unqualified": 0},
        )
        item["total"] = int(item["total"]) + 1
        kind = _result_kind(record["result"])
        if kind in {"qualified", "unqualified"}:
            item[kind] = int(item[kind]) + 1
    return list(categories.values())


def _attach_record_images(
    records: list[dict[str, object]], images: list[dict[str, object]],
) -> None:
    """按锚点行把一张或多张图片列表挂回对应检查记录。"""
    images_by_row: dict[int, list[dict[str, object]]] = {}
    for image in images:
        images_by_row.setdefault(int(image["row"]), []).append(image)
    for record in records:
        record["images"] = images_by_row.get(int(record["row"]), [])  # 保持列表供移动端画廊展示。


def _safety_summary(
    target: Path, worksheet, report_date: str, records: list[dict[str, object]],
    images: list[dict[str, object]],
) -> dict[str, object]:
    """组合看板需要的全局指标、分类摘要、明细和图片元数据。"""
    counts = {"qualified": 0, "unqualified": 0, "pending": 0}
    for record in records:
        counts[_result_kind(record["result"])] += 1
    total = len(records)
    return {
        "file_name": target.name,
        "report_date": report_date,
        "sheet": worksheet.title,
        "total_checks": total,
        "qualified_count": counts["qualified"],
        "unqualified_count": counts["unqualified"],
        "pending_count": counts["pending"],
        "qualification_rate": round(counts["qualified"] / total * 100, 1) if total else 0.0,
        "category_summary": _category_summary(records),
        "records": records,
        "image_count": len(images),
        "images": images,
    }


def analyze(path: str | os.PathLike[str], *, image_dir: str | os.PathLike[str] | None = None) -> dict[str, object]:
    """解析安全检查日报并返回可直接持久化、展示的完整摘要。

    只接受现有 Excel Open XML 格式。工作簿必须以非只读方式打开，因为嵌入图片不会在
    read_only 模式下完整提供；``data_only=True`` 读取公式结果，``keep_links=False``
    避免不必要的外部链接处理。选择第一张尺寸足以容纳规范表的页签，解析首行日期、
    逐行记录、分类统计和图片关联，任何路径都在 finally 中关闭工作簿。
    """
    target = _safety_source_path(path)
    workbook = _open_safety_workbook(target)
    try:
        worksheet = _safety_worksheet(workbook)
        header_row, columns = _header_row(worksheet)
        records, records_by_row = _safety_records(worksheet, header_row, columns)
        output_image_dir = Path(image_dir).resolve() if image_dir else None
        images = _save_images(worksheet, output_image_dir, records_by_row)
        _attach_record_images(records, images)
        return _safety_summary(
            target, worksheet, _report_date_from_first_row(worksheet, workbook.epoch), records, images,
        )
    finally:
        workbook.close()


def run(path: str | os.PathLike[str], out_dir: str | None = None, log=None) -> dict[str, object]:
    """在统一业务输出目录中解析日报并提取图片。

    显式目录会转为绝对路径并创建；默认目录遵循全局设置。结构化摘要和输出目录一起
    返回，调用方负责把摘要及图片文件纳入日清资料的持久化和权限管理。
    """
    if out_dir is None:
        current = settings.get_settings()
        out_dir = paths.resolve_output_dir("daily_safety_check", **current.output_kwargs())
    else:
        out_dir = os.path.abspath(str(out_dir))
        os.makedirs(out_dir, exist_ok=True)
    summary = analyze(path, image_dir=os.path.join(out_dir, "图片"))
    if log:
        log(f"已解析安全检查日报：{summary['total_checks']} 项，其中不合格 {summary['unqualified_count']} 项")
    return {"out_dir": out_dir, "summary": summary}
