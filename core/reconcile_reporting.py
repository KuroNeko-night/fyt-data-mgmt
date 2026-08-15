# -*- coding: utf-8 -*-
"""考勤对账的可信度评估与 Excel 报告渲染。

本模块只消费 ``reconcile_core`` 已经结构化的处理指标（metrics）和异常行，不读取
原始业务文件，也不参与人员匹配；它把报告样式与评分规则从主流程中分离出来，使识别
算法可以独立演进，而前端结构化可信度与 Excel 报告仍复用同一套评分规则。

主要入口：
- ``assess_credibility(metrics)``：按数据源覆盖、名单匹配、总表填写覆盖、劳务文件
  解析、重复姓名、异常占比、名单交集七个维度累计扣分，并夹紧到 0-100 返回结论；
  该函数只读指标，不修改实际对账结果。
- ``write_summary(anomalies, out_path, credibility=None, log_text=None)``：把结构化
  异常写入“对账异常汇总”表；传入可信度结论时再在工作簿首位插入“可信度报告”表。

边界：本模块不读取原始业务文件、不参与人员匹配，也不决定对账结果；评分规则和 Excel
样式集中在本模块，供桌面端与 Web 端复用同一套口径。
"""

from __future__ import annotations

import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


COLS = [
    "姓名",
    "所属劳务公司",
    "异常类型",
    "我司出勤工时",
    "劳务公司工时",
    "差异",
    "差异明细",
    "来源文件",
]


class _CredibilityAssessment:
    """累积可信度检查项和扣分，并在结束时统一限制到百分制范围。

    评分采用“满分起扣”模型：初始 100 分，各维度把检查结论写入
    ``checks`` 并同步从 ``score`` 扣减，最后统一夹紧到 0-100，避免
    中途夹紧造成多个高风险项叠加时失真。
    """

    def __init__(self):
        """初始化满分状态和空检查列表；所有维度共享同一实例累积结论。"""
        self.score = 100
        self.checks = []

    def add(self, item, level, detail, deduct=0):
        """追加一条检查结论，并记录该风险对应的扣分。

        参数：
            item: 检查项目名称，如“数据来源覆盖天数”。
            level: 结论级别，取值“正常/提示/警告/严重”，供前端与 Excel 上色。
            detail: 面向用户的说明文字。
            deduct: 该项目扣分；0 表示仅提示不扣分。

        副作用：向 ``checks`` 追加一条结论，并从 ``score`` 同步扣减。
        """
        self.checks.append({"项目": item, "级别": level, "说明": detail})
        self.score -= deduct

    def result(self):
        """返回前端和 Excel 共用的百分制可信度结果。

        返回：
            dict，包含 score（0-100 的整数）、level（高/中/低）和 checks
            （所有检查结论列表）。该方法只做夹紧和评级，不修改其它状态。
        """
        score = max(0, min(100, self.score))
        level = "高" if score >= 85 else ("中" if score >= 60 else "低")
        return {"score": score, "level": level, "checks": self.checks}


def _assess_source_days(metrics, assessment):
    """评估数据来源的日期覆盖是否符合完整月度数据的基本特征。

    参数：
        metrics: 处理指标字典，读取 ``source_days``。
        assessment: 累积评分的 ``_CredibilityAssessment`` 实例。

    副作用：仅向 ``assessment`` 追加结论并扣分；天数阈值按完整月度口径
    粗判（0 天为格式异常，<10 天疑似只读部分明细，<18 天略少于自然月）。
    """
    days = metrics.get("source_days", 0)
    # 阈值按月度考勤口径划分：18 天以下明显少于一个自然月，10 天以下疑似只读到局部明细。
    if days == 0:
        assessment.add("数据来源覆盖天数", "严重", "未读到任何日期，数据来源可能格式异常或选错文件", 40)
    elif days < 10:
        assessment.add("数据来源覆盖天数", "警告", "仅覆盖 %d 天，明显偏少，可能只读到部分明细" % days, 20)
    elif days < 18:
        assessment.add("数据来源覆盖天数", "提示", "覆盖 %d 天，略少，请确认是否为完整月度数据" % days, 5)
    else:
        assessment.add("数据来源覆盖天数", "正常", "覆盖 %d 天，符合月度考勤预期" % days)


def _assess_source_matching(metrics, assessment):
    """评估数据来源人员能否在待对总表名单中找到。

    参数：
        metrics: 处理指标字典，读取 ``source_people`` 与 ``source_unmatched``。
        assessment: 累积评分的 ``_CredibilityAssessment`` 实例。

    副作用：仅向 ``assessment`` 追加结论并扣分；未匹配率按 50%/20% 两级
    区分“疑似选错文件”和“姓名写法需要抽查”。
    """
    people = metrics.get("source_people", 0)
    unmatched = metrics.get("source_unmatched", 0)
    if people <= 0:
        assessment.add("数据来源→名单匹配", "严重", "数据来源未读到任何人", 30)
        return
    # people > 0 已经由上面的提前返回保证，这里不会再除零。
    rate = unmatched / people
    if rate > 0.5:
        assessment.add(
            "数据来源→名单匹配", "警告",
            "%d/%d 人未在待对表名单中(%.0f%%)，疑似姓名不一致或选错文件"
            % (unmatched, people, rate * 100), 20,
        )
    elif rate > 0.2:
        assessment.add(
            "数据来源→名单匹配", "提示",
            "%d/%d 人未在名单中(%.0f%%)，请抽查这些人的姓名写法"
            % (unmatched, people, rate * 100), 8,
        )
    else:
        assessment.add(
            "数据来源→名单匹配", "正常",
            "%d/%d 人未匹配(%.0f%%)，在合理范围" % (unmatched, people, rate * 100),
        )


def _assess_fill_coverage(metrics, assessment):
    """评估数据来源对待对总表人员的填写覆盖率。

    参数：
        metrics: 处理指标字典，读取 ``zong_people`` 与 ``filled_people``。
        assessment: 累积评分的 ``_CredibilityAssessment`` 实例。

    副作用：仅向 ``assessment`` 追加结论并扣分；总表人数为 0 时说明
    前置流程未产生名单，直接跳过该维度以免除零。
    """
    total_people = metrics.get("zong_people", 0)
    # 总表名单为空时无法计算覆盖率，也说明上游没有可对账对象，跳过该维度。
    if total_people <= 0:
        return
    filled_people = metrics.get("filled_people", 0)
    rate = filled_people / total_people
    if rate < 0.2:
        assessment.add(
            "总表填写覆盖", "警告",
            "仅 %d/%d 人被填入工时(%.0f%%)，数据来源与名单可能不匹配"
            % (filled_people, total_people, rate * 100), 15,
        )
    elif rate < 0.6:
        assessment.add(
            "总表填写覆盖", "提示",
            "%d/%d 人被填入(%.0f%%)，其余无数据来源，请确认是否漏传文件"
            % (filled_people, total_people, rate * 100), 6,
        )
    else:
        assessment.add(
            "总表填写覆盖", "正常",
            "%d/%d 人被填入工时(%.0f%%)" % (filled_people, total_people, rate * 100),
        )


def _assess_labor_total_consistency(meta, assessment):
    """检查劳务表自带合计列与逐日求和是否存在系统性偏差。

    参数：
        meta: 单个劳务文件的解析元数据。
        assessment: 累积评分的 ``_CredibilityAssessment`` 实例。

    副作用：仅向 ``assessment`` 追加结论并扣分；只有同时具备合计列和
    人数时才校验，偏差按人数归一，避免大名单把个别差异放大成严重项。
    """
    if not meta["has_total_col"] or meta["people"] <= 0:
        return
    mismatch = meta["total_sum_mismatch"]
    rate = mismatch / meta["people"]
    filename = meta["file"]
    if rate > 0.3:
        assessment.add(
            "合计列校验·%s" % filename, "警告",
            "%d/%d 人表内合计与逐日求和不符(%.0f%%)，合计列可能识别错误"
            % (mismatch, meta["people"], rate * 100), 12,
        )
    elif mismatch > 0:
        assessment.add(
            "合计列校验·%s" % filename, "提示",
            "%d 人表内合计与逐日求和有差异，属对方原表口径差异或加班另计" % mismatch, 3,
        )


def _assess_labor_files(metrics, assessment):
    """逐个评估劳务文件的页签识别、日期列数量和合计列一致性。

    参数：
        metrics: 处理指标字典，读取 ``labor_meta`` 列表。
        assessment: 累积评分的 ``_CredibilityAssessment`` 实例。

    副作用：仅向 ``assessment`` 追加结论并扣分；每个文件先检查子表是否
    识别成功，再按 28-31 天的日列区间判断漏读/误纳，最后复用合计列校验。
    """
    for meta in metrics.get("labor_meta", []):
        filename = meta["file"]
        item = "劳务文件解析·%s" % filename
        if meta.get("n_candidates", 0) == 0 or meta["sheet"] is None:
            assessment.add(item, "严重", "未识别到考勤子表，该文件未参与对账", 25)
            continue
        # 自然月通常 28-31 天：低于 28 可能漏读日期列，高于 31 可能误把非日期列当成日列。
        if meta["day_cols"] < 28:
            assessment.add(item, "警告", "仅识别到 %d 个日列(通常应≥28)，可能漏读日期列" % meta["day_cols"], 12)
        elif meta["day_cols"] > 31:
            assessment.add(item, "提示", "识别到 %d 个日列(多于31)，请确认是否误纳非日期列" % meta["day_cols"], 5)
        else:
            assessment.add(
                item, "正常",
                "识别 %d 人、%d 个日列，%s" % (
                    meta["people"],
                    meta["day_cols"],
                    "有合计列" if meta["has_total_col"] else "无合计列(逐日求和)",
                ),
            )
        _assess_labor_total_consistency(meta, assessment)


def _assess_duplicate_names(metrics, assessment):
    """提示跨劳务文件出现的重复姓名及当前覆盖语义。

    参数：
        metrics: 处理指标字典，读取 ``labor_duplicate_names``。
        assessment: 累积评分的 ``_CredibilityAssessment`` 实例。

    副作用：仅向 ``assessment`` 追加结论并扣分；重复姓名按“后读取为准”
    覆盖，所以这里只作提示级扣分，提醒确认是否同名不同人。
    """
    duplicate_count = metrics.get("labor_duplicate_names", 0)
    if duplicate_count > 0:
        assessment.add(
            "劳务文件姓名重复", "提示",
            "%d 个姓名在多个劳务文件中重复，已按后读取的为准，请确认无同名不同人"
            % duplicate_count, 5,
        )


def _assess_difference_rate(metrics, assessment):
    """根据双方共有人员中的差异人数判断是否存在系统性错位风险。

    参数：
        metrics: 处理指标字典，读取 ``matched_pairs`` 与 ``diff_people``。
        assessment: 累积评分的 ``_CredibilityAssessment`` 实例。

    副作用：仅向 ``assessment`` 追加结论并扣分；交集人数为 0 时无法
    逐人对账，直接警告并返回，后续比例计算以交集人数为分母。
    """
    matched = metrics.get("matched_pairs", 0)
    difference_people = metrics.get("diff_people", 0)
    if matched <= 0:
        assessment.add("异常占比", "警告", "待对表与劳务公司名单无任何交集，无法逐人对账，请检查是否选错文件", 25)
        return
    # 无交集时对账无法成立，已在上方提前返回；这里分母必不为 0。
    rate = difference_people / matched
    if rate > 0.9:
        assessment.add(
            "异常占比", "警告",
            "%d/%d 双方都有的人中 %.0f%% 存在工时差异，比例过高，更可能是填写/对齐系统性错误而非真实差异，请重点核对"
            % (difference_people, matched, rate * 100), 20,
        )
    elif rate > 0.5:
        assessment.add("异常占比", "提示", "%.0f%% 的人存在工时差异，偏高，建议抽样核对若干人的原始明细" % (rate * 100), 8)
    else:
        assessment.add(
            "异常占比", "正常",
            "%d/%d 人存在差异(%.0f%%)，属正常对账范围"
            % (difference_people, matched, rate * 100),
        )


def _assess_roster_intersection(metrics, assessment):
    """评估双方独有名单占比，识别姓名写法或文件范围不一致。

    参数：
        metrics: 处理指标字典，读取 ``matched_pairs``、``only_us`` 与
        ``only_labor``。
        assessment: 累积评分的 ``_CredibilityAssessment`` 实例。

    副作用：仅向 ``assessment`` 追加结论并扣分；双方名单都为空时跳过，
    避免对空名单计算比例。
    """
    matched = metrics.get("matched_pairs", 0)
    only_us = metrics.get("only_us", 0)
    only_labor = metrics.get("only_labor", 0)
    # 双方都没有名单时避免除零；该维度仅在总名单非空时计算。
    total_names = matched + only_us + only_labor
    if total_names > 0 and (only_us + only_labor) / total_names > 0.5:
        assessment.add(
            "名单交集", "提示",
            "仅一方有的人数较多(我司独有%d、劳务独有%d)，可能是姓名写法不一致或名单范围不同"
            % (only_us, only_labor), 6,
        )


def assess_credibility(metrics):
    """根据处理指标生成可信度评分和逐项结论，不修改实际对账结果。

    参数：
        metrics: 由 ``reconcile_core`` 汇总的处理指标字典。

    返回：
        dict，包含 score、level、checks 三项，与
        ``_CredibilityAssessment.result`` 返回结构一致。

    说明：
        评估顺序固定为“来源覆盖 -> 来源匹配 -> 总表覆盖 -> 劳务文件解析 ->
        重复姓名 -> 异常占比 -> 名单交集”，各维度只向内部评估对象追加结论，
        不会改写 metrics 或任何对账结果。
    """
    assessment = _CredibilityAssessment()
    _assess_source_days(metrics, assessment)
    _assess_source_matching(metrics, assessment)
    _assess_fill_coverage(metrics, assessment)
    _assess_labor_files(metrics, assessment)
    _assess_duplicate_names(metrics, assessment)
    _assess_difference_rate(metrics, assessment)
    _assess_roster_intersection(metrics, assessment)
    return assessment.result()


def _report_styles():
    """集中创建同一工作簿内可复用的边框、字体和对齐样式。

    返回：
        dict，键为样式用途，值为 openpyxl 样式对象；同一个样式对象被多个
        单元格复用，既保证观感一致，也减少重复构造开销。
    """
    thin = Side(style="thin", color="BBBBBB")
    return {
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        "header_fill": PatternFill("solid", fgColor="305496"),
        "header_font": Font(name="微软雅黑", bold=True, color="FFFFFF", size=10),
        "body_font": Font(name="微软雅黑", size=10),
        "center": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "left": Alignment(horizontal="left", vertical="center", wrap_text=True),
    }


def _write_summary_rows(worksheet, anomalies, styles):
    """写入异常正文；异常类型仅在类型列使用低饱和底色。

    参数：
        worksheet: 已建好表头的“对账异常汇总”工作表。
        anomalies: 结构化异常行列表，每行包含 COLS 定义键。
        styles: ``_report_styles`` 返回的复用样式集合。

    说明：
        正文从第 3 行开始（前两行为标题和表头）；异常类型列用固定低饱和
        底色区分四类异常，其余单元格不上色，避免整行底色干扰阅读。
    """
    type_fill = {
        "总工时不一致": PatternFill("solid", fgColor="FCE4D6"),
        "逐日工时不一致": PatternFill("solid", fgColor="FFF2CC"),
        "仅我司名单有": PatternFill("solid", fgColor="E2EFDA"),
        "仅劳务公司有": PatternFill("solid", fgColor="DDEBF7"),
    }
    for row_number, anomaly in enumerate(anomalies, start=3):
        for column, key in enumerate(COLS, start=1):
            cell = worksheet.cell(row_number, column, anomaly.get(key, ""))
            cell.border = styles["border"]
            cell.alignment = styles["left"] if key in ("差异明细", "来源文件") else styles["center"]
            cell.font = styles["body_font"]
        fill = type_fill.get(anomaly.get("异常类型"))
        if fill:
            worksheet.cell(row_number, 3).fill = fill


def _write_credibility_sheet(workbook, credibility, log_text=None):
    """在工作簿首位写入可信度结论、逐项检查和可选运行日志。

    参数：
        workbook: 已包含异常汇总表的工作簿。
        credibility: ``assess_credibility`` 返回的结论 dict。
        log_text: 可选的运行日志文本；多行时逐行写入以保留换行结构。

    返回：
        新建的“可信度报告”工作表。

    说明：
        该表通过索引 0 插入到工作簿首位，保证下载后用户先看到评估结论
        再查看异常清单；运行日志按原样写入纯文本，仅用于事后追溯。
    """
    worksheet = workbook.create_sheet("可信度报告", 0)
    styles = _report_styles()
    level = credibility["level"]
    score = credibility["score"]
    level_color = {"高": "C6EFCE", "中": "FFEB9C", "低": "FFC7CE"}.get(level, "FFFFFF")
    level_font = {"高": "006100", "中": "9C6500", "低": "9C0006"}.get(level, "000000")
    verdict = {
        "高": "结果可信度高，可放心使用，仅需按异常清单核对个别差异。",
        "中": "结果可信度中等，建议先看下方“提示/警告”项，再核对异常清单。",
        "低": "结果可信度低，很可能存在程序识别或文件选择问题，请先排查下方警告项，勿直接采用！",
    }

    worksheet.merge_cells("A1:C1")
    title = worksheet.cell(1, 1, "对账结果可信度报告")
    title.font = Font(name="微软雅黑", bold=True, size=14)
    title.alignment = styles["center"]
    worksheet.row_dimensions[1].height = 30

    worksheet.merge_cells("A2:C2")
    conclusion = worksheet.cell(
        2, 1,
        "综合结论：可信度【%s】  评分 %d/100 —— %s"
        % (level, score, verdict.get(level, "")),
    )
    conclusion.font = Font(name="微软雅黑", bold=True, size=11, color=level_font)
    conclusion.fill = PatternFill("solid", fgColor=level_color)
    conclusion.alignment = styles["left"]
    worksheet.row_dimensions[2].height = 40

    for column, heading in enumerate(("检查项目", "结论级别", "说明"), start=1):
        cell = worksheet.cell(4, column, heading)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["center"]
        cell.border = styles["border"]

    level_fills = {"正常": "E2EFDA", "提示": "FFF2CC", "警告": "FCE4D6", "严重": "FFC7CE"}
    row_number = 5
    for check in credibility["checks"]:
        item = worksheet.cell(row_number, 1, check["项目"])
        level_cell = worksheet.cell(row_number, 2, check["级别"])
        detail = worksheet.cell(row_number, 3, check["说明"])
        for cell in (item, level_cell, detail):
            cell.border = styles["border"]
        item.alignment = detail.alignment = styles["left"]
        item.font = detail.font = styles["body_font"]
        level_cell.alignment = styles["center"]
        level_cell.fill = PatternFill("solid", fgColor=level_fills.get(check["级别"], "FFFFFF"))
        level_cell.font = Font(name="微软雅黑", size=10, bold=check["级别"] in ("警告", "严重"))
        row_number += 1

    row_number += 1
    worksheet.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=3)
    worksheet.cell(row_number, 1, "运行日志（供追溯）").font = Font(name="微软雅黑", bold=True, size=10)
    for line in (log_text or "").splitlines():
        row_number += 1
        worksheet.cell(row_number, 1, line).font = Font(name="Consolas", size=9)

    worksheet.column_dimensions["A"].width = 26
    worksheet.column_dimensions["B"].width = 10
    worksheet.column_dimensions["C"].width = 70
    worksheet.freeze_panes = "A5"
    return worksheet


def write_summary(anomalies, out_path, credibility=None, log_text=None):
    """把结构化异常写入汇总工作簿，并按需在首位附加可信度报告。

    参数：
        anomalies: 结构化异常行列表。
        out_path: 目标 Excel 文件路径；调用方需保证目录存在且路径安全。
        credibility: 可选的可信度结论 dict，传入时额外生成“可信度报告”表。
        log_text: 可选运行日志，仅在生成可信度报告时写入。

    返回：
        out_path 原样返回，便于调用方继续传递输出路径。

    说明：
        默认活动表作为“对账异常汇总”，可信度报告插入到索引 0；异常为空时
        写入“未发现异常”提示而不是留空表。文件由 openpyxl 直接保存到
        out_path（覆盖已存在文件），本模块不负责原子替换与回滚。
    """
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "对账异常汇总"
    styles = _report_styles()

    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
    title = worksheet.cell(
        1, 1,
        "对账异常汇总表  （生成时间：%s）"
        % datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
    )
    title.font = Font(name="微软雅黑", bold=True, size=13)
    title.alignment = styles["center"]
    worksheet.row_dimensions[1].height = 28

    for column, heading in enumerate(COLS, start=1):
        cell = worksheet.cell(2, column, heading)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["center"]
        cell.border = styles["border"]

    _write_summary_rows(worksheet, anomalies, styles)
    if not anomalies:
        worksheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(COLS))
        worksheet.cell(3, 1, "未发现异常，全部一致 ✔").alignment = styles["center"]

    for column, width in enumerate((10, 14, 14, 12, 12, 10, 46, 26), start=1):
        worksheet.column_dimensions[get_column_letter(column)].width = width
    worksheet.freeze_panes = "A3"
    if credibility is not None:
        _write_credibility_sheet(workbook, credibility, log_text)
    workbook.save(out_path)
    workbook.close()
    return out_path
