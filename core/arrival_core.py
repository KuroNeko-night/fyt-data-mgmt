# -*- coding: utf-8 -*-
"""每日主料到料明细的识别、汇总、成品解析与 Excel 输出。

输入既可以是送货计划源表，也可以是已经生成完成的《每日主料到料明细》。源表路径会
识别批次、读取剩余未收数并生成标准工作簿；成品路径只解析表内已有批次和数量，直接
供日清看板使用。两条路径最终都输出同一种结构化批次结果，并复用主数据库补空逻辑。

本模块不依赖桌面或 Web 界面。输出目录、批次记忆和进度反馈分别由 ``paths``、
``settings`` 与 ``common_core.Progress`` 统一管理。
"""
import os
import re
import glob
import datetime
import numbers

import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter

from . import paths as _paths
from . import settings as _settings
from . import shape_detect
from . import common_core as _common       # Progress 进度上报辅助
from . import material_catalog
from .common_core import warn_if_uncached   # 公式未刷新检测(读关键表前告警)

# 旧版送货计划的固定列仅是最后兜底。正常路径优先按表头文字定位，再尝试数据形态识别，
# 防止模板插列或调整顺序后把需求数、剩余未收数读错。
COL_CODE, COL_NAME, COL_SUPPLIER, COL_DEMAND, COL_REMAIN = 2, 3, 5, 7, 12
FIRST_COL, BATCH_STRIDE, DATA_START_ROW = 3, 9, 8
DEFAULT_TOTAL = 566
DEFAULT_TOP_LABEL = "截止16点的数据"
# 输出列宽: 序号/编码/名称/供应商/需求数/剩余未收数/备注
COL_WIDTHS = [7.9, 17.6, 29.6, 34.4, 9.0, 10.5, 11.9]
HEADER_SCAN_ROWS = 8               # 表头最多出现在前几行内

# 表头先去除空白再做子串匹配。供应商列必须排除代码、编号和代号列，避免把供应商编码
# 当成供应商名称写入成品明细。
ALIAS_CODE   = ("物料编码", "物料编号", "材料编码", "材料编号", "物料号", "料号")
ALIAS_NAME   = ("物料名称", "材料名称", "品名")
ALIAS_DEMAND = ("需求数", "需求数量", "需求量", "计划需求", "计划数量")
ALIAS_REMAIN = ("剩余未收数", "剩余未收", "未收数", "未收", "未到货", "未到", "缺料")
CONFIG_PATH = os.path.join(os.path.expanduser("~"), ".arrival_table_config.json")

# 批次号优先读取表内“订单 XXX 批次”标题，失败后才从文件名提取字母数字组合。
RE_TITLE = re.compile(r'订单\s*([A-Za-z0-9]+(?:-[A-Za-z0-9]+)*)\s*批次')
RE_NAME  = re.compile(r'([A-Za-z]{2,}\d{2,}(?:-[A-Za-z0-9]+)*)')
RE_REPORT_DATE = re.compile(r'(?<!\d)(20\d{2})(\d{2})(\d{2})(?!\d)')

# 日清资料入口接收已经生成好的成品表，不重复执行送货计划业务。成品表中每个批次横向
# 占一个区块，批次号、总类数、到货数量和差异都必须从区块内容读取，不能再要求人工
# 重复填写。
FINISHED_TOTAL_LABELS = ("主料总共类", "主料总共类数", "主料总类数")
FINISHED_ARRIVED_LABELS = ("到货数量", "已到货", "到货类数")
FINISHED_MISSING_LABELS = ("差异", "未收料", "缺料类数")

FONT_NAME = "微软雅黑"
thin = Side(style='thin')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
FONT = Font(name=FONT_NAME, size=11)
FONT_B = Font(name=FONT_NAME, size=11, bold=True)
# 表头/标签蓝色填充: 主题色4(蓝) tint≈0.4, 与示例一致
BLUE_FILL = PatternFill(patternType='solid', fgColor=Color(theme=4, tint=0.3999755851924192))


def beijing_date():
    """返回东八区当前日期的紧凑格式，供到料输出文件命名。"""
    # 使用 UTC 加固定八小时，不受服务器本地时区配置影响。
    return (datetime.datetime.now(datetime.timezone.utc) +
            datetime.timedelta(hours=8)).strftime("%Y%m%d")


def _first_ws(wb):
    """返回优先使用的工作表：活动表有效时使用活动表，否则取第一张。

    用户可能重命名工作表，因此不能依赖 ``Sheet1``。少数损坏或特殊工作簿读取活动表
    属性时会异常，此处回退第一张，并把真正的内容判断交给 :func:`_pick_data_ws`。
    """
    try:
        ws = wb.active
        if ws is not None:
            return ws
    except Exception:
        # 这里只选择候选工作表，活动表元数据异常不应阻止后续扫描第一张表。
        pass
    return wb[wb.sheetnames[0]]


def _pick_data_ws(wb, log=None):
    """挑出真正含到货计划数据的工作表，避免静默选择空页。

    单页工作簿直接使用活动页。多页时先尊重用户保存时停留的活动页；若该页没有可识别
    表头，再按工作簿顺序寻找第一个有效页。全部失败时仍返回活动页，让后续形态识别和
    固定列兼容逻辑有机会处理，并通过日志提示人工核对。
    """
    active = _first_ws(wb)
    if len(wb.sheetnames) <= 1:
        return active
    # 活动表本身有效时直接使用，尊重用户在 Excel 中保存的当前页选择。
    if locate_columns(active) is not None:
        if log:
            log("· 多子表: 读取活动表《%s》" % active.title)
        return active
    # 活动表无效时才扫描其余页，避免多页文件中误选说明页或历史页。
    for name in wb.sheetnames:
        ws = wb[name]
        if ws is active:
            continue
        if locate_columns(ws) is not None:
            if log:
                log("· 多子表: 活动表《%s》无有效表头, 改读《%s》"
                    % (active.title, name))
            return ws
    if log:
        log("⚠ 多子表: 各子表均未识别到有效表头, 按活动表《%s》读取(可能选错文件)"
            % active.title)
    return active

def _norm(v):
    """移除中英文空格、换行和制表符，生成表头匹配文本。"""
    if v is None:
        return ""
    s = str(v)
    for ch in ("\r\n", "\n", "\r", "\t", " ", "　"):
        s = s.replace(ch, "")
    return s.strip()

def _match(text, aliases):
    """判断标准化表头是否包含任一受支持别名。"""
    return any(a in text for a in aliases)


def _header_cells(ws, row, start_column=1, end_column=None):
    """读取一段表头并返回 ``(列号, 标准化文字)`` 列表。

    表头识别和成品报表解析原本分别维护两套嵌套循环，容易在新增别名时只修到一处。
    这里统一完成空白清理，并保留原始列号，调用方只需要决定扫描的横向边界。
    """
    last_column = min(ws.max_column or 1, end_column or (ws.max_column or 1))
    return [
        (column, _norm(ws.cell(row=row, column=column).value))
        for column in range(start_column, last_column + 1)
    ]


def _classify_header_cells(cells):
    """把一行表头归类为到料业务字段列。

    “供应商信息”属于明确字段，优先于普通“供应商”；供应商代码、编号和代号不能作为
    名称列。其他字段均取从左到右的首个命中，避免合并表头中的重复文字改变既有口径。
    """
    columns = {"code": 0, "name": 0, "supplier": 0, "demand": 0, "remain": 0}
    supplier_fallback = 0
    for column, text in cells:
        if not text:
            continue
        if not columns["code"] and _match(text, ALIAS_CODE):
            columns["code"] = column
        if not columns["name"] and _match(text, ALIAS_NAME):
            columns["name"] = column
        if not columns["demand"] and _match(text, ALIAS_DEMAND):
            columns["demand"] = column
        if not columns["remain"] and _match(text, ALIAS_REMAIN):
            columns["remain"] = column
        if "供应商" not in text:
            continue
        if "信息" in text and not columns["supplier"]:
            columns["supplier"] = column
        elif not any(word in text for word in ("代码", "编号", "代号")):
            supplier_fallback = supplier_fallback or column
    columns["supplier"] = columns["supplier"] or supplier_fallback
    return columns


def _has_required_arrival_columns(columns):
    """判断编码、需求数和剩余未收数三项必需列是否齐全。"""
    return all(columns.get(name) for name in ("code", "demand", "remain"))


def locate_columns(ws):
    """按表头文字定位送货计划的业务列。

    在前 ``HEADER_SCAN_ROWS`` 行寻找材料编码表头，再在同一行匹配名称、供应商、需求数
    和剩余未收数。供应商信息列优先于普通供应商列，且排除代码类字段。编码、需求和
    剩余未收是计算所需关键列，缺任一项即返回 ``None``；名称和供应商可由主数据库
    后续补空，因此不是致命缺失。
    """
    scan = min(HEADER_SCAN_ROWS, ws.max_row or 1)
    for row in range(1, scan + 1):
        columns = _classify_header_cells(_header_cells(ws, row))
        if _has_required_arrival_columns(columns):
            columns["header_row"] = row
            return columns
    return None


# 数据形态画像:编码/名称/供应商/需求/剩余未收。code 与两数值列为必需(与 locate 一致)。
_SHAPE_PROFILE = [
    ("code", shape_detect.CODE, True),
    ("name", shape_detect.TEXT, False),
    ("supplier", shape_detect.TEXT, False),
    ("demand", shape_detect.NUMBER, True),
    ("remain", shape_detect.NUMBER, True),
]


def _locate_by_shape(ws, log=None):
    """表头文字识别失败时，按数据形态推断列位置。

    返回结构与 :func:`locate_columns` 相同。形态识别属于低确定性兜底，调用层会记录
    提示并在支持人工复核的流程中展示列计划，不能把它当作明确表头识别。
    """
    hr, col, _conf = shape_detect.detect_by_shape(
        ws, _SHAPE_PROFILE, scan_rows=HEADER_SCAN_ROWS, log=log)
    if not hr:
        return None
    return {"code": col["code"], "name": col.get("name", 0),
            "supplier": col.get("supplier", 0), "demand": col["demand"],
            "remain": col["remain"], "header_row": hr}

def detect_batch(path):
    """优先从活动表 A1 标题识别批次号，失败后从文件名识别。

    批次识别只是辅助默认值，读取异常或未命中时返回空字符串，由人工复核界面补填，
    因此这里不会因单个不规范标题中断整批文件扫描。
    """
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        a1 = _first_ws(wb)['A1'].value or ""
        wb.close()
        m = RE_TITLE.search(str(a1))
        if m:
            return m.group(1)
    except Exception:
        # 文件仍可能通过正式处理路径给出更具体错误；批次预识别阶段只回退文件名。
        pass
    m = RE_NAME.search(os.path.basename(path))
    return m.group(1) if m else ""


def detect_report_date(path):
    """从成品到料表文件名读取合法业务日期，识别不到时返回空字符串。"""
    match = RE_REPORT_DATE.search(os.path.basename(str(path)))
    if not match:
        return ""
    try:
        return datetime.date(*(int(part) for part in match.groups())).isoformat()
    except ValueError:
        # 正则只能保证八位数字形态，日期构造负责排除非法月日。
        return ""


def find_plan_files(folder):
    """返回目录内非临时送货计划表，并按修改时间从新到旧排列。"""
    files = [f for f in glob.glob(os.path.join(folder, "*送货计划*.xlsx"))
             if not os.path.basename(f).startswith("~$")]
    return sorted(files, key=os.path.getmtime, reverse=True)


def _resolve_plan_columns(ws, log=None):
    """按文字、数据形态、旧模板固定列的顺序确定送货计划列。"""
    columns = locate_columns(ws)
    if columns is None:
        columns = _locate_by_shape(ws, log=log)
    if columns is not None:
        return columns
    # 固定列只能作为最后兜底，确保插列后的新模板优先走可解释的识别结果。
    return {
        "code": COL_CODE,
        "name": COL_NAME,
        "supplier": COL_SUPPLIER,
        "demand": COL_DEMAND,
        "remain": COL_REMAIN,
        "header_row": 2,
    }


def _scan_plan_rows(ws, columns):
    """读取完整送货计划，分别统计总类数、隐藏行、待核对行和未到明细。"""
    materials = []
    total = hidden = pending = 0
    for row in range(columns["header_row"] + 1, ws.max_row + 1):
        code = ws.cell(row=row, column=columns["code"]).value
        if not _norm(code):
            continue
        total += 1  # 总类数必须来自完整数据行，不能受 Excel 当前筛选视图影响。
        row_dimension = ws.row_dimensions.get(row)
        if row_dimension is not None and row_dimension.hidden:
            hidden += 1  # 隐藏行仍参与业务计算，这里只为日志保留可审计数量。
        remain = ws.cell(row=row, column=columns["remain"]).value
        if isinstance(remain, bool):
            continue  # bool 是数字子类，但 TRUE/FALSE 不是有效的剩余未收数量。
        if remain is None:
            pending += 1  # 公式缓存缺失不能静默当作零，留给人工核对。
            continue
        if not isinstance(remain, numbers.Number) or remain == 0:
            continue
        materials.append([
            str(code),
            ws.cell(row=row, column=columns["name"]).value if columns["name"] else None,
            ws.cell(row=row, column=columns["supplier"]).value if columns["supplier"] else None,
            ws.cell(row=row, column=columns["demand"]).value if columns["demand"] else None,
            remain,
        ])
    return {"materials": materials, "total": total, "pending": pending, "hidden": hidden}


def _log_plan_scan_warnings(path, result, log):
    """把公式缓存和隐藏行提示集中输出，保持扫描函数只负责数据读取。"""
    if not log:
        return
    if result["pending"]:
        log("⚠ 《%s》有 %d 行的剩余未收数读不到值(可能公式未刷新), 未计入未收料, 请核对。"
            % (os.path.basename(path), result["pending"]))
    if result["hidden"]:
        log("· 《%s》包含 %d 条筛选或隐藏物料，已按完整源表参与识别。"
            % (os.path.basename(path), result["hidden"]))


def inspect_plan(path, log=None):
    """扫描完整送货计划并返回主料总类数、未收明细和读取提示。

    列位置依次尝试表头文字、数据形态和旧版固定列。主料总类数按表头下方具有物料编码
    的全部数据行计算，不受 Excel 当前筛选或手工隐藏状态影响；未收明细则只纳入“剩余
    未收数”为非零数字的物料。这样用户保存文件时停留在哪个筛选视图都不会改变结果。

    公式缓存为空的 ``None`` 仍不能静默当作零，会单独计数并提示人工核对。返回字典中的
    ``materials`` 保持历史五列结构，``total`` 是自动识别的主料类数，``pending`` 是无法
    读取剩余未收数的行数，``hidden`` 是扫描到的隐藏物料行数。
    """
    # 剩余未收数常是公式; 若该表未被 Excel 刷新过, data_only 读出 None,
    # 会把整列误当"已收"静默排除(全部物料显示已收)。读表前先醒目告警。
    if log:
        warn_if_uncached(path, log, what="剩余未收数")
    wb = _common.load_data_only(path)   # 跳过内嵌透视缓存解析，只读取公式缓存结果。
    try:
        ws = _pick_data_ws(wb, log=log)
        result = _scan_plan_rows(ws, _resolve_plan_columns(ws, log=log))
    finally:
        wb.close()  # 预扫描和正式处理都会打开文件，及时释放 Windows 文件句柄。
    _log_plan_scan_warnings(path, result, log)
    return result


def extract_unreceived(path, log=None):
    """兼容旧调用方，仅返回完整扫描结果中的非零未收物料列表。"""
    return inspect_plan(path, log=log)["materials"]


def _style(cell, bold=False, fill=False, align=CENTER):
    """为标准到料报表单元格应用统一边框、字体、对齐和可选蓝底。"""
    cell.border = BORDER
    cell.alignment = align
    cell.font = FONT_B if bold else FONT
    if fill:
        cell.fill = BLUE_FILL

def _write_batch(ws, c0, batch_no, materials, total, remark, top_label, pending=0):
    """从起始列写入一个横向批次区块，并返回缺料与已到类别数。

    每个区块固定占七个业务列，外加两个空列形成九列步长。``materials`` 一行代表一个
    未收物料类别，因此差异是明细行数；``pending`` 是“剩余未收数”读不到值（公式未刷新）
    的行数，不能当作已到货。已到类别数按总类别数减差异再减待核对数计算，异常负值仍由
    上层结果质量提示处理。
    """
    diff = len(materials)
    arrived = total - diff - pending  # 公式未刷新的待核对行不得计入已到货，避免误报“全部已到货”。
    # 第一、二行是截止标签与批次号，横跨区块前三列并保持白底。
    for row, val in [(1, top_label), (2, batch_no)]:
        ws.cell(row=row, column=c0, value=val)
        ws.merge_cells(start_row=row, start_column=c0, end_row=row, end_column=c0 + 2)
        for cc in range(c0, c0 + 3):
            _style(ws.cell(row=row, column=cc))
    # 第三至五行采用“两列蓝底标签 + 一列白底数值”的标准模板结构。
    for row, label, val in [(3, "主料总共类", total), (4, "到货数量", arrived), (5, "差异", diff)]:
        ws.cell(row=row, column=c0, value=label)
        ws.merge_cells(start_row=row, start_column=c0, end_row=row, end_column=c0 + 1)
        for cc in range(c0, c0 + 2):
            _style(ws.cell(row=row, column=cc), fill=True)   # 标签蓝底
        ws.cell(row=row, column=c0 + 2, value=val)
        _style(ws.cell(row=row, column=c0 + 2))              # 数值白底
    # 备注存在时横跨整个七列区块，长文本左对齐并自动换行。
    if remark:
        ws.cell(row=6, column=c0, value="备注: " + str(remark))
        ws.merge_cells(start_row=6, start_column=c0, end_row=6, end_column=c0 + 6)
        for cc in range(c0, c0 + 7):
            _style(ws.cell(row=6, column=cc), align=LEFT)
    # 第七行是明细表头，保留“需求数”和“剩余未收数”供管理人员核对数量差额。
    for i, h in enumerate(["序号", "物料编码", "物料名称", "供应商信息",
                           "需求数", "剩余未收数", "备注"]):
        _style(ws.cell(row=7, column=c0 + i, value=h), bold=True, fill=True)
    # 数据行从固定第八行开始，序号仅在当前批次内递增。
    for i, (code, name, supp, dem, remain) in enumerate(materials):
        r = DATA_START_ROW + i
        for j, v in enumerate([i + 1, code, name, supp, dem, remain, None]):
            _style(ws.cell(row=r, column=c0 + j, value=v))
    for i, w in enumerate(COL_WIDTHS):
        ws.column_dimensions[get_column_letter(c0 + i)].width = w
    return diff, arrived

def build_workbook(batches, top_label, out_path):
    """把多个批次横向写入标准工作簿并保存到指定路径。

    ``batches`` 每项包含批次号、未收物料、主料总类数和备注。返回旧协议所需的
    ``(批次号, 缺料类数, 已到类数, 总类数)`` 元组列表，供双端统一结果投影使用。
    """
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Sheet1"  # 成品解析兼容默认页名，但不依赖它进行识别。
    results = []
    col = FIRST_COL
    for b in batches:
        diff, arrived = _write_batch(ws, col, b["batch_no"], b["materials"],
                                     b["total"], b.get("remark", ""), top_label,
                                     b.get("pending", 0))
        results.append((b["batch_no"], diff, arrived, b["total"]))
        col += BATCH_STRIDE  # 七个业务列后留两列间隔，避免相邻批次视觉粘连。
    # 与 purchase/delivery 一致: 目标被 Excel 占用时给出友好提示
    try:
        wb.save(out_path)
    except PermissionError:
        # Windows 上最常见原因是用户正在 Excel 中打开同名输出文件。
        raise PermissionError("无法保存 %s —— 请先在 Excel 里关闭该文件后重试" % out_path)
    return results


# 以下入口同时供 Tauri 与 Web 桥接调用，输出目录必须由参数或统一 paths 系统解析。
def build_batches(rows_data, top_label, log=None, resolver=None, fill_counts=None):
    """把复核行整理成工作簿批次，同时读取未收明细和补充主数据。

    ``rows_data`` 每项包含文件路径、批次号、总类数、备注和是否纳入。返回的 ``batches``
    供工作簿及结构化结果使用，``mem`` 只保存有批次号记录的总类数和备注，供下次人工
    复核预填。传入解析器时只补名称和供应商空值，不覆盖源表内容。
    """
    batches, mem = [], {}
    for row in rows_data:
        if not row.get("include", True):
            # 人工复核取消勾选的文件不读取、不学习，也不写入批次记忆。
            continue
        inspection = inspect_plan(row["path"], log=log)
        materials = inspection["materials"]
        if resolver is not None:
            for material in materials:
                additions = resolver.complete_material(
                    material[0], {"name": material[1], "supplier": material[2]},
                    fields=("name", "supplier"), counts=fill_counts)
                if "name" in additions:
                    material[1] = additions["name"]
                if "supplier" in additions:
                    material[2] = additions["supplier"]
        bn = row.get("batch_no") or detect_batch(row["path"])  # 人工输入优先于自动识别。
        # 界面留空时 total 可能为 None/""/带小数或千分位的文本; int("566.0")、
        # int("5,66") 都会抛 ValueError, 统一 float() 兜一层再取整, 失败回退默认值
        auto_total = int(inspection.get("total", 0) or 0)
        tv = row.get("total")
        try:
            # 人工值优先；未填写时使用本次完整源表识别值，只有空表才回退兼容默认数。
            total = (auto_total or DEFAULT_TOTAL) if tv in (None, "") else int(float(str(tv).replace(",", "")))
        except (ValueError, TypeError):
            total = auto_total or DEFAULT_TOTAL
        remark = row.get("remark", "")
        batches.append({"batch_no": bn, "materials": materials,
                        "total": total, "auto_total": auto_total, "remark": remark,
                        "pending": inspection.get("pending", 0)})
        if bn:
            mem[bn] = {"total": total, "remark": remark}
    return batches, mem


def _plain_result_value(value):
    """把 Excel 单元格值收敛为可安全写入 JSON 与前端展示的简单类型。

    日期时间使用 ISO 文本，常规标量原样保留，其他 openpyxl 或自定义对象回退字符串，
    避免结构化任务结果因不可序列化对象而失败。
    """
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (datetime.date, datetime.datetime, datetime.time)):
        return value.isoformat()
    return str(value)


def _received_quantity(demand, remain):
    """需求数和正缺口均为数字时计算已收数，否则留空等待人工核对。"""
    if isinstance(demand, bool) or isinstance(remain, bool):
        return ""
    try:
        value = float(demand) - float(remain)
    except (TypeError, ValueError):
        return ""
    return int(value) if value.is_integer() else round(value, 4)  # 消除无意义小数尾数并限制浮点噪声。


def _shortage_quantity(value):
    """把正负两种“剩余未收数”统一为正缺口数量。

    无法解析的原始文本仍保留在结果中，便于人工发现异常；布尔值和空值不参与计算。
    """
    if value in (None, "") or isinstance(value, bool):
        return ""
    try:
        number = abs(float(str(value).replace(",", "").strip()))
    except (TypeError, ValueError):
        return _plain_result_value(value)
    return int(number) if number.is_integer() else round(number, 4)


def build_result_batches(batches, results):
    """把内部批次与旧汇总元组转换为统一结构化结果。

    ``zip`` 按同一生成顺序配对批次和汇总；每条物料补齐到五项，避免历史记录字段不足
    引发索引异常。该结构是日清看板和双端结果预览读取到料数据的共同协议。
    """
    detail_batches = []
    for batch, summary in zip(batches, results):
        batch_no, missing_count, arrived_count, total_count = summary
        materials = []
        for item in batch.get("materials", []):
            values = list(item) + [None] * 5  # 历史明细不足五列时用空值安全补齐。
            demand = _plain_result_value(values[3])
            shortage = _shortage_quantity(values[4])
            materials.append({
                "material_code": _plain_result_value(values[0]),
                "material_name": _plain_result_value(values[1]),
                "supplier": _plain_result_value(values[2]),
                "demand_quantity": demand,
                "received_quantity": _received_quantity(demand, shortage),
                "shortage_quantity": shortage,
            })
        detail_batches.append({
            "batch_no": batch_no,
            "missing_count": missing_count,
            "arrived_count": arrived_count,
            "total_count": total_count,
            "missing_materials": materials,
        })
    return detail_batches


def _number(value):
    """解析允许千分位的普通数字，空值、布尔值和非法文本返回 ``None``。"""
    if value in (None, "") or isinstance(value, bool):
        return None
    try:
        return float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return None


def _integer(value):
    """只接受数学意义上的整数，带有效小数的数字返回 ``None``。"""
    number = _number(value)
    return int(number) if number is not None and number.is_integer() else None


def _value_right_of_label(ws, row, column, max_offset=4):
    """读取标签右侧有限范围内的第一个整数。

    成品模板的标签通常横向合并两格，真实数值不一定紧邻标签单元格，因此最多向右扫描
    四列；范围受限可避免误读相邻批次区块。
    """
    for offset in range(1, max_offset + 1):
        value = ws.cell(row=row, column=column + offset).value
        number = _integer(value)
        if number is not None:
            return number
    return None


def _finished_columns(ws, total_row, block_column):
    """在成品批次区块内定位表头行及明细字段列。

    扫描范围限制为总类数标签下方九行和当前区块八列，防止横向相邻批次互相串列。
    编码、需求和剩余未收仍是必需列，名称与供应商允许缺失。
    """
    for row in range(total_row + 1, min(ws.max_row, total_row + 9) + 1):
        columns = _classify_header_cells(
            _header_cells(ws, row, block_column, block_column + 7)
        )
        if _has_required_arrival_columns(columns):
            columns["header_row"] = row
            return columns
    return None


def _finished_materials(ws, columns):
    """读取一个成品批次区块中的未到物料明细。

    数据开始后连续三个空编码行视为区块结束，用于容忍明细间单个空行，同时避免继续
    扫到工作表底部或其他说明区。剩余未收数统一转换为正缺口。
    """
    rows = []
    empty_run = 0
    started = False
    for row in range(columns["header_row"] + 1, ws.max_row + 1):
        code = ws.cell(row=row, column=columns["code"]).value
        if code in (None, ""):
            if started:
                empty_run += 1
                if empty_run >= 3:
                    # 三个连续空行是模板区块结束标记，单个空行仍允许继续读取。
                    break
            continue
        started = True
        empty_run = 0
        shortage = _shortage_quantity(ws.cell(row=row, column=columns["remain"]).value)
        rows.append([
            str(code).strip(),
            ws.cell(row=row, column=columns["name"]).value if columns["name"] else None,
            ws.cell(row=row, column=columns["supplier"]).value if columns["supplier"] else None,
            ws.cell(row=row, column=columns["demand"]).value,
            shortage,
        ])
    return rows


def _finished_metrics(ws, total_row, block_column):
    """读取批次区块的到货类数与差异类数，缺失项返回 ``None``。"""
    arrived = missing = None
    for row in range(total_row + 1, min(ws.max_row, total_row + 4) + 1):
        label = _norm(ws.cell(row=row, column=block_column).value)
        if label in FINISHED_ARRIVED_LABELS:
            arrived = _value_right_of_label(ws, row, block_column)
        elif label in FINISHED_MISSING_LABELS:
            missing = _value_right_of_label(ws, row, block_column)
    return arrived, missing


def _finished_batch_from_anchor(ws, total_row, block_column):
    """把一个“主料总类数”锚点解析为完整批次；说明文本锚点返回 ``None``。"""
    total = _value_right_of_label(ws, total_row, block_column)
    columns = _finished_columns(ws, total_row, block_column)
    if total is None or columns is None:
        return None
    batch_no = _norm(ws.cell(row=max(1, total_row - 1), column=block_column).value)
    if not batch_no:
        raise ValueError(f"《{ws.title}》第 {block_column} 列的到料区块缺少批次号")
    materials = _finished_materials(ws, columns)
    arrived, missing = _finished_metrics(ws, total_row, block_column)
    missing = len(materials) if missing is None else missing
    arrived = total - missing if arrived is None else arrived
    return {
        "batch_no": batch_no,
        "materials": materials,
        "total": total,
        "missing": missing,
        "arrived": arrived,
    }


def _finished_batches_from_sheet(ws):
    """按工作表顺序产出前二十行内可识别的全部成品到料批次。"""
    for row in range(1, min(ws.max_row, 20) + 1):
        for column in range(1, ws.max_column + 1):
            if _norm(ws.cell(row=row, column=column).value) not in FINISHED_TOTAL_LABELS:
                continue
            batch = _finished_batch_from_anchor(ws, row, column)
            if batch is not None:
                yield batch


def analyze_finished_report(path, log=None):
    """解析已经制作完成的《每日主料到料明细》，直接供日清看板展示。

    与 :func:`analyze` 的职责不同：这里不读取送货计划、不要求人工填写批次号或主料
    总类数，也不生成新文件。它扫描所有工作表的前二十行寻找批次区块，指标以成品表
    显示值为准；缺失的到货或差异可由总类数与明细行数推导，并通过数量闭合提示人工。
    """
    log = log or (lambda *a, **k: None)
    wb = _common.load_data_only(path)
    batches = []
    try:
        for ws in wb.worksheets:
            batches.extend(_finished_batches_from_sheet(ws))
    finally:
        wb.close()
    if not batches:
        raise ValueError(
            "未识别到成品每日到料明细，请上传由“每日到料”业务生成的报表"
        )
    results = [
        (batch["batch_no"], batch["missing"], batch["arrived"], batch["total"])
        for batch in batches
    ]
    for batch in batches:
        # 差异类数与可读取明细行数不一致时保留两者，记录提示而不擅自改表内指标。
        if len(batch["materials"]) != batch["missing"]:
            log(
                "⚠ 批次 %s 的差异为 %d 类，但明细读取到 %d 行，请人工核对。"
                % (batch["batch_no"], batch["missing"], len(batch["materials"]))
            )
    return {
        "results": results,
        "batches": build_result_batches(batches, results),
        "source_count": 1,
        "report_date": detect_report_date(path),
        "source_type": "finished_report",
    }


def analyze(rows_data, top_label=None, log=None):
    """只解析到料源文件并返回结构化结果，不生成输出工作簿。

    供不需要输出文件的结构化预览和资料处理使用；批次识别、缺料提取与主数据库补全
    逻辑与正式到料业务保持一致，避免 Web 看板形成第二套口径。该函数不保存批次记忆。
    """
    log = log or (lambda *a, **k: None)
    current = _settings.get_settings()
    if top_label is None:
        top_label = current.arrival.get("top_label", DEFAULT_TOP_LABEL)
    resolver = material_catalog.CatalogResolver()
    fill_counts: dict[str, int] = {}
    batches, _memory = build_batches(
        rows_data, top_label, log=log, resolver=resolver, fill_counts=fill_counts,
    )
    material_catalog.log_fill_summary(log, "到料明细", fill_counts)
    if not batches:
        raise ValueError("没有可处理的到料文件")
    results = []
    for batch in batches:
        missing = len(batch.get("materials", []))
        total = int(batch.get("total", 0) or 0)
        results.append((batch.get("batch_no") or "", missing, total - missing, total))
    return {
        "results": results,
        "batches": build_result_batches(batches, results),
        "source_count": len(batches),
    }


def run(rows_data, top_label=None, out_dir=None, log=None, progress=None):
    """生成标准每日到料工作簿并返回结构化结果。

    ``rows_data`` 结构见 :func:`build_batches`。输出目录未指定时使用统一业务目录；进度
    分为整理批次和写入工作簿两段。成功生成文件后再保存截止标签和批次记忆，配置保存
    失败只记录提示，不把已经成功的业务输出改判为失败。
    """
    log = log or (lambda *a, **k: None)
    st = _settings.get_settings()
    # 工作簿写盘和样式处理更重，因此分配六成进度；阶段权重不代表业务数据比例。
    prog = _common.Progress(progress, stages=[("batch", 40), ("build", 60)])
    if top_label is None:
        top_label = st.arrival.get("top_label", DEFAULT_TOP_LABEL)
    if out_dir is None:
        out_dir = _paths.resolve_output_dir("arrival", **st.output_kwargs())
    prog.stage("batch")
    log("整理批次数据（共 %d 个计划表）..." % len(rows_data))
    resolver = material_catalog.CatalogResolver()
    fill_counts: dict[str, int] = {}
    batches, mem = build_batches(
        rows_data, top_label, log=log, resolver=resolver, fill_counts=fill_counts)
    material_catalog.log_fill_summary(log, "到料明细", fill_counts)
    prog.stage("build")
    fname = "%s每日主料到料明细.xlsx" % beijing_date()
    out_file = _common.unique_path(os.path.join(out_dir, fname))  # 同日重复运行追加序号，不覆盖历史结果。
    results = build_workbook(batches, top_label, out_file)
    for bn, diff, arrived, total in results:
        log("  · 批次 %s：未收料 %d 类，到货 %d，主料 %d 类" % (bn, diff, arrived, total))
    # 只有工作簿保存成功后才更新记忆，避免失败任务把未实际输出的参数设为下次默认值。
    st.arrival["top_label"] = top_label
    st.arrival.setdefault("batches", {}).update(mem)
    if batches:
        st.arrival["last_total"] = batches[-1]["total"]
    if not st.save():
        log("⚠ 到料明细已生成，但批次记忆保存失败，请检查配置目录权限。")
    prog.done()
    log("已保存：%s" % out_file)
    return {
        "out_file": out_file,
        "out_dir": out_dir,
        "results": results,
        "batches": build_result_batches(batches, results),
    }
