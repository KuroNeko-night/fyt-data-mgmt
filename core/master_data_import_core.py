# -*- coding: utf-8 -*-
"""主数据库表格学习：上传批次、字段识别、冲突治理与安全合并。

管理员上传的表格不会直接改写正式主数据。本模块先把可学习的对应关系保存为独立批次，
标记表内冲突和与正式库的冲突，等待管理员逐条决策并确认；合并时再次校验正式库是否
发生变化。批次 JSON 使用文件锁和原子替换，避免并发请求留下半写文件。
"""
from __future__ import annotations

import datetime
import hashlib
import json
import os
import re
import tempfile
import uuid
from collections import defaultdict
from collections.abc import Callable, Iterable, Iterator
from pathlib import Path

from openpyxl import Workbook, load_workbook

from . import common_core, material_catalog, paths, storage_lock


SCHEMA_VERSION = 1
MAX_HEADER_ROWS = 40
MAX_SCAN_COLUMNS = 100
SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}

STATUS_LABELS = {
    "needs_review": "存在冲突",
    "ready_to_confirm": "等待确认",
    "ready": "等待合并",
    "merged": "已合并",
    "rejected": "已拒绝",
    "failed": "分析失败",
}

RELATION_TITLES = {
    "supplier_code": "供应商编码",
    "material_name": "材料名称",
    "material_spec": "材料规格",
    "material_unit": "材料单位",
    "material_supplier": "材料供应商",
}

# 字段识别只依赖列名语义，不绑定具体业务报表。新模块若使用新的同义表头，只需在此
# 补充别名，后续的工作表扫描、关系生成和冲突治理流程无需跟着复制分支。
FIELD_ALIASES = {
    "material_code": {
        "材料编号", "物料编号", "物料编码", "材料编码", "料号", "存货编码",
        "产品编码", "产品编号", "零件号", "编码",
    },
    "material_name": {
        "材料名称", "物料名称", "物料描述", "材料描述", "品名", "产品名称",
        "零件名称", "名称",
    },
    "material_spec": {
        "规格", "规格型号", "型号规格", "材料规格", "物料规格", "型号",
    },
    "material_unit": {"单位", "计量单位", "基本单位", "采购单位"},
    "supplier_name": {
        "供应商", "供应商名称", "供应商信息", "供方", "供方名称", "厂商",
        "厂商名称", "供应单位",
    },
    "supplier_code": {
        "供应商编码", "供应商代码", "供方编码", "供方代码", "厂商编码",
        "厂商代码", "供应商编号",
    },
}


class DuplicateImportError(ValueError):
    """表示同一文件内容已经存在仍有效的导入批次。

    异常携带原批次编号，服务端可以引导管理员回到原批次继续处理，而不是重复生成一组
    完全相同的冲突记录。
    """

    def __init__(self, batch_id: str):
        """构造异常并保存可供 API 返回的既有批次编号。"""
        super().__init__("同一表格已经上传，请在现有批次中继续处理")
        self.batch_id = batch_id


def _now_iso() -> str:
    """返回本机当前时间的秒级 ISO 文本，供批次审计字段使用。"""
    return datetime.datetime.now().isoformat(timespec="seconds")


def import_root(root: str | None = None) -> str:
    """解析主数据导入暂存根目录。

    测试可通过参数使用临时目录，部署环境可由变量覆盖；都未提供时写入应用数据目录，
    从而避免把运行数据混进源码目录。
    """
    if root:
        return os.path.abspath(root)
    override = os.environ.get("FYT_MASTER_DATA_IMPORT_ROOT", "").strip()
    if override:
        return os.path.abspath(override)
    return os.path.join(paths.app_data_dir(), "主数据导入")


def _batches_dir(root: str | None = None) -> str:
    """返回批次 JSON 目录并确保目录已经存在。"""
    target = os.path.join(import_root(root), "batches")
    os.makedirs(target, exist_ok=True)
    return target


def _batch_path(batch_id: str, root: str | None = None) -> str:
    """校验批次编号并生成对应 JSON 路径。

    只接受 UUID4 去掉连字符后的 32 位十六进制文本，既约束内部协议，也防止批次编号被
    当作相对路径实施目录穿越。
    """
    normalized = str(batch_id).strip().lower()
    if not re.fullmatch(r"[0-9a-f]{32}", normalized):
        raise ValueError("主数据导入批次编号无效")
    return os.path.join(_batches_dir(root), normalized + ".json")


def _root_guard(root: str | None = None) -> str:
    """返回导入根目录的全局锁标识路径。

    该路径不承载业务内容，只让“查重并创建批次”成为一个临界区，避免两个并发上传在
    都未看见对方文件时重复创建相同 SHA-256 批次。
    """
    target = import_root(root)
    os.makedirs(target, exist_ok=True)
    return os.path.join(target, "批次索引")


def _atomic_write_json(path: str, data: dict[str, object]) -> None:
    """在目标目录内写临时文件，并用原子替换提交 JSON。

    临时文件与目标文件位于同一目录，保证 ``os.replace`` 不跨文件系统；进程在写入中途
    退出时，旧批次仍保持完整。最终清理用于处理替换前失败留下的临时文件。
    """
    parent = os.path.dirname(path)
    os.makedirs(parent, exist_ok=True)
    descriptor, temp_path = tempfile.mkstemp(
        prefix=os.path.basename(path) + ".", suffix=".tmp", dir=parent)
    try:
        with os.fdopen(descriptor, "w", encoding="utf-8", newline="") as handle:
            json.dump(data, handle, ensure_ascii=False, indent=1)
        os.replace(temp_path, path)  # 同卷原子替换，读者只会看到旧版或完整新版。
    finally:
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except OSError:
                # 清理失败不应覆盖真正的写入异常；残留临时文件不会被批次扫描读取。
                pass


def _read_json(path: str, *, strict: bool = True) -> dict[str, object]:
    """读取批次 JSON，并按调用场景选择严格或容错模式。

    获取单个批次时使用严格模式，把损坏明确报告给管理员；批量列表使用容错模式跳过
    单个损坏文件，保证其他批次仍可管理。
    """
    try:
        with open(path, "r", encoding="utf-8") as handle:
            data = json.load(handle)
    except (OSError, ValueError) as exc:
        if strict:
            raise ValueError("主数据导入批次损坏或无法读取") from exc
        return {}
    if not isinstance(data, dict):
        if strict:
            raise ValueError("主数据导入批次格式无效")
        return {}
    return data


def _file_sha256(path: str) -> str:
    """以流式方式计算文件内容摘要，供重复上传检测使用。"""
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        # 一兆字节分块避免大型主数据表一次性读入内存。
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _text(value: object) -> str:
    """规范化表格值，使相同业务键不会因格式差异被拆成多条关系。

    Excel 整数经 openpyxl 读取后可能表现为浮点数，先去除无意义的小数点；全角空格和
    连续空白也统一压缩，但不改变字母大小写或业务符号。
    """
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).replace("\u3000", " ").strip()
    return re.sub(r"\s+", " ", text)


def _header(value: object) -> str:
    """生成用于表头别名匹配的宽松标准形式。"""
    text = _text(value).lower()
    # 表头中的空格、横线、括号和分隔符通常只影响排版，不应影响字段语义识别。
    return re.sub(r"[\s_\-—－:：()（）\[\]【】/\\]+", "", text)


NORMALIZED_ALIASES = {
    field: {_header(alias) for alias in aliases}
    for field, aliases in FIELD_ALIASES.items()
}


def _field_for_header(value: object) -> str:
    """返回表头对应的内部字段名，无法识别时返回空字符串。"""
    normalized = _header(value)
    if not normalized:
        return ""
    for field, aliases in NORMALIZED_ALIASES.items():
        if normalized in aliases:
            return field
    return ""


def _sheet_layout(rows: list[list[object]]) -> dict[str, object] | None:
    """在工作表前若干行中选择最可能的表头布局。

    材料表必须包含材料编码，随后按名称、规格、单位、供应商等辅助列加分；供应商代码表
    必须同时含供应商名称和编码。额外识别字段只提供有限加分，避免一行普通数据因偶然
    命中多个短标题而超过真正表头。
    """
    best: dict[str, object] | None = None
    for row_index, values in enumerate(rows, start=1):
        columns: dict[str, int] = {}
        for column_index, value in enumerate(values[:MAX_SCAN_COLUMNS]):
            field = _field_for_header(value)
            if field and field not in columns:
                # 同一语义出现多列时取最左列，保持读取规则确定且便于管理员核对源表。
                columns[field] = column_index
        material_score = 0
        if "material_code" in columns:
            material_score = 4 + sum(
                1 for field in ("material_name", "material_spec", "material_unit", "supplier_name")
                if field in columns
            )
        supplier_score = 6 if {"supplier_name", "supplier_code"}.issubset(columns) else 0
        score = max(material_score, supplier_score) + min(len(columns), 4)
        if score and (best is None or score > int(best["score"])):
            # 同分时保留更靠前的行，符合常见表格标题在上、数据在下的结构。
            best = {"header_row": row_index, "columns": columns, "score": score}
    return best


def _iter_openpyxl(path: str) -> Iterator[tuple[str, list[list[object]], Iterable[tuple[int, list[object]]]]]:
    """以只读模式逐工作表提供表头预览和完整行迭代器。

    前 ``MAX_HEADER_ROWS`` 行先缓存用于识别布局，随后通过内部生成器把缓存行与剩余流
    重新拼接，调用方仍能从第一行顺序扫描而无需再次打开工作簿。
    """
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for worksheet in workbook.worksheets:
            iterator = worksheet.iter_rows(values_only=True)
            preview: list[list[object]] = []
            for _ in range(MAX_HEADER_ROWS):
                try:
                    preview.append(list(next(iterator)))
                except StopIteration:
                    break

            def remaining(
                initial: list[list[object]] = preview,
                tail: Iterator[tuple[object, ...]] = iterator,
            ) -> Iterator[tuple[int, list[object]]]:
                """把已消费的预览行与工作表剩余行连接成单次迭代流。"""
                for index, values in enumerate(initial, start=1):
                    yield index, values
                for index, values in enumerate(tail, start=len(initial) + 1):
                    yield index, list(values)

            yield worksheet.title, preview, remaining()
    finally:
        workbook.close()


def _iter_xls(path: str) -> Iterator[tuple[str, list[list[object]], Iterable[tuple[int, list[object]]]]]:
    """使用 xlrd 读取旧版 ``.xls``，并提供与 openpyxl 路径一致的迭代协议。"""
    # xlrd 仅在确实处理旧格式时延迟导入，普通 xlsx 部署不必为启动加载该模块。
    import xlrd

    workbook = xlrd.open_workbook(path, on_demand=True)
    try:
        for worksheet in workbook.sheets():
            preview = [worksheet.row_values(index) for index in range(min(MAX_HEADER_ROWS, worksheet.nrows))]

            def rows(sheet=worksheet) -> Iterator[tuple[int, list[object]]]:
                """逐行返回从一开始的业务行号和单元格值。"""
                for index in range(sheet.nrows):
                    yield index + 1, sheet.row_values(index)

            yield worksheet.name, preview, rows()
    finally:
        workbook.release_resources()


def _iter_sheets(path: str):
    """按文件扩展名选择旧版或现代 Excel 读取器。"""
    return _iter_xls(path) if Path(path).suffix.lower() == ".xls" else _iter_openpyxl(path)


def _candidate_id(relation_type: str, key: str) -> str:
    """为关系类型与键生成稳定候选编号。

    编号用于前端提交冲突决策；同一批次重新读取时必须保持稳定，因此不能使用随机 UUID。
    空字符分隔可防止不同字段拼接后产生边界歧义。
    """
    raw = (relation_type + "\0" + key).encode("utf-8")
    return hashlib.sha1(raw).hexdigest()[:20]


def _source(sheet: str, row: int, value: str) -> dict[str, object]:
    """记录候选值在上传工作簿中的可追溯位置。"""
    return {"sheet": sheet, "row": row, "value": value}


def _append_relation(
    relations: dict[tuple[str, str], dict[str, list[dict[str, object]]]],
    relation_type: str,
    key: object,
    value: object,
    sheet: str,
    row: int,
) -> bool:
    """向关系池加入一条非空键值，并记录其工作表来源。

    返回值表示当前源行是否贡献了有效关系，供分析统计“有效数据行数”。相同值可以有多
    个来源，这些来源不会被去重，因为出现次数和位置是管理员判断候选可信度的依据。
    """
    normalized_key = _text(key)
    normalized_value = _text(value)
    if not normalized_key or not normalized_value:
        return False
    values = relations[(relation_type, normalized_key)]
    values.setdefault(normalized_value, []).append(_source(sheet, row, normalized_value))
    return True


def _sheet_cell(values: list[object], columns: dict[str, int], field: str) -> object:
    """按已识别列读取当前行字段；短行和缺列统一视为空值。"""

    index = columns.get(field)
    return values[index] if isinstance(index, int) and index < len(values) else None


def _collect_sheet_relations(
    sheet_name: str,
    layout: dict[str, object],
    rows: Iterable[tuple[int, list[object]]],
    relations: dict[tuple[str, str], dict[str, list[dict[str, object]]]],
) -> tuple[int, dict[str, object]]:
    """扫描一个已识别页签并写入关系池，返回有效行数和页签审计摘要。"""

    columns = layout["columns"]
    header_row = int(layout["header_row"])
    sheet_rows = 0
    for row_number, values in rows:
        if row_number <= header_row:
            continue

        used = False
        material_code = _sheet_cell(values, columns, "material_code")
        if _text(material_code):
            # 同一物料的名称、规格、单位和供应商关系分别记录，缺一列不影响其他关系学习。
            used |= _append_relation(
                relations, "material_name", material_code,
                _sheet_cell(values, columns, "material_name"), sheet_name, row_number,
            )
            used |= _append_relation(
                relations, "material_spec", material_code,
                _sheet_cell(values, columns, "material_spec"), sheet_name, row_number,
            )
            used |= _append_relation(
                relations, "material_unit", material_code,
                _sheet_cell(values, columns, "material_unit"), sheet_name, row_number,
            )
            used |= _append_relation(
                relations, "material_supplier", material_code,
                _sheet_cell(values, columns, "supplier_name"), sheet_name, row_number,
            )
        # 供应商代码关系独立于材料关系，供应商表只需提供名称和编码即可被学习。
        used |= _append_relation(
            relations,
            "supplier_code",
            _sheet_cell(values, columns, "supplier_name"),
            _sheet_cell(values, columns, "supplier_code"),
            sheet_name,
            row_number,
        )
        if used:
            sheet_rows += 1
    return sheet_rows, {
        "sheet": sheet_name,
        "header_row": header_row,
        "fields": sorted(columns),
        "recognized_rows": sheet_rows,
    }


def _build_candidates(
    relations: dict[tuple[str, str], dict[str, list[dict[str, object]]]],
    catalog: dict[str, object],
) -> list[dict[str, object]]:
    """根据关系池和主数据库快照生成稳定候选，集中处理冲突判定规则。"""

    candidates = []
    for (relation_type, key), value_sources in sorted(relations.items()):
        values = sorted(value_sources)
        current = material_catalog.relation_value(catalog, relation_type, key)
        reasons = []
        if len(values) > 1:
            reasons.append("同一表格中存在多个不同值")
        if current and any(value != current for value in values):
            reasons.append("与正式主数据库当前值不同")
        conflict = bool(reasons)
        # 唯一且未与正式库冲突的关系可以预选；有冲突的关系必须经过管理员决策。
        selected_value = values[0] if len(values) == 1 and not conflict else ""
        candidates.append({
            "id": _candidate_id(relation_type, key),
            "relation_type": relation_type,
            "relation_title": RELATION_TITLES[relation_type],
            "key": key,
            "values": [
                {"value": value, "count": len(value_sources[value]), "sources": value_sources[value]}
                for value in values
            ],
            "current_value": current,
            "expected_current_value": current,
            "conflict": conflict,
            "conflict_reasons": reasons,
            "selected_value": selected_value,
            "decision": None,
        })
    return candidates


def _analyze_workbook(path: str, log: Callable[[str], None] | None = None) -> dict[str, object]:
    """扫描上传工作簿，生成关系候选、冲突原因和可追溯来源。

    公式缓存警告、未识别工作表和有效行统计会随批次保存。分析阶段只读取正式主数据做
    比较，不执行写入；所有真正变更必须经过后续冲突决策、确认和合并流程。
    """
    warnings: list[str] = []
    if Path(path).suffix.lower() != ".xls":
        # openpyxl 的 data_only 模式依赖 Excel 缓存公式值；旧 xls 走 xlrd，不使用该检查。
        common_core.warn_if_uncached(path, warnings.append, what="主数据字段")
    relations: dict[tuple[str, str], dict[str, list[dict[str, object]]]] = defaultdict(dict)
    recognized_sheets = []
    unrecognized_sheets = []
    recognized_rows = 0

    for sheet_name, preview, rows in _iter_sheets(path):
        layout = _sheet_layout(preview)
        if not layout:
            # 未识别页签仍写入分析结果，让管理员知道它被跳过，而不是悄悄忽略。
            unrecognized_sheets.append({
                "sheet": sheet_name,
                "reason": "未找到可识别的材料编号或供应商编码表头",
            })
            continue
        sheet_rows, sheet_summary = _collect_sheet_relations(
            sheet_name, layout, rows, relations,
        )
        recognized_rows += sheet_rows
        recognized_sheets.append(sheet_summary)

    catalog = material_catalog.load()  # 仅作为当前值快照，分析阶段不写正式库。
    candidates = _build_candidates(relations, catalog)

    if log:
        log("已识别 %d 个工作表、%d 行有效数据、%d 条对应关系。" % (
            len(recognized_sheets), recognized_rows, len(candidates)))
    return {
        "candidates": candidates,
        "recognized_sheets": recognized_sheets,
        "unrecognized_sheets": unrecognized_sheets,
        "recognized_rows": recognized_rows,
        "warnings": warnings,
    }


def _summary(batch: dict[str, object]) -> dict[str, object]:
    """生成批次列表需要的轻量摘要，避免传输全部来源和候选明细。"""
    candidates = batch.get("candidates") if isinstance(batch.get("candidates"), list) else []
    conflicts = [item for item in candidates if isinstance(item, dict) and item.get("conflict")]
    unresolved = [item for item in conflicts if not item.get("decision")]
    return {
        "id": batch.get("id"),
        "original_name": batch.get("original_name"),
        "size": batch.get("size", 0),
        "status": batch.get("status"),
        "status_label": STATUS_LABELS.get(str(batch.get("status")), "未知状态"),
        "created_at": batch.get("created_at"),
        "uploader_id": batch.get("uploader_id"),
        "uploader_name": batch.get("uploader_name"),
        "candidate_count": len(candidates),
        "conflict_count": len(conflicts),
        "unresolved_conflict_count": len(unresolved),
        "recognized_rows": batch.get("recognized_rows", 0),
        "recognized_sheet_count": len(batch.get("recognized_sheets") or []),
        "unrecognized_sheet_count": len(batch.get("unrecognized_sheets") or []),
        "confirmed_at": batch.get("confirmed_at", ""),
        "merged_at": batch.get("merged_at", ""),
        "merge_summary": batch.get("merge_summary") or {},
    }


def _public(batch: dict[str, object], *, detail: bool) -> dict[str, object]:
    """生成允许 API 返回的批次结构，并排除源文件绝对路径和摘要值。"""
    result = _summary(batch)
    if detail:
        for field in (
            "recognized_sheets", "unrecognized_sheets", "warnings", "candidates",
            "confirmed_by_name", "merged_by_name", "last_error",
        ):
            result[field] = batch.get(field) or ([] if field.endswith("s") else "")
    return result


def _all_batches(root: str | None = None) -> list[dict[str, object]]:
    """容错读取全部批次，并按创建时间从新到旧排列。"""
    rows = []
    for path in Path(_batches_dir(root)).glob("*.json"):
        data = _read_json(str(path), strict=False)
        if data:
            rows.append(data)
    rows.sort(key=lambda item: str(item.get("created_at") or ""), reverse=True)
    return rows


def list_batches(root: str | None = None) -> dict[str, object]:
    """返回批次列表及各待办状态数量。"""
    items = [_summary(batch) for batch in _all_batches(root)]
    return {
        "items": items,
        "summary": {
            "total": len(items),
            "needs_review": sum(item["status"] == "needs_review" for item in items),
            "ready_to_confirm": sum(item["status"] == "ready_to_confirm" for item in items),
            "ready": sum(item["status"] == "ready" for item in items),
            "merged": sum(item["status"] == "merged" for item in items),
        },
    }


def get_batch(batch_id: str, root: str | None = None) -> dict[str, object]:
    """严格读取单个批次的公开详情。"""
    return _public(_read_json(_batch_path(batch_id, root)), detail=True)


def analyze(
    source_path: str,
    *,
    original_name: str,
    uploader_id: int,
    uploader_name: str,
    batch_id: str | None = None,
    root: str | None = None,
    log: Callable[[str], None] | None = None,
) -> dict[str, object]:
    """分析管理员上传表格并保存暂存批次，不修改正式主数据库。

    文件内容摘要用于全局查重；批次写入与查重共用根目录锁，防止并发重复创建。只有
    ``rejected`` 或 ``failed`` 的旧批次不会阻止重新上传，因为它们已经退出有效流程。
    """
    path = os.path.abspath(source_path)
    if not os.path.isfile(path):
        raise FileNotFoundError("上传表格不存在")
    # 优先校验用户上传时的原始文件名，避免临时保存路径丢失真实扩展名。
    extension = Path(original_name or path).suffix.lower()
    if extension not in SUPPORTED_EXTENSIONS:
        raise ValueError("仅支持 .xlsx、.xlsm 和 .xls 表格")
    digest = _file_sha256(path)
    analysis = _analyze_workbook(path, log=log)
    if not analysis["candidates"]:
        raise ValueError("表格中没有识别到可学习的对应关系，请检查表头和有效数据")

    batch_id = str(batch_id or uuid.uuid4().hex).strip().lower()
    if not re.fullmatch(r"[0-9a-f]{32}", batch_id):
        raise ValueError("主数据导入批次编号无效")
    conflicts = sum(bool(item.get("conflict")) for item in analysis["candidates"])
    batch = {
        "schema_version": SCHEMA_VERSION,
        "id": batch_id,
        "original_name": Path(original_name).name or Path(path).name,
        "source_path": path,
        "sha256": digest,
        "size": os.path.getsize(path),
        "uploader_id": int(uploader_id),
        "uploader_name": _text(uploader_name) or "管理员",
        "created_at": _now_iso(),
        "status": "needs_review" if conflicts else "ready_to_confirm",
        **analysis,
        "confirmed_at": "",
        "confirmed_by_id": None,
        "confirmed_by_name": "",
        "merged_at": "",
        "merged_by_id": None,
        "merged_by_name": "",
        "merge_summary": {},
        "last_error": "",
    }
    # 查重和落盘必须位于同一临界区，否则并发请求可能同时通过查重。
    with storage_lock.file_lock(_root_guard(root)):
        for existing in _all_batches(root):
            if existing.get("sha256") == digest and existing.get("status") not in {"rejected", "failed"}:
                raise DuplicateImportError(str(existing.get("id") or ""))
        _atomic_write_json(_batch_path(batch_id, root), batch)
    return _public(batch, detail=True)


def _mutate_batch(
    batch_id: str,
    mutator: Callable[[dict[str, object]], object],
    root: str | None = None,
) -> tuple[dict[str, object], object]:
    """在批次级文件锁内执行读改写，并返回更新后的批次和回调结果。

    冲突处理、确认和拒绝都复用此入口，确保状态校验与写入之间不会被另一请求穿插。
    回调只修改内存对象，统一由此函数原子落盘。
    """
    path = _batch_path(batch_id, root)
    with storage_lock.file_lock(path):
        batch = _read_json(path)
        result = mutator(batch)
        _atomic_write_json(path, batch)
        return batch, result


def _select_conflict_value(target: dict, decision: str, value: str) -> str:
    """按管理员决策从冲突候选中选择确定值，并拒绝表外注入。"""
    if decision == "keep_current":
        selected = _text(target.get("current_value"))
        if not selected:
            raise ValueError("正式主数据库当前没有可保留的值")
        return selected
    if decision == "use_candidate":
        allowed_values = [_text(item.get("value")) for item in target.get("values", []) if isinstance(item, dict)]
        # 单一候选可省略 value；多个候选必须显式选择其中一个，不能注入表外内容。
        selected = _text(value) or (allowed_values[0] if len(allowed_values) == 1 else "")
        if selected not in allowed_values:
            raise ValueError("请选择上传表格中提供的候选值")
        return selected
    selected = _text(value)
    if not selected:
        raise ValueError("手动规范值不能为空")
    return selected


def _apply_conflict_decision(
    batch: dict[str, object],
    candidate_id: str,
    decision: str,
    value: str,
    actor_id: int,
    actor_name: str,
) -> None:
    """在锁内校验批次状态、写入审计信息并重新计算未决冲突数。"""
    if batch.get("status") in {"ready", "merged", "rejected"}:
        raise ValueError("当前批次状态不允许修改冲突")
    candidates = batch.get("candidates")
    if not isinstance(candidates, list):
        raise ValueError("导入批次候选关系无效")
    target = next((item for item in candidates if isinstance(item, dict) and item.get("id") == candidate_id), None)
    if not target or not target.get("conflict"):
        raise ValueError("没有找到需要处理的冲突")
    selected = _select_conflict_value(target, decision, value)
    target["selected_value"] = selected
    target["decision"] = {
        "type": decision,
        "value": selected,
        "actor_id": int(actor_id),
        "actor_name": _text(actor_name) or "管理员",
        "decided_at": _now_iso(),
    }
    # 决策时重新固定期望旧值，供最终合并执行乐观并发检查。
    target["expected_current_value"] = _text(target.get("current_value"))
    unresolved = any(
        isinstance(item, dict) and item.get("conflict") and not item.get("decision")
        for item in candidates
    )
    batch["status"] = "needs_review" if unresolved else "ready_to_confirm"
    batch["last_error"] = ""


def resolve_conflict(
    batch_id: str,
    candidate_id: str,
    decision: str,
    *,
    value: str = "",
    actor_id: int,
    actor_name: str,
    root: str | None = None,
) -> dict[str, object]:
    """记录管理员对单条冲突关系的决策，并推进批次状态。

    可保留正式库值、采用上传候选、手工填写规范值或忽略该关系。所有冲突完成决策后
    批次只进入“等待确认”，仍不会自动修改正式主数据库。
    """
    allowed = {"keep_current", "use_candidate", "manual", "ignore"}
    if decision not in allowed:
        raise ValueError("不支持的冲突处理方式")

    def mutate(batch: dict[str, object]) -> None:
        """在锁内校验状态、写入审计信息并重新计算未决冲突数。"""
        _apply_conflict_decision(
            batch, candidate_id, decision, value, actor_id, actor_name,
        )

    batch, _ = _mutate_batch(batch_id, mutate, root)
    return _public(batch, detail=True)


def confirm_batch(
    batch_id: str,
    *,
    actor_id: int,
    actor_name: str,
    root: str | None = None,
) -> dict[str, object]:
    """由管理员确认全部候选关系，将批次置为可合并状态。"""
    def mutate(batch: dict[str, object]) -> None:
        """在锁内校验确认前置状态并记录确认人。"""
        if batch.get("status") != "ready_to_confirm":
            raise ValueError("请先处理全部冲突，再确认导入批次")
        batch["status"] = "ready"
        batch["confirmed_at"] = _now_iso()
        batch["confirmed_by_id"] = int(actor_id)
        batch["confirmed_by_name"] = _text(actor_name) or "管理员"
        batch["last_error"] = ""

    batch, _ = _mutate_batch(batch_id, mutate, root)
    return _public(batch, detail=True)


def reject_batch(
    batch_id: str,
    *,
    actor_id: int,
    actor_name: str,
    root: str | None = None,
) -> dict[str, object]:
    """拒绝尚未合并的批次并保留审计记录。"""
    def mutate(batch: dict[str, object]) -> None:
        """在锁内阻止已合并批次回退，并记录拒绝人。"""
        if batch.get("status") == "merged":
            raise ValueError("已合并批次不能拒绝")
        batch["status"] = "rejected"
        batch["rejected_at"] = _now_iso()
        batch["rejected_by_id"] = int(actor_id)
        batch["rejected_by_name"] = _text(actor_name) or "管理员"

    batch, _ = _mutate_batch(batch_id, mutate, root)
    return _public(batch, detail=True)


def _merge_relations(batch: dict[str, object]) -> list[dict[str, str]]:
    """把批次候选转换为正式主数据层接受的关系列表。

    被管理员选择忽略的关系不参与合并；其他关系必须已有确定值。每条关系携带确认时的
    正式库旧值，供 ``material_catalog.apply_relations`` 检测确认后的并发变化。
    """
    relations = []
    for candidate in batch.get("candidates", []):
        if not isinstance(candidate, dict):
            continue
        decision = candidate.get("decision") if isinstance(candidate.get("decision"), dict) else {}
        if decision.get("type") == "ignore":
            continue
        value = _text(candidate.get("selected_value"))
        if not value:
            raise ValueError("导入批次仍有未确定的对应关系")
        relations.append({
            "relation_type": _text(candidate.get("relation_type")),
            "key": _text(candidate.get("key")),
            "value": value,
            "expected_current": _text(candidate.get("expected_current_value")),
        })
    return relations


def merge_batch(
    batch_id: str,
    *,
    actor_id: int | None = None,
    actor_name: str = "系统定期合并",
    root: str | None = None,
) -> dict[str, object]:
    """把已确认批次安全合并进正式主数据库。

    合并具有幂等性：已合并批次再次调用直接返回。正式库若在管理员确认后发生变化，
    主数据层抛出冲突，本函数把受影响候选恢复为待决状态，让管理员基于新值重新确认，
    而不是用旧决定覆盖后来维护的数据。
    """
    path = _batch_path(batch_id, root)
    # 持有批次锁直到正式库合并与批次状态落盘完成，防止同一批次被重复执行。
    with storage_lock.file_lock(path):
        batch = _read_json(path)
        if batch.get("status") == "merged":
            return _public(batch, detail=True)
        if batch.get("status") != "ready":
            raise ValueError("只有已确认的批次可以合并")
        relations = _merge_relations(batch)
        # 正式库写入前由 material_catalog 在批次专属目录保存可恢复备份。
        backup_dir = os.path.join(import_root(root), "backups", str(batch_id))
        try:
            summary = material_catalog.apply_relations(relations, backup_dir=backup_dir)
        except material_catalog.CatalogConflictError as exc:
            # 只重开实际发生并发变化的候选，其余管理员决策保持不变。
            conflicts_by_key = {
                (item["relation_type"], item["key"]): item
                for item in exc.conflicts
            }
            for candidate in batch.get("candidates", []):
                if not isinstance(candidate, dict):
                    continue
                conflict = conflicts_by_key.get((candidate.get("relation_type"), candidate.get("key")))
                if not conflict:
                    continue
                candidate["conflict"] = True
                candidate["current_value"] = conflict["current_value"]
                candidate["expected_current_value"] = conflict["current_value"]
                candidate["conflict_reasons"] = ["确认后正式主数据库发生了变化"]
                candidate["selected_value"] = ""
                candidate["decision"] = None
            batch["status"] = "needs_review"
            batch["last_error"] = "正式主数据已更新，请重新处理新增冲突"
            _atomic_write_json(path, batch)
            return _public(batch, detail=True)
        batch["status"] = "merged"
        batch["merged_at"] = _now_iso()
        batch["merged_by_id"] = int(actor_id) if actor_id is not None else None
        batch["merged_by_name"] = _text(actor_name) or "系统定期合并"
        batch["merge_summary"] = summary
        batch["last_error"] = ""
        _atomic_write_json(path, batch)
        return _public(batch, detail=True)


def merge_ready_batches(
    *, root: str | None = None, limit: int = 5,
) -> dict[str, object]:
    """按创建时间合并有限数量已确认批次，单批失败不阻塞后续批次。

    ``_all_batches`` 返回新到旧，反转后优先处理等待最久的批次。限制单次数量可控制定时
    维护任务的耗时；普通异常收集进结果，由服务端记录和告警，不影响后续独立批次。
    """
    ready = [batch for batch in reversed(_all_batches(root)) if batch.get("status") == "ready"]
    merged = []
    review_required = []
    failed = []
    for batch in ready[:max(1, int(limit))]:  # 即使传入零或负数，也至少尝试处理一个批次。
        batch_id = str(batch.get("id") or "")
        try:
            result = merge_batch(batch_id, root=root)
            if result.get("status") == "merged":
                merged.append(batch_id)
            else:
                review_required.append(batch_id)
        except Exception as exc:
            failed.append({"id": batch_id, "error": str(exc)})
    return {"merged": merged, "review_required": review_required, "failed": failed}


def export_catalog(out_path: str | None = None) -> dict[str, object]:
    """导出当前正式主数据库，便于线下复核与归档。

    导出只读取正式库快照，供应商和材料分别建表并按键排序，使相同数据重复导出时顺序
    稳定。显式路径主要供 Web 隔离任务使用，未提供时写入主数据业务输出目录。
    """
    if out_path is None:
        out_dir = paths.resolve_output_dir("master_data")
        out_path = os.path.join(out_dir, "主数据库.xlsx")
    else:
        out_path = os.path.abspath(out_path)
        os.makedirs(os.path.dirname(out_path), exist_ok=True)
    data = material_catalog.load()
    workbook = Workbook()
    suppliers = workbook.active
    suppliers.title = "供应商代码"
    suppliers.append(["供应商名称", "供应商编码"])
    for name, code in sorted(data["suppliers"].items()):
        suppliers.append([name, code])
    materials = workbook.create_sheet("材料主数据")
    materials.append(["材料编号", "材料名称", "规格", "单位", "供应商"])
    for code, item in sorted(data["materials"].items()):
        materials.append([
            code, item.get("name", ""), item.get("spec", ""),
            item.get("unit", ""), item.get("supplier", ""),
        ])
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.column_dimensions["A"].width = 24
        for column in "BCDE":
            worksheet.column_dimensions[column].width = 20
    workbook.save(out_path)
    # 保存成功后显式关闭工作簿，释放 openpyxl 持有的内存与文件资源。
    workbook.close()
    return {"out_dir": os.path.dirname(out_path), "file": out_path}
