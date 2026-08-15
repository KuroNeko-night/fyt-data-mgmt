# -*- coding: utf-8 -*-
"""我司考勤明细与劳务公司月度工时的识别、填表和对账。

处理先聚合我司“数据来源”中的每人每日实际工时，再填入待对总表；随后识别不同格式的
劳务对账单，对双方名单、总工时和逐日工时进行比较。劳务表中的假、休和空白不参与逐日
比较。输出包括已填写总表、异常汇总和结构化可信度指标。

人工列映射、工作表选择、姓名配对、容差和重复姓名策略通过 ``common_core.Options``
进入同一业务逻辑。预分析只读文件，正式 ``run`` 才会创建输出，源文件不会被覆盖。
"""
import copy
import datetime
import logging
import os
from dataclasses import dataclass

import openpyxl

from . import common_core as cc
from . import reconcile_reporting

# 公共常量/工具统一来自 common_core（保留原内部名作别名，避免改动大量调用点）
TOL = cc.TOL
SKIP_MARKS = cc.SKIP_MARKS
_to_num = cc.to_num
_day_of = cc.day_of
_norm_name = cc.norm_name
_read_sheets = cc.read_sheets

# 报告与可信度实现已迁到独立渲染层；保留原名称兼容既有调用和测试。
COLS = reconcile_reporting.COLS
assess_credibility = reconcile_reporting.assess_credibility
write_summary = reconcile_reporting.write_summary
_write_credibility_sheet = reconcile_reporting._write_credibility_sheet


# ---------------------------------------------------------------------------
# 1) 读取"数据来源"：姓名 / 日期 / 实际工作时间  -> {姓名:{日:工时}}
# ---------------------------------------------------------------------------
SOURCE_WORK_COLUMN_TIERS = (
    ("实际工作时间", "实际工时"),
    ("工作时长", "工作时间"),
    ("工时",),
)


def _source_header_columns(row):
    """从一行表头中识别姓名、日期和实际工时三列。"""

    texts = [str(cell).replace("\n", "") if cell is not None else "" for cell in row]
    name_col = next((column for column, text in enumerate(texts) if "姓名" in text), None)
    date_col = next((column for column, text in enumerate(texts) if "日期" in text), None)
    if name_col is None or date_col is None:
        return None

    for tier in SOURCE_WORK_COLUMN_TIERS:
        for column, text in enumerate(texts):
            if column in (name_col, date_col):
                continue
            if any(keyword in text for keyword in tier):
                return name_col, date_col, column
    return None


def _detect_source_header(rows, roles=None, header=None):
    """定位我司工时来源表的姓名、日期和实际工时列。

    手动角色映射优先，指定表头时只检查该行，否则扫描前五行。工时列按“实际工时”、
    “工作时长”、普通“工时”三级匹配，防止加班工时或标准工时在明确列存在时抢先命中。
    返回零基表头行与列号，无法形成完整三列时返回 ``None``。
    """
    if roles and all(k in roles for k in ("name", "date", "work")):
        hdr0 = (header - 1) if header else 0
        return hdr0, roles["name"], roles["date"], roles["work"]
    candidates = [header - 1] if header else range(min(5, len(rows)))
    for row_index in candidates:
        if row_index < 0 or row_index >= len(rows):
            continue
        row = rows[row_index]
        joined = "".join(str(x) for x in row if x is not None)
        if "姓名" not in joined or "日期" not in joined:
            continue
        columns = _source_header_columns(row)
        if columns is not None:
            return row_index, *columns
    return None


def _source_row_record(row, name_col, date_col, work_col, skip):
    """把一行来源明细规范化为 ``(姓名, 日期键, 工时)``，无效行返回 ``None``。"""

    if name_col >= len(row):
        return None
    name = _norm_name(row[name_col])
    if not name:
        return None
    raw_date = row[date_col] if date_col < len(row) else None
    date_key = cc.norm_date(raw_date)
    if date_key is None:
        date_key = _day_of(raw_date)
    if date_key is None:
        return None
    raw_work = row[work_col] if work_col < len(row) else None
    work = _to_num(raw_work, skip=skip)
    if work is None:
        return None
    return name, date_key, work


def _accumulate_source_sheet(rows, start, columns, data, days_seen, skip):
    """累加一个已识别页签中的有效工时明细，并返回纳入行数。"""

    count = 0
    for row in rows[start:]:
        record = _source_row_record(row, *columns, skip)
        if record is None:
            continue
        name, date_key, work = record
        person = data.setdefault(name, {})
        # 同一员工同一天可能有多段打卡或多个来源文件，工时需要累加而不是覆盖。
        person[date_key] = person.get(date_key, 0.0) + work
        days_seen.add(date_key)
        count += 1
    return count


def _load_source_one(path, data, days_seen, log=None, opts=None):
    """读取一个我司工时来源文件并累加到跨文件汇总。

    每个页签独立判断是否为明细页，支持按文件覆盖工作表、表头、列映射、数据起始行和
    假休标记。日期优先保留完整年月日，只有旧格式缺年月时才回退日号，避免跨月同一日
    被错误相加。返回实际纳入的明细条数。
    """
    def _lg(m):
        """仅在调用方提供日志回调时转发单文件识别信息。"""
        if log:
            log(m)

    opts = opts or cc.DEFAULTS
    fname = os.path.basename(path)
    roles = opts.resolve_roles(path)
    header = opts.resolve_header(path)
    ds_override = opts.resolve_data_start(path)
    want_sheet = opts.resolve_sheet(path)
    skip = opts.skip_set()
    file_cnt = 0
    for sname, rows in _read_sheets(path):
        if want_sheet and sname != want_sheet:
            continue
        if not rows:
            _lg("  · [跳过] %s / %s（空表）" % (fname, sname))
            continue
        det = _detect_source_header(rows, roles=roles, header=header)
        if det is None:
            _lg("  · [跳过] %s / %s（无 姓名/日期/实际工时 列，非考勤明细）" % (fname, sname))
            continue
        hdr_idx, col_name, col_date, col_work = det
        start = (ds_override - 1) if ds_override else (hdr_idx + 1)
        cnt = _accumulate_source_sheet(
            rows,
            start,
            (col_name, col_date, col_work),
            data,
            days_seen,
            skip,
        )
        file_cnt += cnt
        _lg("  · [读取] %s / %s：%d 条明细" % (fname, sname, cnt))
    return file_cnt


def load_source(paths, log=None, opts=None):
    """
    自动识别表头，聚合每人每日"实际工作时间"。支持单个路径或路径列表（多文件）。
    每个文件的每个子表都会自动判断是否为考勤明细，非明细子表自动跳过。
    返回 ``data`` 与 ``days_seen``。日期键可能是完整 ``(年, 月, 日)`` 元组，也可能是
    兼容旧表的纯日号；多个文件按同一姓名和日期累加工时。
    """
    def _lg(m):
        """按需转发跨文件处理日志。"""
        if log:
            log(m)

    opts = opts or cc.DEFAULTS
    if isinstance(paths, str):
        paths = [paths]
    data = {}
    days_seen = set()
    for p in paths:
        _lg("  文件：%s" % os.path.basename(p))
        _load_source_one(p, data, days_seen, log=log, opts=opts)
    return data, days_seen


# ---------------------------------------------------------------------------
# 2) 通用解析"待对数据"(劳务公司)：格式各异，自动识别
# ---------------------------------------------------------------------------
LABOR_TOTAL_EXCLUDES = (
    "天数", "工价", "金额", "工资", "餐补", "薪资", "扣发",
    "保险", "补贴", "单价", "小时工", "备注",
)


def _find_header_cell(row, keyword):
    """返回首个包含指定关键字的零基列号，找不到时返回 ``None``。"""

    for column, cell in enumerate(row):
        if cell is not None and keyword in str(cell):
            return column
    return None


def _find_labor_name_position(rows, roles, header):
    """确定劳务表姓名表头行和姓名列，并保持人工列映射的最高优先级。"""

    name_row = name_col = None
    if header:
        header_index = header - 1
        if 0 <= header_index < len(rows):
            name_row = header_index
            name_col = roles.get("name")
            if name_col is None:
                name_col = _find_header_cell(rows[header_index], "姓名")

    if name_row is None or name_col is None:
        for row_index, row in enumerate(rows[:6]):
            detected_col = _find_header_cell(row, "姓名")
            if detected_col is not None:
                name_row, name_col = row_index, detected_col
                break

    # 人工指定姓名列最终生效，但仍要求能够定位一个真实表头行，以免把普通明细误认成考勤表。
    if "name" in roles:
        name_col = roles["name"]
    if name_row is None or name_col is None:
        return None
    return name_row, name_col


def _day_columns(row, name_col):
    """从一行中提取姓名列右侧的 ``日号 -> 零基列号`` 映射。"""

    columns = {}
    for column, cell in enumerate(row):
        if column <= name_col:
            continue
        day = _day_of(cell)
        if day is not None:
            columns[day] = column
    return columns


def _find_labor_day_layout(rows, name_row, name_col):
    """在姓名表头行及其下一行中选择日期列最多的布局。"""

    candidates = []
    for row_index in (name_row, name_row + 1):
        if row_index < len(rows):
            candidates.append((row_index, _day_columns(rows[row_index], name_col)))
    if not candidates:
        return None
    # ``max`` 在数量相同时保留先出现的姓名表头行，与原先的稳定选择规则一致。
    best = max(candidates, key=lambda item: len(item[1]))
    return best if len(best[1]) >= 10 else None


def _combined_header_text(rows, header_rows, column):
    """合并同一列在多层表头中的文字，供合计列语义判断使用。"""

    parts = []
    for row_index in header_rows:
        if column < len(rows[row_index]) and rows[row_index][column] is not None:
            parts.append(str(rows[row_index][column]))
    return "".join(parts)


def _find_labor_total_column(rows, name_row, day_row, day_cols):
    """在日期区右侧按“工时/出勤、合计/总计”两级优先级寻找合计列。"""

    header_rows = [row for row in (name_row, day_row) if row < len(rows)]
    start_column = max(day_cols.values()) + 1
    stop_column = max((len(rows[row]) for row in header_rows), default=0)
    for keywords in (("工时", "出勤"), ("合计", "总计")):
        for column in range(start_column, stop_column):
            text = _combined_header_text(rows, header_rows, column)
            if not text or any(word in text for word in LABOR_TOTAL_EXCLUDES):
                continue
            if any(word in text for word in keywords):
                return column
    return None


def _find_labor_layout(rows, roles=None, header=None, data_start=None):
    """
    在一个 sheet 内自动识别布局。返回 dict 或 None：
      name_row  含"姓名"的行下标
      name_col  姓名所在列
      day_row   日期/日号所在行下标
      day_cols  {日(int): 列下标}
      total_col 合计/出勤工时列（可能为 None）
      data_start 数据起始行下标
    手动姓名和合计列映射、表头行、数据起始行优先；逐日列始终自动识别。日期行在姓名
    表头及下一行中选择可映射日号最多的一行，至少识别十天才认为是考勤表。合计列只在
    日期区右侧寻找，并排除工资、金额、天数等常见干扰列。
    """
    roles = roles or {}
    name_position = _find_labor_name_position(rows, roles, header)
    if name_position is None:
        return None
    name_row, name_col = name_position
    day_layout = _find_labor_day_layout(rows, name_row, name_col)
    if day_layout is None:
        return None
    day_row, day_cols = day_layout
    total_col = roles.get("total")
    if total_col is None:
        total_col = _find_labor_total_column(rows, name_row, day_row, day_cols)
    start_row = (data_start - 1) if data_start else (max(name_row, day_row) + 1)
    return {
        "name_col": name_col, "day_row": day_row, "day_cols": day_cols,
        "total_col": total_col, "data_start": start_row,
    }


def _collect_labor_candidates(path, roles, header, data_start, wanted_sheet, log):
    """读取并评估全部页签，返回有效候选与被跳过的页签名称。"""

    filename = os.path.basename(path)
    candidates = []
    skipped = []
    for sheet_name, rows in _read_sheets(path):
        if wanted_sheet and sheet_name != wanted_sheet:
            continue
        layout = _find_labor_layout(rows, roles=roles, header=header, data_start=data_start)
        if layout is not None:
            candidates.append((sheet_name, rows, layout))
            continue
        skipped.append(sheet_name)
        log("  · [跳过] %s / %s（非考勤明细）" % (filename, sheet_name))
    return candidates, skipped


def _parse_labor_rows(rows, layout, skip):
    """按已确认的劳务表布局解析人员逐日工时，并统计合计口径不一致人数。"""

    result = {}
    mismatch = 0
    name_col = layout["name_col"]
    day_cols = layout["day_cols"]
    total_col = layout["total_col"]
    for row in rows[layout["data_start"]:]:
        if name_col >= len(row):
            continue
        name = _norm_name(row[name_col])
        if not name or name in ("合计", "合计：", "总计", "总出勤工时"):
            continue
        days = {
            day: value
            for day, column in day_cols.items()
            if column < len(row) and (value := _to_num(row[column], skip=skip)) is not None
        }
        stated_total = (
            _to_num(row[total_col], skip=skip)
            if total_col is not None and total_col < len(row)
            else None
        )
        day_sum = round(sum(days.values()), 2) if days else None
        if stated_total is not None and day_sum is not None and abs(stated_total - day_sum) > TOL:
            mismatch += 1
        total = stated_total if stated_total is not None else day_sum
        if days or total is not None:
            result[name] = {"days": days, "total": total}
    return result, mismatch


def _append_labor_meta(meta, filename, sheet_name, result, layout, mismatch, skipped,
                       candidate_count):
    """把单个劳务文件的解析诊断写入可信度指标列表。"""

    if meta is None:
        return
    meta.append({
        "file": filename,
        "sheet": sheet_name,
        "people": len(result),
        "day_cols": len(layout["day_cols"]) if layout else 0,
        "has_total_col": bool(layout and layout["total_col"] is not None),
        "total_sum_mismatch": mismatch,
        "skipped": skipped,
        "n_candidates": candidate_count,
    })


def load_labor(path, log=None, meta=None, opts=None):
    """
    返回 {姓名: {"days":{日:工时}, "total":合计或None}}。
    合计优先取表内"合计/出勤工时"列；无该列时用逐日数字求和。
    meta: 可选 list，追加本文件的诊断信息（供可信度评估用）。
    支持按文件覆盖姓名、合计列、表头、数据起始行、工作表和假休标记。多个有效页签时
    选择日列最多者，而不是依赖页签顺序。表内合计优先；没有合计时按有效逐日数字求和，
    同时统计表内合计与逐日和不一致人数，作为可信度判断依据。
    """
    def _lg(m):
        """按需转发劳务文件布局识别日志。"""
        if log:
            log(m)

    opts = opts or cc.DEFAULTS
    fname = os.path.basename(path)
    roles = opts.resolve_roles(path)
    header = opts.resolve_header(path)
    ds_override = opts.resolve_data_start(path)
    want_sheet = opts.resolve_sheet(path)
    skip = opts.skip_set()
    candidates, skipped = _collect_labor_candidates(
        path, roles, header, ds_override, want_sheet, _lg,
    )
    if not candidates:
        _lg("  ⚠ %s：未识别到考勤子表" % fname)
        _append_labor_meta(meta, fname, None, {}, None, 0, skipped, 0)
        return {}
    candidates.sort(key=lambda x: len(x[2]["day_cols"]), reverse=True)
    # Python 排序稳定，同日列数平局时保留工作簿中更靠前的页签。
    sname, rows, lay = candidates[0]
    result, mismatch = _parse_labor_rows(rows, lay, skip)
    total_note = "有合计列" if lay["total_col"] is not None else "无合计列(按逐日求和)"
    _lg("  · [读取] %s / %s：识别 %d 人，%d 个日列，%s" % (
        fname, sname, len(result), len(lay["day_cols"]), total_note,
    ))
    if mismatch:
        _lg("      注意：%d 人的表内合计与逐日求和不一致（合计列可能识别有误）" % mismatch)
    _append_labor_meta(meta, fname, sname, result, lay, mismatch, skipped, len(candidates))
    return result


# ---------------------------------------------------------------------------
# 3) 填写待对表·总表
# ---------------------------------------------------------------------------
def _zong_header_rows(header):
    """返回待对总表需要扫描的 1 基表头行。"""

    return [header] if header else range(1, 4)


def _scan_zong_role_columns(ws, rows):
    """扫描总表表头，识别姓名、劳务公司、出勤工时和对账时间列。"""

    columns = {"name": None, "comp": None, "work": None, "check": None}
    for row in rows:
        if row < 1 or row > ws.max_row:
            continue
        for column in range(1, ws.max_column + 1):
            value = ws.cell(row, column).value
            if value is None:
                continue
            text = str(value)
            if columns["name"] is None and text.strip() == "姓名":
                columns["name"] = column
            if columns["comp"] is None and "劳务公司" in text:
                columns["comp"] = column
            if columns["work"] is None and "出勤工时" in text:
                columns["work"] = column
            if columns["check"] is None and "对账时间" in text:
                columns["check"] = column
    return columns


def _apply_zong_role_overrides(columns, roles):
    """把 Options 的零基人工列映射覆盖到 openpyxl 的 1 基列号。"""

    resolved = dict(columns)
    for role in ("name", "comp", "work", "check"):
        if role in roles:
            resolved[role] = roles[role] + 1
    return resolved


def _zong_day_columns(ws, row, work_col):
    """提取一个候选表头行中的日期列，并排除合计列右侧的数字表头。"""

    columns = {}
    for column in range(1, ws.max_column + 1):
        day = _day_of(ws.cell(row, column).value)
        if day is not None and (work_col is None or column < work_col):
            columns[day] = column
    return columns


def _find_zong_day_layout(ws, rows, work_col):
    """选择日期列数量最多的总表表头行，数量相同时保留更靠前的行。"""

    best_row = None
    best_columns = {}
    for row in rows:
        if row < 1 or row > ws.max_row:
            continue
        columns = _zong_day_columns(ws, row, work_col)
        if len(columns) > len(best_columns):
            best_row, best_columns = row, columns
    return best_row, best_columns


def _locate_zong(ws, opts=None, path=""):
    """
    定位我司待对总表中的姓名、劳务公司、逐日工时、合计工时和对账时间列。

    自动识别只扫描工作表前 3 行，这是为了兼容“第 1 行标题、第 2 行表头”一类
    常见模板，同时避免正文中偶然出现“姓名”“出勤工时”等文字时被误认成表头。
    管理员配置的 ``roles``、``header`` 和 ``data_start`` 优先于自动结果；其中角色列
    在配置中使用 0 基下标，而 openpyxl 使用 1 基下标，因此返回前需要加一转换。
    逐日列仍按单元格内容自动识别，以免管理员维护 28～31 个日期列的易错映射。

    返回字典中的 ``day_cols`` 采用 ``{日号: 列号}``，列号均为 openpyxl 的 1 基下标。
    劳务公司、合计和对账时间属于可选列，可以返回 ``None``；姓名列是后续匹配的
    必要主键，缺失时必须明确报错，不能沿用旧实现猜测为固定列。
    """
    opts = opts or cc.DEFAULTS
    roles = opts.resolve_roles(path)
    header = opts.resolve_header(path)
    ds_override = opts.resolve_data_start(path)
    scan_rows = _zong_header_rows(header)
    columns = _scan_zong_role_columns(ws, scan_rows)
    columns = _apply_zong_role_overrides(columns, roles)
    day_row, day_cols = _find_zong_day_layout(
        ws,
        _zong_header_rows(header),
        columns["work"],
    )
    # 数据从日期表头下一行开始；完全找不到日期行时按常见的“第 2 行表头”兼容推定第 3 行。
    data_start = ds_override if ds_override else ((day_row or 2) + 1)
    # 姓名是双方对齐的唯一键，错误猜列会生成看似成功但实际错配的报表，因此宁可中止。
    if columns["name"] is None:
        raise ValueError("总表未能识别到『姓名』列，请检查表头")
    return {
        # 可选列保持 None，让下游显式降级；不能用固定第 3 列冒充劳务公司列。
        "name_col": columns["name"], "comp_col": columns["comp"],
        "day_cols": day_cols, "work_col": columns["work"],
        "check_col": columns["check"], "data_start": data_start,
    }


def _source_month_days(src_data):
    """统计来源数据中每个年月覆盖过的日号集合。"""

    months = {}
    for person in src_data.values():
        for date_key in person:
            if isinstance(date_key, tuple):
                months.setdefault(date_key[:2], set()).add(date_key[2])
    return months


def _target_source_month(month_days):
    """选择覆盖日数最多的月份；并列时保留来源中先出现的月份。"""

    return max(month_days, key=lambda month: len(month_days[month])) if month_days else None


def _source_day_value(person, target_month, day):
    """读取某人的指定日工时，优先完整年月日键并兼容旧版纯日号键。"""

    if target_month is not None:
        full_date = (target_month[0], target_month[1], day)
        if full_date in person:
            return person[full_date]
    return person.get(day)


def _unmapped_person_days(person, target_month, used_days):
    """统计属于目标月份但总表没有对应日期列的来源工时条数。"""

    overflow = 0
    for date_key in person:
        day = date_key[2] if isinstance(date_key, tuple) else date_key
        in_target = (
            not isinstance(date_key, tuple)
            or (target_month is not None and date_key[:2] == target_month)
        )
        if in_target and day not in used_days:
            overflow += 1
    return overflow


def _fill_zong_person(ws, row, layout, person, target_month, checked_at):
    """填写一名人员的逐日工时、合计和对账时间，返回写入格数与溢出天数。"""

    total = 0.0
    filled_cells = 0
    used_days = set()
    for day, column in layout["day_cols"].items():
        value = _source_day_value(person, target_month, day)
        if value is None:
            continue
        ws.cell(row, column).value = value
        total += value
        filled_cells += 1
        used_days.add(day)
    if layout["work_col"]:
        ws.cell(row, layout["work_col"]).value = round(total, 2)
    if layout["check_col"]:
        ws.cell(row, layout["check_col"]).value = checked_at
    return filled_cells, _unmapped_person_days(person, target_month, used_days)


def fill_zong(ws, src_data, log=None, opts=None, path=""):
    """
    把工时来源按姓名和日号写入我司总表，并计算目标月份的出勤工时合计。

    来源键既可能是新版 ``(年, 月, 日)``，也可能是旧版纯日号。本函数先选择覆盖
    日数最多的年月作为总表主月份，再仅把该月数据投射到只有“1～31 日”列的模板。
    这种选择不会把跨月数据悄悄相加；无法落入总表日期列的数据会记录提示。
    """
    def _lg(m):
        """把填表阶段日志转交给调用方；未提供回调时保持静默。"""
        if log:
            log(m)

    lay = _locate_zong(ws, opts=opts, path=path)
    # 无日期列时继续执行会把所有人的合计写成 0，属于危险的“成功结果”，因此直接拒绝。
    if not lay["day_cols"]:
        raise ValueError("总表未能识别到任何『日期』列(日期表头行识别失败)，请检查表头")
    month_days = _source_month_days(src_data)
    target_ym = _target_source_month(month_days)
    if len(month_days) > 1:
        # 总表仅有日号、没有年月维度，跨月内容无法无歧义地写入同一组列，必须告知用户取舍。
        _lg("  注意：数据来源跨 %d 个月份 %s，总表逐日列仅对应主月份 %s，其余月份不并入逐日"
            % (len(month_days), sorted("%d-%02d" % ym for ym in month_days),
               ("%d-%02d" % target_ym) if target_ym else "-"))

    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    filled_people = 0
    filled_cells = 0
    unmatched = []
    for r in range(lay["data_start"], ws.max_row + 1):
        nm = _norm_name(ws.cell(r, lay["name_col"]).value)
        if not nm:
            continue
        if nm not in src_data:
            unmatched.append(nm)
            continue
        person = src_data[nm]
        person_cells, overflow = _fill_zong_person(ws, r, lay, person, target_ym, now)
        filled_cells += person_cells
        if overflow:
            _lg("  · 提示：%s 有 %d 天工时无对应日列，未计入合计" % (nm, overflow))
        filled_people += 1
    _lg("  总表已填 %d 人、%d 个日格子；数据来源中有 %d 人未在总表找到"
        % (filled_people, filled_cells, len(set(src_data) - _zong_names(ws, lay))))
    return {"filled_people": filled_people, "filled_cells": filled_cells,
            "unmatched_in_zong": unmatched, "layout": lay}


def _zong_names(ws, lay):
    """收集总表中的规范化姓名集合，供名单交集、缺失人员和填写覆盖率计算使用。"""
    names = set()
    for r in range(lay["data_start"], ws.max_row + 1):
        nm = _norm_name(ws.cell(r, lay["name_col"]).value)
        if nm:
            names.add(nm)
    return names


# ---------------------------------------------------------------------------
# 4) 对账比较：总工时 + 逐日；劳务公司"假/休/空白"不参与
# ---------------------------------------------------------------------------
def _cached_cell_value(ws, ws_v, row, column):
    """读取工作表单元格；公式文本优先回退到 ``data_only`` 副本的缓存结果。"""

    value = ws.cell(row, column).value
    if isinstance(value, str) and value.startswith("=") and ws_v is not None:
        return ws_v.cell(row, column).value
    return value


def _resolve_company_name(ws, row, layout, name, comp_map):
    """解析总表中的劳务公司名称，并屏蔽没有缓存结果的公式文本。"""

    company = comp_map.get(name)
    if not company and layout["comp_col"]:
        company = ws.cell(row, layout["comp_col"]).value
    if isinstance(company, str) and company.startswith("="):
        return ""
    return str(company).strip() if company is not None else ""


def _project_zong_rows(ws, layout, comp_map, skip, ws_v):
    """把总表行投影为纯内存结构，避免比较阶段反复访问 Excel 单元格。"""

    projected = {}
    for row in range(layout["data_start"], ws.max_row + 1):
        name = _norm_name(ws.cell(row, layout["name_col"]).value)
        if not name:
            continue
        days = {}
        for day, column in layout["day_cols"].items():
            value = _to_num(_cached_cell_value(ws, ws_v, row, column), skip=skip)
            if value is not None:
                days[day] = value
        total = None
        if layout["work_col"]:
            total = _to_num(
                _cached_cell_value(ws, ws_v, row, layout["work_col"]),
                skip=skip,
            )
        if total is None:
            total = round(sum(days.values()), 2)
        projected[name] = {
            "days": days,
            "total": total,
            "comp": _resolve_company_name(ws, row, layout, name, comp_map),
        }
    return projected


def _roster_anomalies(zong, labor):
    """生成双方名单差集异常，顺序保持为“仅我司、仅劳务公司”。"""

    anomalies = []
    zong_names = set(zong)
    labor_names = set(labor)
    for name in sorted(zong_names - labor_names):
        anomalies.append({
            "姓名": name,
            "所属劳务公司": zong[name]["comp"],
            "异常类型": "仅我司名单有",
            "我司出勤工时": zong[name]["total"],
            "劳务公司工时": "",
            "差异": "",
            "差异明细": "该员工不在任何劳务公司对账单中",
            "来源文件": "",
        })
    for name in sorted(labor_names - zong_names):
        source = labor[name].get("source", "")
        anomalies.append({
            "姓名": name,
            "所属劳务公司": source,
            "异常类型": "仅劳务公司有",
            "我司出勤工时": "",
            "劳务公司工时": labor[name]["total"],
            "差异": "",
            "差异明细": "该员工不在我司总表中",
            "来源文件": source,
        })
    return anomalies


def _daily_difference_details(our_days, labor_days, tolerance):
    """返回逐日工时差异文本；劳务侧没有数值的日期不参与比较。"""

    details = []
    for day in sorted(set(our_days) | set(labor_days)):
        labor_value = labor_days.get(day)
        if labor_value is None:
            continue
        our_value = our_days.get(day, 0.0)
        if abs(our_value - labor_value) > tolerance:
            details.append(
                "%d日:我司%s/劳务%s" % (day, _fmt(our_value), _fmt(labor_value))
            )
    return details


def _matched_person_anomalies(name, our_data, labor_data, tolerance):
    """比较一名双方共有人员的总工时与逐日工时，返回零到两条异常。"""

    anomalies = []
    our_total = our_data["total"] or 0.0
    labor_total = labor_data["total"] or 0.0
    source = labor_data.get("source", "")
    common = {
        "姓名": name,
        "所属劳务公司": our_data["comp"],
        "我司出勤工时": round(our_total, 2),
        "劳务公司工时": round(labor_total, 2),
        "差异": round(our_total - labor_total, 2),
        "来源文件": source,
    }
    if abs(our_total - labor_total) > tolerance:
        anomalies.append({**common, "异常类型": "总工时不一致", "差异明细": ""})

    day_details = _daily_difference_details(
        our_data["days"], labor_data["days"], tolerance,
    )
    if day_details:
        anomalies.append({
            **common,
            "异常类型": "逐日工时不一致",
            "差异明细": "；".join(day_details),
        })
    return anomalies


def reconcile(ws, lay, labor, comp_map=None, log=None, tol=None, skip=None, ws_v=None):
    """
    比较已填写的我司总表与合并后的劳务公司数据，返回结构化异常列表。

    ``labor`` 的结构为 ``{姓名: {days: {日号: 工时}, total: 合计, source: 文件名}}``。
    名单差异、总工时差异和逐日差异分别生成记录，便于汇总表按异常类型筛选。
    ``tol`` 控制浮点和表格小数造成的允许误差；``skip`` 交给数值解析器过滤
    “假”“休”等非出勤标记。

    ``ws`` 必须是可写工作簿，以保留公式、样式和最终输出；``ws_v`` 是同一原文件的
    ``data_only`` 副本，仅在 ``ws`` 单元格是公式文本时提供上次由 Excel 缓存的结果。
    若缓存不存在，后续会按逐日数值求和，而不会把公式字符串误当作工时。
    """
    def _lg(m):
        """统一转发对账阶段日志，允许核心逻辑在桌面端和 Web 端复用。"""
        if log:
            log(m)

    tolerance = TOL if tol is None else tol
    zong = _project_zong_rows(ws, lay, comp_map or {}, skip, ws_v)
    anomalies = _roster_anomalies(zong, labor)
    for name in sorted(set(zong) & set(labor)):
        anomalies.extend(_matched_person_anomalies(
            name, zong[name], labor[name], tolerance,
        ))
    _lg("  对账完成：共 %d 条异常" % len(anomalies))
    return anomalies


def _fmt(x):
    """把对账明细中的数值格式化为整数或最多两位小数，减少无意义的小数尾数。"""
    if x is None:
        return "-"
    if float(x).is_integer():
        return str(int(x))
    return str(round(x, 2))


# ---------------------------------------------------------------------------
# 5) 可信度评估与异常工作簿渲染
# ---------------------------------------------------------------------------
# 实现位于 reconcile_reporting.py；本模块顶部保留公开名称兼容旧调用。
# ---------------------------------------------------------------------------
# 人工确认：只读预分析(不写文件),供复核对话框展示识别结果与姓名匹配
# ---------------------------------------------------------------------------
def analyze(target_path, source_paths, labor_paths, opts=None):
    """只读分析三类输入文件，返回人工复核所需的结构计划，不创建任何输出文件。

    plan = {
      "target": {file, sheet, sheets:[可选工作表], name_col, comp_col, work_col,
                 check_col, day_cols, header/ data_start, names:[总表姓名]},
      "sources": [{file, sheet, people}],           # 各数据来源识别概况
      "labor":   [{file, sheet, people, names:[]}],  # 各对账单识别概况
      "labor_names": [...],                          # 对账单合并后的全部姓名
      "only_labor": [...],   # 仅对账单有(待配对到我司)
      "only_zong":  [...],   # 仅我司总表有(可作为配对目标)
    }

    此阶段与正式 ``run`` 共用结构识别和加载函数，保证复核界面展示的工作表、列和姓名
    与最终执行口径一致。所有工作簿都以只读或 ``data_only`` 方式打开并在 ``finally``
    中关闭；人工确认只产生选择数据，不允许绕过正式执行阶段直接修改原文件。
    """
    opts = opts or cc.DEFAULTS
    if isinstance(source_paths, str):
        # 统一成列表，避免单个路径被后续循环按字符逐个处理。
        source_paths = [source_paths]

    # 待对表优先使用人工指定页签，其次寻找约定的“总表”，最后才使用首个工作表。
    want_zong = opts.resolve_sheet(target_path)
    wb = openpyxl.load_workbook(target_path, read_only=True, data_only=True)
    try:
        sheetnames = list(wb.sheetnames)  # 一并返回全部页签，供复核界面允许用户改选。
        if want_zong and want_zong in sheetnames:
            ws = wb[want_zong]; used_sheet = want_zong
        elif "总表" in sheetnames:
            ws = wb["总表"]; used_sheet = "总表"
        else:
            ws = wb.worksheets[0]; used_sheet = ws.title
        lay = _locate_zong(ws, opts=opts, path=target_path)  # 与正式填表完全共用识别规则。
        zong_names = sorted(_zong_names(ws, lay))
    finally:
        wb.close()
    target_info = {
        "file": os.path.basename(target_path), "sheet": used_sheet,
        "sheets": sheetnames,
        "name_col": lay["name_col"], "comp_col": lay["comp_col"],
        "work_col": lay["work_col"], "check_col": lay["check_col"],
        "day_cols": sorted(lay["day_cols"].keys()),
        "data_start": lay["data_start"], "names": zong_names,
    }

    # 工时来源在预分析阶段只展示识别人数，不参与我司与劳务姓名的人工配对。
    sources_info = []
    for p in source_paths:
        one, _ = load_source([p], log=None, opts=opts)
        sources_info.append({"file": os.path.basename(p), "people": len(one)})

    # 每份劳务表分别保留解析概况，同时合并姓名集合，用于计算双方独有名单。
    labor_info = []
    labor_names = set()
    for p in labor_paths:
        meta = []
        one = load_labor(p, log=None, meta=meta, opts=opts)
        m = meta[0] if meta else {}
        labor_names |= set(one.keys())  # 集合只用于预览差集；正式冲突策略仍在 run 中执行。
        labor_info.append({"file": os.path.basename(p),
                           "sheet": m.get("sheet"), "people": len(one),
                           "names": sorted(one.keys())})

    zset = set(zong_names)
    return {
        "target": target_info,
        "sources": sources_info,
        "labor": labor_info,
        "labor_names": sorted(labor_names),
        "only_labor": sorted(labor_names - zset),
        "only_zong": sorted(zset - labor_names),
    }


# ---------------------------------------------------------------------------
# 主流程
# ---------------------------------------------------------------------------
RUN_STAGES = [
    ("read_src", 18),
    ("fill", 27),
    ("read_labor", 18),
    ("compare", 27),
    ("assess", 3),
    ("summary", 7),
]


class _RunReporter:
    """统一管理阶段进度、实时日志和写入可信度报告的完整日志。"""

    def __init__(self, log, progress):
        self._callback = log
        self._progress = cc.Progress(progress, stages=RUN_STAGES)
        self.lines = []

    def log(self, message):
        """记录字符串化日志，并在调用端提供回调时同步转发。"""

        self.lines.append(str(message))
        if self._callback:
            self._callback(message)

    def stage(self, name):
        """进入指定处理阶段并更新总体进度。"""

        self._progress.stage(name)

    def done(self):
        """标记全部处理阶段完成。"""

        self._progress.done()

    def text(self):
        """返回可写入 Excel 可信度页的完整运行日志。"""

        return "\n".join(self.lines)


@dataclass
class _RunContext:
    """保存一次对账任务的只读配置，避免阶段函数携带大量平行参数。"""

    target_path: str
    source_paths: list
    labor_paths: list
    out_dir: str
    timestamp: str
    opts: object
    aliases: dict
    reporter: _RunReporter


@dataclass
class _SourceStage:
    """数据来源阶段的聚合结果。"""

    data: dict
    days_seen: set


@dataclass
class _TargetStage:
    """待对总表阶段的工作簿、布局、缓存值和已填写输出。"""

    workbook: object
    worksheet: object
    values_workbook: object
    values_worksheet: object
    stats: dict
    company_map: dict
    filled_path: str

    def close(self):
        """尽最大努力释放两个工作簿，不让清理异常覆盖真正的业务异常。"""

        _safe_close_workbook(self.workbook)
        _safe_close_workbook(self.values_workbook)


@dataclass
class _LaborStage:
    """劳务文件合并结果及可信度诊断信息。"""

    data: dict
    meta: list
    duplicate_count: int


def _path_list(paths):
    """把单个路径与路径序列统一为新列表，避免后续误遍历字符串字符。"""

    return [paths] if isinstance(paths, str) else list(paths)


def _apply_review_choices(opts, target_path, choices, log):
    """把人工确认转换为本次任务专用配置，并返回姓名别名映射。"""

    aliases = {}
    if not choices:
        return opts, aliases

    aliases = dict(choices.get("aliases") or {})
    target_sheet = choices.get("target_sheet")
    target_roles = choices.get("target_roles") or {}
    if target_sheet or target_roles:
        opts = copy.deepcopy(opts)  # 复核结果只对当前任务生效，绝不能污染共享默认配置。
        filename = os.path.basename(target_path)
        file_mapping = dict(opts.columns.get(filename) or {})
        if target_sheet:
            file_mapping["sheet"] = target_sheet
        if target_roles:
            roles = dict(file_mapping.get("roles") or {})
            for role, column in target_roles.items():
                if column:
                    # 复核界面使用 Excel 的 1 基列号，Options 内部统一保存为 0 基列号。
                    roles[role] = int(column) - 1
            file_mapping["roles"] = roles
        opts.columns[filename] = file_mapping
        log(
            "采用人工确认的待对表结构："
            + ("工作表=%s " % target_sheet if target_sheet else "")
            + ("列映射 %d 项" % len(target_roles) if target_roles else "")
        )
    if aliases:
        log("采用人工姓名配对 %d 组(比对时视为同一人)" % len(aliases))
    return opts, aliases


def _resolve_run_output_dir(target_path, out_dir):
    """解析并创建本次任务输出目录，保留旧版源文件旁输出的兼容回退。"""

    if out_dir is None:
        return _unified_out_dir("reconcile", src=target_path) or cc.make_out_dir(target_path)
    os.makedirs(out_dir, exist_ok=True)
    return out_dir


def _build_run_context(target_path, source_paths, labor_paths, out_dir, opts, choices,
                       reporter):
    """规范化一次运行的路径、选项、人工确认和输出位置。"""

    options = opts or cc.DEFAULTS
    options, aliases = _apply_review_choices(options, target_path, choices, reporter.log)
    output_dir = _resolve_run_output_dir(target_path, out_dir)
    context = _RunContext(
        target_path=target_path,
        source_paths=_path_list(source_paths),
        labor_paths=_path_list(labor_paths),
        out_dir=output_dir,
        timestamp=cc.timestamp(),
        opts=options,
        aliases=aliases,
        reporter=reporter,
    )
    reporter.log("采用选项：" + options.summary())
    reporter.log("输出文件夹：%s" % output_dir)
    return context


def _read_source_stage(context):
    """读取并聚合我司工时来源，返回人员工时和覆盖日期集合。"""

    context.reporter.stage("read_src")
    context.reporter.log("① 读取数据来源 ...")
    for path in context.source_paths:
        cc.warn_if_uncached(path, context.reporter.log, what="工时")
    data, days_seen = load_source(
        context.source_paths,
        log=context.reporter.log,
        opts=context.opts,
    )
    context.reporter.log("   共 %d 人、覆盖 %d 天" % (len(data), len(days_seen)))
    return _SourceStage(data=data, days_seen=days_seen)


def _select_target_sheet(workbook, wanted_sheet):
    """按人工/配置页签、“总表”、首个页签的顺序选择待对工作表。"""

    if wanted_sheet:
        if wanted_sheet not in workbook.sheetnames:
            raise ValueError("待对表中找不到工作表 '%s'" % wanted_sheet)
        return workbook[wanted_sheet]
    return workbook["总表"] if "总表" in workbook.sheetnames else workbook.worksheets[0]


def _safe_close_workbook(workbook):
    """尽最大努力关闭工作簿；该清理动作不能覆盖主流程异常。"""

    if workbook is None:
        return
    try:
        workbook.close()
    except Exception:
        pass


def _extract_company_map(values_worksheet, layout):
    """从 ``data_only`` 总表中提取姓名对应的劳务公司公式缓存值。"""

    company_map = {}
    for row in range(layout["data_start"], values_worksheet.max_row + 1):
        name = _norm_name(values_worksheet.cell(row, layout["name_col"]).value)
        if not name:
            continue
        value = (
            values_worksheet.cell(row, layout["comp_col"]).value
            if layout["comp_col"]
            else None
        )
        if value is not None and not (isinstance(value, str) and value.startswith("=")):
            company_map[name] = str(value).strip()
    return company_map


def _load_target_value_projection(context, wanted_sheet, layout):
    """加载总表公式缓存副本；失败时记录提示并返回空的增强信息。"""

    values_workbook = None
    try:
        values_workbook = cc.load_data_only(context.target_path)
        values_worksheet = _select_target_sheet(values_workbook, wanted_sheet)
        company_map = _extract_company_map(values_worksheet, layout)
        return values_workbook, values_worksheet, company_map
    except Exception as exc:
        _safe_close_workbook(values_workbook)
        context.reporter.log(
            "   ⚠ 读取总表缓存值失败(所属劳务公司/公式列将退回原始值)：%s" % exc
        )
        return None, None, {}


def _prepare_target_stage(context, source):
    """填写待对总表、保存已填写副本，并保留比较阶段需要的工作簿句柄。"""

    context.reporter.stage("fill")
    context.reporter.log("② 填写待对表·总表 ...")
    wanted_sheet = context.opts.resolve_sheet(context.target_path)
    workbook = openpyxl.load_workbook(context.target_path)
    values_workbook = None
    try:
        worksheet = _select_target_sheet(workbook, wanted_sheet)
        stats = fill_zong(
            worksheet,
            source.data,
            log=context.reporter.log,
            opts=context.opts,
            path=context.target_path,
        )
        values_workbook, values_worksheet, company_map = _load_target_value_projection(
            context, wanted_sheet, stats["layout"],
        )
        base = os.path.splitext(os.path.basename(context.target_path))[0]
        filled_path = cc.out_path(
            context.out_dir,
            base,
            "_已填写",
            ".xlsx",
            ts=context.timestamp,
        )
        workbook.save(filled_path)
        context.reporter.log("   已保存：%s" % os.path.basename(filled_path))
        return _TargetStage(
            workbook=workbook,
            worksheet=worksheet,
            values_workbook=values_workbook,
            values_worksheet=values_worksheet,
            stats=stats,
            company_map=company_map,
            filled_path=filled_path,
        )
    except Exception:
        _safe_close_workbook(values_workbook)
        _safe_close_workbook(workbook)
        raise


def _merge_labor_file(target, incoming, source_name, conflict, log):
    """把一个劳务文件并入人员索引，返回本文件触发的重复姓名数量。"""

    duplicates = 0
    for name, info in incoming.items():
        info["source"] = source_name
        if name in target:
            duplicates += 1
            if conflict == "first":
                log("   ⚠ 姓名重复：%s 已存在，按【先者优先】保留先读取的" % name)
                continue
            if conflict == "warn":
                log("   ⚠ 姓名重复：%s 出现在多个劳务文件，按【不覆盖仅提示】保留先者" % name)
                continue
            log("   ⚠ 姓名重复：%s 同时出现在多个劳务公司文件，后者覆盖" % name)
        target[name] = info
    return duplicates


def _apply_name_aliases(labor, aliases, log):
    """在内存索引中应用人工姓名配对，拒绝覆盖已存在的真实人员。"""

    applied = 0
    for labor_name, our_name in aliases.items():
        labor_name = _norm_name(labor_name)
        our_name = _norm_name(our_name)
        if not labor_name or not our_name or labor_name == our_name:
            continue
        if labor_name in labor and our_name not in labor:
            labor[our_name] = labor.pop(labor_name)
            applied += 1
            log("   ↔ 姓名配对：劳务「%s」= 我司「%s」" % (labor_name, our_name))
    if applied:
        log("   已应用 %d 组姓名配对" % applied)


def _read_labor_stage(context):
    """读取全部劳务文件，按冲突策略合并姓名并应用人工配对。"""

    context.reporter.stage("read_labor")
    context.reporter.log("③ 读取待对数据(劳务公司) ...")
    labor = {}
    meta = []
    duplicate_count = 0
    for path in context.labor_paths:
        cc.warn_if_uncached(path, context.reporter.log, what="工时")
        incoming = load_labor(
            path,
            log=context.reporter.log,
            meta=meta,
            opts=context.opts,
        )
        duplicate_count += _merge_labor_file(
            labor,
            incoming,
            os.path.basename(path),
            context.opts.conflict,
            context.reporter.log,
        )
    context.reporter.log("   劳务公司合计 %d 人" % len(labor))
    _apply_name_aliases(labor, context.aliases, context.reporter.log)
    return _LaborStage(data=labor, meta=meta, duplicate_count=duplicate_count)


def _compare_target_stage(context, target, labor):
    """执行双方工时比较，并返回异常列表和本次总表姓名集合。"""

    context.reporter.stage("compare")
    context.reporter.log("④ 对账比较 ...")
    anomalies = reconcile(
        target.worksheet,
        target.stats["layout"],
        labor.data,
        comp_map=target.company_map,
        log=context.reporter.log,
        tol=context.opts.tolerance,
        skip=context.opts.skip_set(),
        ws_v=target.values_worksheet,
    )
    names = _zong_names(target.worksheet, target.stats["layout"])
    return anomalies, names


def _build_run_metrics(source, target, labor, zong_names, anomalies):
    """根据各阶段结构化结果构造可信度评估指标。"""

    labor_names = set(labor.data)
    matched_names = zong_names & labor_names
    difference_people = {
        anomaly["姓名"]
        for anomaly in anomalies
        if anomaly["异常类型"] in ("总工时不一致", "逐日工时不一致")
    }
    return {
        "source_people": len(source.data),
        "source_days": len(source.days_seen),
        "source_unmatched": len(set(source.data) - zong_names),
        "zong_people": len(zong_names),
        "filled_people": target.stats.get("filled_people", 0),
        "labor_meta": labor.meta,
        "labor_duplicate_names": labor.duplicate_count,
        "labor_people": len(labor.data),
        "matched_pairs": len(matched_names),
        "only_us": len(zong_names - labor_names),
        "only_labor": len(labor_names - zong_names),
        "diff_people": len(difference_people),
        "anomaly_count": len(anomalies),
    }


def _assess_run(context, metrics):
    """执行可信度评估，并把需要人工关注的结论写入实时日志。"""

    context.reporter.stage("assess")
    credibility = assess_credibility(metrics)
    context.reporter.log(
        "⑤ 可信度评估：【%s】 %d/100" % (credibility["level"], credibility["score"])
    )
    for check in credibility["checks"]:
        if check["级别"] in ("警告", "严重"):
            context.reporter.log(
                "   [%s] %s：%s" % (check["级别"], check["项目"], check["说明"])
            )
    return credibility


def _write_summary_stage(context, anomalies, credibility):
    """生成含可信度页和运行日志的异常汇总工作簿。"""

    context.reporter.stage("summary")
    context.reporter.log("⑥ 生成异常汇总表（含可信度报告）...")
    summary_path = os.path.join(
        context.out_dir,
        "对账异常汇总_%s.xlsx" % context.timestamp,
    )
    write_summary(
        anomalies,
        summary_path,
        credibility=credibility,
        log_text=context.reporter.text(),
    )
    context.reporter.log("   已保存：%s" % os.path.basename(summary_path))
    context.reporter.done()
    return summary_path


def run(target_path, source_paths, labor_paths, out_dir=None, log=None, opts=None,
        choices=None, progress=None):
    """
    执行“来源工时填表 -> 劳务数据合并 -> 双方对账 -> 可信度评估 -> 报告输出”。

    ``target_path`` 是保留原样式和公式的我司待对表；``source_paths`` 支持单个路径或
    多文件列表；``labor_paths`` 是一组劳务公司对账单。``opts`` 统一承载容差、重复
    姓名冲突策略、排除词和人工列映射。

    ``choices`` 来自 ``analyze`` 后的人工确认，格式为
    ``{"target_sheet": 页签, "target_roles": {角色: 1基列号},
    "aliases": {劳务姓名: 我司姓名}}``。选择只作用于本次运行：函数会深拷贝配置，
    不污染全局设置；姓名别名也仅改写内存中的劳务索引，不修改任何原始表格。

    ``progress`` 接收 0～100 的阶段进度。返回值同时包含两个输出路径、填表统计、
    结构化异常、可信度及其原始指标，供桌面端和 Web 端复用同一结果投影。
    """
    reporter = _RunReporter(log, progress)
    context = _build_run_context(
        target_path,
        source_paths,
        labor_paths,
        out_dir,
        opts,
        choices,
        reporter,
    )
    source = _read_source_stage(context)
    target = _prepare_target_stage(context, source)
    try:
        labor = _read_labor_stage(context)
        anomalies, zong_names = _compare_target_stage(context, target, labor)
    finally:
        # 劳务解析或比较任一阶段失败都必须释放工作簿，避免 Windows 下锁住输入和输出文件。
        target.close()
    metrics = _build_run_metrics(source, target, labor, zong_names, anomalies)
    credibility = _assess_run(context, metrics)
    summary_path = _write_summary_stage(context, anomalies, credibility)
    return {
        "filled_path": target.filled_path,
        "summary_path": summary_path,
        "stats": target.stats,
        "anomalies": anomalies,
        "credibility": credibility,
        "metrics": metrics,
        "parameters": {
            "tolerance": context.opts.tolerance,
            "conflict": context.opts.conflict,
            "skip_extra": sorted(context.opts.skip_extra),
        },
    }


def _unified_out_dir(feature, src=None):
    """
    通过统一路径与设置模块解析业务输出目录，旧环境不可用时返回 ``None``。

    ``src`` 只在“输出到源文件旁”模式下用于确定基准目录。这里保留宽容回退是为了兼容
    早期脚本直接复制单个核心文件运行的场景；正式桌面端和 Web 端均应成功走统一路径。
    本函数不自行创建第二套目录规则，调用方只会在返回 ``None`` 时使用公共兼容助手。
    """
    try:
        from . import paths as _paths
        from . import settings as _settings
        st = _settings.get_settings()
        kw = st.output_kwargs()
        if src and not kw.get("src_path"):
            # 显式配置的 src_path 优先，避免覆盖调用环境已经选定的源文件基准。
            kw["src_path"] = src
        return _paths.resolve_output_dir(feature, **kw)
    except Exception as exc:
        # 兼容入口不能阻断对账主流程；正式环境中的路径问题会由后续输出创建操作暴露，
        # 这里记日志避免静默掩盖真实配置或权限错误。
        logging.getLogger(__name__).warning("统一输出目录不可用：%s", exc)
        return None
