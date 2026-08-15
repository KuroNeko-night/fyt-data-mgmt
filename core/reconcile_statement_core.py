# -*- coding: utf-8 -*-
"""
供应商对账单制作核心
====================
从一份或多份采购清单中识别批次块，先返回批次供人工选择，再把选中批次按供应商
合并成对账单。该两阶段流程避免程序在未确认范围时直接落盘，也让 Web 与桌面端
复用同一批次识别、排除和主数据补全规则。

采购清单的一个批次块由“批次号 + 交付日期”标题定位，固定读取标题起始列右侧六列：
材料编号、材料名称、规格、单位、采购数量和供应商。同一页签可左右并排多个块，也可
纵向连续出现多个块。材料行通过编号形态识别；黄色或橙色填充代表取消、不需对账，
扫描预览会统计这些行，正式生成时才排除。

供应商优先采用人工映射，其次从“供应商名 + 月采购清单明细”的文件名提取，最后从
选中数据行中取出现次数最多的供应商。物料名称、规格、单位和供应商缺失时只从正式
主数据库补空值，不覆盖源表已有内容。输出文件名经过 Windows 保留名、非法字符和
目录穿越校验，并使用唯一文件路径避免覆盖旧报告。
"""
from __future__ import annotations

import os
import re
from collections import Counter

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from . import common_core, material_catalog, paths, settings

EXCEL_SUFFIXES = {".xlsx", ".xlsm"}

_TITLE_PATTERN = re.compile(r"^\s*([A-Z0-9]+(?:-\d+)?)\s*交付日期")
_CODE_PATTERN = re.compile(r"^[A-Za-z]{2,6}\d{2,}[A-Za-z0-9]*$|^\d{5,}$")
# 黄色是常规模板的“不对账”标记，橙色用于部分供应商模板的“取消料”；正常数据行
# 使用的绿色 92D050 刻意不在集合中。比较时只取 RGB 尾六位以兼容带透明度的 ARGB。
_EXCLUDE_FILLS = ("FFFF00", "FFFF99", "FFEB9C", "FFF2CC", "FFC000", "FFA500", "FF9900", "F4B183", "ED7D31")
# 第三方文件可能把使用范围扩到 Excel 最大列；业务块只在前部区域，设置上限可避免
# 每行扫描一万余空单元格而严重拖慢处理。
_MAX_SCAN_COL = 256
_SUPPLIER_STRIP = re.compile(r"[（(]\s*已下单\s*[)）]|已下单")
_INVALID_FILENAME_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')
_WINDOWS_RESERVED_NAMES = {
    "CON", "PRN", "AUX", "NUL",
    *("COM%d" % index for index in range(1, 10)),
    *("LPT%d" % index for index in range(1, 10)),
}


def _text(value) -> str:
    """把任意单元格值转为去首尾空白的业务文本，空值统一为空串。"""
    return "" if value is None else str(value).strip()


def _is_yellow(cell) -> bool:
    """判断单元格填充是否属于采购清单约定的排除颜色。

    优先识别 solid RGB/ARGB；旧模板可能使用 Excel indexed 调色板，因此对黄色和
    橙色常见索引提供有限兜底。主题色和其他索引色不凭视觉近似排除，避免把正常
    蓝色、绿色数据行误删。
    """
    fill = cell.fill
    if fill is None or fill.patternType != "solid" or fill.fgColor is None:
        return False
    rgb = str(getattr(fill.fgColor, "rgb", "") or "")
    if len(rgb) >= 6 and rgb[-6:] in _EXCLUDE_FILLS:
        return True
    # indexed 颜色无法直接取得 RGB，仅接受已确认的黄色和橙色索引。
    if getattr(fill.fgColor, "type", "") == "indexed":
        try:
            index = int(getattr(fill.fgColor, "indexed", -1))
        except (TypeError, ValueError):
            return False
        return index in (6, 44, 45, 46)
    return False


def supplier_from_filename(name: str) -> str | None:
    """从约定格式的采购清单文件名中提取供应商，无法确认时返回 ``None``。

    只有以“数字月采购清单明细”结尾的文件名才参与提取，防止把任意前缀误当供应商；
    提取后移除中英文括号包围的“已下单”标记。
    """
    stem = os.path.splitext(os.path.basename(name))[0]
    match = re.search(r"(\d+月采购清单明细)$", stem)
    if not match:
        return None
    supplier = _SUPPLIER_STRIP.sub("", stem[: match.start()]).strip(" （()）\t")
    return supplier or None


def _safe_filename_component(value: object, fallback: str) -> str:
    """把供应商或月份转换为安全、有限长度的 Windows 文件名片段。

    非法字符、控制字符和连续点号会替换，尾部点号/空格会移除；空值、相对目录符号
    和设备保留名均有显式兜底。最终截到 80 字符，为月份、固定后缀和路径长度留余量。
    """
    text = _INVALID_FILENAME_CHARS.sub("_", _text(value)).rstrip(". ")
    text = re.sub(r"\.{2,}", "_", text)
    if not text or text in {".", ".."}:
        text = fallback
    if text.split(".", 1)[0].upper() in _WINDOWS_RESERVED_NAMES:
        text = "_" + text
    return text[:80]


def _output_path(out_dir: str, supplier: str, month: str) -> str:
    """构造不越出输出目录且不覆盖现有文件的对账单绝对路径。

    供应商和月份先经 Windows 文件名清洗，再用 ``abspath`` 拼出目标路径；
    ``commonpath`` 是目录穿越的最终防线，``unique_path`` 保证重复生成或并发
    输出时不会覆盖已有报告。路径无法约束在输出目录内时抛出 ``ValueError``。
    """
    filename = "%s%s月对单明细.xlsx" % (
        _safe_filename_component(supplier, "未命名供应商"),
        _safe_filename_component(month, "未填写月份"),
    )
    target = os.path.abspath(os.path.join(out_dir, filename))
    # 即使上游过滤规则未来调整，也用 commonpath 做最后一道目录穿越防线。
    if os.path.commonpath((os.path.normcase(out_dir), os.path.normcase(target))) != os.path.normcase(out_dir):
        raise ValueError("对账单输出文件名无效")
    return common_core.unique_path(target)


def _row_title_cells(row) -> list[tuple[str, int]]:
    """从一行中提取所有“批次号 + 交付日期”标题及其起始列。"""
    titles: list[tuple[str, int]] = []
    for cell in row:
        value = cell.value
        # 先做低成本包含判断，再执行正则，减少大表中无关单元格的匹配开销。
        if value is None or "交付日期" not in str(value):
            continue
        match = _TITLE_PATTERN.match(_text(value))
        if match:
            titles.append((match.group(1), cell.column))
    return titles


def _close_blocks(active, completed, title_columns=None) -> None:
    """结束指定列上的活动块；``None`` 表示在工作表结尾结束全部块。"""
    for block in list(active):
        if title_columns is not None and int(block["col"]) not in title_columns:
            continue
        active.remove(block)
        if block["rows"]:
            completed.append(block)


def _start_selected_blocks(active, title_cells, file_index, want_keys, next_order):
    """为人工选择范围内的新标题建立活动块，并返回更新后的顺序号。"""
    for batch, column in title_cells:
        key = "%d:%s" % (file_index, batch)
        if want_keys is not None and key not in want_keys:
            continue
        active.append({
            "batch": batch,
            "col": column,
            "rows": [],
            "order": next_order,
        })
        next_order += 1
    return next_order


def _block_row(block, row):
    """解析当前行在一个活动块中的状态和数据。

    返回 ``("append", 数据)``、``("finish", None)`` 或 ``("skip", None)``。
    显式状态让主扫描循环不再同时承担边界判断、字段提取和排除色识别。
    """
    start_column = int(block["col"])
    if start_column > len(row):
        return "skip", None
    code = _text(row[start_column - 1].value)
    if not code or not _CODE_PATTERN.match(code):
        return ("finish", None) if block["rows"] else ("skip", None)
    field_count = min(6, len(row) - start_column + 1)
    cells = [row[start_column - 1 + offset] for offset in range(field_count)]
    values = [cell.value for cell in cells]
    return "append", {
        "code": _text(values[0]),
        "name": _text(values[1]) if len(values) > 1 else "",
        "spec": _text(values[2]) if len(values) > 2 else "",
        "unit": _text(values[3]) if len(values) > 3 else "",
        "quantity": values[4] if len(values) > 4 else None,
        "supplier": _text(values[5]) if len(values) > 5 else "",
        "excluded": any(_is_yellow(cell) for cell in cells),
    }


def _iter_blocks(worksheet, file_index: int, want_keys: set[str] | None = None):
    """单趟扫描页签并按标题边界产出批次块。

    每个活动块记录批次号、1 基起始列、数据行和发现顺序。同一行可发现多个标题，
    支持客供件一类左右并排布局；同一列再次出现标题时，先结束上一纵向块，再决定
    新块是否属于人工选择范围。数据行要求块首列符合材料编码形态；已经收集到数据后
    遇到空值或非编码即结束该块。返回 ``(batch, start_column, rows)``，并保持标题
    发现顺序，而不是按批次号字典排序。

    ``want_keys`` 使用“文件序号:批次号”解决不同文件批次号相同的问题；未选择的块
    不进入活动列表，但它们的新标题仍参与旧块封口，防止数据跨批次串接。
    """
    active: list[dict[str, object]] = []
    completed: list[dict[str, object]] = []
    next_order = 0
    # max_column 可能为 None 或异常值，至少扫描第一列并限制到业务安全上限。
    used_column_count = max(1, int(getattr(worksheet, "max_column", 1) or 1))
    scan_column_count = min(used_column_count, _MAX_SCAN_COL)
    for row in worksheet.iter_rows(min_col=1, max_col=scan_column_count):
        title_cells = _row_title_cells(row)
        if title_cells:
            # 同一列出现新标题即宣告上一纵向块结束。新块即使未被选择，也必须先
            # 封口旧块，否则后续数据可能继续挂在上一批次名下。
            title_columns = {col0 for _, col0 in title_cells}
            _close_blocks(active, completed, title_columns)
            next_order = _start_selected_blocks(
                active, title_cells, file_index, want_keys, next_order,
            )
            continue
        finished = []
        for block in active:
            state, data = _block_row(block, row)
            if state == "append":
                # 扫描阶段保留排除行，以便界面准确展示总行数和排除数量。
                block["rows"].append(data)
            elif state == "finish":
                finished.append(block)
        for block in finished:
            active.remove(block)
            completed.append(block)
    _close_blocks(active, completed)
    for block in sorted(completed, key=lambda item: int(item["order"])):
        yield block["batch"], block["col"], block["rows"]


def _complete_rows_from_catalog(rows, resolver, fill_counts=None):
    """用主数据只补齐对账所需空字段，并原样返回行列表便于链式调用。"""
    for row in rows:
        # 材料编号是解析块时已确认的主键；正式主数据不得覆盖清单已有业务值。
        resolver.fill_mapping(
            row, fields=("name", "spec", "unit", "supplier"),
            counts=fill_counts)
    return rows


def _validated_source_paths(files) -> list[str]:
    """规范化并校验采购清单路径，避免扫描循环夹杂输入错误处理。"""
    source_paths = [
        os.path.abspath(str(value))
        for value in (files or [])
        if str(value).strip()
    ]
    if not source_paths:
        raise ValueError("请先上传采购清单文件")
    for path in source_paths:
        if not os.path.isfile(path):
            raise FileNotFoundError("找不到文件：%s" % path)
        if os.path.splitext(path)[1].lower() not in EXCEL_SUFFIXES:
            raise ValueError("仅支持 xlsx 或 xlsm 文件：%s" % os.path.basename(path))
    return source_paths


def _scan_source_file(path, file_index, resolver, fill_counts):
    """扫描单个采购清单，返回批次摘要和可用于供应商推断的非空名称。"""
    # 只读模式逐行扫描，避免把整份工作簿载入内存；data_only 取公式缓存值，
    # 调用方需先通过 warn_if_uncached 确认缓存存在。
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    batches: list[dict[str, object]] = []
    suppliers: list[str] = []
    try:
        for worksheet in workbook.worksheets:
            for batch, _, rows in _iter_blocks(worksheet, file_index):
                _complete_rows_from_catalog(rows, resolver, fill_counts)
                suppliers.extend(
                    str(item["supplier"])
                    for item in rows
                    if str(item["supplier"])
                )
                batches.append({
                    "batch": batch,
                    "sheet": worksheet.title,
                    "rows": len(rows),
                    "excluded_rows": sum(1 for item in rows if item["excluded"]),
                })
    finally:
        workbook.close()
    return batches, suppliers


def _supplier_hint(path, suppliers) -> str | None:
    """按文件名优先、数据众数其次的规则生成供应商预填建议。"""
    supplier = supplier_from_filename(os.path.basename(path))
    if supplier is not None or not suppliers:
        return supplier
    return Counter(suppliers).most_common(1)[0][0] or None


def scan(files, log=None, progress=None) -> dict[str, object]:
    """分析采购清单并返回供人工选择的文件、供应商和批次摘要。

    此阶段不生成任何对账单。每个文件会校验存在性与扩展名、检查公式缓存、遍历全部
    页签并统计每批总行数和排除行数。供应商先从文件名识别，失败后才取主数据补全后
    的行内众数。进度按完成文件数线性报告。
    """
    source_paths = _validated_source_paths(files)
    resolver = material_catalog.CatalogResolver()
    # 一批扫描共用解析器和统计，避免逐文件反复加载主数据库并可统一汇报补全效果。
    fill_counts: dict[str, int] = {}
    result_files: list[dict[str, object]] = []
    for index, path in enumerate(source_paths, start=1):
        # 采购数量可能来自公式，data_only 模式下先检查公式缓存，避免数量读到空值。
        common_core.warn_if_uncached(path, log, what="采购数量")
        batches, data_suppliers = _scan_source_file(
            path, index, resolver, fill_counts,
        )
        if not batches:
            raise ValueError("未在 %s 中识别到批次" % os.path.basename(path))
        result_files.append({
            "name": os.path.basename(path),
            "path": path,
            "supplier": _supplier_hint(path, data_suppliers),
            "batches": batches,
        })
        if log:
            log("已扫描 %s：%d 个批次" % (os.path.basename(path), len(batches)))
        if progress:
            progress(round(index / len(source_paths) * 100))
    material_catalog.log_fill_summary(log, "对账单预览", fill_counts)
    return {"files": result_files}


def _group_rows(path, batch_keys: set[str], file_index: int,
                resolver=None, fill_counts=None) -> list[dict[str, object]]:
    """重新读取单个文件，汇总被选批次中未标记排除的正式数据行。

    扫描结果只承担人工选择，正式生成时重新从原文件读取，避免在客户端传回或篡改
    大量明细。每行附加所属批次号，供输出报告展示。
    """
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    grouped: list[dict[str, object]] = []
    try:
        for worksheet in workbook.worksheets:
            for batch, _, rows in _iter_blocks(worksheet, file_index, want_keys=batch_keys):
                if resolver is not None:
                    _complete_rows_from_catalog(rows, resolver, fill_counts)
                for item in rows:
                    if item["excluded"]:
                        # 排除色是业务决定，正式输出阶段才真正丢弃这些行。
                        continue
                    grouped.append({**item, "batch": batch})
    finally:
        workbook.close()
    return grouped


def _normalize_build_inputs(files, selected, month):
    """规范化生成阶段输入，并在读取工作簿前给出明确的缺项错误。"""
    file_list = [
        os.path.abspath(str(value))
        for value in (files or [])
        if str(value).strip()
    ]
    selected_keys = [str(value) for value in (selected or []) if str(value).strip()]
    month_text = _text(month)
    if not file_list:
        raise ValueError("请选择采购清单文件")
    if not selected_keys:
        raise ValueError("请选择要制作对账单的批次")
    if not month_text:
        raise ValueError("请填写月份")
    return file_list, selected_keys, month_text


def _supplier_overrides(supplier_map) -> dict[int, str]:
    """把前端 JSON 的字符串键转换为 1 基文件序号映射。"""
    overrides: dict[int, str] = {}
    if not isinstance(supplier_map, dict):
        return overrides
    for key, value in supplier_map.items():
        try:
            overrides[int(key)] = _text(value)
        except (TypeError, ValueError):
            # 无法解释的键不会对应任何上传文件，忽略比误套到其他文件更安全。
            continue
    return overrides


def _selected_file_indexes(selected_keys) -> set[int]:
    """从“文件序号:批次号”选择键提取需要重新读取的文件序号。"""
    indexes = set()
    for key in selected_keys:
        prefix, separator, _batch = key.partition(":")
        if separator and prefix.isdigit():
            indexes.add(int(prefix))
    return indexes


def _supplier_for_rows(path, file_index, rows, overrides) -> str:
    """按人工映射、文件名、选中数据众数确定最终供应商名称。"""
    supplier = overrides.get(file_index) or supplier_from_filename(os.path.basename(path))
    if supplier:
        return supplier
    counts = Counter(str(item["supplier"]) for item in rows if str(item["supplier"]))
    return counts.most_common(1)[0][0] if counts else "未命名供应商"


def _collect_supplier_groups(file_list, selected_keys, overrides, resolver, fill_counts):
    """重新读取被选文件，把未排除数据按最终供应商归并。"""
    groups: dict[str, list[dict[str, object]]] = {}
    wanted_indexes = _selected_file_indexes(selected_keys)
    selected_set = set(selected_keys)
    for index, path in enumerate(file_list, start=1):
        # 只读取选择键中出现过的文件，未勾选文件不参与归并，保持生成范围与人工确认一致。
        if wanted_indexes and index not in wanted_indexes:
            continue
        rows = _group_rows(
            path, selected_set, index, resolver=resolver, fill_counts=fill_counts,
        )
        if not rows:
            continue
        supplier = _supplier_for_rows(path, index, rows, overrides)
        groups.setdefault(supplier, []).extend(rows)
    return groups


def _reconcile_output_dir(file_list, out_dir):
    """按统一路径策略解析输出目录，并确保显式目录已经创建。"""
    if out_dir is None:
        current = settings.get_settings()
        return paths.resolve_output_dir(
            "reconcile_statement",
            src_path=file_list[0],
            **current.output_kwargs(),
        )
    resolved = os.path.abspath(str(out_dir))
    # 显式目录由调用方指定，可能尚不存在；默认目录则由 resolve_output_dir 统一创建。
    os.makedirs(resolved, exist_ok=True)
    return resolved


def _write_supplier_outputs(groups, out_dir, month, log=None, progress=None):
    """按首次遇到供应商的顺序写出文件，并返回公开输出清单和总行数。"""
    outputs: list[dict[str, object]] = []
    total_rows = 0
    for group_index, (supplier, rows) in enumerate(groups.items(), start=1):
        target = _output_path(out_dir, supplier, month)
        _write_statement(target, supplier, month, rows)
        outputs.append({
            "path": target,
            "name": os.path.basename(target),
            "supplier": supplier,
            "month": month,
            "rows": len(rows),
        })
        total_rows += len(rows)
        if log:
            log("已生成 %s 对账单：%d 行" % (supplier, len(rows)))
        if progress:
            progress(round(group_index / len(groups) * 100))
    return outputs, total_rows


def build(files, selected, month, supplier_map=None, out_dir=None, log=None, progress=None) -> dict[str, object]:
    """按人工确认的批次生成对账单，并把同供应商数据合并到同一文件。

    ``selected`` 中每项为“1 基文件序号:批次号”；``supplier_map`` 是可选的文件序号
    到人工供应商名称映射，优先级高于文件名和行内数据。函数重新读取被选文件并排除
    标色行，使用主数据补空值，随后按最终供应商分组输出。月份只用于标题和文件名，
    为空时拒绝执行。返回输出清单和总数据行数。
    """
    file_list, selected_keys, month = _normalize_build_inputs(files, selected, month)
    resolver = material_catalog.CatalogResolver()
    fill_counts: dict[str, int] = {}
    groups = _collect_supplier_groups(
        file_list,
        selected_keys,
        _supplier_overrides(supplier_map),
        resolver,
        fill_counts,
    )
    if not groups:
        raise ValueError("所选批次中没有可用的数据行")
    material_catalog.log_fill_summary(log, "对账单", fill_counts)
    out_dir = _reconcile_output_dir(file_list, out_dir)
    outputs, total_rows = _write_supplier_outputs(
        groups, out_dir, month, log=log, progress=progress,
    )
    return {"out_dir": out_dir, "files": outputs, "total_rows": total_rows}


def _write_statement(target: str, supplier: str, month: str, rows: list[dict[str, object]]) -> None:
    """按固定八列表样式写出一份供应商月度对账单。

    标题合并 A1:H1，第二行为字段名，正文保留采购数量的原始数值类型；备注列留空供
    后续人工填写。这里不引入模板文件，避免部署包缺少外部模板时无法生成报告。
    """
    # 样式对象在同一工作簿内复用，减少逐单元格创建对象造成的文件体积膨胀。
    thin = Side(style="thin", color="9AA5B1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_font = Font(name="宋体", size=14, bold=True)
    head_fill = PatternFill("solid", fgColor="EAF1FF")
    head_font = Font(name="宋体", size=10, bold=True)
    cell_font = Font(name="宋体", size=10)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    # 列宽根据编号、名称和规格的典型长度固定设置，保证打开即可阅读。
    widths = {"A": 6, "B": 16, "C": 30, "D": 22, "E": 6, "F": 10, "G": 16, "H": 14}
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    worksheet.merge_cells("A1:H1")
    title = worksheet.cell(1, 1, "%s月%s对账单" % (month, supplier))
    title.font = title_font
    title.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[1].height = 26

    headers = ["序号", "材料编号", "材料名称", "规格", "单位", "采购数量", "批次号", "备注"]
    for column, name in enumerate(headers, start=1):
        cell = worksheet.cell(2, column, name)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for index, row in enumerate(rows, start=1):
        values = [
            index, row["code"], row["name"], row["spec"],
            row["unit"], row["quantity"], row["batch"], "",
        ]
        for column, value in enumerate(values, start=1):
            cell = worksheet.cell(index + 2, column, value)
            cell.font = cell_font
            cell.border = border
            if column in (1, 5, 6):
                # 序号、单位和数量采用居中，名称和规格保留默认左对齐以提高可读性。
                cell.alignment = Alignment(horizontal="center", vertical="center")
    workbook.save(target)
    workbook.close()
