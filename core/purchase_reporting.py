# -*- coding: utf-8 -*-
"""采购对账的 Excel 输出层。

本模块从 ``purchase_core.py`` 拆出，负责原表数量列上色、主数据补空写回和汇报单生成。
业务匹配、诊断和运行编排仍由 ``purchase_core.py`` 负责；本模块只在最后阶段打开并保存
工作簿，不参与匹配算法。
"""
import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

GREEN = PatternFill("solid", fgColor="FF92D050")   # 原表数量列绿色表示该行已找到唯一对家。
YELLOW = PatternFill("solid", fgColor="FFFFFF00")  # 原表数量列黄色表示该行仍需人工处理。
# ---------------------------------------------------------------------------
# 对账汇报单（并排）
# ---------------------------------------------------------------------------
def pair_note(a, b):
    """为已配对记录标记弱匹配依据；可信的干净配对返回空串。

    编号相等但批次仅靠单字符增删容错救回时提示“疑似笔误”；编号缺失且批次只近似时
    提示依据较弱；编号和批次都缺失则明确说明只凭名称、规格、数量匹配。可靠配对不加
    备注，避免真正需要处理的提示被大量正常行淹没。
    """
    from . import purchase_core  # 延迟导入，避免与 purchase_core 形成循环依赖。
    na, nb = purchase_core.norm_no(a["no"]), purchase_core.norm_no(b["no"])
    if na and nb:
        if not purchase_core.batch_compat(a["batch"], b["batch"]):
            return "批次疑似笔误(%s↔%s)，请核对" % (a["batch"], b["batch"])
        return ""
    # 至少一侧缺编号时，批次核心完全相等仍可视为较可靠；仅近似包含关系需要提示。
    if purchase_core.batch_core(a["batch"]) != purchase_core.batch_core(b["batch"]):
        return "无编号且批次仅近似(%s↔%s)，请核对" % (a["batch"], b["batch"])
    # 双方编号和批次均为空是最低可信度配对，必须由人工确认。
    if not purchase_core.batch_core(a["batch"]):
        return "无编号且无批次，仅按名称+规格+数量匹配，请核对"
    return ""


_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEAD_FILL = PatternFill("solid", fgColor="FF4472C4")
_HEAD_FONT = Font(bold=True, color="FFFFFFFF")
_SEC_FILL = PatternFill("solid", fgColor="FFD9E1F2")
_MATCH_FILL = PatternFill("solid", fgColor="FFE2EFDA")    # 汇报单淡绿表示可靠匹配。
_UNMATCH_FILL = PatternFill("solid", fgColor="FFFFF2CC")  # 汇报单淡黄表示未匹配。
_FLAG_FILL = PatternFill("solid", fgColor="FFFCE4D6")     # 汇报单淡橙突出需人工核对的弱匹配。
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _put(ws, r, c, val, fill=None, font=None, align=None, border=True):
    """写入报告单元格并统一应用可选底色、字体、对齐和边框，返回该单元格。"""
    cell = ws.cell(r, c, val)
    if fill:
        cell.fill = fill
    cell.font = font or Font()  # 每个单元格使用独立样式对象，避免共享样式被意外修改。
    cell.alignment = align or _CENTER
    if border:
        cell.border = _BORDER
    return cell


def _fill_empty_material_fields(worksheet, row_info, column_map):
    """把主数据补全值写回原表空单元格，绝不覆盖源文件已有内容。"""

    for role in ("name", "spec", "unit"):
        column = column_map.get(role)
        value = row_info.get(role)
        if not column or value is None or not str(value).strip():
            continue
        cell = worksheet.cell(row_info["r"], column)
        if cell.value is None or not str(cell.value).strip():
            cell.value = value


def _save_colored_workbook(workbook, out_path):
    """保存上色副本，并把 Windows 文件占用错误转换成可执行提示。"""

    try:
        workbook.save(out_path)
        return out_path
    except PermissionError:
        raise PermissionError("无法保存 %s —— 请先在 Excel 里关闭该文件后重试" % out_path)


def apply_colors(path, sheet, matched, rows, qty_col, out_path, col_map=None):
    """
    在原工作簿的数量列标记匹配状态，并把主数据补全字段写回空单元格后另存副本。

    此处重新以普通可写模式打开原文件，以保留公式、图片和样式；``rows`` 中的行号来自
    ``load_rows``，因此颜色准确落到原业务行。名称、规格、单位只在原单元格为空时补入，
    严格遵守主数据不覆盖源文件已有值的规则。
    """
    # 上色必须保留原表样式、公式和图片，因此用普通模式完整重开，不能使用只读流式加载。
    wb = openpyxl.load_workbook(path)
    try:
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        columns = col_map or {}
        for flag, info in zip(matched, rows):
            # matched 与 rows 按相同输入顺序构造，zip 可直接把结果映射回原始行号。
            ws.cell(info["r"], qty_col).fill = GREEN if flag else YELLOW
            _fill_empty_material_fields(ws, info, columns)
        return _save_colored_workbook(wb, out_path)
    finally:
        wb.close()  # 确保 Windows 文件句柄释放，失败后用户可关闭文件并立即重试。


# 汇报单列布局：我方 5 列 + 供方 5 列 + 备注；共 12 列
_RH = ["序号",
       "编号", "名称", "规格", "批次", "数量",
       "编号", "名称", "规格", "批次", "数量",
       "备注"]


def _write_row(ws, r, seq, a, b, fill, note):
    """写入一行左右并排明细；任一侧为 ``None`` 时保留该侧空白以维持固定列结构。"""
    _put(ws, r, 1, seq, fill)
    cols_a = ("no", "name", "spec", "batch", "qty")
    for k, key in enumerate(cols_a):
        _put(ws, r, 2 + k, a[key] if a else "", fill,
             align=_LEFT if key in ("name", "spec") else _CENTER)
    for k, key in enumerate(cols_a):
        _put(ws, r, 7 + k, b[key] if b else "", fill,
             align=_LEFT if key in ("name", "spec") else _CENTER)
    # 备注存在时单独使用核对色，即使整行属于已匹配，也能把弱依据从绿色记录中突出。
    _put(ws, r, 12, note, _FLAG_FILL if note else fill, align=_LEFT)


def build_report(rows1, rows2, m1, m2, pairs, out_path, name1, name2):
    """
    创建采购对账汇报工作簿的标题和总体统计，再委托辅助函数写入三个明细章节。

    报告将已匹配双方并排展示，随后分别列出两侧未匹配记录及诊断原因，使人工无需在
    两个原文件间来回查找。``name1`` 和 ``name2`` 仅作为显示名称，不影响匹配逻辑。
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "对账汇报单"
    ncol = len(_RH)

    # 标题跨越固定 12 列；即使双方名称较长也能保持一个完整标题区域。
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    _put(ws, 1, 1, "采购数对账汇报单  （%s ↔ %s）" % (name1, name2),
         _HEAD_FILL, Font(bold=True, size=14, color="FFFFFFFF"))
    ws.row_dimensions[1].height = 26
    # 第二行用双方总行数、匹配数和未匹配数提供报告级快速核验。
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
    _put(ws, 2, 1,
         "%s %d 行 / 对上 %d / 未对上 %d      %s %d 行 / 对上 %d / 未对上 %d      配对 %d 对"
         % (name1, len(rows1), sum(m1), len(m1) - sum(m1),
            name2, len(rows2), sum(m2), len(m2) - sum(m2), len(pairs)),
         _SEC_FILL, Font(bold=True), align=_LEFT)
    return _finish_report(ws, wb, rows1, rows2, m1, m2, pairs,
                          out_path, name1, name2, ncol)


def _section(ws, r, text, ncol):
    """写入跨越报告全宽的章节标题，并返回下一可写行号。"""
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
    _put(ws, r, 1, text, _SEC_FILL, Font(bold=True), align=_LEFT)
    return r + 1


def _header_band(ws, r):
    """写入“双方分组 + 各字段”的两级表头，并返回表头后的首个数据行号。"""
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=1)
    _put(ws, r, 1, "序号", _HEAD_FILL, _HEAD_FONT)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    _put(ws, r, 2, "我方", _HEAD_FILL, _HEAD_FONT)
    ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=11)
    _put(ws, r, 7, "供方", _HEAD_FILL, _HEAD_FONT)
    ws.merge_cells(start_row=r, start_column=12, end_row=r + 1, end_column=12)
    _put(ws, r, 12, "备注", _HEAD_FILL, _HEAD_FONT)
    # 序号和备注纵向合并，只有双方的五个业务字段需要第二级标题。
    for c in range(2, 12):
        _put(ws, r + 1, c, _RH[c - 1], _HEAD_FILL, _HEAD_FONT)
    return r + 2


def _finish_report(ws, wb, rows1, rows2, m1, m2, pairs,
                   out_path, name1, name2, ncol):
    """
    写入已匹配、我方未匹配和供方未匹配三个章节，设置版式并保存报告。

    未匹配诊断只把另一侧的未匹配列表作为具体候选，保证原因说明与报告中可见记录一致。
    已匹配记录按我方原始行号排序，未匹配记录保持各自源表顺序，便于回查。
    """
    from . import purchase_core  # 延迟导入，避免 purchase_reporting 与 purchase_core 循环依赖。
    r = 4
    r = _section(ws, r, "一、已对上（%d 对）" % len(pairs), ncol)
    hb = _header_band(ws, r)
    # 两级表头先创建通用分组，再用调用方名称覆盖，避免辅助函数承担额外参数。
    ws.cell(r, 2).value = name1
    ws.cell(r, 7).value = name2
    r = hb
    seq = 1
    for i, j, _s in sorted(pairs, key=lambda p: rows1[p[0]]["r"]):
        a, b = rows1[i], rows2[j]
        note = pair_note(a, b)
        _write_row(ws, r, seq, a, b, _FLAG_FILL if note else _MATCH_FILL, note)
        r += 1
        seq += 1

    # 先构造双方未匹配池，后续诊断不会引用已经在第一章节使用过的对家。
    um1 = [row for row, ok in zip(rows1, m1) if not ok]
    um2 = [row for row, ok in zip(rows2, m2) if not ok]
    r = _section(ws, r + 1, "二、%s 未对上（%d 条）" % (name1, len(um1)), ncol)
    r = _header_band(ws, r)
    ws.cell(r - 2, 2).value = name1
    for row in um1:
        _write_row(ws, r, row["r"], row, None, _UNMATCH_FILL,
                   purchase_core.diagnose(row, rows2, um2))
        r += 1

    # 对供方记录执行对称诊断，左右展示位置仍固定为“我方列 / 供方列”。
    r = _section(ws, r + 1, "三、%s 未对上（%d 条）" % (name2, len(um2)), ncol)
    r = _header_band(ws, r)
    ws.cell(r - 2, 7).value = name2
    for row in um2:
        _write_row(ws, r, row["r"], None, row, _UNMATCH_FILL,
                   purchase_core.diagnose(row, rows1, um1))
        r += 1

    # 备注和规格需要更宽空间，编号/批次/数量保持紧凑，控制报告总体宽度。
    widths = [6, 13, 14, 16, 14, 7, 13, 14, 16, 14, 7, 30]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A4"
    try:
        wb.save(out_path)
        return out_path
    except PermissionError:
        # 将操作系统级文件占用错误转成用户能够处理的 Excel 关闭提示。
        raise PermissionError("无法保存 %s —— 请先在 Excel 里关闭该文件后重试" % out_path)


