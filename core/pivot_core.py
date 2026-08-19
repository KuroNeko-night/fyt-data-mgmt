# -*- coding: utf-8 -*-
"""采购数据的识别、清洗、复核、单位归并和静态汇总。

旧版本为了模拟人工制作透视表，额外维护了“数量”字段、原生 Excel 透视缓存、OOXML
关系文件和一个重复的“汇总”列。当前业务真正需要的是：

1. 从同一工作表中的多个横向子表直接提取编码、名称、规格、单位和最终采购数量；
2. 按既有规则清洗版本与无效数量；
3. 统一规格和单位，再按物料属性聚合；
4. 输出清洗数据子表和可直接维护的采购主表。

本模块保留原 ``analyze``、``apply_plan``、``run`` 等双阶段公开入口，方便 Web/Tauri
继续使用人工复核协议；但最终文件是普通静态工作簿，不再生成必须依赖 Excel 刷新的透视对象。
"""

from __future__ import annotations

import datetime
import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from . import common_core
from . import incremental_cache
from . import material_catalog
from . import paths as _paths
from . import pivot_reporting
from . import settings as _settings
from .common_core import warn_if_uncached

# 保留分析层和聚类层的公开名称，旧桥接/测试仍可从 ``pivot_core`` 访问。
from .pivot_analysis import (  # noqa: E402,F401
    CODE_ALIASES,
    EXCLUDE_SHEET_TOKENS,
    FIELD_CN,
    FINAL_ALIASES,
    HEADER_SCAN_ROWS,
    INFO_FIELDS,
    KEEP_TOKEN,
    KEY_FIELDS,
    L_CODE,
    L_FINAL,
    L_NAME,
    L_SPEC,
    L_UNIT,
    L_VER,
    MAX_BLOCKS,
    NAME_ALIASES,
    NAME_EXCLUDE,
    NAME_PACKAGING,
    NAME_ZUTUO,
    SPEC_ALIASES,
    UNIT_ALIASES,
    VER_ALIASES,
    _PreviewCell,
    _SheetPreview,
    _analyze_stream_sheet,
    _assign_block_column,
    _block_from_anchor,
    _cell,
    _classify_by_name_and_cols,
    _complete_rows_from_catalog,
    _contains_any,
    _final_has_qty,
    _has_chinese,
    _has_token,
    _header_anchor_columns,
    _is_excluded_sheet,
    _is_valid_code,
    _is_zero,
    _last_col,
    _looks_like_pivot_output,
    _match_anchor,
    _norm,
    _preview_sheet,
    _sheet_name,
    analyze_workbooks,
    classify_sheet,
    clean_rows,
    clean_rows_ex,
    find_all_blocks,
    is_data_sheet,
    normalize_rows,
    normalize_stream_rows,
)
from .pivot_clustering import (  # noqa: E402,F401
    F_CODE,
    F_FINAL,
    F_NAME,
    F_SPEC,
    F_UNIT,
    F_VER,
    _code_order_key,
    _compute_unit_best,
    _is_blank,
    _is_compound_unit,
    _name_unit_prior,
    _norm_key,
    _num,
    _spec_base,
    _spec_gkey,
    _spec_keyof,
    _unit_gkey,
    _unit_key_sample,
    _unit_simplicity,
    aggregate,
    compute_spec_canon,
    compute_unit_best,
    drop_blank_code_rows,
    unify_specs,
    unify_units,
)


# 可信度分析仍是结果投影的一部分；这里保留原名称以兼容旧调用。
assess_confidence = pivot_reporting.assess_confidence
write_confidence_report = pivot_reporting.write_confidence_report


# 输出工作簿只保留两个稳定页签：第一张是可维护主表，第二张是清洗后的明细子表。
SUMMARY_SHEET = "采购汇总"
CLEAN_SHEET = "清洗数据"
L_SUPPLIER = "供应商"
L_DIFF = "差异"
L_RECEIVED = "实收"
L_DATE = "日期"
SUMMARY_HEADERS = [L_CODE, L_NAME, L_SPEC, L_UNIT, L_FINAL,
                   L_SUPPLIER, L_DIFF, L_RECEIVED, L_DATE]
CLEAN_HEADERS = [L_CODE, L_NAME, L_SPEC, L_UNIT, L_FINAL]


# 输出样式集中在此处，避免把表格排版散落到清洗和聚合逻辑中。
_FONT_NAME = "微软雅黑"
_THIN = Side(style="thin", color="D9E1F2")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_FONT = Font(name=_FONT_NAME, size=10)
_FONT_BOLD = Font(name=_FONT_NAME, size=10, bold=True)
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
_TOTAL_FILL = PatternFill(fill_type="solid", fgColor="EAF2F8")


def _style_cell(cell, *, header=False, total=False):
    """给输出单元格应用统一边框、字体、对齐和状态填充。"""
    cell.border = _BORDER
    cell.alignment = _CENTER
    cell.font = _FONT_BOLD if header or total else _FONT
    if header:
        cell.fill = _HEADER_FILL
    elif total:
        cell.fill = _TOTAL_FILL


def _prepare_sheet(worksheet):
    """清空工作表的旧合并和旧内容，保证输出页不残留模板布局。"""
    for merged in list(worksheet.merged_cells.ranges):
        worksheet.unmerge_cells(str(merged))
    if worksheet.max_row:
        worksheet.delete_rows(1, worksheet.max_row)
    worksheet.sheet_view.showGridLines = False


def _finish_table(worksheet, widths, filter_end=None):
    """设置冻结表头、列宽和可选筛选范围，提升普通静态表的可用性。"""
    worksheet.freeze_panes = "A2"
    for column, width in enumerate(widths, start=1):
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = width
    if filter_end:
        worksheet.auto_filter.ref = filter_end


def write_clean_sheet(worksheet, rows):
    """写出五列清洗子表，不再保留内部清洗用的版本和普通数量列。

    版本序号只用于判断是否删除行，普通“数量”列不参与本业务的清洗、聚合或输出；因此
    明细页只保留物料号、名称、规格、单位和最终采购数量五列，正好对应业务后续需要的
    数据集合。行顺序保持来源区块顺序，方便回溯原始表。
    """
    _prepare_sheet(worksheet)
    for column, header in enumerate(CLEAN_HEADERS, start=1):
        cell = worksheet.cell(row=1, column=column, value=header)
        _style_cell(cell, header=True)
    for row_index, record in enumerate(rows, start=2):
        values = [record[F_CODE], record[F_NAME], record[F_SPEC],
                  record[F_UNIT], record[F_FINAL]]
        for column, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row_index, column=column, value=value)
            _style_cell(cell)
    last_row = max(1, len(rows) + 1)
    _finish_table(worksheet, [18, 24, 24, 12, 16], f"A1:E{last_row}")


def write_summary_sheet(worksheet, aggregated, quantity_check=None):
    """写出聚合后的采购主表和可人工填写的补充列。

    主表不再伪装成 Excel 透视表：E 列直接保存聚合后的最终采购数量；供应商、实收和日期
    是后续维护列；差异列只根据“实收 - 最终采购数量”计算，实收为空时保持空白，避免在
    尚未填写实收时显示误导性的负数。总计行只保留一个采购数量合计，不重复制造“汇总”列。
    """
    _prepare_sheet(worksheet)
    for column, header in enumerate(SUMMARY_HEADERS, start=1):
        cell = worksheet.cell(row=1, column=column, value=header)
        _style_cell(cell, header=True)

    for row_index, (code, name, spec, unit, quantity) in enumerate(aggregated, start=2):
        values = [code, name, spec, unit, quantity, "", None, "", ""]
        for column, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row_index, column=column, value=value)
            _style_cell(cell)
        # 实收为空时不展示差异，填写实收后 Excel 自动计算差异。
        worksheet.cell(row=row_index, column=7, value=f'=IF(H{row_index}="","",H{row_index}-E{row_index})')
        _style_cell(worksheet.cell(row=row_index, column=7))

    data_end = len(aggregated) + 1
    total_row = data_end + 1
    total_cell = worksheet.cell(row=total_row, column=1, value="总计")
    _style_cell(total_cell, total=True)
    # 总计写静态值，避免普通预览器因不执行 Excel 公式而显示空白；每个物料行的差异公式
    # 仍保留，以便管理员填写“实收”后能在 Excel 中即时得到差异。
    total_value = worksheet.cell(
        row=total_row,
        column=5,
        value=sum(float(item[4]) for item in aggregated),
    )
    _style_cell(total_value, total=True)
    for column in range(2, len(SUMMARY_HEADERS) + 1):
        _style_cell(worksheet.cell(row=total_row, column=column), total=True)
    if quantity_check:
        check_row = total_row + 2
        check_values = (
            (1, "来源最终采购数", True),
            (2, quantity_check["source_total"], False),
            (4, "最终表采购数", True),
            (5, quantity_check["output_total"], False),
            (7, "自检结果", True),
            (8, quantity_check["status"], False),
            (9, "差异 %s" % _fmt_num(quantity_check["difference"]), False),
        )
        for column, value, label in check_values:
            cell = worksheet.cell(row=check_row, column=column, value=value)
            _style_cell(cell, header=label, total=not label)
    _finish_table(
        worksheet,
        [18, 24, 24, 12, 16, 22, 14, 14, 14],
        f"A1:I{max(1, data_end)}",
    )


def _safe_sheet_name(workbook, base):
    """生成不含 Excel 禁用字符且不重复的工作表名。"""
    for character in "[]:*?/\\":
        base = base.replace(character, " ")
    base = base.strip() or "数据"
    base = base[:28]
    name = base
    suffix = 1
    while name in workbook.sheetnames:
        suffix += 1
        text = str(suffix)
        name = base[:31 - len(text)] + text
    return name


def process_workbook(in_path, out_path=None, log=None):
    """兼容旧的一步式入口，直接走统一的分析、清洗和静态汇总流程。"""
    if out_path is None:
        directory = os.path.dirname(in_path)
        stem = os.path.splitext(os.path.basename(in_path))[0]
        out_path = os.path.join(directory, stem + "_采购汇总.xlsx")
    plan = analyze_workbooks([in_path])
    return apply_plan(plan, _default_choices(plan), out_path, log=log)


def _default_choices(plan):
    """根据分析计划生成默认选择，保持旧双阶段协议的自动路径。"""
    return {
        "sheets": {sheet["id"]: bool(sheet["use"]) for sheet in plan["sheets"]},
        "held": {(item["sid"], item["ridx"]): False for item in plan["held_index"]},
        "unit_overrides": {},
        "spec_overrides": {},
    }


def _sheet_audit_record(sheet, use):
    """构造页签审计记录，后续只补充最终行数和人工恢复数。"""
    return {
        "file": sheet["file"],
        "sheet": sheet["sheet"],
        "use": use,
        "kind": sheet["kind"],
        "confidence": sheet["confidence"],
        "reason": sheet["reason"],
        "cols": sheet["cols"],
        "missing": sheet["missing"],
        "rows": 0,
        "d1": sheet["d1"],
        "d2": sheet["d2"],
        "held_kept": 0,
    }


def _selected_sheet_rows(sheet, selected_held):
    """复制默认保留行，并追加人工明确恢复的疑似删除行。"""
    rows = [list(row) for row in sheet["kept"]]
    kept_count = 0
    kept_total = 0.0
    for row_index, held in enumerate(sheet["held"]):
        if not selected_held.get((sheet["id"], row_index), False):
            continue
        row = list(held["rec"])
        rows.append(row)
        kept_count += 1
        parsed = _num(row[F_FINAL])
        if parsed is not None:
            kept_total += float(parsed)
    return rows, kept_count, kept_total


def _collect_selected_plan_rows(plan, choices):
    """收集最终纳入的页签与记录，并形成稳定的审计汇总。"""
    selected_sheets = choices.get("sheets", {})
    selected_held = choices.get("held", {})
    summary = {
        "audit": [],
        "detail": [],
        "rows": [],
        "processed": 0,
        "skipped": 0,
        "d1": 0,
        "d2": 0,
        "held_kept_n": 0,
        "held_kept_total": 0.0,
        "source_total": 0.0,
        "source_unparsed": 0,
    }

    for sheet in plan["sheets"]:
        use = selected_sheets.get(sheet["id"], sheet["use"])
        audit_record = _sheet_audit_record(sheet, use)
        if not use:
            summary["audit"].append(audit_record)
            summary["skipped"] += 1
            continue

        rows, kept_count, kept_total = _selected_sheet_rows(sheet, selected_held)
        summary["held_kept_n"] += kept_count
        summary["held_kept_total"] += kept_total
        if not rows:
            audit_record.update(use=False, kind="排除:未选中任何行")
            summary["audit"].append(audit_record)
            summary["skipped"] += 1
            continue

        summary["rows"].extend(rows)
        # 自检基准取人工最终选中的源记录，发生清洗、空编码剔除或聚合之前先保存总和。
        # 这样可以发现“源表有数量、最终表却少了数量”的静默丢失，而不是只验证聚合函数自身。
        for row in rows:
            parsed = _num(row[F_FINAL])
            if parsed is None:
                value = row[F_FINAL]
                if value is not None and str(value).strip():
                    summary["source_unparsed"] += 1
                continue
            summary["source_total"] += float(parsed)
        summary["d1"] += sheet["d1"]
        summary["d2"] += sheet["d2"]
        audit_record.update(rows=len(rows), held_kept=kept_count)
        summary["audit"].append(audit_record)
        summary["detail"].append((
            "%s / %s" % (sheet["file"], sheet["sheet"]),
            len(rows),
            sheet["d1"],
            sheet["d2"],
        ))
        summary["processed"] += 1
    return summary


def _quantity_check(source_total, output_total, source_unparsed=0):
    """比较源记录与最终主表的最终采购数，返回可供页面和报告复用的自检结果。

    数量字段允许整数、小数和常见格式化文本，因此比较使用小容差；无法解析的非空源值
    会让检查进入“无法确认”，避免把未知数据误报成通过。差异方向统一为“最终表 - 源表”。
    """
    difference = float(output_total) - float(source_total)
    tolerance = 1e-6
    equal = abs(difference) <= tolerance
    if source_unparsed:
        status = "无法确认"
        passed = False
        message = "源数据中有 %d 个非空最终采购数无法解析，不能确认汇总完整性。" % source_unparsed
    elif equal:
        status = "通过"
        passed = True
        message = "最终表最终采购数汇总与选定数据来源一致。"
    else:
        status = "异常"
        passed = False
        message = "最终表与选定数据来源的最终采购数汇总不一致，请检查被清洗或未纳入的记录。"
    return {
        "status": status,
        "passed": passed,
        "source_total": source_total,
        "output_total": output_total,
        "difference": difference,
        "tolerance": tolerance,
        "source_unparsed": source_unparsed,
        "message": message,
    }


def _build_output_workbook(selected_rows, choices, source_total, source_unparsed):
    """直接构造“采购汇总 + 清洗数据”两页工作簿，不创建透视对象或中间页。"""
    workbook = openpyxl.Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = _safe_sheet_name(workbook, SUMMARY_SHEET)
    clean_sheet = workbook.create_sheet(title=_safe_sheet_name(workbook, CLEAN_SHEET))

    rows = [list(row) for row in selected_rows]
    if rows:
        rows = unify_specs(rows, overrides=choices.get("spec_overrides") or None)
        rows = unify_units(rows, overrides=choices.get("unit_overrides") or None)
        rows, _blank_code_count = drop_blank_code_rows(rows)
    aggregated = aggregate(rows)
    total = sum(float(item[4]) for item in aggregated)
    quantity_check = _quantity_check(source_total, total, source_unparsed)
    write_summary_sheet(summary_sheet, aggregated, quantity_check=quantity_check)
    write_clean_sheet(clean_sheet, rows)
    return workbook, rows, len(aggregated), total, quantity_check


def _build_apply_result(plan, choices, out_path, selection, clean_rows_count,
                        groups, total, quantity_check):
    """组装双端稳定返回协议，并附加可信度结论和人工复核摘要。"""
    result = {
        "processed": selection["processed"],
        "skipped": selection["skipped"],
        "sheets": selection["detail"],
        "out": out_path,
        "files": plan["files"],
        "groups": groups,
        "total": total,
        "source_check": quantity_check,
        "d1": selection["d1"],
        "d2": selection["d2"],
        "audit": selection["audit"],
        "clean_rows": clean_rows_count,
        "review": {
            "plan": plan,
            "choices": choices,
            "details_cached": True,
            "held_kept_n": selection["held_kept_n"],
            "held_kept_total": selection["held_kept_total"],
            "held_total_n": len(plan["held_index"]),
            "unit_conflicts": plan["unit_conflicts"],
            "spec_merges": plan["spec_merges"],
        },
    }
    result.update(assess_confidence(result))
    result["report"] = ""
    return result


def apply_plan(plan, choices, out_path, log=None):
    """按人工最终选择清洗、归并、聚合并写出静态采购结果。"""
    selection = _collect_selected_plan_rows(plan, choices)
    workbook, rows, groups, total, quantity_check = _build_output_workbook(
        selection["rows"],
        choices,
        selection["source_total"],
        selection["source_unparsed"],
    )
    # 只落盘普通工作表；不再有动态透视注入阶段，也不会产生半份 OOXML 文件。
    workbook.save(out_path)
    if log:
        if quantity_check["passed"]:
            log("[自检] 最终采购数汇总校验通过：源数据=%s，最终表=%s。"
                % (_fmt_num(quantity_check["source_total"]), _fmt_num(total)))
        else:
            log("[自检] 最终采购数汇总校验%s：源数据=%s，最终表=%s，差异=%s。"
                % (quantity_check["status"], _fmt_num(quantity_check["source_total"]),
                   _fmt_num(total), _fmt_num(quantity_check["difference"])))
    if log:
        log("已生成采购汇总主表和清洗数据子表：%s" % os.path.basename(out_path))
    return _build_apply_result(
        plan, choices, out_path, selection, len(rows), groups, total, quantity_check,
    )


def process_workbooks(in_paths, out_path, choices=None):
    """兼容旧调用的一步式入口；未传选择时使用分析阶段的默认决策。"""
    plan = analyze_workbooks(in_paths)
    return apply_plan(plan, choices or _default_choices(plan), out_path)


def _fmt_num(value):
    """把数量格式化为整数或最多四位小数。"""
    try:
        number = float(value)
        return str(int(number)) if number == int(number) else ("%.4f" % number).rstrip("0").rstrip(".")
    except (TypeError, ValueError):
        return "" if value is None else str(value)


def _cacheable_result(result):
    """生成不包含逐行计划和元组键的轻量缓存快照。"""
    compact = dict(result)
    review = result.get("review")
    if isinstance(review, dict):
        compact["review"] = {
            "details_cached": False,
            "held_kept_n": review.get("held_kept_n", 0),
            "held_kept_total": review.get("held_kept_total", 0),
            "held_total_n": review.get("held_total_n", 0),
            "unit_conflicts": review.get("unit_conflicts", []),
            "spec_merges": review.get("spec_merges", []),
        }
    return compact


def _materialize_web_cache(cached, out_dir):
    """把跨任务缓存的输出文件复制到当前 Web 任务目录。"""
    import shutil

    result = dict(cached)
    for key in ("out", "report"):
        source = result.get(key)
        if not source or not os.path.isfile(source):
            continue
        target = common_core.unique_path(os.path.join(out_dir, os.path.basename(source)))
        shutil.copy2(source, target)
        result[key] = target
    result["out_dir"] = out_dir
    return result


def _prepare_run(in_paths, out_dir):
    """整理入口参数并计算缓存作用域。"""
    settings = _settings.get_settings()
    paths = [in_paths] if isinstance(in_paths, str) else list(in_paths)
    resolver = material_catalog.CatalogResolver()
    web_root = os.environ.get("FYT_WEB_OUTPUT_ROOT", "").strip()
    if out_dir is not None:
        cache_scope = os.path.abspath(out_dir)
    elif web_root:
        cache_scope = "web-user-cache"
    else:
        cache_scope = {"mode": settings.output_mode, "custom_root": settings.custom_output_root}
    return settings, paths, resolver, web_root, cache_scope


def _try_cached_pivot(cache_key, web_root, out_dir, settings, log, progress):
    """读取缓存并在 Web 场景物化输出文件；缓存异常统一回退完整计算。"""
    try:
        cached = incremental_cache.get(cache_key)
        if not cached:
            return None
        if web_root:
            target_dir = _paths.resolve_output_dir("pivot", **settings.output_kwargs())
            cached = _materialize_web_cache(cached, target_dir)
        cached["cache_hit"] = True
        log("[缓存] 输入文件和处理参数未变化，已复用现有采购汇总结果。")
        if progress:
            progress(100)
        return cached
    except (OSError, ValueError, TypeError) as error:
        log("[缓存] 无法读取缓存，已回退完整处理：%s" % error)
        return None


def _build_pivot_result(paths, choices, out_dir, resolver, log, progress):
    """执行缓存未命中时的分析、复核选择和静态工作簿输出。"""
    fill_counts: dict[str, int] = {}
    progress_reporter = common_core.Progress(progress, stages=[("analyze", 55), ("apply", 45)])
    output_name = "%s采购汇总结果.xlsx" % _beijing_date()
    out_path = common_core.unique_path(os.path.join(out_dir, output_name))
    for path in paths:
        warn_if_uncached(path, log, what="最终采购数量")
    log("① 分析 %d 个文件..." % len(paths))
    progress_reporter.stage("analyze")
    plan = analyze_workbooks(
        paths,
        on_file=progress_reporter.tick,
        resolver=resolver,
        fill_counts=fill_counts,
    )
    material_catalog.log_fill_summary(log, "采购汇总", fill_counts)
    choices = choices if choices is not None else _default_choices(plan)
    log("② 清洗、统一单位并聚合...")
    progress_reporter.stage("apply")
    result = apply_plan(plan, choices, out_path, log=log)
    result["out_dir"] = out_dir
    progress_reporter.done()
    log("   分组 %d 项，合计 %s；可信度【%s】%d/100"
        % (result.get("groups", 0), _fmt_num(result.get("total", 0)),
           result.get("level", "?"), result.get("score", 0)))
    log("已保存：%s" % out_path)
    return result


def _save_pivot_cache(cache_key, result, log):
    """保存采购汇总结果索引；缓存失败不影响已经生成的工作簿。"""
    if not cache_key:
        return
    artifacts = [path for path in (result.get("out"), result.get("report")) if path]
    try:
        incremental_cache.put(cache_key, "pivot", _cacheable_result(result), artifacts)
    except (OSError, ValueError, TypeError) as error:
        log("[缓存] 结果索引保存失败，不影响本次输出：%s" % error)


def run(in_paths, choices=None, out_dir=None, log=None, progress=None):
    """执行采购汇总并返回输出、可信度和结构化复核结果。"""
    log = log or (lambda *args, **kwargs: None)
    settings, paths, resolver, web_root, cache_scope = _prepare_run(in_paths, out_dir)
    cache_key = ""
    if settings.get("enable_incremental_cache", True):
        try:
            cache_key = incremental_cache.make_key(
                "pivot",
                paths,
                {"choices": choices, "output": cache_scope, "catalog": resolver.signature},
                # 自检结果属于正式输出协议；升级版本可防止旧缓存绕过来源数量勾稽。
                engine_version="pivot-v5-source-quantity-check",
            )
            cached = _try_cached_pivot(cache_key, web_root, out_dir, settings, log, progress)
            if cached is not None:
                return cached
        except (OSError, ValueError, TypeError) as error:
            log("[缓存] 无法读取缓存，已回退完整处理：%s" % error)
    if out_dir is None:
        out_dir = _paths.resolve_output_dir("pivot", **settings.output_kwargs())
    result = _build_pivot_result(paths, choices, out_dir, resolver, log, progress)
    _save_pivot_cache(cache_key, result, log)
    return result


def analyze(in_paths, log=None, progress=None):
    """只执行第一阶段分析，返回工作表、疑似删除行和规格/单位复核计划。"""
    if isinstance(in_paths, str):
        in_paths = [in_paths]
    if log:
        log("正在分析 %d 个文件…" % len(in_paths))

    on_file = None
    if progress:
        def on_file(done, total):
            """把按文件完成数转换为整数进度。"""
            progress(int(done * 100 / total) if total else 100)

    resolver = material_catalog.CatalogResolver()
    fill_counts: dict[str, int] = {}
    result = analyze_workbooks(
        in_paths,
        on_file=on_file,
        resolver=resolver,
        fill_counts=fill_counts,
    )
    material_catalog.log_fill_summary(log, "采购汇总", fill_counts)
    return result


def _beijing_date():
    """返回北京时间的紧凑日期字符串，用于输出文件命名。"""
    return (datetime.datetime.now(datetime.timezone.utc) +
            datetime.timedelta(hours=8)).strftime("%Y%m%d")
