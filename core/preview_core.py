# -*- coding: utf-8 -*-
"""
表格文件轻量预览核心
====================
为桌面侧栏和其他快速查看入口读取文件前若干行，而不启动 Excel 或执行业务处理。
xlsx/xlsm 使用 openpyxl 只读流，xls 使用 xlrd，CSV/TXT 探测常见中文编码和分隔符；
所有格式统一返回 ``PreviewData``，界面只负责渲染二维文本。

预览默认限制 30 行、40 列，避免超宽或超大文件拖慢界面。xlsx 会重置错误的工作表
dimension，兼容第三方导出把实际大表标成 ``A1`` 的问题；仍在达到行数上限后立即停止。
CSV 只抽样前 4 KiB/2 KiB 探测编码与方言。任何读取错误都封装到 ``error`` 字段，
不会从预览辅助功能向上抛出并导致主界面崩溃。

模块绝不写入源文件，也不保证预览可替代正式业务读取：公式只展示缓存值，截断标记在
流式格式中是保守估计，正式处理仍应使用各业务核心的严格解析器。
"""
import os
import csv
import datetime

# 行数包含标题和表头；列上限防止异常使用范围或超宽表把界面网格撑爆。
DEFAULT_ROWS = 30
DEFAULT_COLS = 40


class PreviewData(object):
    """一张页签的不可依赖格式预览快照。

    ``rows`` 已全部文本化，``sheets`` 保存工作簿页签列表供界面切换，``truncated``
    表示可能因行数限制未展示全部内容，``error`` 非空时 ``rows`` 通常为空。
    """

    def __init__(self, path, sheet, rows, sheets=None, truncated=False, error=""):
        """保存预览来源、当前页签、二维文本、截断状态和读取错误。"""
        self.path = path
        self.sheet = sheet
        self.rows = rows
        self.sheets = sheets or []
        self.truncated = truncated
        self.error = error

    @property
    def ncols(self):
        """返回预览中实际最宽行的列数，空预览为零。"""
        return max((len(r) for r in self.rows), default=0)

    @property
    def nrows(self):
        """返回当前已经读取的预览行数。"""
        return len(self.rows)


def _fmt(v):
    """把单元格值转换为紧凑、稳定的预览文本。

    空值为空串，布尔值使用 Excel 风格大写英文，日期仅展示年月日，整数浮点去掉
    ``.0``，其他类型使用 ``str``。日期格式化异常时回退原字符串，不影响整页预览。
    """
    if v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, (datetime.datetime, datetime.date)):
        try:
            # 预览只关心业务日期，datetime 的时分秒不在侧栏中展开。
            return v.strftime("%Y-%m-%d")
        except Exception:
            return str(v)
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def list_sheets(path):
    """尽力列出 Excel 页签名，供预览界面切换；失败返回空列表。

    xlsx/xlsm 只读打开并确保关闭，xls 使用 on-demand 模式避免加载全部页内容。CSV/TXT
    没有页签概念，直接返回空列表。该辅助查询不把异常升级为业务错误。
    """
    ext = os.path.splitext(path)[1].lower()
    try:
        if ext in (".xlsx", ".xlsm"):
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True)
            try:
                return list(wb.sheetnames)
            finally:
                wb.close()
        if ext == ".xls":
            import xlrd
            return list(xlrd.open_workbook(path, on_demand=True).sheet_names())
    except Exception:
        # 页签列表只是预览增强能力，正式 read_preview 会给出更具体的错误文本。
        pass
    return []


def read_preview(path, sheet=None, max_rows=DEFAULT_ROWS, max_cols=DEFAULT_COLS):
    """按扩展名读取指定文件的有限区域并统一返回 ``PreviewData``。

    指定页签不存在时，Excel 读取器兼容回落到首个页签；文件不存在、格式不支持或
    读取异常都转为 ``error`` 文本。此入口是界面容错边界，内部读取函数可正常抛错。
    """
    ext = os.path.splitext(path)[1].lower()
    try:
        if not os.path.isfile(path):
            return PreviewData(path, sheet, [], error="文件不存在")
        if ext in (".xlsx", ".xlsm"):
            return _read_xlsx(path, sheet, max_rows, max_cols)
        if ext == ".xls":
            return _read_xls(path, sheet, max_rows, max_cols)
        if ext in (".csv", ".txt"):
            return _read_csv(path, max_rows, max_cols)
        return PreviewData(path, sheet, [], error="不支持预览的类型:%s" % ext)
    except Exception as e:
        # 预览失败不影响文件继续作为业务输入，界面可在原位置展示错误原因。
        return PreviewData(path, sheet, [], error=str(e))


def _read_xlsx(path, sheet, max_rows, max_cols):
    """以只读计算值模式预览 xlsx/xlsm，并返回工作簿全部页签名。"""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        names = list(wb.sheetnames)
        # 指定名称无效时回落首表，避免历史界面仍记住已被删除页签而无法预览文件。
        ws = wb[sheet] if (sheet and sheet in names) else wb[names[0]]
        # 某些导出把 dimension 错写为 A1；重置后按实际 XML 流式遍历，仍只取有限行。
        try:
            ws.reset_dimensions()
        except Exception:
            pass
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                break
            rows.append([_fmt(c) for c in row[:max_cols]])
        # 只读流无法低成本确认后面是否还有一行；“取满上限”保守标记为可能截断。
        truncated = len(rows) >= max_rows
        return PreviewData(path, ws.title, rows, sheets=names, truncated=truncated)
    finally:
        wb.close()


def _read_xls(path, sheet, max_rows, max_cols):
    """用 xlrd 预览旧 xls 的指定页签，并依据已知行数精确判断截断。"""
    import xlrd
    book = xlrd.open_workbook(path)
    names = book.sheet_names()
    sh = book.sheet_by_name(sheet) if (sheet and sheet in names) else book.sheet_by_index(0)
    rows = []
    for r in range(min(sh.nrows, max_rows)):
        rows.append([_fmt(sh.cell(r, c).value) for c in range(min(sh.ncols, max_cols))])
    return PreviewData(path, sh.name, rows, sheets=list(names),
                       truncated=sh.nrows > max_rows)


def _read_csv(path, max_rows, max_cols):
    """探测常见编码和分隔符后预览 CSV/TXT 的有限行列。"""
    # 先抽样少量原始字节判断 UTF-8；失败后回落 GBK，正式读取用 replace 保证可展示。
    with open(path, "rb") as fb:
        raw = fb.read(4096)
    enc = "utf-8-sig"
    try:
        raw.decode("utf-8-sig")
    except Exception:
        enc = "gbk"
    rows = []
    with open(path, "r", encoding=enc, errors="replace", newline="") as f:
        # Sniffer 只看 2 KiB，并限制逗号、制表符和分号，避免猜出罕见错误分隔符。
        sample = f.read(2048)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
        except Exception:
            # 无法判断时使用标准逗号 CSV 方言，仍给用户一个可检查的预览。
            dialect = csv.excel
        reader = csv.reader(f, dialect)
        for i, row in enumerate(reader):
            if i >= max_rows:
                break
            rows.append([_fmt(c) for c in row[:max_cols]])
    return PreviewData(path, "", rows, truncated=len(rows) >= max_rows)
