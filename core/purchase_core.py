# -*- coding: utf-8 -*-
"""
采购数对账业务核心
======================
把"我方对账单"与"供应商对单明细"逐行匹配，生成三样产物：
  · 两张原表副本，数量列上色（绿=对上，黄=未对上）；
  · 对账汇报单：已对上并排 + 双方各自未对上 + 未对上原因，供人工一眼核对。

匹配规则（经真实数据验证，与人工上色结果 168/168 完全一致）：
  · 必需：材料名称 + 规格 + 数量 三者一致（规格忽略大小写/空格/*×；数量按数值比）；
  · 材料编号：两边都有则必须相等，且批次不得矛盾（容忍"漏打/多打一位"笔误）；
  · 批次号：仅当某侧缺编号时作兜底判据，并作优先级评分让"编号+批次都吻合"先锁定；
  · 一对一全局配对：先最大化成功配对数量，再最大化编号/批次可信评分；每行只用一次。

表头行与列位置自动识别，不写死列号。主数据只补空字段，匹配、原因诊断、原表上色和
汇报单结构均在本模块实现，桌面端与 Web 端不重复这些业务算法。
"""
import os
import re

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import paths as _paths
from . import settings as settings_mod
from . import common_core as _common       # 公共文本清洗、只读加载与分阶段进度上报。
from .common_core import warn_if_uncached  # 读取关键公式列前检测未保存的公式缓存。
from . import header_detect
from . import shape_detect
from . import material_catalog

# 表头关键词映射到业务角色；公共引擎先精确匹配再包含匹配，并保证每列只承担一个角色。
HEADER_KEYS = {
    "no":    ["材料编号", "物料编号", "存货编码", "物料编码", "材料编码", "产品编号",
              "商品编码", "货号", "物料代码", "材料代码", "编号", "编码", "料号"],
    "name":  ["材料名称", "物料名称", "存货名称", "产品名称", "名称", "品名"],
    "spec":  ["规格型号", "规格", "型号"],
    "unit":  ["计量单位", "单位"],
    "qty":   ["采购数量", "对账数量", "结算数量", "数量合计", "数量"],
    "batch": ["批次号", "生产批次", "批次", "批号"],
    "note":  ["备注", "说明"],
}


def norm_name(v):
    """规范化材料名称，忽略首尾和内部普通空格，但保留其他可能有业务意义的符号。"""
    return "" if v is None else str(v).strip().replace(" ", "")


def norm_spec(v):
    """统一规格的大小写、空格及 ``*``/``×`` 乘号写法，用于跨模板等价比较。"""
    if v is None:
        return ""
    return str(v).strip().upper().replace(" ", "").replace("*", "X").replace("×", "X")


def norm_no(v):
    """
    规范化材料编号，使 Excel 数字 ``123.0``、文本 ``00123`` 和 ``123`` 能归为同值。

    编号可能被 Excel 自动读取为浮点，也可能由不同系统补前导零；若直接转字符串会把
    同一材料误报成“对方无此编号”。含字母、横杠等非纯数字编号保留其结构，只执行
    公共全半角、零宽字符和大小写清洗。
    """
    if v is None:
        return ""
    # Excel 数值编号常成为 float；仅整数浮点去掉 .0，真实小数编号不武断改写。
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    s = _common.clean_str(v).upper()  # 全角转半角并清除零宽字符、NBSP，再统一英文大小写。
    # 纯数字去前导零；全零保留为单个“0”，非纯数字编号不改变内部格式。
    if s.isdigit():
        s = s.lstrip("0") or "0"
    return s


def batch_core(v):
    """去除批次开头的供应商字母前缀，提取可跨双方比较的批次核心。"""
    if v is None:
        return ""
    s = str(v).strip().upper()
    s2 = re.sub(r"^[A-Z]+", "", s)  # 只去开头字母，保留数字后的字母和子批次信息。
    return s2 or s


def batch_compat(a, b):
    """判断两侧批次核心是否相等或互为非空子串，以兼容主批次与子批次粒度差异。"""
    ca, cb = batch_core(a), batch_core(b)
    return bool(ca and cb and (ca == cb or ca in cb or cb in ca))


def _one_indel(a, b):
    """是否仅差一次"插入/删除"(长度差恰为1，短串=长串删一个字符)。

    只容忍漏打或多打一位，例如 ``26004`` 与 ``2604``。等长替换不属于此容错范围，
    因为 ``21004`` 与 ``21014`` 更可能是两个真实批次，放宽会制造错误配对。
    """
    la, lb = len(a), len(b)
    if abs(la - lb) != 1:
        return False
    if la > lb:
        # 后续统一验证“短串是否等于长串删除一位”，交换可避免维护两套分支。
        a, b = b, a
    for k in range(len(b)):
        if a == b[:k] + b[k + 1:]:
            return True
    return False


def _split_batch(v):
    """按最后一个横杠把批次核心拆成主批次和子批次尾号。"""
    core = batch_core(v)
    if "-" in core:
        p, s = core.rsplit("-", 1)
        return p, s
    return core, ""


def batch_consistent(a, b):
    """编号已相等时，批次是否可信为"同一行"。

    当材料编号已经相等时，仍需防止不同批次被名称、规格、数量偶然拉到一起。本规则
    只容忍供应商漏打/多打一位的笔误，拒绝真正不同的批次：
      · 子批次尾号两侧都有且不同 → 不同批，拒绝(26004-01 vs 26004-02)；
      · 前缀被改写一位(等长替换)→ 不同批，拒绝(21004-01 vs 21014-01)；
      · 前缀差异超过一次增删 → 不同批，拒绝(26004-01 vs 26010-01)。
    """
    if batch_compat(a, b):
        # 完全一致或主/子批次包含关系已经足够可信，无需进入笔误容错。
        return True
    pa, sa = _split_batch(a)
    pb, sb = _split_batch(b)
    if sa and sb and sa != sb:
        # 双方都明确给出不同子批次尾号时必须拒绝，不能被主批次近似规则救回。
        return False
    if not pa or not pb:
        # 至少一侧缺主批次时没有证据证明矛盾，保持兼容以免编号相等的行全部漏配。
        return True
    return _one_indel(pa, pb)


def qty_eq(a, b):
    """优先按浮点容差比较数量；无法数值化时退回去除两端空白的文本精确比较。"""
    if a is None or b is None:
        return False
    try:
        # 极小容差只吸收二进制浮点尾差，不容忍真实采购数量偏差。
        return abs(float(a) - float(b)) < 1e-9
    except (TypeError, ValueError):
        return str(a).strip() == str(b).strip()


# ---------------------------------------------------------------------------
# 表头/列自动识别
# ---------------------------------------------------------------------------
def detect_layout(ws, scan_rows=12, log=None):
    """在前若干行识别采购表头及业务列，失败时返回 ``(None, {})``。

    公共引擎负责候选评分，本函数只提供采购字段别名和最低条件。数量列必须存在，
    名称或编号至少存在一个；名称缺失时后续可由主数据库按编号补齐。保持薄封装可让
    多个业务模块共享表头修正，而不在这里复制识别算法。
    """
    header_row, columns = header_detect.detect_layout(
        ws, HEADER_KEYS, require=("qty",), scan_rows=scan_rows, log=log)
    if header_row and ("name" in columns or "no" in columns):
        return header_row, columns
    return None, {}


# 数据形态画像只用于分析/人工确认兜底：角色顺序接近典型列序，末位布尔值表示必需列。
_SHAPE_PROFILE = [
    ("no", shape_detect.CODE, False),
    ("name", shape_detect.TEXT, True),
    ("spec", shape_detect.TEXT, False),
    ("unit", shape_detect.TEXT, False),
    ("qty", shape_detect.NUMBER, True),
]


def detect_layout_or_shape(ws, scan_rows=12, log=None):
    """先按表头文字识别，失败后再用数据形态画像兜底。

    返回 ``(header_row, col_map, source)``，其中来源为 ``header``、``shape`` 或 ``None``。
    形态识别只说明“像某种列”，可信度低于明确表头，调用方必须交给用户确认，不能
    在无人知情时直接据此落盘。
    """
    hr, col = detect_layout(ws, scan_rows=scan_rows, log=log)
    if hr:
        return hr, col, "header"
    hr2, col2, _conf = shape_detect.detect_by_shape(
        ws, _SHAPE_PROFILE, scan_rows=scan_rows, log=log)
    if hr2:
        return hr2, col2, "shape"
    return None, {}, None


def _select_purchase_sheet(workbook, requested_sheet, log):
    """选择采购业务页签，并返回 ``(工作表, 表头行, 角色列映射)``。

    显式指定页签时直接识别；未指定时先识别首张表，识别失败且工作簿含多个页签时再顺序
    寻找后续有效页签，以兼容封面/说明页位于首位的工作簿。单表识别失败也原样返回，
    由 ``load_rows`` 抛出统一的可读错误。
    """

    if requested_sheet:
        worksheet = workbook[requested_sheet]
        header_row, columns = detect_layout(worksheet, log=log)
        return worksheet, header_row, columns

    first_name = workbook.sheetnames[0]
    worksheet = workbook[first_name]
    header_row, columns = detect_layout(worksheet, log=log)
    if header_row or len(workbook.sheetnames) == 1:
        return worksheet, header_row, columns

    for sheet_name in workbook.sheetnames[1:]:
        candidate = workbook[sheet_name]
        candidate_header, candidate_columns = detect_layout(candidate, log=log)
        if candidate_header:
            if log:
                log("· 多子表: 首表《%s》无有效表头, 改读《%s》" % (first_name, sheet_name))
            return candidate, candidate_header, candidate_columns
    return worksheet, header_row, columns


def _purchase_cell(worksheet, row, columns, role):
    """按角色读取采购业务单元格；模板没有该可选角色时返回 ``None``。"""

    column = columns.get(role)
    return worksheet.cell(row, column).value if column else None


def _purchase_item(worksheet, row, columns, resolver, fill_counts):
    """把工作表一行转换为采购匹配记录，空行、汇总行和无名称行返回 ``None``。

    ``resolver`` 用于按材料编号只补空字段，补全统计累加进 ``fill_counts``；名称与编号
    同时为空视为空行，名称含“合计/小计/总计”视为汇总行，均不参与后续逐行配对。
    """

    name = _purchase_cell(worksheet, row, columns, "name")
    material_no = _purchase_cell(worksheet, row, columns, "no")
    name_empty = name is None or not str(name).strip()
    number_empty = material_no is None or not str(material_no).strip()
    if name_empty and number_empty:
        return None
    # 汇总行（合计/小计/总计）不是逐行明细，必须排除，否则汇总数量会被当作可配对明细。
    if isinstance(name, str) and any(word in name for word in ("合计", "小计", "总计")):
        return None

    item = {
        "r": row,
        "no": material_no,
        "name": name,
        "spec": _purchase_cell(worksheet, row, columns, "spec"),
        "unit": _purchase_cell(worksheet, row, columns, "unit"),
        "qty": _purchase_cell(worksheet, row, columns, "qty"),
        "batch": _purchase_cell(worksheet, row, columns, "batch"),
    }
    if resolver is not None:
        resolver.fill_mapping(
            item,
            code_key="no",
            field_keys={"name": "name", "spec": "spec", "unit": "unit"},
            fields=("name", "spec", "unit"),
            counts=fill_counts,
        )
    if item["name"] is None or not str(item["name"]).strip():
        return None
    return item


def _read_purchase_rows(worksheet, header_row, columns, resolver, fill_counts):
    """读取一个已识别页签的有效采购行，并统计数量为空的记录数。"""

    rows = []
    missing_quantity = 0
    for row in range(header_row + 1, worksheet.max_row + 1):
        item = _purchase_item(worksheet, row, columns, resolver, fill_counts)
        if item is None:
            continue
        if item["qty"] is None:
            missing_quantity += 1
        rows.append(item)
    return rows, missing_quantity


def load_rows(path, sheet=None, log=None, resolver=None, fill_counts=None):
    """
    读取采购表有效明细并返回 ``(rows, layout)``。

    每条明细包含原始行号、编号、名称、规格、单位、数量和批次。未指定页签时先尝试
    首张表，识别失败才顺序寻找后续页签，兼容封面/说明页位于首位的工作簿。读取采用
    ``data_only`` 计算值模式；材料名称、规格和单位可由主数据库只补空值。
    """
    # 数量公式未保存缓存时 data_only 会读到 None，使该行永远无法配对，因此提前告警。
    if log:
        warn_if_uncached(path, log, sheet=sheet, what="数量")
    # 公共加载器跳过易损坏的内嵌透视缓存，只读取本功能需要的单元格计算值。
    wb = _common.load_data_only(path)
    try:
        ws, header_row, col = _select_purchase_sheet(wb, sheet, log)
        if not header_row:
            raise ValueError("未能在 %s / %s 中识别表头（需含“名称”“数量”列）"
                             % (os.path.basename(path), ws.title))
        rows, none_qty = _read_purchase_rows(
            ws, header_row, col, resolver, fill_counts,
        )
        layout = {"sheet": ws.title, "header_row": header_row, "col": col}
    finally:
        wb.close()  # Windows 下显式释放句柄，后续才能正常复制和保存同一文件。
    # 空数量仍保留在 rows 供报告展示，但明确提示其不能正确参与数值配对。
    if none_qty and log:
        log("⚠ 《%s》有 %d 行的数量未取到值(可能公式未刷新), 这些行将无法正确对账, 请核对。"
            % (os.path.basename(path), none_qty))
    return rows, layout


# ---------------------------------------------------------------------------
# 匹配引擎
# ---------------------------------------------------------------------------
def _can_match(a, b):
    """
    判断两条业务行是否具备配对资格，不处理“每行只能使用一次”的全局约束。

    名称、规格和数量必须等价；双方都有材料编号时编号必须相同且批次不能矛盾。
    编号缺失时才允许批次作为兜底；双方编号和批次都缺失时退化为名称、规格、数量匹配，
    后续 ``pair_note`` 会把这种弱依据明确标为需要人工核对。
    """
    if norm_name(a["name"]) != norm_name(b["name"]):
        return False
    if norm_spec(a["spec"]) != norm_spec(b["spec"]):
        return False
    if not qty_eq(a["qty"], b["qty"]):
        return False
    na, nb = norm_no(a["no"]), norm_no(b["no"])
    if na and nb:
        # 编号是强键，双方都有时绝不允许不同编号；批次再用受限笔误规则排除真实冲突。
        return na == nb and batch_consistent(a["batch"], b["batch"])
    # 至少一侧缺编号时才让批次承担识别作用，避免用模糊批次覆盖明确编号冲突。
    if batch_compat(a["batch"], b["batch"]):
        return True
    # 双方编号和批次都空时仍允许弱匹配，以免旧表完全无法处理；报告会强制提示人工核对。
    if not na and not nb and not batch_core(a["batch"]) and not batch_core(b["batch"]):
        return True
    return False


def _pair_score(a, b):
    """计算候选边可信评分：编号相等加 2，批次兼容加 1，供全局匹配二级优化。"""
    na, nb = norm_no(a["no"]), norm_no(b["no"])
    s = 2 if (na and nb and na == nb) else 0
    if batch_compat(a["batch"], b["batch"]):
        s += 1
    return s


def match_rows(rows1, rows2):
    """对双方明细执行确定性的一对一全局最优配对。

    返回两个与输入等长的匹配布尔列表，以及 ``(左索引, 右索引, 评分)`` 配对列表。
    先按规范化名称和规格分桶，跨桶不可能满足基础条件，因此可显著缩小图规模；每个桶
    再用最小费用最大流求解，优先最大化配对数量，其次最大化编号/批次评分总和，避免
    局部贪心先占用某条记录的唯一对家而导致总体少配。
    """
    m1 = [False] * len(rows1)
    m2 = [False] * len(rows2)
    pairs = []

    # 分桶只使用名称和规格；数量、编号和批次继续由桶内候选边规则判断。
    buckets1, buckets2 = {}, {}
    for i, row in enumerate(rows1):
        buckets1.setdefault((norm_name(row["name"]), norm_spec(row["spec"])), []).append(i)
    for j, row in enumerate(rows2):
        buckets2.setdefault((norm_name(row["name"]), norm_spec(row["spec"])), []).append(j)

    for key, left in buckets1.items():
        right = buckets2.get(key, [])
        if not right:
            continue
        for i, j, score in _optimal_bucket_pairs(rows1, rows2, left, right):
            m1[i] = m2[j] = True
            pairs.append((i, j, score))
    # 求解顺序不应影响报告阅读顺序，最终按我方原表索引稳定排列。
    pairs.sort(key=lambda pair: pair[0])
    return m1, m2, pairs


class _ResidualNetwork:
    """容量均为一的残量网络，用于同名同规格桶内的一对一最优匹配。"""

    def __init__(self, node_count):
        """构建包含 ``node_count`` 个节点的空残量网络，所有边由后续 ``add_edge`` 注入。"""
        self.graph = [[] for _ in range(node_count)]

    def add_edge(self, start, end, capacity, cost, meta=None):
        """添加正向边及对应反向边，业务元数据只保存在正向候选边。"""

        self.graph[start].append([end, len(self.graph[end]), capacity, cost, meta])
        self.graph[end].append([start, len(self.graph[start]) - 1, 0, -cost, None])

    def shortest_path(self, source, sink):
        """使用队列版 Bellman-Ford 查找当前残量网络中的最短增广路。"""

        distances = [None] * len(self.graph)
        previous = [None] * len(self.graph)
        in_queue = [False] * len(self.graph)
        queue = [source]
        distances[source] = 0
        in_queue[source] = True
        head = 0
        while head < len(queue):
            node = queue[head]
            head += 1
            in_queue[node] = False
            for edge_index, edge in enumerate(self.graph[node]):
                target, _reverse, capacity, cost, _meta = edge
                if capacity <= 0:
                    continue
                candidate = distances[node] + cost
                if distances[target] is None or candidate < distances[target]:
                    distances[target] = candidate
                    previous[target] = (node, edge_index)
                    if not in_queue[target]:
                        queue.append(target)
                        in_queue[target] = True
        return previous if previous[sink] is not None else None

    def augment(self, source, sink, previous):
        """沿最短路径发送一个单位流，并同步开放各边的反向撤销容量。"""

        node = sink
        while node != source:
            parent, edge_index = previous[node]
            edge = self.graph[parent][edge_index]
            edge[2] -= 1
            self.graph[node][edge[1]][2] += 1
            node = parent

    def maximize_flow(self, source, sink):
        """持续增广直到汇点不可达，实现最大基数、最小费用匹配。"""

        while True:
            previous = self.shortest_path(source, sink)
            if previous is None:
                return
            self.augment(source, sink, previous)


def _build_bucket_network(rows1, rows2, left, right):
    """构造同名同规格桶的二分图残量网络及节点位置。

    节点布局为“源点 -> 左侧行 -> 右侧行 -> 汇点”，仅 ``_can_match`` 通过的候选建立
    左右边，边上保存评分和原始索引作为业务元数据。返回网络、源点、汇点及左侧节点范围。
    """

    left_count = len(left)
    right_count = len(right)
    source = 0
    left_base = 1
    right_base = left_base + left_count
    sink = right_base + right_count
    network = _ResidualNetwork(sink + 1)
    for left_index in range(left_count):
        network.add_edge(source, left_base + left_index, 1, 0)
    for right_index in range(right_count):
        network.add_edge(right_base + right_index, sink, 1, 0)
    for left_index, source_row in enumerate(left):
        for right_index, target_row in enumerate(right):
            if not _can_match(rows1[source_row], rows2[target_row]):
                continue
            score = _pair_score(rows1[source_row], rows2[target_row])
            # 评分权重远大于顺序项，输入顺序只在总评分完全相同时作为稳定决胜依据。
            cost = -score * 100000 + left_index * max(1, right_count) + right_index
            network.add_edge(
                left_base + left_index,
                right_base + right_index,
                1,
                cost,
                (source_row, target_row, score),
            )
    return network, source, sink, left_base, left_count


def _matched_bucket_edges(network, left_base, left_count):
    """提取最终承载一个单位流的左右候选边业务元数据。"""

    result = []
    for left_index in range(left_count):
        for edge in network.graph[left_base + left_index]:
            if edge[4] is not None and edge[2] == 0:
                result.append(edge[4])
    return result


def _optimal_bucket_pairs(rows1, rows2, left, right):
    """
    在一个同名同规格桶内，用最小费用最大流求最大基数、最大评分的二分图匹配。

    网络结构为“源点 -> 左侧行 -> 右侧行 -> 汇点”，所有容量均为 1，因此天然保证
    每行最多参与一次。只为 ``_can_match`` 通过的候选建立左右边；边费用以负评分为主，
    输入顺序为极小的同分决胜项，使相同文件在不同运行中得到稳定配对。
    """
    network, source, sink, left_base, left_count = _build_bucket_network(
        rows1, rows2, left, right,
    )
    network.maximize_flow(source, sink)
    return _matched_bucket_edges(network, left_base, left_count)


# ---------------------------------------------------------------------------
# 未对上原因诊断
# ---------------------------------------------------------------------------
def _same_name_spec(row, pool):
    """筛出与当前记录名称和规格均等价的候选记录。"""

    return [
        candidate
        for candidate in pool
        if norm_name(row["name"]) == norm_name(candidate["name"])
        and norm_spec(row["spec"]) == norm_spec(candidate["spec"])
    ]


def _diagnose_numbered_row(row, others_all, others_avail, material_no):
    """按编号、名称规格、批次和数量逐级诊断有材料编号的未匹配记录。"""

    same_number_all = [item for item in others_all if norm_no(item["no"]) == material_no]
    if not same_number_all:
        return "对方无此编号(%s)" % row["no"]
    same_number_available = [
        item for item in others_avail if norm_no(item["no"]) == material_no
    ]
    if not same_number_available:
        return "对方此编号明细已全部配对，本行多出"
    same_material = _same_name_spec(row, same_number_available)
    if not same_material:
        candidate = same_number_available[0]
        return "同编号但名称/规格不同(对方:%s %s)" % (
            candidate["name"], candidate["spec"],
        )
    same_batch = [
        item
        for item in same_material
        if batch_consistent(row["batch"], item["batch"])
    ]
    if not same_batch:
        return "同编号同规格但对方无此批次(%s)" % (row["batch"] or "空")
    quantities = "/".join(sorted({str(item["qty"]) for item in same_batch}))
    return "同编号同批次但数量不符(对方数量:%s)" % quantities


def _diagnose_unnumbered_row(row, others_avail):
    """按名称规格、批次和数量诊断缺少材料编号的未匹配记录。"""

    same_material = _same_name_spec(row, others_avail)
    if not same_material:
        return "对方(未对上项中)无此料/规格"
    same_batch = [
        item for item in same_material if batch_compat(row["batch"], item["batch"])
    ]
    if not same_batch:
        return "同名同规格但对方无此批次(%s)" % (row["batch"] or "空")
    quantities = "/".join(sorted({str(item["qty"]) for item in same_batch}))
    return "同名同规格同批次但数量不符(对方数量:%s)" % quantities


def diagnose(row, others_all, others_avail):
    """为一条未匹配记录逐级定位最接近的对方记录并生成可行动的原因说明。

    定位顺序为编号、名称规格、批次、数量。判断“对方是否存在该编号”时查看全部对方行，
    但具体名称、批次和数量诊断只使用仍未配对的记录，绝不引用已被其他行占用的对家。
    否则报告会给出用户在对方未匹配清单中根本找不到的数量，造成错误排查方向。
    """
    material_no = norm_no(row["no"])
    if material_no:
        return _diagnose_numbered_row(row, others_all, others_avail, material_no)
    return _diagnose_unnumbered_row(row, others_avail)


def _quantity_conflict_candidate(left, right):
    """判断两条未匹配记录是否除数量外具备同料同批次关系。"""

    if norm_name(left["name"]) != norm_name(right["name"]):
        return False
    if norm_spec(left["spec"]) != norm_spec(right["spec"]):
        return False
    left_no, right_no = norm_no(left["no"]), norm_no(right["no"])
    if left_no and right_no and left_no != right_no:
        return False
    return batch_compat(left["batch"], right["batch"]) and not qty_eq(
        left["qty"], right["qty"],
    )


def _best_quantity_conflict(left, rows2, matched2):
    """为一条我方记录选择批次核心最接近且顺序稳定的数量疑点对家。"""

    best = None
    for index, right in enumerate(rows2):
        if matched2[index] or not _quantity_conflict_candidate(left, right):
            continue
        exact_batch = batch_core(left["batch"]) == batch_core(right["batch"])
        if best is None or (exact_batch and not best[1]):
            best = (right, exact_batch)
    return best[0] if best is not None else None


def quantity_conflicts(rows1, rows2, matched1, matched2):
    """从双方未匹配记录中提取“只有数量不同”的高价值人工疑点。

    候选必须名称规格一致、双方都有编号时编号相等、批次兼容，且数量确实不同。
    已匹配记录不再重复报告。每条我方记录只保留一个最接近对家，批次核心完全相等者
    优先，避免同料多批次产生大量排列组合式提示。
    """
    out = []
    for index, left in enumerate(rows1):
        if matched1[index]:
            continue
        right = _best_quantity_conflict(left, rows2, matched2)
        if right is not None:
            out.append((left, right))
    return out


from .purchase_reporting import apply_colors, build_report, GREEN, YELLOW
def _out_name(path):
    """在来源文件主名后追加“对账结果”，非 xlsx 输入统一输出为 xlsx。"""
    base, ext = os.path.splitext(os.path.basename(path))
    return "%s_对账结果%s" % (base, ext if ext.lower() == ".xlsx" else ".xlsx")


def _read_purchase_inputs(files, sheets, names, log):
    """使用同一主数据快照读取双方采购表，并记录布局与补全统计。

    双方共用一个 ``CatalogResolver``，保证补全口径一致且统计不重复；``files``、
    ``sheets``、``names`` 三个序列必须等长。返回 ``(rows, layouts)``，分别保存双方
    明细行列表与各自工作表布局。
    """

    # 两侧共用同一个 CatalogResolver，使主数据补全快照一致、补全统计不重复。
    resolver = material_catalog.CatalogResolver()
    fill_counts = {}
    rows = []
    layouts = []
    for path, sheet, display_name in zip(files, sheets, names):
        side_rows, layout = load_rows(
            path,
            sheet,
            log=log,
            resolver=resolver,
            fill_counts=fill_counts,
        )
        rows.append(side_rows)
        layouts.append(layout)
    # 保持既有日志顺序：先说明主数据库补全情况，再列出双方实际参与匹配的行数。
    material_catalog.log_fill_summary(log, "采购数对账", fill_counts)
    for side_rows, layout, display_name in zip(rows, layouts, names):
        log("%s 有效行: %d（表头第%d行，工作表 %s）" % (
            display_name, len(side_rows), layout["header_row"], layout["sheet"],
        ))
    return rows, layouts


def _log_purchase_match_summary(rows, matched, pairs, names, log):
    """记录配对总数及双方已匹配、未匹配数量。"""

    log("配对成功 %d 对" % len(pairs))
    for side_rows, side_matched, display_name in zip(rows, matched, names):
        matched_count = sum(side_matched)
        log("%s 对上 %d / 未对上 %d" % (
            display_name, matched_count, len(side_rows) - matched_count,
        ))


def _log_quantity_conflicts(conflicts, names, log):
    """把仅数量不同的高价值疑点写入任务日志。"""

    if not conflicts:
        return
    log("\n【数量不一致疑点】%d 处：" % len(conflicts))
    for left, right in conflicts:
        log("  %s行%s(量%s) ↔ %s行%s(量%s) | %s %s %s" % (
            names[0], left["r"], left["qty"],
            names[1], right["r"], right["qty"],
            left["no"], left["name"], left["spec"],
        ))


def _resolve_purchase_output_dir(out_dir):
    """按统一设置解析采购对账输出目录，并确保显式目录存在。"""

    if out_dir is None:
        current = settings_mod.get_settings()
        return _paths.resolve_output_dir("purchase", **current.output_kwargs())
    os.makedirs(out_dir, exist_ok=True)
    return out_dir


def _write_colored_purchase_copies(files, rows, layouts, matched, out_dir, progress, log):
    """生成双方原表上色副本，并按完成侧数更新当前进度阶段。"""

    outputs = []
    total = len(files)
    for index, (path, side_rows, layout, side_matched) in enumerate(
        zip(files, rows, layouts, matched), start=1,
    ):
        output = apply_colors(
            path,
            layout["sheet"],
            side_matched,
            side_rows,
            layout["col"]["qty"],
            os.path.join(out_dir, _out_name(path)),
            col_map=layout["col"],
        )
        outputs.append(output)
        progress.tick(index, total)
        log("已生成上色表：%s" % output)
    return outputs


# ---------------------------------------------------------------------------
# 统一入口（与其余功能一致：run(..., out_dir=None, log=None) -> dict）
# ---------------------------------------------------------------------------
def run(file1, file2, sheet1=None, sheet2=None, name1="我方", name2="供方",
        out_dir=None, log=None, progress=None):
    """执行采购数对账，生成两张状态上色副本和一张并排汇报单。

    ``file1``/``file2`` 分别为我方和供方表格；``sheet1``/``sheet2`` 可显式指定页签，
    未指定时自动识别。``name1``/``name2`` 只影响报告文案，不写死具体企业名称。
    ``progress`` 接收 0～100 的阶段进度。

    返回结构同时包含原始投影行、匹配布尔数组、配对索引、数量疑点、布局和三个输出
    路径，供统一业务结果层生成前端预览，不需要前端重新解析 Excel。
    """
    def _lg(msg):
        """把核心处理日志转交桌面端或 Web 任务上下文。"""
        if log:
            log(msg)

    # 权重按典型耗时分配：读取 30%、匹配诊断 20%、两表写盘 35%、汇报单 15%。
    prog = _common.Progress(progress, stages=[
        ("read", 30), ("match", 20), ("color", 35), ("report", 15)])
    prog.stage("read")
    files = (file1, file2)
    names = (name1, name2)
    rows, layouts = _read_purchase_inputs(files, (sheet1, sheet2), names, _lg)
    rows1, rows2 = rows
    lay1, lay2 = layouts

    prog.stage("match")
    # 全局匹配先最大化对数，再偏好编号和批次更强的组合。
    m1, m2, pairs = match_rows(rows1, rows2)
    matched = (m1, m2)
    _log_purchase_match_summary(rows, matched, pairs, names, _lg)

    # 从尚未匹配部分提取“除数量外均相近”的记录，作为最值得人工优先核查的疑点。
    qc = quantity_conflicts(rows1, rows2, m1, m2)
    _log_quantity_conflicts(qc, names, _lg)
    out_dir = _resolve_purchase_output_dir(out_dir)
    _lg("输出文件夹：%s" % out_dir)

    prog.stage("color")
    o1, o2 = _write_colored_purchase_copies(
        files, rows, layouts, matched, out_dir, prog, _lg,
    )

    prog.stage("report")
    # 汇报单使用固定业务名称；与现有行为一致，若同名文件打开会得到清晰占用提示。
    report = os.path.join(out_dir, "采购数对账汇报单.xlsx")
    rp = build_report(rows1, rows2, m1, m2, pairs, report, name1, name2)
    _lg("已生成汇报单：%s" % rp)
    prog.done()
    return {"rows1": rows1, "rows2": rows2, "matched1": m1, "matched2": m2,
            "pairs": pairs, "qty_conflicts": qc, "out1": o1, "out2": o2,
            "report": rp, "out_dir": out_dir, "layout1": lay1, "layout2": lay2,
            "name1": name1, "name2": name2}
