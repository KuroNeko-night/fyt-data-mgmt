# -*- coding: utf-8 -*-
"""日清生产计划文件的安全解析与摘要。

本模块只读解析管理员上传的生产、订单和发运工作簿，把不同版式归一化为前端看板与
日清报告可复用的结构化摘要。解析设置了预览、列数和表格行数上限，防止异常大表拖垮
请求；原始文件不在此处修改，也不把 Excel 解析逻辑复制到前端。
"""

from __future__ import annotations

import datetime as dt
import os
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles.numbers import is_date_format
from openpyxl.utils.datetime import from_excel


SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm"}
MAX_PREVIEW_ROWS = 12
MAX_PREVIEW_COLUMNS = 24
MAX_TABLE_ROWS = 240
_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")


def _cell_text(value: Any, *, number_format: str = "", epoch: Any = None, force_excel_date: bool = False) -> str:
    """把单元格值转换为稳定文本，并识别 Excel 日期序列号。

    Excel 日期可能以日期对象、带日期格式的数字或横向模板中的普通格式数字出现。
    ``force_excel_date`` 用于模板已确认是日期行的情况；数值范围限制可避免把普通产量
    误当成日期。转换失败时回退原文本，保留数据供人工查看。
    """
    if value is None:
        return ""
    # bool 是 int 的子类，必须排除；否则 True 可能进入日期序列号判断。
    if isinstance(value, (int, float)) and not isinstance(value, bool) and epoch is not None and (force_excel_date or is_date_format(number_format or "")) and 20000 < value < 80000:
        try:
            converted = from_excel(value, epoch=epoch)
            return converted.date().isoformat() if hasattr(converted, "date") else str(converted)
        except (TypeError, ValueError, OverflowError):
            # 工作簿日期系统或损坏格式无法转换时保留原值，不让单个单元格中断整表。
            pass
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except (AttributeError, TypeError, ValueError):
            pass
    return str(value).strip()


def _table_kind(sheet_name: str, headers: list[str]) -> str:
    """根据页签名和表头组合判断工作表的业务类别。

    明确的规范页签名优先，其次使用多个字段共同判断，最后回退“通用数据”。规则集中
    在这里便于新增模板时维护，调用方不需要知道每类工作表的命名细节。
    """
    joined = "".join(headers).replace(" ", "")
    name = sheet_name.replace(" ", "")
    if name in {"主料异常", "辅料异常", "包装异常", "海外历史记录", "防错异常", "问题一览表"}:
        return "现场问题"
    if name == "班组":
        return "班组名册"
    if name == "Sheet1" and ({"物料编码", "物料名称", "供应商信息"}.intersection(headers) or {"材料编号", "材料名称", "供应商"}.issubset(set(headers))):
        # 每日到料成品常保留默认 Sheet1 名称，只能依靠材料字段组合识别。
        return "到料明细"
    if "零星订单" in name or {"运输方式", "是否拼箱"}.issubset(set(headers)):
        return "零星订单"
    if "正式订单" in name or ("订单号" in headers and "发运完成时间" in joined):
        return "正式订单"
    if {"计划", "实际", "差异"}.issubset(set(headers)) or "生产计划" in name:
        return "生产计划"
    if "物料号" in headers and any("实际" in header for header in headers):
        return "生产实绩"
    return "通用数据"


def _find_header_index(rows: list[list[str]]) -> int | None:
    """在前二十四行中寻找至少命中两个业务标记的传统表头行。

    生产模板顶部常含标题、日期和合并单元格，不能假设第一行就是表头。要求两个标记
    可以减少普通数据行被误判；横向矩阵没有传统表头时返回 ``None``，由专门逻辑处理。
    """
    tokens = (
        "计划", "实际", "差异", "订单号", "物料号", "物料描述", "求和项:实际包装数量",
        "材料编号", "材料名称", "供应商", "数量", "物料编码", "物料名称", "供应商信息",
        "需求数", "剩余未收数", "批次号", "发运时间", "发运完成时间", "缺件", "危包", "运输方式",
    )
    for index, row in enumerate(rows[:24]):
        normalized = {cell.replace(" ", "").replace("\n", "") for cell in row if cell}
        if len(normalized.intersection(tokens)) >= 2:
            return index
    return None


def _as_number(value: object) -> float | None:
    """解析允许千分位的数字；空值、布尔值和非法文本返回 ``None``。"""
    if value in (None, "") or isinstance(value, bool):
        return None
    try:
        return float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return None


def _number_label(value: int | float | None) -> int | float:
    """把统计结果压缩为整数或最多四位小数，空值按展示零处理。"""
    if value is None:
        return 0
    # 聚合值既可能来自浮点单元格，也可能来自 sum(bool) 等整数统计；先统一类型，
    # 避免 Python 3.11 在整数对象上调用 float.is_integer() 时抛出属性错误。
    number = float(value)
    return int(number) if number.is_integer() else round(number, 4)


def _row_label(row: list[str]) -> str:
    """取一行首个非空单元格作为横向矩阵的行标签。"""
    for value in row:
        if value:
            return value.replace(" ", "").replace("\n", "")
    return ""


def _find_row(rows: list[list[str]], label: str) -> tuple[int, list[str]] | None:
    """按规范化后的首个非空值定位矩阵行，并返回行号和整行。"""
    for index, row in enumerate(rows):
        normalized = _row_label(row)
        if normalized == label:
            return index, row
    return None


def _fill_forward(values: list[str], start: int) -> list[str]:
    """从指定列开始向右填充合并单元格语义。

    横向计划模板通常只在一组班次的第一列填写日期，其余列为空；读取模式不会自动把
    合并值复制到每列，因此需要沿用最近一个非空日期。
    """
    result = [""] * len(values)
    current = ""
    for index in range(start, len(values)):
        value = values[index].strip() if index < len(values) else ""
        if value:
            current = value
        result[index] = current
    return result


def _matrix_records(
    date_row: list[str],
    shift_row: list[str],
    plan_row: list[str],
    actual_row: list[str],
    difference_row: list[str],
    label_column: int,
) -> list[dict[str, object]]:
    """把横向矩阵的每个有效班次列转换为统一记录。"""
    dates = _fill_forward(date_row, label_column + 1)  # 合并日期单元格只在首列有值，先向右展开。
    records: list[dict[str, object]] = []
    width = max(len(date_row), len(shift_row), len(plan_row), len(actual_row))  # 以最长业务行覆盖所有班次列。
    for column in range(label_column + 1, width):
        current_date = dates[column] if column < len(dates) else ""  # 越界列视为无日期，避免补齐数据被误算。
        shift = shift_row[column].strip() if column < len(shift_row) else ""  # 班次为空通常是说明或合计列。
        plan = _as_number(plan_row[column] if column < len(plan_row) else None)  # 计划是识别业务列的必要条件。
        actual = _as_number(actual_row[column] if column < len(actual_row) else None)  # 空值代表尚未填报而非零。
        difference = _as_number(difference_row[column] if column < len(difference_row) else None)  # 优先采用模板差异。
        if not current_date or not _DATE_RE.match(current_date) or not shift or plan is None:
            continue  # 缺少日期、班次或计划的列不属于可展示生产记录。
        actual_reported = actual is not None  # 单独保留填报状态，避免把空值压成零后丢失语义。
        difference_value = difference if actual_reported and difference is not None else ((actual - plan) if actual_reported else None)
        records.append({
            "date": current_date, "shift": shift, "plan": _number_label(plan),
            "actual": _number_label(actual), "difference": _number_label(difference_value),
            "actual_reported": actual_reported,
        })
    return records


def _matrix_day_summary(records: list[dict[str, object]], current_date: str) -> dict[str, object]:
    """汇总某一天的计划、已填报产量和未填报计划量。"""
    day_rows = [item for item in records if item["date"] == current_date]  # 只聚合当前日期，保留班次明细。
    reported_rows = [item for item in day_rows if item.get("actual_reported")]  # 完成率分母只使用已填报班次。
    plan_total = sum(float(item["plan"]) for item in day_rows)  # 总计划包含尚未填报的班次。
    reported_plan_total = sum(float(item["plan"]) for item in reported_rows)
    actual_total = sum(float(item["actual"]) for item in reported_rows)
    difference_total = sum(float(item["difference"]) for item in reported_rows)
    return {
        "date": current_date,
        "plan": _number_label(plan_total), "actual": _number_label(actual_total),
        "difference": _number_label(difference_total), "reported_plan": _number_label(reported_plan_total),
        "unreported_plan": _number_label(plan_total - reported_plan_total),
        "reported_shift_count": len(reported_rows),
        "unreported_shift_count": len(day_rows) - len(reported_rows),
        "completion_rate": round(actual_total / reported_plan_total * 100, 1) if reported_plan_total else 0,
        "shifts": day_rows,
    }


def _matrix_highlights(shifts: list[dict[str, object]]) -> list[str]:
    """将班次差异转换成最多六条可直接行动的提示。"""
    highlights: list[str] = []
    for item in shifts:
        if not item.get("actual_reported"):
            highlights.append(f"{item['shift']}尚未填报实际产量（计划 {float(item['plan']):g} 台）")  # 优先提醒数据缺口。
        elif float(item["difference"]) < 0:
            highlights.append(f"{item['shift']}较计划少完成 {abs(float(item['difference'])):g} 台")  # 负差异需要管理层关注。
        elif float(item["difference"]) > 0:
            highlights.append(f"{item['shift']}较计划多完成 {float(item['difference']):g} 台")  # 正差异保留为现场进展提示。
    return highlights[:6]  # 限制看板密度，完整明细仍在班次列表中。


def _matrix_insights(rows: list[list[str]], report_date: str | None) -> dict[str, object]:
    """解析生产计划模板中“日期/班次/计划/实际/差异”横向矩阵。

    每一列代表某天的一个班次。实际产量为空表示尚未填报，不能与实际为零混为一谈；
    完成率只以已填报班次对应的计划为分母，避免未到填报时点的班次把当天完成率人为
    拉低。若上传日期不在文件中，则聚焦文件内最近日期并明确标记来源。
    """
    date_entry = _find_row(rows, "日期")
    shift_entry = _find_row(rows, "班次")
    plan_entry = _find_row(rows, "计划")
    actual_entry = _find_row(rows, "实际")
    difference_entry = _find_row(rows, "差异")
    if not (date_entry and shift_entry and plan_entry and actual_entry):
        return {}
    _, date_row = date_entry
    _, shift_row = shift_entry
    _, plan_row = plan_entry
    _, actual_row = actual_entry
    difference_row = difference_entry[1] if difference_entry else []
    label_column = next((index for index, value in enumerate(date_row) if value == "日期"), 0)
    records = _matrix_records(date_row, shift_row, plan_row, actual_row, difference_row, label_column)
    if not records:
        return {}
    available_dates = sorted({str(item["date"]) for item in records})
    focus_date = report_date if report_date in available_dates else available_dates[-1]
    focus_source = "上传日期" if report_date in available_dates else "文件最近日期"
    daily = [_matrix_day_summary(records, current_date) for current_date in available_dates]  # 先做完整日序列，再选焦点日。
    focus = next(item for item in daily if item["date"] == focus_date)
    return {
        "focus_date": focus_date,
        "focus_date_source": focus_source,
        "has_focus_date": report_date in available_dates if report_date else True,
        "plan_total": focus["plan"],
        "actual_total": focus["actual"],
        "difference_total": focus["difference"],
        "reported_plan_total": focus["reported_plan"],
        "unreported_plan_total": focus["unreported_plan"],
        "reported_shift_count": focus["reported_shift_count"],
        "unreported_shift_count": focus["unreported_shift_count"],
        "completion_rate": focus["completion_rate"],
        "daily": daily,
        "shift_summary": focus["shifts"],
        "highlights": _matrix_highlights(focus["shifts"]),
    }


def _tabular_insights(headers: list[str], rows: list[list[str]], report_date: str | None) -> dict[str, object]:
    """解析一行一个班次的传统“计划/实际/差异”表格。

    该函数是横向矩阵解析的备用路径，指标口径与 :func:`_matrix_insights` 保持一致，尤其
    是区分“实际为零”和“尚未填报”，以及完成率仅使用已填报计划作为分母。
    """
    normalized = [value.replace(" ", "").replace("\n", "") for value in headers]  # 表头规范化后再做稳定索引。
    if not {"计划", "实际"}.issubset(set(normalized)):
        return {}
    indexes = {label: normalized.index(label) for label in ("计划", "实际")}
    difference_index = normalized.index("差异") if "差异" in normalized else -1
    shift_index = normalized.index("班次") if "班次" in normalized else -1
    shifts = _tabular_shift_records(rows, indexes, difference_index, shift_index, report_date)
    if not shifts:
        return {}
    summary = _tabular_summary(shifts, report_date)  # 与横向矩阵保持同一指标口径。
    return {
        **summary,
        "shift_summary": shifts,
        "highlights": [
            *(f"{item['shift']}尚未填报实际产量（计划 {float(item['plan']):g} 台）" for item in shifts if not item.get("actual_reported")),
            *(f"{item['shift']}较计划少完成 {abs(float(item['difference'])):g} 台" for item in shifts if item.get("actual_reported") and float(item["difference"]) < 0),
        ][:6],
    }


def _tabular_shift_records(
    rows: list[list[str]], indexes: dict[str, int], difference_index: int,
    shift_index: int, report_date: str | None,
) -> list[dict[str, object]]:
    """将传统逐行生产计划转换为班次记录，跳过无法确认产量语义的行。"""
    shifts: list[dict[str, object]] = []
    for row in rows:
        plan = _as_number(row[indexes["计划"]] if indexes["计划"] < len(row) else None)  # 计划列为空时无法识别业务行。
        actual = _as_number(row[indexes["实际"]] if indexes["实际"] < len(row) else None)  # 空实际代表尚未填报。
        if plan is None and actual is None:
            continue  # 说明行、空行和纯文本行不生成虚假的零产量记录。
        difference = _as_number(row[difference_index] if 0 <= difference_index < len(row) else None)
        actual_reported = actual is not None
        difference_value = difference if actual_reported and difference is not None else ((actual - (plan or 0)) if actual_reported else None)
        shifts.append({
            "date": report_date or "",
            "shift": row[shift_index] if 0 <= shift_index < len(row) and row[shift_index] else "合计",
            "plan": _number_label(plan), "actual": _number_label(actual),
            "difference": _number_label(difference_value), "actual_reported": actual_reported,
        })
    return shifts


def _tabular_summary(shifts: list[dict[str, object]], report_date: str | None) -> dict[str, object]:
    """汇总传统表格班次并生成与矩阵格式一致的单日结构。"""
    reported = [item for item in shifts if item.get("actual_reported")]  # 只把已填报班次纳入完成率分母。
    plan_total = sum(float(item["plan"]) for item in shifts)
    reported_plan_total = sum(float(item["plan"]) for item in reported)
    actual_total = sum(float(item["actual"]) for item in reported)
    difference_total = sum(float(item["difference"]) for item in reported)
    day = {
        "date": report_date or "", "plan": _number_label(plan_total), "actual": _number_label(actual_total),
        "difference": _number_label(difference_total), "reported_plan": _number_label(reported_plan_total),
        "unreported_plan": _number_label(plan_total - reported_plan_total),
        "reported_shift_count": len(reported), "unreported_shift_count": len(shifts) - len(reported),
        "completion_rate": round(actual_total / reported_plan_total * 100, 1) if reported_plan_total else 0,
        "shifts": shifts,
    }
    return {
        "focus_date": report_date or "", "focus_date_source": "上传日期", "has_focus_date": True,
        "plan_total": day["plan"], "actual_total": day["actual"], "difference_total": day["difference"],
        "reported_plan_total": day["reported_plan"], "unreported_plan_total": day["unreported_plan"],
        "reported_shift_count": day["reported_shift_count"], "unreported_shift_count": day["unreported_shift_count"],
        "completion_rate": day["completion_rate"], "daily": [day],
    }


def _team_and_batch_insights(rows: list[list[str]], focus_date: str) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    """从“班组/CASE”横向区块提取聚焦日期的班组与批次产量分布。

    日期同样使用向右填充；只有名称以“组”结尾或包含“班组”的行才按班组统计，避免
    把小计和其他指标行计入。批次名称来自对应列的 CASE 行，保留完整名称供图表标注。
    """
    entry = _find_row(rows, "班组/CASE")
    if not entry:
        return [], []
    date_index, date_row = entry
    label_column = next((index for index, value in enumerate(date_row) if value == "班组/CASE"), 0)
    dates = _fill_forward(date_row, label_column + 1)
    # 班次行与 CASE 行共同确定模板相对结构；当前聚合按班组和批次输出，班次值预留给扩展。
    shift_row = rows[date_index + 1] if date_index + 1 < len(rows) else []
    case_row = rows[date_index + 2] if date_index + 2 < len(rows) else []
    team_totals: dict[str, float] = {}
    batch_totals: dict[str, float] = {}
    for row in rows[date_index + 3:]:
        label = _row_label(row)
        if not label or label in {"小计：", "小计", "台数"}:
            continue
        if not (label.endswith("组") or "班组" in label):
            continue
        total = 0.0
        for column in range(label_column + 1, len(dates)):
            if dates[column] != focus_date:
                continue
            amount = _as_number(row[column] if column < len(row) else None)
            if amount is None:
                continue
            total += amount
            case = case_row[column].strip() if column < len(case_row) else ""
            if case:
                batch_totals[case] = batch_totals.get(case, 0.0) + amount
        team_totals[label] = total
    teams = [{"team": key, "quantity": _number_label(value)} for key, value in sorted(team_totals.items(), key=lambda item: (-item[1], item[0]))]
    batches = [{"batch": key, "quantity": _number_label(value)} for key, value in sorted(batch_totals.items(), key=lambda item: (-item[1], item[0]))]
    return teams, batches


def _meaningful(value: object) -> str:
    """过滤业务表中表示“无值”的常见占位符。"""
    text = str(value or "").strip()
    return "" if text in {"/", "-", "无", "None"} else text


def _month_key(value: object, year: int) -> str:
    """把“8月”或“8月份”等月份标题转换为 ``YYYY-MM``。"""
    text = _meaningful(value)
    match = re.search(r"(\d{1,2})\s*月份?", text)
    if not match:
        return ""
    month = int(match.group(1))
    return f"{year:04d}-{month:02d}" if 1 <= month <= 12 else ""


def _order_date(value: object, year: int) -> str:
    """解析订单模板中的完整日期、Excel 序列号或“月.日”简写。

    无法识别的非空文本原样返回，便于管理员在明细中发现源表问题；语法符合“月.日”
    但日期本身非法时返回空值，避免生成不存在的日期。
    """
    text = _meaningful(value)
    if not text:
        return ""
    if _DATE_RE.match(text):
        return text
    try:
        number = float(text)
    except ValueError:
        number = None
    if number is not None and 20000 < number < 80000:
        # 订单日期从已转成文本的表格行读取，需要在此再次兼容日期序列号。
        try:
            converted = from_excel(number)
            return converted.date().isoformat() if isinstance(converted, dt.datetime) else converted.isoformat()
        except (TypeError, ValueError, OverflowError):
            pass
    match = re.fullmatch(r"(\d{1,2})\.(\d{1,2})", text)
    if match:
        try:
            return dt.date(year, int(match.group(1)), int(match.group(2))).isoformat()
        except ValueError:
            return ""
    return text


def _new_formal_order(row: list[str], current_month: str, year: int) -> dict[str, object]:
    """根据正式订单首行创建订单对象，续行字段由后续合并阶段补充。"""
    return {
        "sequence": _meaningful(row[0]), "month": current_month,  # 月份可能来自前置分组行。
        "order_no": _meaningful(row[2]), "country": _meaningful(row[3]),
        "order_type": _meaningful(row[4]), "quantity": _number_label(_as_number(row[5])),
        "shipment_date": _order_date(row[6], year), "closed_text": _meaningful(row[20]),
        "note": _meaningful(row[21]), "missing_actual_completion": _order_date(row[17], year),
        "hazardous_actual_completion": _order_date(row[18], year),
        "container_actual_completion": _order_date(row[19], year),
        "missing_parts": [], "hazardous_packages": [],  # 明细必须保留列表，供后续闭环统计。
    }


def _merge_formal_order_row(order: dict[str, object], row: list[str], year: int) -> None:
    """把正式订单的一行续行信息合并到当前订单，不用空值覆盖既有字段。"""
    for key, value in (
        ("country", row[3]), ("order_type", row[4]), ("closed_text", row[20]), ("note", row[21]),
    ):
        if _meaningful(value):
            order[key] = _meaningful(value)  # 合并单元格续行只在有值时补充订单级字段。
    quantity = _as_number(row[5])
    if quantity is not None:
        order["quantity"] = _number_label(quantity)  # 后续行若明确给出数量，以最后一个有效值为准。
    for key, column in (
        ("shipment_date", 6), ("missing_actual_completion", 17),
        ("hazardous_actual_completion", 18), ("container_actual_completion", 19),
    ):
        if _meaningful(row[column]):
            order[key] = _order_date(row[column], year)  # 日期字段沿用同一年度解析规则。
    missing_code = _meaningful(row[7])
    if missing_code:
        order["missing_parts"].append({
            "material_code": missing_code, "material_name": _meaningful(row[8]),
            "quantity": _number_label(_as_number(row[9])), "shipment_order_no": _meaningful(row[10]),
            "shipment_date": _order_date(row[11], year),
        })  # 缺件明细按出现顺序保存，导出时保持原表阅读顺序。
    hazardous_code = _meaningful(row[12])
    if hazardous_code:
        order["hazardous_packages"].append({
            "material_code": hazardous_code, "material_name": _meaningful(row[13]),
            "quantity": _number_label(_as_number(row[14])), "shipment_order_no": _meaningful(row[15]),
            "shipment_date": _order_date(row[16], year),
        })  # 危包明细与缺件明细分开保存，避免两种业务被混为一类。


def _finalize_formal_order(order: dict[str, object], report_date: str | None) -> None:
    """在所有续行合并完成后计算正式订单及其缺件状态。"""
    if not order.get("month"):
        order["month"] = str(order.get("shipment_date") or report_date or "")[:7]  # 没有月份时从发运日期回填。
    missing_completion = str(order.get("missing_actual_completion") or "")
    hazardous_completion = str(order.get("hazardous_actual_completion") or "")
    for item in order["missing_parts"]:
        item["completed"] = bool(item.get("shipment_date") or missing_completion)  # 明细发运或整体完成即闭环。
    for item in order["hazardous_packages"]:
        item["completed"] = bool(item.get("shipment_date") or hazardous_completion)
    closed_text = str(order.get("closed_text") or "")
    explicit_closed = any(token in closed_text for token in ("是", "关闭", "完结", "完成"))
    order["completed"] = explicit_closed  # 整单状态只认模板明确关闭语义，不从缺件状态反推。
    order["status"] = "已关闭" if explicit_closed else ("已发运待关闭" if order.get("shipment_date") else "待发运")
    order["outstanding_missing_count"] = sum(not bool(item.get("completed")) for item in order["missing_parts"])
    order["outstanding_hazardous_count"] = sum(not bool(item.get("completed")) for item in order["hazardous_packages"])


def _formal_orders(rows: list[list[str]], report_date: str | None) -> list[dict[str, object]]:
    """解析正式订单台账及其跨行缺件、危包明细。

    模板使用合并单元格语义：订单号只出现在首行，后续行继续列出该订单的缺件或危包。
    ``current`` 保存最近订单，使这些续行附着到正确订单；同一订单号再次出现时复用原
    对象。订单关闭状态只依据明确关闭文本，不用缺件完成情况反推整单关闭。
    """
    year = int((report_date or dt.date.today().isoformat())[:4])
    orders: list[dict[str, object]] = []
    by_order: dict[str, dict[str, object]] = {}
    current: dict[str, object] | None = None
    current_month = ""
    for source in rows:
        row = list(source) + [""] * max(0, 22 - len(source))  # 补齐模板列，后续固定索引可安全读取。
        if _meaningful(row[7]) == "物料号" or _meaningful(row[2]) == "订单号":
            continue
        month = _month_key(row[1], year)
        if month:
            # 月份通常只在分组首行填写，后续订单沿用最近出现的月份。
            current_month = month
        order_no = _meaningful(row[2])
        if order_no:
            current = by_order.get(order_no)
            if current is None:
                current = _new_formal_order(row, current_month, year)  # 首次出现时建立订单主对象。
                by_order[order_no] = current  # 后续跨行记录通过订单号复用同一对象。
                orders.append(current)  # 保持订单首次出现的顺序，报告阅读顺序稳定。
        if current is None:
            continue
        _merge_formal_order_row(current, row, year)  # 订单主字段、缺件和危包在同一处集中合并。
    for order in orders:
        _finalize_formal_order(order, report_date)  # 全部续行读取完毕后再判定闭环状态。
    return orders


def _new_sporadic_order(row: list[str]) -> dict[str, object]:
    """从零星订单首行创建主对象，列表字段等待后续续行共同补充。"""
    return {
        "order_no": _meaningful(row[0]),
        "transport_mode": _meaningful(row[1]),
        "country": _meaningful(row[2]),
        "order_type": _meaningful(row[3]),
        "consolidated": _meaningful(row[4]),
        "container_quantity": _number_label(_as_number(row[5])),
        "container_type": _meaningful(row[6]),
        "shipment_dates": [],
        "pallets": [],
        "driver_plate": _meaningful(row[13]),
        "driver_name": _meaningful(row[14]),
        "driver_phone": _meaningful(row[15]),
        "notes": [],
    }


def _append_unique(values: list[str], value: str) -> None:
    """向保持源顺序的文本列表追加非空且未出现的值。"""
    if value and value not in values:
        values.append(value)


def _merge_sporadic_order_row(order: dict[str, object], row: list[str], year: int) -> None:
    """把零星订单主行或续行的运输、日期、托盘和备注合并到当前订单。"""
    for key, column in (
        ("transport_mode", 1), ("country", 2), ("order_type", 3), ("consolidated", 4),
        ("container_type", 6), ("driver_plate", 13), ("driver_name", 14),
        ("driver_phone", 15),
    ):
        value = _meaningful(row[column])
        if value:
            order[key] = value  # 续行显式填写的值覆盖较早空值或旧值，符合台账最后填写口径。
    _append_unique(order["shipment_dates"], _order_date(row[12], year))
    _append_unique(order["notes"], _meaningful(row[16]))
    if any(_meaningful(row[column]) for column in range(7, 12)):
        # 托盘数量或任一尺寸存在时才创建明细，空续行不会制造零值托盘。
        order["pallets"].append({
            "pallet_count": _number_label(_as_number(row[7])),
            "length_mm": _number_label(_as_number(row[8])),
            "width_mm": _number_label(_as_number(row[9])),
            "height_mm": _number_label(_as_number(row[10])),
            "volume_cbm": _number_label(_as_number(row[11])),
        })


def _finalize_sporadic_order(order: dict[str, object], report_date: str | None) -> None:
    """汇总零星订单日期、托数、体积和最终发运状态。"""
    dates = order["shipment_dates"]
    order["shipment_date"] = dates[-1] if dates else ""  # 看板主日期使用最后一次实际发运日期。
    order["month"] = (dates[0] if dates else (report_date or ""))[:7]
    order["pallet_count"] = _number_label(sum(
        float(item.get("pallet_count") or 0) for item in order["pallets"]
    ))
    order["volume_cbm"] = _number_label(sum(
        float(item.get("volume_cbm") or 0) for item in order["pallets"]
    ))
    note_text = "；".join(order["notes"])
    order["note"] = note_text
    order["completed"] = bool(dates) or any(token in note_text for token in ("完结", "完成", "已发"))
    order["status"] = "已发运" if order["completed"] else "待发运"


def _sporadic_orders(rows: list[list[str]], report_date: str | None) -> list[dict[str, object]]:
    """解析零星订单、托盘尺寸、分次发运日期和司机信息。

    一个订单可跨多行记录不同托盘或不同发运日期，处理方式与正式订单相同：订单号非空
    时切换当前订单，后续空订单号行继续附着。日期和备注去重，托数与体积在全部行读取
    完毕后汇总。
    """
    year = int((report_date or dt.date.today().isoformat())[:4])
    orders: list[dict[str, object]] = []
    by_order: dict[str, dict[str, object]] = {}
    current: dict[str, object] | None = None
    for source in rows:
        row = list(source) + [""] * max(0, 17 - len(source))  # 固定列模板不足时补空值。
        order_no = _meaningful(row[0])
        if order_no and order_no != "订单号":
            current = by_order.get(order_no)
            if current is None:
                current = _new_sporadic_order(row)
                by_order[order_no] = current
                orders.append(current)  # 保持订单首现顺序，报告与源台账阅读顺序一致。
        if current is None:
            continue
        _merge_sporadic_order_row(current, row, year)
    for order in orders:
        _finalize_sporadic_order(order, report_date)
    return orders


def _monthly_order_summary(
    formal: list[dict[str, object]], sporadic: list[dict[str, object]],
) -> list[dict[str, object]]:
    """按月份汇总正式订单与零星订单的数量、完成数和体积指标。"""
    months = sorted({str(item.get("month") or "") for item in formal + sporadic if item.get("month")})
    summary = []
    for month in months:
        formal_rows = [item for item in formal if item.get("month") == month]
        sporadic_rows = [item for item in sporadic if item.get("month") == month]
        summary.append({
            "month": month, "formal_total": len(formal_rows),
            "formal_completed": sum(bool(item.get("completed")) for item in formal_rows),
            "formal_quantity": _number_label(sum(float(item.get("quantity") or 0) for item in formal_rows)),
            "sporadic_total": len(sporadic_rows),
            "sporadic_completed": sum(bool(item.get("completed")) for item in sporadic_rows),
            "sporadic_pallets": _number_label(sum(float(item.get("pallet_count") or 0) for item in sporadic_rows)),
            "sporadic_volume_cbm": _number_label(sum(float(item.get("volume_cbm") or 0) for item in sporadic_rows)),
        })
    return summary


def _order_ledger_insights(sheets: list[dict[str, object]], report_date: str | None) -> dict[str, object]:
    """汇总工作簿内所有正式订单与零星订单页签，并生成月度指标。"""
    formal: list[dict[str, object]] = []
    sporadic: list[dict[str, object]] = []
    for sheet in sheets:
        kind = str(sheet.get("kind") or "")
        rows = sheet.get("table_rows") if isinstance(sheet.get("table_rows"), list) else []
        if kind == "正式订单":
            formal.extend(_formal_orders(rows, report_date))
        elif kind == "零星订单":
            sporadic.extend(_sporadic_orders(rows, report_date))
    return {
        "formal_orders": formal, "sporadic_orders": sporadic,
        "monthly_summary": _monthly_order_summary(formal, sporadic),
        "missing_parts": [dict(item, order_no=order["order_no"]) for order in formal for item in order["missing_parts"]],
        "hazardous_packages": [dict(item, order_no=order["order_no"]) for order in formal for item in order["hazardous_packages"]],
    }


def _shipping_summary(ledger: dict[str, object]) -> list[dict[str, object]]:
    """把两类订单压缩成看板使用的发运总数、完成数和数量指标。"""
    result = []
    for kind, key, quantity_key in (
        # 正式订单以台数汇总，零星订单以托数汇总，两类数量含义不同但结构统一。
        ("正式订单", "formal_orders", "quantity"), ("零星订单", "sporadic_orders", "pallet_count"),
    ):
        rows = ledger.get(key) if isinstance(ledger.get(key), list) else []
        if rows:
            completed = sum(bool(item.get("completed")) for item in rows)
            quantity = sum(float(item.get(quantity_key) or 0) for item in rows)
            result.append({
                "type": kind, "total": len(rows), "completed": completed,
                "pending": len(rows) - completed, "quantity": _number_label(quantity),
            })
    return result


def _empty_production_insights(report_date: str | None) -> dict[str, object]:
    """返回字段齐全的空生产洞察，保持前端图表协议稳定。"""
    return {
        "focus_date": report_date or "",
        "focus_date_source": "暂无生产计划矩阵",
        "has_focus_date": False,
        "plan_total": 0,
        "actual_total": 0,
        "difference_total": 0,
        "reported_plan_total": 0,
        "unreported_plan_total": 0,
        "reported_shift_count": 0,
        "unreported_shift_count": 0,
        "completion_rate": 0,
        "daily": [],
        "shift_summary": [],
        "team_summary": [],
        "batch_summary": [],
        "highlights": [],
    }


def _production_plan_insights(
    headers: list[object],
    rows: list[list[str]],
    report_date: str | None,
) -> dict[str, object]:
    """按“横向矩阵优先、传统表格兜底”的规则解析一个生产计划页。"""
    insights = _matrix_insights(rows, report_date) or _tabular_insights(headers, rows, report_date)
    if not insights:
        return {}
    focus_date = str(insights.get("focus_date") or report_date or "")
    teams, batches = _team_and_batch_insights(rows, focus_date)
    if teams:
        insights["team_summary"] = teams
    if batches:
        insights["batch_summary"] = batches
    return insights


def _actual_packaging_quantity(headers: list[object], rows: list[list[str]]) -> float:
    """汇总生产实绩中的实际包装数量；缺少目标列时返回零。"""
    quantity_index = next(
        (
            index
            for index, value in enumerate(headers)
            if "实际包装数量" in str(value).replace(" ", "")
        ),
        -1,
    )
    if quantity_index < 0:
        return 0.0
    return sum(
        _as_number(row[quantity_index]) or 0
        for row in rows
        if quantity_index < len(row)
    )


def _build_insights(sheets: list[dict[str, object]], report_date: str | None) -> dict[str, object]:
    """组合生产计划、生产实绩和订单发运三类结构化洞察。

    生产计划优先解析横向矩阵，失败后尝试传统表格；生产实绩只能补充已有计划洞察，
    不能单独伪造计划完成率。无计划数据时返回字段齐全的空结构，前端无需到处判断键是否
    存在。订单台账始终独立生成并附加到结果。
    """
    insights: dict[str, object] = {}
    order_ledger = _order_ledger_insights(sheets, report_date)
    for sheet in sheets:
        kind = str(sheet.get("kind") or "")
        rows = sheet.get("table_rows") if isinstance(sheet.get("table_rows"), list) else []
        headers = sheet.get("table_headers") if isinstance(sheet.get("table_headers"), list) else []
        if kind == "生产计划":
            # 同一工作簿可能存在多个计划页；成功解析的新页覆盖旧页，无法解析则保留旧结果。
            insights = _production_plan_insights(headers, rows, report_date) or insights
        elif kind == "生产实绩" and insights:
            actual_quantity = _actual_packaging_quantity(headers, rows)
            if actual_quantity:
                # 实绩表的包装数量作为辅助校验值，不覆盖计划表中的分班实际数据。
                insights["actual_quantity_from_result"] = _number_label(actual_quantity)

    if not insights:
        insights = _empty_production_insights(report_date)
    insights.setdefault("team_summary", [])
    insights.setdefault("batch_summary", [])
    insights["order_ledger"] = order_ledger
    insights["shipping_summary"] = _shipping_summary(order_ledger)
    return insights


def _read_plan_sheet(worksheet: Any, workbook: Any) -> tuple[dict[str, object], int]:
    """读取单个生产计划页签的预览与结构化数据，返回页签记录和源行数。

    每个页签最多保留 ``MAX_PREVIEW_ROWS`` 行首屏预览、``MAX_TABLE_ROWS`` 行结构化
    数据；``row_count`` 仍统计源工作表真实行数，用于截断提示。
    """
    rows: list[list[str]] = []
    table_rows: list[list[str]] = []
    row_count = 0
    for raw_row in worksheet.iter_rows():
        row_count += 1
        raw_values = [cell.value for cell in raw_row[:MAX_PREVIEW_COLUMNS]]
        # 横向日期模板的单元格可能没有日期格式，看到日期语义行时强制尝试序列号转换。
        force_excel_date = any(value in {"日期", "班组/CASE"} for value in raw_values)
        values = [_cell_text(cell.value, number_format=cell.number_format, epoch=workbook.epoch, force_excel_date=force_excel_date) for cell in raw_row[:MAX_PREVIEW_COLUMNS]]
        if len(rows) < MAX_PREVIEW_ROWS:
            rows.append(values)
        if len(table_rows) < MAX_TABLE_ROWS and any(values):
            # 分析区忽略全空行，但 row_count 保留源工作表真实行数用于截断提示。
            table_rows.append(values)
    header_index = _find_header_index(table_rows)
    if header_index is not None:
        headers = table_rows[header_index]
        data_rows = table_rows[header_index + 1:]
    else:
        # 横向日期矩阵没有传统表头，保留所有非空行，避免前端只剩下首屏 12 行。
        headers, data_rows = [], table_rows
    sheet_kind = _table_kind(worksheet.title, headers)
    # 没有传统表头时，用前二十四行整体语义补判横向生产计划或到料成品表。
    flattened = "".join(cell.replace(" ", "") for row in table_rows[:24] for cell in row if cell)
    if sheet_kind == "通用数据" and all(token in flattened for token in ("计划", "实际", "差异", "班次")):
        sheet_kind = "生产计划"
    if sheet_kind == "通用数据" and all(token in flattened for token in ("主料总共类", "到货数量", "剩余未收数")):
        sheet_kind = "到料明细"
    sheet = {
        "sheet": worksheet.title,
        "rows": row_count,
        "columns": worksheet.max_column,
        "preview": rows,
        "kind": sheet_kind,
        "table_headers": headers,
        "table_rows": data_rows,
        "table_truncated": row_count > MAX_TABLE_ROWS,
    }
    return sheet, row_count


def analyze(path: str | os.PathLike[str], report_date: str | None = None) -> dict[str, object]:
    """读取生产计划工作簿的可展示摘要，不修改原文件。

    仅支持 openpyxl 可安全读取的现代格式。每个页签保留少量首屏预览，并最多保留
    ``MAX_TABLE_ROWS`` 行供结构化分析；截断状态会返回前端提示管理员。工作簿始终在
    ``finally`` 中关闭，单个损坏文件会转换为业务可读错误。
    """
    target = Path(path).resolve()
    if target.suffix.lower() not in SUPPORTED_EXTENSIONS:
        raise ValueError("生产计划仅支持 .xlsx 或 .xlsm 文件")
    if not target.is_file():
        raise ValueError("生产计划文件不存在")
    try:
        # 只读、公式缓存值且不加载外部链接，降低大表内存占用和外部引用风险。
        workbook = load_workbook(target, read_only=True, data_only=True, keep_links=False)
    except Exception as exc:  # pragma: no cover - openpyxl 已给出具体错误
        raise ValueError("生产计划文件无法读取，请确认文件没有损坏") from exc
    sheets: list[dict[str, object]] = []
    total_rows = 0
    try:
        for worksheet in workbook.worksheets:
            sheet, row_count = _read_plan_sheet(worksheet, workbook)
            sheets.append(sheet)
            total_rows += row_count
    finally:
        workbook.close()
    return {
        "file_name": target.name,
        "sheet_count": len(sheets),
        "row_count": total_rows,
        "sheets": sheets,
        "insights": _build_insights(sheets, report_date),
    }
