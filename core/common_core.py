# -*- coding: utf-8 -*-
"""
公共核心模块 —— 考勤填报 与 工时对账 两功能共用
=================================================
把两个功能原先各自实现、且细节不一致的工具函数统一到这里，避免同名不同义：
  · 姓名规范化：统一为“去掉所有空格”（原填报只去首尾，会导致“张 三”对不上）
  · 日期、时间、数字解析：各一份权威实现
  · read_sheets：统一的多格式读取（.xlsx/.xlsm/.xls）
  · 常量：标准工时、容差、假/休标记
  · Options：高级选项载体（供用户手动干预识别）
  · 输出路径：两功能统一「源文件目录/output/名字_后缀_时间戳.ext」

兼容 Windows 10/11 + Python 3.13。
"""
import os
import re
import datetime
import threading
import unicodedata

import openpyxl

# ---------------- 统一常量 ----------------
# 标准工作时长（小时）：实际工时超出此值的部分记为加班，不足记 0。
STANDARD_WORKDAY_HOURS = 9.0
# 数值容差（工时以 0.5 为最小单位，容差取 0.01 判断相等）
TOL = 0.01
# 视为“非工时”的标记（假/休/空白等），不参与逐日对比
SKIP_MARKS = {"假", "休", "旷", "缺", "调休", "病假", "事假", "年假", ""}

# 中文录入/Excel 常见脏字符:全角数字·全角标点·全角冒号来自中文输入法;
# 零宽字符(​/﻿)与软连字符肉眼不可见,却让"同一个值"对不上、数字转不出。
_INVISIBLE = dict.fromkeys(
    [0x200b, 0x200c, 0x200d, 0xfeff, 0x00ad, 0x2060], None)   # 删除


def clean_str(v):
    """单元格文本清洗:NFKC 归一(全角→半角:数字/字母/冒号/句点/括号/逗号,
    NBSP→普通空格) + 删零宽/软连字符,再去首尾空白。None→''。
    只做"形态"归一,不动语义(不去前导零、不改大小写、不拆单位)。"""
    if v is None:
        return ""
    s = unicodedata.normalize("NFKC", str(v)).translate(_INVISIBLE)
    return s.strip()


def _num_str(v):
    """把文本形态的数字规整成可 float 的串:清洗 + 去千分位逗号。
    仅当去逗号后为合法数字才去(避免误伤真含逗号的非数字文本)。"""
    s = clean_str(v)
    if "," in s:
        t = s.replace(",", "")
        if re.fullmatch(r"[+-]?\d+(\.\d+)?", t):
            return t
    return s


# ---------------- 列角色定义（供可视化列映射界面使用） ----------------
# 每种“文件类型”需要指定哪些列角色。key=角色内部名，value=(中文名, 是否必填)
ROLE_DEFS = {
    "att_source": [("name", "姓名", True), ("date", "日期", True),
                   ("on", "上班1打卡时间", True), ("off", "下班1打卡时间", True)],
    "att_target": [("name", "姓名", True), ("date", "日期", True),
                   ("sys_on", "上班时间(系统)", False), ("act_on", "上班时间(实际)", False),
                   ("sys_off", "下班时间(系统)", False), ("act_off", "下班时间(实际)", False),
                   ("rest", "休息时间", False), ("work", "实际工作时间", False),
                   ("ot", "加班", False)],
    "rec_source": [("name", "姓名", True), ("date", "日期", True),
                   ("work", "实际工作时间", True)],
    "rec_zong": [("name", "姓名", True), ("comp", "所属劳务公司", False),
                 ("work", "出勤工时", False), ("check", "对账时间", False)],
    "rec_labor": [("name", "姓名", True), ("total", "合计/出勤工时列", False)],
}
KIND_TITLES = {
    "att_source": "填报·系统数据表", "att_target": "填报·待填考勤表",
    "rec_source": "对账·数据来源", "rec_zong": "对账·待对总表",
    "rec_labor": "对账·劳务对账单",
}


# ---------------- 高级选项 ----------------
class Options:
    """高级选项载体：让用户在程序识别有误时手动干预。

    workday_hours : 每日标准工时（小时），加班基准。
    overtime      : 是否计算加班列。
    conflict      : 同一(姓名,日期)在多文件重复时的策略 last/first/warn。
    header_row    : 全局手动表头行(1-based)；None=自动。（per-file 映射优先）
    sheet_name    : 全局手动工作表名；None=自动/全部。（per-file 映射优先）
    tolerance     : 对账工时比对容差(小时)。
    data_start    : 手动数据起始行(1-based)；None=表头下一行。（per-file 映射优先）
    skip_extra    : 追加的“非工时”标记词集合（与内置 SKIP_MARKS 合并）。
    columns       : per-file 列映射 {文件名basename: {"sheet":名或None, "header":行1based或None,
                    "data_start":行或None, "roles":{角色:列0based}}}。手动映射优先于自动识别。
    auto_actual   : 是否自动按半小时算“实际上/下班时间”（上班进位、下班退位）。默认 True。
    night_shift   : 是否启用两班制夜班识别（跨零点 +24 修正）。默认 True。
    night_start_hour   : 实际上班钟点 ≥ 此值判为夜班。默认 17.0。
    night_workday_hours: 夜班标准工时（加班基准）。默认 11.0。
    night_max_hours    : 夜班合理工时上限，超过判异常（防漏打卡）。默认 16.0。
    """
    def __init__(self, workday_hours=STANDARD_WORKDAY_HOURS, overtime=True,
                 conflict="last", header_row=None, sheet_name=None, tolerance=TOL,
                 data_start=None, skip_extra=None, columns=None,
                 auto_actual=True, night_shift=True, night_start_hour=17.0,
                 night_workday_hours=11.0, night_max_hours=16.0):
        self.workday_hours = float(workday_hours)
        self.overtime = bool(overtime)
        # 自动按半小时算“实际上/下班时间”：上班进位、下班退位，再据此算实际工时。
        self.auto_actual = bool(auto_actual)
        # 两班制：按实际上班钟点区分白/夜班；夜班跨零点自动 +24 修正。
        self.night_shift = bool(night_shift)              # 是否启用夜班识别
        self.night_start_hour = float(night_start_hour)   # 上班打卡≥此钟点 → 夜班
        self.night_workday_hours = float(night_workday_hours)  # 夜班标准工时(加班基准)
        self.night_max_hours = float(night_max_hours)     # 夜班合理工时上限(超则判异常)
        self.conflict = conflict if conflict in ("last", "first", "warn") else "last"
        self.header_row = header_row
        self.sheet_name = (sheet_name or None)
        self.tolerance = float(tolerance)
        self.data_start = data_start
        self.skip_extra = set(skip_extra) if skip_extra else set()
        self.columns = columns if columns else {}

    def skip_set(self):
        """内置 + 自定义 的“非工时”标记词集合。"""
        return SKIP_MARKS | self.skip_extra

    def file_map(self, path):
        """取某文件的 per-file 列映射；无则 None。path 可为完整路径或 basename。"""
        if not self.columns:
            return None
        return self.columns.get(os.path.basename(path)) or self.columns.get(path)

    def resolve_sheet(self, path):
        """该文件应处理的工作表名：per-file > 全局 > None(自动)。"""
        fm = self.file_map(path)
        if fm and fm.get("sheet"):
            return fm["sheet"]
        return self.sheet_name

    def resolve_header(self, path):
        """该文件表头行(1-based)：per-file > 全局 > None。"""
        fm = self.file_map(path)
        if fm and fm.get("header"):
            return fm["header"]
        return self.header_row

    def resolve_data_start(self, path):
        """该文件数据起始行(1-based)：per-file > 全局 > None。"""
        fm = self.file_map(path)
        if fm and fm.get("data_start"):
            return fm["data_start"]
        return self.data_start

    def resolve_roles(self, path):
        """该文件手动列映射 {角色:列0based}；无则 {}。"""
        fm = self.file_map(path)
        return dict(fm["roles"]) if (fm and fm.get("roles")) else {}

    def summary(self):
        """一行文字，用于日志追溯本次采用的选项。"""
        cn = {"last": "后者覆盖", "first": "先者优先", "warn": "不覆盖仅提示"}
        parts = ["标准工时=%g" % self.workday_hours,
                 "加班=%s" % ("算" if self.overtime else "不算"),
                 "重复=%s" % cn.get(self.conflict, self.conflict),
                 "容差=%g" % self.tolerance]
        if self.header_row:
            parts.append("表头行=%d" % self.header_row)
        if self.sheet_name:
            parts.append("工作表=%s" % self.sheet_name)
        if self.skip_extra:
            parts.append("额外假休标记=%s" % "/".join(sorted(self.skip_extra)))
        parts.append("实际时间=%s" % ("自动半小时进退位" if self.auto_actual else "不自动"))
        if self.night_shift:
            parts.append("夜班=启用(≥%g点/标准%gh/上限%gh)"
                         % (self.night_start_hour, self.night_workday_hours, self.night_max_hours))
        else:
            parts.append("夜班=不识别")
        if self.columns:
            parts.append("列映射=%d个文件" % len(self.columns))
        return "；".join(parts)


DEFAULTS = Options()


# ---------------- 统一解析工具 ----------------
def norm_name(v):
    """姓名规范化：去掉所有空格（含中间空格）。两功能一致，避免“张 三”对不上。"""
    if v is None:
        return ""
    return re.sub(r"\s+", "", str(v))


# Excel 日期序列号的合理区间:约 1990-01-01(32874) ~ 2100-12-31(73415)。
# 下界远高于 31,确保绝不会把"当月第几天"整数误当日期序列;上界防脏数据。
_EXCEL_SERIAL_MIN = 32874
_EXCEL_SERIAL_MAX = 73415


def _serial_to_date(n):
    """Excel 日期序列号 -> datetime.date;越界或失败返回 None。
    日期列一旦丢失单元格格式,openpyxl 会读到原始序列号(如 46143=2026-05-01),
    不还原就会导致该列每一行都对不上、整表静默漏读。"""
    try:
        f = float(n)
    except (ValueError, TypeError):
        return None
    if not (_EXCEL_SERIAL_MIN <= f <= _EXCEL_SERIAL_MAX):
        return None
    try:
        from openpyxl.utils.datetime import from_excel
        d = from_excel(f)
        return d.date() if isinstance(d, datetime.datetime) else d
    except Exception:
        return None


def norm_date(v):
    """把各种日期形式统一成 (year, month, day) 元组，无法解析返回 None。"""
    if v is None:
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return (v.year, v.month, v.day)
    # 裸数字(int/float):按 Excel 日期序列号还原(丢格式的日期列常见)
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        d = _serial_to_date(v)
        return (d.year, d.month, d.day) if d else None
    s = clean_str(v)
    if not s or s == "-":
        return None
    s = s.split()[0]                           # "2026-05-01 08:00:00" 取日期段
    digits = "".join(ch for ch in s if ch.isdigit())
    def make_date(year, month, day):
        try:
            d = datetime.date(int(year), int(month), int(day))
            return (d.year, d.month, d.day)
        except (TypeError, ValueError):
            return None

    if len(digits) == 8:                       # 20260501
        try:
            return make_date(digits[0:4], digits[4:6], digits[6:8])
        except ValueError:
            pass
    cn = re.fullmatch(r"(\d{4})年(\d{1,2})月(\d{1,2})日?", s)
    if cn:
        return make_date(*cn.groups())
    for sep in ("-", "/", "."):                # 2026-05-01 / 2026/5/1
        if sep in s:
            parts = s.split(sep)
            if len(parts) == 3:
                try:
                    return make_date(parts[0], parts[1], parts[2])
                except ValueError:
                    pass
    # 纯数字文本形式的序列号("46143")兜底
    if s.isdigit():
        d = _serial_to_date(int(s))
        if d:
            return (d.year, d.month, d.day)
    return None


def day_of(v):
    """把各种日期表示转成“当月第几天”(1~31)。识别失败返回 None。"""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.day
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        n = int(v)
        if 1 <= n <= 31:                       # 直接就是"第几天"
            return n
        d = _serial_to_date(v)                 # 否则按 Excel 日期序列号还原
        return d.day if d else None
    s = clean_str(v)
    s = s.split()[0] if s else s               # "2026-05-01 08:00:00" 取日期段
    date_value = norm_date(s)
    if date_value:
        return date_value[2]
    m = re.search(r"(\d{1,2})\s*日?\s*$", s)   # 结尾数字或“5日”
    if m:
        n = int(m.group(1))
        if 1 <= n <= 31:
            return n
    # 纯数字文本形式的序列号("46143")兜底
    if s.isdigit():
        d = _serial_to_date(int(s))
        if d:
            return d.day
    return None


def parse_time(v):
    """把单元格值解析成 datetime.time；无效返回 None。"""
    if v is None:
        return None
    if isinstance(v, datetime.time):
        return v
    if isinstance(v, datetime.datetime):
        return v.time()
    s = clean_str(v)                           # 全角冒号"："→":"、去零宽等
    if not s or s in ("-", "—"):
        return None
    if ":" in s:
        parts = s.split(":")
        try:
            h = int(parts[0]); m = int(parts[1])
            sec = int(parts[2]) if len(parts) > 2 else 0
            return datetime.time(h, m, sec)
        except (ValueError, IndexError):
            return None
    return None


def to_hours(t):
    """datetime.time -> 小时数(float)。None -> None。"""
    if t is None:
        return None
    return t.hour + t.minute / 60.0 + t.second / 3600.0


def round_half_hour(t, mode):
    """按半小时把 datetime.time 取整到最近的整点/半点。

    mode="up"  进位（用于上班）：向上取到 :00 或 :30。7:56→8:00，7:31→8:00，7:30→7:30。
    mode="down"退位（用于下班）：向下取到 :00 或 :30。8:13→8:00，8:24→8:00，8:30→8:30。
    恰好落在整点/半点则不变。t 为 None 返回 None。
    """
    import math
    if t is None:
        return None
    total = t.hour * 3600 + t.minute * 60 + t.second      # 当日秒数
    step = 1800                                            # 半小时
    if mode == "up":
        secs = int(math.ceil(total / float(step))) * step
    else:                                                  # down
        secs = (total // step) * step
    secs = max(0, min(secs, 86400 - step))                # 兜底钳制（考勤为日间，进位不会到 24:00）
    return datetime.time(secs // 3600, (secs % 3600) // 60)


def fmt_time(t):
    """datetime.time -> 'HH:MM' 字符串；None -> ''。"""
    if t is None:
        return ""
    return "%02d:%02d" % (t.hour, t.minute)


def parse_rest(v):
    """休息时间解析为小时数(float)。空/无效 -> 0。"""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = clean_str(v)
    if not s or s in ("-", "—"):
        return 0.0
    try:
        return float(_num_str(v))
    except ValueError:
        return 0.0


def to_num(v, skip=None):
    """把单元格值转成 float；无法转（含假/休/空白）返回 None。
    skip: 可选“非工时”标记词集合，默认用内置 SKIP_MARKS。"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = clean_str(v)
    marks = skip if skip is not None else SKIP_MARKS
    if s == "" or s in marks:
        return None
    try:
        return float(_num_str(v))
    except ValueError:
        return None


# ---------------- 公式未刷新检测（贯穿性根因） ----------------
def detect_uncached_formula(path, sheet=None, sample_rows=400):
    """检测工作簿里是否有"公式单元格但无缓存值"。

    openpyxl 以 data_only=True 读取时，公式单元格若从未被 Excel 打开并保存过，
    缓存值为 None —— 下游会把它误当空/0，造成静默丢数。本函数用 data_only=False
    再扫一遍：某格是公式(值以 '=' 开头)却在 data_only 读取下为 None，即判定该表
    "公式未刷新"。返回命中的列(0-based)集合；空集表示无此问题。
    仅 .xlsx/.xlsm 适用（.xls 无此机制），其余或读取失败一律返回 set()（不误报）。
    """
    ext = os.path.splitext(path)[1].lower()
    if ext not in (".xlsx", ".xlsm"):
        return set()
    cols = set()
    try:
        wb_f = openpyxl.load_workbook(path, data_only=False, read_only=True)
        wb_v = openpyxl.load_workbook(path, data_only=True, read_only=True)
        try:
            names = [sheet] if sheet and sheet in wb_f.sheetnames else wb_f.sheetnames
            for nm in names:
                wf, wv = wb_f[nm], wb_v[nm]
                fit = wf.iter_rows(max_row=sample_rows, values_only=True)
                vit = wv.iter_rows(max_row=sample_rows, values_only=True)
                for frow, vrow in zip(fit, vit):
                    for c, fval in enumerate(frow):
                        if (isinstance(fval, str) and fval.startswith("=")
                                and (c >= len(vrow) or vrow[c] is None)):
                            cols.add(c)
        finally:
            wb_f.close(); wb_v.close()
    except Exception:
        return set()          # 检测本身失败绝不阻断主流程，也不误报
    return cols


def warn_if_uncached(path, log, sheet=None, what="数据"):
    """便捷封装：检测到公式未刷新则通过 log 发醒目警告，返回是否命中。
    各核心在读关键表前调用，把"公式未刷新→静默丢数"变成用户可见的提示。"""
    cols = detect_uncached_formula(path, sheet)
    if cols:
        log("⚠ 警告：《%s》中%s所在列含未刷新的公式（读取值为空），可能导致漏算或算错。"
            % (os.path.basename(path), what))
        log("  请先用 Excel 打开该表、按 Ctrl+S 保存一次以刷新公式后重试。")
        return True
    return False


# ---------------- 读取加速：跳过输入文件内嵌的 Excel 透视缓存 ----------------
import contextlib


_PIVOT_CACHE_LOCK = threading.RLock()


@contextlib.contextmanager
def _skip_pivot_cache_parse():
    """临时让 openpyxl 读取时不解析工作簿内嵌的 PivotTable 缓存。

    背景：源文件常自带 Excel 透视表，openpyxl 加载时会 eager 反序列化其
    pivotCache（实测占某些透视源读取耗时的 ~64%）。而本程序读输入只取单元格
    值(data_only)，从不访问 ws._pivots，这些缓存是纯粹的死负载。

    做法：把 WorkbookParser.pivot_caches 属性临时替换为"取任意键都返回 None"
    的空映射——read_worksheets 里 `pivot_caches[pivot.cacheId]` 仍可取值(得到
    None)、`ws.add_pivot` 照常调用，只是不再触发昂贵的缓存解析。退出即还原。
    openpyxl 内部结构若与预期不符(版本差异)，则安全退化为不做任何改动。
    """
    # WorkbookParser.pivot_caches 是进程级类属性。多个 Worker 若交错替换/还原，
    # 后退出的线程会把前一线程的临时 property 当成“原值”永久写回。
    # 整个替换窗口必须串行；RLock 保留同线程嵌套读取能力。
    with _PIVOT_CACHE_LOCK:
        try:
            from openpyxl.reader.workbook import WorkbookParser
        except Exception:
            yield                      # 拿不到内部类：老实走原路径，不加速也不报错
            return
        if not isinstance(getattr(WorkbookParser, "pivot_caches", None), property):
            yield
            return

        class _NullCaches(dict):
            def __missing__(self, key):       # 任意 cacheId 都返回 None，不解析
                return None

        original = WorkbookParser.pivot_caches
        WorkbookParser.pivot_caches = property(lambda self: _NullCaches())
        try:
            yield
        finally:
            WorkbookParser.pivot_caches = original     # 无论如何都还原，避免污染全局


def _load_workbook_safe(path, **kwargs):
    """openpyxl.load_workbook 的容错封装:把"文件损坏/伪装成 xlsx"这类底层异常
    (BadZipFile / InvalidFileException)转成一句人话的 ValueError,而非让上层裸崩
    出 'File is not a zip file'。用户误选了损坏文件或把 .xls 改名成 .xlsx 时很常见。"""
    import zipfile
    try:
        return openpyxl.load_workbook(path, **kwargs)
    except (zipfile.BadZipFile, OSError) as e:
        raise ValueError("无法打开 %s —— 文件可能已损坏,或不是真正的 xlsx"
                         "(例如把 .xls 直接改名为 .xlsx)。请在 Excel 里另存为 .xlsx 后重试。"
                         % os.path.basename(path)) from e
    except Exception as e:
        # openpyxl 的 InvalidFileException 及其它读取期异常,同样给清晰提示
        if e.__class__.__name__ == "InvalidFileException":
            raise ValueError("无法打开 %s —— 不是受支持的 Excel 格式,请另存为 .xlsx 后重试。"
                             % os.path.basename(path)) from e
        raise


def load_data_only(path, **kwargs):
    """openpyxl.load_workbook(path, data_only=True) 的加速封装。

    等价于常规 data_only 读取，但跳过输入文件内嵌透视缓存的解析(见
    _skip_pivot_cache_parse)。仅用于"只读单元格值、不碰透视对象"的场景。
    额外 kwargs 透传给 load_workbook。"""
    with _skip_pivot_cache_parse():
        return _load_workbook_safe(path, data_only=True, **kwargs)


def load_data_only_stream(path):
    """以只读流式模式打开工作簿，并修复错误的 ``dimension`` 声明。

    部分业务导出文件会把实际数千行的工作表声明成 ``A1:A1``。openpyxl 的
    只读模式会信任该声明而静默漏行，因此每张表必须先 ``reset_dimensions``。
    返回的工作簿由调用方负责 ``close()``；适合只顺序读取单元格值、不访问样式、
    合并区域或透视对象的大文件路径。
    """
    with _skip_pivot_cache_parse():
        wb = _load_workbook_safe(path, data_only=True, read_only=True)
    try:
        for ws in wb.worksheets:
            reset = getattr(ws, "reset_dimensions", None)
            if reset is not None:
                reset()
    except Exception:
        wb.close()
        raise
    return wb


# ---------------- 统一多格式读取 ----------------
def read_sheets(path):
    """读取任意工作簿为 [(sheet_name, rows)]，rows 为二维列表。
    支持 .xlsx / .xlsm / .xls（.xls 需 xlrd==1.2.0）。两功能共用。"""
    ext = os.path.splitext(path)[1].lower()
    result = []
    if ext in (".xlsx", ".xlsm"):
        # 注意:不能用 read_only=True。实测部分导出文件的 <dimension> 标签错误
        # (标为单格 A1:A1),read_only 会信任它、iter_rows 只吐 1 行导致整表漏读;
        # 且 read_only 不把行补齐到最大列宽(尾部空格被截),破坏表头/列定位。
        # 常规模式会重算真实用量,稳妥优先。
        # 但仍跳过内嵌透视缓存解析(只读单元格值,源表若自带透视表可省大量耗时)。
        with _skip_pivot_cache_parse():
            wb = _load_workbook_safe(path, data_only=True)
        try:
            for ws in wb.worksheets:
                result.append((ws.title,
                               [list(row) for row in ws.iter_rows(values_only=True)]))
        finally:
            wb.close()
    elif ext == ".xls":
        import xlrd  # 打包时通过 xlrd==1.2.0 支持 .xls
        book = xlrd.open_workbook(path)
        for sh in book.sheets():
            rows = []
            for r in range(sh.nrows):
                row = []
                for c in range(sh.ncols):
                    cell = sh.cell(r, c)
                    val = cell.value
                    if cell.ctype == 3:        # xlrd 类型3=日期
                        try:
                            y, mo, d, h, mi, s = xlrd.xldate_as_tuple(val, book.datemode)
                            val = datetime.datetime(y, mo, d, h, mi, s)
                        except Exception:
                            pass
                    row.append(val)
                rows.append(row)
            result.append((sh.name, rows))
    else:
        raise ValueError("不支持的文件类型：%s" % ext)
    return result


# ---------------- 统一输出路径 ----------------
def make_out_dir(src_path):
    """在源文件所在目录建 output 子文件夹并返回其路径（两功能统一）。"""
    base_dir = os.path.dirname(os.path.abspath(src_path))
    out_dir = os.path.join(base_dir, "output")
    if not os.path.isdir(out_dir):
        os.makedirs(out_dir)
    return out_dir


def timestamp():
    """当前时间戳字符串 YYYYMMDD_HHMM，用于输出文件名。"""
    return datetime.datetime.now().strftime("%Y%m%d_%H%M")


def unique_path(path):
    """若 path 已存在,自动在扩展名前加 ' (2)'/' (3)'… 直到不冲突,返回可安全写入的路径。
    时间戳只精确到分钟,同一分钟内重复生成会同名覆盖(静默丢结果); 用它兜底防覆盖。"""
    if not os.path.exists(path):
        return path
    root, ext = os.path.splitext(path)
    i = 2
    while True:
        cand = "%s (%d)%s" % (root, i, ext)
        if not os.path.exists(cand):
            return cand
        i += 1


def out_path(out_dir, base_name, suffix, ext=".xlsx", ts=None):
    """拼输出路径：out_dir/base_name_suffix_时间戳.ext（两功能统一命名）。
    base_name 传源文件名(不含扩展名)；ts 不传则自动取当前时间。
    已存在则自动避让,不覆盖同分钟的上一次结果。"""
    if ts is None:
        ts = timestamp()
    fname = "%s%s_%s%s" % (base_name, suffix, ts, ext)
    return unique_path(os.path.join(out_dir, fname))


# ---------------- 供“列映射”界面读取表头预览 ----------------
def sheet_names(path):
    """只取工作表名列表（供界面下拉）。"""
    return [name for name, _ in read_sheets(path)]


def preview_rows(path, sheet=None, limit=8):
    """取某工作表前 limit 行（含表头，二维列表）。sheet=None 取第一个。
    返回 (sheet_name, rows)。供列映射界面显示表头与样例数据。"""
    sheets = read_sheets(path)
    if not sheets:
        return None, []
    chosen = None
    if sheet:
        for name, rows in sheets:
            if name == sheet:
                chosen = (name, rows); break
    if chosen is None:
        chosen = sheets[0]
    name, rows = chosen
    return name, [list(r) for r in rows[:limit]]


def apply_saved_mapping(opts, path, mapping):
    """把字段映射中心记录合并进 Options.columns，供本次任务优先使用。"""
    if opts is None or not path or not mapping:
        return False
    base = os.path.basename(path)
    fm = dict(opts.columns.get(base) or {})
    if mapping.get("sheet") and not fm.get("sheet"):
        fm["sheet"] = mapping["sheet"]
    if mapping.get("header") and not fm.get("header"):
        fm["header"] = int(mapping["header"])
    roles = mapping.get("roles") or {}
    if roles:
        merged = dict(fm.get("roles") or {})
        for key, value in roles.items():
            merged.setdefault(str(key), int(value))
        fm["roles"] = merged
    if fm:
        opts.columns[base] = fm
        return True
    return False


def auto_apply_mapping(opts, path, role_kind):
    """按当前文件结构尝试复用字段映射，返回匹配记录或 None。"""
    if not opts or not path:
        return None
    from . import mapping_store
    sheet, rows = preview_rows(path, sheet=opts.resolve_sheet(path), limit=20)
    mapping = mapping_store.find_for_rows(sheet, rows, role_kind)
    if mapping and apply_saved_mapping(opts, path, mapping):
        return mapping
    return None


def cell_text(v):
    """单元格值转成简短显示文本（供界面）。"""
    if v is None:
        return ""
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d %H:%M").replace(" 00:00", "")
    if isinstance(v, datetime.date):
        return v.strftime("%Y-%m-%d")
    s = str(v)
    return s if len(s) <= 18 else s[:17] + "…"


# ---------------- 进度上报 ----------------
class Progress:
    """把"阶段 + 阶段内 i/n"折算成 0~100 的整数进度，避免各 core 散落魔法数字。

    用法：
        p = Progress(progress, stages=[("读取", 30), ("填表", 50), ("汇总", 20)])
        p.stage("读取");  ... ;  p.tick(i, n)   # 在本阶段区间内按 i/n 插值
        p.stage("填表");  ... ;  p.done()        # 收尾时补到 100

    · progress 为 None 时全程 no-op（不上报进度的旧调用与测试不受影响）。
    · stages 的权重之和不必等于 100，内部按比例归一。
    · 进度只增不减：回调只在百分比真正变大时触发，避免抖动与无谓刷新。
    """

    def __init__(self, cb, stages):
        self._cb = cb if callable(cb) else None
        total_w = sum(max(0, w) for _, w in stages) or 1
        # 预计算每个阶段的 [起点, 跨度]（都是 0~100 的浮点）
        self._span = {}
        acc = 0.0
        for name, w in stages:
            frac = max(0, w) / total_w * 100.0
            self._span[name] = (acc, frac)
            acc += frac
        self._base = 0.0        # 当前阶段起点
        self._range = 0.0       # 当前阶段跨度
        self._last = -1         # 上次已发出的整数百分比

    def stage(self, name):
        """进入某阶段：立即把进度推到该阶段起点。"""
        base, rng = self._span.get(name, (self._base, 0.0))
        self._base, self._range = base, rng
        self._emit(base)

    def tick(self, i, n):
        """在当前阶段内按 i/n 插值上报（i 从 0 到 n）。n<=0 时忽略。"""
        if n and n > 0:
            frac = min(max(i, 0), n) / float(n)
            self._emit(self._base + self._range * frac)

    def done(self):
        """收尾：补到 100。"""
        self._emit(100.0)

    def _emit(self, pct):
        if self._cb is None:
            return
        v = int(pct)
        if v > self._last:      # 只增不减，减少无谓刷新
            self._last = v
            try:
                self._cb(v)
            except Exception:
                pass            # 进度回调异常绝不能影响主流程
