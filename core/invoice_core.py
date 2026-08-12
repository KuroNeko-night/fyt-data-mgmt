# -*- coding: utf-8 -*-
"""增值税发票筛选、人工复核与统计输出核心。

递归扫描资料文件夹里的 PDF，识别“增值税专用发票”，按开票月份筛选，
抽取字段汇总。可靠字段全自动；费用项目/备注给“PDF 原始种子”供人工精修。
另把所有专用发票原始 PDF 复制到复核文件夹（宽松判定，含存疑清单），供人工二次核对。
PDF 自动抽取只提供候选数据，存疑文件、费用项目和备注仍保留人工复核入口；任何扫描件、
字段残缺或金额关系异常都不能静默丢弃。桌面端和 Web 结果展示复用本模块输出，不自行
编写第二套发票号码、税率或金额解析逻辑。

对外主接口：
  scan(root, log=None)                 -> ScanResult      （扫描+抽取，不筛月份）
  filter_month(items, ym)              -> list[Invoice]
  detect_month(items)                  -> "YYYY-MM"
  generate(result, rows, ym, out_dir=None, log=None) -> dict
        （统一出口：写汇总表 + 导出复核文件夹，输出目录经 paths 统一解析）
"""
import os
import re
import glob

from . import common_core as _common  # 公共唯一路径、分阶段进度与输出辅助。

try:
    from pypdf import PdfReader
except Exception:
    PdfReader = None

BUYER = "重庆峰运通供应链管理有限公司"

_COMPANY = re.compile(
    r"[一-龥A-Za-z0-9（）()]{2,45}"
    r"(?:有限责任公司|有限公司|股份公司|分公司|公司|银行|事务所|合作社|个体工商户|中心)")
_DATE = re.compile(r"(\d{4})年(\d{1,2})月(\d{1,2})日")  # 月日允许一位，提取后统一补零。
_NUM = re.compile(r"发票号码[:：]?(\d{20}|\d{8})")  # 锚定全电 20 位或旧版 8 位票号。
_NUM_LOOSE = re.compile(r"\d{20}|\d{8}")  # 无锚点时的宽松兜底。
_MONEY = re.compile(r"¥\s*([\d\s]+\.\s*\d\s*\d)")
_RATE = re.compile(r"(\d+)%")
_SKIP_LINE = ("开户", "账号", "地址", "电话")


class Invoice(object):
    """一张发票的结构化抽取结果，字段均为可跨进程和 JSON 序列化的基础类型。"""
    __slots__ = ("path", "num", "date", "seller", "amount", "tax", "total",
                 "rate", "item_seed", "note_seed", "special")

    def __init__(self, **kw):
        """按固定槽位接收字段；缺失键保存为空值，避免不同版式产生动态属性。"""
        for k in self.__slots__:
            setattr(self, k, kw.get(k))

    def as_row(self):
        """转换为人工复核和 Excel 写入字典，费用项目与备注先采用自动种子。"""
        return dict(num=self.num, date=self.date, seller=self.seller,
                    item=self.item_seed or "", amount=self.amount,
                    tax=self.tax, total=self.total, rate=self.rate,
                    note=self.note_seed or "")


def _norm(t):
    """移除普通空格和制表符，生成适合跨 PDF 断词匹配的紧凑文本。"""
    return re.sub(r"[ \t]+", "", t or "")


def _seller(raw):
    """
    从逐行 PDF 文本中寻找第一个非买方公司名称，作为销售方候选。

    开户行、账号、地址和电话行可能包含银行或公司文字，先排除；公司正则命中后再去掉
    粘连在名称前的税号字符，并排除固定买方“峰运通”，降低把购买方当销售方的风险。
    """
    for ln in raw.splitlines():
        if any(k in ln for k in _SKIP_LINE):
            continue
        for m in _COMPANY.finditer(ln):
            c = re.sub(r"^[0-9A-Za-z]+", "", m.group(0))  # 去除粘连在公司名前的税号前缀。
            if not c or "峰运通" in c or c == BUYER:
                continue
            return c
    return ""


def _money3(raw):
    """
    提取全文最后三个人民币金额，依次解释为不含税金额、税额和价税合计。

    PDF 文本层常把金额数字拆入空格，正则允许并在转换前移除。发票明细区可能有多笔单价
    和金额，合计区通常位于末尾，因此采用最后三个。少于三个时返回空值并交给存疑流程。
    """
    vals = [float(re.sub(r"\s+", "", m.group(1))) for m in _MONEY.finditer(raw)]
    if len(vals) >= 3:
        return vals[-3], vals[-2], vals[-1]
    return None, None, None


# 合法标准税率按两位和一位拆分，解析粘连数字时必须优先尝试更长后缀。
_RATE2 = ("17", "16", "13", "11", "10")
_RATE1 = ("9", "6", "5", "3", "1", "0")


def _one_rate(numstr):
    """从可能粘连型号数字的百分数字符串末尾恢复合法增值税税率。"""
    for v in _RATE2:
        if numstr.endswith(v):
            return v
    for v in _RATE1:
        if numstr.endswith(v):
            return v
    return None


def _rate(nn):
    """
    提取正文中的合法税率：单税率返回小数，多税率返回降序拼接文本。

    单税率用数值便于 Excel 百分比格式和后续计算；多税率无法用一个数表达，保留
    ``9%+6%`` 等文本。重复税率按首次发现去重，最终多税率按数值从大到小稳定输出。
    """
    found = []
    for m in _RATE.finditer(nn):
        v = _one_rate(m.group(1))
        if v is not None and v not in found:
            found.append(v)
    if not found:
        return ""
    if len(found) == 1:
        return round(int(found[0]) / 100.0, 2)
    return "+".join(v + "%" for v in sorted(found, key=lambda x: -int(x)))


_STD_RATES = (0.17, 0.16, 0.13, 0.11, 0.10, 0.09, 0.06, 0.05, 0.03, 0.01)


def _deduction(nn):
    """提取差额征税备注中的扣除额；未找到时返回空值。"""
    m = re.search(r"扣除额[:：]([\d.]+)元", nn)
    return float(m.group(1)) if m else None


def _snap(r):
    """把反推税率吸附到最接近的标准档位，偏差超过 0.006 时拒绝猜测。"""
    if r is None or r <= 0:
        return None
    best = min(_STD_RATES, key=lambda s: abs(s - r))
    return best if abs(best - r) <= 0.006 else None


def _derive_rate(amount, tax, total, ded):
    """
    在正文未印税率时按金额关系反推，并只返回足够接近标准档位的结果。

    差额征税先以“价税合计 - 扣除额”为含税计税基础，再把税额占含税额比例换算为
    不含税税率；普通情况直接使用税额除以不含税金额。无法形成正数基础时返回空值。
    """
    if not tax:
        return None
    if ded and total:
        base = total - ded  # 差额征税下扣除后的含税计税基础。
        if base > 0:
            ratio = tax / base
            if ratio < 1:
                return _snap(ratio / (1 - ratio))
    if amount:
        return _snap(tax / amount)
    return None


def _item_seed(raw):
    """
    从星号分隔的发票项目行提取“税收分类/品名”中文片段，作为人工复核初始值。

    全电发票文本常形如“开票人*生产生活服务*设备租赁费型号”，先移除 2～4 位纯中文
    开票人段，再保留前两个业务段并截断后续型号字母数字。该结果仅是种子，允许人工修改。
    """
    for ln in raw.splitlines():
        s = ln.strip()
        if s.count("*") >= 2:
            # 星号分隔通常比 PDF 文本列位置稳定，优先利用其还原税收分类与品名。
            parts = [p for p in s.split("*") if p.strip()]
            # 首段若是常见长度的纯中文姓名则视为开票人，不写入费用项目。
            if parts and re.match(r"^[一-龥]{2,4}$", parts[0]):
                parts = parts[1:]
            seg = "/".join(p.strip() for p in parts[:2] if p.strip())
            seg = re.sub(r"[0-9A-Za-z]+.*$", "", seg).strip("/ ")
            if seg:
                return seg
    return ""


def _note_seed(raw):
    """从备注区提取扣除额和库区等常见信息，拼成供人工精修的初始备注。"""
    n = _norm(raw)
    bits = []
    m = re.search(r"扣除额[:：]([\d.]+)元", n)
    if m:
        bits.append("扣除额 %s 元" % m.group(1))
    m = re.search(r"库区[:：]?([A-Za-z0-9\-]+)", n)
    if m:
        bits.append("库区 " + m.group(1))
    return "；".join(bits)


def _find_num(nn):
    """
    优先从“发票号码”锚点提取全电 20 位或旧版 8 位票号，失败后全文宽松查找。

    宽松查找优先 20 位，避免其中某段 8 位数字被旧版规则提前截走；无号码的已识别发票
    不参与去重和统计，而是进入存疑清单等待人工确认。
    """
    m = _NUM.search(nn)
    if m:
        return m.group(1)
    m20 = re.search(r"\d{20}", nn)  # 宽松兜底仍优先完整全电票号。
    if m20:
        return m20.group(0)
    m8 = _NUM_LOOSE.search(nn)  # 最后再接受旧版 8 位号码。
    return m8.group(0) if m8 else ""


def _extract_one(raw, path):
    """
    从单页 PDF 文本提取一张增值税专用或普通发票；无法确认类型时返回 ``None``。

    日期统一为 ``YYYY-MM-DD``，金额取合计区，税率优先读取正文，缺失时再按金额关系
    反推。字段不完整仍返回 ``Invoice``，由扫描阶段按专票金额完整性和勾稽关系标记存疑，
    避免自动抽取层直接丢掉可能有效的发票。
    """
    nn = _norm(raw)
    # “专用发票”优先判断，防止同时出现普通字样的说明文本把专票错误降级。
    special = "专用发票" in nn
    normal = ("普通发票" in nn) and not special
    if not (special or normal):
        return None
    md = _DATE.search(nn)
    # 月日补零后，按 YYYY-MM 前缀筛选不会漏掉原文“6月1日”等一位写法。
    date = "%s-%02d-%02d" % (md.group(1), int(md.group(2)), int(md.group(3))) if md else ""
    amount, tax, total = _money3(raw)
    rate = _rate(nn)
    if rate == "":
        # 只有正文完全未提取到税率时才反推，不能覆盖发票上明确印出的多税率文本。
        d = _derive_rate(amount, tax, total, _deduction(nn))
        if d is not None:
            rate = round(d, 2)
    return Invoice(
        path=path, num=_find_num(nn), date=date,
        seller=_seller(raw), amount=amount, tax=tax, total=total,
        rate=rate, item_seed=_item_seed(raw),
        note_seed=_note_seed(raw), special=special)


def extract(path):
    """逐页解析单个 PDF，返回发票列表和页级存疑原因列表。

    多页 PDF(含多张发票)逐页识别,避免只抽第一页而漏计;
    多页 PDF 可能每页各是一张发票，不能只读取首页。无文本层扫描件、单页解析异常和
    “具备多个发票信号但类型未确认”的页面都回传原因，供 ``scan`` 写入存疑清单。
    """
    if PdfReader is None:
        raise RuntimeError("缺少 pypdf 依赖")
    # 整份 PDF 损坏或加密时让异常上抛，由目录扫描统一记录文件级失败。
    reader = PdfReader(path)
    invoices = []
    notes = []
    pages = reader.pages
    multi = len(pages) > 1
    for pno, page in enumerate(pages, start=1):
        try:
            raw = page.extract_text() or ""
        except Exception as e:
            # 单页异常不阻断同一 PDF 的其他页面，但必须留下页码和原因。
            notes.append("第%d页文本读取异常:%s" % (pno, e))
            continue
        if not raw.strip():
            # 当前功能不做 OCR；无文本层页面只能进入人工复核，不能假定不是发票。
            notes.append("第%d页无文本层(疑似扫描件),请人工核对" % pno)
            continue
        inv = _extract_one(raw, path)
        if inv is not None:
            invoices.append(inv)
        elif _looks_like_invoice(raw):
            # 宽松信号只用于防漏提示，不直接把未确认页面计入发票统计。
            notes.append("第%d页疑似发票但未能确认类型" % pno)
    return invoices, notes


class ScanResult(object):
    """一次扫描的结果：识别到的发票 + 存疑文件清单。

    invoices : 去重后的 Invoice（含专用/普通，供界面按 special 过滤）
    suspects : list[(path, reason)] —— 解析失败、或疑似发票但字段残缺的文件，
               供人工二次核对，避免漏掉一张专用发票。
    """
    __slots__ = ("invoices", "suspects")

    def __init__(self, invoices, suspects):
        """保存去重发票和 ``(文件路径, 原因)`` 存疑记录，不额外复制列表。"""
        self.invoices = invoices
        self.suspects = suspects


# 宽松信号只决定是否列入存疑，不影响专用/普通发票的正式类型判定。
_LOOSE_HINT = ("电子发票", "增值税", "发票号码", "价税合计", "税率")


def _looks_like_invoice(raw):
    """正文命中至少两个发票特征时判为疑似发票，供防漏复核使用。"""
    nn = _norm(raw)
    return sum(1 for k in _LOOSE_HINT if k in nn) >= 2


def scan(root, log=None, progress=None):
    """递归扫描目录内全部 PDF，按票号去重并返回日期排序结果和存疑清单。

    去重以发票号码为唯一键，先扫描到的版本保留；缺号码发票只进入存疑。专票金额字段
    残缺或“金额 + 税额”与价税合计偏差超过 0.01 时仍保留统计候选，同时要求人工核对。
    ``progress`` 按已完成 PDF 数量线性上报 0～100。
    """
    def _lg(m):
        """统一转发目录扫描日志；没有回调时保持静默。"""
        if log:
            log(m)
    prog = _common.Progress(progress, stages=[("scan", 100)])
    prog.stage("scan")
    # 路径排序使重复票号“先者保留”规则在同一目录结构下稳定可复现。
    pdfs = sorted(glob.glob(os.path.join(root, "**", "*.pdf"), recursive=True))
    _lg("发现 %d 个 PDF，开始识别…" % len(pdfs))
    n_pdf = len(pdfs)
    by_num = {}
    suspects = []
    n_err = n_dup = 0
    for _pi, p in enumerate(pdfs):
        try:
            invs, notes = extract(p)
        except Exception as e:
            n_err += 1
            _lg("  跳过（解析失败）：%s —— %s" % (os.path.basename(p), e))
            suspects.append((p, "解析失败：%s" % e))
            continue
        # 页级异常和疑似漏网同时记录日志与结构化清单，便于最终复核文件夹追踪。
        for note in notes:
            _lg("  存疑：%s —— %s" % (os.path.basename(p), note))
            suspects.append((p, note))
        for inv in invs:
            if not inv.num:
                suspects.append((p, "识别为发票但缺发票号码"))
                continue
            if inv.special and (inv.amount is None or inv.total is None):
                suspects.append((p, "专用发票但金额字段残缺"))
            # 金额勾稽异常可能是抽取到了明细金额；保留候选但明确要求人工核对。
            elif (inv.amount is not None and inv.tax is not None
                    and inv.total is not None
                    and abs(inv.amount + inv.tax - inv.total) > 0.01):
                suspects.append((p, "金额+税额与价税合计不符,请核对是否取错金额"))
            if inv.num in by_num:
                n_dup += 1
                _lg("  重复发票号 %s，已忽略：%s" % (inv.num, os.path.basename(p)))
                continue
            by_num[inv.num] = inv
        prog.tick(_pi + 1, n_pdf)  # 每处理完一份 PDF 推进一格，与页数无关。
    # 日期为空的存量候选排在前部，随后按日期和票号稳定排序供复核界面展示。
    items = sorted(by_num.values(), key=lambda i: (i.date or "", i.num))
    n_spec = sum(1 for i in items if i.special)
    _lg("识别发票 %d 张（专用 %d ·普通 %d）；去重 %d，失败 %d，存疑 %d。"
        % (len(items), n_spec, len(items) - n_spec, n_dup, n_err, len(suspects)))
    prog.done()
    return ScanResult(items, suspects)


def filter_month(items, ym):
    """按 ``YYYY-MM`` 日期前缀筛选发票；月份为空时返回输入的浅列表副本。"""
    if not ym:
        return list(items)
    return [i for i in items if (i.date or "").startswith(ym)]


def export_review_folder(result, out_dir, log=None):
    """把全部专用发票原始 PDF 按月份复制到复核目录，并写入存疑清单。

    复核目录不只包含目标统计月份，而是包含扫描结果中的所有专票，使用户能检查目标月
    是否漏收、其他月份是否正确排除。未知日期进入“未知月份”；同名 PDF 递增编号保留，
    不覆盖不同来源文件。单个复制失败记录日志但不阻断其他发票和存疑清单生成。
    """
    import shutil

    def _lg(m):
        """转发复核文件复制日志；未提供回调时保持静默。"""
        if log:
            log(m)
    review = os.path.join(out_dir, "专用发票复核")
    specials = [i for i in result.invoices if i.special]
    n = 0
    for inv in specials:
        # 只取年月建立目录；无日期时保留明确的“未知月份”人工处理入口。
        ym = (inv.date or "未知月份")[:7] or "未知月份"
        sub = os.path.join(review, ym.replace("-", "年") + "月" if "-" in ym else ym)
        if not os.path.isdir(sub):
            os.makedirs(sub)
        dst = _unique_path(os.path.join(sub, os.path.basename(inv.path)))
        try:
            shutil.copy2(inv.path, dst)
            n += 1
        except Exception as e:
            # 单个源文件可能被删除或占用，不影响其余复核材料继续导出。
            _lg("  复制失败：%s —— %s" % (os.path.basename(inv.path), e))
    _write_suspects(review, result.suspects, specials)
    _lg("已导出专用发票 %d 张到复核文件夹，存疑 %d 个。"
        % (n, len(result.suspects)))
    return review


def _unique_path(path):
    """为复核 PDF 生成未占用路径，同名时在扩展名前递增追加序号。"""
    if not os.path.exists(path):
        return path
    base, ext = os.path.splitext(path)
    i = 2
    while os.path.exists("%s (%d)%s" % (base, i, ext)):
        i += 1
    return "%s (%d)%s" % (base, i, ext)


def _write_suspects(review, suspects, specials):
    """
    写入 UTF-8 存疑清单，列出文件名、原因和原始绝对路径，并附已导出专票数量。

    绝对路径只存在于本机复核文本，不进入客户界面；它帮助管理员直接找到未复制或
    未统计的原始文件。没有存疑项时仍写明确结论，避免用户误以为清单生成失败。
    """
    if not os.path.isdir(review):
        os.makedirs(review)
    path = os.path.join(review, "存疑清单.txt")
    lines = ["增值税专用发票复核 —— 存疑清单",
             "（下列文件程序未纳入统计，请人工确认是否有被漏掉的专用发票）", ""]
    if suspects:
        for p, reason in suspects:
            lines.append("· %s\n    原因：%s\n    路径：%s" %
                         (os.path.basename(p), reason, p))
    else:
        lines.append("（无存疑文件）")
    lines += ["", "本次已导出专用发票 %d 张。" % len(specials)]
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def detect_month(items):
    """统计有效开票日期，返回出现次数最多的 ``YYYY-MM`` 主月份；无日期返回空串。"""
    from collections import Counter
    c = Counter((i.date or "")[:7] for i in items if i.date)
    # Counter 在次数相同时保持首次出现顺序，与扫描结果的稳定日期排序共同保证可复现。
    return c.most_common(1)[0][0] if c else ""


# 汇总表列顺序和宽度属于交付模板协议，人工复核字典字段必须映射到这些固定列。
HEADERS = ["序号", "发票号码", "开票日期", "销售方名称", "费用项目",
           "不含税金额（元）", "税额（元）", "价税合计（元）",
           "税率/征收方式", "备注"]
_WIDTHS = [5.1, 22.6, 10.4, 35.5, 23.9, 17.1, 10.9, 15.0, 16.2, 24.6]


def write_xlsx(rows, out_path, ym=""):
    """
    把人工确认后的发票字典写成固定格式汇总工作簿，并追加金额合计行。

    日期转换为真正的 Excel 日期以支持筛选，单一数值税率使用百分比格式；多税率文本
    保持原样。函数创建父目录但不负责生成唯一文件名，调用方应在写入前完成防覆盖处理。
    """
    import datetime
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    # 目标月份转为无横杠页签名；未指定时使用稳定中文名称。
    ws.title = (ym or "").replace("-", "") or "增值税发票"
    thin = Side(style="thin")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    song = "宋体"

    # 标题跨越全部十列，第二行固定表头，数据从第三行开始。
    ws.merge_cells("A1:J1")
    t = ws.cell(1, 1, "增值税发票")
    t.font = Font(name=song, bold=True, size=16)
    t.alignment = center
    ws.row_dimensions[1].height = 20.25
    ws.row_dimensions[2].height = 16.5
    for c, (h, w) in enumerate(zip(HEADERS, _WIDTHS), start=1):
        cell = ws.cell(2, c, h)
        cell.font = Font(name=song, size=11)
        cell.alignment = center
        cell.border = box
        ws.column_dimensions[chr(64 + c)].width = w

    r = 3
    for i, row in enumerate(rows, start=1):
        # 顺序必须与 HEADERS 一一对应；费用项目和备注已经包含人工复核后的最终值。
        vals = [i, row.get("num", ""), _as_date(row.get("date"), datetime),
                row.get("seller", ""), row.get("item", ""), row.get("amount"),
                row.get("tax"), row.get("total"), row.get("rate", ""),
                row.get("note", "")]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(r, c, v)
            cell.font = Font(name=song, size=11)
            cell.alignment = center
            cell.border = box
            if c == 3 and isinstance(v, datetime.datetime):
                # 真日期使用 Excel 日期格式；解析失败的原文本仍按普通文本保留。
                cell.number_format = "yyyy-mm-dd"
            elif c == 9 and isinstance(v, float):
                # 单税率以小数存储并显示为百分比，多税率字符串不套数值格式。
                cell.number_format = "0%"
        r += 1

    _write_total(ws, r, rows, box, center, song, Font)
    out_dir = os.path.dirname(out_path)
    if out_dir and not os.path.isdir(out_dir):
        os.makedirs(out_dir)
    wb.save(out_path)
    return out_path


def _as_date(s, datetime):
    """把规范 ``YYYY-MM-DD`` 文本转换为 Excel 可识别的 datetime，其他值原样返回。"""
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", str(s or ""))
    if m:
        return datetime.datetime(*[int(x) for x in m.groups()])
    return s or ""


def _write_total(ws, r, rows, box, center, song, Font):
    """写入汇总尾行：前五列合并，三列金额求和，税率与备注列显示短横线。"""
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c = ws.cell(r, 1, "合计")
    c.font = Font(name=song, bold=True, size=11)
    c.alignment = center
    for col, key in ((6, "amount"), (7, "tax"), (8, "total")):
        # 空金额按零参与合计，但对应发票已在扫描阶段进入存疑，汇总仍可正常生成。
        s = sum(x.get(key) or 0 for x in rows)
        cell = ws.cell(r, col, round(s, 2))
        cell.font = Font(name=song, bold=True, size=11)
        cell.alignment = center
    for col in (9, 10):
        ws.cell(r, col, "-").alignment = center
    for col in range(1, 11):
        ws.cell(r, col).border = box


# ---------------------------------------------------------------------------
# 统一出口（与其余功能一致：generate(...) -> dict，输出目录经 paths 统一解析）
# ---------------------------------------------------------------------------
def generate(result, rows, ym, out_dir=None, log=None, progress=None):
    """生成最终汇总表和专用发票复核目录，并返回路径及数量统计。

    result : scan() 的 ScanResult（用于导出复核文件夹与存疑清单）
    rows   : 复核对话框最终确认的行 dict 列表（已含人工精修的费用项目/备注）
    ym     : 目标月份 'YYYY-MM'，决定 sheet 名与汇总表文件名
    out_dir: 不传则经 settings + paths 统一解析到 <文档>/…/输出/增值税发票统计/<时间戳>/
    ``rows`` 必须是人工复核后的最终字典列表，``result`` 则保留扫描所得原始 PDF 路径和
    存疑信息，两者职责不同。输出目录未指定时走统一设置；汇总文件使用唯一名称防覆盖，
    复核目录按月份继续累积同名避让文件。进度分为汇总表和 PDF 复核两个等权阶段。
    """
    def _lg(msg):
        """统一转发最终生成阶段日志。"""
        if log:
            log(msg)
    # 汇总表与复核文件夹各占一半，后者包含多文件复制但内部当前没有更细粒度回调。
    prog = _common.Progress(progress, stages=[("xlsx", 50), ("review", 50)])
    if out_dir is None:
        # 正式入口始终遵循 settings/paths 输出策略，不在发票模块维护第二套目录规则。
        from . import paths as _paths, settings as _settings
        st = _settings.get_settings()
        out_dir = _paths.resolve_output_dir("invoice", **st.output_kwargs())
    elif not os.path.isdir(out_dir):
        os.makedirs(out_dir)
    _lg("输出文件夹：%s" % out_dir)

    prog.stage("xlsx")
    # 文件名只使用月份中文片段；同月重复运行通过 unique_path 保留之前人工确认的结果。
    mm = "%d月" % int(ym[5:7]) if ym and len(ym) >= 7 else ""
    xlsx = _common.unique_path(os.path.join(out_dir, "%s统计增值税发票.xlsx" % mm))
    write_xlsx(rows, xlsx, ym)
    _lg("已生成汇总表：%s（%d 张专用发票）" % (os.path.basename(xlsx), len(rows)))

    prog.stage("review")
    review = export_review_folder(result, out_dir, log=log)
    prog.done()
    return {"xlsx": xlsx, "review_dir": review, "out_dir": out_dir,
            "count": len(rows), "suspects": len(result.suspects)}
