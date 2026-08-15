# -*- coding: utf-8 -*-
"""跨业务模块共用的 Excel 安全读取与公式缓存检查。"""

from __future__ import annotations

import contextlib
import datetime
import os
import threading
import zipfile
from typing import Iterator

import openpyxl


_PIVOT_CACHE_LOCK = threading.RLock()


def _formula_sheets(workbook, requested: str | None) -> list[str]:
    """解析公式检测的页签范围；不存在的指定页签兼容回退为全部页签。"""

    if requested and requested in workbook.sheetnames:
        return [requested]
    return list(workbook.sheetnames)


def _uncached_columns(formula_sheet, value_sheet, sample_rows: int) -> set[int]:
    """比较同一页签的公式视图和缓存值视图，返回缺缓存的 0 基列号。"""

    columns: set[int] = set()
    formula_rows = formula_sheet.iter_rows(max_row=sample_rows, values_only=True)
    value_rows = value_sheet.iter_rows(max_row=sample_rows, values_only=True)
    for formula_row, value_row in zip(formula_rows, value_rows):
        for column, formula_value in enumerate(formula_row):
            if not isinstance(formula_value, str) or not formula_value.startswith("="):
                continue
            if column >= len(value_row) or value_row[column] is None:
                columns.add(column)
    return columns


def detect_uncached_formula(
    path: str,
    sheet: str | None = None,
    sample_rows: int = 400,
) -> set[int]:
    """检测公式文本存在但 Excel 未保存计算缓存的列，检测失败返回空集。"""

    if os.path.splitext(path)[1].lower() not in (".xlsx", ".xlsm"):
        return set()
    formula_book = value_book = None
    try:
        # 公式缓存检查只读取工作表单元格。部分业务工作簿会携带数百个外部链接，
        # openpyxl 默认恢复这些链接时会解析大量 externalLink XML，既耗时又占内存；
        # 它们不参与公式文本与缓存值的比较，因此必须显式跳过。
        formula_book = openpyxl.load_workbook(
            path, data_only=False, read_only=True, keep_links=False,
        )
        value_book = openpyxl.load_workbook(
            path, data_only=True, read_only=True, keep_links=False,
        )
        columns: set[int] = set()
        for sheet_name in _formula_sheets(formula_book, sheet):
            columns.update(
                _uncached_columns(formula_book[sheet_name], value_book[sheet_name], sample_rows)
            )
        return columns
    except Exception:
        # 这是预警工具：检测失败不能阻断主业务，也不能在证据不足时制造公式告警。
        return set()
    finally:
        if formula_book is not None:
            formula_book.close()
        if value_book is not None:
            value_book.close()


def warn_if_uncached(path: str, log, sheet: str | None = None, what: str = "数据") -> bool:
    """发现公式缓存风险时记录客户可执行的修复步骤。

    ``log`` 允许为 ``None``（各业务入口默认不传回调）；此时只返回是否命中，不再写日志，
    避免把“公式未刷新”的预警升级成 ``TypeError`` 中断整个业务任务。
    """
    if not detect_uncached_formula(path, sheet):
        return False
    if log is not None:
        log(
            "⚠ 警告：《%s》中%s所在列含未刷新的公式（读取值为空），可能导致漏算或算错。"
            % (os.path.basename(path), what)
        )
        log("  请先用 Excel 打开该表、按 Ctrl+S 保存一次以刷新公式后重试。")
    return True


@contextlib.contextmanager
def skip_pivot_cache_parse() -> Iterator[None]:
    """临时阻止 openpyxl 反序列化输入文件中业务不需要的透视缓存。"""

    # WorkbookParser.pivot_caches 是进程级类属性，必须串行替换并保证 finally 恢复。
    with _PIVOT_CACHE_LOCK:
        try:
            from openpyxl.reader.workbook import WorkbookParser
        except Exception:
            yield
            return
        current = getattr(WorkbookParser, "pivot_caches", None)
        if not isinstance(current, property):
            yield
            return

        class _NullCaches(dict):
            """满足读取器索引协议，但不触发任何真实缓存解析。"""

            def __missing__(self, key):
                return None

        WorkbookParser.pivot_caches = property(lambda self: _NullCaches())
        try:
            yield
        finally:
            WorkbookParser.pivot_caches = current


def load_workbook_safe(path: str, **kwargs):
    """打开工作簿，并把损坏或伪装格式异常转换为中文业务错误。"""

    try:
        return openpyxl.load_workbook(path, **kwargs)
    except (zipfile.BadZipFile, OSError) as exc:
        raise ValueError(
            "无法打开 %s —— 文件可能已损坏,或不是真正的 xlsx"
            "(例如把 .xls 直接改名为 .xlsx)。请在 Excel 里另存为 .xlsx 后重试。"
            % os.path.basename(path)
        ) from exc
    except Exception as exc:
        # InvalidFileException 在不同 openpyxl 版本中的公开导入路径不同，按类名兼容。
        if exc.__class__.__name__ == "InvalidFileException":
            raise ValueError(
                "无法打开 %s —— 不是受支持的 Excel 格式,请另存为 .xlsx 后重试。"
                % os.path.basename(path)
            ) from exc
        raise


def load_data_only(path: str, **kwargs):
    """以 data_only 模式加载工作簿，并跳过不参与业务计算的缓存与外部链接。"""

    # data_only 调用方只消费当前文件已保存的单元格结果，不维护外部链接关系。
    # 使用 setdefault 保留少数明确需要外部链接的未来调用方主动覆盖的能力。
    kwargs.setdefault("keep_links", False)
    with skip_pivot_cache_parse():
        return load_workbook_safe(path, data_only=True, **kwargs)


def load_data_only_stream(path: str):
    """以只读模式打开大表，跳过外部链接并修复错误的 dimension 声明。"""

    with skip_pivot_cache_parse():
        # 某些评审表实际只有数百行，却内嵌数百个外部工作簿链接。默认恢复链接会让
        # 一个几 MB 的文件耗时数分钟并占用近 GB 内存；业务计算不使用这些关系。
        workbook = load_workbook_safe(
            path, data_only=True, read_only=True, keep_links=False,
        )
    try:
        for worksheet in workbook.worksheets:
            reset = getattr(worksheet, "reset_dimensions", None)
            if reset is not None:
                reset()
    except Exception:
        # 不能把仅部分修复的工作簿句柄交给调用方；先释放 Windows 文件锁再重抛。
        workbook.close()
        raise
    return workbook


def _read_modern_sheets(path: str) -> list[tuple[str, list[list[object]]]]:
    """读取 xlsx/xlsm 的全部页签值，并确保工作簿句柄及时关闭。"""

    workbook = load_data_only(path)
    try:
        return [
            (worksheet.title, [list(row) for row in worksheet.iter_rows(values_only=True)])
            for worksheet in workbook.worksheets
        ]
    finally:
        workbook.close()


def _xls_cell_value(cell, datemode: int):
    """恢复旧 xls 日期单元格；损坏日期保留原始值供上层决定。"""

    if cell.ctype != 3:
        return cell.value
    try:
        year, month, day, hour, minute, second = __import__("xlrd").xldate_as_tuple(
            cell.value, datemode
        )
        return datetime.datetime(year, month, day, hour, minute, second)
    except Exception:
        return cell.value


def _read_legacy_sheets(path: str) -> list[tuple[str, list[list[object]]]]:
    """读取旧 xls 全部页签，并把日期类型恢复为 datetime。"""

    import xlrd

    book = xlrd.open_workbook(path)
    result: list[tuple[str, list[list[object]]]] = []
    for sheet in book.sheets():
        rows = [
            [_xls_cell_value(sheet.cell(row, column), book.datemode) for column in range(sheet.ncols)]
            for row in range(sheet.nrows)
        ]
        result.append((sheet.name, rows))
    return result


def read_sheets(path: str) -> list[tuple[str, list[list[object]]]]:
    """把 xlsx、xlsm 或 xls 读取为 ``[(页签名, 二维行列表), ...]``。"""

    extension = os.path.splitext(path)[1].lower()
    if extension in (".xlsx", ".xlsm"):
        # 普通模式会重新发现真实边界，避免错误 dimension 声明让只读迭代静默漏行。
        return _read_modern_sheets(path)
    if extension == ".xls":
        return _read_legacy_sheets(path)
    raise ValueError("不支持的文件类型：%s" % extension)


__all__ = [
    "detect_uncached_formula",
    "load_data_only",
    "load_data_only_stream",
    "load_workbook_safe",
    "read_sheets",
    "skip_pivot_cache_parse",
    "warn_if_uncached",
]
