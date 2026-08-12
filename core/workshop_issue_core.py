# -*- coding: utf-8 -*-
"""车间现场问题的模板协议、图片规范化与标准报表导出。

五类现场问题拥有不同的录入字段、必填项、图片要求和 Excel 列布局。本模块集中维护
这些规则，供服务端校验、看板归一化和报表导出共同使用，避免前端隐藏字段后仍可写入
不属于该类型的数据，也避免导出格式与录入规范脱节。
"""

from __future__ import annotations

import os
from dataclasses import dataclass
from datetime import date, datetime
from math import ceil
from collections.abc import Mapping

import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageOps, UnidentifiedImageError

from . import paths, settings


IMAGE_FORMATS = {"JPEG": (".jpg", "image/jpeg"), "PNG": (".png", "image/png"), "WEBP": (".jpg", "image/jpeg")}
MAX_IMAGE_EDGE = 2400
MAX_IMAGE_PIXELS = 40_000_000
WORKSHOP_ISSUE_CATEGORY_ORDER = (
    "main_material", "auxiliary_material", "packaging", "overseas", "error_proofing",
)
WORKSHOP_ISSUE_CATEGORY_LABELS = {
    "main_material": "主料异常",
    "auxiliary_material": "辅料异常",
    "packaging": "包装异常",
    "overseas": "海外问题",
    "error_proofing": "防错异常",
}
WORKSHOP_ISSUE_CATEGORY_SHEETS = {
    "main_material": "主料异常",
    "auxiliary_material": "辅料异常",
    "packaging": "包装异常",
    "overseas": "海外历史记录",
    "error_proofing": "防错异常",
}

# 数据库保留这些字段作为五种模板的稳定超集；每项同时声明最大文本长度和中文标签。
# 具体类型允许写哪些字段仍由 WORKSHOP_ISSUE_TEMPLATES 约束，不能把数据库存在某列
# 理解为所有类型都允许提交该字段。
WORKSHOP_ISSUE_TEMPLATE_FIELDS = {
    "issue_source": (80, "问题源"),
    "model": (80, "车型"),
    "country": (80, "国家"),
    "batch_no": (120, "批次号"),
    "team": (80, "班组"),
    "material_code": (160, "物料编码"),
    "material_name": (240, "物料名称"),
    "cause_analysis": (2000, "原因分析"),
    "corrective_action": (2000, "纠正措施"),
    "responsibility_party": (120, "责任方"),
    "external_inspection_owner": (120, "外检责任人"),
    "discoverer": (80, "发现人"),
    "issue_level": (40, "问题等级"),
    "quantity": (80, "故障数量"),
    "issue_type": (80, "故障类别"),
    "completion_date": (40, "完成时间"),
    "recurring": (40, "是否复发"),
    "record_count": (40, "记录次数"),
    "happened_at": (80, "发生时间"),
    "handling_time": (80, "处理时间"),
    "responsible_person": (120, "责任人"),
    "updated_by_name": (120, "更新人"),
    "carrier": (160, "承运商"),
    "supplier": (160, "供应商"),
    "tracking_status": (80, "状态"),
}

_MAIN_COLUMNS = (
    ("issue_date", "发现日期"), ("model", "车型"), ("batch_no", "批次号"),
    ("team", "班组"), ("material_code", "物料编码"), ("material_name", "物料名称"),
    ("cause", "故障描述"), ("cause_analysis", "原因分析"), ("corrective_action", "纠正措施"),
    ("images", "故障照片"), ("discoverer", "发现人"), ("issue_level", "问题\n等级"),
    ("quantity", "故障\n数量"), ("issue_type", "故障类别"), ("completion_date", "完成时间"),
    ("recurring", "是否\n复发"), ("carrier", "承运商"), ("supplier", "供应商"),
    ("external_inspection_owner", "外检责任人"),
)
_AUXILIARY_COLUMNS = (
    ("issue_source", "问题源"), ("issue_date", "发现日期"), ("model", "车型"),
    ("batch_no", "批次号"), ("material_code", "物料编码"), ("material_name", "物料名称"),
    ("cause", "故障描述"), ("cause_analysis", "原因分析"), ("corrective_action", "纠正措施"),
    ("images", "故障照片"), ("discoverer", "发现人"), ("issue_level", "问题\n等级"),
    ("quantity", "故障\n数量"), ("issue_type", "故障类别"), ("completion_date", "完成时间"),
    ("recurring", "是否\n复发"),
)
_PACKAGING_COLUMNS = (
    ("issue_source", "问题源"), ("issue_date", "发现日期"), ("model", "车型"),
    ("batch_no", "批次号"), ("team", "班组"), ("material_code", "物料编码"),
    ("material_name", "物料名称"), ("cause", "故障描述"), ("cause_analysis", "原因分析"),
    ("corrective_action", "纠正措施"), ("images", "故障照片"), ("responsibility_party", "责任方"),
    ("discoverer", "发现人"), ("issue_level", "问题\n等级"), ("quantity", "故障\n数量"),
    ("issue_type", "故障类别"), ("completion_date", "完成时间"), ("recurring", "是否\n复发"),
)
_OVERSEAS_COLUMNS = (
    ("issue_date", "发现日期"), ("model", "车型"), ("country", "国家"), ("batch_no", "批次号"),
    ("material_code", "物料编码"), ("material_name", "物料名称"), ("quantity", "数量"),
    ("cause", "问题描述"), ("cause_analysis", "原因分析"), ("corrective_action", "整改措施"),
    ("completion_date", "完成时间"), ("responsible_person", "负责人"), ("issue_type", "故障类别"),
    ("issue_level", "故障等级"), ("record_count", "记录次数"), ("images", "涉及照片"),
)
_ERROR_PROOFING_COLUMNS = (
    ("happened_at", "发生时间"), ("batch_no", "批次号"), ("material_name", "物料名称"),
    ("material_code", "物料号"), ("cause", "问题描述"), ("cause_analysis", "原因分析"),
    ("corrective_action", "纠正措施"), ("tracking_status", "状态"), ("handling_time", "处理时间"),
    ("responsible_person", "责任人"), ("updated_by_name", "更新人"), ("notes", "备注"),
)
# 每个模板同时服务三个场景：API 必填和图片校验、前端动态表单、标准 Excel 复刻。
# columns 的字段顺序必须与规范表格一致；owner_fields 按优先级决定看板主要负责人。
WORKSHOP_ISSUE_TEMPLATES = {
    "main_material": {
        "label": WORKSHOP_ISSUE_CATEGORY_LABELS["main_material"], "sheet": WORKSHOP_ISSUE_CATEGORY_SHEETS["main_material"],
        "title": "主料故障信息", "columns": _MAIN_COLUMNS, "header_mode": "grouped", "prefix_columns": 0,
        "merge_prefix": True, "title_end_offset": 1, "detached_fields": ("external_inspection_owner",),
        "header_font_size": 14, "row_heights": (24, 48, 60),
        "required": ("cause", "discoverer"), "requires_images": True,
        "owner_fields": ("external_inspection_owner", "discoverer"),
    },
    "auxiliary_material": {
        "label": WORKSHOP_ISSUE_CATEGORY_LABELS["auxiliary_material"], "sheet": WORKSHOP_ISSUE_CATEGORY_SHEETS["auxiliary_material"],
        "title": "辅料故障信息", "columns": _AUXILIARY_COLUMNS, "header_mode": "grouped", "prefix_columns": 1,
        "merge_prefix": True, "title_end_offset": 0, "detached_fields": (),
        "header_font_size": 14, "row_heights": (24, 45, 44),
        "required": ("cause", "discoverer"), "requires_images": True,
        "owner_fields": ("discoverer",),
    },
    "packaging": {
        "label": WORKSHOP_ISSUE_CATEGORY_LABELS["packaging"], "sheet": WORKSHOP_ISSUE_CATEGORY_SHEETS["packaging"],
        "title": "包装故障信息", "columns": _PACKAGING_COLUMNS, "header_mode": "grouped", "prefix_columns": 1,
        "merge_prefix": False, "title_end_offset": 0, "detached_fields": (),
        "header_font_size": 16, "row_heights": (55, 55, 55),
        "required": ("cause", "responsibility_party", "discoverer"), "requires_images": True,
        "owner_fields": ("responsibility_party", "discoverer"),
    },
    "overseas": {
        "label": WORKSHOP_ISSUE_CATEGORY_LABELS["overseas"], "sheet": WORKSHOP_ISSUE_CATEGORY_SHEETS["overseas"],
        "title": "海外问题统计", "columns": _OVERSEAS_COLUMNS, "header_mode": "title_row", "prefix_columns": 0,
        "header_font_size": 14, "row_heights": (24, 24, 60),
        "required": ("cause", "material_name", "responsible_person"), "requires_images": True,
        "owner_fields": ("responsible_person", "primary_owner"),
    },
    "error_proofing": {
        "label": WORKSHOP_ISSUE_CATEGORY_LABELS["error_proofing"], "sheet": WORKSHOP_ISSUE_CATEGORY_SHEETS["error_proofing"],
        "title": "防错异常处理", "columns": _ERROR_PROOFING_COLUMNS, "header_mode": "title_row", "prefix_columns": 0,
        "header_font_size": 14, "row_heights": (24, 24, 45),
        "required": ("happened_at", "cause", "tracking_status", "responsible_person"), "requires_images": False,
        "owner_fields": ("responsible_person",),
    },
}

_TEMPLATE_COLUMN_WIDTHS = {
    "main_material": (4.56, 14.44, 7.78, 21.27, 18.33, 22.54, 28.67, 18.33, 62.89, 18.0, 30.56, 11.33, 11.33, 13.0, 15.41, 13.18, 11.33, 11.33, 13.0, 16.89),
    "auxiliary_material": (11.33, 16.11, 11.33, 11.33, 24.0, 18.33, 23.11, 27.78, 11.33, 13.0, 16.78, 11.33, 11.33, 13.0, 15.44, 11.33, 11.33),
    "packaging": (11.33, 14.67, 17.22, 11.33, 30.0, 18.33, 24.6, 39.0, 46.67, 60.44, 52.0, 26.03, 17.67, 11.33, 13.0, 13.0, 13.0, 19.36, 11.33),
    "overseas": (11.33, 13.0, 13.0, 13.0, 13.0, 14.0, 17.78, 11.33, 45.0, 56.67, 63.24, 14.0, 11.33, 13.0, 13.0, 13.0, 13.0),
    "error_proofing": (18.11, 18.56, 45.22, 14.0, 53.11, 79.89, 37.44, 11.33, 25.78, 15.44, 11.33, 13.0),
    "overview": (11.33, 13.0, 13.0, 13.56, 11.33, 13.0, 17.78, 12.67, 22.22, 19.67, 31.89, 11.33, 13.0, 13.0, 13.0, 13.0, 13.0, 13.0, 13.0),
}

# 从导出列自动派生接口字段白名单，排除所有类型共用的日期、描述和图片字段，避免维护
# 两套清单后发生偏差。
WORKSHOP_ISSUE_CATEGORY_ALLOWED_FIELDS = {
    category: frozenset(field for field, _ in template["columns"] if field not in {"issue_date", "cause", "images"})
    for category, template in WORKSHOP_ISSUE_TEMPLATES.items()
}
WORKSHOP_ISSUE_IMAGE_REQUIRED_CATEGORIES = frozenset(
    category for category, template in WORKSHOP_ISSUE_TEMPLATES.items() if template["requires_images"]
)

# 旧分类映射只用于数据库升级、历史记录展示和回收站恢复。新建记录必须使用正式五类，
# 不得把下列旧名称重新暴露为可选择的问题类型。
LEGACY_WORKSHOP_CATEGORY_NAMES = {
    "main_material": "main_material", "auxiliary_material": "auxiliary_material", "packaging": "packaging",
    "overseas": "overseas", "error_proofing": "error_proofing", "quality": "main_material",
    "material": "main_material", "equipment": "error_proofing", "process": "error_proofing",
    "safety": "error_proofing", "people": "error_proofing", "other": "main_material",
}


def normalize_workshop_category(value: object, issue: Mapping[str, object] | None = None) -> str:
    """将旧版分类和历史记录转换成正式五类模板之一。

    正式分类值直接返回；历史记录再依据国家、记录次数、防错时间字段和问题源推断，
    最后才使用旧名称映射。无法判断时回退主料异常，保证旧记录可展示，但该回退不用于
    放宽新建接口的分类校验。
    """
    raw = str(value or "").strip()
    if raw in WORKSHOP_ISSUE_TEMPLATES:
        return raw
    item = issue or {}
    source = str(item.get("issue_source") or "")
    if str(item.get("country") or "").strip() or str(item.get("record_count") or "").strip():
        # 国家和记录次数是海外问题模板的辨识度最高字段，优先于旧分类名。
        return "overseas"
    if any(str(item.get(field) or "").strip() for field in ("happened_at", "handling_time", "updated_by_name")):
        # 发生时间、处理时间和更新人只存在于防错异常规范中。
        return "error_proofing"
    if "辅料" in source:
        return "auxiliary_material"
    if raw == "packaging" or "包装" in source:
        return "packaging"
    return LEGACY_WORKSHOP_CATEGORY_NAMES.get(raw, "main_material")


def workshop_issue_primary_owner(category: object, values: Mapping[str, object], fallback: object = "") -> str:
    """按模板定义的责任字段优先级选择看板主要负责人。

    不同问题类型的责任主体不同，例如包装问题优先责任方，海外问题优先负责人。只有
    模板字段均为空时才使用通用 ``fallback``，避免硬编码某个字段适用于所有类型。
    """
    normalized = normalize_workshop_category(category, values)
    template = WORKSHOP_ISSUE_TEMPLATES[normalized]
    for field in template["owner_fields"]:
        value = str(values.get(field) or "").strip()
        if value:
            return value
    return str(fallback or "").strip()


def workshop_issue_severity(values: Mapping[str, object], fallback: object = "normal") -> str:
    """根据模板中的问题等级推导看板统一严重程度。

    源表可能使用字母、英文或中文描述，统一映射为 ``critical``、``important`` 和
    ``normal`` 三档。这里只改变展示分组，不改写管理员提交的原始等级文本。
    """
    raw = str(values.get("issue_level") or fallback or "normal").strip().lower()
    if raw in {"a", "critical", "重大", "重大/升级", "升级"} or "重大" in raw or "升级" in raw:
        return "critical"
    if raw in {"b", "important", "重点"} or "重点" in raw:
        return "important"
    return "normal"


def normalize_image(source_path, target_stem):
    """校验并规范化手机图片，返回可持久化的图片元数据。

    图片先用 ``verify`` 检查文件结构，再重新打开进行 EXIF 方向纠正、像素上限检查和
    缩放。透明 PNG 保留透明通道，其他格式统一为压缩 JPEG；元数据只在成功写入后返回。
    """
    try:
        with Image.open(source_path) as source:
            # verify 只校验编码结构并会使图像对象不可继续解码，因此下面必须重新打开。
            source.verify()
        with Image.open(source_path) as source:
            image_format = str(source.format or "").upper()
            if image_format not in IMAGE_FORMATS:
                raise ValueError("仅支持 JPG、PNG 和 WebP 图片")
            source = ImageOps.exif_transpose(source)  # 手机照片常依靠 EXIF 标记方向，先转正像素。
            if source.width * source.height > MAX_IMAGE_PIXELS:
                raise ValueError("图片像素过大，请先在手机中缩小后再上传")
            # thumbnail 保持宽高比且绝不放大，小图不因规范化损失清晰度。
            source.thumbnail((MAX_IMAGE_EDGE, MAX_IMAGE_EDGE), Image.Resampling.LANCZOS)
            has_alpha = source.mode in {"RGBA", "LA"} or "transparency" in source.info
            extension, content_type = IMAGE_FORMATS[image_format]
            if has_alpha and image_format == "PNG":
                # 只有透明 PNG 保留透明度；WebP 统一转 JPEG，减少 Excel 兼容问题。
                extension, content_type = ".png", "image/png"
                normalized = source.convert("RGBA")
                save_options = {"optimize": True}
            else:
                extension, content_type = ".jpg", "image/jpeg"
                if has_alpha:
                    # JPEG 不支持透明通道，先铺白底避免透明区域被渲染成黑色。
                    background = Image.new("RGB", source.size, "white")
                    background.paste(source, mask=source.getchannel("A"))
                    normalized = background
                else:
                    normalized = source.convert("RGB")
                save_options = {"quality": 88, "optimize": True, "progressive": True}
            target = os.path.abspath(str(target_stem) + extension)
            os.makedirs(os.path.dirname(target), exist_ok=True)
            normalized.save(target, **save_options)
            width, height = normalized.size
    except (UnidentifiedImageError, OSError) as exc:
        raise ValueError("图片无法读取或文件已经损坏") from exc
    return {
        "path": target,
        "content_type": content_type,
        "size": os.path.getsize(target),
        "width": width,
        "height": height,
    }


_EMU_PER_PIXEL = 9525
_HEADER_GRAY = "D9D9D9"
_HEADER_ORANGE = "ED7D31"
_HEADER_YELLOW = "FFFF00"
_HEADER_RED = "FF0000"


def _add_thumbnails(sheet, images, row: int, column: int) -> int:
    """在模板图片单元格中按两列排列缩略图，返回实际插入数量。

    只处理仍存在的受管图片文件。每张图限制在 72×56 像素内，两列排列，超过两张后
    向下扩展；返回数量用于调用方计算对应数据行高度，实现多图片适配。
    """
    available = [
        item for item in images
        if isinstance(item, Mapping) and os.path.isfile(str(item.get("path") or ""))
    ]
    for index, item in enumerate(available):
        image = ExcelImage(str(item.get("path")))
        scale = min(72 / max(1, image.width), 56 / max(1, image.height), 1)  # 最后一项防止放大小图。
        image.width = max(1, int(image.width * scale))
        image.height = max(1, int(image.height * scale))
        col_offset = (index % 2) * 76
        row_offset = (index // 2) * 60
        # openpyxl 锚点偏移使用 EMU 单位，列号和行号则从零开始。
        marker = AnchorMarker(
            col=column - 1, colOff=col_offset * _EMU_PER_PIXEL,
            row=row - 1, rowOff=row_offset * _EMU_PER_PIXEL,
        )
        image.anchor = OneCellAnchor(
            _from=marker,
            ext=XDRPositiveSize2D(
                cx=image.width * _EMU_PER_PIXEL,
                cy=image.height * _EMU_PER_PIXEL,
            ),
        )
        sheet.add_image(image)
    return len(available)


def _issue_value(issue: Mapping[str, object], field: str) -> object:
    """读取普通导出字段；图片列留空以便稍后插入真正的 Excel 图片对象。"""
    if field == "images":
        return ""
    return issue.get(field, "")


@dataclass(frozen=True)
class _TemplateStyles:
    """保存单张现场问题工作表复用的样式对象，避免为每个单元格重复创建样式。"""

    border: Border
    gray_fill: PatternFill
    orange_fill: PatternFill
    yellow_fill: PatternFill
    white_fill: PatternFill
    title_font: Font
    header_font: Font
    black_header_font: Font
    body_font: Font


def _template_styles(template: Mapping[str, object]) -> _TemplateStyles:
    """按模板字号创建一次表头与正文样式集合。"""
    line = Side(style="thin", color="000000")
    header_size = int(template.get("header_font_size", 14))
    return _TemplateStyles(
        border=Border(left=line, right=line, top=line, bottom=line),
        gray_fill=PatternFill("solid", fgColor=_HEADER_GRAY),
        orange_fill=PatternFill("solid", fgColor=_HEADER_ORANGE),
        yellow_fill=PatternFill("solid", fgColor=_HEADER_YELLOW),
        white_fill=PatternFill("solid", fgColor="FFFFFF"),
        title_font=Font(name="微软雅黑", color=_HEADER_RED, bold=True, size=header_size),
        header_font=Font(name="微软雅黑", color=_HEADER_RED, bold=True, size=header_size),
        black_header_font=Font(name="微软雅黑", color="000000", bold=True, size=header_size),
        body_font=Font(name="微软雅黑", color="222222", size=12),
    )


def _style_header_cell(cell, *, fill, font, border: Border) -> None:
    """应用现场问题表头的公共居中、换行和边框规则。"""
    cell.fill = fill
    cell.font = font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border


def _write_grouped_header(
    sheet, template: Mapping[str, object], export_columns: list[tuple[str, str]],
    styles: _TemplateStyles,
) -> None:
    """写入主料、辅料和包装模板的两层分组表头。"""
    prefix = int(template["prefix_columns"]) + 1  # 分组模板均带问题编号，因此在业务前缀列数上加一。
    merge_prefix = bool(template.get("merge_prefix", True))
    for column in range(1, prefix + 1):
        if merge_prefix:
            sheet.merge_cells(start_row=1, start_column=column, end_row=2, end_column=column)
        _style_header_cell(
            sheet.cell(1, column, export_columns[column - 1][1]),
            fill=styles.orange_fill, font=styles.black_header_font, border=styles.border,
        )
        if not merge_prefix:
            # 部分规范保留上下两个独立前缀格；空的下格仍需完整样式以维持打印边界。
            _style_header_cell(
                sheet.cell(2, column), fill=styles.orange_fill,
                font=styles.black_header_font, border=styles.border,
            )

    title_start = prefix + 1
    title_end = len(export_columns) - int(template.get("title_end_offset", 0))
    sheet.merge_cells(start_row=1, start_column=title_start, end_row=1, end_column=title_end)
    title_cell = sheet.cell(1, title_start, str(template["title"]))
    title_cell.fill = styles.gray_fill
    title_cell.font = styles.title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    for column in range(title_start, title_end + 1):
        sheet.cell(1, column).fill = styles.gray_fill
        sheet.cell(1, column).border = styles.border

    detached_fields = set(template.get("detached_fields", ()))
    highlighted_fields = {"cause_analysis", "corrective_action", "completion_date"}
    for column, (field, label) in enumerate(export_columns[title_start - 1:], title_start):
        detached = field in detached_fields
        fill = (
            styles.white_fill if detached
            else styles.yellow_fill if field in highlighted_fields
            else styles.gray_fill
        )
        _style_header_cell(
            sheet.cell(2, column, label), fill=fill,
            font=styles.black_header_font if detached else styles.header_font,
            border=styles.border,
        )
        if detached:
            # 主标题范围外的独立字段上下两格都保持白底，不能继承主标题灰色填充。
            _style_header_cell(
                sheet.cell(1, column), fill=styles.white_fill,
                font=styles.black_header_font, border=styles.border,
            )


def _write_simple_header(
    sheet, template: Mapping[str, object], export_columns: list[tuple[str, str]],
    styles: _TemplateStyles,
) -> None:
    """写入海外和防错模板的整行标题与普通字段表头。"""
    last_column = len(export_columns)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    title_cell = sheet.cell(1, 1, str(template["title"]))
    title_cell.fill = styles.gray_fill
    title_cell.font = styles.black_header_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    for column in range(1, last_column + 1):
        sheet.cell(1, column).fill = styles.gray_fill
        sheet.cell(1, column).border = styles.border
    highlighted_fields = {"quantity", "cause_analysis", "issue_level"}
    for column, (field, label) in enumerate(export_columns, 1):
        _style_header_cell(
            sheet.cell(2, column, label),
            fill=styles.yellow_fill if field in highlighted_fields else styles.gray_fill,
            font=styles.black_header_font, border=styles.border,
        )


def _normalized_export_issue(category: str, issue: Mapping[str, object]) -> dict[str, object]:
    """生成不修改数据库原记录的导出副本，并补齐历史海外负责人字段。"""
    normalized = dict(issue)
    normalized["category"] = category
    if category == "overseas" and not str(normalized.get("responsible_person") or "").strip():
        # 历史海外记录可能只有通用主要负责人，按当前模板优先级投影到负责人列。
        normalized["responsible_person"] = workshop_issue_primary_owner(category, normalized)
    return normalized


def _write_issue_rows(
    sheet, category: str, issues: list[Mapping[str, object]],
    export_columns: list[tuple[str, str]], styles: _TemplateStyles, body_height: float,
) -> None:
    """写入问题正文、嵌入图片，并按图片行数调整记录高度。"""
    image_column = next(
        (index for index, (field, _) in enumerate(export_columns, 1) if field == "images"), 0,
    )
    for index, issue in enumerate(issues, 1):
        row = index + 2
        normalized = _normalized_export_issue(category, issue)
        values = [
            index if field == "__index__" else _issue_value(normalized, field)
            for field, _ in export_columns
        ]
        sheet.append(values)
        images = normalized.get("images") if isinstance(normalized.get("images"), list) else []
        inserted = _add_thumbnails(sheet, images, row, image_column) if image_column else 0
        # 每两张图片占一层 60 点高度；没有图片或图片较少时仍遵守模板正文最低高度。
        sheet.row_dimensions[row].height = max(
            body_height, ceil(inserted / 2) * 60 if inserted else body_height,
        )
        for column in range(1, len(export_columns) + 1):
            cell = sheet.cell(row, column)
            cell.font = styles.body_font
            cell.border = styles.border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _configure_template_sheet(
    sheet, template: Mapping[str, object], category: str, width_key: str | None,
    last_column: int,
) -> None:
    """设置列宽、冻结区域、筛选和打印分页等工作表级属性。"""
    widths = _TEMPLATE_COLUMN_WIDTHS[width_key or category]
    for column, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    # 主料冻结问题编号列，其他模板只冻结两行表头；显式 None 用于问题一览表关闭冻结。
    sheet.freeze_panes = template.get(
        "freeze_panes", "B3" if category == "main_material" else "A3",
    )
    sheet.auto_filter.ref = f"A2:{get_column_letter(last_column)}{max(2, sheet.max_row)}"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1  # 打印时横向压到一页，纵向允许自然分页。
    sheet.page_setup.fitToHeight = 0
    sheet.print_title_rows = "1:2"


def _write_template_sheet(
    workbook,
    category: str,
    issues: list[Mapping[str, object]],
    *,
    template_override: Mapping[str, object] | None = None,
    width_key: str | None = None,
) -> None:
    """按某类规范模板创建工作表、写入问题记录并嵌入多张图片。

    ``template_override`` 供问题一览表复用主料版式；``width_key`` 可选择另一组列宽。
    分组表头和单标题表头分别对应现行规范文件，不能统一成普通表头而丢失模板结构。
    """
    template = template_override or WORKSHOP_ISSUE_TEMPLATES[category]
    columns = list(template["columns"])
    show_index = category != "error_proofing"  # 防错规范表本身没有“问题编号”列。
    export_columns = ([('__index__', '问题编号')] if show_index else []) + columns
    sheet = workbook.create_sheet(str(template["sheet"]))
    sheet.sheet_view.showGridLines = False

    styles = _template_styles(template)
    last_column = len(export_columns)
    if template["header_mode"] == "grouped":
        _write_grouped_header(sheet, template, export_columns, styles)
    else:
        _write_simple_header(sheet, template, export_columns, styles)

    title_height, header_height, body_height = template.get("row_heights", (24, 42, 46))
    sheet.row_dimensions[1].height = title_height
    sheet.row_dimensions[2].height = header_height
    _write_issue_rows(sheet, category, issues, export_columns, styles, float(body_height))
    _configure_template_sheet(sheet, template, category, width_key, last_column)


def _write_overview_sheet(workbook) -> None:
    """按标准模板保留空白的“问题一览表”，供后续人工汇总使用。

    规范文件要求该页存在，但其内容不是五类问题明细的简单重复，因此只复用主料表样式
    创建空白页，不由程序擅自填充或推断汇总规则。
    """
    template = {
        **WORKSHOP_ISSUE_TEMPLATES["main_material"],
        "sheet": "问题一览表",
        "columns": _MAIN_COLUMNS[:-1],
        "title_end_offset": 0,
        "detached_fields": (),
        "row_heights": (24.75, 54, 60),
        "freeze_panes": None,
    }
    _write_template_sheet(
        workbook, "main_material", [], template_override=template, width_key="overview",
    )
    workbook["问题一览表"].auto_filter.ref = None


def _parse_report_date(value: object, label: str) -> date:
    """严格解析报表日期，并在错误中指出开始或结束日期。"""
    text = str(value or "").strip()
    try:
        return datetime.strptime(text, "%Y-%m-%d").date()
    except ValueError as exc:
        raise ValueError(f"{label}无效，应为 YYYY-MM-DD") from exc


def _report_range(issues, issue_date=None, start_date=None, end_date=None) -> tuple[str, str]:
    """确定导出日期范围并返回标准 ISO 日期。

    ``issue_date`` 是旧版单日接口的兼容参数；未显式给范围时，优先覆盖传入问题列表的
    最早至最晚日期，完全无数据时默认当天。开始日期晚于结束日期会明确拒绝。
    """
    if issue_date and not start_date:
        # 新参数优先，只有未提供 start_date 时才使用旧版 issue_date。
        start_date = issue_date
    if not start_date:
        available = sorted(
            str(issue.get("issue_date") or "").strip()
            for issue in issues
            if isinstance(issue, Mapping) and str(issue.get("issue_date") or "").strip()
        )
        start_date = available[0] if available else datetime.now().strftime("%Y-%m-%d")
        if not end_date and available:
            end_date = available[-1]
    if not end_date:
        end_date = start_date
    parsed_start = _parse_report_date(start_date, "开始日期")
    parsed_end = _parse_report_date(end_date, "结束日期")
    if parsed_start > parsed_end:
        raise ValueError("开始日期不能晚于结束日期")
    return parsed_start.isoformat(), parsed_end.isoformat()


def run(issues, issue_date=None, start_date=None, end_date=None, out_dir=None, log=None):
    """按标准模板把指定日期范围内的五类现场问题导出为 Excel。

    调用方应已按日期和权限筛选问题，本函数负责分类、版式和图片嵌入。五个正式分类
    工作表始终生成，即使某类没有记录也保留模板；最后追加规范要求的空白问题一览表。
    """
    if out_dir is None:
        st = settings.get_settings()
        out_dir = paths.resolve_output_dir("workshop_issue", **st.output_kwargs())
    os.makedirs(out_dir, exist_ok=True)
    report_start, report_end = _report_range(
        issues, issue_date=issue_date, start_date=start_date, end_date=end_date,
    )
    range_label = report_start if report_start == report_end else f"{report_start}至{report_end}"
    output = os.path.join(out_dir, f"异常问题报告-{range_label}.xlsx")

    grouped: dict[str, list[Mapping[str, object]]] = {
        category: [] for category in WORKSHOP_ISSUE_CATEGORY_ORDER
    }
    for issue in issues:
        if not isinstance(issue, Mapping):
            continue
        # 历史分类先归一化，确保导出只出现规范中的五个页签。
        category = normalize_workshop_category(issue.get("category"), issue)
        grouped[category].append(issue)

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)  # 删除默认空白页，后续按规范顺序创建全部页签。
    for category in WORKSHOP_ISSUE_CATEGORY_ORDER:
        _write_template_sheet(workbook, category, grouped[category])
    _write_overview_sheet(workbook)

    workbook.save(output)
    if log:
        log(f"已按标准异常问题模板导出 {sum(len(rows) for rows in grouped.values())} 条现场问题：{output}")
    return {
        "out_dir": out_dir,
        "out_file": output,
        "count": sum(len(rows) for rows in grouped.values()),
        "start_date": report_start,
        "end_date": report_end,
        "category_counts": {category: len(grouped[category]) for category in WORKSHOP_ISSUE_CATEGORY_ORDER},
    }
