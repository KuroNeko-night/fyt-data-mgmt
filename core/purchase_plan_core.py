# -*- coding: utf-8 -*-
"""采购计划导入与实收差异提取业务核心。

输入一个带供应商代码子表的模板文件和一份或多份辅料清单总表；
每个批次按文件名中的批次号（如 26036-02、26178A）生成一个输出文件。
模板中的“仓库编号、采购员编号、预计到货日期”列保留模板里已填写的值，未填写则留空。
材料名称含“原厂”的记录按业务约定排除，不进入输出。

正式生成采用“临时目录全部准备成功后再提交”的事务式流程，避免多个批次处理中途失败时
留下不完整的一半结果。业务文件中的材料属性优先，主数据库只补空值；全部输出成功后
才学习新材料和供应商代码，使主数据不会被失败任务的中间状态污染。
"""
from __future__ import annotations

import os
import re
import shutil
import tempfile
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

from . import common_core, material_catalog, paths, settings
from .supplier_batch_core import _best_sheet, _is_original, _quantity, _supplier, _text

EXCEL_SUFFIXES = {".xlsx", ".xlsm"}
# 批次号允许五位数字、可选字母尾缀和可选“-两位”子批次，例如 26178A、26036-02。
BATCH_PATTERN = re.compile(r"(\d{5}[A-Z]?(?:-\d{2})?)")
MAIN_SHEET_KEYWORD = "模板"
SUPPLIER_SHEET_NAME = "Sheet1"
# 模板主表固定写入列：3 产品编号、4 产品名称、5 规格、6 供应商编码、7 供应商、8 数量。
DATA_FONT = Font(name="宋体", size=10)
# 透视或汇总模板常带这些伪材料编号，必须在业务行提取时过滤。
SUMMARY_IDS = {"(空白)", "总计"}


def _validate_files(values, label: str) -> list[str]:
    """校验采购计划输入文件存在且为受支持的 Excel 格式，并返回绝对路径。"""
    result = [os.path.abspath(str(value)) for value in (values or []) if str(value).strip()]
    if not result:
        raise ValueError("%s不能为空" % label)
    for path in result:
        if not os.path.isfile(path):
            raise FileNotFoundError("找不到文件：%s" % path)
        if Path(path).suffix.lower() not in EXCEL_SUFFIXES:
            raise ValueError("%s仅支持 xlsx 或 xlsm 文件：%s" % (label, os.path.basename(path)))
    return result


def _batch_name(path: str) -> str:
    """按统一正则从文件名中提取首个业务批次号；无法识别时返回空字符串。"""
    match = BATCH_PATTERN.search(Path(path).stem)
    return match.group(1) if match else ""


def _read_supplier_codes(template_path: str) -> dict[str, str]:
    """
    从模板约定子表读取 ``供应商名称 -> 供应商编码`` 映射，名称位于 B 列、编码位于 C 列。

    该子表是采购计划模板的必要组成部分；完全缺失或没有任何有效映射时立即报错，避免
    后续生成采购系统无法导入的空供应商编码。重复名称按表格从上到下采用最后有效行。
    """
    workbook = openpyxl.load_workbook(template_path, read_only=True, data_only=True)
    try:
        if SUPPLIER_SHEET_NAME not in workbook.sheetnames:
            raise ValueError("模板中未找到供应商代码子表 %s" % SUPPLIER_SHEET_NAME)
        sheet = workbook[SUPPLIER_SHEET_NAME]
        mapping: dict[str, str] = {}
        for row in sheet.iter_rows(min_row=1, values_only=True):
            name = row[1] if len(row) > 1 else None
            code = row[2] if len(row) > 2 else None
            if name is not None and code is not None:
                # 使用规范化名称作为键；后出现的维护记录覆盖旧行，符合表格人工修订习惯。
                mapping[str(name).strip()] = str(code).strip()
        if not mapping:
            raise ValueError("模板供应商代码子表为空，请先维护供应商代码")
        return mapping
    finally:
        workbook.close()


def _main_sheet(workbook) -> object:
    """优先选择名称含“模板”的采购计划主表；旧模板无约定名称时兼容使用首张表。"""
    for worksheet in workbook.worksheets:
        if MAIN_SHEET_KEYWORD in str(worksheet.title):
            return worksheet
    return workbook.worksheets[0]


def _row_value(values, column):
    """按可选 1 基列号安全读取只读行，缺列或短行时返回 ``None``。"""

    return values[int(column) - 1] if column and int(column) <= len(values) else None


@dataclass
class _BatchRowContext:
    """保存一个批次逐行转换所需的主数据快照及累计统计。"""

    supplier_codes: dict[str, str]
    known_materials: set[str]
    known_suppliers: set[str]
    resolver: object = None
    fill_counts: dict[str, int] | None = None
    excluded_original: int = 0
    unknown_materials: set[str] = field(default_factory=set)
    unknown_suppliers: set[str] = field(default_factory=set)
    learned_items: list[dict[str, object]] = field(default_factory=list)


def _purchase_row_item(values, columns, context):
    """把一行辅料清单转换为模板业务列；无采购意义的行返回 ``None``。"""

    code = _text(_row_value(values, columns.get("code")))
    if not code or code in SUMMARY_IDS:
        return None
    item = {
        "code": code,
        "name": _text(_row_value(values, columns.get("name"))),
        "spec": _text(_row_value(values, columns.get("spec"))),
        "unit": _text(_row_value(values, columns.get("unit"))),
        "supplier": _supplier(_row_value(values, columns.get("supplier"))),
    }
    if context.resolver is not None:
        # 正式主数据只补空值，源业务表中已经明确填写的材料属性始终优先。
        context.resolver.fill_mapping(item, counts=context.fill_counts)

    name = str(item["name"])
    if _is_original(name):
        context.excluded_original += 1
        return None
    quantity = _quantity(_row_value(values, columns.get("qty")))
    supplier = str(item["supplier"])
    if quantity is None or quantity <= 0 or not supplier:
        return None

    supplier_code = context.supplier_codes.get(supplier)
    if not supplier_code and context.resolver is not None:
        supplier_code = context.resolver.complete_supplier_code(
            supplier, counts=context.fill_counts,
        )
    if not supplier_code:
        raise ValueError(
            "模板与主数据档案中都缺少供应商“%s”的代码，请先在模板子表或主数据档案维护"
            % supplier
        )

    material_known = (
        context.resolver.has_material(code)
        if context.resolver is not None
        else code in context.known_materials
    )
    supplier_known = (
        context.resolver.has_supplier(supplier)
        if context.resolver is not None
        else supplier in context.known_suppliers
    )
    if not material_known:
        context.unknown_materials.add(code)
    if not supplier_known:
        context.unknown_suppliers.add(supplier)

    spec = str(item["spec"])
    unit = str(item["unit"])
    context.learned_items.append({
        "code": code,
        "name": name,
        "spec": spec,
        "unit": unit,
        "supplier": supplier,
    })
    return [code, name, spec, supplier_code, supplier, quantity]


def _collect_purchase_rows(worksheet, layout, context):
    """扫描一个批次业务页签，返回可写入采购计划模板的有效记录。"""

    columns = layout["columns"]
    max_column = max(int(value) for value in columns.values())
    rows = []
    for values in worksheet.iter_rows(
        min_row=int(layout["header_row"]) + 1,
        max_col=max_column,
        values_only=True,
    ):
        if all(value is None or _text(value) == "" for value in values):
            continue
        converted = _purchase_row_item(values, columns, context)
        if converted is not None:
            rows.append(converted)
    return rows


def _clear_purchase_template(sheet):
    """清除模板中的旧业务值，同时保留管理员维护的三个默认参数列。"""

    keep_columns = {1, 2, 10}
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=11):
        for cell in row:
            if cell.column not in keep_columns:
                cell.value = None
    return keep_columns


def _write_purchase_rows(sheet, rows, keep_columns):
    """写入采购计划业务行，并清除新数据末尾残留的模板示例参数。"""

    for row_index, values in enumerate(rows, start=2):
        for column, value in enumerate(values, start=3):
            cell = sheet.cell(row=row_index, column=column, value=value)
            cell.font = DATA_FONT
    first_extra = 2 + len(rows)
    if first_extra > sheet.max_row:
        return
    for row in sheet.iter_rows(min_row=first_extra, max_row=sheet.max_row, max_col=11):
        for cell in row:
            if cell.column in keep_columns:
                cell.value = None


def _write_purchase_output(template_path, target, rows):
    """复制并填写采购计划模板，保留原有样式、验证和隐藏子表。"""

    shutil.copy(template_path, target)
    output = openpyxl.load_workbook(target)
    try:
        sheet = _main_sheet(output)
        keep_columns = _clear_purchase_template(sheet)
        _write_purchase_rows(sheet, rows, keep_columns)
        output.save(target)
    finally:
        output.close()


def _convert_batch(
    template_path: str,
    batch_path: str,
    target: str,
    batch_name: str,
    supplier_codes: dict[str, str],
    known_materials: set[str],
    known_suppliers: set[str],
    log=None,
    resolver=None,
    fill_counts=None,
) -> tuple[int, int, set[str], set[str], list[dict[str, object]]]:
    """
    把一个辅料批次清单转换到模板副本，写入指定临时路径并返回处理统计。

    返回值依次为有效行数、原厂排除数、新材料集合、新供应商集合和可学习材料明细。
    函数只准备文件和学习候选，不直接更新主数据库；调用方确认所有批次均成功提交后
    才统一学习，保证输出与主数据状态的一致性。
    """
    # data_only 依赖 Excel 保存的公式缓存，先提示未刷新公式，避免数量被静默当成空值。
    common_core.warn_if_uncached(batch_path, log, what="最终采购数量")
    workbook = openpyxl.load_workbook(batch_path, read_only=True, data_only=True)
    try:
        worksheet, layout = _best_sheet(workbook)
        context = _BatchRowContext(
            supplier_codes=supplier_codes,
            known_materials=known_materials,
            known_suppliers=known_suppliers,
            resolver=resolver,
            fill_counts=fill_counts,
        )
        rows = _collect_purchase_rows(worksheet, layout, context)
    finally:
        workbook.close()

    if not rows:
        raise ValueError("批次 %s 没有可制作的采购记录" % batch_name)

    _write_purchase_output(template_path, target, rows)
    return (
        len(rows),
        context.excluded_original,
        context.unknown_materials,
        context.unknown_suppliers,
        context.learned_items,
    )


def _reserve_output_path(out_dir: str, batch_name: str, reserved: set[str]) -> str:
    """
    为批次预留一个既不覆盖磁盘文件、也不与本次其他批次冲突的输出路径。

    ``reserved`` 使用 ``normcase`` 后的路径，覆盖 Windows 大小写不敏感行为；重复名称
    依次添加 ``(2)``、``(3)``。这里只预留名称，真正提交前仍会再次检查并发占用。
    """
    base = os.path.join(out_dir, "%s.xlsx" % batch_name)
    root, extension = os.path.splitext(base)
    target = base
    index = 2
    while os.path.exists(target) or os.path.normcase(target) in reserved:
        target = "%s (%d)%s" % (root, index, extension)
        index += 1
    reserved.add(os.path.normcase(target))
    return target


@dataclass
class _PurchasePlanContext:
    """保存采购计划批量生成阶段共享的模板、主数据和输出配置。"""

    template_path: str
    batches: list[str]
    out_dir: str
    supplier_codes: dict[str, str]
    template_supplier_codes: dict[str, str]
    resolver: object
    fill_counts: dict[str, int]
    known_materials: set[str]
    known_suppliers: set[str]


@dataclass
class _PurchasePlanTotals:
    """累计全部批次的输出文件、行数、排除项和主数据学习候选。"""

    files: list[str] = field(default_factory=list)
    rows: int = 0
    excluded_original: int = 0
    new_materials: set[str] = field(default_factory=set)
    new_suppliers: set[str] = field(default_factory=set)
    learned_items: list[dict[str, object]] = field(default_factory=list)

    def add_batch(self, final_target, converted):
        """合并一个已完整写入临时文件的批次统计。"""

        rows, excluded, materials, suppliers, items = converted
        self.files.append(final_target)
        self.rows += rows
        self.excluded_original += excluded
        self.new_materials.update(materials)
        self.new_suppliers.update(suppliers)
        self.learned_items.extend(items)


def _resolve_purchase_plan_output_dir(batches, out_dir):
    """解析采购计划输出目录，并为源文件旁模式提供首批次定位基准。"""

    if out_dir is None:
        current = settings.get_settings()
        return paths.resolve_output_dir(
            "purchase_plan",
            src_path=os.path.abspath(str(batches[0])),
            **current.output_kwargs(),
        )
    resolved = os.path.abspath(str(out_dir))
    os.makedirs(resolved, exist_ok=True)
    return resolved


def _prepare_purchase_plan_context(template_paths, batch_paths, out_dir, log, progress):
    """校验输入并冻结任务开始时的供应商代码与主数据快照。"""

    templates = _validate_files(template_paths, "采购计划模板")
    if len(templates) != 1:
        raise ValueError("请选择 1 个采购计划模板文件")
    batches = _validate_files(batch_paths, "辅料清单总表")
    if progress:
        progress(5)
    template_supplier_codes = _read_supplier_codes(templates[0])
    if log:
        log("已读取模板供应商代码 %d 条。" % len(template_supplier_codes))

    catalog_before = material_catalog.load()
    resolver = material_catalog.CatalogResolver(catalog_before)
    supplier_codes = {
        # 主数据库仅提供兜底，模板中的当前维护值对同名供应商具有最终优先级。
        **{str(name): str(code) for name, code in catalog_before.get("suppliers", {}).items()},
        **template_supplier_codes,
    }
    if progress:
        progress(15)
    return _PurchasePlanContext(
        template_path=templates[0],
        batches=batches,
        out_dir=_resolve_purchase_plan_output_dir(batches, out_dir),
        supplier_codes=supplier_codes,
        template_supplier_codes=template_supplier_codes,
        resolver=resolver,
        fill_counts={},
        known_materials=set(catalog_before.get("materials", {})),
        known_suppliers=set(catalog_before.get("suppliers", {})),
    )


def _plan_purchase_outputs(context):
    """为全部批次预留不覆盖历史文件且任务内不冲突的最终路径。"""

    reserved = set()
    plans = []
    for path in context.batches:
        batch_name = _batch_name(path)
        if not batch_name:
            raise ValueError("无法从文件名识别批次号：%s" % os.path.basename(path))
        target = _reserve_output_path(context.out_dir, batch_name, reserved)
        plans.append((path, batch_name, target))
    return plans


def _commit_purchase_outputs(prepared, committed):
    """原子提交全部临时文件，并就地记录已提交路径供异常回滚。"""

    for staged_target, final_target in prepared:
        if os.path.exists(final_target):
            raise FileExistsError("输出文件已被其他任务占用：%s" % os.path.basename(final_target))
        os.replace(staged_target, final_target)
        committed.append(final_target)


def _rollback_purchase_outputs(committed):
    """尽最大努力删除本次已提交文件，不触碰任务开始前存在的历史结果。"""

    for target in committed:
        try:
            os.remove(target)
        except OSError:
            pass


def _convert_purchase_plans(context, plans, log, progress):
    """在同文件系统临时目录中转换全部批次，成功后统一原子提交。"""

    totals = _PurchasePlanTotals()
    staging = tempfile.mkdtemp(prefix=".purchase_plan_", dir=context.out_dir)
    prepared = []
    committed = []
    try:
        for index, (path, batch_name, final_target) in enumerate(plans, start=1):
            staged_target = os.path.join(staging, os.path.basename(final_target))
            converted = _convert_batch(
                context.template_path,
                path,
                staged_target,
                batch_name,
                context.supplier_codes,
                context.known_materials,
                context.known_suppliers,
                log=log,
                resolver=context.resolver,
                fill_counts=context.fill_counts,
            )
            prepared.append((staged_target, final_target))
            totals.add_batch(final_target, converted)
            if log:
                log("批次 %s：生成 %d 行采购计划。" % (batch_name, converted[0]))
            if progress:
                progress(15 + round(index / len(plans) * 80))
        _commit_purchase_outputs(prepared, committed)
        return totals
    except Exception:
        _rollback_purchase_outputs(committed)
        raise
    finally:
        shutil.rmtree(staging, ignore_errors=True)


def _learn_purchase_plan_data(context, totals, log):
    """在文件事务成功后学习供应商代码和材料关系。"""

    material_catalog.log_fill_summary(log, "采购计划", context.fill_counts)
    learned_suppliers = material_catalog.learn_suppliers(
        context.template_supplier_codes, log=log,
    )
    learned_materials = (
        material_catalog.learn_materials(totals.learned_items, log=log)
        if totals.learned_items
        else 0
    )
    if log and learned_suppliers:
        log("主数据档案新增供应商代码 %d 条。" % learned_suppliers)
    if log and learned_materials:
        log("主数据档案新增材料 %d 条。" % learned_materials)


def _log_purchase_plan_totals(totals, log):
    """输出面向管理员的生成、排除和新主数据摘要。"""

    if not log:
        return
    log("已生成 %d 个批次采购计划、共 %d 行。" % (len(totals.files), totals.rows))
    if totals.excluded_original:
        log("已排除 %d 条材料名称含“原厂”的记录。" % totals.excluded_original)
    if totals.new_materials:
        log(
            "发现 %d 个档案外新材料：%s。已自动学习，可在主数据档案中查看和管理。"
            % (len(totals.new_materials), "、".join(sorted(totals.new_materials)[:10]))
        )
    if totals.new_suppliers:
        log(
            "发现 %d 个档案外新供应商：%s。已自动学习，可在主数据档案中查看和管理。"
            % (len(totals.new_suppliers), "、".join(sorted(totals.new_suppliers)[:10]))
        )


def run(
    template_paths,
    batch_paths,
    out_dir=None,
    log=None,
    progress=None,
) -> dict[str, object]:
    """
    按唯一采购计划模板批量生成各批次文件，并在全部提交成功后更新主数据。

    所有批次先写入输出目录内的隐藏临时目录；只有每个文件都转换成功，才逐个原子移动
    到最终路径。提交阶段若发生并发占用或其他异常，会删除本次已经移动的文件，保证
    调用者看到“全部成功”或“没有本次结果”，而不是难以判断的一半成功状态。
    材料名称含“原厂”的记录始终排除。
    """
    context = _prepare_purchase_plan_context(
        template_paths, batch_paths, out_dir, log, progress,
    )
    plans = _plan_purchase_outputs(context)
    totals = _convert_purchase_plans(context, plans, log, progress)
    _learn_purchase_plan_data(context, totals, log)
    if progress:
        progress(100)
    _log_purchase_plan_totals(totals, log)
    return {
        "out_dir": context.out_dir,
        "files": totals.files,
        "generated": len(totals.files),
        "rows": totals.rows,
        "excluded_original_count": totals.excluded_original,
        "new_materials": sorted(totals.new_materials),
        "new_suppliers": sorted(totals.new_suppliers),
    }


def _extra_columns(worksheet, layout) -> dict[str, int]:
    """
    在已识别业务表头的同一行额外定位“实收”和“差异”列。

    通用供应商批次布局不依赖这两个字段，因此仅在差异提取入口按需扩展扫描范围。
    “差异”优先判断，避免名称同时含“实收”和“差异”时被较宽泛的实收规则抢占。
    """
    columns: dict[str, int] = {}
    header_row = int(layout["header_row"])
    for values in worksheet.iter_rows(
        min_row=header_row, max_row=header_row,
        # 常见模板把附加列放在既有业务列右侧，额外读取四列兼顾变体并限制无关扫描。
        max_col=max(int(value) for value in layout["columns"].values()) + 4,
        values_only=True,
    ):
        for index, value in enumerate(values, start=1):
            text = _text(value)
            if "差异" in text and "diff" not in columns:
                columns["diff"] = index
            elif "实收" in text and "actual" not in columns:
                columns["actual"] = index
    return columns


@dataclass
class _DiffAccumulator:
    """累计实收差异业务行、原厂排除数量和主数据补全统计。"""

    resolver: object
    fill_counts: dict[str, int]
    rows: list[list[object]] = field(default_factory=list)
    excluded_original: int = 0


def _diff_row(values, columns, diff_col, actual_col, batch_name, accumulator):
    """解析一行实收差异；非材料、原厂和零差异行直接忽略。"""

    code = _text(_row_value(values, columns.get("code")))
    if not code or code in SUMMARY_IDS:
        return
    item = {
        "code": code,
        "name": _text(_row_value(values, columns.get("name"))),
        "spec": _text(_row_value(values, columns.get("spec"))),
        "unit": _text(_row_value(values, columns.get("unit"))),
    }
    accumulator.resolver.fill_mapping(
        item,
        fields=("name", "spec", "unit"),
        counts=accumulator.fill_counts,
    )
    if _is_original(str(item["name"])):
        accumulator.excluded_original += 1
        return
    difference = _quantity(_row_value(values, diff_col))
    if difference is None or difference == 0:
        return
    accumulator.rows.append([
        batch_name,
        code,
        str(item["name"]),
        item["spec"],
        item["unit"],
        _quantity(_row_value(values, columns.get("qty"))) or 0,
        _quantity(_row_value(values, actual_col)) if actual_col else None,
        difference,
    ])


def _read_diff_batch(path, batch_name, accumulator, log):
    """读取一个批次的非零实收差异，并合并到任务累计器。"""

    common_core.warn_if_uncached(path, log, what="实收差异")
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet, layout = _best_sheet(workbook)
        extra_columns = _extra_columns(worksheet, layout)
        diff_col = extra_columns.get("diff")
        actual_col = extra_columns.get("actual")
        if not diff_col:
            raise ValueError("批次 %s 中未识别到“差异”列" % batch_name)
        columns = layout["columns"]
        max_column = max(
            max(int(value) for value in columns.values()),
            diff_col,
            actual_col or 0,
        )
        for values in worksheet.iter_rows(
            min_row=int(layout["header_row"]) + 1,
            max_col=max_column,
            values_only=True,
        ):
            if all(value is None or _text(value) == "" for value in values):
                continue
            _diff_row(
                values,
                columns,
                diff_col,
                actual_col,
                batch_name,
                accumulator,
            )
    finally:
        workbook.close()


def _resolve_diff_output_dir(batches, out_dir):
    """解析实收差异报告目录，并确保显式目录存在。"""

    if out_dir is None:
        current = settings.get_settings()
        return paths.resolve_output_dir(
            "purchase_plan",
            src_path=os.path.abspath(str(batches[0])),
            **current.output_kwargs(),
        )
    resolved = os.path.abspath(str(out_dir))
    os.makedirs(resolved, exist_ok=True)
    return resolved


def _write_diff_report(out_dir, rows):
    """把结构化实收差异写成带筛选友好样式的独立工作簿。"""

    from openpyxl.styles import Alignment, Border, Font as XlFont, PatternFill, Side

    target = common_core.unique_path(os.path.join(out_dir, "实收差异清单.xlsx"))
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "差异清单"
    thin = Side(style="thin", color="9AA5B1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="FFF3E0")
    headers = ["批次", "材料编号", "材料名称", "规格", "单位", "最终采购数量", "实收", "差异"]
    for column, name in enumerate(headers, start=1):
        cell = worksheet.cell(1, column, name)
        cell.font = XlFont(name="宋体", size=10, bold=True)
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for row_index, values in enumerate(rows, start=2):
        for column, value in enumerate(values, start=1):
            cell = worksheet.cell(row_index, column, value)
            cell.font = XlFont(name="宋体", size=10)
            cell.border = border
    widths = {"A": 13, "B": 16, "C": 30, "D": 22, "E": 8, "F": 13, "G": 10, "H": 10}
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width
    try:
        workbook.save(target)
    finally:
        workbook.close()
    return target


def diff(batch_paths, out_dir=None, log=None, progress=None) -> dict[str, object]:
    """
    汇总一个或多个批次辅料清单中的非零实收差异，生成独立差异清单工作簿。

    每条记录保留批次、材料属性、计划数量、实收和差异；材料属性缺失时由主数据库
    只补空值。没有“差异”列的批次属于不兼容输入并明确报错；没有任何非零差异时
    不生成空报告。原厂材料同采购计划主流程一致地排除。
    """
    batches = _validate_files(batch_paths, "辅料清单总表")
    if progress:
        progress(10)
    accumulator = _DiffAccumulator(
        resolver=material_catalog.CatalogResolver(),
        fill_counts={},
    )
    for index, path in enumerate(batches, start=1):
        batch_name = _batch_name(path)
        if not batch_name:
            raise ValueError("无法从文件名识别批次号：%s" % os.path.basename(path))
        _read_diff_batch(path, batch_name, accumulator, log)
        if progress:
            progress(10 + round(index / len(batches) * 80))
    if not accumulator.rows:
        raise ValueError("所选批次中没有实收差异记录")

    out_dir = _resolve_diff_output_dir(batches, out_dir)
    target = _write_diff_report(out_dir, accumulator.rows)
    material_catalog.log_fill_summary(log, "实收差异清单", accumulator.fill_counts)
    if log:
        log("已生成实收差异清单 %d 条，排除原厂记录 %d 条。" % (
            len(accumulator.rows), accumulator.excluded_original,
        ))
    return {
        "out_dir": out_dir,
        "path": target,
        "rows": len(accumulator.rows),
        "excluded_original_count": accumulator.excluded_original,
    }
