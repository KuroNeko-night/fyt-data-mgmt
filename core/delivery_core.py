# -*- coding: utf-8 -*-
"""
送货计划表制作业务核心
====================
以"物料清单"为主表逐行生成送货计划，按物料号从"物料明细表(含供应商)"查供应商
代码与名称；KD/SUB 列按用户选择统一填写；CASE/CASE托数/班组 可从一张已做好的
往期送货计划(参考表)按物料编码带出；到货/收货等跟单列留空供后续人工填写。

输入（物料清单必需，其余可选；两份主/供表顺序任意，程序自动辨识）：
  · 物料清单：含 物料号 + 数量（可再含中/英文描述）——决定输出的行与需求数；
  · 供应商明细（可选）：含 零部件代码 + 供应商代码 + 供应商名称——供按编码查供应商；
    不提供时，供应商两列留空供人工补填；若物料清单自带供应商列则就地取用。
  · 参考送货计划（可选）：往期做好的送货计划，按物料编码带出 CASE/CASE托数/班组。

输出固定 16 列送货计划，样式与现行模板一致。表头和输入列自动识别，不在前端复制
物料、供应商和参考计划的合并规则；主数据库只补空值，源表已有值始终优先。
"""
import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import paths as _paths
from . import settings as settings_mod
from . import header_detect
from . import shape_detect
from . import common_core as _common  # 公共文本清洗、计算值加载和公式缓存预警。
from . import material_catalog

# 表头关键词映射到业务角色；共享识别器先精确匹配再包含匹配，每列只归一个角色。
HEADER_KEYS = {
    # “下阶物料”是 SAP KD 清单编码列；“下阶物料描述”只应落到名称角色。
    "code":  ["物料号", "物料编码", "零部件代码", "零部件编码", "物料编号", "下阶物料",
              "零件号", "零件编号", "零件代码", "料号", "编码"],
    "cname": ["物料中文描述", "物料英文描述", "下阶物料描述", "物料名称", "零部件名称", "中文描述", "名称", "品名"],
    "ename": ["物料英文描述", "英文描述", "英文名称"],
    "qty":   ["需求数", "需求数量", "计划数量", "数量"],
    "sup_code": ["供应商代码", "供应商编码", "供方代码"],
    "sup_name": ["供应商名称", "供应商信息", "供方名称", "供应商"],
    "attr":  ["属性", "KD/SUB", "KD/SUB属性"],
}

# 输出列顺序是现行业务模板协议，公式列和人工跟单列均依赖这些固定位置。
OUT_HEADERS = ["序号", "物料编码", "物料名称", "供应商代码", "供应商信息", "KD/SUB",
               "需求数", "计划到货日期", "实际收货数", "实际收货日期", "第二次到货日期",
               "剩余未收数", "CASE", "CASE托数", "班组", "备注"]


def norm_code(v):
    """
    规范化物料编码的数值尾缀、全半角和不可见字符，同时保留文本前导零。

    Excel 可能把纯数字编码读成整数浮点，此时去掉无意义的 ``.0``；文本编码交给
    ``clean_str``，不执行整数化，以免 ``00123`` 的业务前导零被破坏。
    """
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return _common.clean_str(v)


def cell_text(v):
    """把单元格值转为展示文本；空值为空串，整数浮点去除尾随 ``.0``。"""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


# ---------------------------------------------------------------------------
# 表头/列自动识别
# ---------------------------------------------------------------------------
# “委外供应商属性”等列虽然含“供应商”，实际不是供应商代码/名称，包含匹配时必须排除。
_EXCLUDE_CONTAINS = {"sup_code": ["属性"], "sup_name": ["属性"]}


def detect_layout(ws, scan_rows=12, log=None):
    """调用共享表头识别器，在工作表前若干行定位送货业务列。

    物料编码是唯一必要角色；数量和供应商列是否存在由后续分类逻辑判断。此处仅传入
    本业务别名和误匹配排除规则，保持公共识别算法为单一事实来源。
    """
    return header_detect.detect_layout(
        ws, HEADER_KEYS, require=("code",), scan_rows=scan_rows,
        exclude_contains=_EXCLUDE_CONTAINS, log=log)


# 数据形态画像只作人工确认兜底，角色顺序接近典型列序，布尔值表示是否为必要列。
_SHAPE_PROFILE = [
    ("code", shape_detect.CODE, True),
    ("cname", shape_detect.TEXT, False),
    ("qty", shape_detect.NUMBER, False),
    ("sup_code", shape_detect.CODE, False),
    ("sup_name", shape_detect.TEXT, False),
]


def detect_layout_or_shape(ws, scan_rows=12, log=None):
    """先识别表头文字，失败后再按数据形态推断列角色。

    返回 (header_row, col_map, source):source 为 "header"(表头识别)或
    ``shape`` 结果只说明数据分布相似，必须在界面交由用户核对，不得直接静默落盘。
    """
    hr, col = detect_layout(ws, scan_rows=scan_rows, log=log)
    if hr:
        return hr, col, "header"
    hr2, col2, _conf = shape_detect.detect_by_shape(
        ws, _SHAPE_PROFILE, scan_rows=scan_rows, log=log)
    if hr2:
        return hr2, col2, "shape"
    return None, {}, None


def list_sheets(path):
    """返回 xlsx/xlsm 页签名供界面选择；不支持格式或读取失败时返回空列表。"""
    ext = os.path.splitext(path)[1].lower()
    if ext not in (".xlsx", ".xlsm"):
        return []  # openpyxl 不读取旧 xls，此辅助函数不承担多格式兼容。
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()
    except Exception:
        # 页签预览是辅助能力，错误由正式加载阶段提供更具体提示，此处不阻断选文件界面。
        return []


def _select_delivery_sheet(workbook, requested_sheet, log):
    """选择送货输入页签并返回已识别的表头布局。

    显式指定页签时严格使用该页，名称错误由 openpyxl 抛出；自动模式先检查首表，仅在
    首表无法识别编码列时顺序寻找后续页签，以兼容封面、说明页位于最前面的工作簿。
    """
    if requested_sheet:
        worksheet = workbook[requested_sheet]
        header_row, columns = detect_layout(worksheet, log=log)
        return worksheet, header_row, columns

    first_name = workbook.sheetnames[0]
    worksheet = workbook[first_name]
    header_row, columns = detect_layout(worksheet, log=log)
    if header_row or len(workbook.sheetnames) == 1:
        return worksheet, header_row, columns

    for sheet_name in workbook.sheetnames[1:]:
        candidate = workbook[sheet_name]
        candidate_header, candidate_columns = detect_layout(candidate, log=log)
        if not candidate_header:
            continue
        if log:
            log("· 多子表: 首表《%s》无有效表头, 改读《%s》" % (
                first_name, sheet_name,
            ))
        return candidate, candidate_header, candidate_columns
    return worksheet, header_row, columns


def _delivery_optional_value(worksheet, row, columns, role):
    """读取一个可选业务角色的单元格；模板缺少该列时返回 ``None``。"""
    column = columns.get(role)
    return worksheet.cell(row, column).value if column else None


def _is_delivery_summary_row(name):
    """判断名称是否表示合计、小计或总计等非物料汇总行。"""
    return isinstance(name, str) and any(
        token in name for token in ("合计", "小计", "总计")
    )


def _delivery_row(worksheet, row, columns):
    """把一个工作表数据行投影为送货业务记录；无效行返回 ``None``。"""
    code = worksheet.cell(row, columns["code"]).value
    if code is None or not norm_code(code):
        # 编码是供应商和参考计划关联主键，空编码行不能可靠进入输出。
        return None
    chinese_name = _delivery_optional_value(worksheet, row, columns, "cname")
    if _is_delivery_summary_row(chinese_name):
        # 部分汇总行仍带编码或数量，必须按名称排除，避免生成伪物料记录。
        return None
    return {
        "r": row,
        "code": code,
        "cname": chinese_name,
        "ename": _delivery_optional_value(worksheet, row, columns, "ename"),
        "qty": _delivery_optional_value(worksheet, row, columns, "qty"),
        "sup_code": _delivery_optional_value(worksheet, row, columns, "sup_code"),
        "sup_name": _delivery_optional_value(worksheet, row, columns, "sup_name"),
        "attr": _delivery_optional_value(worksheet, row, columns, "attr"),
    }


def _read_delivery_rows(worksheet, header_row, columns):
    """顺序读取表头后的有效送货业务行，并保持源表原始顺序。"""
    rows = []
    last_row = worksheet.max_row or header_row
    for row in range(header_row + 1, last_row + 1):
        record = _delivery_row(worksheet, row, columns)
        if record is not None:
            rows.append(record)
    return rows


def load_sheet(path, sheet=None, log=None):
    """读取一个送货相关工作表，自动识别列并返回有效业务行和布局。

    rows: [{r, code, cname, ename, qty, sup_code, sup_name, attr}]，按角色缺省为 None。
    未指定页签时先读首张表，若无法识别再顺序寻找其他页签，兼容封面或说明页占首位。
    空编码行和名称含合计/小计/总计的汇总行会被过滤；其他字段缺失保持 ``None``，
    留给主数据库补全或后续分类判断。
    """
    # 只需要计算值，不访问公式和样式，因此使用跳过透视缓存的公共加载器。
    wb = _common.load_data_only(path)
    try:
        ws, header_row, col = _select_delivery_sheet(wb, sheet, log)
        if not header_row:
            raise ValueError("未能在 %s / %s 中识别表头（需含“物料号/编码”列）"
                             % (os.path.basename(path), ws.title))
        rows = _read_delivery_rows(ws, header_row, col)
        return rows, {"sheet": ws.title, "header_row": header_row, "col": col}
    finally:
        wb.close()


def analyze(path, sheet=None, log=None):
    """只读预检文件的页签、表头、角色列和数据规模，不生成输出文件。

    供 UI 在用户选文件后立刻反馈"能否认出各列 / 是靠形态兜底(需核对)",
    以便在点"生成"前就发现列错位。返回 dict:
      ok           - 是否识别成功(拿到必需列)
      sheet        - 命中的子表名
      header_row   - 表头所在行(1-based)
      roles        - {角色: 列号}
      source       - "header"(表头文字识别) / "shape"(形态兜底,需人工核对) / None
      n_rows       - 表头之后的数据行数(粗计,不排合计行)
      sheets       - 该簿全部子表名
      error        - 失败原因(ok=False 时)

    预检允许形态兜底，以便界面展示候选列并要求人工确认；正式 ``load_sheet`` 仍只接受
    明确表头识别结果，防止低可信映射绕过复核直接用于生成。
    """
    res = {"ok": False, "sheet": "", "header_row": 0, "roles": {},
           "source": None, "n_rows": 0, "sheets": [], "error": ""}
    try:
        res["sheets"] = list_sheets(path)
        # 形态探测依赖随机 cell 和真实 max_row；错误 dimension 文件在只读模式会被截成一行。
        wb = _common.load_data_only(path)
        try:
            # 指定页签无效时预检首表并在返回中告知实际页签，不让底层 KeyError 泄漏到界面。
            ws = wb[sheet] if (sheet and sheet in wb.sheetnames) else wb[wb.sheetnames[0]]
            hr, col, src = detect_layout_or_shape(ws, log=log)
            res["sheet"] = ws.title
            if not hr:
                res["error"] = "未能识别表头(需含“物料号/编码”列)"
                return res
            res["ok"] = True
            res["header_row"] = hr
            res["roles"] = dict(col)
            res["source"] = src
            res["n_rows"] = max(0, (ws.max_row or hr) - hr)
        finally:
            wb.close()
    except Exception as e:
        # analyze 以 ok/error 表达失败，便于文件一选中就展示问题而不终止整个界面。
        res["error"] = str(e)
    return res


def _has_supplier(layout):
    """判断布局是否至少含供应商代码或名称，可否作为供应商信息来源。"""
    return "sup_code" in layout["col"] or "sup_name" in layout["col"]


def _has_qty(layout):
    """判断布局是否含需求数量列，这是物料清单主表的重要特征。"""
    return "qty" in layout["col"]


def classify(lay_a, lay_b, n_a=0, n_b=0, log=None):
    """辨识两份顺序任意的输入中哪份是物料主表、哪份是供应商来源。

    规则(按可靠度依次):
      1) 仅一份带供应商列 -> 它作供应商来源, 另一份作主表(最可靠);
      2) 两份都带供应商列 -> 用"含数量列"区分: 有数量的那份作主表;
      3) 数量列也无法区分 -> 按行数多者作主表, 并 log 明确警告(而非静默择 A);
      4) 两份都不带供应商列 -> 报错提示"未找到供应商列"。

    返回两个 ``a``/``b`` 键。只有最后的行数规则属于低可信降级，因此必须写日志提醒
    用户核对；供应商列和数量列能明确区分时不产生多余警告。
    """
    a_sup, b_sup = _has_supplier(lay_a), _has_supplier(lay_b)
    if not a_sup and not b_sup:
        # 传入两份文件代表用户期望其中一份提供供应商，两份都没有时应视为选错文件。
        raise ValueError("两份表都未找到供应商列(供应商代码/名称)，无法确定供应商来源，"
                         "请确认是否选错文件。")
    if a_sup and not b_sup:
        return "b", "a"  # B 无供应商，作为主表；A 是供应商来源。
    if b_sup and not a_sup:
        return "a", "b"  # A 无供应商，作为主表；B 是供应商来源。
    # 两份都含供应商时，用数量列区分；主物料清单通常必须提供需求数。
    a_qty, b_qty = _has_qty(lay_a), _has_qty(lay_b)
    if a_qty and not b_qty:
        return "a", "b"
    if b_qty and not a_qty:
        return "b", "a"
    # 最后才按行数降级，行数相等稳定选择 A；日志明确说明此结果需要人工核对。
    if log:
        log("⚠ 两份表都含供应商列且都含/都不含数量列，无法可靠区分主表与供应商表，"
            "已按行数多者(%d vs %d)作主表，请核对结果。" % (n_a, n_b))
    return ("a", "b") if n_a >= n_b else ("b", "a")


# ---------------------------------------------------------------------------
# 输出样式与固定列协议
# ---------------------------------------------------------------------------
_HEAD_FILL = PatternFill("solid", fgColor="FFBDD7EE")     # 新模板表头浅蓝
_HEAD_FONT = Font(name="微软雅黑", size=11, bold=True, color="FF000000")
_DATA_FONT = Font(name="微软雅黑", size=11)                # 新模板数据字体
_TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="FF000000")
_THIN = Side(style="thin", color="FF000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
# 前五列按内容设置不同宽度，其余跟单列保持统一宽度。
_WIDTHS = [13, 15.83, 39.91, 13, 45.75, 13, 13, 13, 13, 13, 13, 13, 13, 13, 13, 13]

# 关键输出列使用 1 基常量，公式引用和字段位置保持与 OUT_HEADERS 同步。
_C_QTY = 7        # 需求数
_C_RECV = 9       # 实际收货数
_C_LEFT = 12      # 剩余未收数（= 实际收货数 - 需求数）
_C_CASE = 13      # CASE
_C_CASE_QTY = 14  # CASE托数
_C_TEAM = 15      # 班组

_TITLE_ROW_H = 39   # 标题行高
_HEAD_ROW_H = 33    # 表头行高
_DATA_ROW_H = 22    # 数据行统一行高（与参考送货计划一致）


def build_plan_sheet(ws, master_rows, sup_map, order_type=None,
                     case_map=None, log=None, report_missing=True):
    """把物料主表写成固定 16 列送货计划，并返回产出与匹配统计。

    sup_map    : 归一编码 -> (供应商代码, 供应商名称)。
    order_type : "SUB" / "KD"，统一填入 KD/SUB 列；None 则留空。
    case_map   : 归一编码 -> (CASE, CASE托数, 班组)，来自参考送货计划；None 则不填。
    report_missing : 是否在日志里汇报"未匹配供应商"的编码。未提供供应商来源时
                     每行都会记入 missing，此时应传 False 以免刷无意义警告。

    版式：第 1 行合并 A1:P1 作标题（留空，仅套样式）；第 2 行表头；数据自第 3 行起。
    数据行严格保持物料清单原顺序，不按供应商重排，以便用户回查源表。剩余未收数沿用
    现行模板口径，写入“实际收货数 - 需求数”公式；到货、收货与备注列保持空白供跟单。
    返回 (写入行数, 未匹配供应商的编码列表, CASE 命中数)。
    """
    ncol = len(OUT_HEADERS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    t = ws.cell(1, 1, None)  # 标题内容由用户后续填写，程序只保留模板视觉区域。
    t.font = _TITLE_FONT
    t.alignment = _CENTER
    ws.row_dimensions[1].height = _TITLE_ROW_H
    for c in range(1, ncol + 1):  # 第二行固定写入业务表头。
        cell = ws.cell(2, c, OUT_HEADERS[c - 1])
        cell.font = _HEAD_FONT
        cell.fill = _HEAD_FILL
        cell.alignment = _CENTER
        cell.border = _BORDER
    ws.row_dimensions[2].height = _HEAD_ROW_H

    case_map = case_map or {}
    recs, missing, hit_case = _prepare_plan_records(master_rows, sup_map, case_map)
    _write_plan_rows(ws, recs, order_type, start_row=3)

    for c, w in enumerate(_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    if log and missing and report_missing:
        # 最多展示前八个缺失编码，保留可操作线索且避免大批缺失淹没任务日志。
        log("有 %d 个物料在供应商明细中未找到供应商，已留空：%s%s"
            % (len(missing), "、".join(missing[:8]), " 等" if len(missing) > 8 else ""))
    return len(recs), missing, hit_case


def _prepare_plan_records(master_rows, sup_map, case_map):
    """连接供应商和参考计划信息，保持物料清单顺序并统计缺失与命中。"""

    records = []
    missing = []
    hit_case = 0
    for row in master_rows:
        code = norm_code(row["code"])
        supplier = sup_map.get(code)
        if supplier is None:
            # 映射不存在才算缺失；映射存在但某字段为空交给人工补写，不重复报警。
            missing.append(code)
        case_info = case_map.get(code)
        if case_info:
            hit_case += 1
        supplier_code, supplier_name = supplier if supplier else (None, None)
        records.append({
            "row": row,
            "supplier_code": supplier_code,
            "supplier_name": supplier_name,
            "case_info": case_info,
        })
    return records, missing, hit_case


def _write_plan_rows(ws, records, order_type, *, start_row):
    """把已连接的内部记录写入固定 16 列协议，并生成剩余未收数公式。"""

    for index, record in enumerate(records, start=1):
        row_number = start_row + index - 1
        source = record["row"]
        case_value, case_quantity, team = record["case_info"] or (None, None, None)
        values = [
            index, source["code"], source.get("cname"),
            record["supplier_code"], record["supplier_name"], order_type or None,
            source.get("qty"), None, None, None, None,
            # 通过列常量生成公式，表头位置变化时无需查找散落的字母列号。
            "=%s%d-%s%d" % (
                get_column_letter(_C_RECV), row_number,
                get_column_letter(_C_QTY), row_number,
            ),
            case_value, case_quantity, team, None,
        ]
        for column, value in enumerate(values, start=1):
            cell = ws.cell(row_number, column, value)
            cell.font = _DATA_FONT
            cell.alignment = _CENTER
            cell.border = _BORDER
        ws.row_dimensions[row_number].height = _DATA_ROW_H


def build_supplier_map(sup_rows, log=None):
    """从供应商明细建立 ``归一物料编码 -> (供应商代码, 供应商名称)`` 映射。

    同一编码出现不同供应商时保留首条有效记录并统计冲突。这一策略保持文件从上到下的
    人工优先顺序，同时通过日志公开歧义，避免后出现的重复行悄悄改变整个送货计划。
    """
    m = {}
    conflicts = 0
    for row in sup_rows:
        code = norm_code(row["code"])
        if not code:
            continue
        # 空字符串统一为 None，便于区分“映射存在但缺某字段”和“编码完全未映射”。
        pair = (cell_text(row.get("sup_code")) or None,
                cell_text(row.get("sup_name")) or None)
        if code in m:
            # 重复且内容相同不算冲突；不同内容只计数，不覆盖首次选择。
            if m[code] != pair:
                conflicts += 1
            continue
        m[code] = pair
    if log and conflicts:
        log("注意：供应商明细中有 %d 个物料存在多个不同供应商，已取首个。" % conflicts)
    return m


def _complete_from_catalog(master_rows, sup_map, resolver, fill_counts=None):
    """
    使用主数据库补全物料名称、供应商名称及供应商代码，现有业务表值始终优先。

    物料名称写回 ``master_rows`` 的内存投影；供应商信息写回 ``sup_map``，供输出阶段
    统一连接。主数据解析器只返回缺失字段的补充值，不覆盖源表或供应商明细中的明确值。
    """
    for row in master_rows:
        code = norm_code(row.get("code"))
        supplier_code, supplier_name = sup_map.get(code, (None, None))
        # 把当前名称和供应商交给解析器，只有为空的字段才会出现在 additions 中。
        additions = resolver.complete_material(
            code,
            {"name": row.get("cname"), "supplier": supplier_name},
            fields=("name", "supplier"),
            counts=fill_counts,
        )
        if "name" in additions:
            row["cname"] = additions["name"]
        if "supplier" in additions:
            supplier_name = additions["supplier"]
        # 供应商名称确定后再补代码，支持供应商来自主数据而非输入明细的情况。
        supplier_code = resolver.complete_supplier_code(
            supplier_name, supplier_code, counts=fill_counts) or None
        if supplier_code or supplier_name:
            # 至少有一个有效字段才建立映射，避免把所有编码都变成 (None, None) 假命中。
            sup_map[code] = (supplier_code, supplier_name or None)


# ---------------------------------------------------------------------------
# 参考送货计划：按物料编码带出 CASE / CASE托数 / 班组
# ---------------------------------------------------------------------------
_REF_KEYS = {"code": ["物料编码", "物料号", "零部件代码", "编码"],
             "case": ["CASE"], "case_qty": ["CASE托数", "托数"],
             "team": ["班组"]}


def _match_ref_header(ws, scan_rows=8):
    """在参考计划前若干行定位包含物料编码和 CASE 信息的明细表头。

    参考表常有多个 sheet（如透视 Sheet2 + 明细"零件到货计划"）。透视汇总表虽也含
    "班组"、且"计数项:物料编码"会被物料编码子串命中，但它没有 CASE 列——故以 CASE 为
    判据可稳妥排除透视表。code 匹配优先精确表头，避免误取"计数项:物料编码"之类。
    CASE 是区分明细表和透视汇总页的硬条件，同时要求班组或 CASE 托数至少一项。
    编码角色仅接受精确表头，防止“计数项:物料编码”等透视统计列误命中。
    """
    for hr in range(1, min(scan_rows, ws.max_row) + 1):
        col = _reference_columns(ws, hr)
        # CASE 加班组/托数组合能有效排除只有计数和分组字段的透视汇总页。
        if "code" in col and "case" in col and ("team" in col or "case_qty" in col):
            return hr, col
    return None, {}


def _reference_columns(ws, header_row):
    """识别参考计划单行表头，编码严格匹配，其余字段允许带业务前缀。"""

    columns = {}
    for column in range(1, ws.max_column + 1):
        text = cell_text(ws.cell(header_row, column).value)
        if not text:
            continue
        for role, keys in _REF_KEYS.items():
            if role in columns:
                continue
            # “计数项:物料编码”不能作为明细编码列，编码角色只接受精确表头。
            if role == "code":
                matched = any(key == text for key in keys)
            else:
                matched = any(key == text or key in text for key in keys)
            if matched:
                columns[role] = column
    return columns


def _find_reference_detail_sheet(workbook):
    """按工作簿顺序返回首个符合参考计划明细结构的页签及布局。"""
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        header_row, columns = _match_ref_header(worksheet)
        if header_row:
            return worksheet, header_row, columns
    return None


def _reference_case_record(worksheet, row, columns):
    """读取一条非空参考记录，返回编码及 CASE、托数、班组；无有效补充值时返回空。"""
    code = norm_code(worksheet.cell(row, columns["code"]).value)
    if not code:
        return None
    case_value = cell_text(worksheet.cell(row, columns["case"]).value)
    case_quantity = _delivery_optional_value(worksheet, row, columns, "case_qty")
    team = cell_text(_delivery_optional_value(worksheet, row, columns, "team"))
    if not case_value and case_quantity is None and not team:
        # 空补充行不应制造“已命中参考资料”的假象，后续同编码有效行仍可被采用。
        return None
    return code, (case_value or None, case_quantity, team or None)


def _read_reference_case_map(worksheet, header_row, columns):
    """顺序读取参考明细，同一编码保留第一条具有有效补充信息的记录。"""
    mapping = {}
    last_row = worksheet.max_row or header_row
    for row in range(header_row + 1, last_row + 1):
        record = _reference_case_record(worksheet, row, columns)
        if record is None:
            continue
        code, values = record
        if code not in mapping:
            # 靠前记录通常是当前有效安排；重复项不应被表尾历史或汇总内容覆盖。
            mapping[code] = values
    return mapping


def build_case_map(path, log=None):
    """从可选参考计划读取 ``物料编码 -> (CASE, CASE托数, 班组)`` 映射。

    自动跳过透视/汇总页，采用第一个符合明细结构的页签；同一编码多次出现时保留首值。
    参考表是增强输入，读取失败或无匹配页签时记录提示并返回空字典，不阻断主计划生成。
    """
    def _lg(msg):
        """把参考计划读取提示转交调用方；未提供日志回调时保持静默。"""
        if log:
            log(msg)
    if not path:
        return {}
    try:
        # 参考表只读取计算值，不需要样式和透视对象，使用公共加速加载器。
        wb = _common.load_data_only(path)
    except Exception as e:
        _lg("参考送货计划读取失败，已跳过 CASE/班组：%s" % e)
        return {}
    try:
        detail = _find_reference_detail_sheet(wb)
        if detail is None:
            _lg("参考送货计划里未找到含「物料编码 + CASE/班组」的表，已跳过。")
            return {}
        ws, header_row, columns = detail
        mapping = _read_reference_case_map(ws, header_row, columns)
        _lg("参考送货计划：从工作表「%s」读到 %d 条 CASE/班组 记录。" % (
            ws.title, len(mapping),
        ))
        return mapping
    finally:
        wb.close()


def run(file_a, file_b=None, sheet_a=None, sheet_b=None, out_dir=None, log=None,
        order_type=None, ref_plan=None):
    """执行送货计划生成，支持可选供应商明细和可选往期参考计划。

    供应商明细为可选：
      · 传了 file_b —— 两份输入顺序任意，自动辨识主表/供应商来源（原行为）；
      · 未传 file_b —— file_a 即物料清单主表；若它自带供应商列则就地取用，
        否则供应商代码/名称两列留空供人工补填（不视为"未匹配"报警）。

    两份输入都提供时允许顺序任意，由 ``classify`` 辨识主表和供应商来源；只提供
    ``file_a`` 时它直接作为主表，并优先使用自带供应商列，再由主数据库补空值。
    ``order_type`` 规范化为大写后统一写入 KD/SUB 列；``ref_plan`` 仅用于补充 CASE、
    CASE 托数和班组。返回输出路径、行数、匹配统计和实际采用的输入来源。
    """
    def _lg(msg):
        """统一转发主流程日志，使核心不依赖具体界面。"""
        if log:
            log(msg)

    # 此时尚未辨识哪份是主表，因此两份都检查数量公式缓存，避免需求数静默变空。
    from . import common_core
    common_core.warn_if_uncached(file_a, _lg, sheet_a, what="需求数/数量")
    if file_b:
        common_core.warn_if_uncached(file_b, _lg, sheet_b, what="需求数/数量")

    rows_a, lay_a = load_sheet(file_a, sheet_a, log=_lg)
    if file_b:
        rows_b, lay_b = load_sheet(file_b, sheet_b, log=_lg)
        master_key, sup_key = classify(lay_a, lay_b, len(rows_a), len(rows_b), log=_lg)
        # 用分类键映射完整三元组，保持主表行、布局和文件路径始终成组切换。
        pack = {"a": (rows_a, lay_a, file_a), "b": (rows_b, lay_b, file_b)}
        master_rows, _lm, master_file = pack[master_key]
        sup_rows, _ls, sup_file = pack[sup_key]
        _lg("主表(物料清单)：%s —— %d 行" % (os.path.basename(master_file), len(master_rows)))
        _lg("供应商来源：%s —— %d 行" % (os.path.basename(sup_file), len(sup_rows)))
        sup_map = build_supplier_map(sup_rows, log=_lg)
    else:
        # 单文件模式允许主表自带供应商；完全没有时不是错误，主数据和人工维护可继续补充。
        master_rows, master_file, sup_file = rows_a, file_a, ""
        _lg("主表(物料清单)：%s —— %d 行" % (os.path.basename(master_file), len(master_rows)))
        if _has_supplier(lay_a):
            sup_map = build_supplier_map(rows_a, log=_lg)
            _lg("未单独提供供应商明细，已从物料清单自带的供应商列带出。")
        else:
            sup_map = {}
            _lg("未提供供应商明细，将优先从主数据库补全，仍缺失的内容可稍后人工填写。")
    # 输入解析结束后统一补全，确保源文件或供应商表中的现有信息始终优先。
    resolver = material_catalog.CatalogResolver()
    fill_counts: dict[str, int] = {}
    _complete_from_catalog(master_rows, sup_map, resolver, fill_counts)
    material_catalog.log_fill_summary(_lg, "送货计划", fill_counts)
    # 只要主数据或输入提供任一映射，就启用匹配统计；完全无来源时不把所有行误报缺失。
    supplier_used = bool(sup_map)

    # 空白保持 None，使 Excel 单元格真正留空；非空值统一大写以稳定展示。
    ot = (order_type or "").strip().upper() or None
    if ot:
        _lg("KD/SUB 列统一填：%s" % ot)
    case_map = build_case_map(ref_plan, log=_lg) if ref_plan else {}

    if out_dir is None:
        # 默认走统一输出设置；显式目录保持既有调用方选择并确保目录存在。
        st = settings_mod.get_settings()
        out_dir = _paths.resolve_output_dir("delivery", **st.output_kwargs())
    else:
        os.makedirs(out_dir, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    n, missing, hit_case = build_plan_sheet(
        ws, master_rows, sup_map, order_type=ot, case_map=case_map, log=_lg,
        report_missing=supplier_used)
    if supplier_used:
        matched = n - len(missing)
        _lg("已生成 %d 行，供应商匹配 %d / %d。" % (n, matched, n))
    else:
        # 完全无供应商来源时两列按设计留空，不能把这种模式误解为 n 条匹配失败。
        missing, matched = [], 0
        _lg("已生成 %d 行（供应商两列留空）。" % n)
    if case_map:
        _lg("CASE/班组 按物料编码匹配 %d / %d 行。" % (hit_case, n))

    # 文件名遵循现行业务约定；若 Excel 占用同名结果，保存异常会给出明确关闭提示。
    plan_path = os.path.join(out_dir, "送货计划.xlsx")
    try:
        wb.save(plan_path)
    except PermissionError:
        raise PermissionError("无法保存 %s —— 请先在 Excel 里关闭该文件后重试" % plan_path)
    _lg("已生成送货计划：%s" % plan_path)
    return {"plan_path": plan_path, "out_dir": out_dir, "rows": n,
            "matched": matched, "missing": missing, "order_type": ot,
            "case_hit": hit_case, "case_used": bool(case_map),
            "supplier_used": supplier_used,
            "master_file": master_file, "supplier_file": sup_file}
