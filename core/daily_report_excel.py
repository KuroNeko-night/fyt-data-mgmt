# -*- coding: utf-8 -*-
"""基于日清快照生成管理层 Excel 报告。

工作表只消费结构化快照，不重新读取原始业务表格。所有样式、列宽、空态和下载统计
集中在此模块，快照聚合规则的变化不会与导出格式相互缠绕。
"""

from __future__ import annotations

import os
from typing import Any, Mapping

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import common_core, paths, settings
from .daily_report_snapshot import _integer, _text, _validate_date



_THIN = Side(style="thin", color="D5DDE8")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_TITLE_FONT = Font(name="微软雅黑", size=18, bold=True, color="17365D")
_SECTION_FONT = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
_HEAD_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
_CELL_FONT = Font(name="微软雅黑", size=10, color="24344D")
_SECTION_FILL = PatternFill("solid", fgColor="17365D")
_HEAD_FILL = PatternFill("solid", fgColor="2F6BFF")
_SOFT_FILL = PatternFill("solid", fgColor="EAF1FF")
_WARNING_FILL = PatternFill("solid", fgColor="FFF4E7")


def _style_table(sheet, header_row: int, max_row: int, max_col: int) -> None:
    """为日清导出表统一应用表头、边框和自动换行样式。

    显式限制 ``max_col``，避免空态占位行或工作表历史格式把样式扩散到无关列；数据行
    使用顶部对齐和自动换行，以容纳问题说明、整改措施等长文本。
    """
    for cell in sheet[header_row]:
        if cell.column > max_col:
            break
        cell.font = _HEAD_FONT
        cell.fill = _HEAD_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER
    for row in sheet.iter_rows(min_row=header_row + 1, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.font = _CELL_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = _BORDER


def _set_widths(sheet, widths: list[int]) -> None:
    """按业务预设顺序设置工作表列宽。

    不根据内容自动估宽，是为了避免一条异常长备注把整张表撑到难以浏览；各写入函数
    根据字段语义为名称、说明和数字列提供稳定宽度。
    """
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _create_table_sheet(workbook, title: str, headers: list[str]):
    """创建无网格线的明细工作表并写入第一行表头。"""
    sheet = workbook.create_sheet(title)
    sheet.sheet_view.showGridLines = False
    sheet.append(headers)
    return sheet


def _finish_table_sheet(
    sheet,
    headers: list[str],
    widths: list[int],
    *,
    filter_last_row: int | None = None,
) -> None:
    """统一完成明细表的样式、冻结窗格、筛选范围和列宽设置。

    空数据工作表也会写入一行可读占位信息，因此筛选范围至少包含第一行。带合计行的
    工作表可通过 ``filter_last_row`` 排除合计，防止用户筛选时把总计混入普通记录。
    """
    _style_table(sheet, 1, sheet.max_row, len(headers))
    sheet.freeze_panes = "A2"  # 冻结表头，长明细滚动后仍能辨认字段。
    last_row = sheet.max_row if filter_last_row is None else filter_last_row
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, last_row)}"
    _set_widths(sheet, widths)


def _write_summary_sheet(
    workbook,
    snapshot: Mapping[str, Any],
    arrival: Mapping[str, Any],
    workshop: Mapping[str, Any],
    safety_checks: Mapping[str, Any],
    production_ledger: Mapping[str, Any],
) -> None:
    """写入供管理层快速阅读的日清概览和统计口径。

    概览只放当天最关键的六项指标，详细记录由后续工作表承载。统计口径随报告写入，
    让离线文件也能说明数据按什么日期、状态和来源计算，而不依赖网页帮助文案。
    """
    report_date = _validate_date(snapshot.get("date"))
    definitions = snapshot.get("definitions") if isinstance(snapshot.get("definitions"), Mapping) else {}
    summary = workbook.active  # 复用 Workbook 默认工作表，避免生成无意义的 Sheet 页。
    summary.title = "日清概览"
    summary.sheet_view.showGridLines = False
    summary.merge_cells("A1:F1")
    summary["A1"] = f"峰运通日清报告 · {report_date}"
    summary["A1"].font = _TITLE_FONT
    summary["A1"].alignment = Alignment(vertical="center")
    summary.row_dimensions[1].height = 32
    summary["A2"] = "生成时间"
    summary["B2"] = _text(snapshot.get("generated_at"))
    summary["A3"] = "业务时区"
    summary["B3"] = _text(snapshot.get("timezone")) or "Asia/Shanghai"
    summary.merge_cells("A5:F5")
    summary["A5"] = "当日核心指标"
    summary["A5"].font = _SECTION_FONT
    summary["A5"].fill = _SECTION_FILL
    summary["A5"].alignment = Alignment(horizontal="left")
    indicators = [
        ("到料完成率", f"{float(arrival.get('completion_rate', 0) or 0):.1f}%"),
        ("未收料类数", arrival.get("missing_categories", 0)),
        ("现场问题", workshop.get("issue_count", 0)),
        ("安全不合格", safety_checks.get("unqualified_count", 0)),
        ("月度生产订单", production_ledger.get("formal_total", 0)),
        (
            "待关闭/待发运",
            _integer(production_ledger.get("formal_pending"))
            + _integer(production_ledger.get("sporadic_pending")),
        ),
    ]
    # 指标标题与数值上下排列，每项占一列，便于管理层横向扫读。
    for index, (label, value) in enumerate(indicators, start=1):
        summary.cell(6, index, label)
        summary.cell(7, index, value)
        summary.cell(6, index).font = Font(name="微软雅黑", size=10, bold=True, color="24344D")
        summary.cell(6, index).fill = _SOFT_FILL
        summary.cell(7, index).font = Font(name="微软雅黑", size=16, bold=True, color="17365D")
        for row_index in (6, 7):
            summary.cell(row_index, index).border = _BORDER
            summary.cell(row_index, index).alignment = Alignment(horizontal="center", vertical="center")
    summary.merge_cells("A9:F9")
    summary["A9"] = "统计口径"
    summary["A9"].font = _SECTION_FONT
    summary["A9"].fill = _SECTION_FILL
    summary["A10"] = "到料"
    summary["B10"] = _text(definitions.get("arrival"))
    summary["A11"] = "现场问题"
    summary["B11"] = _text(definitions.get("workshop"))
    summary.merge_cells("B10:F10")
    summary.merge_cells("B11:F11")
    for row in range(10, 12):
        for column in range(1, 7):
            summary.cell(row, column).font = _CELL_FONT
            summary.cell(row, column).border = _BORDER
            summary.cell(row, column).alignment = Alignment(vertical="top", wrap_text=True)
    # 数量不闭合属于需要人工确认的数据质量问题，醒目标记但不阻止整份日报导出。
    if _integer(arrival.get("invalid_batch_count")):
        summary.merge_cells("A13:F13")
        summary["A13"] = f"提示：{arrival.get('invalid_batch_count')} 个批次数量关系需要人工核对。"
        summary["A13"].fill = _WARNING_FILL
        summary["A13"].font = Font(name="微软雅黑", size=10, bold=True, color="24344D")
        summary["A13"].alignment = Alignment(wrap_text=True)
    _set_widths(summary, [16, 20, 20, 20, 16, 16])


def _write_arrival_sheets(workbook, batches: list[object]) -> int:
    """写入到料批次汇总和未到物料明细，返回实际明细行数。

    批次表用于快速比较完成率，未到物料表用于落实到具体物料、供应商和数量差额。
    返回行数供 ``run`` 生成结构化结果摘要，不能使用工作表最大行数，因为空态也会写入
    一行提示文字。
    """
    arrival_headers = ["批次", "提交人", "任务", "完成时间", "主料总类数", "已到货", "未收料", "完成率", "缺料供应商汇总"]
    arrival_sheet = _create_table_sheet(workbook, "每日到料", arrival_headers)
    for batch in batches:
        if not isinstance(batch, Mapping):
            continue
        arrival_sheet.append([
            batch.get("batch_no"), batch.get("uploader"), batch.get("job_title"),
            batch.get("completed_at"), batch.get("total_count"), batch.get("arrived_count"),
            # Excel 百分比单元格需要保存 0～1 的数值，快照中则使用 0～100 便于前端展示。
            batch.get("missing_count"), float(batch.get("completion_rate", 0) or 0) / 100,
            "；".join(
                f"{item.get('supplier')} 缺口 {item.get('shortage_quantity')}"
                for item in batch.get("supplier_distribution", [])
                if isinstance(item, Mapping)
            ),
        ])
    # 保留空态工作表而不是直接省略，离线阅读者能明确区分“无数据”和“导出遗漏”。
    if arrival_sheet.max_row == 1:
        arrival_sheet.append(["当日没有到料任务结果"])
    _finish_table_sheet(
        arrival_sheet,
        arrival_headers,
        [18, 14, 24, 24, 14, 12, 12, 12, 36],
    )
    for row in range(2, arrival_sheet.max_row + 1):
        arrival_sheet.cell(row, 8).number_format = "0.0%"

    missing_headers = ["批次", "提交人", "物料编码", "物料名称", "供应商", "需求数", "已收数", "缺口数"]
    missing_sheet = _create_table_sheet(workbook, "未到物料", missing_headers)
    missing_detail_count = 0
    for batch in batches:
        if not isinstance(batch, Mapping):
            continue
        materials = batch.get("missing_materials") if isinstance(batch.get("missing_materials"), list) else []
        for material in materials:
            if not isinstance(material, Mapping):
                continue
            missing_sheet.append([
                batch.get("batch_no"), batch.get("uploader"),
                material.get("material_code"), material.get("material_name"),
                material.get("supplier"), material.get("demand_quantity"),
                material.get("received_quantity"), material.get("shortage_quantity"),
            ])
            missing_detail_count += 1
    if missing_sheet.max_row == 1:
        missing_sheet.append(["当日任务没有可用的未到物料明细"])
    _finish_table_sheet(missing_sheet, missing_headers, [18, 14, 18, 30, 28, 12, 12, 12])
    return missing_detail_count


def _write_safety_and_workshop_sheets(
    workbook,
    safety_checks: Mapping[str, Any],
    issues: list[object],
) -> None:
    """写入安全检查和现场问题的可追踪明细。

    安全检查使用管理员最后上传版本中的记录；现场问题写入固定超集列，五类模板不适用
    的字段保持为空。这样既遵守每类问题的录入白名单，也能让跨类型报表保持稳定列序。
    """
    safety_headers = ["检查类别", "检查项目", "安全标准要求", "检查结果", "问题描述", "整改措施", "责任人", "图片数"]
    safety_sheet = _create_table_sheet(workbook, "安全检查日报", safety_headers)
    records = safety_checks.get("records") if isinstance(safety_checks.get("records"), list) else []
    for record in records:
        if not isinstance(record, Mapping):
            continue
        safety_sheet.append([
            record.get("category"), record.get("check_item"), record.get("standard"), record.get("result"),
            record.get("problem_description"), record.get("corrective_action"), record.get("owner"),
            # 报告记录图片数量而非本地路径，避免导出文件泄露服务端存储结构。
            len(record.get("images") if isinstance(record.get("images"), list) else []),
        ])
    if safety_sheet.max_row == 1:
        safety_sheet.append(["当日没有上传安全检查日报"])
    _finish_table_sheet(safety_sheet, safety_headers, [16, 24, 42, 14, 42, 42, 16, 10])

    # 字段顺序与规范问题表的业务阅读顺序保持一致；不同类型不会被要求填写全部字段。
    workshop_headers = [
        "提交时间", "提交人", "问题原因", "主要负责人", "次要负责人", "备注", "图片数", "问题分类", "严重程度",
        "问题源", "车型", "国家", "批次号", "班组", "物料编码", "物料名称", "原因分析", "纠正措施", "责任方", "外检责任人", "发现人",
        "问题等级", "故障数量", "故障类别", "完成时间", "是否复发", "记录次数", "发生时间", "处理时间", "责任人", "更新人", "承运商", "供应商", "模板处理状态",
        "问题状态", "解决情况", "解决时间", "解决人", "最后更新时间",
    ]
    workshop_sheet = _create_table_sheet(workbook, "现场问题", workshop_headers)
    for issue in issues:
        if not isinstance(issue, Mapping):
            continue
        workshop_sheet.append([
            issue.get("created_at"), issue.get("uploader"), issue.get("cause"),
            issue.get("primary_owner"), issue.get("secondary_owner"), issue.get("notes"),
            issue.get("image_count", 0), issue.get("category", "other"), issue.get("severity", "normal"),
            issue.get("issue_source", ""), issue.get("model", ""), issue.get("country", ""), issue.get("batch_no", ""), issue.get("team", ""),
            issue.get("material_code", ""), issue.get("material_name", ""), issue.get("cause_analysis", ""),
            issue.get("corrective_action", ""), issue.get("responsibility_party", ""), issue.get("external_inspection_owner", ""), issue.get("discoverer", ""),
            issue.get("issue_level", ""), issue.get("quantity", ""), issue.get("issue_type", ""),
            issue.get("completion_date", ""), issue.get("recurring", ""), issue.get("record_count", ""), issue.get("happened_at", ""),
            issue.get("handling_time", ""), issue.get("responsible_person", ""), issue.get("updated_by_name", ""), issue.get("carrier", ""),
            issue.get("supplier", ""), issue.get("tracking_status", ""),
            "已解决" if issue.get("resolution_status") == "resolved" else "处理中",
            issue.get("resolution_note", ""), issue.get("resolved_at", ""), issue.get("resolved_by_name", ""), issue.get("updated_at", ""),
        ])
    if workshop_sheet.max_row == 1:
        workshop_sheet.append(["当日没有已发布的现场问题"])
    _finish_table_sheet(
        workshop_sheet,
        workshop_headers,
        [24, 14, 42, 16, 16, 38, 10, 14, 12, 14, 14, 14, 18, 14, 18, 28, 34, 34, 16, 18, 14, 12, 12, 16, 16, 12, 12, 18, 18, 16, 16, 18, 24, 14, 12, 38, 22, 16, 22],
    )


def _write_attendance_sheets(workbook, attendance: Mapping[str, Any]) -> None:
    """按参会人员、生产班组和汇总三个层次写入考勤数据。

    参会人员按姓名逐人记录；生产人员只按班组和班次记录编制、出勤及备注。两个模型
    不混在一张明细表中，最后的考勤汇总仅在有可用汇总数据时生成。
    """
    people = attendance.get("people") if isinstance(attendance.get("people"), list) else []
    production_groups = attendance.get("production_groups") if isinstance(attendance.get("production_groups"), list) else []
    if people:
        attendance_headers = ["人员类型", "姓名", "单位/班组", "班次", "是否出勤", "状态", "原因", "更新时间"]
        attendance_sheet = _create_table_sheet(workbook, "每日考勤", attendance_headers)
        for person in people:
            if not isinstance(person, Mapping):
                continue
            attendance_sheet.append([
                person.get("person_type"), person.get("name"), person.get("unit"), person.get("shift"),
                "出勤" if person.get("present") else "缺勤", person.get("status"),
                person.get("reason"), person.get("updated_at"),
            ])
        _finish_table_sheet(attendance_sheet, attendance_headers, [14, 16, 22, 12, 12, 14, 34, 24])
    if production_groups:
        production_headers = ["生产班组", "班次", "人员编制", "出勤", "差异", "备注", "更新时间"]
        production_sheet = _create_table_sheet(workbook, "生产出勤", production_headers)
        for group in production_groups:
            if not isinstance(group, Mapping):
                continue
            production_sheet.append([
                group.get("group_name"), group.get("shift_name"), group.get("staffing_count", 0),
                group.get("attendance_count", 0), group.get("difference", 0), group.get("note"), group.get("updated_at"),
            ])
        # 合计行是展示结果，不属于可筛选的班组记录，完成样式时会从筛选范围排除。
        production_sheet.append([
            "合计", "", attendance.get("production_staffing_count", attendance.get("production_total", 0)),
            attendance.get("production_present_count", 0), attendance.get("production_difference", 0), "", "",
        ])
        _finish_table_sheet(
            production_sheet,
            production_headers,
            [24, 14, 14, 12, 12, 48, 24],
            filter_last_row=max(2, production_sheet.max_row - 1),
        )
    unit_summary = attendance.get("unit_summary") if isinstance(attendance.get("unit_summary"), list) else []
    if unit_summary:
        summary_headers = ["人员类型", "单位/班组", "班次", "人员编制", "出勤", "差异", "缺勤原因"]
        summary_sheet = _create_table_sheet(workbook, "考勤汇总", summary_headers)
        for item in unit_summary:
            if not isinstance(item, Mapping):
                continue
            summary_sheet.append([
                item.get("person_type"), item.get("unit"), item.get("shift"), item.get("total", 0),
                item.get("present", 0), item.get("difference", 0), "；".join(item.get("reasons", [])),
            ])
        _finish_table_sheet(summary_sheet, summary_headers, [14, 22, 14, 14, 12, 12, 44])


def _write_brief_and_plan_sheets(workbook, snapshot: Mapping[str, Any]) -> None:
    """按需写入重点事项、通报和生产计划上传摘要。

    这两类工作表没有数据时不创建，避免报告出现大量空页；生产计划这里只记录文件级
    来源摘要，订单、发运、缺件等已解析信息由专门的台账工作表完整展示。
    """
    brief_items = snapshot.get("brief_items") if isinstance(snapshot.get("brief_items"), list) else []
    if brief_items:
        brief_headers = ["类别", "单位", "责任人", "标题", "事项说明", "完成日期", "进展", "状态"]
        brief_sheet = _create_table_sheet(workbook, "重点事项与通报", brief_headers)
        for item in brief_items:
            if not isinstance(item, Mapping):
                continue
            brief_sheet.append([
                item.get("category"), item.get("unit"), item.get("owner"), item.get("title"),
                item.get("description"), item.get("due_date"), item.get("progress"), item.get("status"),
            ])
        _finish_table_sheet(brief_sheet, brief_headers, [14, 18, 16, 28, 48, 14, 20, 14])

    production_plans = snapshot.get("production_plans") if isinstance(snapshot.get("production_plans"), list) else []
    if production_plans:
        plan_headers = ["日期", "文件名", "上传者", "上传时间", "工作表数", "总行数"]
        plan_sheet = _create_table_sheet(workbook, "生产计划", plan_headers)
        for plan in production_plans:
            if not isinstance(plan, Mapping):
                continue
            plan_summary = plan.get("summary") if isinstance(plan.get("summary"), Mapping) else {}
            plan_sheet.append([
                plan.get("report_date"), plan.get("original_name"), plan.get("uploaded_by_name"),
                plan.get("created_at"), plan_summary.get("sheet_count", 0), plan_summary.get("row_count", 0),
            ])
        _finish_table_sheet(plan_sheet, plan_headers, [14, 36, 16, 24, 14, 14])


def _write_formal_ledger_sheet(workbook, rows: list) -> None:
    """写入正式订单台账工作表。"""
    headers = ["月份", "订单号", "国家", "类型", "数量", "发运完成时间", "状态", "未完成缺件", "未完成危包", "备注"]
    sheet = _create_table_sheet(workbook, "月度生产订单台账", headers)
    for order in rows:
        if not isinstance(order, Mapping):
            continue
        sheet.append([
            order.get("month"), order.get("order_no"), order.get("country"), order.get("order_type"),
            order.get("quantity"), order.get("shipment_date"), order.get("status"),
            order.get("outstanding_missing_count", 0), order.get("outstanding_hazardous_count", 0), order.get("note"),
        ])
    if sheet.max_row == 1:
        sheet.append(["本月没有生产订单台账数据"])
    _finish_table_sheet(sheet, headers, [12, 22, 16, 14, 12, 18, 16, 14, 14, 42])


def _write_ledger_detail_sheet(workbook, sheet_name: str, key: str, label: str, rows: list) -> None:
    """写入缺件或危包明细工作表；两类结构一致，按 ``label`` 区分。"""
    headers = ["订单号", "类别", "物料号", "零件名称", "数量", "发运订单号", "发运完成时间", "是否完成"]
    sheet = _create_table_sheet(workbook, sheet_name, headers)
    for item in rows:
        if not isinstance(item, Mapping):
            continue
        sheet.append([
            item.get("order_no"), label, item.get("material_code"), item.get("material_name"),
            item.get("quantity"), item.get("shipment_order_no"), item.get("shipment_date"),
            "已完成" if item.get("completed") else "未完成",
        ])
    if sheet.max_row == 1:
        sheet.append([f"本月没有{label}记录"])
    _finish_table_sheet(sheet, headers, [22, 12, 18, 30, 12, 24, 18, 12])


def _write_sporadic_ledger_sheet(workbook, rows: list) -> None:
    """写入零星订单明细工作表。"""
    headers = ["月份", "订单号", "运输方式", "国家", "订单类型", "托数", "体积(CBM)", "发运时间", "司机车牌", "司机姓名", "司机电话", "状态", "备注"]
    sheet = _create_table_sheet(workbook, "零星订单明细", headers)
    for order in rows:
        if not isinstance(order, Mapping):
            continue
        sheet.append([
            order.get("month"), order.get("order_no"), order.get("transport_mode"), order.get("country"),
            order.get("order_type"), order.get("pallet_count"), order.get("volume_cbm"),
            "、".join(order.get("shipment_dates") or []), order.get("driver_plate"), order.get("driver_name"),
            order.get("driver_phone"), order.get("status"), order.get("note"),
        ])
    if sheet.max_row == 1:
        sheet.append(["本月没有零星订单数据"])
    _finish_table_sheet(sheet, headers, [12, 22, 14, 14, 14, 10, 14, 22, 16, 14, 20, 14, 38])


def _write_production_ledger_sheets(workbook, production_ledger: Mapping[str, Any]) -> None:
    """写入正式订单、缺件/危包和零星订单明细。

    正式订单与零星订单字段差异较大，分别建表；订单内嵌的缺件、危包也拆为独立明细，
    使管理人员可以直接筛选未完成项，而不必再次打开原始生产计划。
    """
    formal_orders = production_ledger.get("formal_orders") if isinstance(production_ledger.get("formal_orders"), list) else []
    _write_formal_ledger_sheet(workbook, formal_orders)
    # 缺件与危包结构一致，复用同一段写入逻辑，同时保留独立工作表便于业务筛选。
    for sheet_name, key, label in (
        ("订单缺件明细", "missing_parts", "缺件"),
        ("订单危包明细", "hazardous_packages", "危包"),
    ):
        rows = production_ledger.get(key) if isinstance(production_ledger.get(key), list) else []
        _write_ledger_detail_sheet(workbook, sheet_name, key, label, rows)
    sporadic_orders = production_ledger.get("sporadic_orders") if isinstance(production_ledger.get("sporadic_orders"), list) else []
    _write_sporadic_ledger_sheet(workbook, sporadic_orders)


def run(snapshot: Mapping[str, Any], out_dir: str | None = None, log=None) -> dict[str, object]:
    """根据已构建快照生成日清 Excel，并返回统一输出信息。

    未显式指定目录时使用全局输出设置；Web/Tauri 任务传入隔离运行目录时则只在该目录
    内创建文件。所有工作表都读取同一个快照，不再解析原始上传文件。保存失败会关闭
    工作簿资源并向上抛出可操作的中文错误，成功后返回文件路径和关键记录数量。
    """
    report_date = _validate_date(snapshot.get("date"))
    if out_dir is None:
        current = settings.get_settings()
        out_dir = paths.resolve_output_dir("daily_report", **current.output_kwargs())
    else:
        # 调用方传入的任务目录可能尚未创建；转绝对路径可使返回结果不依赖当前工作目录。
        out_dir = os.path.abspath(str(out_dir))
        os.makedirs(out_dir, exist_ok=True)
    # 不覆盖用户已打开或已留存的同名报告，冲突时由公共工具追加安全序号。
    out_file = common_core.unique_path(os.path.join(out_dir, f"日清报告-{report_date}.xlsx"))

    arrival = snapshot.get("arrival") if isinstance(snapshot.get("arrival"), Mapping) else {}
    workshop = snapshot.get("workshop") if isinstance(snapshot.get("workshop"), Mapping) else {}
    safety_checks = snapshot.get("safety_checks") if isinstance(snapshot.get("safety_checks"), Mapping) else {}
    production_ledger = snapshot.get("production_ledger") if isinstance(snapshot.get("production_ledger"), Mapping) else {}
    batches = arrival.get("batches") if isinstance(arrival.get("batches"), list) else []
    issues = workshop.get("issues") if isinstance(workshop.get("issues"), list) else []

    workbook = Workbook()
    attendance = snapshot.get("attendance") if isinstance(snapshot.get("attendance"), Mapping) else {}
    # 各写入函数按业务主题组织，避免导出协议继续堆积在单个长函数中。
    _write_summary_sheet(workbook, snapshot, arrival, workshop, safety_checks, production_ledger)
    missing_detail_count = _write_arrival_sheets(workbook, batches)
    _write_safety_and_workshop_sheets(workbook, safety_checks, issues)
    _write_attendance_sheets(workbook, attendance)
    _write_brief_and_plan_sheets(workbook, snapshot)
    _write_production_ledger_sheets(workbook, production_ledger)

    try:
        workbook.save(out_file)
    except PermissionError as exc:
        # Windows 上最常见原因是同名 Excel 正被占用，转换成用户可直接处理的提示。
        raise PermissionError(f"无法保存 {out_file}，请先关闭正在打开的同名文件") from exc
    finally:
        # 无论保存成功或异常都释放 openpyxl 资源，避免批量任务长期占用文件句柄。
        workbook.close()
    if log:
        log(f"已生成日清报告：{os.path.basename(out_file)}")
    return {
        "out_file": out_file,
        "out_dir": out_dir,
        "date": report_date,
        "arrival_batches": len(batches),
        "arrival_missing_materials": missing_detail_count,
        "workshop_issues": len(issues),
    }
