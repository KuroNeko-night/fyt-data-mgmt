# -*- coding: utf-8 -*-
"""
发票与采购供应商匹配核心
========================
汇总一份或多份发票台账中的“销售方名称 + 价税合计”，再与采购明细中的供应商集合
做名称交集和差集，生成按供应商聚合的票货匹配表。结果分为正常、无票采购（采购有
供应商但发票无销售方）和有发票无采购三种状态。

该功能只回答“某供应商是否同时出现在票、货两侧”以及发票侧总金额，不尝试把每张
发票匹配到具体采购行。原因是当前采购明细只稳定提供供应商，采购数量对账也没有
统一金额字段，缺少可靠的金额级关联键。供应商名称采用去首尾空白后的精确匹配，
不会擅自做简称或模糊归并，以免把不同公司错误合并；名称标准化应由主数据维护完成。

所有输入只读取首个页签的计算值，公式缓存缺失会提前提示。输出目录遵循统一设置，
文件名带秒级时间戳并通过 ``unique_path`` 防止覆盖同秒产生的既有报告。
"""
from __future__ import annotations

import os
import re
from collections import defaultdict
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from . import common_core, paths, settings

EXCEL_SUFFIXES = {".xlsx", ".xlsm"}


def _text(value) -> str:
    """将单元格值转换为去首尾空白的名称文本，空值统一为空串。"""
    return "" if value is None else str(value).strip()


def _number(value) -> float | None:
    """把数值或带千分位的数值文本转为浮点数，非数值返回 ``None``。

    布尔值虽然是 Python 的整数子类，但不具备金额语义，因此显式排除。解析失败不在
    此处报错，由读取函数按零金额保留销售方记录，使供应商存在性仍能参与匹配。
    """
    if isinstance(value, bool) or value is None:
        return None
    try:
        return float(str(value).replace(",", ""))
    except (TypeError, ValueError):
        return None


def _find_header_row(worksheet, keyword: str, scan_rows: int = 6) -> tuple[int, dict[str, int]]:
    """在页签顶部定位发票表头或采购供应商表头。

    ``keyword`` 为空时要求同一行同时出现销售方和价税合计；非空时寻找包含该词、
    且不含“编码/代码”的供应商名称列，避免把供应商编码误当名称。每个角色只采用
    首个命中列，返回 1 基行列号；扫描范围内没有完整角色则返回 ``(0, {})``。
    """
    for row_index in range(1, min(scan_rows, worksheet.max_row or scan_rows) + 1):
        columns: dict[str, int] = {}
        for cell in worksheet[row_index]:
            text = _text(cell.value)
            if not text:
                continue
            # if/elif 保证一个含多个关键词的异常表头不会同时占用多个角色。
            if "销售方" in text and "seller" not in columns:
                columns["seller"] = cell.column
            elif "价税合计" in text and "total" not in columns:
                columns["total"] = cell.column
            elif keyword and keyword in text and "供应商" not in columns and "编码" not in text and "代码" not in text:
                columns["supplier"] = cell.column
        if keyword and "supplier" in columns:
            return row_index, columns
        if not keyword and "seller" in columns and "total" in columns:
            return row_index, columns
    return 0, {}


def _read_invoices(path: str, log=None) -> list[dict[str, object]]:
    """读取首个页签的有效销售方和价税合计记录。

    合计行按销售方文本排除；金额无法解析时记为零，但仍保留销售方，使“有发票”
    状态不会因为一个坏金额单元格而消失。多份台账中的同名销售方由 ``match`` 汇总。
    """
    common_core.warn_if_uncached(path, log, what="价税合计")
    # 仅顺序读取计算值，不修改样式，read_only 可控制大型发票台账的内存占用。
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        header_row, columns = _find_header_row(worksheet, "")
        if not header_row:
            raise ValueError("未在 %s 中识别到“销售方名称”和“价税合计”列" % os.path.basename(path))
        invoices: list[dict[str, object]] = []
        for values in worksheet.iter_rows(
            min_row=header_row + 1, max_col=max(columns.values()), values_only=True):
            # values 是从第一列到最大必要列的元组，按 1 基列号换算为 0 基索引。
            seller = _text(values[columns["seller"] - 1] if columns["seller"] <= len(values) else None)
            if not seller or "合计" in seller:
                continue
            total = _number(values[columns["total"] - 1] if columns["total"] <= len(values) else None)
            # None 使用零兜底只影响金额汇总，不改变该销售方已经出现在发票侧的事实。
            invoices.append({"seller": seller, "total": total or 0})
        return invoices
    finally:
        workbook.close()


def _read_purchase(path: str, log=None) -> list[dict[str, object]]:
    """读取首个页签中所有非空、非合计的供应商名称。

    兼容供应商批次表和采购计划导入输出；这里只保留名称，不读取数量或金额，因为
    后续匹配按供应商集合判断存在性，重复供应商会在入口转为集合去重。
    """
    common_core.warn_if_uncached(path, log, what="供应商")
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        header_row, columns = _find_header_row(worksheet, "供应商", scan_rows=12)
        if not header_row:
            raise ValueError("未在 %s 中识别到“供应商”列" % os.path.basename(path))
        supplier_column = columns["supplier"]
        suppliers: list[dict[str, object]] = []
        for values in worksheet.iter_rows(
            min_row=header_row + 1, max_col=supplier_column, values_only=True):
            supplier = _text(values[supplier_column - 1] if supplier_column <= len(values) else None)
            if not supplier or "合计" in supplier:
                continue
            suppliers.append({"supplier": supplier})
        return suppliers
    finally:
        workbook.close()


def _validate_inputs(invoice_paths, purchase_paths) -> tuple[list[str], list[str]]:
    """规范化并一次性校验两侧输入路径。"""
    invoices = [os.path.abspath(str(value)) for value in (invoice_paths or []) if str(value).strip()]  # 去掉空选择并固定绝对路径。
    purchases = [os.path.abspath(str(value)) for value in (purchase_paths or []) if str(value).strip()]
    if not invoices or not purchases:
        raise ValueError("请同时选择发票台账与采购明细文件")  # 两侧缺一不可，避免生成误导性空报告。
    for path in invoices + purchases:
        if not os.path.isfile(path):
            raise FileNotFoundError("找不到文件：%s" % path)  # 读取前先完整检查，失败不会留下半成品。
        if os.path.splitext(path)[1].lower() not in EXCEL_SUFFIXES:
            raise ValueError("仅支持 xlsx 或 xlsm 文件：%s" % os.path.basename(path))
    return invoices, purchases


def _collect_supplier_totals(invoice_paths, purchase_paths, log, progress):
    """读取发票金额和采购供应商集合，返回匹配所需的两个事实集合。"""
    invoice_totals: dict[str, float] = defaultdict(float)
    for path in invoice_paths:
        for item in _read_invoices(path, log=log):
            invoice_totals[str(item["seller"])] += float(item["total"])  # 同名销售方跨文件累计为一行。
    purchase_suppliers: set[str] = set()
    for path in purchase_paths:
        for item in _read_purchase(path, log=log):
            purchase_suppliers.add(str(item["supplier"]))  # 集合去重，避免重复采购行放大供应商数量。
    if progress:
        progress(60)  # 两侧读取结束后进入报告构建阶段。
    if not invoice_totals:
        raise ValueError("发票台账中没有有效记录")
    if not purchase_suppliers:
        raise ValueError("采购明细中没有识别到供应商")
    return invoice_totals, purchase_suppliers


def _supplier_sets(invoice_totals, purchase_suppliers):
    """生成稳定排序的匹配、无票采购和有票无采购集合。"""
    invoice_suppliers = set(invoice_totals)  # 发票侧供应商来自累计金额字典键。
    return (
        sorted(invoice_suppliers & purchase_suppliers),
        sorted(purchase_suppliers - invoice_suppliers),
        sorted(invoice_suppliers - purchase_suppliers),
    )


def _write_match_report(out_dir, invoice_totals, both, no_invoice, no_purchase):
    """按稳定模板写出匹配报告并返回文件路径。"""
    thin = Side(style="thin", color="9AA5B1")  # 统一边框对象，减少工作簿样式数量。
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="EAF1FF")
    warn_fill = PatternFill("solid", fgColor="FFF3E0")
    head_font = Font(name="宋体", size=10, bold=True)
    cell_font = Font(name="宋体", size=10)
    workbook = openpyxl.Workbook()
    try:
        summary = workbook.active
        summary.title = "票货匹配"
        headers = ["供应商", "发票价税合计（元）", "采购明细供应商", "状态"]
        for column, width in {"A": 22, "B": 18, "C": 16, "D": 22}.items():
            summary.column_dimensions[column].width = width  # 先固定列宽，避免中文标题被截断。
        for column, name in enumerate(headers, start=1):
            cell = summary.cell(1, column, name)
            cell.font = head_font; cell.fill = head_fill
            cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = border
        row_index = 2
        for supplier, amount, marker, status, warning in (
            *[(name, invoice_totals[name], "✓", "正常", False) for name in both],
            *[(name, 0, "✓", "无票采购", True) for name in no_invoice],
            *[(name, invoice_totals[name], "—", "有发票无采购", True) for name in no_purchase],
        ):
            for column, value in enumerate([supplier, round(amount, 2), marker, status], start=1):
                cell = summary.cell(row_index, column, value)
                cell.font = cell_font; cell.border = border
                if warning:
                    cell.fill = warn_fill  # 异常行统一底色，便于打开报告快速定位。
            row_index += 1
        target = common_core.unique_path(os.path.join(out_dir, "票货匹配表_%s.xlsx" % datetime.now().strftime("%Y%m%d_%H%M%S")))
        workbook.save(target)  # 保存完成后再返回，调用方可立即提供下载。
        return target
    finally:
        workbook.close()  # 即使保存失败也释放 openpyxl 句柄。


def match(invoice_paths, purchase_paths, out_dir=None, log=None, progress=None) -> dict[str, object]:
    """汇总票货两侧供应商并生成匹配状态报告。

    两类文件都至少需要一份，且只接受 xlsx/xlsm。发票侧按销售方累计全部价税合计，
    采购侧只保留唯一供应商集合；随后通过集合交集和双向差集得到三种状态。进度在
    输入校验后报 10%，读取汇总后报 60%，报告保存完成由调用结束隐含为完成。

    返回输出路径、三类数量和两类异常供应商名称列表，前端可直接展示摘要，不需要
    再读取生成的 Excel。
    """
    invoices_input, purchase_input = _validate_inputs(invoice_paths, purchase_paths)
    if progress:
        progress(10)  # 输入全部合法后再报告校验阶段完成。
    invoice_totals, purchase_suppliers = _collect_supplier_totals(invoices_input, purchase_input, log, progress)
    both, no_invoice, no_purchase = _supplier_sets(invoice_totals, purchase_suppliers)

    if out_dir is None:
        current = settings.get_settings()
        # “输出到源旁”模式以第一份发票台账定位，其他模式仍由统一设置决定。
        out_dir = paths.resolve_output_dir(
            "invoice_match", src_path=os.path.abspath(invoices_input[0]),
            **current.output_kwargs())
    else:
        out_dir = os.path.abspath(str(out_dir))
        os.makedirs(out_dir, exist_ok=True)

    target = _write_match_report(out_dir, invoice_totals, both, no_invoice, no_purchase)

    if log:
        log("匹配完成：正常 %d 家、无票采购 %d 家、有发票无采购 %d 家。"
            % (len(both), len(no_invoice), len(no_purchase)))
        if no_invoice:
            log("无票采购（有采购明细但无发票）：%s" % "、".join(no_invoice))
        if no_purchase:
            log("有发票无采购：%s" % "、".join(no_purchase))
    return {
        "out_dir": out_dir, "path": target,
        "matched": len(both), "no_invoice": len(no_invoice),
        "no_purchase": len(no_purchase),
        "no_invoice_suppliers": no_invoice,
        "no_purchase_suppliers": no_purchase,
    }
