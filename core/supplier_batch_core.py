# -*- coding: utf-8 -*-
"""
供应商批次表业务核心。

模块负责识别一个或多个辅料批次清单，依次使用当前清单、历史采购明细和主数据库
补全材料的供应商归属，再把“供应商选择”和“每批交付日期”交给人工复核。正式执行
阶段按供应商分别生成采购明细，任何材料名称中含“原厂”的记录都会在进入复核结果前
排除。桌面端和 Web 端都只调用本模块，不在界面层重复表格识别或输出规则。
"""
from __future__ import annotations

import math
import os
import re
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from . import common_core, material_catalog, paths, settings


# 输出表头保持与既有采购明细模板一致，避免供应商收到的文件结构随前端变化。
HEADERS = ["材料编号", "材料名称", "规格", "单位", "求和项:最终采购数量", "供应商"]
# 这些值来自空白、公式错误或人工占位，不能被当成真实供应商进入选择列表。
INVALID_SUPPLIERS = {"", "0", "#N/A", "#VALUE!", "未匹配", "无"}
EXCEL_SUFFIXES = {".xlsx", ".xlsm"}
# 行数较多时按批次拆分页签，减少单个页签连续多个区块造成的查找和打印困难。
SPLIT_THRESHOLD = 30

# 每个业务角色按“更具体、更可信”到“较宽泛”的顺序列出别名，顺序会参与布局评分。
_ROLE_ALIASES = {
    "code": ("材料编号", "物料编号", "材料编码", "物料编码", "料号"),
    "name": ("材料名称", "物料名称", "品名"),
    "spec": ("规格", "型号"),
    "unit": ("单位",),
    "qty": ("最终采购数量", "采购数量", "需求数量", "数量"),
    "supplier": ("供应商名称", "供应商信息", "供应商"),
}


def _text(value) -> str:
    """把任意单元格值安全转换为去除首尾空白的文本，空单元格统一返回空串。"""
    if value is None:
        return ""
    return str(value).strip()


def _header(value) -> str:
    """移除表头中的空白和中英文冒号，使“求和项:最终采购数量”等变体可统一匹配。"""
    return re.sub(r"[\s:：]", "", _text(value))


def _supplier(value) -> str:
    """清洗供应商名称，并把 Excel 错误值、数字占位和未匹配标记归一为空值。"""
    text = _text(value)
    # 错误值使用大写比较以兼容 #n/a 等写法；正常中文名称保留用户原始大小写和标点。
    return "" if text.upper() in INVALID_SUPPLIERS else text


def _quantity(value) -> float | None:
    """解析去除千位逗号后的有限数值；是否大于零由调用方按业务场景判断。"""
    if isinstance(value, bool) or value is None:
        # Python 中 bool 是 int 的子类，必须显式排除，不能把 TRUE/FALSE 当作 1/0 数量。
        return None
    try:
        number = float(str(value).replace(",", ""))
    except (TypeError, ValueError):
        return None
    # NaN 和无穷值可以通过 float 转换，但写入采购数量没有业务意义，因此视为无效。
    return number if math.isfinite(number) else None


def _is_original(name: object) -> bool:
    """判断材料名称是否含“原厂”；先去除空白以覆盖“（ 原 厂 ）”等录入变体。"""
    return "原厂" in re.sub(r"\s+", "", _text(name))


def _validate_files(values, label: str) -> list[str]:
    """
    校验一组 Excel 输入文件并返回绝对路径列表。

    绝对化能让 Web 子进程、桌面桥接和直接脚本运行使用相同定位结果。校验在打开
    工作簿前完成，以便一次给出清晰的“缺文件/格式不支持”错误，而不是暴露底层库异常。
    """
    result = [os.path.abspath(str(value)) for value in (values or []) if str(value).strip()]
    if not result:
        raise ValueError("%s不能为空" % label)
    for path in result:
        if not os.path.isfile(path):
            raise FileNotFoundError("找不到文件：%s" % path)
        # 只接受与模板一致的工作簿格式，避免把 CSV 或旧版 xls 误当可解析对象传给 openpyxl。
        if Path(path).suffix.lower() not in EXCEL_SUFFIXES:
            raise ValueError("%s仅支持 xlsx 或 xlsm 文件：%s" % (label, os.path.basename(path)))
    return result


def _record_header_match(columns, strengths, role, column_index, strength):
    """在单个角色中只保留当前扫描行里的最强别名匹配列。"""
    if strength > strengths.get(role, 0):
        strengths[role] = strength
        columns[role] = column_index


def _row_layout(values):
    """扫描一行表头，返回字段列映射和每个角色的别名强度。"""
    columns: dict[str, int] = {}
    strengths: dict[str, int] = {}
    for column_index, value in enumerate(values, start=1):
        text = _header(value)
        if not text:
            continue
        for role, aliases in _ROLE_ALIASES.items():
            for alias_index, alias in enumerate(aliases):
                if alias in text:
                    # 越靠前的别名越具体，分值越高；同一角色只保留当前行中最强匹配列。
                    _record_header_match(
                        columns, strengths, role, column_index,
                        len(aliases) - alias_index,
                    )
                    break
    return columns, strengths


def _detect_layout(worksheet) -> dict[str, object] | None:
    """
    扫描工作表前十二行，返回字段覆盖完整且别名匹配质量最高的表头布局。

    材料编号、材料名称、单位和数量是生成采购明细的最低条件；规格和供应商允许缺失，
    后续可由主数据库或历史表补齐。候选行按别名在列表中的优先级累积分，含供应商列
    额外加权，使完整业务表优先于同一页中可能出现的简略说明表。
    """
    best = None
    # 只看顶部有限区域，避免大表扫描成本，也避免正文内容拼出一组“伪表头”。
    scan_rows = min(12, worksheet.max_row or 12)
    for row_index, values in enumerate(
        worksheet.iter_rows(min_row=1, max_row=scan_rows, values_only=True), start=1
    ):
        columns, strengths = _row_layout(values)
        # 缺少任一必要角色时即使分数较高也不能安全生成采购明细。
        if not {"code", "name", "unit", "qty"}.issubset(columns):
            continue
        # 供应商列对本功能尤其关键，给予足够加分，优先采用无需外部映射的完整表头。
        score = sum(strengths.values()) + (20 if "supplier" in columns else 0)
        candidate = {"header_row": row_index, "columns": columns, "score": score}
        if best is None or score > int(best["score"]):
            # 分数相同保留更靠前的候选行，使同一文件多次分析得到稳定结果。
            best = candidate
    return best


def _best_sheet(workbook) -> tuple[object, dict[str, object]]:
    """在工作簿全部页签中选择布局评分最高的业务表，评分相同时采用更靠前页签。"""
    candidates = []
    for index, worksheet in enumerate(workbook.worksheets):
        layout = _detect_layout(worksheet)
        if layout:
            # 使用 -index 让 max 在同分时选择工作簿中更靠前的页签，结果稳定且符合阅读顺序。
            candidates.append((int(layout["score"]), -index, worksheet, layout))
    if not candidates:
        raise ValueError("未识别到材料编号、材料名称、单位和最终采购数量列")
    _, _, worksheet, layout = max(candidates, key=lambda item: (item[0], item[1]))
    return worksheet, layout


def _iter_data_rows(worksheet, layout: dict[str, object]):
    """
    从识别到的表头下一行开始顺序产生数据行，连续五十个空白行后提前停止。

    少量空行可能是批次分组或排版间隔，不能遇到第一行空白就结束；但部分模板的
    ``max_row`` 会因历史格式残留而非常大，连续空白阈值可避免遍历整张空白尾部。
    """
    columns = layout["columns"]
    # 只读取最右侧业务列以内的数据，减少 read_only 模式下无关格式列的解析量。
    max_column = max(int(value) for value in columns.values())
    blank_run = 0
    for values in worksheet.iter_rows(
        min_row=int(layout["header_row"]) + 1,
        max_col=max_column,
        values_only=True,
    ):
        if all(value is None or _text(value) == "" for value in values):
            blank_run += 1
            if blank_run >= 50:
                break
            continue
        blank_run = 0
        yield values


def _cell(values, columns: dict[str, int], role: str):
    """按 1 基列映射读取 values 元组；缺少角色或行长度不足时返回 ``None``。"""
    index = columns.get(role)
    return values[int(index) - 1] if index and int(index) <= len(values) else None


def _batch_name(path: str) -> str:
    """从文件名移除“辅料清单总表”及后缀，得到人工复核和区块标题使用的批次名。"""
    stem = Path(path).stem
    name = re.sub(r"辅料清单总表.*$", "", stem, flags=re.IGNORECASE).strip(" _-")
    return name or stem


def _infer_file_supplier(path: str) -> str:
    """从“供应商 + 可选月份 + 采购清单明细”的历史文件名中推断供应商。"""
    stem = Path(path).stem
    # 非贪婪捕获避免把紧邻“采购清单明细”的月份也吸收到供应商名称中。
    matched = re.match(r"(.+?)(?:\d{1,2}月)?采购清单明细", stem)
    return _supplier(matched.group(1) if matched else "")


def _merge_history_sheet(worksheet, layout, file_supplier, mapping, conflicts):
    """把一张历史明细页签中的有效 JBC 供应关系并入映射并统计冲突。"""
    columns = layout["columns"]
    for values in _iter_data_rows(worksheet, layout):
        code = _text(_cell(values, columns, "code"))
        # 供应商批次表只处理项目约定的 JBC 辅料编号，排除标题、合计和其他物料。
        if not code.upper().startswith("JBC"):
            continue
        # 文件级供应商可覆盖表内空列和错误公式；无法推断时再使用每行供应商。
        supplier = file_supplier or _supplier(_cell(values, columns, "supplier"))
        if not supplier:
            continue
        if code in mapping and mapping[code] != supplier:
            conflicts.add(code)
        mapping[code] = supplier  # 冲突时按约定让后读取文件成为本次有效记录。


def _history_mapping(history_paths: list[str], log=None) -> tuple[dict[str, str], int]:
    """
    读取历史采购明细，建立 ``材料编号 -> 供应商`` 映射并统计冲突编号数。

    历史文件按用户选择顺序处理；同一材料出现不同供应商时记录冲突，但采用最后一次
    有效记录，与“越新的文件通常越靠后选择”的使用习惯一致。文件名能推断供应商时
    优先采用文件级名称，否则读取表内供应商列。
    """
    mapping: dict[str, str] = {}
    conflicts: set[str] = set()
    for path in history_paths:
        file_supplier = _infer_file_supplier(path)
        # 历史表只参与读取映射，不需要样式和公式文本，使用只读计算值模式降低内存占用。
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            for worksheet in workbook.worksheets:
                layout = _detect_layout(worksheet)
                if layout:
                    _merge_history_sheet(
                        worksheet, layout, file_supplier, mapping, conflicts,
                    )
        finally:
            workbook.close()
    if log and conflicts:
        log("历史明细中有 %d 个材料编号对应多个供应商，已按文件选择顺序采用最后一次记录。" % len(conflicts))
    return mapping, len(conflicts)


def _read_batch_item(values, columns, resolver, fill_counts):
    """
    清洗单个当前批次行，返回 ``(item, status)``。

    ``status`` 为 ``"ok"`` 时 ``item`` 是可直接进入采购复核的完整记录；``"skip"``
    表示空行、非 JBC 编号、无正数数量或缺单位的无采购意义行；``"excluded"`` 表示
    材料名称含“原厂”需要单独计数。主数据库补全发生在原厂判断之前，因此源表名称为
    空但档案已标记原厂的记录也会被正确排除。
    """
    code = _text(_cell(values, columns, "code"))
    if not code.upper().startswith("JBC"):
        # 跳过空行、标题行、合计行及不属于该业务范围的物料编号。
        return None, "skip"
    item = {
        "code": code,
        "name": _text(_cell(values, columns, "name")),
        "spec": _text(_cell(values, columns, "spec")),
        "unit": _text(_cell(values, columns, "unit")),
    }
    if resolver is not None:
        # 正式主数据遵循“仅补空值”原则，fill_counts 用于最终记录各字段补全数量。
        resolver.fill_mapping(
            item, fields=("name", "spec", "unit"), counts=fill_counts,
        )
    quantity = _quantity(_cell(values, columns, "qty"))
    # 无效、零或负数量不构成采购需求；缺少单位时输出也无法交付供应商执行。
    if quantity is None or quantity <= 0 or not item["unit"]:
        return None, "skip"
    if _is_original(item["name"]):
        # 排除发生在供应商归属之前，使“原厂”记录不会出现在人工供应商复核清单。
        return None, "excluded"
    item.update({
        "qty": quantity,
        # 保留当前清单直接提供的供应商，后续其优先级高于历史记录和主数据库。
        "source_supplier": _supplier(_cell(values, columns, "supplier")),
    })
    return item, "ok"


def _read_batch(path: str, log=None, resolver=None, fill_counts=None) -> dict[str, object]:
    """
    读取单个当前批次清单，清洗有效采购行并用主数据库补齐名称、规格和单位。

    数量必须是正数且单位有效；材料编号仅接受 JBC 前缀。主数据库只补空字段，不覆盖
    当前清单已有内容。名称补全后再执行“原厂”排除，确保源表名称为空但主数据库已经
    标记为原厂的记录也不会进入输出。
    """
    if log:
        # 数量若由公式生成且缓存未刷新，data_only 会读为空值，处理前先向用户明确告警。
        common_core.warn_if_uncached(path, log, what="最终采购数量")
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet, layout = _best_sheet(workbook)
        columns = layout["columns"]
        items = []
        excluded_original = 0
        for values in _iter_data_rows(worksheet, layout):
            item, status = _read_batch_item(values, columns, resolver, fill_counts)
            if status == "excluded":
                excluded_original += 1
            elif item is not None:
                items.append(item)
        return {
            "batch": _batch_name(path),
            "file": os.path.basename(path),
            "sheet": worksheet.title,
            "items": items,
            "excluded_original": excluded_original,
        }
    finally:
        workbook.close()


def _current_supplier_mapping(
    batches: list[dict[str, object]],
) -> tuple[dict[str, str], int]:
    """
    从全部当前批次中提取材料供应商映射，并统计同一编号对应不同供应商的冲突数。

    当前清单比历史记录更接近本次业务，因此后续供应商分配优先使用此映射。同一材料
    在多个当前批次中冲突时采用后扫描批次的值，同时把冲突数量带到人工复核结果中，
    让用户知道该自动归属需要重点检查。
    """
    mapping: dict[str, str] = {}
    conflicts: set[str] = set()
    for batch in batches:
        for item in batch["items"]:
            supplier = str(item["source_supplier"])
            code = str(item["code"])
            if not supplier:
                # 空供应商不应覆盖同一材料在其他批次中已经出现的有效归属。
                continue
            if code in mapping and mapping[code] != supplier:
                conflicts.add(code)
            mapping[code] = supplier  # 冲突按批次输入顺序以后者为本次有效映射。
    return mapping, len(conflicts)


def _assign_suppliers(
    batches: list[dict[str, object]],
    current_map: dict[str, str],
    history_map: dict[str, str],
    resolver: material_catalog.CatalogResolver,
    fill_counts: dict[str, int],
) -> tuple[list[dict[str, object]], dict[str, dict[str, int]]]:
    """
    按“当前批次 -> 历史明细 -> 主数据库”的优先级确定每条材料的供应商。

    返回未匹配材料明细和按供应商、批次统计的行数。主数据库仍只补充前两级都没有
    给出的供应商，不能被动覆盖业务文件中的明确归属。未匹配记录留给人工发现和维护，
    不会被写入任意供应商文件，从而避免错误采购。
    """
    unmatched: list[dict[str, object]] = []
    supplier_counts: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for batch in batches:
        batch_name = str(batch["batch"])
        for item in batch["items"]:
            code = str(item["code"])
            # 本次业务文件最可信，其次才是历史供应关系；二者都为空时再请求主数据补全。
            supplier = current_map.get(code) or history_map.get(code)
            additions = resolver.complete_material(
                code, {"supplier": supplier}, fields=("supplier",), counts=fill_counts,
            )
            # 使用显式占位便于前端统计，但该占位不会进入有效供应商集合。
            supplier = supplier or additions.get("supplier") or "未匹配"
            item["supplier"] = supplier
            if supplier == "未匹配":
                unmatched.append({
                    "batch": batch_name,
                    "code": item["code"],
                    "name": item["name"],
                })
            else:
                supplier_counts[str(supplier)][batch_name] += 1
    return unmatched, supplier_counts


def _supplier_summaries(
    supplier_counts: dict[str, dict[str, int]],
) -> list[dict[str, object]]:
    """
    把嵌套计数转换为排序稳定、可直接供人工复核界面展示的列表结构。

    供应商和批次均按名称排序，避免集合或字典构建顺序变化导致复核页面选项跳动；
    ``rows`` 是该供应商跨批次总行数，``batches`` 保留每个批次的明细规模。
    """
    return [
        {
            "name": supplier,
            "rows": sum(by_batch.values()),
            "batches": [
                {"batch": name, "rows": by_batch[name]} for name in sorted(by_batch)
            ],
        }
        for supplier, by_batch in sorted(supplier_counts.items())
    ]


def _collection_inputs(batch_paths, history_paths):
    """
    校验正式批次输入与可选历史明细，并返回统一的绝对路径列表。

    ``batch_paths`` 必须非空，``history_paths`` 允许为空；两个列表都会经过存在性和
    扩展名校验，历史明细为空时直接返回空列表，供后续读取阶段跳过。返回元组与读取
    顺序一一对应，调用方按同样的顺序消费文件。
    """
    batches = _validate_files(batch_paths, "当前批次清单")
    history = _validate_files(history_paths, "历史供应商明细") if history_paths else []
    return batches, history


def _read_current_batches(batch_paths, resolver, fill_counts, log, progress):
    """
    读取全部当前批次，并把 20%～65% 的进度按文件数线性上报。

    ``resolver`` 与 ``fill_counts`` 由调用方创建并贯穿本次任务，确保各批次补全使用
    同一个主数据库解析器，补全统计不会因重复加载而失真。返回批次顺序与
    ``batch_paths`` 一致，后续供应商冲突按该顺序取最后一次有效值。
    """
    batches = []
    total = len(batch_paths)
    for index, path in enumerate(batch_paths, start=1):
        batches.append(_read_batch(
            path,
            log=log,
            resolver=resolver,
            fill_counts=fill_counts,
        ))
        if progress:
            progress(20 + round(index / total * 45))
    return batches


def _ensure_batch_items(batches):
    """
    拒绝所有批次都无可制作记录的输入，避免生成空成品或空工作簿。

    输入为空意味着正式执行只会产生无内容的目录或零行文件，还会让后续“至少一家
    供应商”校验失去意义，因此在读取完成后立即失败。
    """
    if not any(batch["items"] for batch in batches):
        raise ValueError("批次清单中没有可制作的正数采购记录")


def _resolve_batch_suppliers(batches, history_map, resolver, fill_counts):
    """
    合并当前、历史与主数据供应关系，返回复核摘要和两层冲突数量。

    先聚合当前批次映射，再按“当前 -> 历史 -> 主数据”优先级分配供应商；没有任何
    供应商被识别时抛出 ``ValueError``，因为“未匹配”不能作为供应商写出。返回值中
    ``suppliers`` 已按名称排序，``unmatched`` 保持输入扫描顺序便于人工逐条排查。
    """
    current_map, current_conflicts = _current_supplier_mapping(batches)
    unmatched, supplier_counts = _assign_suppliers(
        batches,
        current_map,
        history_map,
        resolver,
        fill_counts,
    )
    suppliers = _supplier_summaries(supplier_counts)
    if not suppliers:
        # 不能把“未匹配”当成供应商写出；用户必须先补充业务资料或维护主数据。
        raise ValueError("未识别到供应商，请补充带供应商列的批次清单或历史供应商明细")
    return suppliers, unmatched, current_conflicts


def _build_collection_result(batches, suppliers, unmatched, history_map,
                             history_conflicts, current_conflicts):
    """
    组装分析与执行共同使用的稳定中间协议。

    返回字典是 ``analyze`` 与 ``run`` 共用的业务快照：同一输入、同一扫描顺序下
    内容稳定，复核阶段展示的供应商、未匹配数量和原厂排除数量与执行阶段完全一致。
    """
    return {
        "suppliers": suppliers,
        "batches": batches,
        "unmatched": unmatched,
        "history_mapping_count": len(history_map),
        # 当前与历史冲突来自不同可信层，只合并展示数量，不改变既定来源优先级。
        "supplier_conflicts": history_conflicts + current_conflicts,
        "excluded_original_count": sum(
            int(batch["excluded_original"]) for batch in batches
        ),
    }


def _collect(batch_paths, history_paths=None, log=None, progress=None) -> dict[str, object]:
    """
    聚合扫描、主数据补全、供应商归属和人工复核所需的全部中间数据。

    ``analyze`` 与 ``run`` 都调用本函数，确保复核阶段看到的供应商、未匹配材料、冲突
    和原厂排除数量与最终执行完全一致。函数只读取输入和主数据库，不创建输出文件。
    """
    batches_input, history_input = _collection_inputs(batch_paths, history_paths)
    if progress:
        progress(5)
    history_map, history_conflicts = _history_mapping(history_input, log=log)
    if progress:
        progress(20)

    # 单次任务复用同一个解析器和计数器，减少重复加载主数据并给出统一补全统计。
    resolver = material_catalog.CatalogResolver()
    fill_counts: dict[str, int] = {}
    batches = _read_current_batches(
        batches_input, resolver, fill_counts, log, progress,
    )
    _ensure_batch_items(batches)
    suppliers, unmatched, current_conflicts = _resolve_batch_suppliers(
        batches, history_map, resolver, fill_counts,
    )
    if progress:
        progress(70)
    material_catalog.log_fill_summary(log, "供应商批次表", fill_counts)
    return _build_collection_result(
        batches,
        suppliers,
        unmatched,
        history_map,
        history_conflicts,
        current_conflicts,
    )


def analyze(batch_paths, history_paths=None, log=None, progress=None) -> dict[str, object]:
    """
    只读扫描输入，返回供应商选择、批次交付日期和风险提示所需的人工复核计划。

    未匹配明细最多返回前 100 条以控制桌面桥接和 Web JSON 体积，同时单独返回完整计数；
    正式执行仍会重新调用相同聚合逻辑，避免长时间复核后沿用过期的内存数据。
    """
    data = _collect(batch_paths, history_paths, log=log, progress=progress)
    if progress:
        progress(100)
    return {
        "suppliers": data["suppliers"],
        "batches": [{
            "batch": batch["batch"],
            "file": batch["file"],
            "sheet": batch["sheet"],
            "rows": len(batch["items"]),
        } for batch in data["batches"]],
        "unmatched": data["unmatched"][:100],  # 预览限量，不影响完整未匹配数量和执行结果。
        "unmatched_count": len(data["unmatched"]),
        "history_mapping_count": data["history_mapping_count"],
        "supplier_conflicts": data["supplier_conflicts"],
        "excluded_original_count": data["excluded_original_count"],
    }


def _safe_name(value: str, fallback: str) -> str:
    """替换 Windows 禁用字符、去除危险尾缀并截断长度，生成跨平台安全名称片段。"""
    # 同时遵守 Windows 更严格的文件名约束，Linux 部署包生成的文件也可安全回传 Windows。
    cleaned = re.sub(r"[<>:\"/\\|?*]", "_", value).strip(" .")
    return cleaned[:100] or fallback


def _batch_group(names: list[str]) -> str:
    """单批次直接返回名称；多批次提取足够长的公共前缀，否则使用明确的“多批次”。"""
    if len(names) == 1:
        return names[0]
    # commonprefix 是字符级公共前缀，长度过短通常没有识别意义，因此设置四字符下限。
    prefix = os.path.commonprefix(names).rstrip(" _-")
    return prefix if len(prefix) >= 4 else "多批次"


def _combined_sheet_name(names: list[str]) -> str:
    """
    为多个批次共用的页签生成可读名称，并满足 Excel 页签最多 31 个字符的限制。

    若批次具有共同前缀，则用“公共前缀 + 各自后缀”表达差异；否则退回批次组名称。
    最终仍经过跨平台安全字符清洗，防止斜杠等字符导致 create_sheet 失败。
    """
    if len(names) == 1:
        return _safe_name(names[0], "批次")[:31]
    prefix = os.path.commonprefix(names).rstrip(" _-")
    suffixes = [name[len(prefix):].strip(" _-") for name in names]
    if prefix and all(suffixes):
        return _safe_name(prefix + "&".join(suffixes), "批次汇总")[:31]
    return _safe_name(_batch_group(names), "批次汇总")[:31]


def _style_sheet(worksheet) -> None:
    """设置统一列宽、网格线及横向单页宽度打印参数，不改写具体业务单元格。"""
    widths = {"A": 18, "B": 36, "C": 34, "D": 11, "E": 24, "F": 14}
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width
    worksheet.sheet_view.showGridLines = True
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1  # 打印时横向压到一页，纵向页数由明细长度自然扩展。
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True


def _write_block_title(worksheet, start_row, batch_name, delivery_date, border):
    """写入横跨六列、含批次与交付日期的黄色标题行。"""
    worksheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
    title = worksheet.cell(start_row, 1, "%s交付日期%s" % (batch_name, delivery_date))
    title.fill = PatternFill("solid", fgColor="FFFF00")
    title.font = Font(name="宋体", size=11)
    title.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[start_row].height = 24
    # 合并单元格仍逐列设置边框，确保 Excel 渲染和纸质打印时四周边线完整。
    for column in range(1, 7):
        worksheet.cell(start_row, column).border = border


def _write_block_header(worksheet, header_row, border):
    """按单一事实来源 ``HEADERS`` 写入表头行。"""
    for column, value in enumerate(HEADERS, start=1):
        cell = worksheet.cell(header_row, column, value)
        cell.font = Font(name="宋体", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    worksheet.row_dimensions[header_row].height = 30


def _write_block_item(worksheet, row_index, item, border):
    """写入一条供应商批次明细并返回下一行行号。"""
    values = [
        item["code"], item["name"], item["spec"], item["unit"],
        round(float(item["qty"]), 2), item["supplier"],
    ]
    for column, value in enumerate(values, start=1):
        cell = worksheet.cell(row_index, column, value)
        cell.font = Font(name="宋体", size=10)
        # 数量右对齐便于比较位数，供应商居中，其余描述字段左对齐并允许换行。
        cell.alignment = Alignment(
            horizontal="right" if column == 5 else "center" if column == 6 else "left",
            vertical="center",
            wrap_text=True,
        )
        cell.border = border
    worksheet.row_dimensions[row_index].height = 22
    return row_index + 1


def _write_block(
    worksheet,
    start_row: int,
    batch_name: str,
    delivery_date: str,
    items: list[dict[str, object]],
) -> int:
    """
    从 ``start_row`` 起写入一个“批次标题 + 表头 + 明细”区块，并返回最后使用行号。

    标题同时包含明确批次名和人工确认的交付日期，解决多个区块合并到同一页签后日期
    无法对应批次的问题。返回最后行而非下一行，让调用方自行决定区块间需要留几行。
    """
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    _write_block_title(worksheet, start_row, batch_name, delivery_date, border)
    header_row = start_row + 1
    _write_block_header(worksheet, header_row, border)
    row_index = header_row + 1
    for item in items:
        row_index = _write_block_item(worksheet, row_index, item, border)
    return row_index - 1


def _write_supplier_workbook(
    path: str,
    supplier: str,
    batches: list[dict[str, object]],
    batch_dates: dict[str, str],
) -> int:
    """
    筛选一个供应商的全部批次记录，并按总行数决定分批次页签或合并页签输出。

    超过 ``SPLIT_THRESHOLD`` 时每个批次独立成页，减少大表查找成本；较少记录则合并
    到一个页签并在区块间留空行，方便供应商一次查看。此处再次执行“原厂”过滤，作为
    输出边界的防御性校验，确保上游未来调整时也不会意外把原厂材料写出。
    返回实际写出的材料行数；没有该供应商记录时不创建空工作簿内容。
    """
    per_batch = []
    for batch in batches:
        items = [
            item for item in batch["items"]
            if item.get("supplier") == supplier and not _is_original(item.get("name"))
        ]
        if items:
            per_batch.append((str(batch["batch"]), items))
    total_rows = sum(len(items) for _, items in per_batch)
    if not total_rows:
        return 0

    workbook = openpyxl.Workbook()
    # 删除 openpyxl 自动创建的空白页，保证成品只包含业务页签。
    workbook.remove(workbook.active)
    try:
        if total_rows > SPLIT_THRESHOLD:
            # 大批量按批次拆页，每页标题仍保留该批次自己的交付日期。
            for batch_name, items in per_batch:
                worksheet = workbook.create_sheet(_safe_name(batch_name, "批次")[:31])
                _style_sheet(worksheet)
                _write_block(worksheet, 1, batch_name, batch_dates[batch_name], items)
        else:
            # 小批量合并到一个页签，降低供应商在多个短页签间切换的成本。
            worksheet = workbook.create_sheet(_combined_sheet_name([name for name, _ in per_batch]))
            _style_sheet(worksheet)
            start_row = 1
            for batch_name, items in per_batch:
                start_row = _write_block(
                    worksheet, start_row, batch_name, batch_dates[batch_name], items,
                ) + 2  # 在上一批次末行后留一空行，再开始下一个标题区块。
        # 所有页签写入完成后再执行唯一一次落盘；目标是 unique_path 生成的新路径，保存失败不会影响既有文件。
        workbook.save(path)
    finally:
        workbook.close()
    return total_rows


def _selected_supplier_names(
    available: list[str], selected_suppliers,
) -> list[str]:
    """
    清洗并校验人工选择的供应商，按首次出现顺序去重后返回。

    ``None`` 表示调用端没有提供选择，兼容自动生成全部已识别供应商；显式空列表则
    表示用户尚未完成必要复核，必须报错。任何不在本次扫描结果中的名称都被拒绝，
    防止客户端提交陈旧或伪造选项生成空文件。
    """
    if selected_suppliers is None:
        return available
    # dict 保序特性用于稳定去重，让输出文件顺序与用户勾选顺序一致。
    selected = list(dict.fromkeys(_supplier(value) for value in selected_suppliers))
    selected = [value for value in selected if value]
    if not selected:
        raise ValueError("请至少选择一个需要制作批次表的供应商")
    unknown = [value for value in selected if value not in available]
    if unknown:
        raise ValueError("所选供应商不在本次扫描结果中：%s" % "、".join(unknown))
    return selected


def _normalized_batch_dates(
    batch_names: list[str], batch_dates,
) -> dict[str, str]:
    """
    校验人工填写的批次交付日期，并按本次批次扫描顺序返回规范字典。

    交付日期暂以业务文本保存，允许“8月15日前”等人工约定写法，因此不强制解析成
    日历日期；但必须逐批填写、不得包含未知批次、控制字符或超长文本。按扫描顺序重建
    返回字典，可使桌面端、Web 端和输出文件在相同输入下保持稳定排列。
    """
    if not isinstance(batch_dates, dict):
        raise ValueError("请填写每个批次的交付日期")
    # 同时清洗键和值，抵御前端多余空白；空批次键不会进入后续未知项判断。
    normalized = {
        _text(batch): _text(value)
        for batch, value in batch_dates.items()
        if _text(batch)
    }
    unknown = [batch for batch in normalized if batch not in batch_names]
    if unknown:
        raise ValueError("交付日期中包含未知批次：%s" % "、".join(unknown))
    missing = [batch for batch in batch_names if not normalized.get(batch)]
    if missing:
        raise ValueError("请填写以下批次的交付日期：%s" % "、".join(missing))
    # 控制字符可能破坏 Excel XML 或制造不可见内容，长度上限则保护标题排版。
    invalid = [
        batch for batch in batch_names
        if len(normalized[batch]) > 50 or re.search(r"[\x00-\x1f]", normalized[batch])
    ]
    if invalid:
        raise ValueError("交付日期格式不正确：%s" % "、".join(invalid))
    return {batch: normalized[batch] for batch in batch_names}


def _supplier_output_dir(batch_paths, out_dir) -> str:
    """
    按统一设置解析供应商批次表输出目录；显式目录则绝对化并确保存在。

    默认路径把第一个批次文件作为“源文件旁输出”模式的定位依据，其他输出模式仍由
    ``settings`` 和 ``paths`` 的统一规则决定，不在本模块维护第二套目录约定。
    """
    if out_dir is None:
        current = settings.get_settings()
        return paths.resolve_output_dir(
            "supplier_batch",
            # _collect 已校验至少有一个批次文件，因此此处访问首项是安全的。
            src_path=os.path.abspath(str(batch_paths[0])),
            **current.output_kwargs(),
        )
    resolved = os.path.abspath(str(out_dir))
    os.makedirs(resolved, exist_ok=True)
    return resolved


def _generate_supplier_files(
    selected: list[str],
    batches: list[dict[str, object]],
    batch_dates: dict[str, str],
    out_dir: str,
    progress=None,
) -> tuple[list[str], int]:
    """
    按人工选择顺序逐供应商生成工作簿，返回非空文件路径和累计明细行数。

    文件名由安全供应商名与批次组名组成；若同名文件已存在，使用公共 ``unique_path``
    生成不覆盖历史结果的新路径。进度使用剩余 70%～100% 区间按供应商数均匀推进。
    """
    group_name = _safe_name(_batch_group([str(batch["batch"]) for batch in batches]), "批次")
    files: list[str] = []
    total_rows = 0
    for index, supplier in enumerate(selected, start=1):
        filename = "%s%s批次采购清单明细.xlsx" % (_safe_name(supplier, "供应商"), group_name)
        # 始终生成唯一名称，避免用户重复执行时悄悄覆盖已确认并发送过的采购文件。
        # 同一目录下并发执行时，unique_path 的序号机制也会错开目标，防止相互覆盖。
        target = common_core.unique_path(os.path.join(out_dir, filename))
        rows = _write_supplier_workbook(target, supplier, batches, batch_dates)
        if rows:
            # 理论上已选供应商都有记录；仍只返回非空文件，防御数据在执行阶段发生变化。
            files.append(target)
            total_rows += rows
        if progress:
            progress(70 + round(index / len(selected) * 30))
    return files, total_rows


def _prepare_run_review(data, selected_suppliers, batch_dates):
    """
    按本次重新扫描结果校验供应商选择和逐批交付日期。

    供应商选择只允许落在本次扫描结果内，交付日期必须覆盖全部输入批次；返回的
    ``selected`` 保留用户勾选顺序，``normalized_dates`` 按扫描顺序重建，保证输出
    文件与标题排列稳定。
    """
    available = [str(item["name"]) for item in data["suppliers"]]
    selected = _selected_supplier_names(available, selected_suppliers)
    batch_names = [str(batch["batch"]) for batch in data["batches"]]
    # 即使只选择部分供应商，也必须为所有输入批次填写日期，保证每个标题都可追溯。
    normalized_dates = _normalized_batch_dates(batch_names, batch_dates)
    return selected, normalized_dates


def _log_generation_summary(log, data, files, total_rows):
    """
    汇报正常产出、原厂排除和未匹配三类正式执行结果。

    ``log`` 允许为空；这里只报告结果供界面和日志使用，不改变任何业务数据或返回
    结构，未匹配记录也不会因此被写成供应商文件。
    """
    if not log:
        return
    log("已生成 %d 家供应商、%d 行批次明细。" % (len(files), total_rows))
    if data["excluded_original_count"]:
        log("已排除 %d 条材料名称含“原厂”的记录。" % data["excluded_original_count"])
    if data["unmatched"]:
        log("另有 %d 条记录未匹配供应商，未生成到供应商批次表。" % len(data["unmatched"]))


def _build_run_result(out_dir, files, selected, normalized_dates, total_rows, data):
    """
    生成供应商批次表正式入口的稳定返回结构。

    仅把已确认的输入和统计值封装为返回字典，不做额外计算；调用方按该结构展示
    输出目录、文件列表和排除/未匹配统计。
    """
    return {
        "out_dir": out_dir,
        "files": files,
        "suppliers": selected,
        "batch_dates": normalized_dates,
        "generated": len(files),
        "rows": total_rows,
        "excluded_original_count": data["excluded_original_count"],
        "unmatched_count": len(data["unmatched"]),
    }


def run(
    batch_paths,
    history_paths=None,
    selected_suppliers=None,
    batch_dates=None,
    out_dir=None,
    log=None,
    progress=None,
) -> dict[str, object]:
    """
    执行供应商批次表正式生成，并返回输出文件、复核选择和排除统计。

    本入口会重新扫描输入，随后校验供应商选择和每个批次的交付日期，再按供应商分别
    写出工作簿。材料名称含“原厂”的记录在读取和输出两层均会排除；未匹配供应商的
    记录只计入提示，不会被放进任何成品。调用方可通过 ``out_dir`` 覆盖统一输出设置，
    ``log`` 和 ``progress`` 仅报告状态，不参与业务判断。
    """
    data = _collect(batch_paths, history_paths, log=log, progress=progress)
    selected, normalized_dates = _prepare_run_review(
        data, selected_suppliers, batch_dates,
    )
    out_dir = _supplier_output_dir(batch_paths, out_dir)
    files, total_rows = _generate_supplier_files(
        selected, data["batches"], normalized_dates, out_dir, progress,
    )
    _log_generation_summary(log, data, files, total_rows)
    return _build_run_result(
        out_dir, files, selected, normalized_dates, total_rows, data,
    )
