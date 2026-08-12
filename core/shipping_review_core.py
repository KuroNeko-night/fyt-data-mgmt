# -*- coding: utf-8 -*-
"""发运评审对比：过滤包装计划、聚合物料数量并与评审表核对。

该模块是桌面端和 Web 端共用的唯一业务事实源。输入文件始终只读：包装日计划中的
``BOX状态=已作废`` 记录只在内存中排除，不会改写源工作簿。两侧都可能因 BOX、供应商
或其他维度把同一物料拆成多行，因此比较前必须先按物料号汇总，禁止采用“最后一行
覆盖前一行”的实现。
"""

from __future__ import annotations

import os
import re
from collections import Counter, defaultdict
from decimal import Decimal, InvalidOperation
from typing import Iterable, Mapping

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import common_core
from . import material_catalog
from . import paths
from . import settings


PACKAGE_HEADERS = {
    "code": ("物料号",),
    "name": ("物料描述",),
    "status": ("BOX状态", "BOX 状态"),
    "quantity": ("实际包装数量",),
}
REVIEW_HEADERS = {
    "code": ("Part No", "Part No.", "PartNo"),
    "name": ("Chinese Name", "Chinese name", "ChineseName"),
    "quantity": ("总数",),
}
OPTIONAL_PACKAGE_HEADERS = {
    "box_no": ("BOX号", "BOX编号", "BOX NO", "BOX No", "箱号"),
}

_TITLE_FILL = PatternFill("solid", fgColor="FF1F4E78")
_HEADER_FILL = PatternFill("solid", fgColor="FF5B9BD5")
_SECTION_FILL = PatternFill("solid", fgColor="FFD9EAF7")
_SUCCESS_FILL = PatternFill("solid", fgColor="FFE2F0D9")
_WARNING_FILL = PatternFill("solid", fgColor="FFFFF2CC")
_DANGER_FILL = PatternFill("solid", fgColor="FFFCE4D6")
_INFO_FILL = PatternFill("solid", fgColor="FFDDEBF7")
_THIN = Side(style="thin", color="FFD9E2F3")
_BORDER = Border(bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _text(value: object) -> str:
    """返回用于展示的规范文本，清除全角和不可见字符但保留内部业务符号。"""

    return common_core.clean_str(value)


def _header_key(value: object) -> str:
    """生成表头匹配键；忽略空白、大小写和常见标点差异。"""

    return re.sub(r"[\s._·-]+", "", _text(value)).casefold()


def _material_code(value: object) -> str:
    """规范物料号的 Excel 表现，同时保留文本编码中的前导零。

    数值单元格 ``123.0`` 会恢复为 ``123``；文本 ``00123`` 仍保持前导零，因为它可能
    是正式物料编码。内部空白只在匹配键中移除，报告仍使用首次出现的规范展示值。
    """

    if value is None or isinstance(value, bool):
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = _text(value)
    match = re.fullmatch(r"([+-]?\d+)\.0+", text)
    return match.group(1) if match else text


def _code_key(value: object) -> str:
    """生成只用于双方关联的物料号键，不改变报告中的展示写法。"""

    return re.sub(r"\s+", "", _material_code(value)).casefold()


def _decimal(value: object, *, file_name: str, sheet: str, row: int, label: str) -> Decimal:
    """把数量转换为精确十进制；有物料的无效数量必须明确报错而不是按零吞掉。"""

    if isinstance(value, bool) or value is None or _text(value) == "":
        raise ValueError(
            f"《{file_name}》工作表“{sheet}”第 {row} 行的“{label}”为空，请补充后重试。"
        )
    raw = _text(value).replace(",", "")
    try:
        number = Decimal(raw)
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(
            f"《{file_name}》工作表“{sheet}”第 {row} 行的“{label}”不是有效数字：{_text(value)}"
        ) from exc
    if not number.is_finite():
        raise ValueError(
            f"《{file_name}》工作表“{sheet}”第 {row} 行的“{label}”不是有限数字。"
        )
    return number


def _json_number(value: Decimal) -> int | float:
    """把精确计算结果转换为 JSON 和 openpyxl 都可直接使用的数值。"""

    if value == value.to_integral_value():
        return int(value)
    return float(value)


def _match_headers(values: Iterable[object], required: Mapping[str, Iterable[str]]) -> dict[str, int] | None:
    """在一行中匹配全部必需表头，返回零基列号；缺少任意角色即判定失败。"""

    normalized = {_header_key(value): index for index, value in enumerate(values) if _header_key(value)}
    columns: dict[str, int] = {}
    for role, aliases in required.items():
        column = next((normalized.get(_header_key(alias)) for alias in aliases if _header_key(alias) in normalized), None)
        if column is None:
            return None
        columns[role] = column
    return columns


def _optional_headers(values: Iterable[object], aliases: Mapping[str, Iterable[str]]) -> dict[str, int]:
    """识别审计需要但不影响业务执行的可选列。"""

    normalized = {_header_key(value): index for index, value in enumerate(values) if _header_key(value)}
    result: dict[str, int] = {}
    for role, names in aliases.items():
        column = next((normalized.get(_header_key(name)) for name in names if _header_key(name) in normalized), None)
        if column is not None:
            result[role] = column
    return result


def _sheet_layout(worksheet, required: Mapping[str, Iterable[str]], *, max_rows: int = 60):
    """在页签前若干行中定位表头，避免把说明行或合并标题误当作数据。"""

    # 流式读取器为修复错误 dimension 声明会把 max_row 重置为 None，因此不能依赖它做 min。
    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=max_rows, values_only=True),
        start=1,
    ):
        columns = _match_headers(row, required)
        if columns is not None:
            return row_number, columns, row
    return None


def _select_package_sheet(workbook, requested: str | None = None):
    """选择包含包装四个必需字段的唯一页签。"""

    if requested:
        if requested not in workbook.sheetnames:
            raise ValueError(f"包装日计划中不存在工作表“{requested}”。")
        layout = _sheet_layout(workbook[requested], PACKAGE_HEADERS)
        if layout is None:
            raise ValueError(f"包装工作表“{requested}”缺少物料号、物料描述、BOX状态或实际包装数量。")
        return workbook[requested], layout
    candidates = [(worksheet, layout) for worksheet in workbook.worksheets if (layout := _sheet_layout(worksheet, PACKAGE_HEADERS))]
    if not candidates:
        raise ValueError("包装日计划中未找到同时包含物料号、物料描述、BOX状态和实际包装数量的工作表。")
    if len(candidates) > 1:
        names = "、".join(worksheet.title for worksheet, _ in candidates)
        raise ValueError(f"包装日计划中有多个可用工作表（{names}），请人工指定包装工作表。")
    return candidates[0]


def _select_review_sheet(workbook, requested: str | None = None):
    """优先使用人工指定页签，否则采用工作簿保存时的活动页签。

    活动页签不符合评审格式时才自动搜索候选；唯一候选可安全使用，多候选则要求人工
    指定，避免打开工作簿后的当前视图与业务数据页不一致时静默选错。
    """

    if requested:
        if requested not in workbook.sheetnames:
            raise ValueError(f"发运评审表中不存在工作表“{requested}”。")
        layout = _sheet_layout(workbook[requested], REVIEW_HEADERS)
        if layout is None:
            raise ValueError(f"评审工作表“{requested}”缺少 Part No、Chinese Name 或总数。")
        return workbook[requested], layout
    active = workbook.active
    if active is not None:
        layout = _sheet_layout(active, REVIEW_HEADERS)
        if layout is not None:
            return active, layout
    candidates = [(worksheet, layout) for worksheet in workbook.worksheets if (layout := _sheet_layout(worksheet, REVIEW_HEADERS))]
    if not candidates:
        raise ValueError("发运评审文件中未找到同时包含 Part No、Chinese Name 和总数的工作表。")
    if len(candidates) > 1:
        names = "、".join(worksheet.title for worksheet, _ in candidates)
        raise ValueError(f"发运评审文件中有多个候选工作表（{names}），请人工指定评审工作表。")
    return candidates[0]


def _new_item(code: str) -> dict[str, object]:
    """创建按物料号聚合的内部记录；Decimal 和集合只在 Core 内部流转。"""

    return {"code": code, "quantity": Decimal("0"), "names": set(), "rows": 0}


def _add_name(item: dict[str, object], value: object) -> None:
    """把非空名称加入集合，保留同一物料存在多个名称的审计证据。"""

    text = _text(value)
    if text:
        item["names"].add(text)


def _fill_empty_names(items: Mapping[str, dict[str, object]], resolver, counts: dict[str, int]) -> None:
    """仅当某物料一侧完全没有名称时从主数据库补全，不混入或覆盖源名称。"""

    for item in items.values():
        if item["names"]:
            continue
        addition = resolver.complete_material(
            item["code"], {"name": ""}, fields=("name",), counts=counts,
        )
        if addition.get("name"):
            item["names"].add(addition["name"])


def _read_package(path: str, requested_sheet: str | None, resolver, fill_counts, log, progress):
    """读取包装计划，记录状态审计，并建立描述维度透视和物料号汇总。"""

    workbook = common_core.load_data_only_stream(path)
    try:
        worksheet, (header_row, columns, header_values) = _select_package_sheet(workbook, requested_sheet)
        optional = _optional_headers(header_values, OPTIONAL_PACKAGE_HEADERS)
        common_core.warn_if_uncached(path, log, sheet=worksheet.title, what="实际包装数量")
        by_code: dict[str, dict[str, object]] = {}
        pivot: dict[tuple[str, str], dict[str, object]] = {}
        status_counts: Counter[str] = Counter()
        excluded: list[dict[str, object]] = []
        source_rows = kept_rows = 0
        file_name = os.path.basename(path)
        # dimension 不可信时无法预知总行数，进度仍由阶段切换保证单调，逐行 tick 退化为提示性进度。
        data_total = max(1, (worksheet.max_row or header_row + 1) - header_row)
        for offset, values in enumerate(
            worksheet.iter_rows(min_row=header_row + 1, values_only=True), start=1,
        ):
            code = _material_code(values[columns["code"]] if columns["code"] < len(values) else None)
            if not code:
                # 完全空白、汇总和说明行不属于物料数据；含数量但无物料号则不能安全聚合。
                if any(_text(value) for value in values):
                    quantity_value = values[columns["quantity"]] if columns["quantity"] < len(values) else None
                    if _text(quantity_value):
                        raise ValueError(f"《{file_name}》工作表“{worksheet.title}”第 {header_row + offset} 行有数量但物料号为空。")
                progress.tick(offset, data_total)
                continue
            source_rows += 1
            row_number = header_row + offset
            name = _text(values[columns["name"]] if columns["name"] < len(values) else None)
            status = _text(values[columns["status"]] if columns["status"] < len(values) else None)
            status_counts[status or "（空白）"] += 1
            quantity = _decimal(
                values[columns["quantity"]] if columns["quantity"] < len(values) else None,
                file_name=file_name, sheet=worksheet.title, row=row_number, label="实际包装数量",
            )
            if status == "已作废":
                excluded.append({
                    "row": row_number,
                    "box_no": _text(values[optional["box_no"]]) if "box_no" in optional and optional["box_no"] < len(values) else "",
                    "code": code,
                    "name": name,
                    "quantity": _json_number(quantity),
                    "status": status,
                })
                progress.tick(offset, data_total)
                continue
            kept_rows += 1
            key = _code_key(code)
            item = by_code.setdefault(key, _new_item(code))
            item["quantity"] += quantity
            item["rows"] += 1
            _add_name(item, name)
            pivot_key = (key, name)
            pivot_item = pivot.setdefault(pivot_key, {"code": code, "name": name, "quantity": Decimal("0"), "rows": 0})
            pivot_item["quantity"] += quantity
            pivot_item["rows"] += 1
            progress.tick(offset, data_total)
        _fill_empty_names(by_code, resolver, fill_counts)
        # 描述为空的透视行只在主数据库确有名称时补空，不覆盖任何源描述。
        for (key, name), item in list(pivot.items()):
            if name:
                continue
            aggregate_names = by_code[key]["names"]
            if aggregate_names:
                item["name"] = sorted(aggregate_names)[0]
        return {
            "sheet": worksheet.title,
            "header_row": header_row,
            "source_rows": source_rows,
            "kept_rows": kept_rows,
            "obsolete_rows": len(excluded),
            "status_counts": dict(status_counts),
            "excluded": excluded,
            "items": by_code,
            "pivot": list(pivot.values()),
        }
    finally:
        workbook.close()


def _read_review(path: str, requested_sheet: str | None, resolver, fill_counts, log, progress):
    """读取评审表并按 Part No 汇总；供应商拆行必须累加而不是覆盖。"""

    workbook = common_core.load_data_only_stream(path)
    try:
        worksheet, (header_row, columns, _header_values) = _select_review_sheet(workbook, requested_sheet)
        common_core.warn_if_uncached(path, log, sheet=worksheet.title, what="总数")
        items: dict[str, dict[str, object]] = {}
        source_rows = 0
        file_name = os.path.basename(path)
        data_total = max(1, (worksheet.max_row or header_row + 1) - header_row)
        for offset, values in enumerate(
            worksheet.iter_rows(min_row=header_row + 1, values_only=True), start=1,
        ):
            code = _material_code(values[columns["code"]] if columns["code"] < len(values) else None)
            if not code:
                progress.tick(offset, data_total)
                continue
            source_rows += 1
            row_number = header_row + offset
            quantity = _decimal(
                values[columns["quantity"]] if columns["quantity"] < len(values) else None,
                file_name=file_name, sheet=worksheet.title, row=row_number, label="总数",
            )
            key = _code_key(code)
            item = items.setdefault(key, _new_item(code))
            item["quantity"] += quantity
            item["rows"] += 1
            _add_name(item, values[columns["name"]] if columns["name"] < len(values) else None)
            progress.tick(offset, data_total)
        _fill_empty_names(items, resolver, fill_counts)
        return {
            "sheet": worksheet.title,
            "header_row": header_row,
            "source_rows": source_rows,
            "items": items,
        }
    finally:
        workbook.close()


def _name_state(package_names: set[str], review_names: set[str]) -> str:
    """分别判断名称缺失与名称不一致，避免把空白双方误判为一致。"""

    if not package_names and not review_names:
        return "双方名称缺失"
    if not review_names:
        return "评审名称缺失"
    if not package_names:
        return "包装名称缺失"
    package_keys = {_header_key(name) for name in package_names}
    review_keys = {_header_key(name) for name in review_names}
    return "名称一致" if package_keys == review_keys else "名称不一致"


def _row_status(package_item, review_item, name_state: str, difference: Decimal) -> str:
    """按单侧、数量和名称三个维度生成互斥状态。"""

    if package_item is None:
        return "仅发运评审表"
    if review_item is None:
        return "仅包装日计划"
    quantity_diff = difference != 0
    name_diff = name_state != "名称一致"
    if quantity_diff and name_diff:
        return "数量及名称差异"
    if quantity_diff:
        return "数量差异"
    if name_diff:
        return name_state
    return "一致"


def _compare(package_items, review_items, progress):
    """以物料号并集生成完整对比行和状态统计。"""

    rows = []
    keys = sorted(set(package_items) | set(review_items))
    for index, key in enumerate(keys, start=1):
        package_item = package_items.get(key)
        review_item = review_items.get(key)
        package_qty = package_item["quantity"] if package_item else Decimal("0")
        review_qty = review_item["quantity"] if review_item else Decimal("0")
        difference = package_qty - review_qty
        package_names = set(package_item["names"]) if package_item else set()
        review_names = set(review_item["names"]) if review_item else set()
        name_state = _name_state(package_names, review_names)
        status = _row_status(package_item, review_item, name_state, difference)
        rows.append({
            "index": index,
            "code": package_item["code"] if package_item else review_item["code"],
            "package_name": "；".join(sorted(package_names)),
            "review_name": "；".join(sorted(review_names)),
            "package_quantity": _json_number(package_qty),
            "review_quantity": _json_number(review_qty),
            "difference": _json_number(difference),
            "status": status,
            "name_status": name_state,
            "package_rows": int(package_item["rows"]) if package_item else 0,
            "review_rows": int(review_item["rows"]) if review_item else 0,
        })
        progress.tick(index, max(1, len(keys)))
    return rows


def _counts(rows: list[dict[str, object]]) -> dict[str, int]:
    """汇总前端指标和报告顶部卡片所需计数。"""

    statuses = Counter(str(row["status"]) for row in rows)
    both = [row for row in rows if row["package_rows"] and row["review_rows"]]
    return {
        "total_materials": len(rows),
        "full_match": statuses["一致"],
        "quantity_match": sum(1 for row in both if Decimal(str(row["difference"])) == 0),
        "quantity_diff": sum(1 for row in both if Decimal(str(row["difference"])) != 0),
        "name_issues": sum(1 for row in both if row["name_status"] != "名称一致"),
        "only_package": statuses["仅包装日计划"],
        "only_review": statuses["仅发运评审表"],
        "exceptions": sum(1 for row in rows if row["status"] != "一致"),
    }


def _style_title(worksheet, title: str, end_column: int) -> None:
    """写入报告页签统一标题带。"""

    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    cell = worksheet.cell(1, 1, title)
    cell.fill = _TITLE_FILL
    cell.font = Font(name="微软雅黑", size=16, bold=True, color="FFFFFFFF")
    cell.alignment = _CENTER
    worksheet.row_dimensions[1].height = 32
    worksheet.sheet_view.showGridLines = False


def _style_header(cells) -> None:
    """设置明细表头样式。"""

    for cell in cells:
        cell.fill = _HEADER_FILL
        cell.font = Font(name="微软雅黑", bold=True, color="FFFFFFFF")
        cell.alignment = _CENTER
        cell.border = _BORDER


def _status_fill(status: str):
    """把业务状态映射为报告颜色。"""

    if status == "一致":
        return _SUCCESS_FILL
    if status in ("数量差异", "数量及名称差异"):
        return _DANGER_FILL
    if status.startswith("仅"):
        return _WARNING_FILL
    return _INFO_FILL


def _write_compare_sheet(worksheet, rows, counts, *, exceptions_only: bool = False) -> None:
    """写入总表或异常表；两页使用同一列契约，便于人工筛选和复核。"""

    title = "发运评审异常明细" if exceptions_only else "发运评审对比总表"
    selected = [row for row in rows if row["status"] != "一致"] if exceptions_only else rows
    _style_title(worksheet, title, 10)
    worksheet.merge_cells("A2:J2")
    worksheet["A2"] = (
        f"物料 {counts['total_materials']} 项｜完整一致 {counts['full_match']} 项｜"
        f"数量差异 {counts['quantity_diff']} 项｜名称问题 {counts['name_issues']} 项｜"
        f"单侧物料 {counts['only_package'] + counts['only_review']} 项"
    )
    worksheet["A2"].fill = _SECTION_FILL
    worksheet["A2"].font = Font(name="微软雅黑", bold=True, color="FF1F4E78")
    worksheet["A2"].alignment = _CENTER
    headers = ["序号", "物料号", "包装物料描述", "评审中文名称", "包装数量", "评审总数", "差异（包装－评审）", "状态", "包装来源行数", "评审来源行数"]
    for column, value in enumerate(headers, start=1):
        worksheet.cell(4, column, value)
    _style_header(worksheet[4])
    for row_number, item in enumerate(selected, start=5):
        values = [item["index"], item["code"], item["package_name"], item["review_name"], item["package_quantity"], item["review_quantity"], item["difference"], item["status"], item["package_rows"], item["review_rows"]]
        for column, value in enumerate(values, start=1):
            cell = worksheet.cell(row_number, column, value)
            cell.font = Font(name="微软雅黑", size=10)
            cell.alignment = _LEFT if column in (2, 3, 4) else _CENTER
            cell.border = _BORDER
        worksheet.cell(row_number, 8).fill = _status_fill(str(item["status"]))
    last_row = max(4, 4 + len(selected))
    worksheet.auto_filter.ref = f"A4:J{last_row}"
    worksheet.freeze_panes = "A5"
    widths = [8, 18, 34, 34, 14, 14, 18, 18, 14, 14]
    for column, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(column)].width = width
    for row_number in range(5, last_row + 1):
        worksheet.row_dimensions[row_number].height = 28
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True


def _write_pivot_sheet(worksheet, pivot_rows) -> None:
    """写入过滤后的包装侧透视结果，保留物料描述维度。"""

    _style_title(worksheet, "包装日计划透视", 4)
    headers = ["物料号", "物料描述", "实际包装数量合计", "来源行数"]
    for column, value in enumerate(headers, start=1):
        worksheet.cell(3, column, value)
    _style_header(worksheet[3])
    ordered = sorted(pivot_rows, key=lambda item: (_code_key(item["code"]), _text(item["name"])))
    for row_number, item in enumerate(ordered, start=4):
        values = [item["code"], item["name"], _json_number(item["quantity"]), item["rows"]]
        for column, value in enumerate(values, start=1):
            cell = worksheet.cell(row_number, column, value)
            cell.alignment = _LEFT if column in (1, 2) else _CENTER
            cell.font = Font(name="微软雅黑", size=10)
            cell.border = _BORDER
    last_row = max(3, 3 + len(ordered))
    worksheet.auto_filter.ref = f"A3:D{last_row}"
    worksheet.freeze_panes = "A4"
    for column, width in enumerate((20, 42, 22, 14), start=1):
        worksheet.column_dimensions[get_column_letter(column)].width = width


def _write_audit_sheet(worksheet, package_data) -> None:
    """写入 BOX 状态分布和被排除的已作废明细，确保过滤过程可追溯。"""

    _style_title(worksheet, "包装日计划过滤审计", 6)
    worksheet["A3"] = "BOX 状态"
    worksheet["B3"] = "记录数"
    _style_header(worksheet[3][0:2])
    status_rows = sorted(package_data["status_counts"].items(), key=lambda item: (-item[1], item[0]))
    for row_number, (status, count) in enumerate(status_rows, start=4):
        worksheet.cell(row_number, 1, status)
        worksheet.cell(row_number, 2, count)
        for cell in worksheet[row_number][0:2]:
            cell.border = _BORDER
            cell.alignment = _CENTER
    start = 5 + len(status_rows)
    worksheet.merge_cells(start_row=start, start_column=1, end_row=start, end_column=6)
    worksheet.cell(start, 1, f"已排除“已作废”记录（{package_data['obsolete_rows']} 条）")
    worksheet.cell(start, 1).fill = _SECTION_FILL
    worksheet.cell(start, 1).font = Font(name="微软雅黑", bold=True, color="FF1F4E78")
    worksheet.cell(start, 1).alignment = _LEFT
    headers = ["源行号", "BOX号", "物料号", "物料描述", "实际包装数量", "BOX状态"]
    for column, value in enumerate(headers, start=1):
        worksheet.cell(start + 1, column, value)
    _style_header(worksheet[start + 1])
    for row_number, item in enumerate(package_data["excluded"], start=start + 2):
        values = [item["row"], item["box_no"], item["code"], item["name"], item["quantity"], item["status"]]
        for column, value in enumerate(values, start=1):
            cell = worksheet.cell(row_number, column, value)
            cell.alignment = _LEFT if column in (2, 3, 4) else _CENTER
            cell.border = _BORDER
    worksheet.freeze_panes = "A4"
    for column, width in enumerate((12, 22, 20, 42, 18, 14), start=1):
        worksheet.column_dimensions[get_column_letter(column)].width = width


def _write_report(report_path: str, rows, counts, package_data) -> str:
    """原子生成四页签对比报告；失败时不留下半成品。"""

    workbook = Workbook()
    total_sheet = workbook.active
    total_sheet.title = "对比总表"
    _write_compare_sheet(total_sheet, rows, counts)
    _write_compare_sheet(workbook.create_sheet("异常明细"), rows, counts, exceptions_only=True)
    _write_pivot_sheet(workbook.create_sheet("包装透视"), package_data["pivot"])
    _write_audit_sheet(workbook.create_sheet("过滤审计"), package_data)
    temp_path = report_path + ".tmp.xlsx"
    try:
        workbook.save(temp_path)
        os.replace(temp_path, report_path)
    finally:
        workbook.close()
        if os.path.exists(temp_path):
            os.remove(temp_path)
    return report_path


def _resolve_output_dir(out_dir: str | None) -> str:
    """解析统一输出目录；显式目录用于 Web 用户隔离和测试。"""

    if out_dir is None:
        current = settings.get_settings()
        return paths.resolve_output_dir("shipping_review", **current.output_kwargs())
    os.makedirs(out_dir, exist_ok=True)
    return os.path.abspath(out_dir)


def run(
    package_plan,
    review_workbook,
    package_sheet=None,
    review_sheet=None,
    out_dir=None,
    log=None,
    progress=None,
) -> dict[str, object]:
    """执行发运评审对比并返回报告路径、统计、异常和结构化预览数据。

    ``review_sheet`` 为空时使用评审工作簿保存时的活动页签；``package_sheet`` 通常无需
    指定，只有包装工作簿存在多个同结构页签时才要求人工选择。
    """

    package_path = os.path.abspath(str(package_plan))
    review_path = os.path.abspath(str(review_workbook))
    for label, file_path in (("包装日计划", package_path), ("发运评审表", review_path)):
        if not os.path.isfile(file_path):
            raise FileNotFoundError(f"找不到{label}：{file_path}")
        if os.path.splitext(file_path)[1].lower() not in (".xlsx", ".xlsm"):
            raise ValueError(f"{label}仅支持 .xlsx 或 .xlsm 文件。")

    def _log(message: str) -> None:
        if callable(log):
            log(message)

    prog = common_core.Progress(progress, stages=[
        ("package", 35), ("review", 25), ("compare", 20), ("report", 20),
    ])
    resolver = material_catalog.CatalogResolver()
    fill_counts: dict[str, int] = {}

    prog.stage("package")
    _log("正在读取包装日计划并排除已作废 BOX……")
    package_data = _read_package(
        package_path, _text(package_sheet) or None, resolver, fill_counts, _log, prog,
    )
    _log(
        "包装日计划保留 %d 行，排除已作废 %d 行，汇总 %d 个物料。"
        % (package_data["kept_rows"], package_data["obsolete_rows"], len(package_data["items"]))
    )

    prog.stage("review")
    _log("正在读取发运评审表并按 Part No 汇总……")
    review_data = _read_review(
        review_path, _text(review_sheet) or None, resolver, fill_counts, _log, prog,
    )
    _log("评审工作表“%s”汇总 %d 个物料。" % (review_data["sheet"], len(review_data["items"])))
    material_catalog.log_fill_summary(_log, "发运评审对比", fill_counts)

    prog.stage("compare")
    _log("正在逐项比较物料名称与汇总数量……")
    detail_rows = _compare(package_data["items"], review_data["items"], prog)
    counts = _counts(detail_rows)
    exceptions = [row for row in detail_rows if row["status"] != "一致"]

    prog.stage("report")
    output_dir = _resolve_output_dir(out_dir)
    report_path = os.path.join(output_dir, "发运评审对比报告.xlsx")
    _write_report(report_path, detail_rows, counts, package_data)
    prog.done()
    _log(
        "对比完成：完整一致 %d 项，数量差异 %d 项，名称问题 %d 项，单侧物料 %d 项。"
        % (
            counts["full_match"], counts["quantity_diff"], counts["name_issues"],
            counts["only_package"] + counts["only_review"],
        )
    )
    _log("已生成报告：%s" % report_path)
    return {
        "report_path": report_path,
        "out_dir": output_dir,
        "package_sheet": package_data["sheet"],
        "review_sheet": review_data["sheet"],
        "source_rows": package_data["source_rows"],
        "kept_rows": package_data["kept_rows"],
        "obsolete_rows": package_data["obsolete_rows"],
        "review_rows": review_data["source_rows"],
        "package_materials": len(package_data["items"]),
        "review_materials": len(review_data["items"]),
        "status_counts": package_data["status_counts"],
        "counts": counts,
        "details": detail_rows,
        "exceptions": exceptions,
        "quality": {
            "score": round(100 * counts["full_match"] / max(1, counts["total_materials"])),
            "name_filled": int(fill_counts.get("name", 0)),
        },
        "parameters": {
            "package_file": os.path.basename(package_path),
            "review_file": os.path.basename(review_path),
            "package_sheet": package_data["sheet"],
            "review_sheet": review_data["sheet"],
        },
    }


__all__ = ["run"]
