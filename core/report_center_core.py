# -*- coding: utf-8 -*-
"""
公共任务报表中心
================
接收桌面或 Web 服务已经完成权限过滤的统一任务记录，按业务模块和状态汇总任务数、
完成数、失败/中断数与成功率，并生成“汇总 + 明细”Excel 报告。核心层只负责报表
投影，不直接读取用户数据库，避免绕过服务端的数据归属和管理员权限检查。

状态 ``ok``/``completed`` 计为完成，``failed``/``interrupted`` 计为失败；等待、运行中
和取消只进入任务总数，不计成功或失败。模块标题和状态标题通过显式字典翻译，未知值
保留原文本，便于新增功能在映射尚未更新时仍可导出。

时间范围起点辅助函数支持 7 天、30 天、当月以及历史兼容别名；实际筛选应由持有任务
数据的调用层完成，``build_report`` 假定传入列表已经符合用户选择范围。
"""
from __future__ import annotations

import os
import tempfile
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from . import common_core

FEATURE_TITLES = {
    # 报表中的客户可读模块名；未知 feature 会回退原值而不是丢弃记录。
    "attendance": "考勤填报", "reconcile": "工时对账", "arrival": "到料明细",
    "pivot": "销售透视", "purchase": "采购对账", "shipping_review": "发运评审对比",
    "delivery": "送货计划",
    "supplier_batch": "供应商批次表", "purchase_plan": "采购计划导入",
    "invoice": "发票统计", "rename": "批量重命名", "text": "文本工具",
    "pdf": "PDF 工具", "excel": "Excel 工具", "compare": "表格比对",
    "currency": "金额大写", "library": "数据库", "mappings": "字段映射",
    "templates": "模板中心", "settings": "设置",
}

STATUS_TITLES = {
    # 同时兼容桌面历史的 ok 和 Web 任务的 completed 状态。
    "ok": "完成", "running": "处理中", "queued": "等待", "failed": "失败",
    "interrupted": "中断", "cancelled": "取消", "completed": "完成",
}
DONE_STATUSES = {"ok", "completed"}
FAILED_STATUSES = {"failed", "interrupted"}


def _text(value) -> str:
    """把任务字段转换为去首尾空白的展示文本，空值统一为空串。"""
    return "" if value is None else str(value).strip()


def range_start(range_key: str) -> datetime:
    """把预设范围键转换为本地时间的查询起点。

    ``7d`` 与历史别名 ``week`` 都表示向前七天；``month`` 与 ``monthly`` 都表示本月
    首日零点；``30d`` 表示滚动三十天。未知键回退到 2000-01-01，作为“全部历史”
    的兼容下界。函数不负责时区转换，调用层应与任务记录的时间口径保持一致。
    """
    now = datetime.now()
    if range_key == "7d":
        return now - timedelta(days=7)
    if range_key == "30d":
        return now - timedelta(days=30)
    if range_key == "month":
        return now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    if range_key == "week":
        return now - timedelta(days=7)
    if range_key == "monthly":
        return now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    return datetime(2000, 1, 1)


def _parse_time(text: str) -> datetime | None:
    """尽力解析 ISO 时间文本为无时区 datetime，失败返回 ``None``。

    结尾 ``Z`` 先换成 UTC 偏移形式，再按第一个正偏移符号截去时区部分，保持现有
    本地筛选协议。该辅助函数不应替代需要严格时区换算的服务端业务日期逻辑。
    """
    try:
        parsed = datetime.fromisoformat(str(text).replace("Z", "+00:00"))
    except (TypeError, ValueError):
        return None
    if parsed.tzinfo is not None:
        # 统一剥离时区并返回 naive 时间；兼容负偏移（如 -05:00），避免与 naive 时间比较时抛 TypeError。
        parsed = parsed.replace(tzinfo=None)
    return parsed


def _report_rows(items: list[dict[str, object]]) -> tuple[dict[str, object], dict[str, object]]:
    """把任务记录转换为明细行和按业务模块聚合的统计。"""

    from collections import Counter

    feature_stats: dict[str, Counter] = {}
    details: list[dict[str, object]] = []
    for item in items:
        feature = _text(item.get("feature")) or "其他"
        status = _text(item.get("status"))
        stats = feature_stats.setdefault(feature, Counter())
        stats["total"] += 1
        if status in DONE_STATUSES:
            stats["done"] += 1
        elif status in FAILED_STATUSES:
            stats["failed"] += 1
        details.append({
            "started_at": _text(item.get("started_at")),
            "feature": FEATURE_TITLES.get(feature, feature),
            "title": _text(item.get("title")),
            "status": STATUS_TITLES.get(status, status or "未知"),
            "files": int(item.get("files") or 0),
        })
    return feature_stats, details


def _style_report_header(worksheet, row_index: int, count: int, *, font, fill, border) -> None:
    """统一设置报表页签表头样式。"""

    for column in range(1, count + 1):
        cell = worksheet.cell(row_index, column)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border


def _write_report_summary(workbook, feature_stats, details, range_label, *, border, title_font, head_font, cell_font, head_fill) -> None:
    """写入汇总页，顶部展示总指标，下部展示模块分布。"""

    summary = workbook.create_sheet("汇总")
    for column, width in {"A": 22, "B": 14, "C": 14, "D": 14, "E": 14}.items():
        summary.column_dimensions[column].width = width
    total = len(details)
    done = sum(stats["done"] for stats in feature_stats.values())
    failed = sum(stats["failed"] for stats in feature_stats.values())
    summary.cell(1, 1, "峰运通业务报表").font = title_font
    summary.cell(2, 1, "统计范围：%s" % range_label).font = cell_font
    summary.cell(3, 1, "生成时间：%s" % datetime.now().strftime("%Y-%m-%d %H:%M")).font = cell_font
    for row, label, value in (
        (5, "任务总数", total),
        (6, "已完成", done),
        (7, "失败/中断", failed),
        (8, "成功率", "%.1f%%" % (done / total * 100 if total else 0)),
    ):
        summary.cell(row, 1, label).font = head_font
        summary.cell(row, 2, value)
        for column in (1, 2):
            summary.cell(row, column).border = border
    summary.cell(10, 1, "模块分布").font = head_font
    header_row = 11
    for column, name in enumerate(("模块", "任务数", "完成", "失败", "成功率"), start=1):
        summary.cell(header_row, column, name)
    _style_report_header(summary, header_row, 5, font=head_font, fill=head_fill, border=border)
    for row_index, (feature, stats) in enumerate(sorted(feature_stats.items()), start=header_row + 1):
        values = [
            FEATURE_TITLES.get(feature, feature),
            stats["total"],
            stats["done"],
            stats["failed"],
            "%.1f%%" % (stats["done"] / stats["total"] * 100 if stats["total"] else 0),
        ]
        for column, value in enumerate(values, start=1):
            cell = summary.cell(row_index, column, value)
            cell.font = cell_font
            cell.border = border


def _write_report_details(workbook, details, *, border, head_font, cell_font, head_fill) -> None:
    """写入客户可读明细页，不包含任务 ID 或绝对路径。"""

    detail = workbook.create_sheet("明细")
    for column, width in {"A": 19, "B": 14, "C": 40, "D": 10, "E": 10}.items():
        detail.column_dimensions[column].width = width
    for column, name in enumerate(("开始时间", "模块", "任务", "状态", "结果文件"), start=1):
        detail.cell(1, column, name)
    _style_report_header(detail, 1, 5, font=head_font, fill=head_fill, border=border)
    for index, row in enumerate(details, start=2):
        for column, value in enumerate((row["started_at"], row["feature"], row["title"], row["status"], row["files"]), start=1):
            cell = detail.cell(index, column, value)
            cell.font = cell_font
            cell.border = border


def _save_report_workbook(workbook, out_path: str) -> None:
    """在目标目录内保存临时文件并原子替换正式报表。"""

    parent = os.path.dirname(os.path.abspath(out_path))
    descriptor, temp_path = tempfile.mkstemp(prefix="report_", suffix=".xlsx", dir=parent)
    os.close(descriptor)  # openpyxl 需要自行打开路径，先释放 mkstemp 创建的 Windows 句柄。
    try:
        workbook.save(temp_path)
        os.replace(temp_path, out_path)
    finally:
        workbook.close()
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except OSError:
                # 清理失败不能覆盖真正的保存异常；临时文件前缀可供维护人员识别。
                pass


def build_report(items: list[dict[str, object]], out_path: str, range_label: str) -> int:
    """把已筛选任务列表汇总为两页 Excel，并返回写入的明细行数。

    每条任务提取开始时间、模块、标题、状态和结果文件数；按原始 feature 统计总量、
    完成和失败。汇总页展示总体指标及模块分布，明细页逐条保留客户可读名称。函数
    不创建父目录，也不执行权限或日期筛选，调用方必须先完成这些工作。
    """
    feature_stats, details = _report_rows(items)

    workbook = openpyxl.Workbook()
    # 删除默认空页，确保报告只有明确命名的“汇总”和“明细”。
    workbook.remove(workbook.active)
    # 两个页签复用样式对象，减少重复样式记录并保持视觉一致。
    thin = Side(style="thin", color="9AA5B1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="EAF1FF")
    title_font = Font(name="宋体", size=14, bold=True)
    head_font = Font(name="宋体", size=10, bold=True)
    cell_font = Font(name="宋体", size=10)

    _write_report_summary(
        workbook, feature_stats, details, range_label,
        border=border, title_font=title_font, head_font=head_font,
        cell_font=cell_font, head_fill=head_fill,
    )
    _write_report_details(
        workbook, details, border=border, head_font=head_font,
        cell_font=cell_font, head_fill=head_fill,
    )
    _save_report_workbook(workbook, out_path)
    return len(details)


def unique_report_path(out_dir: str, range_label: str) -> str:
    """生成带范围标签和秒级时间戳的唯一报表路径。

    ``range_label`` 应由调用层提供客户可读且已校验的短文本；最终仍使用公共唯一路径
    规则避免同秒重复导出覆盖旧报告。
    """
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = "业务报表_%s_%s.xlsx" % (range_label, stamp)
    return common_core.unique_path(os.path.join(out_dir, filename))
