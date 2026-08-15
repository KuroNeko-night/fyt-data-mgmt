# -*- coding: utf-8 -*-
"""
通用 Excel 表格比对核心
=======================
用于核对程序输出与人工结果、两个数据版本或交接表。用户指定两表共有的关键列后，
模块按关键值配对记录，因此两份表的行顺序不同也不影响结果；输出同时包含单元格
差异、仅在单边出现的行、重复键和空键统计，并可生成带颜色标记的 Excel 报告。

比对前会做有边界的值归一化：普通数字与等价数字文本视为相同，文本去除首尾空白，
Decimal 数值采用微小容差；但 ``001`` 一类带前导零的编码保持文本身份，避免物料号
被错误折叠成数字 ``1``。重复关键值不会简单按行号硬配对，而是先消去内容完全一致
的记录，再以“差异列最少”为原则配对剩余记录，降低同键多行乱序造成的误报。

模块不依赖桌面或 Web 界面，表头探测、比对和导出均可独立测试。读取数据时使用
公共 ``load_data_only``，因此正式执行前会由入口检查公式缓存，提醒用户刷新公式值。
"""
import os
import numbers
import re
from decimal import Decimal, InvalidOperation

import openpyxl
from openpyxl.styles import PatternFill, Font

from . import paths as _paths
from . import settings as settings_mod
from . import common_core as _common       # load_data_only 加速读取

# 报告使用通用中文字体；颜色分别表达“内容不同”“仅单边存在”和普通表头。
FONT_NAME = "微软雅黑"
FILL_DIFF = PatternFill("solid", fgColor="FFC7CE")
FILL_ONLY = PatternFill("solid", fgColor="FFEB9C")
FILL_HEAD = PatternFill("solid", fgColor="D9E1F2")
# Decimal 比较仍保留容差，兼容 Excel 公式和二进制浮点产生的极小尾差。
FLOAT_TOL = 1e-9


def _norm_cell(v):
    """把单元格值归一化为适合内容比较、同时保留编码语义的对象。

    空值变为空串，布尔值保持布尔类型，数值和普通数字文本转为 ``Decimal``，其他
    文本只去首尾空白。带前导零的整数文本不转数字，因为它更可能是物料、人员或
    订单编码。千分位仅在数值识别时去除，返回原文本时仍保留用户看到的写法。
    """
    if v is None:
        return ""
    if isinstance(v, bool):
        # bool 是 int 的子类，必须在 numbers.Number 之前处理，避免 True 被当成 1。
        return v
    if isinstance(v, numbers.Number):
        try:
            # 先转字符串再构造 Decimal，避免直接吸收二进制浮点的长尾误差。
            return Decimal(str(v))
        except InvalidOperation:
            return str(v)
    s = str(v).strip()
    if s == "":
        return ""
    numeric = s.replace(",", "")
    # “001”“0008”常是编码而非数量，不能与 1/8 混同；普通“10.0”仍可数值化。
    if re.fullmatch(r"[+-]?\d+(?:\.\d+)?", numeric):
        digits = numeric.lstrip("+-")
        if not (digits.startswith("0") and len(digits.split(".", 1)[0]) > 1):
            try:
                return Decimal(numeric)
            except InvalidOperation:
                pass
    return s


def _eq(a, b):
    """判断两个归一化值是否等价；仅数值类型应用统一容差。"""
    if isinstance(a, Decimal) and isinstance(b, Decimal):
        return abs(a - b) <= Decimal(str(FLOAT_TOL))
    return a == b


def _key_str(v):
    """将关键列值转换为稳定的索引字符串。

    Excel 数值 ``10``、``10.0`` 统一为 ``"10"``，避免存储类型差异把同一记录拆到
    两边；文本值只去首尾空白，因此 ``"001"`` 和 ``"1,234"`` 的编码语义不变。
    布尔值单独处理，防止它沿数值分支变成 ``1`` 或 ``0``。
    """
    if v is None:
        return ""
    if isinstance(v, bool):
        return str(v)
    if isinstance(v, numbers.Number):
        f = float(v)
        # 整数浮点去掉“.0”；非整数使用 repr，获得适合索引的稳定短表示。
        return str(int(f)) if f.is_integer() else repr(f)
    return str(v).strip()


def _open_ws(path, sheet=None):
    """打开指定页签并返回工作簿与工作表，由调用方负责关闭工作簿。"""
    # 比对只看计算后的值，不改样式或公式；公共加载器还能跳过无关的透视缓存解析。
    wb = _common.load_data_only(path)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    return wb, ws


def _cell_text(v):
    """取得表头探测用的简单文本，不执行数字归一化或编码推断。"""
    return "" if v is None else str(v).strip()


def _is_numeric_text(s):
    """判断文本去除千分位后是否可转数字，用于区分数据行与表头标签。"""
    try:
        float(s.replace(",", ""))
        return True
    except (ValueError, AttributeError):
        return False


def _looks_like_header(texts):
    """用轻量统计特征判断一组非空文本是否像字段表头。

    合格候选至少两列，且不少于七成是 25 字以内的短标签、不多于一半像数字、
    不少于七成互不重复。这些阈值用于排除填满单元格的标题横幅、元信息和正文数据，
    不试图理解具体业务词汇，因此可供任意表格比对使用。
    """
    n = len(texts)
    if n < 2:
        return False
    short = sum(1 for t in texts if len(t) <= 25)
    numeric = sum(1 for t in texts if _is_numeric_text(t))
    uniq = len(set(texts))
    return (short / n >= 0.7          # 长句比例过高时更像标题或说明。
            and numeric / n <= 0.5    # 数字占多数时更像正文数据。
            and uniq / n >= 0.7)      # 大量重复值常见于合并标题或占位行。


def _best_header_index(rows):
    """从候选行中选择表头并返回 0 基下标，空输入返回 ``None``。

    首选所有“像表头”候选中非空字段最多的一行；完全没有候选时回退到全局非空最多
    的旧规则，以兼容缺少规范表头的历史导出。``read_table`` 和 ``read_headers``
    共用此函数，确保前端关键列下拉与正式比对读取的是同一行。
    """
    best_i, best_cnt = None, -1  # 兼容回退：所有行中的最大非空数。
    hdr_i, hdr_cnt = None, 0     # 首选规则：表头候选中的最大非空数。
    for i, rw in enumerate(rows):
        texts = [t for v in rw for t in (_cell_text(v),) if t != ""]
        cnt = len(texts)
        if cnt > best_cnt:
            best_cnt, best_i = cnt, i
        if _looks_like_header(texts) and cnt > hdr_cnt:
            hdr_cnt, hdr_i = cnt, i
    return hdr_i if hdr_i is not None else best_i


def _detect_header_row(ws, scan_rows=15):
    """扫描工作表前若干行并返回 openpyxl 使用的 1 基表头行号。"""
    limit = min(scan_rows, ws.max_row or 0)
    if limit < 1:
        return None
    rows = [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            for r in range(1, limit + 1)]
    idx = _best_header_index(rows)
    return None if idx is None else idx + 1


def read_table(path, sheet=None, scan_rows=15):
    """读取一张工作表，返回唯一表头列表和按表头映射的数据行。

    空表返回两个空列表。空表头按列号命名为“列N”；重复名称从第二次出现起添加
    ``(2)``、``(3)`` 后缀，保证行字典不会因重名列静默丢值。表头后整行为空的
    记录会跳过，其他行保留原始单元格值，具体等价判断统一留给 ``compare``。
    """
    wb, ws = _open_ws(path, sheet)
    try:
        hr = _detect_header_row(ws, scan_rows)
        if hr is None:
            return [], []
        headers, seen = [], {}
        for c in range(1, ws.max_column + 1):
            name = str(ws.cell(hr, c).value).strip() if ws.cell(hr, c).value is not None else ""
            if name == "":
                name = "列%d" % c
            if name in seen:
                # 字典无法容纳两个同名键，稳定后缀可同时保留两列并供用户单独选择。
                seen[name] += 1
                name = "%s(%d)" % (name, seen[name])
            else:
                seen[name] = 1
            headers.append(name)
        rows = []
        for r in range(hr + 1, ws.max_row + 1):
            vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            # 用比较归一化判断空行，纯空白文本也视为空，避免产生无意义记录。
            if all(_norm_cell(v) == "" for v in vals):
                continue
            rows.append({headers[i]: vals[i] for i in range(len(headers))})
        return headers, rows
    finally:
        wb.close()


def read_headers(path, sheet=None, scan_rows=15):
    """以只读模式取得唯一表头列表，供界面快速生成关键列和比较列选项。

    本函数只遍历前 ``scan_rows`` 行，不读取完整数据区。它与正式读取共享表头选择和
    重名处理规则，防止界面显示的列名在执行时找不到。
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        # 某些第三方导出把工作表 dimension 错写为单格，read_only 因而只返回一行。
        # 重置维度可按实际 XML 内容重新遍历；后续仍限制扫描行数，开销保持可控。
        try:
            ws.reset_dimensions()
        except Exception:
            pass
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= scan_rows:
                break
            rows.append(row)
        if not rows:
            return []
        idx = _best_header_index(rows)
        best = rows[idx] if idx is not None else rows[0]
        headers, seen = [], {}
        for c, v in enumerate(best, 1):
            name = str(v).strip() if v is not None else ""
            if name == "":
                name = "列%d" % c
            if name in seen:
                seen[name] += 1
                name = "%s(%d)" % (name, seen[name])
            else:
                seen[name] = 1
            headers.append(name)
        return headers
    finally:
        wb.close()


def common_columns(headers_a, headers_b):
    """返回两表交集列，并保持 A 表顺序供界面稳定展示。"""
    setb = set(headers_b)
    return [h for h in headers_a if h in setb]


def _index_by_key(rows, key):
    """按归一化关键值建立一对多索引，并单独收集空键行。

    值为行列表而非单行，是因为订单号、物料号等业务键可能合法重复；重复组会在
    ``compare`` 中继续逐行配对，不能在建索引时覆盖或丢弃。
    """
    idx, blank = {}, []
    for row in rows:
        k = row.get(key)
        ks = _key_str(k)
        if ks == "":
            blank.append(row)
            continue
        idx.setdefault(ks, []).append(row)
    return idx, blank


def compare(headers_a, rows_a, headers_b, rows_b, key, columns=None, log=None):
    """按关键列和内容相似度配对两表记录，返回结构化差异。

    ``key`` 必须同时存在；``columns`` 默认取两表公共列并排除关键列。返回值中的
    ``diffs`` 是配对行内不同的单元格，``only_a``/``only_b`` 是无法找到对家的行，
    ``dup_a``/``dup_b`` 记录各表重复键，空键行只计数不参与配对。

    对每个关键值的重复组分两轮处理：第一轮消去所有比较列完全一致的行，解决同键
    记录乱序；第二轮把剩余 A 行与差异列数量最少的 B 行配对。该策略是局部贪心，
    目标是减少明显误报而不是推断业务上的唯一对应关系，因此重复键仍会单独提示。
    """
    if key not in headers_a or key not in headers_b:
        raise ValueError("关键列「%s」需同时存在于两份表" % key)
    if columns is None:
        # 默认只比较两表都具备的字段，避免版本新增列被误判为每行差异。
        columns = [c for c in common_columns(headers_a, headers_b) if c != key]
    missing_columns = [c for c in columns if c not in headers_a or c not in headers_b]
    if missing_columns:
        raise ValueError("比较列需同时存在于两份表：%s" % "、".join(missing_columns))

    idx_a, blank_a = _index_by_key(rows_a, key)
    idx_b, blank_b = _index_by_key(rows_b, key)
    dup_a = sorted(k for k, v in idx_a.items() if len(v) > 1)
    dup_b = sorted(k for k, v in idx_b.items() if len(v) > 1)

    diffs, only_a, only_b = [], [], []
    matched = 0
    for k in sorted(set(idx_a) | set(idx_b)):
        group_a, group_b = list(idx_a.get(k, [])), list(idx_b.get(k, []))
        # 先消去完全相同的记录；使用 B 的位置列表，才能正确处理内容也重复的多行。
        remaining_b = list(range(len(group_b)))
        paired = []
        for ra in group_a:
            exact = next((pos for pos in remaining_b
                          if all(_eq(_norm_cell(ra.get(col)),
                                     _norm_cell(group_b[pos].get(col)))
                                 for col in columns)), None)
            if exact is not None:
                # 一个 B 行只能消费一次，否则多个 A 重复行会错误配到同一条记录。
                paired.append((ra, group_b[exact]))
                remaining_b.remove(exact)
            else:
                paired.append((ra, None))
        leftovers_a = []
        for ra, rb in paired:
            if rb is None:
                leftovers_a.append(ra)
            else:
                matched += 1
        # 未精确命中的同键行选择差异列最少的剩余对家；没有对家才是真正单边记录。
        for ra in leftovers_a:
            if not remaining_b:
                only_a.append({"key": k, "row": ra})
                continue
            best = min(remaining_b, key=lambda pos: sum(
                not _eq(_norm_cell(ra.get(col)), _norm_cell(group_b[pos].get(col)))
                for col in columns))
            rb = group_b[best]
            remaining_b.remove(best)
            matched += 1
            for col in columns:
                va, vb = ra.get(col), rb.get(col)
                if not _eq(_norm_cell(va), _norm_cell(vb)):
                    diffs.append({"key": k, "column": col, "a": va, "b": vb})
        for pos in remaining_b:
            only_b.append({"key": k, "row": group_b[pos]})

    if log:
        log("· 比对完成:差异 %d 处,只在A %d 行,只在B %d 行"
            % (len(diffs), len(only_a), len(only_b)))
        if dup_a or dup_b:
            log("⚠ 关键列有重复值(A:%d B:%d),重复键只按首条比对"
                % (len(dup_a), len(dup_b)))
        if blank_a or blank_b:
            log("⚠ 关键列为空的行已跳过(A:%d B:%d)" % (len(blank_a), len(blank_b)))

    counts = {"diffs": len(diffs), "only_a": len(only_a), "only_b": len(only_b),
              "dup_a": len(dup_a), "dup_b": len(dup_b),
              "blank_a": len(blank_a), "blank_b": len(blank_b),
               "matched": matched}
    return {"diffs": diffs, "only_a": only_a, "only_b": only_b,
            "dup_a": dup_a, "dup_b": dup_b, "blank_a": len(blank_a),
            "blank_b": len(blank_b), "columns": columns, "key": key,
            "counts": counts}


def _style_header(ws, headers):
    """写入报告表头并统一应用中文字体、粗体和浅蓝底色。"""
    for c, name in enumerate(headers, 1):
        cell = ws.cell(1, c, name)
        cell.font = Font(name=FONT_NAME, bold=True)
        cell.fill = FILL_HEAD


def export_report(result, out_dir=None, out_name="差异报告.xlsx", log=None):
    """将结构化比对结果导出为四页带颜色提示的 Excel 报告。

    “概要”保存统计口径，“差异明细”逐单元格列出 A/B 值并标红，两个单边页保留
    对应原始整行并标黄。没有单边记录时仍创建只有关键列表头的空页，使报告结构
    固定，便于人工查看和后续程序读取。
    """
    # 只取文件名，拒绝调用方通过 ``out_name`` 携带目录片段越出输出目录。
    out_name = os.path.basename(out_name) or "差异报告.xlsx"
    out_dir = out_dir or os.getcwd()
    if not os.path.isdir(out_dir):
        # 允许调用方传入尚不存在的业务输出目录，避免保存时才抛出路径错误。
        os.makedirs(out_dir)
    path = os.path.join(out_dir, out_name)
    wb = openpyxl.Workbook()

    # 概要页只展示统计，不复制所有明细，适合首先判断两份表的总体一致性。
    ws = wb.active; ws.title = "概要"
    _style_header(ws, ["项目", "数量"])
    cn = result["counts"]
    for name, val in [("关键列", result["key"]), ("比较列数", len(result["columns"])),
                      ("值差异(单元格)", cn["diffs"]), ("配对成功(行)", cn["matched"]),
                      ("只在A的行", cn["only_a"]), ("只在B的行", cn["only_b"]),
                      ("A重复键", cn["dup_a"]), ("B重复键", cn["dup_b"]),
                      ("A关键列空行", cn["blank_a"]), ("B关键列空行", cn["blank_b"])]:
        ws.append([name, val])

    # 一个差异单元格占一行，同一业务记录可能对应多行，便于按列筛选问题。
    ws = wb.create_sheet("差异明细")
    _style_header(ws, [result["key"], "列名", "A 值", "B 值"])
    for d in result["diffs"]:
        ws.append([d["key"], d["column"], d["a"], d["b"]])
        for c in (3, 4):
            ws.cell(ws.max_row, c).fill = FILL_DIFF

    # 单边页以第一条记录的键顺序作为列顺序；读取阶段已经保证所有行字段一致。
    for title, items in [("只在A", result["only_a"]), ("只在B", result["only_b"])]:
        ws = wb.create_sheet(title)
        cols = list(items[0]["row"].keys()) if items else [result["key"]]
        _style_header(ws, cols)
        for it in items:
            ws.append([it["row"].get(c) for c in cols])
            for c in range(1, len(cols) + 1):
                ws.cell(ws.max_row, c).fill = FILL_ONLY

    wb.save(path)
    if log:
        log("· 报告已生成:%s" % path)
    return path


def run(file_a, file_b, key, sheet_a=None, sheet_b=None, columns=None,
        out_dir=None, log=None, progress=None):
    """执行“读取两表、内容比对、导出报告”的完整业务流程。

    ``sheet_a``、``sheet_b`` 可固定页签，``columns`` 可限制比较字段；未指定输出目录
    时遵循全局输出设置。进度权重为读取 40%、比对 35%、导出 25%，返回的原始
    结构中追加 ``report_path`` 和 ``out_dir``，供双端直接展示并提供下载。
    """
    from . import common_core
    # 三段任务都可能受文件规模影响，显式分段可避免长时间停留在同一进度值。
    prog = common_core.Progress(progress, stages=[
        ("read", 40), ("compare", 35), ("export", 25)])
    if log:
        # data_only 读取公式缓存；缓存缺失会把公式结果当空值，必须在比对前提示用户。
        common_core.warn_if_uncached(file_a, log, sheet_a, what="比对数据")
        common_core.warn_if_uncached(file_b, log, sheet_b, what="比对数据")
    prog.stage("read")
    ha, ra = read_table(file_a, sheet_a)
    prog.tick(1, 2)
    hb, rb = read_table(file_b, sheet_b)
    prog.tick(2, 2)
    if log:
        log("· A《%s》%d 行,B《%s》%d 行"
            % (os.path.basename(file_a), len(ra), os.path.basename(file_b), len(rb)))
    prog.stage("compare")
    result = compare(ha, ra, hb, rb, key, columns=columns, log=log)
    # 未显式指定时走全局目录策略，与其他业务统一支持固定目录、按时间及源旁输出。
    if out_dir is None:
        st = settings_mod.get_settings()
        out_dir = _paths.resolve_output_dir("compare", **st.output_kwargs())
    stem = "差异报告_%s_vs_%s.xlsx" % (
        os.path.splitext(os.path.basename(file_a))[0],
        os.path.splitext(os.path.basename(file_b))[0])
    prog.stage("export")
    report = export_report(result, out_dir=out_dir, out_name=stem, log=log)
    result["report_path"] = report
    result["out_dir"] = os.path.dirname(report)
    prog.done()
    return result
