# -*- coding: utf-8 -*-
"""
数据库（自带存储）—— 自动分类 + 归档 + 索引
==============================================
用户把各处来的 Excel 拖进来，程序据"文件名 + 表头"双阶段评分自动判定用途，
复制归档到 <文档>/峰运通数据管理系统/数据库/<类别>/，记录最后更新日期等元信息；
未能识别的表统一进"未识别"文件夹。各功能页可据类别自动取到所需的表。

类别（10 类 + 未识别）：
  att_source   填报·系统数据表（打卡来源）
  att_target   填报·待填考勤表
  rec_source   对账·数据来源（工时明细）
  rec_zong     对账·待对总表
  rec_labor    对账·劳务对账单
  pivot_src    透视·采购数据表
  arrival_plan 到料·送货计划表
  purchase_stmt 采购·采购对账单（我方/供方通用）
  deliv_bom    送货·物料清单（KD/SUB 物料清单）
  deliv_supp   送货·供应商明细（含供应商代码/名称）
  unknown      未识别

兼容 Windows 10/11 + Python 3.13。
"""
import os
import re
import json
import shutil
import datetime

from . import paths
from .storage_lock import file_lock

# 类别顺序（也用于页面展示顺序）
CATEGORIES = ["att_source", "att_target", "rec_source", "rec_zong",
              "rec_labor", "pivot_src", "arrival_plan",
              "purchase_stmt", "deliv_bom", "deliv_supp"]
UNKNOWN = "unknown"

CATEGORY_TITLES = {
    "att_source": "填报 · 系统数据表",
    "att_target": "填报 · 待填考勤表",
    "rec_source": "对账 · 数据来源",
    "rec_zong": "对账 · 待对总表",
    "rec_labor": "对账 · 劳务对账单",
    "pivot_src": "透视 · 采购数据表",
    "arrival_plan": "到料 · 送货计划表",
    "purchase_stmt": "采购 · 采购对账单",
    "deliv_bom": "送货 · 物料清单",
    "deliv_supp": "送货 · 供应商明细",
    "unknown": "未识别",
}
# 类别 → 归档子文件夹名（中文，便于用户直接翻看）
CATEGORY_DIRS = dict(CATEGORY_TITLES)
CATEGORY_DIRS["unknown"] = "未识别"


EXCEL_EXT = (".xlsx", ".xlsm", ".xls")


def _norm(s):
    """去空白/制表符，便于表头匹配。"""
    return re.sub(r"\s+", "", str(s)) if s is not None else ""


def scan_headers(path, max_rows=15, max_cells=60, log=None):
    """轻量读取：只取每个子表前若干行的单元格文本，用于表头识别。
    返回 {sheet_name: set(归一化文本)}。尽量快、低内存。
    传 log 时，整表打不开(损坏/加密/被占用)会记一条告警而非静默返回空
    ——否则文件读不出会被当成"无法识别"归入 unknown，用户无从分辨。"""
    ext = os.path.splitext(path)[1].lower()
    out = {}
    try:
        if ext in (".xlsx", ".xlsm"):
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            try:
                for ws in wb.worksheets:
                    # 有些文件声明的 <dimension> 不准，read_only 会漏读；
                    # reset_dimensions 让 openpyxl 按实际内容重算范围。
                    try:
                        ws.reset_dimensions()
                    except Exception:
                        pass
                    toks = set()
                    for i, row in enumerate(ws.iter_rows(values_only=True)):
                        if i >= max_rows:
                            break
                        for v in row[:max_cells]:
                            t = _norm(v)
                            if t:
                                toks.add(t)
                    out[ws.title] = toks
            finally:
                wb.close()
        elif ext == ".xls":
            import xlrd
            book = xlrd.open_workbook(path)
            for sh in book.sheets():
                toks = set()
                for r in range(min(max_rows, sh.nrows)):
                    for c in range(min(max_cells, sh.ncols)):
                        t = _norm(sh.cell(r, c).value)
                        if t:
                            toks.add(t)
                out[sh.name] = toks
    except Exception as e:
        # 整表读取失败:记告警(带异常类型),让"读不出"区别于"识别不了"。
        if log:
            try:
                log("⚠ 无法读取 %s(%s: %s),将按未识别处理"
                    % (os.path.basename(path), type(e).__name__, e))
            except Exception:
                pass
    return out


def _any_tok(tokens, needles):
    """tokens 中是否有任一包含 needle 子串。"""
    for tk in tokens:
        for nd in needles:
            if nd in tk:
                return True
    return False


def _score_sheet(fname, tokens, ext):
    """对单个子表(tokens)结合文件名给各类别打分。返回 {cat: (score, [signals])}。"""
    fn = _norm(fname)
    res = {}

    def has(*needles):
        return _any_tok(tokens, needles)

    def fnhas(*needles):
        return any(nd in fn for nd in needles)

    # att_source —— 唯一强特征"上班1打卡"
    s, sig = 0, []
    if has("上班1打卡"):
        s += 70; sig.append("含“上班1打卡”列")
    if has("姓名"):
        s += 12; sig.append("含“姓名”")
    if fnhas("打卡", "考勤机", "每日统计", "系统导出"):
        s += 15; sig.append("文件名含打卡/统计")
    res["att_source"] = (s, sig)

    # att_target —— 休息时间 + 系统/实际 列(且非 att_source)
    s, sig = 0, []
    if not has("上班1打卡"):
        if has("休息时间"):
            s += 42; sig.append("含“休息时间”列")
        if has("实际工作时间", "实际工时"):
            s += 22; sig.append("含“实际工作时间”")
        if has("上班时间", "下班时间"):
            s += 16; sig.append("含上/下班时间列")
        if has("姓名") and has("日期"):
            s += 12; sig.append("含姓名+日期")
        if fnhas("考勤表", "待填", "考勤"):
            s += 12; sig.append("文件名含考勤表")
    res["att_target"] = (s, sig)

    # rec_source —— 姓名/日期/实际工时 明细,但无 att_target 富特征;
    # 文件名含"对账单/劳务/结算"的是劳务方账单(rec_labor),不是本方来源,排除。
    s, sig = 0, []
    if (not has("上班1打卡") and not has("休息时间") and not has("所属劳务公司")
            and not fnhas("对账单", "劳务", "结算")):
        if has("实际工作时间", "实际工时"):
            s += 46; sig.append("含“实际工作时间”明细")
        if has("姓名") and has("日期"):
            s += 22; sig.append("含姓名+日期")
        if fnhas("明细", "工时", "已填写"):
            s += 10; sig.append("文件名含明细/工时")
    res["rec_source"] = (s, sig)

    # rec_zong —— 待对总表
    s, sig = 0, []
    if has("所属劳务公司"):
        s += 55; sig.append("含“所属劳务公司”")
    if has("出勤工时"):
        s += 30; sig.append("含“出勤工时”")
    if has("对账时间"):
        s += 28; sig.append("含“对账时间”")
    if s and has("姓名"):
        s += 10
    if fnhas("总表", "待对"):
        s += 10; sig.append("文件名含总表/待对")
    res["rec_zong"] = (s, sig)

    # rec_labor —— 劳务对账单(格式各异,弱特征)
    s, sig = 0, []
    if fnhas("劳务", "对账单", "工时单", "结算"):
        s += 42; sig.append("文件名含劳务/对账单")
    if has("姓名"):
        s += 18; sig.append("含“姓名”")
    if has("合计", "小计"):
        s += 14; sig.append("含合计列")
    if ext == ".xls":
        s += 8
    res["rec_labor"] = (s, sig)

    # pivot_src —— 采购数据表
    s, sig = 0, []
    if has("最终采购数量"):
        s += 45; sig.append("含“最终采购数量”")
    if has("材料编号", "物料编号", "物料号", "料号", "物料编码"):
        s += 30; sig.append("含材料/物料编号")
    if fnhas("采购量核算", "pfep", "包装方案", "组托辅材", "包材用量", "组托"):
        s += 35; sig.append("文件名含采购/PFEP/组托")
    if has("规格") and has("单位"):
        s += 10; sig.append("含规格+单位")
    res["pivot_src"] = (s, sig)

    # arrival_plan —— 送货计划表
    s, sig = 0, []
    if fnhas("送货计划"):
        s += 55; sig.append("文件名含“送货计划”")
    if has("剩余未收", "未收数", "未收货", "未收料"):
        s += 30; sig.append("含未收料列")
    if has("供应商"):
        s += 14; sig.append("含“供应商”")
    if has("编码", "物料编码", "物料编号"):
        s += 10
    res["arrival_plan"] = (s, sig)

    # purchase_stmt —— 采购对账单(我方/供方通用)。
    # 强特征:批次号+采购数量+材料编号三件套;排除透视表(有"最终采购数量")与
    # 含姓名/工时的劳务账单,避免抢 pivot_src / rec_labor。
    s, sig = 0, []
    if (has("批次号") and has("采购数量") and not has("最终采购数量")
            and not has("姓名") and not has("实际工时", "出勤工时")):
        s += 55; sig.append("含批次号+采购数量")
        if has("材料编号", "物料编号", "材料号", "物料号"):
            s += 25; sig.append("含材料编号")
        if has("材料名称", "物料名称"):
            s += 10; sig.append("含材料名称")
        if fnhas("对账单", "对单", "结算单"):
            s += 12; sig.append("文件名含对账/对单")
    res["purchase_stmt"] = (s, sig)

    # deliv_bom —— 送货·物料清单。以"物料中文/英文描述"为唯一强特征(透视表只有物料号)。
    s, sig = 0, []
    if has("物料中文描述", "物料英文描述", "中文描述", "英文描述"):
        s += 55; sig.append("含物料中/英文描述")
        if (has("物料号", "物料编码") and has("数量")):
            s += 25; sig.append("含物料号+数量")
        if fnhas("物料清单", "bom", "kd", "sub"):
            s += 12; sig.append("文件名含物料清单/KD")
    res["deliv_bom"] = (s, sig)

    # deliv_supp —— 送货·供应商明细。两条判据(取高者):
    #   A) 经典表:唯一强特征"零部件代码";
    #   B) SAP KD 表:用"下阶物料/物料"作编码,靠 供应商代码+供应商名称 双列定性,
    #      并要求带 库区/结算/科室 等供应商表特征列。排除到料"送货计划"(文件名)与
    #      劳务单(姓名),避免抢 arrival_plan / rec_labor。
    s, sig = 0, []
    if has("零部件代码"):
        s += 50; sig.append("含零部件代码")
        if has("供应商代码", "供应商名称", "供应商"):
            s += 25; sig.append("含供应商代码/名称")
        if has("库区", "结算方式", "属性", "订单情况"):
            s += 12; sig.append("含库区/结算/属性列")
    elif (has("供应商代码") and has("供应商名称") and not has("姓名")
          and not fnhas("送货计划")):
        # 供应商代码+名称双列同现,基本只出现在供应商明细/缺账表
        s += 55; sig.append("含供应商代码+供应商名称双列")
        if has("下阶物料", "下阶物料描述", "物料", "料号"):
            s += 12; sig.append("含物料/下阶物料列")
        if has("库区", "结算方式", "科室", "订单情况", "发货数量", "签收数量", "批次号"):
            s += 12; sig.append("含库区/结算/发签数等供应商表列")
    res["deliv_supp"] = (s, sig)

    return res


ACCEPT_THRESHOLD = 50          # 最高分低于此值 → 未识别


def classify(path, log=None):
    """对一个文件分类。返回 dict：category / confidence / signals / sheet。
    传 log 时，文件整体读取失败会记告警(见 scan_headers)。"""
    ext = os.path.splitext(path)[1].lower()
    fname = os.path.basename(path)
    sheets = scan_headers(path, log=log)
    if not sheets:
        sheets = {"": set()}
    best = {"category": UNKNOWN, "score": 0, "signals": [], "sheet": ""}
    # 每个类别记录其"最高分子表"——支持一个文件多子表分属不同功能(多标签)。
    per_cat = {}          # cat -> {"score", "signals", "sheet"}
    for sname, tokens in sheets.items():
        scored = _score_sheet(fname, tokens, ext)
        for cat, (sc, sig) in scored.items():
            if sc > best["score"]:
                best = {"category": cat, "score": sc, "signals": sig, "sheet": sname}
            if cat not in per_cat or sc > per_cat[cat]["score"]:
                per_cat[cat] = {"score": sc, "signals": sig, "sheet": sname}
    if best["score"] < ACCEPT_THRESHOLD:
        conf = int(best["score"])          # 记录原始分供参考
        return {"category": UNKNOWN, "confidence": conf, "categories": [],
                "signals": best["signals"], "sheet": best["sheet"], "sheets": {}}
    # 所有达到阈值的类别都作为标签,按分数降序;主类别 = 最高分(=best,首位)。
    labels = sorted((c for c, v in per_cat.items() if v["score"] >= ACCEPT_THRESHOLD),
                    key=lambda c: per_cat[c]["score"], reverse=True)
    # 各标签命中的子表名(供多子表分属不同功能时,读取器定位用)。
    sheet_map = {c: per_cat[c]["sheet"] for c in labels}
    return {"category": best["category"], "confidence": min(100, int(best["score"])),
            "signals": best["signals"], "sheet": best["sheet"],
            "categories": labels, "sheets": sheet_map}


# ---------------- 索引读写 ----------------
def _load_index():
    p = paths.library_index_path()
    try:
        with open(p, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, dict) and "items" in data:
            return data
    except Exception:
        pass
    return {"items": []}


def _save_index(idx):
    """原子写索引：临时文件 + os.replace，避免写一半坏掉索引。
    失败返回 False 上报调用方（不再静默吞异常导致孤立文件）。"""
    p = paths.library_index_path()
    tmp = p + ".tmp"
    try:
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(idx, f, ensure_ascii=False, indent=2)
            f.flush()
            try:
                os.fsync(f.fileno())
            except OSError:
                pass
        os.replace(tmp, p)
        return True
    except Exception:
        try:
            if os.path.exists(tmp):
                os.remove(tmp)
        except OSError:
            pass
        return False


def _cat_dir(category):
    d = os.path.join(paths.library_dir(), CATEGORY_DIRS.get(category, category))
    paths._ensure(d)
    return d


def _now_str():
    return datetime.datetime.now().strftime("%Y-%m-%d %H:%M")


def list_items(category=None):
    """列出库中条目（可按类别过滤）。返回列表，每项含 name/category/path/updated/size/confidence/signals。"""
    items = _load_index()["items"]
    if category:
        # 多标签:命中主类别或任一附加标签都算(旧条目无 categories 时回退到 category)。
        items = [it for it in items
                 if category in (it.get("categories") or [it.get("category")])]
    return items


def _item_cats(it):
    """条目的全部类别标签(兼容旧索引:无 categories 时用单个 category)。"""
    return it.get("categories") or [it.get("category", UNKNOWN)]


def counts():
    """各类别条目数量 {category: n}，含 unknown。多标签文件在其每个标签下各计一次
    (与"从数据库选择"里的可见性一致)。"""
    c = {cat: 0 for cat in CATEGORIES}
    c[UNKNOWN] = 0
    for it in _load_index()["items"]:
        for cat in _item_cats(it):
            c[cat] = c.get(cat, 0) + 1
    return c


def storage_stats():
    """归档存储概况：返回 (条目数, 占用字节数)。

    只统计索引在册的归档表（count 与 size 始终一致）；孤立残留文件不计入，
    避免出现"0 张表却占用 X MB"这类自相矛盾的展示。"""
    items = _load_index()["items"]
    total = 0
    for it in items:
        p = it.get("path", "")
        try:
            if p and os.path.isfile(p):
                total += os.path.getsize(p)
        except OSError:
            pass
    return len(items), total


def human_size(nbytes):
    """把字节数格式化成 B/KB/MB/GB 便于展示。"""
    n = float(nbytes)
    for unit in ("B", "KB", "MB", "GB"):
        if n < 1024 or unit == "GB":
            return ("%.0f %s" % (n, unit)) if unit == "B" else ("%.1f %s" % (n, unit))
        n /= 1024
    return "%.1f GB" % n


def import_file(path, log=None):
    """分类并复制归档一个文件。同类同名→新版替换旧版(旧的存 .bak)。
    返回条目 dict（含 category/confidence 等）。不删除源文件（由调用方决定）。"""
    def _lg(m):
        if log:
            log(m)
    info = classify(path, log=_lg)
    cat = info["category"]
    fname = os.path.basename(path)
    dst_dir = _cat_dir(cat)
    dst = os.path.join(dst_dir, fname)

    with file_lock(paths.library_index_path()):
        idx = _load_index()
        items = idx["items"]
        replaced = False
        bak = dst + ".bak"
        if os.path.exists(dst):
            if os.path.exists(bak):
                os.remove(bak)
            # 只有备份成功才允许覆盖，保证任何失败路径都可恢复旧版本。
            shutil.copy2(dst, bak)
            replaced = True
        items = [it for it in items
                 if not (it.get("category") == cat and it.get("name") == fname)]

        part = dst + ".part"
        try:
            shutil.copy2(path, part)
            os.replace(part, dst)
            size = os.path.getsize(dst)
            cats = info.get("categories") or [cat]
            item = {
                "name": fname, "category": cat, "categories": cats, "path": dst,
                "updated": _now_str(), "size": size,
                "confidence": info["confidence"], "signals": info["signals"],
                "sheet": info.get("sheet", ""), "sheets": info.get("sheets", {}),
                "origin": os.path.abspath(path),
            }
            items.append(item)
            idx["items"] = items
            if _save_index(idx) is False:
                raise IOError("索引保存失败")
        except Exception:
            try:
                if os.path.exists(part):
                    os.remove(part)
                if replaced and os.path.exists(bak):
                    os.replace(bak, dst)
                elif os.path.exists(dst):
                    os.remove(dst)
            except OSError:
                pass
            raise IOError("索引或归档保存失败，已回滚：%s" % fname)
    _lg("%s → 【%s】可信度 %d%s"
        % (fname, CATEGORY_TITLES.get(cat, cat), info["confidence"],
           "（替换旧版）" if replaced else ""))
    return item


def import_many(pathlist, log=None):
    """批量导入。返回 [item,...]。"""
    out = []
    for p in pathlist:
        try:
            out.append(import_file(p, log=log))
        except Exception as e:
            if log:
                log("导入失败 %s：%s" % (os.path.basename(p), e))
    return out


def remove_item(category, name, delete_file=True):
    """从索引移除一项；delete_file 时连同归档文件与 .bak 一起删。"""
    with file_lock(paths.library_index_path()):
        idx = _load_index()
        keep, gone = [], []
        for it in idx["items"]:
            if category in _item_cats(it) and it.get("name") == name:
                gone.append(it)
            else:
                keep.append(it)
        idx["items"] = keep
        if gone and _save_index(idx) is False:
            raise IOError("数据库索引保存失败，未删除归档文件")
        if delete_file:
            for it in gone:
                for p in (it.get("path"), (it.get("path") or "") + ".bak"):
                    try:
                        if p and os.path.exists(p):
                            os.remove(p)
                    except OSError:
                        pass
        return len(gone)


def reclassify(category, name, new_category):
    """手动把一项改判到另一类别：移动归档文件并更新索引。"""
    if new_category not in CATEGORIES and new_category != UNKNOWN:
        return False
    with file_lock(paths.library_index_path()):
        idx = _load_index()
        for it in idx["items"]:
            if category in _item_cats(it) and it.get("name") == name:
                src = it.get("path", "")
                dst = os.path.join(_cat_dir(new_category), name)
                moved = False
                try:
                    # 源文件已丢失时不能只改索引，否则会产生“看得见但打不开”的条目。
                    if not src or not os.path.isfile(src):
                        return False
                    if os.path.abspath(src) != os.path.abspath(dst):
                        shutil.move(src, dst)
                        moved = True
                    it["category"] = new_category
                    it["categories"] = [new_category]
                    it["path"] = dst
                    it["updated"] = _now_str()
                    it["confidence"] = 100
                    it["signals"] = ["人工指定类别"]
                    if _save_index(idx) is False:
                        raise IOError("索引保存失败")
                    return True
                except Exception:
                    if moved:
                        try:
                            shutil.move(dst, src)
                        except OSError:
                            pass
                    return False
    return False


def latest_in(category):
    """取某类别中最近更新的一项 path（供功能页"载入最新"）。无则 None。"""
    items = sorted(list_items(category), key=lambda x: x.get("updated", ""), reverse=True)
    return items[0]["path"] if items else None
