# -*- coding: utf-8 -*-
"""
Excel 文件工具核心
==================
提供工作簿合并、按页签拆分、Excel/CSV 转换和同结构表纵向合并四类通用能力。
所有操作只依赖 openpyxl；旧版 ``.xls`` 读取按可选依赖 xlrd 降级处理，输出目录
统一交给 ``core.paths``，因此桌面端和 Web 端无需各自实现文件操作规则。

能力与格式边界：
  · ``merge_books``：每个源页签复制为结果中的独立页签；
  · ``split_sheets``：每个页签生成一个独立 xlsx；
  · ``convert``：旧 xls/CSV 转 xlsx，或 Excel 各页签转 UTF-8 BOM CSV；
  · ``stack_tables``：按规范化表头对齐同结构表，并增加来源文件列。

xlsx/xlsm 合并会复制单元格样式、行列尺寸、合并区域和冻结窗格，但 openpyxl 无法
可靠地把图表、图片等绘图对象跨工作簿搬运；拆分采用“复制整簿后删除其他页签”，
因此可保留这些对象。xls 和 CSV 只能按值读取，任何格式降级都应在日志中明确说明。

纯数据读取刻意不使用 openpyxl 的只读模式：部分外部系统把工作表使用范围错误写成
``A1``，只读模式会静默漏掉其余数据。公共加载器以常规模式读取并跳过无关透视缓存，
在正确性和性能之间取得更安全的平衡。
"""
import os
import csv
import copy as _copy

import openpyxl

from . import paths as _paths
from . import common_core as _common  # 提供安全的非只读计算值加载和统一文本清洗。

try:
    import xlrd  # 仅用于读取旧二进制 .xls，不参与 xlsx 写入。
    _HAS_XLRD = True
except Exception:
    # xlrd 是可选能力；现代 xlsx/CSV 操作不应因其缺失而无法启动。
    _HAS_XLRD = False


class ExcelToolError(Exception):
    """可直接转换为界面提示的文件工具业务异常。"""


def _safe_sheet_title(name, used):
    """生成符合 Excel 约束、且在当前输出工作簿内不重复的页签名。

    Excel 页签名最多 31 个字符，且不能包含方括号、冒号、星号、问号及正反斜杠。
    重名比较不区分大小写，
    因为 Excel 本身也按不区分大小写处理页签；冲突时追加递增后缀，并为后缀预留
    长度，确保最终名称仍不超过限制。
    """
    bad = set('[]:*?/\\')
    t = "".join("_" if c in bad else c for c in (name or "Sheet"))[:31] or "Sheet"
    base = t
    i = 1
    while t.lower() in used:
        suffix = "_%d" % i
        # 先截断基础名称再拼后缀，否则长文件名会让最终页签再次超过 31 字符。
        t = base[:31 - len(suffix)] + suffix
        i += 1
    used.add(t.lower())
    return t


def _read_sheets(path):
    """把一个支持的表格文件读取为 ``[(页签名, 行值列表)]``。

    xlsx/xlsm 通过公共 ``load_data_only`` 取得公式缓存值；不得改用 read_only，
    因为部分导出文件的 XML 使用范围错误标成 ``A1``，只读模式会信任该范围并静默
    漏掉整张表。xls 使用 xlrd，并专门把日期单元格从序列号恢复为 datetime；CSV
    复用编码探测函数并包装成单页签结构。未知扩展名抛出业务异常。
    """
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        wb = _common.load_data_only(path)
        out = []
        for ws in wb.worksheets:
            # 此辅助结构只承载值，不保留单元格坐标或样式；格式操作使用其他路径。
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
            out.append((ws.title, rows))
        wb.close()
        return out
    if ext == ".xls":
        if not _HAS_XLRD:
            raise ExcelToolError("未安装 xlrd,无法读取老式 .xls 文件")
        book = xlrd.open_workbook(path)
        out = []
        for sh in book.sheets():
            rows = []
            for r in range(sh.nrows):
                # xls 把日期存成受工作簿日期制影响的浮点序列号，必须结合 datemode
                # 逐格还原；转换失败时保留原值，比丢弃整行更利于人工发现异常。
                cells = []
                for c in range(sh.ncols):
                    cell = sh.cell(r, c)
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            cells.append(xlrd.xldate.xldate_as_datetime(
                                cell.value, book.datemode))
                        except Exception:
                            cells.append(cell.value)
                    else:
                        cells.append(cell.value)
                rows.append(cells)
            out.append((sh.name, rows))
        return out
    if ext == ".csv":
        return [(os.path.splitext(os.path.basename(path))[0], _read_csv(path))]
    raise ExcelToolError("不支持的文件类型:%s" % ext)


def _write_rows(ws, rows):
    """顺序写入二维行值；``None`` 行按空行处理，不复制任何格式。"""
    for r in rows:
        ws.append(list(r) if r is not None else [])


def _copy_sheet(src, dst):
    """把一个页签的值和基础格式复制到另一个工作簿的页签。

    复制范围包括单元格值、字体、填充、边框、对齐、保护、数字格式、行列尺寸、
    隐藏状态、合并区域、冻结窗格和网格线设置。样式对象不能在工作簿间共享内部
    引用，因此逐项浅复制为新对象。图表、图片、批注形状等绘图对象不在此函数能力
    范围内；需要保留它们的拆分操作采用整簿深复制方案。
    """
    for row in src.iter_rows():
        for c in row:
            d = dst.cell(row=c.row, column=c.column, value=c.value)
            if c.has_style:
                # 对每种样式分别复制，避免目标簿引用源簿的样式表索引。
                d.font = _copy.copy(c.font)
                d.fill = _copy.copy(c.fill)
                d.border = _copy.copy(c.border)
                d.alignment = _copy.copy(c.alignment)
                d.protection = _copy.copy(c.protection)
                d.number_format = c.number_format
    # 行列维度不是单元格样式的一部分，必须单独迁移。
    for key, dim in src.column_dimensions.items():
        nd = dst.column_dimensions[key]
        nd.width = dim.width
        nd.hidden = dim.hidden
        if dim.width is None and dim.hidden:
            # 某些隐藏列没有显式宽度，补 Excel 默认值可避免目标簿解除隐藏后宽度异常。
            nd.width = 8.43
    for idx, dim in src.row_dimensions.items():
        nd = dst.row_dimensions[idx]
        nd.height = dim.height
        nd.hidden = dim.hidden
    # 先复制普通单元格再重建合并区域，避免合并单元格占位对象干扰写值。
    for rng in list(src.merged_cells.ranges):
        dst.merge_cells(str(rng))
    # 视图层只保留业务最常用的冻结与网格线，其他窗口状态不跨簿复制。
    dst.freeze_panes = src.freeze_panes
    dst.sheet_view.showGridLines = src.sheet_view.showGridLines


def merge_books(files, out_dir=None, out_name="合并工作簿.xlsx",
                keep_formula=False, log=None):
    """把多个工作簿的所有页签合并进一个新工作簿。

    结果页签按“源文件名-原页签名”命名并安全去重。默认 ``keep_formula=False``，
    xlsx/xlsm 读取公式缓存值，可避免页签改名后跨页公式仍引用旧名称；开启后保留
    公式原文，仅能保证表内公式安全。现代格式复制基础样式，xls/CSV 则明确降级为
    仅数据。至少需要两个文件，返回输出文件、目录和统一 ``out_files`` 列表。
    """
    log = log or (lambda *_: None)
    if len(files) < 2:
        raise ExcelToolError("合并至少需要 2 个 Excel 文件")
    out_dir = out_dir or _paths.resolve_output_dir("excel_tools")
    wb = openpyxl.Workbook()
    # 新工作簿的默认空页没有业务意义，先删除再按源页签顺序创建。
    wb.remove(wb.active)
    used = set()
    total = 0
    for f in files:
        stem = os.path.splitext(os.path.basename(f))[0]
        ext = os.path.splitext(f)[1].lower()
        if ext in (".xlsx", ".xlsm"):
            # data_only 的取值与 keep_formula 相反：保留公式时读取公式文本，否则读缓存值。
            src_wb = openpyxl.load_workbook(f, data_only=not keep_formula)
            for sn in src_wb.sheetnames:
                title = _safe_sheet_title("%s-%s" % (stem, sn), used)
                ws = wb.create_sheet(title=title)
                _copy_sheet(src_wb[sn], ws)
                total += 1
            log("并入 %s(%d 个表,含格式%s)"
                % (os.path.basename(f), len(src_wb.sheetnames),
                   "、保留公式" if keep_formula else ""))
            src_wb.close()
        else:
            # 旧 xls 与 CSV 不具备可由 openpyxl 跨簿复制的样式模型，只写值并记录降级。
            sheets = _read_sheets(f)
            for name, rows in sheets:
                title = _safe_sheet_title("%s-%s" % (stem, name), used)
                _write_rows(wb.create_sheet(title=title), rows)
                total += 1
            log("并入 %s(%d 个表,仅数据)" % (os.path.basename(f), len(sheets)))
    out_file = os.path.join(out_dir, out_name)
    wb.save(out_file)
    log("已合并 %d 个文件、共 %d 个工作表 → %s" % (len(files), total, out_file))
    return {"out_file": out_file, "out_dir": out_dir, "out_files": [out_file]}


def split_sheets(file, out_dir=None, log=None):
    """把一个多页工作簿拆成每页一个 xlsx 文件。

    xlsx/xlsm 先完整载入一次，再为每个目标页深复制整本工作簿并删除其他页，这比
    跨簿复制更能保留图表、图片和复杂对象，也避免为每个页签反复从磁盘读取原文件。
    xls 只能按值重建。单页工作簿直接提示无需拆分；输出文件名会过滤 Windows 与
    Excel 文件名中的非法字符。
    """
    log = log or (lambda *_: None)
    out_dir = out_dir or _paths.resolve_output_dir("excel_tools")
    stem = os.path.splitext(os.path.basename(file))[0]
    ext = os.path.splitext(file)[1].lower()

    def _safe(name):
        """过滤 Windows 文件名非法字符，同时保留可读的原页签名称。"""
        return "".join("_" if c in '\\/:*?"<>|' else c for c in name)

    outs = []
    if ext in (".xlsx", ".xlsm"):
        # 只读盘一次；每个输出在内存深复制，保留绘图关系并减少磁盘解析次数。
        full = openpyxl.load_workbook(file)
        names = list(full.sheetnames)
        if len(names) < 2:
            full.close()
            raise ExcelToolError("该工作簿只有 1 个工作表,无需拆分")
        for target in names:
            wb = _copy.deepcopy(full)
            for sn in list(wb.sheetnames):
                if sn != target:
                    del wb[sn]
            # 原页可能是隐藏页；作为唯一页签时 Excel 要求至少一个可见页，因此强制可见。
            wb[target].sheet_state = "visible"
            wb.active = 0
            of = os.path.join(out_dir, "%s_%s.xlsx" % (stem, _safe(target)))
            wb.save(of)
            wb.close()
            outs.append(of)
            log("导出工作表「%s」(含格式)→ %s" % (target, os.path.basename(of)))
        full.close()
    else:
        sheets = _read_sheets(file)
        if len(sheets) < 2:
            raise ExcelToolError("该工作簿只有 1 个工作表,无需拆分")
        for name, rows in sheets:
            of = os.path.join(out_dir, "%s_%s.xlsx" % (stem, _safe(name)))
            wb = openpyxl.Workbook()
            _write_rows(wb.active, rows)
            wb.active.title = _safe_sheet_title(name, set())
            wb.save(of)
            outs.append(of)
            log("导出工作表「%s」(仅数据,老 .xls 无格式)→ %s"
                % (name, os.path.basename(of)))
    log("已按工作表拆分为 %d 个文件" % len(outs))
    return {"out_files": outs, "out_dir": out_dir, "out_file": outs[0] if outs else ""}


def _read_csv(path, log=None):
    """读取 CSV 行值，并在常见中文编码间按顺序降级。

    首选 ``utf-8-sig`` 以兼容 Excel 导出的 BOM，其次尝试 GBK 和普通 UTF-8。
    GBK 对任意字节都较宽容，可能“成功”得到乱码，因此解码后抽样检查替代字符和
    异常控制字符；可疑时记录警告但保留结果，让用户仍有机会检查和转换文件。
    """
    log = log or (lambda *_: None)
    for enc in ("utf-8-sig", "gbk", "utf-8"):
        try:
            with open(path, "r", encoding=enc, newline="") as f:
                rows = [row for row in csv.reader(f)]
        except (UnicodeDecodeError, UnicodeError):
            # 只有编码错误才尝试下一编码；文件不存在等 I/O 错误应原样上抛。
            continue
        # 最多抽取前 50 行、5000 字符，足以发现明显乱码且不随大文件线性增加开销。
        sample = "".join(str(c) for r in rows[:50] for c in r)[:5000]
        if sample:
            bad = sum(1 for ch in sample
                      if ord(ch) == 0xFFFD or (ord(ch) < 32 and ch not in "\t\n\r"))
            if bad and bad / float(len(sample)) > 0.02:
                log("警告:%s 以 %s 解码后疑似乱码,编码可能不是 utf-8/gbk,请核对"
                    % (os.path.basename(path), enc))
        return rows
    raise ExcelToolError("无法识别 CSV 编码:%s" % os.path.basename(path))


def convert(files, target, out_dir=None, log=None):
    """批量执行 xlsx 与 CSV 方向的值级格式转换。

    目标为 xlsx 时，每个源文件生成一本工作簿，xls 的全部页签都会保留为独立页；
    目标为 CSV 时，每个 Excel 页签各生成一个 UTF-8 BOM 文件，多页工作簿在文件名
    中追加页签名。转换只保证数据值，不承诺公式、样式或绘图对象。与目标格式相同
    的 CSV 会跳过；若最终没有任何输出则给出明确业务提示。
    """
    log = log or (lambda *_: None)
    if not files:
        raise ExcelToolError("请先选择文件")
    out_dir = out_dir or _paths.resolve_output_dir("excel_tools")
    outs = []
    for f in files:
        ext = os.path.splitext(f)[1].lower()
        stem = os.path.splitext(os.path.basename(f))[0]
        if target == "xlsx":
            # 删除 openpyxl 默认页，后续按源结构创建，避免多出一个空 Sheet。
            wb = openpyxl.Workbook(); wb.remove(wb.active); used = set()
            if ext == ".csv":
                ws = wb.create_sheet(_safe_sheet_title(stem, used))
                _write_rows(ws, _read_csv(f, log))
            else:
                for name, rows in _read_sheets(f):
                    _write_rows(wb.create_sheet(_safe_sheet_title(name, used)), rows)
            of = os.path.join(out_dir, stem + ".xlsx")
            wb.save(of); outs.append(of)
            log("%s → %s" % (os.path.basename(f), os.path.basename(of)))
        else:  # csv
            if ext == ".csv":
                continue  # 同格式转换既不产生新内容，也避免覆盖原文件。
            sheets = _read_sheets(f)
            for name, rows in sheets:
                suffix = ("_" + name) if len(sheets) > 1 else ""
                of = os.path.join(out_dir, "%s%s.csv" % (stem, suffix))
                # utf-8-sig 带 BOM，可让 Windows Excel 默认按 UTF-8 打开中文内容。
                with open(of, "w", encoding="utf-8-sig", newline="") as fh:
                    w = csv.writer(fh)
                    for r in rows:
                        w.writerow(["" if c is None else c for c in r])
                outs.append(of)
                log("%s[%s] → %s" % (os.path.basename(f), name, os.path.basename(of)))
    if not outs:
        raise ExcelToolError("没有可转换的文件(目标格式与源相同?)")
    return {"out_files": outs, "out_dir": out_dir, "out_file": outs[0]}


def _stack_header_keys(header, filename):
    """生成忽略空白和大小写的字段键，并拒绝同一表头内的重复字段。

    空字段使用位置占位，只有处于相同列位的空表头才会被视为同一字段；否则无法证明
    两份模板的空列语义一致，不能冒险按位置合并。
    """
    keys, seen = [], set()
    for index, value in enumerate(header, 1):
        text = "".join(_common.clean_str(value).split()).lower()
        key = text or "__blank_%d" % index
        if key in seen:
            raise ExcelToolError(
                "%s 的表头存在重复列「%s」，无法安全纵向合并"
                % (os.path.basename(filename), value or "空列")
            )
        seen.add(key)
        keys.append(key)
    return keys


def _stack_reorder_plan(base_keys, header, filename, log):
    """验证后续表头集合并返回按首表字段顺序取值的位置计划。"""
    keys = _stack_header_keys(header, filename)
    if set(keys) != set(base_keys):
        missing = [key for key in base_keys if key not in keys]
        extra = [key for key in keys if key not in base_keys]
        raise ExcelToolError(
            "%s 的表头与首表不一致（缺少:%s；新增:%s），请先统一结构"
            % (
                os.path.basename(filename),
                ",".join(missing) or "无",
                ",".join(extra) or "无",
            )
        )
    if keys != base_keys:
        log("%s 的表头顺序不同，已按列名重排后合并" % os.path.basename(filename))
    source_positions = {key: index for index, key in enumerate(keys)}
    return [source_positions[key] for key in base_keys]


def _stack_align_row(row, *, base_cols, reorder, filename, log):
    """按表头计划重排行，并把最终宽度固定到首表列数。"""
    values = list(row)
    if reorder is not None:
        # 尾部缺列按 None 补齐，既避免索引越界，也保持每个字段仍落在首表对应列。
        values = [values[index] if index < len(values) else None for index in reorder]
    if len(values) == base_cols:
        return values
    log(
        "警告:%s 某行列数为 %d,与首表 %d 列不一致,已补齐/截断对齐"
        % (os.path.basename(filename), len(values), base_cols)
    )
    if len(values) < base_cols:
        return values + [None] * (base_cols - len(values))
    return values[:base_cols]


def stack_tables(files, has_header=True, out_dir=None,
                 out_name="纵向合并.xlsx", log=None):
    """把多个同结构文件的首个页签纵向合并为一张可追溯明细表。

    有表头模式下，以首个有效文件建立基准字段；后续表头经清理空白和大小写后必须
    与基准字段集合一致，顺序不同会按列名重排，缺列或新增列则拒绝合并，避免数据
    静默错位。无表头模式只按首行列数对齐。每条输出记录追加“来源文件”，便于追溯
    原始数据。行列数异常会补空或截断，并通过日志提示。
    """
    log = log or (lambda *_: None)
    if len(files) < 2:
        raise ExcelToolError("纵向合并至少需要 2 个文件")
    out_dir = out_dir or _paths.resolve_output_dir("excel_tools")
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "合并"
    header_written = False
    total_rows = 0
    base_cols = None
    base_keys = None
    try:
        for f in files:
            sheets = _read_sheets(f)
            if not sheets:
                continue
            rows = sheets[0][1]
            if not rows:
                log("跳过空表 %s" % os.path.basename(f))
                continue
            start, reorder = 0, None
            if has_header:
                if not header_written:
                    # 第一份有效表决定可读表头、输出顺序和固定数据宽度。
                    base_keys = _stack_header_keys(rows[0], f)
                    base_cols = len(rows[0])
                    ws.append(list(rows[0]) + ["来源文件"])
                    header_written = True
                else:
                    reorder = _stack_reorder_plan(base_keys, rows[0], f, log)
                start = 1
            elif base_cols is None:
                base_cols = len(rows[0])  # 无表头模式只能以第一行宽度作为安全对齐基准。
            for source_row in rows[start:]:
                row = _stack_align_row(
                    source_row,
                    base_cols=base_cols,
                    reorder=reorder,
                    filename=f,
                    log=log,
                )
                ws.append(row + [os.path.basename(f)])
                total_rows += 1
            log("追加 %s:%d 行" % (os.path.basename(f), len(rows) - start))
        out_file = os.path.join(out_dir, out_name)
        wb.save(out_file)
    finally:
        # 表头验证或保存失败时也释放工作簿持有的临时资源，便于 Windows 立即重试。
        wb.close()
    log("已纵向合并 %d 个文件、共 %d 行数据 → %s" % (len(files), total_rows, out_file))
    return {"out_file": out_file, "out_dir": out_dir, "out_files": [out_file]}
