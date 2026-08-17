# -*- coding: utf-8 -*-
"""
考勤填表业务核心
================
读取一份或多份系统导出的每日打卡统计表，以“规范化姓名 + 业务日期”为联合键，
将上、下班打卡时间写入现有考勤模板，并计算实际时间、工时和加班。模块只处理
工作簿数据，不依赖桌面或 Web 界面，因此双端均通过同一入口复用本实现。

主要规则：
  · 源表支持 xlsx、xlsm 和 xls，由 ``common_core.read_sheets`` 统一读取；
  · 上班实际时间向上取整到半小时，下班实际时间向下取整到半小时；
  · 白班下班早于上班视为异常，不擅自按跨日补 24 小时；只有明确判定为夜班时
    才允许跨零点，防止漏打下班卡被误算成超长正常工时；
  · 异常行保留真实计算值、在原表标黄，并汇总到“异常核对报告”供人工复核；
  · 未匹配记录同样进入异常报告，但“假、休、调休”等非工时标记不误报；
  · 模板原有样式由 openpyxl 保留，公共解析、选项和输出目录规则来自 common_core。

本模块不推断或修改考勤制度。标准工时、夜班阈值、冲突策略等均由 ``Options``
传入，输出参数会随结果返回，便于前端展示并保留本次处理口径。
"""
import os
from dataclasses import dataclass

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

from . import attendance_source as _attendance_source
from . import common_core as cc
from .common_core import Options, norm_name, norm_date, parse_time, to_hours, parse_rest

# 异常统一使用纯黄色：既能在原模板中醒目提示，又与报告页表头保持一致。
_HL_FILL = PatternFill(fill_type="solid", fgColor="FFFF00")

# 旧调用方可能仍从本模块导入该名称，因此只做别名转发，不复制公共常量值。
STANDARD_WORKDAY_HOURS = cc.STANDARD_WORKDAY_HOURS


# ---------- 读取源表（每日统计表） ----------
def _detect_source_header(rows, opts, path=""):
    """兼容旧内部入口，实际识别规则由独立来源读取模块维护。

    参数原样透传给 attendance_source.detect_source_header；返回其识别到的
    ``(表头行索引, 列号映射)``，未识别到有效表头时返回 ``None``。
    """

    return _attendance_source.detect_source_header(rows, opts, path)


def load_source(path, opts=None):
    """读取单个每日统计文件，返回该文件首个有效页签解析出的打卡记录字典。

    该名称是历史公开入口，实现已委托给独立来源模块；本函数只做透传，便于来源
    识别规则单独演进，调用方不需要感知拆分细节。文件无有效表头时抛出 ValueError。
    """

    return _attendance_source.load_source(path, opts)


def load_source_multi(paths, opts=None, log=None):
    """合并多个来源文件，返回 ``(打卡记录字典, 来源统计字典)``。

    重复姓名日期的覆盖策略由 ``opts.conflict`` 决定，读取失败与冲突告警也由
    独立来源模块统一记录；本函数只保持公开入口稳定，不在此处重复实现合并口径。
    """

    return _attendance_source.load_source_multi(paths, opts, log)


# ---------- 异常行标黄 + 报告子表 ----------
def _highlight_row(ws, r, cols):
    """把目标字段覆盖的连续单元格区间标黄，便于在原模板中定位异常。

    只处理已识别且列号非空的角色；不涂满整行，避免破坏模板右侧可能存在的说明、
    签字或其他非业务区域。
    """
    used = [c for c in cols.values() if c]
    if not used:
        return
    for c in range(min(used), max(used) + 1):
        ws.cell(r, c).fill = _HL_FILL


def _write_report_sheet(wb, anomalies):
    """把异常明细写入独立核对页；没有异常时不创建空报告。

    若输入模板已含同名页，先删除再重建，确保本次输出不会混入历史异常。原始业务
    页上的行号、姓名和日期一并保存，使人工可从报告快速返回对应位置复核。
    """
    if not anomalies:
        return
    title = "异常核对报告"
    if title in wb.sheetnames:
        # 同名页通常来自重复处理旧输出；重建比追加更能保证报告属于当前计算结果。
        del wb[title]
    ws = wb.create_sheet(title)
    head = ["工作表", "行号", "姓名", "日期", "班次", "实际上班", "实际下班",
            "休息时间", "算出工时", "异常原因"]
    ws.append(head)
    for c in range(1, len(head) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True)
        cell.fill = _HL_FILL
        cell.alignment = Alignment(horizontal="center")
    for a in anomalies:
        ws.append([a["sheet"], a["row"], a["name"], a["date"], a.get("shift", ""),
                   a["act_on"], a["act_off"], a["rest"], a["hours"], a["reason"]])
    # 固定列宽优先保证异常原因可读，同时避免自动宽度扫描大表增加处理耗时。
    widths = [14, 6, 10, 12, 8, 10, 10, 10, 10, 34]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"


def _row_rest_mark(ws, r, cols, skip):
    """识别目标行中人工填写的非工时标记，避免将休假误报为缺卡。

    现有考勤模板可能把“假、休、调休”等写在系统或实际时间四个格中的任意一个，
    因此逐格检查；返回命中的原词便于调用方区分“找到标记”和“没有标记”。
    """
    for k in ("sys_on", "act_on", "sys_off", "act_off"):
        c = cols.get(k)
        if not c:
            continue
        txt = norm_name(ws.cell(r, c).value)
        if txt in skip and txt != "":
            return txt
    return ""


def compute_shift(a_on, a_off, rest, opts):
    """根据实际打卡时间计算班次、净工时、异常原因和加班基准。

    ``a_on``、``a_off`` 是已按业务规则取整的 ``datetime.time``，``rest`` 为小时数。
    返回 ``(hours, reason, shift, base)``：工时保留两位小数，异常时也保留真实负值；
    ``reason`` 为空代表可直接计算加班；``base`` 是对应班次的标准工时。

    班次只根据上班小时是否达到 ``night_start_hour`` 判定。夜班允许下班小于上班并
    补 24 小时，这是明确的跨零点场景；白班绝不自动补 24 小时，否则一次漏打卡
    可能被伪装成正常夜班。白班和夜班还分别受最大时长约束，用于捕捉漏卡或班次
    阈值配置错误。函数只返回异常说明，不自行修正可疑数据，最终决定留给人工。
    """
    on_h = to_hours(a_on)
    off_h = to_hours(a_off)
    is_night = opts.night_shift and on_h >= opts.night_start_hour
    shift = "夜班" if is_night else "白班"
    base = opts.night_workday_hours if is_night else opts.workday_hours
    raw = off_h - on_h  # 先保留同一自然日的有符号差值，随后按班次决定能否跨零点。
    reason = ""
    if is_night:
        if raw < 0:
            # 夜班的负差值是预期跨日表现，补一天后才是连续工作时长。
            raw += 24
        hours = round(raw - rest, 2)
        if raw > opts.night_max_hours:
            # 先判断原始跨度，再判断扣休息后的负值，错误信息更接近真实问题来源。
            reason = "夜班时长 %.2fh 超上限 %.2fh（疑漏打卡，请核对）" % (raw, opts.night_max_hours)
        elif hours < 0:
            reason = "夜班工时不足扣休息（时长 %.2fh 少于休息 %.2fh）" % (raw, rest)
    else:
        # 白班不做跨日补偿：负值和过长时长都必须交由人工确认，不能猜测班次。
        hours = round(raw - rest, 2)
        day_max = opts.day_max_hours
        if raw < 0:
            reason = "下班早于上班（实际下班 %s 早于实际上班 %s）" % (
                cc.fmt_time(a_off), cc.fmt_time(a_on))
        elif raw > day_max:
            reason = "白班时长 %.2fh 超上限 %.2fh（疑漏打卡，或本应为跨零点夜班，请核对）" % (raw, day_max)
        elif hours < 0:
            reason = "工时不足扣休息（时长 %.2fh 少于休息 %.2fh）" % (raw, rest)
    return hours, reason, shift, base


# ---------- 填写目标表（保留原格式，用 openpyxl 写回） ----------
def find_target_columns(ws, opts=None, path=""):
    """在待填写模板中定位考勤字段，返回表头行和 1 基列号映射。

    表头行优先采用管理员配置，否则默认为第一行。自动识别同时支持中英文括号、
    换行和“系统上班时间”等常见变体；先精确匹配再做包含匹配。管理员保存的角色
    映射最后覆盖自动结果，因为它代表人工确认；映射存储协议使用 0 基列号，而
    openpyxl 使用 1 基列号，所以覆盖时必须加一。
    """
    opts = opts or cc.DEFAULTS
    hr = opts.resolve_header(path) or 1
    def _norm_hdr(v):
        """清理表头换行并统一括号，降低仅有排版差异时的识别失败。"""
        t = norm_name(v).replace("\n", "")
        return t.replace("（", "(").replace("）", ")")

    # 字典值保留 openpyxl 的 1 基列号；重复同名表头按最后一列生效，与 Python
    # 字典构造规则一致，若模板确有歧义应由人工角色映射明确指定。
    header = {_norm_hdr(ws.cell(hr, c).value): c
              for c in range(1, ws.max_column + 1)}

    def col(*keys):
        """按候选词顺序先精确、后包含匹配，并返回首个可靠列号。"""
        for k in keys:
            if k in header:
                return header[k]
        for k in keys:
            for h, c in header.items():
                if k and k in h:
                    return c
        return None
    # 每个角色给出从具体到宽松的候选词；实际列缺失时保留 None，写表阶段按需跳过。
    cols = {
        "name": col("姓名"), "date": col("日期"),
        "sys_on": col("上班时间(系统)", "上班(系统)", "系统上班时间", "上班时间系统"),
        "act_on": col("上班时间(实际)", "上班(实际)", "实际上班时间", "上班时间实际"),
        "sys_off": col("下班时间(系统)", "下班(系统)", "系统下班时间", "下班时间系统"),
        "act_off": col("下班时间(实际)", "下班(实际)", "实际下班时间", "下班时间实际"),
        "rest": col("休息时间"), "work": col("实际工作时间"), "ot": col("加班"),
    }
    roles = opts.resolve_roles(path)
    for k, c0 in roles.items():
        if k in cols:
            # 配置层统一采用 0 基列号，转换后才能直接传给 ws.cell。
            cols[k] = c0 + 1
    return hr, cols


@dataclass
class _SheetFillStats:
    """单个考勤页签的写入计数，结束后再一次性合并到工作簿总计。"""

    matched: int = 0
    filled_time: int = 0
    computed_work: int = 0
    unmatched: int = 0
    filled_actual: int = 0
    anomalies: int = 0

    def merge_into(self, totals):
        """把页签计数累加到对外返回的工作簿统计字典。"""
        for key in (
            "matched", "filled_time", "computed_work",
            "unmatched", "filled_actual", "anomalies",
        ):
            totals[key] += getattr(self, key)


@dataclass
class _AttendanceRowContext:
    """一行考勤处理所需的共享工作簿、位置、统计与异常容器。"""

    ws: object
    row: int
    cols: dict[str, int]
    name: str
    day: tuple[int, int, int]
    stats: _SheetFillStats
    anomalies: list[dict[str, object]]
    log: object


@dataclass
class _AttendanceFillContext:
    """单个考勤页签填充阶段所需的全部输入。"""

    ws: object
    start: int
    cols: dict[str, int]
    source_data: dict[object, object]
    opts: Options
    skip: object
    anomalies: list[dict[str, object]]
    log: object


def _select_target_sheets(workbook, target_path, opts):
    """应用管理员的页签选择；未指定时返回工作簿全部页签。

    页签名精确匹配，找不到指定页签时抛出 ValueError，而不是静默改写其它页。
    """
    wanted = opts.resolve_sheet(target_path)
    if not wanted:
        return workbook.worksheets
    selected = [sheet for sheet in workbook.worksheets if sheet.title == wanted]
    if not selected:
        raise ValueError("目标表中找不到工作表 '%s'" % wanted)
    return selected


def _write_matched_times(ctx: _AttendanceRowContext, on_text, off_text, opts):
    """写入系统打卡时间，并按设置生成实际时间。

    系统时间只在有值且不是 ``-``/``—`` 占位符时写入；自动生成实际时间时，
    上班按半小时向上取整、下班按半小时向下取整，与考勤制度保持一致。写入计数
    累加到 ``ctx.stats``，本函数不返回业务结果。
    """
    ws, row, cols = ctx.ws, ctx.row, ctx.cols
    # 横线是模板占位符而非真实打卡，跳过以免把占位符当成有效时间写入。
    if cols["sys_on"] and on_text and on_text not in ("-", "—"):
        ws.cell(row, cols["sys_on"]).value = on_text
        ctx.stats.filled_time += 1
    if cols["sys_off"] and off_text and off_text not in ("-", "—"):
        ws.cell(row, cols["sys_off"]).value = off_text
    if not opts.auto_actual:
        return
    # 自动生成实际时间时按半小时整段取整：上班向上、下班向下，统一双端展示口径。
    if cols["act_on"]:
        rounded_on = cc.round_half_hour(parse_time(on_text), "up")
        if rounded_on is not None:
            ws.cell(row, cols["act_on"]).value = cc.fmt_time(rounded_on)
            ctx.stats.filled_actual += 1
    if cols["act_off"]:
        rounded_off = cc.round_half_hour(parse_time(off_text), "down")
        if rounded_off is not None:
            ws.cell(row, cols["act_off"]).value = cc.fmt_time(rounded_off)

def _compute_row_work(ctx: _AttendanceRowContext, opts):
    """根据实际上下班时间计算工时；返回该行是否具备完整可计算时间。

    工时写入“实际工作时间”列；无异常且开启加班时才写“加班”列，异常行只标黄
    并进入核对报告，不生成加班值，避免把可疑数据直接当成有效结果。休息时间列
    缺失时按 0 处理。
    """
    ws, row, cols = ctx.ws, ctx.row, ctx.cols
    if not (cols["work"] and cols["act_on"] and cols["act_off"]):
        return False
    actual_on = parse_time(ws.cell(row, cols["act_on"]).value)
    actual_off = parse_time(ws.cell(row, cols["act_off"]).value)
    if actual_on is None or actual_off is None:
        return False
    rest = parse_rest(ws.cell(row, cols["rest"]).value) if cols["rest"] else 0.0
    hours, reason, shift, base = compute_shift(actual_on, actual_off, rest, opts)
    ws.cell(row, cols["work"]).value = hours
    ctx.stats.computed_work += 1
    if reason:
        _record_shift_anomaly(
            ctx, actual_on, actual_off, rest, hours, shift, reason,
        )
    elif opts.overtime and cols["ot"]:
        # 仅无异常时进入此分支；异常行不生成加班值，避免把待核对数据直接当有效结果。
        overtime = round(hours - base, 2)
        ws.cell(row, cols["ot"]).value = overtime if overtime > 0 else 0
    return True


def _record_shift_anomaly(
    ctx: _AttendanceRowContext,
    actual_on,
    actual_off,
    rest,
    hours,
    shift,
    reason,
):
    """标记工时异常并添加到跨页签核对报告。

    异常行会在原表标黄，并把工作簿、行号、姓名、日期及计算结果写入统一报告，
    供人工从报告页快速定位复核；本函数只记录和提示，不修正任何数据。
    """
    ws, row, cols = ctx.ws, ctx.row, ctx.cols
    _highlight_row(ws, row, cols)
    ctx.stats.anomalies += 1
    ctx.anomalies.append({
        "sheet": ws.title,
        "row": row,
        "name": ctx.name,
        "date": "%04d-%02d-%02d" % ctx.day,
        "shift": shift,
        "act_on": cc.fmt_time(actual_on),
        "act_off": cc.fmt_time(actual_off),
        "rest": rest,
        "hours": hours,
        "reason": reason,
    })
    ctx.log(
        "    ! 异常：%s 第%d行 %s(%s) —— %s"
        % (ws.title, row, ctx.name, shift, reason)
    )


def _record_missing_punch_anomaly(ctx: _AttendanceRowContext, skip):
    """源数据未命中且无完整人工实际时间时，记录缺卡异常。

    “假、休、调休”等非工时标记视为合理缺卡，不误报；其余行标黄并写入核对报告。
    """
    ws, row, cols = ctx.ws, ctx.row, ctx.cols
    if _row_rest_mark(ws, row, cols, skip):
        return
    _highlight_row(ws, row, cols)
    ctx.stats.anomalies += 1
    ctx.anomalies.append({
        "sheet": ws.title,
        "row": row,
        "name": ctx.name,
        "date": "%04d-%02d-%02d" % ctx.day,
        "shift": "-",
        "act_on": "",
        "act_off": "",
        "rest": "",
        "hours": "",
        "reason": "未匹配到打卡数据（系统数据中查无此人此日，或姓名/日期不一致、缺卡）",
    })
    ctx.log("    ! 异常：%s 第%d行 %s —— 未匹配到打卡数据" % (ws.title, row, ctx.name))


def _fill_attendance_sheet(fill_ctx: _AttendanceFillContext):
    """处理一个有效数据页签，返回该页签独立统计。

    从 ``start`` 起逐行扫描到工作表末尾；姓名或日期为空的行跳过，避免把表尾
    说明文字误当业务行。匹配、写入、工时计算和缺卡异常全部累加到本页签统计中。
    """
    ws = fill_ctx.ws
    cols = fill_ctx.cols
    stats = _SheetFillStats()
    for row in range(fill_ctx.start, ws.max_row + 1):
        name = norm_name(ws.cell(row, cols["name"]).value)
        day = norm_date(ws.cell(row, cols["date"]).value)
        # 姓名或日期为空的行不是业务行，跳过可避免把表尾说明误配为缺卡。
        if not name or day is None:
            continue
        key = (name, day)
        ctx = _AttendanceRowContext(
            ws=ws, row=row, cols=cols, name=name, day=day,
            stats=stats, anomalies=fill_ctx.anomalies, log=fill_ctx.log,
        )
        matched = key in fill_ctx.source_data
        if matched:
            on_text, off_text = fill_ctx.source_data[key]
            stats.matched += 1
            _write_matched_times(ctx, on_text, off_text, fill_ctx.opts)
        else:
            stats.unmatched += 1
        computed = _compute_row_work(ctx, fill_ctx.opts)
        if not matched and not computed:
            _record_missing_punch_anomaly(ctx, fill_ctx.skip)
    return stats


def _log_sheet_result(log, sheet_title, stats):
    """输出单个页签的客户可读处理摘要。"""
    anomaly_text = (
        "，异常 %d 行(已标黄)" % stats.anomalies
        if stats.anomalies else ""
    )
    log(
        "工作表 '%s'：匹配 %d 行，填打卡 %d 处，算实际时间 %d 处，"
        "算工时 %d 行%s"
        % (
            sheet_title,
            stats.matched,
            stats.filled_time,
            stats.filled_actual,
            stats.computed_work,
            anomaly_text,
        )
    )


def fill_workbook(target_path, source_data, out_path, opts=None, log=None):
    """将合并后的打卡数据写入一个考勤模板，并保存独立输出文件。

    页签选择、单行写入、工时计算和异常记录已拆成独立步骤；所有步骤仍共用同一个打开
    的工作簿，保持原模板样式、计数口径和异常报告结构不变。无论写入是否成功，
    工作簿都会在 finally 中关闭；异常不在本层吞掉，由统一入口向调用方抛出。
    """
    opts = opts or cc.DEFAULTS
    log = log or (lambda *a, **k: None)
    skip = opts.skip_set()
    workbook = openpyxl.load_workbook(target_path)
    try:
        totals = {
            "sheets": [],
            "matched": 0,
            "filled_time": 0,
            "computed_work": 0,
            "unmatched": 0,
            "filled_actual": 0,
            "anomalies": 0,
        }
        anomalies = []
        data_start = opts.resolve_data_start(target_path)
        for ws in _select_target_sheets(workbook, target_path, opts):
            header_row, cols = find_target_columns(ws, opts, target_path)
            if not cols["name"] or not cols["date"]:
                log("跳过工作表 '%s'（未找到姓名/日期列）" % ws.title)
                continue
            # 数据起始行未配置时按“表头行 + 1”处理，确保表头本身不会进入业务行。
            sheet_stats = _fill_attendance_sheet(_AttendanceFillContext(
                ws=ws,
                start=data_start if data_start else header_row + 1,
                cols=cols,
                source_data=source_data,
                opts=opts,
                skip=skip,
                anomalies=anomalies,
                log=log,
            ))
            totals["sheets"].append((
                ws.title,
                sheet_stats.matched,
                sheet_stats.filled_time,
                sheet_stats.computed_work,
                sheet_stats.unmatched,
            ))
            sheet_stats.merge_into(totals)
            _log_sheet_result(log, ws.title, sheet_stats)
        _write_report_sheet(workbook, anomalies)
        if anomalies:
            log(
                "⚠ 共 %d 行异常数据需人工核对，已标黄并汇总到子表『异常核对报告』。"
                % len(anomalies)
            )
        workbook.save(out_path)
        return totals
    finally:
        workbook.close()


# ---------- 统一入口：与对账功能同构 ----------
def run(targets, sources, opts=None, log=None, out_dir=None, progress=None):
    """执行一批考勤填表任务，并返回双端可消费的结构化结果。

    ``targets`` 是一个或多个待填模板，``sources`` 是一个或多个打卡来源；字符串会
    自动提升为单元素列表。未传输出目录时使用统一路径配置，显式传入时仅确保目录
    存在。全部目标共享同一时间戳，便于识别为同一批处理，但每个模板生成独立文件。

    进度按“读取合并 30% + 模板写入 70%”分段，写入阶段再按目标数量均分。返回值
    包含输出文件、源表统计、每个模板统计和本次关键参数，前端无需重新解析 Excel
    即可展示处理结果与计算口径。
    """
    opts = opts or cc.DEFAULTS
    log = log or (lambda *a, **k: None)
    if isinstance(targets, str):
        targets = [targets]
    if isinstance(sources, str):
        sources = [sources]
    if not targets:
        raise ValueError("请选择待填写的考勤表模板")
    if not sources:
        raise ValueError("请选择打卡来源表")
    # 写回工作簿和保存文件通常比读取更耗时，因此把大部分进度权重分配给填表阶段。
    prog = cc.Progress(progress, stages=[("read", 30), ("fill", 70)])
    ts = cc.timestamp()
    if out_dir is None:
        # 统一目录解析异常时兼容旧部署逻辑，保证历史环境仍能完成核心业务。
        out_dir = _unified_out_dir("attendance", ts, src=targets[0], log=log) or cc.make_out_dir(targets[0])
    else:
        os.makedirs(out_dir, exist_ok=True)
    log("采用选项：" + opts.summary())
    prog.stage("read")
    log("① 读取并合并系统数据（%d 个文件）..." % len(sources))
    data, sstat = load_source_multi(sources, opts=opts, log=log)
    log("   合并后共 %d 条打卡记录。" % sstat["records"])

    prog.stage("fill")
    n_tgt = len(targets)
    out_files, results = [], []
    for i, tgt in enumerate(targets, 1):
        log("\n② 填写第 %d/%d 个待填表：%s" % (i, n_tgt, os.path.basename(tgt)))
        base = os.path.splitext(os.path.basename(tgt))[0]
        op = cc.out_path(out_dir, base, "_已填写", ".xlsx", ts=ts)
        stats = fill_workbook(tgt, data, op, opts=opts, log=log)
        log("   匹配 %d 行、填打卡 %d 处、算工时 %d 行、未匹配 %d 行、异常 %d 行"
            % (stats["matched"], stats["filled_time"], stats["computed_work"],
               stats["unmatched"], stats["anomalies"]))
        log("   已保存：%s" % op)
        out_files.append(op); results.append((tgt, op, stats))
        prog.tick(i, n_tgt)  # 以完成的目标数推进，避免大文件处理中虚构精细进度。
    prog.done()
    return {
        "out_files": out_files,
        "out_dir": out_dir,
        "source_stat": sstat,
        "results": results,
        "parameters": {
            "workday_hours": opts.workday_hours,
            "overtime": opts.overtime,
            "conflict": opts.conflict,
            "auto_actual": opts.auto_actual,
            "night_shift": opts.night_shift,
            "night_start_hour": opts.night_start_hour,
            "night_workday_hours": opts.night_workday_hours,
            "night_max_hours": opts.night_max_hours,
            "day_max_hours": opts.day_max_hours,
            "skip_extra": sorted(opts.skip_extra),
        },
    }


def _unified_out_dir(feature, ts=None, src=None, log=None):
    """按当前设置解析统一输出目录，失败时返回 ``None`` 触发旧逻辑回退。

    ``src`` 仅在“输出到源文件旁”模式中补充定位依据。这里采用延迟导入并容忍异常，
    是为了兼容早期独立分发过的考勤模块；正式项目环境应正常走 ``core.paths``。
    回退只影响输出位置，不改变任何业务计算；失败原因写入日志，避免静默吞掉配置错误。
    """
    try:
        from . import paths as _paths
        from . import settings as _settings
        st = _settings.get_settings()
        kw = st.output_kwargs()
        if src and not kw.get("src_path"):
            # 调用方显式配置的 src_path 优先，本函数只在缺失时用首个目标模板补全。
            kw["src_path"] = src
        return _paths.resolve_output_dir(feature, ts=ts, **kw)
    except Exception as error:
        # 由 run 使用 common_core.make_out_dir 接管，旧环境不会因路径模块不可用而中断。
        if log:
            log("统一输出目录不可用，回退到源文件旁输出：%s" % error)
        return None
